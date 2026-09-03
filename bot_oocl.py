import time
import threading
import subprocess
import socket
import re
import os
import html
import urllib.request
import urllib.parse
import openpyxl
import calendar
import math
from itertools import product
from datetime import datetime, timedelta, date as date_type
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.action_chains import ActionChains
import random
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from bot_runtime_utils import wait_for_terminal_enter
from remark_rules import charge_amount_to_usd, get_manifest_code, is_china_destination

# ── Timestamp print ──
import sys, io
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
_orig_print = print
def print(*args, **kwargs):
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    _orig_print(f"[{ts}]", *args, **kwargs)

def r():
    time.sleep(random.uniform(0.01, 0.04))

# ========================== DEBUG FLAGS ==========================
ENABLE_TIMING = False
ENABLE_DEBUG  = os.environ.get("OOCL_DEBUG", "0").strip().lower() in ("1", "true", "yes", "y", "on")
OOCL_ESPOT_SCAN_NEXT_WEEK = os.environ.get("OOCL_ESPOT_SCAN_NEXT_WEEK", "0").strip().lower() in ("1", "true", "yes", "y", "on")

def debug_print(*args, **kwargs):
    if ENABLE_DEBUG:
        print(*args, **kwargs)

# ========================== CONFIG ==========================
OOCL_URL           = "https://freightsmart.oocl.com/en/"
NEW_UI_URL         = "https://freightsmart.oocl.com/digital/"
EQUOTE_NEW_URL     = "https://freightsmart.oocl.com/ui/my-quotation"
OOCL_APP_URL_MARKERS = ("/ui", "/digital", "/dashboard", "booking-request", "search-result")
LOGIN_EMAIL        = os.environ.get("OOCL_EMAIL", "celine@pio-logistics.vn")
LOGIN_PASSWORD     = os.environ.get("OOCL_PASSWORD", "XvntXvnt*3536")
OOCL_LOGIN_URL     = (
    "https://freightsmart.oocl.com/api/admin/keycloak/sso/token"
    f"?login_hint={urllib.parse.quote(LOGIN_EMAIL, safe='@')}&ui_locales=en"
)
OOCL_MANUAL_LOGIN_WAIT_PATH = os.environ.get(
    "OOCL_MANUAL_LOGIN_WAIT_PATH",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), ".oocl_manual_login_wait"),
)
EXCEL_PATH         = os.environ.get("EXCEL_PATH", "input_gia.xlsx")
DRIVER_PATH        = os.path.join(os.path.dirname(os.path.abspath(__file__)), "msedgedriver.exe")
FILTER_POL         = os.environ.get("FILTER_POL", "").strip().upper()
FILTER_POD         = os.environ.get("FILTER_POD", "").strip().upper()
SINGLE_ROW         = os.environ.get("SINGLE_ROW", "").strip()
DATE_OFFSET_DAYS = parse_date_offset_days()
try:
    OOCL_POST_CONNECT_SLEEP = max(0.0, float(os.environ.get("OOCL_POST_CONNECT_SLEEP", "0.5")))
    OOCL_TAB_NAV_SLEEP = max(0.0, float(os.environ.get("OOCL_TAB_NAV_SLEEP", "0.35")))
    OOCL_FOCUS_SLEEP = max(0.0, float(os.environ.get("OOCL_FOCUS_SLEEP", "0.1")))
    OOCL_VISUAL_SWITCH_SLEEP = max(0.0, float(os.environ.get("OOCL_VISUAL_SWITCH_SLEEP", "0.06")))
    OOCL_TAB_CLOSE_SLEEP = max(0.0, float(os.environ.get("OOCL_TAB_CLOSE_SLEEP", "0.05")))
except ValueError:
    OOCL_POST_CONNECT_SLEEP = 0.5
    OOCL_TAB_NAV_SLEEP = 0.35
    OOCL_FOCUS_SLEEP = 0.1
    OOCL_VISUAL_SWITCH_SLEEP = 0.06
    OOCL_TAB_CLOSE_SLEEP = 0.05
try:
    OOCL_LOGIN_WAIT_SECONDS = max(20, int(os.environ.get("OOCL_LOGIN_WAIT_SECONDS", "45")))
except ValueError:
    OOCL_LOGIN_WAIT_SECONDS = 45
try:
    OOCL_LOGIN_MANUAL_WAIT_SECONDS = max(60, int(os.environ.get("OOCL_LOGIN_MANUAL_WAIT_SECONDS", "900")))
except ValueError:
    OOCL_LOGIN_MANUAL_WAIT_SECONDS = 900
try:
    OOCL_EQUOTE_RESULT_WAIT_SECONDS = max(4, int(os.environ.get("OOCL_EQUOTE_RESULT_WAIT_SECONDS", "7")))
except ValueError:
    OOCL_EQUOTE_RESULT_WAIT_SECONDS = 7
try:
    OOCL_EQUOTE_COLLECT_WAIT_SECONDS = max(3, int(os.environ.get("OOCL_EQUOTE_COLLECT_WAIT_SECONDS", "6")))
except ValueError:
    OOCL_EQUOTE_COLLECT_WAIT_SECONDS = 6
try:
    OOCL_EQUOTE_SAME_SIGNATURE_GRACE_SECONDS = max(1.0, float(os.environ.get("OOCL_EQUOTE_SAME_SIGNATURE_GRACE_SECONDS", "2.5")))
except ValueError:
    OOCL_EQUOTE_SAME_SIGNATURE_GRACE_SECONDS = 2.5
try:
    OOCL_ESPOT_SAME_SIGNATURE_GRACE_SECONDS = max(1.0, float(os.environ.get("OOCL_ESPOT_SAME_SIGNATURE_GRACE_SECONDS", "2.5")))
except ValueError:
    OOCL_ESPOT_SAME_SIGNATURE_GRACE_SECONDS = 2.5

def _excel_formula_from_parts(parts):
    tokens = []
    for raw in parts:
        try:
            amount = float(raw)
        except (TypeError, ValueError):
            continue
        if abs(amount) < 1e-9:
            continue
        sign = "-" if amount < 0 else "+"
        amount = abs(amount)
        if abs(amount - round(amount)) < 1e-9:
            text = str(int(round(amount)))
        else:
            text = f"{amount:.2f}".rstrip("0").rstrip(".")
        tokens.append((sign if tokens else ("-" if sign == "-" else "")) + text)
    return "=" + "".join(tokens) if tokens else None

def _excel_value_with_formula_fallback(formula, price):
    if formula:
        return formula
    if isinstance(price, str):
        text = price.strip()
        if text.startswith("="):
            return text
        try:
            price = float(text.replace(",", ""))
        except ValueError:
            return price or 0
    return _excel_formula_from_parts([price]) or price or 0

def _ceil_money_value(value):
    try:
        return int(math.ceil(float(value) - 1e-9))
    except Exception:
        return 0


def _parse_price_display(text):
    m = re.search(r'([\d,]+(?:\.\d+)?)', str(text or ""))
    if not m:
        return 0
    return _ceil_money_value(m.group(1).replace(",", ""))

SHEET_NAME         = "Sheet1"
EDGE_EXE           = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
EDGE_DEBUG_PORT    = int(os.environ.get("OOCL_EDGE_PORT", "9527"))
EDGE_USER_DATA_DIR = os.environ.get("OOCL_EDGE_PROFILE") or os.environ.get("OOCL_EDGE_USER_DATA_DIR") or r"C:\edge_oocl"
OOCL_WORKER_INDEX  = max(0, int(os.environ.get("OOCL_WORKER_INDEX", "0") or "0"))
OOCL_WORKER_COUNT  = max(1, int(os.environ.get("OOCL_WORKER_COUNT", "1") or "1"))
PORT_SEARCH_ALIAS  = {
    "Antwerp":  "Antwerpen",
    "COCHIN":   "KOCHI",
    "Cochin":   "Kochi",
    "PIPAVAV":   "RAJULA",
}

# Mapping (PORT_NAME_UPPER, COUNTRY_UPPER) → UN/LOCODE cho các cảng trùng tên
PORT_COUNTRY_CODE = {
    ("HAMBURG", "GERMANY"):     "DEHAM",
    ("CHARLESTON", "USA"):      "USCHS",
    ("CHARLESTON", "UNITED STATES"): "USCHS",
    ("VALENCIA", "SPAIN"):      "ESVLC",
    ("VICTORIA", "CANADA"):     "CAVIC",
    ("PORTLAND", "USA"):        "USPDX",
    ("PORTLAND", "UNITED STATES"): "USPDX",
    ("GEORGETOWN", "MALAYSIA"): "MYGEP",
    ("TRIPOLI", "LIBYA"):       "LYTIP",
    ("TRIPOLI", "LEBANON"):     "LBKYE",
    ("NEWCASTLE", "AUSTRALIA"): "AUNTL",
}
PORT_COUNTRY_SEARCH_QUERY = {
    ("NEWCASTLE", "AUSTRALIA"): "NEWCASTLE NEW SOUTH WALES AUSTRALIA",
}
PORT_SELECT_DELAY  = random.uniform(0.01, 0.04)
MAX_QUALIFIED_CARDS = 3
OOCL_BLOCK_MARKERS = (
    "CONNECTION FROM CURRENT IP ADDRESS IS BLOCKED",
    "CURRENT IP",
    "IP ADDRESS IS BLOCKED",
    "CLIENT IP",
)

EU_COUNTRIES = {
    "GERMANY","FRANCE","NETHERLANDS","BELGIUM","SPAIN","ITALY",
    "UNITED KINGDOM","UK","ENGLAND","SCOTLAND","WALES","IRELAND",
    "POLAND","PORTUGAL","SWEDEN","DENMARK","NORWAY","FINLAND",
    "AUSTRIA","SWITZERLAND","GREECE","TURKEY","RUSSIA","UKRAINE",
    "CZECH","HUNGARY","ROMANIA","BULGARIA","CROATIA","SERBIA",
    "SLOVENIA","SLOVAKIA","LATVIA","ESTONIA","LITHUANIA","ISRAEL",
    "MALTA","CYPRUS","LUXEMBOURG","ICELAND","ALBANIA","MONTENEGRO",
}
CN_COUNTRIES = {"CHINA"}
JP_COUNTRIES = {"JAPAN"}

SURCHARGE_EXCLUDE_EQUOTE = [
    "20' HEAVY WEIGHT CHARGE",
    "HIGH SECURITY SEAL CHARGE",
    "OUTBOUND DOCUMENTATION FEE",
    "TERMINAL HANDLING CHARGE AT ORIGIN",
    "ONLY IF WEIGHT",
]

EQUOTE_SURCHARGE_HARD_EXCLUDE = [
    "TERMINAL HANDLING CHARGE AT ORIGIN",
    "OUTBOUND DOCUMENTATION FEE",
    "HIGH SECURITY SEAL CHARGE",
    "ADVANCE MANIFEST SECURITY CHARGE",
]

EQUOTE_OWS_KEYWORDS = [
    "20' HEAVY WEIGHT CHARGE",
    "HEAVY WEIGHT CHARGE",
    "ONLY IF WEIGHT",
]

# ========================== HELPERS ==========================

def fast_input(driver, element, text):
    """
    Nhập text trigger React dropdown.
    Dùng JS setValue + send_keys ký tự cuối để trigger onChange.
    """
    try:
        # Set value qua JS native setter
        driver.execute_script("""
            let el  = arguments[0];
            let val = arguments[1];
            let setter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value').set;
            setter.call(el, val);
            el.dispatchEvent(new Event('input',   {bubbles:true}));
            el.dispatchEvent(new Event('change',  {bubbles:true}));
            el.dispatchEvent(new KeyboardEvent('keydown',
                {bubbles:true, key:val.slice(-1)}));
            el.dispatchEvent(new KeyboardEvent('keyup',
                {bubbles:true, key:val.slice(-1)}));
        """, element, text)
        # Gửi thêm 1 ký tự Space rồi Backspace để chắc chắn React re-render
        element.send_keys(Keys.SPACE)
        element.send_keys(Keys.BACK_SPACE)
    except Exception as e:
        # Fallback: send_keys thường
        try:
            element.clear()
            element.send_keys(text)
        except Exception:
            pass

def smart_wait(driver, condition_fn, timeout=10, poll=0.1, after=0.1, label=""):
    """
    Polling liên tục mỗi poll giây cho đến khi condition_fn(driver) = True.
    Khi thỏa → chờ thêm after giây rồi trả về True.
    Timeout → trả về False.
    """
    end = time.time() + timeout
    while time.time() < end:
        try:
            if condition_fn(driver):
                time.sleep(after)
                return True
        except Exception:
            pass
        time.sleep(poll)
    if label:
        print(f"[WARN] smart_wait timeout ({timeout}s): {label}")
    return False

def is_port_open(port, host="127.0.0.1"):
    try:
        with socket.create_connection((host, port), timeout=1):
            return True
    except (OSError, ConnectionRefusedError):
        return False

def safe_click(driver, element, delay=None):
    if delay is None: delay = random.uniform(0.01, 0.04)
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
        time.sleep(delay)
        element.click()
    except Exception:
        try:
            driver.execute_script("arguments[0].click();", element)
        except Exception:
            pass

def get_pod_region(country):
    c = country.strip().upper()
    if c in EU_COUNTRIES: return "EU"
    if c in CN_COUNTRIES: return "CN"
    if c in JP_COUNTRIES: return "JP"
    return "OTHER"

def calculate_valid_for_espot(etd_dates):
    if not etd_dates:
        return None
    dates = [d.date() if hasattr(d, 'date') else d for d in etd_dates]
    latest = max(dates)
    year, month = latest.year, latest.month
    last_day = calendar.monthrange(year, month)[1]
    for ms in [7, 14, 21, last_day]:
        d = date_type(year, month, ms)
        if d >= latest:
            return d
    if month == 12:
        return date_type(year + 1, 1, 7)
    return date_type(year, month + 1, 7)

def parse_sailing_date(text):
    try:
        clean = re.sub(r'\s*\([^)]+\)', '', text).strip()
        parts = clean.split()
        if len(parts) >= 2:
            year = datetime.now().year
            dt = datetime.strptime(f"{parts[0]} {parts[1]} {year}", "%d %b %Y").date()
            if dt < datetime.now().date() and (datetime.now().date() - dt).days > 180:
                dt = datetime.strptime(f"{parts[0]} {parts[1]} {year+1}", "%d %b %Y").date()
            return dt
    except Exception:
        pass
    return None

def parse_date_from_text(text):
    try:
        parts = text.strip().split()
        if len(parts) >= 2:
            year = datetime.now().year
            dt = datetime.strptime(f"{parts[0]} {parts[1]} {year}", "%d %b %Y")
            if dt < datetime.now() and (datetime.now() - dt).days > 180:
                dt = datetime.strptime(f"{parts[0]} {parts[1]} {year+1}", "%d %b %Y")
            return dt
    except Exception:
        pass
    return None

def format_etd_for_excel(etd_dates):
    if not etd_dates:
        return ""
    dates = [d.date() if hasattr(d, 'date') else d for d in etd_dates]
    fmt = [(d.day, d.strftime("%b")) for d in dates]
    if len(fmt) == 1:
        return f"{fmt[0][0]}-{fmt[0][1]}"
    if len(fmt) == 2:
        return f"{fmt[0][0]}-{fmt[0][1]} & {fmt[1][0]}-{fmt[1][1]}"
    months = [f[1] for f in fmt]
    if months[0] == months[1] == months[2]:
        return f"{fmt[0][0]}, {fmt[1][0]}, {fmt[2][0]}-{fmt[2][1]}"
    if months[0] == months[1]:
        return f"{fmt[0][0]}, {fmt[1][0]}-{fmt[1][1]}, {fmt[2][0]}-{fmt[2][1]}"
    return f"{fmt[0][0]}-{fmt[0][1]}, {fmt[1][0]}-{fmt[1][1]}, {fmt[2][0]}-{fmt[2][1]}"

def format_transit_for_excel(transit_list):
    if not transit_list:
        return ""
    unique = set(transit_list)
    if len(unique) == 1:
        return str(next(iter(unique)))
    return f"{min(transit_list)}-{max(transit_list)}"

def normalize_port_text(text):
    raw = (text or "").strip()
    if not raw:
        return ""
    clean = re.sub(r'\s*\(.*?\)', '', raw).strip()
    return clean.split(",")[0].strip().upper()

def _close_tab_safe(driver, handle, keep_tab, timeout=15):
    """Đóng 1 tab an toàn với timeout tổng (mặc định 15s).
    - Pre-check window_handles trước khi switch để tránh treo 120s.
    - Toàn bộ thao tác được wrap trong thread với timeout cứng.
    """
    result = {"done": False}

    def _do_close():
        try:
            # Pre-check: tab còn tồn tại không?
            try:
                handles = driver.window_handles
            except Exception:
                return
            if handle not in handles:
                if keep_tab in handles:
                    driver.switch_to.window(keep_tab)
                return

            driver.switch_to.window(handle)
            try:
                driver.execute_script("window.stop();")
                driver.execute_script("window.location.href = 'about:blank';")
                time.sleep(0.25)
            except Exception:
                pass
            try:
                driver.execute_script("window.close();")
                time.sleep(0.2)
            except Exception:
                pass
            try:
                if handle in driver.window_handles:
                    driver.close()
                    time.sleep(0.2)
            except Exception:
                pass
        except Exception as e:
            print(f"[WARN] _close_tab_safe {handle[-6:]}: {e}")
        finally:
            try:
                if keep_tab in driver.window_handles:
                    driver.switch_to.window(keep_tab)
            except Exception:
                pass
            result["done"] = True

    t = threading.Thread(target=_do_close, daemon=True)
    t.start()
    t.join(timeout=timeout)
    if not result["done"]:
        print(f"[WARN] _close_tab_safe {handle[-6:]}: timeout sau {timeout}s → bỏ qua")

# ========================== BASE ==========================

class OOCLIpBlockedError(RuntimeError):
    pass


class OOCLSetupTabsError(RuntimeError):
    """OOCL could not initialize both E-Spot and E-Quote tabs."""
    pass


class OOCLManualLoginTimeout(RuntimeError):
    """User did not finish OOCL OTP/CAPTCHA inside the manual window."""
    pass


class OOCLBaseScraper:

    def _is_oocl_ip_blocked(self):
        try:
            url = (self.driver.current_url or "").lower()
            if "freightsmart.oocl.com" not in url:
                return False
            body = (self.driver.find_element(By.TAG_NAME, "body").text or "").upper()
            return "IP ADDRESS IS BLOCKED" in body or (
                "CONNECTION FROM CURRENT IP ADDRESS IS BLOCKED" in body
            )
        except Exception:
            return False

    def _check_driver_alive(self, timeout=8):
        """Kiểm tra nhanh driver/msedgedriver còn sống không.
        Trả về True nếu OK, False nếu driver chết."""
        result = {"alive": False}
        def _ping():
            try:
                _ = self.driver.title
                result["alive"] = True
            except Exception:
                pass
        t = threading.Thread(target=_ping, daemon=True)
        t.start()
        t.join(timeout=timeout)
        return result["alive"]

    def _abort_if_oocl_blocked(self, context=""):
        if self._is_oocl_ip_blocked():
            msg = "OOCL IP BLOCKED"
            if context:
                msg += f" ({context})"
            try:
                print(f"[BLOCK] {msg}. Dừng bot OOCL, không retry để tránh block nặng hơn.")
                print(f"[BLOCK] URL: {self.driver.current_url}")
            except Exception:
                print(f"[BLOCK] {msg}.")
            raise OOCLIpBlockedError(msg)

    def _current_url_lower(self):
        try:
            return (self.driver.current_url or "").lower()
        except Exception:
            return ""

    def _is_oocl_login_url(self):
        url = self._current_url_lower()
        return any(k in url for k in (
            "/en/login",
            "/login?logintype",
            "/api/admin/keycloak/sso/token",
            "exiamfw.home.oocl.com",
            "/openid-connect/auth",
            "/login-actions/",
        ))

    def _is_visible(self, element):
        try:
            return bool(element and element.is_displayed())
        except Exception:
            return False

    def _find_visible_xpath(self, xpaths, timeout=5, clickable=False):
        candidates = xpaths if isinstance(xpaths, (list, tuple)) else [xpaths]
        end = time.time() + timeout
        while time.time() < end:
            for xp in candidates:
                try:
                    elems = self.driver.find_elements(By.XPATH, xp)
                except Exception:
                    continue
                for elem in elems:
                    try:
                        if not elem.is_displayed():
                            continue
                        if clickable and not elem.is_enabled():
                            continue
                        return elem
                    except StaleElementReferenceException:
                        continue
                    except Exception:
                        continue
            time.sleep(0.15)
        return None

    def _has_visible_login_controls(self):
        controls = [
            "//input[@type='password']",
            "//input[@id='email-input' or @type='email' or @name='email']",
            "//input[@id='username' or @name='username']",
            "//button[@id='kc-login' or @name='login']",
            "//button[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
            "//a[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
        ]
        return self._find_visible_xpath(controls, timeout=1, clickable=False) is not None

    def _is_oocl_app_ready(self, success_url_keywords=OOCL_APP_URL_MARKERS):
        url = self._current_url_lower()
        if self._is_oocl_login_url():
            return False
        if not any(k.lower() in url for k in success_url_keywords):
            return False
        return not self._has_visible_login_controls()

    def _oocl_manual_login_check_present(self):
        try:
            body = (self.driver.find_element(By.TAG_NAME, "body").text or "").lower()
        except Exception:
            body = ""
        markers = (
            "multi-factor",
            "mfa",
            "verification code",
            "email verification",
            "verify your identity",
            "authentication code",
            "enter the code",
            "sent a code",
            "code to your email",
            "one-time",
            "one time",
            "otp",
            "authenticator",
            "captcha",
            "security check",
        )
        if any(m in body for m in markers):
            return True
        try:
            return bool(self.driver.find_elements(
                By.XPATH,
                "//input[contains(translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'otp') "
                "or contains(translate(@id,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'otp') "
                "or contains(translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'code') "
                "or contains(translate(@id,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'code') "
                "or contains(translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'verification') "
                "or contains(translate(@id,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'verification') "
                "or contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'code') "
                "or contains(translate(@aria-label,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'code') "
                "or @autocomplete='one-time-code']"
            ))
        except Exception:
            return False

    def _set_input_value(self, element, value, remove_readonly=False):
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
        except Exception:
            pass
        if remove_readonly:
            try:
                self.driver.execute_script("arguments[0].removeAttribute('readonly');", element)
            except Exception:
                pass
        try:
            element.click()
            r()
            element.send_keys(Keys.CONTROL + "a")
            r()
            element.send_keys(Keys.DELETE)
            r()
            element.send_keys(value)
            r()
        except Exception:
            pass
        try:
            self.driver.execute_script("""
                const input = arguments[0];
                const value = arguments[1];
                const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                setter.call(input, value);
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
                input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: value.slice(-1) || 'a' }));
            """, element, value)
        except Exception:
            pass
        try:
            return (element.get_attribute("value") or "").strip() == str(value).strip()
        except Exception:
            return False

    def _click_first_login_button(self, timeout=8):
        button_xpaths = [
            "//button[@id='kc-login']",
            "//button[@name='login']",
            "//input[@id='email-login-bt']",
            "//button[@id='email-login-bt']",
            "//button[@type='submit']",
            "//button[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
            "//button[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
            "//button[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'next')]",
            "//button[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'continue')]",
            "//input[@type='button' and contains(translate(@value,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
            "//input[@type='submit']",
        ]
        btn = self._find_visible_xpath(button_xpaths, timeout=timeout, clickable=True)
        if btn:
            safe_click(self.driver, btn, delay=0.2)
            return True
        return False

    def _wait_for_oocl_login_or_app(self, timeout=None, success_url_keywords=OOCL_APP_URL_MARKERS):
        timeout = timeout or OOCL_LOGIN_WAIT_SECONDS
        end = time.time() + timeout
        last_log = time.time()
        while time.time() < end:
            self._abort_if_oocl_blocked("wait login/app")
            if self._is_oocl_app_ready(success_url_keywords):
                return "APP"
            # The new OTP page still exposes a visible readonly username.  OTP
            # must win over USER/PASSWORD or the bot restarts the login flow.
            if self._oocl_manual_login_check_present():
                return "MANUAL"
            pwd = self._find_visible_xpath("//input[@type='password' or @id='password' or @name='password']", timeout=0.5, clickable=True)
            if pwd:
                return "PASSWORD"
            email_or_user = self._find_visible_xpath([
                "//input[@id='email-input' or @type='email' or @name='email']",
                "//input[@id='username' or @name='username']",
                "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'email')]",
                "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'user id')]",
            ], timeout=0.5, clickable=True)
            if email_or_user:
                return "USER"
            if time.time() - last_log > 10:
                print("[INFO] OOCL login: đang chờ form/redirect ổn định...")
                last_log = time.time()
            time.sleep(0.4)
        return "TIMEOUT"

    def _wait_manual_login_done(
        self,
        reason="MFA/CAPTCHA",
        success_url_keywords=OOCL_APP_URL_MARKERS,
        input_func=None,
    ):
        """Pause all Selenium activity while the user completes OTP/CAPTCHA."""
        manual_input = input_func or input
        marker_created = False
        try:
            with open(OOCL_MANUAL_LOGIN_WAIT_PATH, "w", encoding="utf-8") as marker:
                marker.write(str(os.getpid()))
            marker_created = True
        except OSError as exc:
            print(f"[WARN] Không tạo được cờ chờ OOCL OTP: {type(exc).__name__}")

        print(f"[PAUSE] OOCL đang yêu cầu {reason}.")
        print("[PAUSE] Bot đã DỪNG thao tác browser; hãy nhập OTP/xử lý xác thực trên Edge OOCL.")
        print(
            "[PAUSE] Sau khi hoàn tất, quay lại terminal và nhấn ENTER để bot tiếp tục "
            f"(tối đa {OOCL_LOGIN_MANUAL_WAIT_SECONDS // 60} phút)."
        )
        try:
            resumed = wait_for_terminal_enter(
                manual_input,
                OOCL_LOGIN_MANUAL_WAIT_SECONDS,
                "[PAUSE] Nhấn ENTER sau khi OOCL đã đăng nhập xong... ",
            )
            if not resumed:
                raise OOCLManualLoginTimeout(
                    f"OOCL {reason} chưa hoàn tất sau {OOCL_LOGIN_MANUAL_WAIT_SECONDS} giây"
                )
        finally:
            if marker_created:
                try:
                    os.remove(OOCL_MANUAL_LOGIN_WAIT_PATH)
                except OSError:
                    pass

        self._abort_if_oocl_blocked(f"manual login {reason} resumed")
        resumed_url = self._current_url_lower()
        is_espot_workspace = "/digital" in resumed_url or "search-result" in resumed_url
        if is_espot_workspace and self._is_oocl_app_ready(success_url_keywords):
            print("[OK] OOCL manual login xong; session đã ở trong app.")
            return True

        print(f"[INFO] Sau ENTER chưa ở trang check; chuyển tới {NEW_UI_URL}")
        try:
            self.driver.get(NEW_UI_URL)
        except TimeoutException:
            try:
                self.driver.execute_script("window.stop();")
            except Exception:
                pass

        state = self._wait_for_oocl_login_or_app(
            timeout=OOCL_LOGIN_WAIT_SECONDS,
            success_url_keywords=success_url_keywords,
        )
        if state == "APP":
            print("[OK] OOCL manual login xong; đã vào lại trang check.")
            return True

        print(f"[ERROR] OOCL chưa xác nhận đăng nhập sau ENTER (state={state}).")
        return False

    def _get_espot_result_state(self):
        """Đọc trạng thái kết quả E-Spot để tránh dùng lại DOM của tuyến trước."""
        try:
            return self.driver.execute_script("""
                const visible = (el) => {
                    if (!el) return false;
                    const st = window.getComputedStyle(el);
                    return st && st.display !== 'none' && st.visibility !== 'hidden'
                        && el.getClientRects && el.getClientRects().length > 0;
                };
                const dates = Array.from(document.querySelectorAll(
                    'div.date-selector div.date-row div.date-item'
                )).filter(visible);
                const dateText = dates.map(el => (el.innerText || '')
                    .replace(/\\s+/g, ' ').trim()).filter(Boolean).join('|');
                const resultEls = Array.from(document.querySelectorAll(
                    'div.search-result-card, div.rate-card, div.schedule-card, '
                    + 'div[class*="result-card"], div[class*="schedule-card"]'
                )).filter(visible);
                const resultText = resultEls.map(el => (el.innerText || '')
                    .replace(/\\s+/g, ' ').trim()).filter(Boolean).join('|').slice(0, 5000);
                const emptyEls = Array.from(document.querySelectorAll(
                    'div.empty-spot, div.no-result, div[class*="no-result"], '
                    + 'div[class*="noResult"]'
                )).filter(visible);
                const loading = Array.from(document.querySelectorAll(
                    '.ant-spin-spinning, [class*="loading"], [class*="Loading"], '
                    + '[aria-busy="true"]'
                )).some(visible);
                const signature = [dateText, resultText].filter(Boolean).join('||');
                return {
                    dateCount: dates.length,
                    signature: signature.slice(0, 8000),
                    empty: emptyEls.length > 0,
                    loading: !!loading
                };
            """) or {}
        except Exception:
            return {}

    def _get_equote_result_state(self):
        try:
            return self.driver.execute_script("""
                const visible = (el) => {
                    if (!el) return false;
                    const st = window.getComputedStyle(el);
                    return st && st.display !== 'none' && st.visibility !== 'hidden'
                        && el.getClientRects && el.getClientRects().length > 0;
                };
                const rows = Array.from(document.querySelectorAll(
                    'div.ant-table-body table tbody tr.ant-table-row.ant-table-row-level-0, tr.ant-table-row'
                )).filter(visible);
                const rowText = rows.map(r => (r.innerText || '').replace(/\\s+/g, ' ').trim())
                    .filter(Boolean)
                    .join('|')
                    .slice(0, 5000);
                const emptyEls = Array.from(document.querySelectorAll(
                    'div.ant-empty, div[class*="empty"], div[class*="no-data"], div[class*="noData"]'
                )).filter(visible);
                const bodyText = ((document.body && document.body.innerText) || '').toLowerCase();
                const emptyByText = [
                    'no data', 'no result', 'no records', 'no quotation',
                    'no matching', 'not found', 'no available'
                ].some(t => bodyText.includes(t));
                const loading = Array.from(document.querySelectorAll(
                    '.ant-spin-spinning, .ant-table-placeholder .ant-spin, [class*="loading"], [class*="Loading"]'
                )).some(visible);
                return {
                    rowCount: rows.length,
                    signature: rowText || '',
                    empty: emptyEls.length > 0 || emptyByText,
                    loading: !!loading
                };
            """) or {}
        except Exception:
            return {}

    def _wait_for_results(self, tab_type="espot", timeout=12, previous_signature=None):
        """
        Chờ kết quả load xong thay vì sleep cứng.
        tab_type: 'espot' hoặc 'equote'
        """
        try:
            if tab_type == "espot":
                # Smart Navigate giữ nguyên URL search-result và DOM cũ trong lúc React
                # tải tuyến mới. Không được xem URL/card cũ là kết quả vừa load xong.
                started = time.time()
                last_sig = previous_signature or ""
                sig_since = started

                def _espot_ready(d):
                    nonlocal last_sig, sig_since
                    state = self._get_espot_result_state()
                    if state.get("loading"):
                        return False
                    elapsed = time.time() - started
                    sig = state.get("signature") or ""
                    if sig != last_sig:
                        last_sig = sig
                        sig_since = time.time()
                    if previous_signature is not None:
                        # Khi bấm Search, React xóa lịch cũ trước rồi mới render lịch
                        # mới. Trạng thái rỗng trung gian không phải là kết quả mới.
                        if (sig and sig != previous_signature and elapsed >= 0.8
                                and time.time() - sig_since >= 0.5):
                            return True
                        if elapsed < OOCL_ESPOT_SAME_SIGNATURE_GRACE_SECONDS:
                            return False
                        if state.get("empty"):
                            return True
                        # Có tuyến có lịch/giá giống hệt tuyến trước.
                        return bool(sig) and sig == previous_signature
                    if sig:
                        return elapsed >= 0.25
                    return state.get("empty") and elapsed >= 0.8

                WebDriverWait(self.driver, timeout).until(_espot_ready)
                time.sleep(0.2)
                print("[OK] E-Spot: kết quả đã load.")
            else:
                timeout = timeout or OOCL_EQUOTE_RESULT_WAIT_SECONDS
                started = time.time()

                def _equote_ready(_d):
                    state = self._get_equote_result_state()
                    if state.get("loading"):
                        return False
                    if state.get("empty"):
                        return True
                    sig = state.get("signature") or ""
                    if sig:
                        if not previous_signature or sig != previous_signature:
                            return True
                        if time.time() - started >= OOCL_EQUOTE_SAME_SIGNATURE_GRACE_SECONDS:
                            return True
                    return False

                WebDriverWait(self.driver, timeout).until(_equote_ready)
                time.sleep(0.25)
                print("[OK] E-Quote: kết quả đã load.")
            return True
        except TimeoutException:
            print(f"[WARN] _wait_for_results ({tab_type}): timeout {timeout}s, tiếp tục.")
            return False
        except Exception as e:
            print(f"[WARN] _wait_for_results: {e}")
            return False

    def _navigate_next_espot(self, pol, pod, country,
                              prev_pol, prev_pod,
                              pol_country, pod_country):
        """
        Smart navigate: dùng lại trang search-result, chỉ sửa field thay đổi.
        Trả về True nếu thành công, False nếu cần fallback reload.
        """
        PANEL_XPATH  = "/html/body/div[1]/div/div[2]/div[2]/div[1]"
        PANEL_TITLE  = "/html/body/div[1]/div/div[2]/div[2]/div[1]/div[1]"
        POL_INPUT    = "/html/body/div[1]/div/div[2]/div[2]/div[1]/div[2]/div[1]/div[1]/div/span/input"
        POD_INPUT    = "/html/body/div[1]/div/div[2]/div[2]/div[1]/div[2]/div[1]/div[2]/div/span/input"
        POL_CLEAR    = "/html/body/div[1]/div/div[2]/div[2]/div[1]/div[2]/div[1]/div[1]/div/span/span[2]/span"
        POD_CLEAR    = "/html/body/div[1]/div/div[2]/div[2]/div[1]/div[2]/div[1]/div[2]/div/span/span[2]/span"
        POL_POPUP    = "/html/body/div[6]/div/div/div/div[2]/div"
        POD_POPUP    = "/html/body/div[7]/div/div/div/div[2]/div"
        # Dùng CSS class thay vì XPath index cứng cho Search button
        SEARCH_BTN_XPATH = "//div[contains(@class,'search-button')]//button[.//span[text()='Search']]"

        try:
            cur_url = self.driver.current_url or ""
            if "search-result" not in cur_url:
                return False

            # ── Bước 1: Kiểm tra panel đóng/mở ──
            panel = WebDriverWait(self.driver, 6).until(
                EC.presence_of_element_located((By.XPATH, PANEL_XPATH)))
            classes = panel.get_attribute("class") or ""
            is_collapsed = "collapsed" in classes

            if is_collapsed:
                print("[DEBUG] Searching Criteria đang đóng → click mở.")
                # Click vào title div (chính xác hơn click cả panel)
                try:
                    title_el = self.driver.find_element(By.XPATH, PANEL_TITLE)
                    self.driver.execute_script("arguments[0].click();", title_el)
                except Exception:
                    self.driver.execute_script("arguments[0].click();", panel)
                # Chờ fixed time thay vì check class (tránh StaleElement)
                time.sleep(1.5)
                print("[OK] Panel click sent.")
            else:
                print("[DEBUG] Searching Criteria đang mở sẵn.")

            time.sleep(OOCL_FOCUS_SLEEP)

            # Kiểm tra xem content của panel có visible không
            try:
                WebDriverWait(self.driver, 4).until(
                    EC.visibility_of_element_located((By.XPATH, POL_INPUT)))
            except TimeoutException:
                print("[WARN] Panel content chưa visible sau click.")
                return False

            pol_changed = (pol != prev_pol)
            pod_changed = (pod != prev_pod)

            # ── Bước 2: Sửa POL nếu thay đổi ──
            if pol_changed:
                print(f"[INFO] POL thay đổi: {prev_pol} → {pol}, xóa và nhập lại.")
                try:
                    clear_btn = self.driver.find_element(By.XPATH, POL_CLEAR)
                    self.driver.execute_script("arguments[0].click();", clear_btn)
                    time.sleep(0.4)
                except Exception:
                    pass
                pol_ok = self._enter_port_result_page(
                    POL_INPUT, POL_POPUP, pol, pol_country, "POL (result)")
                if pol_ok is False:
                    print("[WARN] POL (result): chọn port thất bại → fallback reload.")
                    return False
            else:
                print(f"[INFO] POL không đổi ({pol}), bỏ qua.")

            # ── Bước 3: Sửa POD nếu thay đổi ──
            if pod_changed:
                print(f"[INFO] POD thay đổi: {prev_pod} → {pod}, xóa và nhập lại.")
                try:
                    clear_btn = self.driver.find_element(By.XPATH, POD_CLEAR)
                    self.driver.execute_script("arguments[0].click();", clear_btn)
                    time.sleep(0.4)
                except Exception:
                    pass
                pod_ok = self._enter_port_result_page(
                    POD_INPUT, POD_POPUP, pod, pod_country, "POD (result)")
                if pod_ok is False:
                    print("[WARN] POD (result): chọn port thất bại → fallback reload.")
                    return False
            else:
                print(f"[INFO] POD không đổi ({pod}), bỏ qua.")

            # ── Bước 4: Click Search ──
            search_btn = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, SEARCH_BTN_XPATH)))
            self.driver.execute_script("arguments[0].click();", search_btn)
            print("[OK] Smart navigate: đã nhấn Search.")
            return True

        except Exception as e:
            print(f"[WARN] _navigate_next_espot thất bại: {e} → fallback reload.")
            return False

    def _enter_port_result_page(self, input_xpath, popup_xpath,
                                 port_name, country, label):
        """Nhập cảng vào Searching Criteria panel và chọn từ popup."""
        clean         = port_name.split(',')[0].strip().upper()
        country_upper = country.upper()
        port_code     = PORT_COUNTRY_CODE.get((clean, country_upper), "")
        country_search = PORT_COUNTRY_SEARCH_QUERY.get((clean, country_upper), f"{clean} {country_upper}".strip())

        try:
            inp = WebDriverWait(self.driver, 6).until(
                EC.element_to_be_clickable((By.XPATH, input_xpath)))
        except TimeoutException:
            print(f"[WARN] {label}: không tìm thấy input.")
            return

        self.driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", inp)
        safe_click(self.driver, inp, delay=0.1)
        inp.send_keys(Keys.CONTROL + "a")
        inp.send_keys(Keys.DELETE)
        fast_input(self.driver, inp, clean)   # ← thay send_keys(clean)

        # Chờ dropdown với items match port name (không chỉ chờ items bất kỳ — tránh stale)
        def _dropdown_has_match(d):
            items = d.find_elements(By.CSS_SELECTOR, "div.location-item div.location-text")
            for it in items:
                try:
                    if clean.lower() in it.text.strip().lower():
                        return True
                except Exception:
                    pass
            return False

        dropdown_matched = smart_wait(self.driver, _dropdown_has_match,
            timeout=8, poll=0.15, after=0.15, label=f"{label} dropdown match")

        if not dropdown_matched:
            print(f"[WARN] {label}: dropdown chưa có '{clean}' sau 8s, thử nhập lại...")
            safe_click(self.driver, inp, delay=0.2)
            inp.send_keys(Keys.CONTROL + "a")
            inp.send_keys(Keys.DELETE)
            time.sleep(0.3)
            fast_input(self.driver, inp, clean)
            dropdown_matched = smart_wait(self.driver, _dropdown_has_match,
                timeout=8, poll=0.15, after=0.15, label=f"{label} dropdown match retry")
            if not dropdown_matched:
                print(f"[ERROR] {label}: dropdown vẫn không có '{clean}' sau retry.")
                return False

        js = """
        let portName    = arguments[0].toUpperCase();
        let countryName = arguments[1].toUpperCase();
        let xpaths      = arguments[2];

        // Tìm trong các popup XPath
        for (let xp of xpaths) {
            let res = document.evaluate(
                xp, document, null,
                XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            let container = res.singleNodeValue;
            if (!container) continue;
            let items = container.querySelectorAll('div.location-item');
            if (!items || items.length === 0) continue;
            let fallback = null;
            for (let item of items) {
                let textEl = item.querySelector('div.location-text');
                if (!textEl) continue;
                let raw  = textEl.textContent.trim();
                let port = raw.split(',')[0].trim().toUpperCase();
                if (port === portName || port.includes(portName)) {
                    if (countryName && raw.toUpperCase().includes(countryName)) {
                        item.click();
                        return {clicked:true, matched:raw, xpath:xp};
                    }
                    if (!fallback) fallback = {el:item, text:raw};
                }
            }
            // Virtual list: scroll dropdown xuống để render thêm items
            if (countryName && fallback) {
                let scrollHolder = container.querySelector('.rc-virtual-list-holder');
                if (!scrollHolder) scrollHolder = container.querySelector('[style*="overflow"]');
                if (scrollHolder) {
                    let scrollStep = 150;
                    let maxScroll = scrollHolder.scrollHeight;
                    for (let pos = scrollStep; pos <= maxScroll; pos += scrollStep) {
                        scrollHolder.scrollTop = pos;
                        let t0 = Date.now(); while (Date.now() - t0 < 50) {}
                        let newItems = container.querySelectorAll('div.location-item');
                        for (let ni of newItems) {
                            let textEl = ni.querySelector('div.location-text');
                            if (!textEl) continue;
                            let raw = textEl.textContent.trim();
                            let port = raw.split(',')[0].trim().toUpperCase();
                            if ((port === portName || port.includes(portName))
                                && raw.toUpperCase().includes(countryName)) {
                                ni.click();
                                return {clicked:true, matched:raw, xpath:xp};
                            }
                        }
                    }
                }
            }
            if (fallback) {
                if (countryName) {
                    return {clicked:false, reason:'port_only_no_country', matched:fallback.text};
                }
                fallback.el.click();
                return {clicked:true, matched:fallback.text, xpath:xp};
            }
        }
        // Fallback: tìm trong ant-popover, ant-select-dropdown
        let drops = Array.from(document.querySelectorAll(
            'div.ant-popover, div[class*="popover"], div.ant-select-dropdown')).filter(d => {
            let st = window.getComputedStyle(d);
            return st.display !== 'none' && d.offsetParent !== null;
        });
        for (let drop of drops) {
            let items = drop.querySelectorAll('div.location-item');
            let fallback = null;
            for (let item of items) {
                let textEl = item.querySelector('div.location-text');
                if (!textEl) continue;
                let raw  = textEl.textContent.trim();
                let port = raw.split(',')[0].trim().toUpperCase();
                if (port === portName || port.includes(portName)) {
                    if (countryName && raw.toUpperCase().includes(countryName)) {
                        item.click();
                        return {clicked:true, matched:raw, xpath:'fallback-dropdown'};
                    }
                    if (!fallback) fallback = {el:item, text:raw};
                }
            }
            if (fallback) {
                if (countryName) {
                    return {clicked:false, reason:'port_only_no_country', matched:fallback.text};
                }
                fallback.el.click();
                return {clicked:true, matched:fallback.text, xpath:'fallback-dropdown'};
            }
        }
        return {clicked:false};
        """

        popup_xpaths = [
            popup_xpath,
            *[f"/html/body/div[{i}]/div/div/div/div[2]/div" for i in range(5, 12)],
        ]

        clicked = False
        country_retry_done = False
        for attempt in range(12):
            try:
                res = self.driver.execute_script(js, clean, country_upper, popup_xpaths)
                if res and res.get("clicked"):
                    print(f"[OK] {label} chọn: '{res.get('matched')}' "
                          f"via {res.get('xpath')} | attempt {attempt+1}")
                    clicked = True
                    break
                elif res and res.get("reason") == 'port_only_no_country' and not country_retry_done:
                    country_retry_done = True
                    print(f"[WARN] {label}: tìm thấy '{res.get('matched')}' nhưng country không khớp, thử search '{country_search}'...")
                    try:
                        inp2 = self.driver.find_element(By.XPATH, input_xpath)
                        safe_click(self.driver, inp2, delay=0.1)
                        inp2.send_keys(Keys.CONTROL + "a")
                        inp2.send_keys(Keys.DELETE)
                        time.sleep(0.3)
                        fast_input(self.driver, inp2, country_search)
                    except: pass
                    smart_wait(self.driver,
                        lambda d: any(clean in it.text.upper() and country_upper in it.text.upper()
                                      for it in d.find_elements(By.CSS_SELECTOR, "div.location-item")),
                        timeout=6, poll=0.1, after=0.15, label=f"{label} dropdown+country")
                    res2 = self.driver.execute_script(js, clean, country_upper, popup_xpaths)
                    if res2 and res2.get("clicked"):
                        print(f"[OK] {label} chọn: '{res2.get('matched')}' | attempt {attempt+1} (retry+country)")
                        clicked = True
                        break
                    if port_code:
                        print(f"[WARN] {label}: thử search bằng port code '{port_code}' cho {clean}/{country_upper}...")
                        try:
                            inp3 = self.driver.find_element(By.XPATH, input_xpath)
                            safe_click(self.driver, inp3, delay=0.1)
                            inp3.send_keys(Keys.CONTROL + "a")
                            inp3.send_keys(Keys.DELETE)
                            time.sleep(0.3)
                            fast_input(self.driver, inp3, port_code)
                            smart_wait(self.driver,
                                lambda d: any(clean in it.text.upper() and country_upper in it.text.upper()
                                              for it in d.find_elements(By.CSS_SELECTOR, "div.location-item")),
                                timeout=6, poll=0.1, after=0.15, label=f"{label} dropdown+code")
                            res3 = self.driver.execute_script(js, clean, country_upper, popup_xpaths)
                            if res3 and res3.get("clicked"):
                                print(f"[OK] {label} chọn: '{res3.get('matched')}' | attempt {attempt+1} (retry+code {port_code})")
                                clicked = True
                                break
                        except Exception as e:
                            debug_print(f"[DEBUG] {label} retry code {port_code}: {e}")
            except Exception as e:
                debug_print(f"[DEBUG] {label} JS attempt {attempt+1}: {e}")
            time.sleep(0.5)

        if not clicked:
            print(f"[WARN] {label}: không tìm thấy option sau 12 lần thử.")
            return False

        # Chờ dropdown đóng
        time.sleep(0.15)
        return True

    def _get_espot_calendar_dates(self):
        dates = []
        try:
            items = self.driver.find_elements(
                By.CSS_SELECTOR, "div.date-selector div.date-item")
            for item in items:
                try:
                    txt = item.find_element(
                        By.CSS_SELECTOR, "div.date-text").text.strip()
                    dates.append(datetime.strptime(txt, "%Y-%m-%d").date())
                except Exception:
                    continue
        except Exception:
            pass
        return dates

    def _click_espot_week_in_popover(self, target_date):
        """
        OOCL UI mới: bấm mũi tên tuần chỉ mở lịch tháng.
        Click ngày đầu tuần trong popover để đổi thanh 7 ngày.
        """
        target_str = target_date.strftime("%Y-%m-%d")

        def _visible_month_popover(d):
            for pop in d.find_elements(By.CSS_SELECTOR, ".ant-popover"):
                try:
                    if pop.is_displayed() and pop.find_elements(By.CSS_SELECTOR, ".month-calendar-body"):
                        return pop
                except Exception:
                    continue
            return None

        try:
            pop = WebDriverWait(self.driver, 4).until(_visible_month_popover)
            for _ in range(4):
                target_cells = pop.find_elements(
                    By.CSS_SELECTOR, f"td[title='{target_str}']")
                target_cell = next((c for c in target_cells if c.is_displayed()), None)
                if target_cell:
                    classes = target_cell.get_attribute("class") or ""
                    if "disabled" in classes:
                        print(f"[WARN] Week picker: ngày {target_str} bị disabled.")
                        return False
                    self.driver.execute_script("arguments[0].click();", target_cell)
                    WebDriverWait(self.driver, 5).until(
                        lambda d: target_date in self._get_espot_calendar_dates())
                    print(f"[OK] Week picker: đã chọn tuần bắt đầu {target_str}.")
                    return True

                shown = []
                for td in pop.find_elements(By.CSS_SELECTOR, "td[title]"):
                    try:
                        shown.append(datetime.strptime(
                            td.get_attribute("title"), "%Y-%m-%d").date())
                    except Exception:
                        pass
                if not shown:
                    break

                direction = "right" if target_date > max(shown) else "left"
                nav_btn = None
                for btn in pop.find_elements(By.CSS_SELECTOR, "button.month-nav-button"):
                    try:
                        if btn.find_elements(By.CSS_SELECTOR, f"span[aria-label='{direction}']"):
                            nav_btn = btn
                            break
                    except Exception:
                        continue
                if not nav_btn or nav_btn.get_attribute("disabled"):
                    break
                self.driver.execute_script("arguments[0].click();", nav_btn)
                time.sleep(0.4)
                pop = WebDriverWait(self.driver, 4).until(_visible_month_popover)
        except Exception as e:
            print(f"[WARN] Week picker chọn {target_str} thất bại: {e}")
        return False

    def _move_espot_calendar_week(self, direction="next"):
        """
        Chuyển E-Spot calendar sang tuần trước/sau.
        Hỗ trợ UI cũ lẫn UI mới có bảng chọn tuần sau khi bấm mũi tên.
        """
        before = self._get_espot_calendar_dates()
        if not before:
            print("[WARN] Không đọc được tuần hiện tại của E-Spot calendar.")
            return False

        target = (max(before) + timedelta(days=1)
                  if direction == "next"
                  else min(before) - timedelta(days=7))
        btn_index = 1 if direction == "next" else 0

        try:
            buttons = WebDriverWait(self.driver, 4).until(
                lambda d: d.find_elements(By.CSS_SELECTOR, "div.date-selector button.nav-button"))
            if len(buttons) <= btn_index:
                print(f"[WARN] Không thấy nút {direction} tuần E-Spot.")
                return False
            btn = buttons[btn_index]
            if btn.get_attribute("disabled"):
                print(f"[WARN] Nút {direction} tuần E-Spot đang disabled.")
                return False
            self.driver.execute_script("arguments[0].click();", btn)
        except Exception as e:
            print(f"[WARN] Click nút {direction} tuần E-Spot thất bại: {e}")
            return False

        try:
            WebDriverWait(self.driver, 1.2).until(
                lambda d: self._get_espot_calendar_dates()
                and self._get_espot_calendar_dates() != before)
            print(f"[OK] Calendar đã chuyển {direction} bằng nút mũi tên.")
            return True
        except Exception:
            pass

        return self._click_espot_week_in_popover(target)

    def __init__(self):
        self.driver = None
        self.wait   = None
        self.edge_user_data_dir = EDGE_USER_DATA_DIR

    def _is_debug_endpoint_ready(self):
        try:
            with urllib.request.urlopen(
                f"http://127.0.0.1:{EDGE_DEBUG_PORT}/json/version", timeout=2
            ) as resp:
                body = (resp.read() or b"").decode("utf-8", errors="ignore")
                return "webSocketDebuggerUrl" in body
        except Exception:
            return False

    def launch_edge_if_needed(self):
        if is_port_open(EDGE_DEBUG_PORT) and self._is_debug_endpoint_ready():
            print(f"[OK] Edge đã mở sẵn trên port {EDGE_DEBUG_PORT}, bỏ qua bước mở Edge.")
            return True
        if is_port_open(EDGE_DEBUG_PORT) and not self._is_debug_endpoint_ready():
            print("[WARN] Port mở nhưng endpoint lỗi, chờ thêm 3s...")
            time.sleep(3)
            if self._is_debug_endpoint_ready():
                print("[OK] Edge endpoint ổn sau khi chờ.")
                return True
        print("[INFO] Mở Edge mới với remote-debugging-port...")
        try:
            subprocess.Popen([
                EDGE_EXE,
                f"--remote-debugging-port={EDGE_DEBUG_PORT}",
                f"--user-data-dir={self.edge_user_data_dir}",
                "--new-window",
                "--no-first-run",
                "--no-default-browser-check",
                "--disable-background-timer-throttling",
                "--disable-renderer-backgrounding",
                "--disable-backgrounding-occluded-windows",
                OOCL_LOGIN_URL,
            ], stdin=subprocess.DEVNULL, stdout=subprocess.DEVNULL,
               stderr=subprocess.DEVNULL)
            for i in range(40):
                time.sleep(0.5)
                if is_port_open(EDGE_DEBUG_PORT) and self._is_debug_endpoint_ready():
                    print(f"[OK] Edge sẵn sau {(i+1)*0.5:.1f}s.")
                    return True
            print("[ERROR] Edge không khởi động được sau 20s.")
            return False
        except FileNotFoundError:
            print(f"[ERROR] Không thấy Edge tại: {EDGE_EXE}")
            return False

    def init_browser(self):
        opts = Options()
        opts.use_chromium = True
        opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{EDGE_DEBUG_PORT}")
        for attempt in range(1, 4):
            try:
                # Wrap webdriver.Edge() trong thread timeout để tránh treo vô hạn
                _driver_holder = {"driver": None, "error": None}
                def _create_session():
                    try:
                        _driver_holder["driver"] = webdriver.Edge(
                            service=Service(executable_path=DRIVER_PATH),
                            options=opts,
                        )
                    except Exception as ex:
                        _driver_holder["error"] = ex
                _t = threading.Thread(target=_create_session, daemon=True)
                _t.start()
                _t.join(timeout=30)
                if _t.is_alive() or _driver_holder["driver"] is None:
                    if _driver_holder["error"]:
                        raise _driver_holder["error"]
                    raise TimeoutException(f"webdriver.Edge() treo quá 30s (attempt {attempt})")
                self.driver = _driver_holder["driver"]
                self.driver.set_page_load_timeout(30)
                try:
                    self.driver.maximize_window()
                except Exception:
                    try:
                        info = self.driver.execute_cdp_cmd("Browser.getWindowForTarget", {})
                        self.driver.execute_cdp_cmd("Browser.setWindowBounds", {
                            "windowId": info["windowId"],
                            "bounds": {"windowState": "maximized"}
                        })
                    except Exception:
                        pass
                self.wait = WebDriverWait(self.driver, 20)
                time.sleep(OOCL_POST_CONNECT_SLEEP)
                _ = self.driver.current_window_handle
                print(f"[OK] Kết nối Selenium thành công (attempt {attempt}).")
                return True
            except Exception as e:
                print(f"[WARN] init_browser attempt {attempt}: {e}")
                try:
                    if self.driver: self.driver.quit()
                except Exception:
                    pass
                self.driver = None
                self.wait = None
                time.sleep(1)
        print("[ERROR] Không kết nối được Selenium sau 3 lần thử.")
        return False

    def close_all_tabs_on_finish(self):
        """Close every tab in OOCL's dedicated Edge profile after the run."""
        drv = getattr(self, "driver", None)
        result = {"closed": 0}

        def _close_tabs():
            try:
                handles = list(drv.window_handles)
                for handle in reversed(handles):
                    try:
                        if handle not in drv.window_handles:
                            continue
                        drv.switch_to.window(handle)
                        drv.close()
                        result["closed"] += 1
                    except Exception:
                        continue
            except Exception:
                pass
            finally:
                try:
                    drv.quit()
                except Exception:
                    pass

        worker = None
        if drv is not None:
            worker = threading.Thread(target=_close_tabs, daemon=True)
            worker.start()
            worker.join(timeout=5)

        # Selenium can hang while Edge remains alive. Force-stop only the
        # dedicated OOCL profile/port, plus this bot's own msedgedriver. This
        # also releases stdout inherited from main.py.
        profile = str(self.edge_user_data_dir or "").replace("'", "''")
        ps = f"""
$profile = '{profile}'.ToLower()
$port = '--remote-debugging-port={EDGE_DEBUG_PORT}'
Get-CimInstance Win32_Process -Filter "name = 'msedge.exe'" |
    Where-Object {{
        $cmd = ([string]$_.CommandLine).ToLower()
        ($profile -and $cmd.Contains($profile)) -or $cmd.Contains($port)
    }} |
    ForEach-Object {{ Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }}
Get-CimInstance Win32_Process -Filter "name = 'msedgedriver.exe'" |
    Where-Object {{ $_.ParentProcessId -eq {os.getpid()} }} |
    ForEach-Object {{ Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }}
"""
        kwargs = {
            "stdin": subprocess.DEVNULL,
            "stdout": subprocess.DEVNULL,
            "stderr": subprocess.DEVNULL,
            "timeout": 12,
        }
        if hasattr(subprocess, "CREATE_NO_WINDOW"):
            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        needs_force_close = bool(worker is not None and worker.is_alive()) or is_port_open(EDGE_DEBUG_PORT)
        if needs_force_close:
            try:
                subprocess.run(["powershell", "-NoProfile", "-Command", ps], **kwargs)
            except Exception as exc:
                print(f"[WARN] OOCL force-close Edge lỗi: {type(exc).__name__}")

        if worker is not None and worker.is_alive():
            worker.join(timeout=2)
        print(f"[OK] OOCL đã đóng browser sau khi hoàn tất ({result['closed']} tab qua Selenium).")

    def keep_browser_on_finish(self):
        """Detach Selenium while keeping the dedicated OOCL Edge/session open."""
        drv = getattr(self, "driver", None)
        if drv is not None:
            try:
                service = getattr(drv, "service", None)
                process = getattr(service, "process", None)
                if process is not None and process.poll() is None:
                    process.kill()
                    process.wait(timeout=5)
                if service is not None:
                    service.process = None
            except Exception as exc:
                print(f"[WARN] Không dừng được msedgedriver OOCL: {type(exc).__name__}")
        self.driver = None
        self.wait = None
        print("[OK] OOCL hoàn tất; giữ nguyên Edge và session đăng nhập cho lần chạy sau.")

    def _ensure_session_alive(self):
        try:
            _ = self.driver.current_window_handle
            _ = self.driver.current_url
            return True
        except Exception:
            print("[WARN] Session died. Thử reconnect...")
            return self.init_browser()

    def _find_authenticated_oocl_tab(self):
        """Return an existing signed-in OOCL app tab without forcing a new SSO."""
        try:
            handles = list(self.driver.window_handles)
            original = self.driver.current_window_handle
        except Exception:
            return None

        for handle in handles:
            try:
                self.driver.switch_to.window(handle)
                self._abort_if_oocl_blocked("scan authenticated tab")
                if self._is_oocl_app_ready():
                    return handle
            except OOCLIpBlockedError:
                raise
            except Exception:
                continue

        try:
            if original in self.driver.window_handles:
                self.driver.switch_to.window(original)
        except Exception:
            pass
        return None

    def _find_existing_manual_login_tab(self):
        """Return an already-open OTP/CAPTCHA tab without restarting SSO."""
        try:
            handles = list(self.driver.window_handles)
            original = self.driver.current_window_handle
        except Exception:
            return None

        for handle in handles:
            try:
                self.driver.switch_to.window(handle)
                self._abort_if_oocl_blocked("scan existing OTP tab")
                if self._oocl_manual_login_check_present():
                    return handle
            except OOCLIpBlockedError:
                raise
            except Exception:
                continue

        try:
            if original in self.driver.window_handles:
                self.driver.switch_to.window(original)
        except Exception:
            pass
        return None

    def clean_tabs_and_open_fresh(self):
        try:
            handles = list(self.driver.window_handles)
            print(f"[DEBUG] Số tab hiện có: {len(handles)}")
            for i, h in enumerate(handles):
                try:
                    self.driver.switch_to.window(h)
                    print(f"[DEBUG]   Tab[{i}]: {self.driver.current_url}")
                except Exception:
                    print(f"[DEBUG]   Tab[{i}]: (không đọc được URL)")

            keep_tab = handles[0]
            self.driver.switch_to.window(keep_tab)
            print(f"[INFO] Navigate tab đầu tiên về OOCL SSO token...")
            try:
                self.driver.get(OOCL_LOGIN_URL)
            except TimeoutException:
                print("[WARN] Timeout load OOCL SSO token, ép stop...")
                try: self.driver.execute_script("window.stop();")
                except Exception: pass
            time.sleep(OOCL_TAB_NAV_SLEEP)
            self._abort_if_oocl_blocked("clean_tabs_and_open_fresh")
            print(f"[OK] Tab giữ lại: {self.driver.current_url}")

            closed = 0
            for h in reversed(handles[1:]):
                try:
                    if h not in self.driver.window_handles:
                        continue
                    self.driver.switch_to.window(h)
                    cur = self.driver.current_url or ""
                    # Bỏ qua tab NTP
                    if "ntp.msn.com" in cur:
                        print(f"[DEBUG] Bỏ qua tab NTP {h[-6:]}")
                        continue
                    # Set timeout ngắn 3s để không bị treo 20s
                    self.driver.set_page_load_timeout(3)
                    try:
                        self.driver.close()
                        time.sleep(OOCL_TAB_CLOSE_SLEEP)
                        closed += 1
                    except Exception:
                        print(f"[WARN] Bỏ qua tab {h[-6:]} không đóng được.")
                    finally:
                        self.driver.set_page_load_timeout(30)
                except Exception as e:
                    print(f"[WARN] Xử lý tab {h[-6:]}: {e}")
                finally:
                    try:
                        if keep_tab in self.driver.window_handles:
                            self.driver.switch_to.window(keep_tab)
                    except Exception:
                        pass

            remaining = list(self.driver.window_handles)
            self.driver.switch_to.window(remaining[0])
            self._focus_tab(remaining[0])
            print(f"[OK] Đã đóng {closed} tab, còn {len(remaining)} tab.")
            print(f"[OK] URL hiện tại: {self.driver.current_url}")
            return True

        except Exception as e:
            print(f"[ERROR] clean_tabs_and_open_fresh: {e}")
            return False
        
    def _focus_tab(self, handle):
        try:
            self.driver.switch_to.window(handle)
            time.sleep(0.2)
            try:
                import ctypes
                import ctypes.wintypes
                def enum_windows_callback(hwnd, results):
                    title = ctypes.create_unicode_buffer(256)
                    ctypes.windll.user32.GetWindowTextW(hwnd, title, 256)
                    if "FreightSmart" in title.value or "OOCL" in title.value or "Edge" in title.value:
                        results.append(hwnd)
                    return True
                results = []
                WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.py_object)
                ctypes.windll.user32.EnumWindows(WNDENUMPROC(enum_windows_callback), results)
                if results:
                    ctypes.windll.user32.SetForegroundWindow(results[0])
                    ctypes.windll.user32.ShowWindow(results[0], 9)
            except Exception:
                pass
            self.driver.execute_script("window.focus();")
            time.sleep(0.3)
        except Exception as e:
            print(f"[WARN] _focus_tab: {e}")

    def _activate_tab(self, handle):
        """
        Switch Selenium context + ép browser coi tab là foreground (visible).
        Fix vấn đề React không render khi tab chạy ngầm.
        """
        try:
            self.driver.switch_to.window(handle)
            time.sleep(0.3)

            # Ép Edge coi tab này là visible
            self.driver.execute_script("""
                try {
                    Object.defineProperty(document, 'hidden', {
                        value: false, configurable: true, writable: true});
                    Object.defineProperty(document, 'visibilityState', {
                        value: 'visible', configurable: true, writable: true});
                    document.dispatchEvent(new Event('visibilitychange', {bubbles: true}));
                    window.dispatchEvent(new Event('focus', {bubbles: true}));
                } catch(e) {}
            """)
            time.sleep(0.2)

            # Bring Edge window lên foreground
            try:
                import ctypes, ctypes.wintypes
                def _enum_cb(hwnd, out):
                    buf = ctypes.create_unicode_buffer(256)
                    ctypes.windll.user32.GetWindowTextW(hwnd, buf, 256)
                    if any(k in buf.value for k in ("FreightSmart","OOCL","Edge")):
                        out.append(hwnd)
                    return True
                found = []
                PROC = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.py_object)
                ctypes.windll.user32.EnumWindows(PROC(_enum_cb), found)
                if found:
                    ctypes.windll.user32.ShowWindow(found[0], 9)   # SW_RESTORE
                    ctypes.windll.user32.SetForegroundWindow(found[0])
            except Exception:
                pass

            self.driver.execute_script("window.focus();")
            time.sleep(0.3)
        except Exception as e:
            print(f"[WARN] _activate_tab: {e}")

    def _switch_to_tab_by_keyboard(self, handle):
        try:
            handle = str(handle)
            all_handles = [str(h) for h in self.driver.window_handles]

            if handle not in all_handles:
                print(f"[WARN] _switch_to_tab_by_keyboard: "
                      f"handle {handle[-6:]} không tồn tại, bỏ qua.")
                return

            # Dùng Selenium switch — không dùng pyautogui, không ảnh hưởng window khác
            self.driver.switch_to.window(handle)
            self.driver.execute_script("""
                try {
                    Object.defineProperty(document,'hidden',
                        {value:false,configurable:true,writable:true});
                    Object.defineProperty(document,'visibilityState',
                        {value:'visible',configurable:true,writable:true});
                    document.dispatchEvent(
                        new Event('visibilitychange',{bubbles:true}));
                    window.dispatchEvent(new Event('focus',{bubbles:true}));
                } catch(e) {}
            """)
            tab_idx = all_handles.index(handle) + 1
            print(f"[DEBUG] Switch → tab {tab_idx} (handle {handle[-6:]})")

        except Exception as e:
            print(f"[WARN] _switch_to_tab_by_keyboard: {e}")

    def _find_first_xpath(self, xpaths, timeout=8, clickable=False):
        candidates = xpaths if isinstance(xpaths, (list, tuple)) else [xpaths]
        end = time.time() + timeout
        while time.time() < end:
            for xp in candidates:
                try:
                    cond = EC.element_to_be_clickable if clickable else EC.presence_of_element_located
                    return WebDriverWait(self.driver, 1).until(cond((By.XPATH, xp)))
                except TimeoutException:
                    continue
                except StaleElementReferenceException:
                    continue
            time.sleep(0.15)
        return None

    def _type_to_xpath(self, xpaths, value, timeout=8, use_ctrl_a=False):
        end = time.time() + timeout
        while time.time() < end:
            el = self._find_first_xpath(xpaths, timeout=1.2, clickable=True)
            if not el:
                continue
            try:
                el.click(); r()
                if use_ctrl_a:
                    el.send_keys(Keys.CONTROL + "a"); r()
                    el.send_keys(Keys.DELETE); r()
                else:
                    el.clear()
                el.send_keys(value); r()
                return True
            except StaleElementReferenceException:
                continue
            except Exception:
                continue
        return False

    def check_and_login(self):
        for attempt in range(3):
            try:
                if not self._ensure_session_alive():
                    print(f"[ERROR] Session không khả dụng (attempt {attempt+1})")
                    continue
                self._abort_if_oocl_blocked("before login")
                authenticated_tab = self._find_authenticated_oocl_tab()
                if authenticated_tab:
                    self._focus_tab(authenticated_tab)
                    print("[OK] Tái sử dụng session OOCL đã đăng nhập; bỏ qua SSO/OTP.")
                    return True
                manual_tab = self._find_existing_manual_login_tab()
                if manual_tab:
                    self._focus_tab(manual_tab)
                    print("[INFO] OOCL đang có sẵn màn OTP/CAPTCHA; giữ nguyên mã hiện tại.")
                    return self._wait_manual_login_done("MFA/CAPTCHA")
                if not self.clean_tabs_and_open_fresh():
                    print(f"[WARN] clean_tabs thất bại (attempt {attempt+1})")
                    continue
                self._abort_if_oocl_blocked("after open OOCL")

                print("[INFO] Chờ OOCL SSO token chuyển tới form login/app...")
                state = self._wait_for_oocl_login_or_app(timeout=OOCL_LOGIN_WAIT_SECONDS)
                if state == "APP":
                    print("[OK] Đã đăng nhập OOCL sẵn.")
                    return True

                if state == "TIMEOUT":
                    print("[INFO] SSO token chưa ra form rõ ràng, mở lại token login OOCL...")
                    try:
                        self.driver.get(OOCL_LOGIN_URL)
                    except TimeoutException:
                        try:
                            self.driver.execute_script("window.stop();")
                        except Exception:
                            pass
                    time.sleep(1.5)

                if not self._do_login():
                    continue
                self._abort_if_oocl_blocked("after _do_login")
                return True

            except OOCLIpBlockedError:
                raise
            except OOCLManualLoginTimeout:
                raise
            except Exception as e:
                print(f"[ERROR] check_and_login attempt {attempt+1}: {e}")
                time.sleep(1)

        print("[ERROR] check_and_login thất bại sau 3 lần.")
        return False

    def _do_login(self, success_url_keywords=OOCL_APP_URL_MARKERS):
        try:
            self._abort_if_oocl_blocked("do_login start")
            print("[INFO] Đang xử lý form đăng nhập OOCL...")
            self.driver.execute_script("window.focus();")

            PWD_XPATHS = [
                "//input[@id='password']",
                "//input[@type='password']",
                "//input[@name='password']",
            ]

            USER_XPATHS = [
                "//input[@id='email-input' or @type='email' or @name='email']",
                "//input[@id='username' or @name='username']",
                "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'email')]",
                "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'user id')]",
            ]

            if self._is_oocl_app_ready(success_url_keywords):
                print("[OK] OOCL đã ở trong app, bỏ qua login.")
                return True

            state = self._wait_for_oocl_login_or_app(
                timeout=OOCL_LOGIN_WAIT_SECONDS,
                success_url_keywords=success_url_keywords,
            )
            if state == "APP":
                print("[OK] Đăng nhập OOCL đã sẵn.")
                return True
            if state == "MANUAL":
                return self._wait_manual_login_done("MFA/CAPTCHA", success_url_keywords)
            if state == "TIMEOUT":
                print("[WARN] Chưa thấy form login rõ ràng, mở trực tiếp trang login OOCL...")
                try:
                    self.driver.get(OOCL_LOGIN_URL)
                except TimeoutException:
                    try:
                        self.driver.execute_script("window.stop();")
                    except Exception:
                        pass
                state = self._wait_for_oocl_login_or_app(
                    timeout=OOCL_LOGIN_WAIT_SECONDS,
                    success_url_keywords=success_url_keywords,
                )

            if state == "USER":
                user_input = self._find_visible_xpath(USER_XPATHS, timeout=8, clickable=True)
                if not user_input:
                    print("[ERROR] Không tìm thấy ô User ID/Email OOCL.")
                    return False

                current_user = (user_input.get_attribute("value") or "").strip()
                readonly = bool(user_input.get_attribute("readonly"))
                generic_readonly_user = readonly and current_user.upper() in {"FS", "OOCL"}

                if generic_readonly_user:
                    print(f"[INFO] Ô username OOCL đang readonly='{current_user}', giữ nguyên và chờ password.")
                else:
                    if not self._set_input_value(user_input, LOGIN_EMAIL, remove_readonly=readonly):
                        print("[ERROR] Không điền được User ID/Email OOCL.")
                        return False
                    print("[INFO] Đã điền User ID/Email OOCL.")
                    if not self._click_first_login_button(timeout=5):
                        try:
                            user_input.send_keys(Keys.ENTER)
                        except Exception:
                            pass

                time.sleep(1.5)
                self._abort_if_oocl_blocked("after login user/email")
                state = self._wait_for_oocl_login_or_app(
                    timeout=OOCL_LOGIN_WAIT_SECONDS,
                    success_url_keywords=success_url_keywords,
                )
                if state == "APP":
                    print("[OK] OOCL login xong sau bước User ID/Email.")
                    return True
                if state == "MANUAL":
                    return self._wait_manual_login_done("MFA/CAPTCHA", success_url_keywords)

            pwd = self._find_visible_xpath(PWD_XPATHS, timeout=12, clickable=True)
            if not pwd:
                if self._oocl_manual_login_check_present():
                    return self._wait_manual_login_done("MFA/CAPTCHA", success_url_keywords)
                print(f"[ERROR] Không tìm thấy password input. State={state}, URL={self.driver.current_url}")
                return False

            if not self._set_input_value(pwd, LOGIN_PASSWORD):
                print("[ERROR] Không điền được password OOCL.")
                return False
            print("[INFO] Đã điền password OOCL.")

            if not self._click_first_login_button(timeout=8):
                try:
                    pwd.send_keys(Keys.ENTER)
                except Exception:
                    pass
            time.sleep(0.5)
            self._abort_if_oocl_blocked("after login submit")

            end = time.time() + OOCL_LOGIN_WAIT_SECONDS
            last_log = time.time()
            while time.time() < end:
                self._abort_if_oocl_blocked("login wait done")
                if self._is_oocl_app_ready(success_url_keywords):
                    print("[OK] Đăng nhập thành công! URL:", self.driver.current_url)
                    return True
                if self._oocl_manual_login_check_present():
                    return self._wait_manual_login_done("MFA/CAPTCHA", success_url_keywords)
                try:
                    body = (self.driver.find_element(By.TAG_NAME, "body").text or "").lower()
                except Exception:
                    body = ""
                if any(msg in body for msg in (
                    "invalid username or password",
                    "incorrect password",
                    "could not find your oocl account",
                    "please enter an valid email",
                )):
                    print("[ERROR] OOCL báo lỗi credential/email trên màn login.")
                    return False
                if time.time() - last_log > 10:
                    print("[INFO] OOCL login: đang chờ redirect vào app...")
                    last_log = time.time()
                time.sleep(0.5)

            print(f"[ERROR] Timeout login. URL: {self.driver.current_url}")
            return False
        except TimeoutException:
            self._abort_if_oocl_blocked("login timeout")
            print(f"[ERROR] Timeout login. URL: {self.driver.current_url}")
            return False
        except OOCLManualLoginTimeout:
            # Manual OTP/CAPTCHA already owns its full 15-minute deadline.
            # Do not downgrade it to a normal login failure, otherwise callers
            # may restart SSO and create another unwanted OTP wait cycle.
            raise
        except OOCLIpBlockedError:
            raise
        except Exception as e:
            print(f"[ERROR] Login: {e}")
            return False


# ========================== COMBINED BOT ==========================

class OOCLCombinedBot(OOCLBaseScraper):

    def _stop_setup_page_load(self):
        """Stop the current page before retrying tab setup in the same Edge."""
        try:
            if self.driver is not None:
                self.driver.execute_script("window.stop();")
                time.sleep(0.5)
        except Exception as exc:
            print(f"[WARN] Không dừng được page load trước setup retry: {type(exc).__name__}")

    def _restart_oocl_edge_for_setup(self):
        """Restart only OOCL's dedicated Edge/profile, then restore login state."""
        print("[RECOVERY] Restart Edge OOCL để khôi phục setup E-Spot/E-Quote...")
        try:
            self.close_all_tabs_on_finish()
        except Exception as exc:
            print(f"[WARN] Restart OOCL: lỗi đóng Edge cũ: {type(exc).__name__}")
        finally:
            self.driver = None
            self.wait = None

        if not self.launch_edge_if_needed():
            print("[ERROR] Recovery OOCL: không mở lại được Edge.")
            return False
        if not self.init_browser():
            print("[ERROR] Recovery OOCL: không kết nối lại được Selenium.")
            return False
        try:
            if not self.check_and_login():
                print("[ERROR] Recovery OOCL: login lại thất bại.")
                return False
            return True
        except OOCLIpBlockedError:
            raise
        except Exception as exc:
            print(f"[ERROR] Recovery OOCL: login/session lỗi: {exc}")
            return False

    def setup_tabs_with_recovery(self):
        """
        Create usable E-Spot/E-Quote tabs with bounded recovery:
        one retry in the current Edge, then one dedicated Edge restart.
        """
        failures = []

        for attempt in (1, 2):
            try:
                if self.setup_tabs():
                    if attempt > 1:
                        print("[RECOVERY] Setup tabs thành công sau retry trong Edge hiện tại.")
                    return True
                failures.append(f"setup_tabs returned False (session attempt {attempt})")
            except OOCLIpBlockedError:
                raise
            except Exception as exc:
                failures.append(f"{type(exc).__name__}: {exc}")

            if attempt == 1:
                print("[RECOVERY] Setup tabs lỗi/timeout, dừng page và thử lại Edge hiện tại (1/1)...")
                self._stop_setup_page_load()

        print("[RECOVERY] Setup tabs vẫn lỗi; restart Edge OOCL và thử lần cuối (1/1)...")
        if self._restart_oocl_edge_for_setup():
            try:
                if self.setup_tabs():
                    print("[RECOVERY] Setup tabs thành công sau restart Edge OOCL.")
                    return True
                failures.append("setup_tabs returned False after Edge restart")
            except OOCLIpBlockedError:
                raise
            except Exception as exc:
                failures.append(f"{type(exc).__name__}: {exc}")
        else:
            failures.append("OOCL Edge restart/login recovery failed")

        raise OOCLSetupTabsError(
            "OOCL không tạo được E-Spot/E-Quote sau recovery: " + " | ".join(failures)
        )

    def _debug_sailing_card_html(self, card_el, idx):
        """Dùng 1 lần để xem cấu trúc HTML thật"""
        try:
            print(f"\n=== DEBUG SAILING CARD #{idx} ===")
            print("Inner HTML snippet:")
            print(card_el.get_attribute("innerHTML")[:2000])
            print("Full text:")
            print(repr(card_el.text))
            print("="*60)
        except:
            pass
        
    def _read_espot_charge_breakdown_usd(self):
        """
        Đọc popup E-Spot > tab Charge Breakdown
        Trả về surcharge USD cần cộng thêm:
        {"20GP": int, "40GP": int, "40HQ": int}
        Chỉ cộng các surcharge USD visible trong bảng Ocean Surcharges.
        """
        result = {"20GP": 0, "40GP": 0, "40HQ": 0}

        try:
            js = r"""
            function moneyToInt(txt) {
                txt = (txt || "").trim().toUpperCase();
                if (!txt.includes("USD")) return 0;
                const m = txt.match(/USD\s*([\d,\.]+)/i);
                if (!m) return 0;
                return Math.ceil(parseFloat(m[1].replace(/,/g, "")) || 0);
            }

            const out = {"20GP": 0, "40GP": 0, "40HQ": 0};

            // tìm title Ocean Surcharges
            const titles = Array.from(document.querySelectorAll("span.table-title"));
            let titleEl = null;
            for (const t of titles) {
                if ((t.textContent || "").trim() === "Ocean Surcharges") {
                    titleEl = t;
                    break;
                }
            }
            if (!titleEl) return out;

            // table-content chứa title này
            let wrap = titleEl.closest("div.table-content");
            if (!wrap) return out;

            // tbody rows. Bảng dùng rowspan trên 4 cột đầu, nên row tiếp theo
            // của cùng 1 phụ phí chỉ có 3 ô (unit, price, remarks). Không filter
            // theo length — quét theo nội dung ô để lấy đúng 20GP/40GP/40HQ.
            const rows = wrap.querySelectorAll("tbody.ant-table-tbody > tr.ant-table-row");
            for (const tr of rows) {
                const tds = tr.querySelectorAll("td.ant-table-cell");
                if (!tds || tds.length < 2) continue;

                let unitText = "";
                let priceText = "";

                for (let i = 0; i < tds.length - 1; i++) {
                    const txt = (tds[i].textContent || "").trim().toUpperCase();
                    if (txt === "20GP" || txt === "40GP" || txt === "40HQ") {
                        unitText = txt;
                        priceText = (tds[i + 1].textContent || "").trim();
                        break;
                    }
                }

                if (!unitText || !priceText) continue;

                const val = moneyToInt(priceText);
                if (!val) continue;

                if (unitText === "20GP") out["20GP"] += val;
                else if (unitText === "40GP") out["40GP"] += val;
                else if (unitText === "40HQ") out["40HQ"] += val;
            }

            return out;
            """
            result = self.driver.execute_script(js) or result
            print(f"[INFO] Charge Breakdown surcharge USD: 20={result['20GP']} 40={result['40GP']} 40HQ={result['40HQ']}")
            return result

        except Exception as e:
            print(f"[WARN] _read_espot_charge_breakdown_usd: {e}")
            return result

    def _read_espot_prices(self, excel_path, sheet_name, row, country="", pod=""):
        """
        Phase đọc giá E-Spot đúng flow:
        1. Tick Available Inventory Only
        2. Thu thập ETD đủ điều kiện
        3. Với từng ETD:
           - parse cards/options
           - chọn option đạt chuẩn
           - click Details
           - đọc popup Overview / Charge Breakdown / Conditional Charges
        4. Tổng hợp kết quả tốt nhất
        """
        print("[INFO] ── E-Spot: Bắt đầu phase đọc giá ──")
        self.driver.switch_to.window(self.espot_tab)

        cur_url = self.driver.current_url or ""
        if "search-result" not in cur_url:
            print(f"[WARN] E-Spot: URL không phải search-result ({cur_url}), bỏ qua đọc giá.")
            return None

        self._ensure_available_inventory_checkbox()
        time.sleep(0.5)

        qualified = self._collect_espot_calendar_dates()
        if not qualified:
            print("[INFO] E-Spot: Không có ETD nào có giá.")
            return None

        popup_results = self._collect_espot_prices_via_popup(
            qualified,
            country=country,
            pod=pod
        )
        if not popup_results:
            print("[INFO] E-Spot: Không đọc được popup result nào.")
            return None

        # chọn giá thấp nhất theo 20GP
        popup_results = sorted(
            popup_results,
            key=lambda x: (
                int(x.get("price_20") or 99999999),
                int(x.get("transit") or 99999999),
                x.get("date_str") or ""
            )
        )

        best_price_20 = int(popup_results[0]["price_20"])

        same_best = [x for x in popup_results if int(x.get("price_20") or 0) == best_price_20]
        # Sort lại same_best theo thời gian để ETD và Vessel theo đúng chuẩn biên bản
        same_best.sort(key=lambda x: x.get("date_str") or "")

        etd_dates = []
        transit_list = []
        vessel_lines = []
        transship_ports = []
        ft_pod = ""
        remark = ""

        for item in same_best:
            try:
                etd_d = datetime.strptime(item["date_str"], "%Y-%m-%d")
                etd_dates.append(etd_d)
            except Exception:
                pass

            if item.get("transit"):
                transit_list.append(int(item["transit"]))

            if item.get("vessel"):
                _etd_dt = datetime.strptime(item['date_str'], '%Y-%m-%d')
                vessel_lines.append(
                    f"{item['vessel']} / ETD: {_etd_dt.day}-{_etd_dt.strftime('%b')} / "
                    f"Transit time: {item.get('transit', 0)} Days / "
                    f"Transshipment Port: {item.get('transshipment') or 'DIRECT'}"
                )

            if item.get("transshipment"):
                transship_ports.append(item["transshipment"])

            if item.get("ft_pod") and not ft_pod:
                ft_pod = item["ft_pod"]

            if item.get("remark") and not remark:
                remark = item["remark"]

        valid_date = calculate_valid_for_espot(etd_dates)
        valid_text = f"{valid_date.day}-{valid_date.strftime('%b')}" if valid_date else ""
        etd_text   = format_etd_for_excel(etd_dates)
        transit_text = format_transit_for_excel(transit_list)

        result = {
            "price_20":    same_best[0]["price_20"],
            "price_40":    same_best[0]["price_40"],
            "price_40hq":  same_best[0]["price_40hq"],
            "formula_20":  same_best[0].get("formula_20"),
            "formula_40":  same_best[0].get("formula_40"),
            "formula_40hq": same_best[0].get("formula_40hq"),
            "etd_dates":   etd_dates,
            "standard_etd_ok": any(
                d.date() >= (datetime.now().date() + timedelta(days=DATE_OFFSET_DAYS))
                for d in etd_dates
            ),
            "etd_text":    etd_text,
            "transit_text": transit_text,
            "valid_text":  valid_text,
            "source":      "E-Spot",
            "vessel_info": "\n".join(vessel_lines),
            "remark":      remark,
            "ft_pod":      ft_pod,
            "transshipment": ", ".join(sorted(set(transship_ports))) if transship_ports else "DIRECT",
        }
        print(f"[OK] E-Spot kết quả: {result}")
        return result

    def _extract_espot_card_data(self):

        """
        Deprecated.
        Không dùng nữa vì logic mới của E-Spot phải click Details để đọc popup.
        """
        print("[WARN] _extract_espot_card_data() đã deprecated, không nên dùng nữa.")
        return None

        """
        Sau khi click vào 1 ETD trong calendar, đọc card E-Spot đầu tiên (giá thấp nhất).
        Trả về dict: {price_20, price_40, price_40hq, etd, eta, transit, vessel}
        hoặc None nếu không có kết quả.
        """
        try:
            # Chờ section E-Spot load
            WebDriverWait(self.driver, 8).until(
                lambda d: (
                    len(d.find_elements(By.CSS_SELECTOR,
                        "div.search-results div.result-item")) > 0
                    or len(d.find_elements(By.CSS_SELECTOR,
                        "div.empty-spot")) > 0
                ))
            time.sleep(0.3)

            # Kiểm tra "no results"
            empty = self.driver.find_elements(
                By.CSS_SELECTOR,
                "div.search-results div[data-v-46025bec] > div > div.empty-spot")
            # Tìm section E-Spot (section thứ 2 trong search-results)
            result_sections = self.driver.find_elements(
                By.CSS_SELECTOR, "div.search-results > div[data-v-46025bec]")

            espot_section = None
            for sec in result_sections:
                titles = sec.find_elements(By.CSS_SELECTOR, "span.moc-heading-3 > span[data-v-46025bec]")
                # Tìm title chứa text "E-Spot"
                for t in titles:
                    if "E-Spot" in (t.text or ""):
                        espot_section = sec
                        break
                if espot_section:
                    break
            # Fallback: lấy section thứ 2
            if not espot_section and len(result_sections) >= 2:
                espot_section = result_sections[1]

            if not espot_section:
                print("[WARN] Không tìm thấy section E-Spot trong bảng kết quả.")
                return None

            # Kiểm tra empty trong section này
            empty_spot = espot_section.find_elements(By.CSS_SELECTOR, "div.empty-spot")
            if empty_spot:
                print("[INFO] E-Spot: No results cho ETD này.")
                return None

            # Lấy card đầu tiên (giá thấp nhất — product-card-container)
            cards = espot_section.find_elements(
                By.CSS_SELECTOR, "div.product-card-container")
            if not cards:
                return None

            card = cards[0]

            # ── Đọc giá 20GP / 40GP / 40HQ ──
            price_20 = price_40 = price_40hq = 0
            price_infos = card.find_elements(By.CSS_SELECTOR, "div.price-info")
            for pi in price_infos:
                try:
                    label = pi.find_element(
                        By.CSS_SELECTOR, "span.moc-body-1").text.strip().upper()
                    val_text = pi.find_element(
                        By.CSS_SELECTOR, "span.box-price").text.strip()
                    val = _parse_price_display(val_text)
                    if "20" in label:
                        price_20 = val
                    elif "40GP" in label:
                        price_40 = val
                    elif "40HQ" in label or "40HC" in label:
                        price_40hq = val
                except Exception:
                    continue

            # ── Đọc ETD–ETA ──
            etd_str = eta_str = ""
            try:
                left_sec = card.find_element(By.CSS_SELECTOR, "div.left-section")
                info_divs = left_sec.find_elements(By.XPATH, ".//div[@data-v-a5fd22b9='']")
                for div in info_divs:
                    try:
                        lbl = div.find_element(
                            By.CSS_SELECTOR, "div.moc-body-1.grey-text").text.strip()
                        val = div.find_element(
                            By.CSS_SELECTOR, "div.moc-body-2.text").text.strip()
                        if "ETD" in lbl and "ETA" in lbl:
                            parts = val.split("-")
                            if len(parts) == 2:
                                etd_str = parts[0].strip()
                                eta_str = parts[1].strip()
                    except Exception:
                        continue
            except Exception:
                pass

            # ── Đọc Transit Time ──
            transit = 0
            try:
                for div in card.find_elements(By.XPATH, ".//div[@data-v-a5fd22b9='']"):
                    try:
                        lbl = div.find_element(
                            By.CSS_SELECTOR, "div.moc-body-1.grey-text").text.strip()
                        val = div.find_element(
                            By.CSS_SELECTOR, "div.moc-body-2.text").text.strip()
                        if "Transit" in lbl:
                            m = re.search(r'\d+', val)
                            if m: transit = int(m.group())
                    except Exception:
                        continue
            except Exception:
                pass

            # ── Đọc Vessel ──
            vessel = ""
            try:
                vessel = card.find_element(
                    By.CSS_SELECTOR, "div.bold-text").text.strip()
            except Exception:
                pass

            return {
                "price_20":   price_20,
                "price_40":   price_40,
                "price_40hq": price_40hq,
                "etd":        etd_str,
                "eta":        eta_str,
                "transit":    transit,
                "vessel":     vessel,
            }

        except Exception as e:
            print(f"[WARN] _extract_espot_card_data: {e}")
            return None

    def _collect_espot_calendar_dates(self):
        """
        Đọc bảng lịch tàu, thu thập các date-item có giá E-Spot (price-text thứ 2 ≠ '-').
        Scroll forward tối đa 1 lần nếu ETD tiếp theo có thể nằm trong 9 ngày.
        Trả về list[(date_str, price_int, element)] đã lọc theo giá rẻ nhất + span ≤ 9 ngày.
        """
        SCROLL_BACK  = "/html/body/div[1]/div/div[2]/div[2]/div[2]/div/div/div[1]/div[2]/button[1]"
        SCROLL_FWD   = "/html/body/div[1]/div/div[2]/div[2]/div[2]/div/div/div[1]/div[2]/button[2]"
        CALENDAR_CSS = "div.date-selector div.date-row div.date-item"

        def _parse_price(txt):
            val = _parse_price_display(txt)
            return val if val else None

        def _read_current_window():
            """Đọc tất cả date-item trong cửa sổ hiện tại."""
            items = []
            try:
                els = self.driver.find_elements(By.CSS_SELECTOR, CALENDAR_CSS)
                for el in els:
                    try:
                        date_str = el.find_element(
                            By.CSS_SELECTOR, "div.date-text").text.strip()
                        price_els = el.find_elements(
                            By.CSS_SELECTOR, "div.price-text")
                        # price_text[0]=E-Quote, price_text[1]=E-Spot
                        spot_price = None
                        if len(price_els) >= 2:
                            spot_price = _parse_price(price_els[1].text)
                        if date_str and spot_price is not None:
                            items.append((date_str, spot_price, el))
                    except Exception:
                        continue
            except Exception as e:
                print(f"[WARN] _read_current_window: {e}")
            return items

        # ── Đọc cửa sổ đầu tiên ──
        time.sleep(0.25)
        all_items = _read_current_window()
        print(f"[INFO] Calendar cửa sổ 1: {[(d,p) for d,p,_ in all_items]}")

        # ── Quyết định có nên scroll tiếp không ──
        if all_items and OOCL_ESPOT_SCAN_NEXT_WEEK:
            first_date = datetime.strptime(all_items[0][0], "%Y-%m-%d").date()
            latest_allowed = first_date + timedelta(days=9)

            # Lấy date cuối cùng của cửa sổ hiện tại
            try:
                last_els = self.driver.find_elements(By.CSS_SELECTOR, CALENDAR_CSS)
                last_date_str = last_els[-1].find_element(
                    By.CSS_SELECTOR, "div.date-text").text.strip()
                last_date = datetime.strptime(last_date_str, "%Y-%m-%d").date()
            except Exception:
                last_date = first_date + timedelta(days=6)

            # Scroll tiếp nếu cửa sổ kế có thể có ngày trong phạm vi 9 ngày
            next_window_start = last_date + timedelta(days=1)
            if next_window_start <= latest_allowed:
                try:
                    if not self._move_espot_calendar_week("next"):
                        raise Exception("không chọn được tuần kế tiếp")
                    time.sleep(0.8)
                    next_items = _read_current_window()
                    print(f"[INFO] Calendar cửa sổ 2: {[(d,p) for d,p,_ in next_items]}")
                    # Chỉ lấy các date ≤ latest_allowed
                    for d, p, el in next_items:
                        try:
                            dt = datetime.strptime(d, "%Y-%m-%d").date()
                            if dt <= latest_allowed:
                                all_items.append((d, p, el))
                        except Exception:
                            pass
                    # Scroll lại để cửa sổ hiển thị đúng nếu không có thêm gì
                    if not any(datetime.strptime(d, "%Y-%m-%d").date() <= latest_allowed
                               for d, p, _ in next_items):
                        try:
                            self._move_espot_calendar_week("prev")
                            time.sleep(0.5)
                        except Exception:
                            pass
                except Exception as e:
                    print(f"[WARN] Không scroll lịch tiếp: {e}")
            else:
                print(f"[INFO] Không cần scroll: next_window_start={next_window_start} > allowed={latest_allowed}")
        elif all_items:
            print("[OPTIMIZE] E-Spot: tuần hiện tại đã có giá -> không scan tuần kế để tăng tốc.")
        else:
            # Không có gì trong cửa sổ đầu, thử scroll 1 lần
            try:
                moved = self._move_espot_calendar_week("next")
                if not moved:
                    # Nút next có thể còn disabled trong khoảnh khắc React vừa thay
                    # lịch của tuyến cũ bằng tuyến mới. Đọc lại và thử đúng một lần.
                    time.sleep(1.2)
                    all_items = _read_current_window()
                    if all_items:
                        print(f"[INFO] Calendar cập nhật trễ: {[(d,p) for d,p,_ in all_items]}")
                    else:
                        moved = self._move_espot_calendar_week("next")
                if not all_items and not moved:
                    raise Exception("không chọn được tuần kế tiếp sau retry")
                if all_items:
                    moved = False
                if moved:
                    time.sleep(0.8)
                    all_items = _read_current_window()
                    print(f"[INFO] Calendar sau scroll (không có ban đầu): {[(d,p) for d,p,_ in all_items]}")
            except Exception as e:
                print(f"[WARN] Scroll lần đầu thất bại: {e}")

        if not all_items:
            return []

        # ── Lọc: chỉ lấy ETD có giá thấp nhất ──
        min_price = min(p for _, p, _ in all_items)
        qualified = [(d, p, el) for d, p, el in all_items if p == min_price]

        # ── Kiểm tra span ≤ 9 ngày ──
        qualified_dates = [datetime.strptime(d, "%Y-%m-%d").date() for d, _, _ in qualified]
        if qualified_dates:
            span = (max(qualified_dates) - min(qualified_dates)).days
            if span > 9:
                # Chỉ lấy từ ngày đầu đến ngày đầu + 9
                cutoff = min(qualified_dates) + timedelta(days=9)
                qualified = [(d, p, el) for d, p, el in qualified
                             if datetime.strptime(d, "%Y-%m-%d").date() <= cutoff]

        # chỉ trả date + price, không tin vào WebElement cũ sau khi scroll
        normalized = [(d, p, None) for d, p, _ in qualified]

        print(f"[INFO] ETD đủ điều kiện (giá {min_price}): {[d for d,_,_ in normalized]}")
        return normalized

    def _parse_espot_cards_current_etd(self):
        """
        Parse toàn bộ card/options của E-Spot ở ETD hiện tại.
        Chỉ lấy option có nút Details đỏ / enabled.
        Return list:
        [
          {
            "card_index": 1,
            "option_index": 1,
            "price_20": 1475,
            "price_40": 2500,
            "price_40hq": 2500,
            "transit": 42,
            "vessel": "LL3 OOCL UNITED KINGDOM 037 W",
            "detail_button_xpath": "...",
            "option_root_xpath": "...",
          },
          ...
        ]
        """
        js = r"""
        function getXPath(el) {
            if (!el) return "";
            if (el.id) return '//*[@id="' + el.id + '"]';
            const parts = [];
            while (el && el.nodeType === Node.ELEMENT_NODE) {
                let index = 1;
                let sib = el.previousElementSibling;
                while (sib) {
                    if (sib.nodeName === el.nodeName) index++;
                    sib = sib.previousElementSibling;
                }
                parts.unshift(el.nodeName.toLowerCase() + '[' + index + ']');
                el = el.parentElement;
            }
            return '/' + parts.join('/');
        }

        function parseIntSafe(txt) {
            const m = (txt || '').match(/([\d,]+(?:\.\d+)?)/);
            if (!m) return 0;
            return Math.ceil(parseFloat(m[1].replace(/,/g, '')) || 0);
        }

        function text(el, sel) {
            const node = sel ? el.querySelector(sel) : el;
            return node ? (node.textContent || '').trim() : '';
        }

        function parseTransit(mainCard) {
            let transit = 0;
            const blocks = mainCard.querySelectorAll('div');
            for (const b of blocks) {
                const t = (b.textContent || '').trim();
                if (t === 'Transit Time') {
                    const nxt = b.nextElementSibling;
                    if (nxt) {
                        const m = (nxt.textContent || '').match(/(\d+)/);
                        if (m) transit = parseInt(m[1], 10);
                    }
                }
            }
            return transit;
        }

        const out = [];
        const searchResults = document.querySelector('div.search-results');
        if (!searchResults) return out;

        // section E-Spot = section có title "E-Spot"
        let espotSection = null;
        const sections = Array.from(searchResults.children);
        for (const sec of sections) {
            if ((sec.textContent || '').includes('E-Spot')) {
                if (sec.querySelector('.result-item')) {
                    espotSection = sec;
                }
            }
        }
        if (!espotSection) return out;

        const resultItems = espotSection.querySelectorAll('div.result-item');
        let cardIndex = 0;

        for (const resultItem of resultItems) {
            const container = resultItem.querySelector('div.product-card-container');
            if (!container) continue;

            const productCards = Array.from(container.children)
                .filter(x => x.classList && x.classList.contains('product-card'));

            if (!productCards.length) continue;

            // main card = card đầu có left-section
            let mainCard = null;
            for (const c of productCards) {
                if (c.querySelector('div.left-section')) {
                    mainCard = c;
                    break;
                }
            }
            if (!mainCard) continue;

            cardIndex += 1;

            const vessel = text(mainCard, 'div.bold-text');
            const transit = parseTransit(mainCard);

            // options = các product-card có right-section
            const optionCards = productCards.filter(c => c.querySelector('div.right-section'));
            let optionIndex = 0;

            for (const opt of optionCards) {
                optionIndex += 1;

                const btn = opt.querySelector('div.right-section button');
                if (!btn) continue;

                const btnText = (btn.textContent || '').trim().toUpperCase();
                const btnClass = btn.getAttribute('class') || '';
                const disabled = btn.disabled || btnClass.includes('disable-button');

                // Chỉ giữ Details đỏ / enabled
                if (disabled) continue;
                if (btnText !== 'DETAILS') continue;

                let p20 = 0, p40 = 0, p40hq = 0;
                const pis = opt.querySelectorAll('div.right-section div.price-info');
                for (const pi of pis) {
                    const label = text(pi, 'span.moc-body-1').toUpperCase();
                    const price = parseIntSafe(text(pi, 'span.box-price'));
                    if (label.includes('20GP')) p20 = price;
                    else if (label.includes('40GP')) p40 = price;
                    else if (label.includes('40HQ')) p40hq = price;
                }

                if (!p20) continue;

                out.push({
                    card_index: cardIndex,
                    option_index: optionIndex,
                    price_20: p20,
                    price_40: p40,
                    price_40hq: p40hq,
                    transit: transit,
                    vessel: vessel,
                    detail_button_xpath: getXPath(btn),
                    option_root_xpath: getXPath(opt)
                });
            }
        }

        return out;
        """
        try:
            data = self.driver.execute_script(js) or []
            print(f"[INFO] E-Spot: parse được {len(data)} option hợp lệ ở ETD hiện tại.")
            for x in data:
                print(f"[DEBUG]   card={x['card_index']} opt={x['option_index']} "
                      f"20={x['price_20']} transit={x['transit']} vessel={x['vessel']}")
            return data
        except Exception as e:
            print(f"[WARN] _parse_espot_cards_current_etd: {e}")
            return []

    def _choose_best_espot_option(self, options):
        """
        Chọn option tốt nhất:
        1) giá 20GP thấp nhất
        2) nếu trùng giá thì transit ngắn hơn
        """
        if not options:
            return None
        best = sorted(
            options,
            key=lambda x: (
                int(x.get("price_20") or 99999999),
                int(x.get("transit") or 99999999),
                int(x.get("card_index") or 99999999),
                int(x.get("option_index") or 99999999),
            )
        )[0]
        print(f"[OK] Best option: card={best['card_index']} opt={best['option_index']} "
              f"20={best['price_20']} transit={best['transit']} vessel={best['vessel']}")
        return best

    def _click_espot_details(self, option):
        """
        Click đúng nút Details của option đã chọn.
        """
        try:
            xp = option.get("detail_button_xpath")
            if not xp:
                return False
            btn = WebDriverWait(self.driver, 6).until(
                EC.element_to_be_clickable((By.XPATH, xp))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.2)
            self.driver.execute_script("arguments[0].click();", btn)

            WebDriverWait(self.driver, 8).until(
                EC.visibility_of_element_located(
                    (By.CSS_SELECTOR, "div.ant-modal-content"))
            )
            print("[OK] Đã click Details và popup E-Spot đã mở.")
            time.sleep(0.5)
            return True
        except Exception as e:
            print(f"[WARN] _click_espot_details: {e}")
            return False

    def _parse_espot_popup_overview(self):
        """
        Đọc thông tin ở popup tab Overview.
        Return:
        {
          vessel_name: "",
          transit: 42,
          transshipment_port: "SINGAPORE",
          ft_pod: "7 COMBINED"
        }
        """
        result = {
            "vessel_name": "",
            "transit": 0,
            "transshipment_port": "",
            "ft_pod": "",
        }

        js = r"""
        const out = {
            vessel_name: "",
            transit: 0,
            transshipment_port: "",
            ft_pod: ""
        };

        const modal = document.querySelector('div.ant-modal-content');
        if (!modal) return out;

        function formatFreeTime(txt) {
            txt = (txt || '').replace(/\s+/g, ' ').trim();
            if (!txt) return '';

            const dem = txt.match(/\bDEM(?:URRAGE)?\b\s*(\d+)\s*(?:CD|CALENDAR\s*DAYS?)?/i);
            const det = txt.match(/\bDET(?:ENTION)?\b\s*(\d+)\s*(?:CD|CALENDAR\s*DAYS?)?/i);
            if (dem && det) return `${dem[1]} DEM + ${det[1]} DET`;

            const dd2in1 = txt.match(/(\d+)\s+Calendar\s+days?\s+for\s+DD2in1/i);
            if (dd2in1) return `${dd2in1[1]} COMBINED`;

            const combined = txt.match(/\b(?:COMBINED|DD2IN1)\b.*?(\d+)\s*(?:CD|CALENDAR\s*DAYS?)?/i)
                || txt.match(/(\d+)\s*(?:CD|CALENDAR\s*DAYS?).*\b(?:COMBINED|DD2IN1)\b/i);
            if (combined) return `${combined[1]} COMBINED`;

            return '';
        }

        // vessel name: ưu tiên li[2] span.svvd
        const vesselNode = modal.querySelector('span.svvd');
        if (vesselNode) {
            let txt = (vesselNode.textContent || '').trim();
            txt = txt.replace(/\s+/g, ' ').trim();
            // bỏ service code đầu nếu có
            txt = txt.replace(/^[A-Z0-9]{2,6}\s+/, '').trim();
            out.vessel_name = txt;
        }

        // transit
        const shipTips = modal.querySelectorAll('div.ship-tips');
        for (const tip of shipTips) {
            const title = (tip.querySelector('span.title') || {}).textContent || '';
            const content = (tip.querySelector('span.content') || {}).textContent || '';
            if (title.trim().toUpperCase() === 'EST. TRANSIT TIME') {
                const m = content.match(/(\d+)/);
                if (m) out.transit = parseInt(m[1], 10);
            }
        }

        // transshipment port: lấy title của timeline item có text "Transship:"
        const timelineItems = modal.querySelectorAll('ul.ant-timeline li.ant-timeline-item');
        for (const li of timelineItems) {
            const raw = (li.textContent || '').trim();
            if (raw.includes('Transship:')) {
                const titleNode = li.querySelector('div.title span');
                if (titleNode) {
                    let port = (titleNode.textContent || '').trim().toUpperCase();
                    port = port.split('-')[0].trim();
                    port = port.split(',')[0].trim();
                    out.transshipment_port = port;
                    break;
                }
            }
        }

        // free time: Destination
        const detailItems = modal.querySelectorAll('div.content-detail-item');
        for (const item of detailItems) {
            const itemName = (item.querySelector('p.item-name') || {}).textContent || '';
            if (itemName.trim().toUpperCase() === 'FREE TIME') {
                const rows = item.querySelectorAll('div.item-content');
                let destinationText = '';
                for (const row of rows) {
                    const sub = (row.querySelector('span.sub-content') || {}).textContent || '';
                    const main = (row.querySelector('span.main-content') || {}).textContent || '';
                    if (sub.trim().toUpperCase() === 'DESTINATION') {
                        destinationText += ' ' + main;
                    }
                }
                const ft = formatFreeTime(destinationText);
                if (ft) out.ft_pod = ft;
            }
        }

        if (!out.ft_pod) {
            const fullText = (modal.innerText || modal.textContent || '').replace(/\s+/g, ' ').trim();
            const destChunk = fullText.match(/Destination[\s\S]{0,400}?(?:DEM|DET|Demurrage|Detention|Combined|DD2in1)[\s\S]{0,160}/i);
            if (destChunk) {
                const ft = formatFreeTime(destChunk[0]);
                if (ft) out.ft_pod = ft;
            }
        }

        return out;
        """
        try:
            data = self.driver.execute_script(js) or result
            result.update(data)
            print(f"[INFO] Popup Overview: vessel={result['vessel_name']} transit={result['transit']} "
                  f"ts={result['transshipment_port']} ft={result['ft_pod']}")
            return result
        except Exception as e:
            print(f"[WARN] _parse_espot_popup_overview: {e}")
            return result

    def _go_to_espot_charge_breakdown_tab(self):
        try:
            tab = WebDriverWait(self.driver, 6).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//div[contains(@class,'ant-tabs-tab-btn') and normalize-space()='Charge Breakdown']"
                ))
            )
            self.driver.execute_script("arguments[0].click();", tab)
            time.sleep(0.6)
            print("[OK] Đã mở tab Charge Breakdown.")
            return True
        except Exception as e:
            print(f"[WARN] _go_to_espot_charge_breakdown_tab: {e}")
            return False

    def _parse_espot_charge_breakdown(self, country="", pod=""):
        """
        Đọc phụ phí trong Charge Breakdown.
        Chỉ cộng surcharge dạng USD vào giá total.
        Return:
        {
          add_20: ...,
          add_40: ...,
          add_40hq: ...
        }
        """
        result = {"add_20": 0, "add_40": 0, "add_40hq": 0, "non_usd_origin_thc": []}
        js = r"""
        const out = {add_20:0, add_40:0, add_40hq:0, non_usd_origin_thc:[]};

        function parseMoney(txt) {
            const m = (txt || '').toUpperCase().match(/\b(USD|VND|EUR|AUD|CHF|CNY|GBP)\b\s*([\d,\.]+)/);
            if (!m) return null;
            return {currency:m[1], amount:Math.ceil(parseFloat(m[2].replace(/,/g,'')) || 0)};
        }

        const modal = document.querySelector('div.ant-modal-content');
        if (!modal) return out;

        const tables = modal.querySelectorAll('div.ant-table-wrapper table');
        for (const tb of tables) {
            const titleWrap = tb.closest('div.table-content');
            if (!titleWrap) continue;
            const title = titleWrap.querySelector('span.table-title');
            const titleText = title ? (title.textContent || '').trim().toUpperCase() : '';

            // Bỏ Ocean Freight, chỉ lấy surcharge
            if (titleText.includes('OCEAN FREIGHT')) continue;

            // Bảng surcharge dùng rowspan trên 4 cột đầu (name/code/payment/per),
            // nên row đầu của 1 phụ phí có 7 ô (full), 2 row tiếp theo (40GP/40HQ)
            // chỉ có 3 ô (unit, price, remarks). Không hard-code index — quét theo
            // nội dung ô: tìm ô là 20GP/40GP/40HQ và lấy ô KẾ TIẾP làm price.
            const rows = tb.querySelectorAll('tbody tr.ant-table-row');
            let currentChargeName = '';
            for (const tr of rows) {
                const tds = tr.querySelectorAll('td');
                if (tds.length < 2) continue;

                let unit = '';
                let priceText = '';
                for (let i = 0; i < tds.length - 1; i++) {
                    const t = (tds[i].textContent || '').trim().toUpperCase();
                    if (t === '20GP' || t === '40GP' || t === '40HQ') {
                        unit = t;
                        priceText = (tds[i + 1].textContent || '').trim();
                        break;
                    }
                }
                if (!unit || !priceText) continue;

                if (tds.length > 3) {
                    const candidate = (tds[0].textContent || '').trim();
                    if (candidate && !/^(20GP|40GP|40HQ)$/i.test(candidate)) currentChargeName = candidate;
                }

                const money = parseMoney(priceText);
                if (!money || !money.amount) continue;
                const fullText = (titleText + ' ' + currentChargeName + ' ' + (tr.textContent || '')).toUpperCase();
                const isOriginThc = /TERMINAL HANDLING CHARGE AT ORIGIN|ORIGIN THC|\bOTHC\b|THC\/L/.test(fullText)
                    && !/DESTINATION|IMPORT|DISCHARGE|DTHC/.test(fullText);
                if (money.currency !== 'USD') {
                    if (isOriginThc) out.non_usd_origin_thc.push({unit:unit, currency:money.currency, amount:money.amount});
                    continue;
                }
                const val = money.amount;

                if (unit === '20GP') out.add_20 += val;
                else if (unit === '40GP') out.add_40 += val;
                else if (unit === '40HQ') out.add_40hq += val;
            }
        }
        return out;
        """
        try:
            data = self.driver.execute_script(js) or result
            result.update(data)
            if is_china_destination(country, pod):
                for item in result.get("non_usd_origin_thc", []):
                    amount_usd = charge_amount_to_usd(item.get("amount"), item.get("currency"))
                    unit = str(item.get("unit") or "").upper()
                    if unit == "20GP":
                        result["add_20"] += amount_usd
                    elif unit == "40GP":
                        result["add_40"] += amount_usd
                    elif unit == "40HQ":
                        result["add_40hq"] += amount_usd
                    print(f"[INFO] E-Spot +O.THC CHINA: {item.get('amount')} {item.get('currency')} -> {amount_usd:.2f} USD/{unit}")
            print(f"[INFO] Charge Breakdown surcharge USD: 20={result['add_20']} "
                  f"40={result['add_40']} 40HQ={result['add_40hq']}")
            return result
        except Exception as e:
            print(f"[WARN] _parse_espot_charge_breakdown: {e}")
            return result

    def _go_to_espot_conditional_tab(self):
        try:
            tab = WebDriverWait(self.driver, 6).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//div[contains(@class,'ant-tabs-tab-btn') and normalize-space()='Conditional Charges']"
                ))
            )
            self.driver.execute_script("arguments[0].click();", tab)
            time.sleep(0.6)
            print("[OK] Đã mở tab Conditional Charges.")
            return True
        except Exception as e:
            print(f"[WARN] _go_to_espot_conditional_tab: {e}")
            return False

    def _parse_espot_conditional_charges(self):
        result = {"has_origin_thc": True, "ows_items": []}

        js = r"""
        const out = {has_origin_thc: true, ows_items: []};

        function parseUSD(txt) {
            const m = (txt || '').match(/USD\s*([\d,\.]+)/i);
            return m ? Math.ceil(parseFloat(m[1].replace(/,/g,'')) || 0) : 0;
        }

        const modal = document.querySelector('div.ant-modal-content');
        if (!modal) return out;

        const rows = modal.querySelectorAll('tbody tr.ant-table-row');
        let currentName = "";

        for (const tr of rows) {
            const tds = tr.querySelectorAll('td.ant-table-cell');
            if (tds.length === 0) continue;

            // Lấy name từ row có rowspan
            if (tds[0].getAttribute('rowspan') || tds[0].textContent.trim().includes('Heavy Weight')) {
                currentName = tds[0].textContent.trim();
            }

            const name   = currentName || (tds[0].textContent || '').trim();
            const code   = tds.length > 1 ? (tds[1].textContent || '').trim().toUpperCase() : '';
            const unit   = Array.from(tds).find(td => td.textContent.trim() === '20GP');
            const priceTd = unit ? unit.nextElementSibling : null;
            const remarkTd = priceTd ? priceTd.nextElementSibling : null;

            const priceText = priceTd ? priceTd.textContent.trim() : '';
            const remark    = remarkTd ? remarkTd.textContent.trim() : '';

            const fullText = (name + " " + code + " " + remark).toUpperCase();

            if (/HEAVY WEIGHT|OVERWEIGHT|CWC|CWX/i.test(fullText)) {
                const amt = parseUSD(priceText);
                if (amt > 0) {
                    let threshold = "";
                    const m = remark.match(/AT OR ABOVE\s+(\d+)/i);
                    if (m) {
                        const ton = Math.round(parseInt(m[1]) / 1000);
                        threshold = ` (>${ton}TONS)`;
                    }
                    out.ows_items.push({
                        amount_20: amt,
                        text: `OWS $${amt}/20'${threshold}`.trim()
                    });
                }
            }

            if (fullText.includes("TERMINAL HANDLING CHARGE AT ORIGIN")) {
                out.has_origin_thc = true;
            }
        }
        return out;
        """

        try:
            data = self.driver.execute_script(js) or result
            result.update(data)
            
            print(f"[INFO] Conditional Charges: tìm thấy {len(result['ows_items'])} OWS")
            for item in result['ows_items']:
                print(f"   → {item['text']}")
            
            return result
        except Exception as e:
            print(f"[WARN] _parse_espot_conditional_charges: {e}")
            return result

    def _close_espot_popup(self):
        try:
            btn = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button.ant-modal-close"))
            )
            self.driver.execute_script("arguments[0].click();", btn)
            WebDriverWait(self.driver, 5).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.ant-modal-content"))
            )
            print("[OK] Đã đóng popup E-Spot.")
            time.sleep(0.3)
            return True
        except Exception as e:
            print(f"[WARN] _close_espot_popup: {e}")
            try:
                body = self.driver.find_element(By.TAG_NAME, "body")
                body.send_keys(Keys.ESCAPE)
                time.sleep(0.5)
            except Exception:
                pass
            return False

    def _build_espot_remark(self, country, conditional_data, pod=""):
        manifest = self._get_manifest_remark_by_country(country)
        if is_china_destination(country, pod):
            prefix = "INCLUDED O.THC, SUBJECT TO BILL, SEAL"
        else:
            prefix = "SUBJECT TO THC, BILL, SEAL"
        if manifest:
            prefix += f", {manifest}"

        parts = [prefix]
        seen = set()
        for item in (conditional_data or {}).get("ows_items", []):
            txt = (item or {}).get("text", "").strip()
            if txt and txt not in seen:
                seen.add(txt)
                parts.append(txt)

        return ", ".join(parts)

    def _ensure_available_inventory_checkbox(self):
        """
        Tick checkbox 'Available Inventory Only' bằng JS tìm theo text,
        không dùng xpath index cứng (dễ sai).
        """
        try:
            js = """
            let labels = document.querySelectorAll('label.ant-checkbox-wrapper');
            for (let lbl of labels) {
                if (lbl.textContent.trim().includes('Available Inventory Only')) {
                    let inp = lbl.querySelector('input[type="checkbox"]');
                    if (!inp) { lbl.click(); return {found:true, action:'click_label'}; }
                    if (!inp.checked) {
                        lbl.click();
                        return {found:true, action:'ticked'};
                    }
                    return {found:true, action:'already_checked'};
                }
            }
            return {found:false};
            """
            res = self.driver.execute_script(js)
            if res and res.get("found"):
                print(f"[OK] Đã tick 'Available Inventory Only' ({res.get('action')}).")
                if res.get("action") == "ticked":
                    time.sleep(0.25)
                return True
            # Fallback: tìm div chứa text
            els = self.driver.find_elements(By.CSS_SELECTOR, "div.checkbox-group")
            for el in els:
                if "Available Inventory Only" in (el.text or ""):
                    inp = None
                    try: inp = el.find_element(By.CSS_SELECTOR, "input[type='checkbox']")
                    except Exception: pass
                    if inp and not inp.is_selected():
                        self.driver.execute_script("arguments[0].click();", el)
                    elif inp is None:
                        self.driver.execute_script("arguments[0].click();", el)
                    print("[OK] Tick 'Available Inventory Only' qua fallback CSS.")
                    time.sleep(0.5)
                    return True
            print("[WARN] Không tìm thấy checkbox 'Available Inventory Only'.")
            return False
        except Exception as e:
            print(f"[WARN] _ensure_available_inventory_checkbox: {e}")
            return False

    def __init__(self):
        super().__init__()
        self.espot_tab       = None
        self.equote_tab      = None
        self._prev_espot_row = None  # (pol, pod) của row vừa xử lý xong
        self._prev_equote_row = None  # (pol, pod) của row vừa xử lý xong

    def _focus_tab(self, handle):
        try:
            self.driver.switch_to.window(handle)
            time.sleep(0.2)
            try:
                import ctypes
                import ctypes.wintypes
                def enum_windows_callback(hwnd, results):
                    title = ctypes.create_unicode_buffer(256)
                    ctypes.windll.user32.GetWindowTextW(hwnd, title, 256)
                    if "FreightSmart" in title.value or "OOCL" in title.value or "Edge" in title.value:
                        results.append(hwnd)
                    return True
                results = []
                WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.py_object)
                ctypes.windll.user32.EnumWindows(WNDENUMPROC(enum_windows_callback), results)
                if results:
                    ctypes.windll.user32.SetForegroundWindow(results[0])
                    ctypes.windll.user32.ShowWindow(results[0], 9)
            except Exception:
                pass
            self.driver.execute_script("window.focus();")
            time.sleep(0.3)
        except Exception as e:
            print(f"[WARN] _focus_tab: {e}")

    def setup_tabs(self):
        print("[INFO] ── Bước 4: Setup 2 tab E-Spot và E-Quote ──")
        try:
            if not self._ensure_session_alive():
                print("[ERROR] setup_tabs: session không khả dụng.")
                return False
            self._abort_if_oocl_blocked("setup_tabs start")

            import ctypes, ctypes.wintypes

            def _get_edge_hwnd():
                _hwnds = []
                @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.wintypes.HWND, ctypes.wintypes.LPARAM)
                def _cb(hwnd, lp):
                    buf = ctypes.create_unicode_buffer(512)
                    ctypes.windll.user32.GetWindowTextW(hwnd, buf, 512)
                    if (any(k in buf.value for k in ("FreightSmart","OOCL","Edge"))
                            and ctypes.windll.user32.IsWindowVisible(hwnd)):
                        _hwnds.append(hwnd)
                    return True
                ctypes.windll.user32.EnumWindows(_cb, 0)
                return _hwnds[0] if _hwnds else None

            def _focus_edge_no_resize():
                hwnd = _get_edge_hwnd()
                if not hwnd:
                    return
                class WINDOWPLACEMENT(ctypes.Structure):
                    _fields_ = [
                        ("length",           ctypes.c_uint),
                        ("flags",            ctypes.c_uint),
                        ("showCmd",          ctypes.c_uint),
                        ("ptMinPosition",    ctypes.c_long * 2),
                        ("ptMaxPosition",    ctypes.c_long * 2),
                        ("rcNormalPosition", ctypes.c_long * 4),
                    ]
                wp = WINDOWPLACEMENT()
                wp.length = ctypes.sizeof(WINDOWPLACEMENT)
                ctypes.windll.user32.GetWindowPlacement(hwnd, ctypes.byref(wp))
                ctypes.windll.user32.keybd_event(0, 0, 0, 0)
                if wp.showCmd == 2:
                    ctypes.windll.user32.ShowWindow(hwnd, 3)
                elif wp.showCmd == 3:
                    pass
                else:
                    ctypes.windll.user32.ShowWindow(hwnd, 5)
                ctypes.windll.user32.SetForegroundWindow(hwnd)
                time.sleep(0.3)

            def _visual_switch(handle):
                try:
                    handle = str(handle)
                    self.driver.switch_to.window(handle)
                    self.driver.execute_script("""
                        try {
                            Object.defineProperty(document,'hidden',
                                {value:false,configurable:true,writable:true});
                            Object.defineProperty(document,'visibilityState',
                                {value:'visible',configurable:true,writable:true});
                            document.dispatchEvent(
                                new Event('visibilitychange',{bubbles:true}));
                            window.dispatchEvent(new Event('focus',{bubbles:true}));
                        } catch(e) {}
                    """)
                    time.sleep(OOCL_VISUAL_SWITCH_SLEEP)
                except Exception as e:
                    print(f"[WARN] _visual_switch: {e}")

            def _is_hidden(handle):
                try:
                    self.driver.switch_to.window(handle)
                    return bool(self.driver.execute_script("return document.hidden;"))
                except Exception:
                    return True

            # ════════════════════════════════════
            # Bước 1: Scan tab hiện có
            # ════════════════════════════════════
            handles = list(self.driver.window_handles)
            print(f"[DEBUG] setup_tabs: có {len(handles)} tab ban đầu")

            espot_handle  = None
            equote_handle = None
            trash_handles = []

            for h in handles:
                try:
                    self.driver.switch_to.window(h)
                    url = self.driver.current_url or ""
                    self._abort_if_oocl_blocked("setup_tabs scan")
                    print(f"[DEBUG]   Tab {h[-6:]}: {url}")

                    if "my-quotation" in url:
                        if equote_handle is None:
                            equote_handle = h
                            print(f"[DEBUG]   → E-Quote")
                        else:
                            if "ntp.msn.com" not in url:
                                trash_handles.append(h)
                            else:
                                print(f"[DEBUG]   → Bỏ qua tab NTP (không đóng)")
                    elif ("freightsmart.oocl.com/ui" in url
                          or "freightsmart.oocl.com/digital" in url
                          or "freightsmart.oocl.com/en" in url
                          or "search-result" in url):
                        if espot_handle is None:
                            espot_handle = h
                            print(f"[DEBUG]   → E-Spot")
                        else:
                            if "ntp.msn.com" not in url:
                                trash_handles.append(h)
                            else:
                                print(f"[DEBUG]   → Bỏ qua tab NTP (không đóng)")
                    else:
                        if "ntp.msn.com" not in url:
                            trash_handles.append(h)
                        else:
                            print(f"[DEBUG]   → Bỏ qua tab NTP (không đóng)")
                except OOCLIpBlockedError:
                    raise
                except Exception:
                    trash_handles.append(h)

            print(f"[DEBUG] E-Spot: {espot_handle[-6:] if espot_handle else 'None'} | "
                  f"E-Quote: {equote_handle[-6:] if equote_handle else 'None'} | "
                  f"Trash: {len(trash_handles)}")

            # ════════════════════════════════════
            # Bước 2: Đóng tab thừa
            # ════════════════════════════════════
            anchor = espot_handle or equote_handle or handles[0]
            for h in trash_handles:
                try:
                    if h not in self.driver.window_handles:
                        continue
                    self.driver.switch_to.window(h)
                    self.driver.set_page_load_timeout(3)
                    try:
                        self.driver.close()
                        time.sleep(OOCL_TAB_CLOSE_SLEEP)
                        print(f"[DEBUG] Đóng tab thừa {h[-6:]}")
                    except Exception:
                        print(f"[WARN] Bỏ qua tab thừa {h[-6:]} không đóng được.")
                    finally:
                        self.driver.set_page_load_timeout(30)
                except Exception as e:
                    print(f"[WARN] Xử lý tab {h[-6:]}: {e}")
                finally:
                    try:
                        if anchor in self.driver.window_handles:
                            self.driver.switch_to.window(anchor)
                    except Exception:
                        pass                
            # ════════════════════════════════════
            # Bước 3: Tạo tab còn thiếu
            # ════════════════════════════════════

            # Nếu không có E-Spot → dùng tab còn lại hoặc mở mới
            if not espot_handle:
                remaining = self.driver.window_handles
                if remaining:
                    espot_handle = remaining[0]
                    self.driver.switch_to.window(espot_handle)
                    self.driver.execute_script(f"window.location.href = '{NEW_UI_URL}';")
                    time.sleep(OOCL_TAB_NAV_SLEEP)
                    print(f"[DEBUG] Navigate tab {espot_handle[-6:]} về E-Spot URL")

            # Nếu không có E-Quote → mở tab mới bằng JS
            if not equote_handle:
                self.driver.switch_to.window(espot_handle)
                before = set(self.driver.window_handles)
                self.driver.execute_script("window.open('about:blank','_blank');")
                WebDriverWait(self.driver, 5).until(
                    lambda d: len(d.window_handles) > len(before))
                after = set(self.driver.window_handles)
                equote_handle = list(after - before)[-1]
                self.driver.switch_to.window(equote_handle)
                self.driver.execute_script(f"window.location.href = '{EQUOTE_NEW_URL}';")
                time.sleep(OOCL_TAB_NAV_SLEEP)
                print(f"[DEBUG] Mở tab E-Quote mới: {equote_handle[-6:]}")

            self.espot_tab  = espot_handle
            self.equote_tab = equote_handle

            # ════════════════════════════════════
            # Bước 4: Visual switch từng tab để Edge focus
            # ════════════════════════════════════
            _visual_switch(self.espot_tab)
            h1 = _is_hidden(self.espot_tab)
            print(f"[OK] E-Spot  tab = {self.espot_tab[-6:]} | "
                  f"URL: {self.driver.current_url} | hidden={h1}")

            _visual_switch(self.equote_tab)
            h2 = _is_hidden(self.equote_tab)
            print(f"[OK] E-Quote tab = {self.equote_tab[-6:]} | "
                  f"URL: {self.driver.current_url} | hidden={h2}")

            # ════════════════════════════════════
            # Bước 5: Xác nhận
            # ════════════════════════════════════
            all_tabs = list(self.driver.window_handles)
            print(f"[OK] Tổng số tab sau setup: {len(all_tabs)} (mong đợi: 2)")
            for h in all_tabs:
                try:
                    self.driver.switch_to.window(h)
                    label  = "E-Spot" if h == self.espot_tab else "E-Quote"
                    url    = self.driver.current_url
                    hidden = self.driver.execute_script("return document.hidden;")
                    print(f"  [{label}] {url} | hidden={hidden}")
                except Exception:
                    pass

            # Focus lại E-Spot để bắt đầu
            _visual_switch(self.espot_tab)
            return True

        except OOCLIpBlockedError:
            raise
        except Exception as e:
            print(f"[ERROR] setup_tabs: {e}")
            return False

    def _input_port_popover(self, input_xpath, port_name, label,
                            dropdown_xpath, country=""):
        try:
            print(f"[INFO] Nhập {label}: '{port_name}'")
            input_xpaths = input_xpath if isinstance(input_xpath, (list, tuple)) else [input_xpath]
            port_clean = port_name.split(',')[0].strip().upper()
            country_upper = country.strip().upper()
            port_code = PORT_COUNTRY_CODE.get((port_clean, country_upper), "")
            country_search = PORT_COUNTRY_SEARCH_QUERY.get((port_clean, country_upper), f"{port_name.strip()} {country.strip()}".strip())
            field = None
            for xp in input_xpaths:
                try:
                    field = WebDriverWait(self.driver, 3).until(
                        EC.element_to_be_clickable((By.XPATH, xp)))
                    if field: break
                except TimeoutException:
                    continue
            if not field:
                print(f"[ERROR] {label}: không tìm thấy input.")
                return False

            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", field)
            safe_click(self.driver, field, delay=0.3)
            time.sleep(0.3)
            # Xóa + paste nhanh
            field.send_keys(Keys.CONTROL + "a")
            field.send_keys(Keys.DELETE)
            fast_input(self.driver, field, port_name)

            def _dropdown_has_target_port(d):
                for it in d.find_elements(By.CSS_SELECTOR, "div.location-item div.location-text"):
                    try:
                        txt = (it.text or "").strip().upper()
                        port_part = txt.split(",")[0].strip()
                        if port_part == port_clean or port_clean in port_part or port_part in port_clean:
                            return True
                    except Exception:
                        pass
                return False

            # Chờ dropdown có đúng port mới. Nếu chỉ chờ "có item bất kỳ" thì OOCL hay trả list cũ
            # của route trước, làm bot retry 3-4 lần vô ích.
            dropdown_appeared = smart_wait(self.driver,
                _dropdown_has_target_port,
                timeout=5, poll=0.1, after=0.1, label=f"{label} dropdown target")

            if not dropdown_appeared:
                print(f"[WARN] {label}: dropdown chưa có '{port_clean}' sau 5s, thử nhập lại...")
                safe_click(self.driver, field, delay=0.2)
                field.send_keys(Keys.CONTROL + "a")
                field.send_keys(Keys.DELETE)
                time.sleep(0.3)
                fast_input(self.driver, field, port_name)
                dropdown_appeared = smart_wait(self.driver,
                    _dropdown_has_target_port,
                    timeout=5, poll=0.1, after=0.1, label=f"{label} dropdown target retry")
                if not dropdown_appeared:
                    print(f"[WARN] {label}: dropdown vẫn chưa match sau retry, thử JS fallback.")

            js = """
            let portName    = arguments[0].toUpperCase();
            let countryName = arguments[1].toUpperCase();
            let dropXpath   = arguments[2];

            let result = document.evaluate(
                dropXpath, document, null,
                XPathResult.FIRST_ORDERED_NODE_TYPE, null);
            let container = result.singleNodeValue;
            let usedFallback = false;
            if (!container) {
                let drops = Array.from(document.querySelectorAll('div.ant-select-dropdown'));
                drops = drops.filter(d => {
                    let st = window.getComputedStyle(d);
                    return st.display !== 'none' && st.visibility !== 'hidden' && d.offsetParent !== null;
                });
                if (drops.length > 0) { container = drops[drops.length - 1]; usedFallback = true; }
            }
            if (!container) return {clicked: false, reason: 'no_container'};

            let items = container.querySelectorAll('div.location-item');
            if (!items || items.length === 0) return {clicked: false, reason: 'no_items'};

            let fallback = null;
            let debug_texts = [];
            for (let item of items) {
                let textEl = item.querySelector('div.location-text');
                if (!textEl) continue;
                let raw = textEl.textContent.trim();
                debug_texts.push(raw);
                let port = raw.split(',')[0].trim().toUpperCase();
                if (port === portName || port.includes(portName)) {
                    if (countryName && raw.toUpperCase().includes(countryName)) {
                        item.click();
                        return {clicked: true, matched: raw, matchType: 'port+country', usedFallback: usedFallback};
                    }
                    if (!fallback) fallback = {el: item, text: raw};
                }
            }

            // Virtual list: scroll dropdown xuống để render thêm items
            if (countryName && fallback) {
                let scrollHolder = container.querySelector('.rc-virtual-list-holder');
                if (!scrollHolder) scrollHolder = container.querySelector('[style*="overflow"]');
                if (scrollHolder) {
                    let scrollStep = 150;
                    let maxScroll = scrollHolder.scrollHeight;
                    for (let pos = scrollStep; pos <= maxScroll; pos += scrollStep) {
                        scrollHolder.scrollTop = pos;
                        // Chờ DOM update (sync trick)
                        let t0 = Date.now(); while (Date.now() - t0 < 50) {}
                        let newItems = container.querySelectorAll('div.location-item');
                        for (let item of newItems) {
                            let textEl = item.querySelector('div.location-text');
                            if (!textEl) continue;
                            let raw = textEl.textContent.trim();
                            if (debug_texts.indexOf(raw) === -1) debug_texts.push(raw);
                            let port = raw.split(',')[0].trim().toUpperCase();
                            if ((port === portName || port.includes(portName))
                                && raw.toUpperCase().includes(countryName)) {
                                item.click();
                                return {clicked: true, matched: raw, matchType: 'port+country+scroll', usedFallback: usedFallback};
                            }
                        }
                    }
                }
            }

            if (fallback) {
                if (countryName) {
                    return {clicked: false, reason: 'port_only_no_country', matched: fallback.text, texts: debug_texts.slice(0,10)};
                }
                fallback.el.click();
                return {clicked: true, matched: fallback.text, matchType: 'port_only', usedFallback: usedFallback};
            }
            return {clicked: false, reason: 'no_match', texts: debug_texts.slice(0,10)};
            """

            clicked = False
            country_retry_done = False
            for attempt in range(8):
                try:
                    res = self.driver.execute_script(js, port_name.strip(), country.strip(), dropdown_xpath)
                    if res and res.get("clicked"):
                        print(f"[OK] {label} chọn: '{res.get('matched')}' ({res.get('matchType')}) | attempt {attempt+1}")
                        clicked = True
                        break
                    else:
                        reason = res.get("reason") if res else "unknown"
                        # Tìm thấy port nhưng country không khớp → thử search kèm country
                        if reason == 'port_only_no_country' and not country_retry_done:
                            country_retry_done = True
                            print(f"[WARN] {label}: tìm thấy '{res.get('matched')}' nhưng country không khớp, thử search '{port_name.strip()} {country.strip()}'...")
                            for xp in (input_xpath if isinstance(input_xpath, (list, tuple)) else [input_xpath]):
                                try:
                                    fld = self.driver.find_element(By.XPATH, xp)
                                    if fld:
                                        safe_click(self.driver, fld, delay=0.2)
                                        fld.send_keys(Keys.CONTROL + "a")
                                        fld.send_keys(Keys.DELETE)
                                        time.sleep(0.3)
                                        fast_input(self.driver, fld, port_name.strip() + " " + country.strip())
                                        break
                                except: continue
                            smart_wait(self.driver,
                                lambda d: len(d.find_elements(By.CSS_SELECTOR, "div.location-item")) > 0,
                                timeout=6, poll=0.1, after=0.15, label=f"{label} dropdown+country")
                            # Thử JS lại với dropdown mới; nếu vẫn không match country thì thử port code.
                            res2 = self.driver.execute_script(js, port_name.strip(), country.strip(), dropdown_xpath)
                            if res2 and res2.get("clicked"):
                                print(f"[OK] {label} chọn: '{res2.get('matched')}' ({res2.get('matchType')}) | attempt {attempt+1} (retry+country)")
                                clicked = True
                                break
                            if port_code:
                                print(f"[WARN] {label}: thử search bằng port code '{port_code}' cho {port_clean}/{country_upper}...")
                                for xp in (input_xpath if isinstance(input_xpath, (list, tuple)) else [input_xpath]):
                                    try:
                                        fld = self.driver.find_element(By.XPATH, xp)
                                        if fld:
                                            safe_click(self.driver, fld, delay=0.2)
                                            fld.send_keys(Keys.CONTROL + "a")
                                            fld.send_keys(Keys.DELETE)
                                            time.sleep(0.3)
                                            fast_input(self.driver, fld, port_code)
                                            break
                                    except: continue
                                smart_wait(self.driver,
                                    lambda d: len(d.find_elements(By.CSS_SELECTOR, "div.location-item")) > 0,
                                    timeout=6, poll=0.1, after=0.15, label=f"{label} dropdown+code")
                                res3 = self.driver.execute_script(js, port_name.strip(), country.strip(), dropdown_xpath)
                                if res3 and res3.get("clicked"):
                                    print(f"[OK] {label} chọn: '{res3.get('matched')}' ({res3.get('matchType')}) | attempt {attempt+1} (retry+code {port_code})")
                                    clicked = True
                                    break
                        else:
                            debug_print(f"[DEBUG] {label} attempt {attempt+1}: {reason}")
                            if res and res.get("texts"):
                                debug_print(f"[DEBUG]   samples: {res.get('texts')}")
                except Exception as e:
                    debug_print(f"[DEBUG] {label} JS error attempt {attempt+1}: {e}")
                time.sleep(0.4)

            if not clicked:
                print(f"[WARN] {label}: không click được option sau 8 lần thử.")
                return False

            time.sleep(0.5); r()
            return True

        except Exception as e:
            print(f"[ERROR] _input_port_popover {label}: {e}")
            return False

    def _select_date_espot(self, date_input_xpath, date_popup_xpath, days_offset=None):
        if days_offset is None:
            days_offset = DATE_OFFSET_DAYS
        target_date = (datetime.now() + timedelta(days=days_offset)).date()
        target_str  = target_date.strftime("%Y-%m-%d")
        print(f"[INFO] Chọn ngày ETD: {target_str} (hôm nay +{days_offset})")
        try:
            inp = WebDriverWait(self.driver, 8).until(
                EC.element_to_be_clickable((By.XPATH, date_input_xpath)))
            safe_click(self.driver, inp, delay=0.3)
            WebDriverWait(self.driver, 8).until(
                EC.presence_of_element_located((By.XPATH, date_popup_xpath)))
            time.sleep(0.5)

            cell = None
            try:
                cell = self.driver.find_element(
                    By.CSS_SELECTOR,
                    f"td.ant-picker-cell[title='{target_str}']:not(.ant-picker-cell-disabled)")
            except Exception:
                pass

            if not cell:
                print(f"[WARN] Ngày {target_str} disabled, tìm ngày gần nhất...")
                for extra in range(1, 15):
                    alt_str = (datetime.now() + timedelta(days=days_offset + extra)).date().strftime("%Y-%m-%d")
                    try:
                        cell = self.driver.find_element(
                            By.CSS_SELECTOR,
                            f"td.ant-picker-cell[title='{alt_str}']:not(.ant-picker-cell-disabled)")
                        print(f"[INFO] Dùng ngày thay thế: {alt_str}")
                        break
                    except Exception:
                        continue

            if cell:
                safe_click(self.driver, cell, delay=0.2)
                print(f"[OK] Đã chọn ngày.")
                time.sleep(0.5)
            else:
                print("[WARN] Không tìm được ô ngày, nhấn Escape.")
                self.driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)

        except Exception as e:
            print(f"[ERROR] _select_date_espot: {e}")

    def _set_containers_espot(self, container_box_xpath, container_popup_xpath):
        try:
            print("[INFO] Mở popup chọn container...")
            box = None
            CONT_BOX_SELECTORS = [
                (By.CSS_SELECTOR, "div.cargo-input-wrap div.input-container"),
                (By.CSS_SELECTOR, "div.cargo-input-wrap div.placeholder-wrap"),
                (By.CSS_SELECTOR, "div.cargo-block div.input-container"),
                (By.XPATH, container_box_xpath),
            ]
            for by, sel in CONT_BOX_SELECTORS:
                try:
                    els = self.driver.find_elements(by, sel)
                    visible = [e for e in els if e.is_displayed()]
                    if visible:
                        box = visible[0]
                        print(f"[DEBUG] Tìm thấy CONT_BOX bằng: {sel}")
                        break
                except Exception:
                    continue

            if not box:
                print("[ERROR] Không tìm thấy ô container bằng bất kỳ selector nào.")
                return

            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", box)
            time.sleep(0.3)
            self.driver.execute_script("arguments[0].click();", box)
            print("[DEBUG] Đã click vào ô container.")
            time.sleep(1.5)

            popup_visible = False
            for attempt in range(6):
                els = self.driver.find_elements(
                    By.CSS_SELECTOR, "div.ant-popover input.ant-input-number-input")
                visible = [e for e in els if e.is_displayed()]
                if visible:
                    popup_visible = True
                    print(f"[OK] Popup container hiện tại attempt {attempt+1}.")
                    break
                print(f"[DEBUG] Popup chưa hiện (attempt {attempt+1}), thử click lại...")
                self.driver.execute_script("arguments[0].click();", box)
                time.sleep(0.8)

            if not popup_visible:
                print("[ERROR] Không tìm thấy popup container sau 6 lần thử.")
                return

            time.sleep(0.3)
            inputs = self.driver.find_elements(
                By.CSS_SELECTOR, "div.ant-popover input.ant-input-number-input")
            visible_inputs = [i for i in inputs if i.is_displayed()]
            print(f"[DEBUG] Tìm thấy {len(visible_inputs)} input container visible.")

            labels = ["20GP", "40GP", "40HQ"]
            for idx, inp in enumerate(visible_inputs[:3]):
                label = labels[idx] if idx < len(labels) else f"input[{idx}]"
                try:
                    self.driver.execute_script("""
                        let input = arguments[0];
                        let nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                            window.HTMLInputElement.prototype, 'value').set;
                        nativeInputValueSetter.call(input, '1');
                        input.dispatchEvent(new Event('input', {bubbles: true}));
                        input.dispatchEvent(new Event('change', {bubbles: true}));
                    """, inp)
                    print(f"[OK] Đã nhập {label} = 1")
                    r()
                except Exception as e:
                    print(f"[WARN] Không nhập được {label}: {e}")

            if len(visible_inputs) < 3:
                print(f"[WARN] Chỉ tìm thấy {len(visible_inputs)}/3 input container.")

            time.sleep(0.3)
            self.driver.find_element(By.TAG_NAME, "body").click()
            time.sleep(0.5)

        except Exception as e:
            print(f"[ERROR] _set_containers_espot: {e}")

    def perform_unified_search(self, pol, pod, country=""):
        """
        Nhập liệu trên tab E-Spot.
        Row đầu: reload trang. Row tiếp: smart navigate dùng lại search-result.
        """
        print(f"\n[INFO] ── E-SPOT: Nhập liệu {pol} → {pod} ──")
        self.driver.switch_to.window(self.espot_tab)
        time.sleep(0.2)
        self._switch_to_tab_by_keyboard(self.espot_tab)

        pol_s = PORT_SEARCH_ALIAS.get(pol, pol)
        pod_s = PORT_SEARCH_ALIAS.get(pod, pod)

        # ── Smart navigate nếu đang ở search-result và có row trước ──
        if self._prev_espot_row is not None:
            prev_pol, prev_pod = self._prev_espot_row
            cur_url = self.driver.current_url or ""
            if "search-result" in cur_url:
                print(f"[INFO] E-Spot: thử smart navigate (search-result → reuse).")
                prev_result_sig = (self._get_espot_result_state().get("signature") or "")
                ok = self._navigate_next_espot(
                    pol_s, pod_s, country,
                    PORT_SEARCH_ALIAS.get(prev_pol, prev_pol),
                    PORT_SEARCH_ALIAS.get(prev_pod, prev_pod),
                    "VIETNAM", country)
                if ok:
                    print("[OK] E-Spot: Smart navigate thành công. Chờ kết quả...")
                    fresh = self._wait_for_results(
                        "espot", timeout=12, previous_signature=prev_result_sig)
                    if fresh:
                        print("[OK] E-Spot: Phase nhập liệu hoàn tất.")
                        self._prev_espot_row = (pol, pod)
                        return True
                    print("[WARN] E-Spot: kết quả Smart Navigate chưa đổi → fallback reload.")
                else:
                    print("[WARN] E-Spot: Smart navigate thất bại → fallback reload.")

        # ── Fallback: reload trang ──
        print(f"[INFO] E-Spot: navigate (refresh) về {NEW_UI_URL}")
        try:
            self.driver.get(NEW_UI_URL)
            WebDriverWait(self.driver, 15).until(
                lambda d: (
                    "freightsmart.oocl.com/ui" in d.current_url
                    or "freightsmart.oocl.com/digital" in d.current_url
                )
                          and "my-quotation" not in d.current_url)
            time.sleep(1.0)
        except TimeoutException:
            try: self.driver.execute_script("window.stop();")
            except Exception: pass

        POL_XPATH   = ["/html/body/div[1]/div/div[2]/div/div[1]/div[3]/div/div[1]/div[1]/div/span/input"]
        POD_XPATH   = ["/html/body/div[1]/div/div[2]/div/div[1]/div[3]/div/div[1]/div[2]/div/span/input"]
        DATE_XPATH  = "/html/body/div[1]/div/div[2]/div/div[1]/div[3]/div[1]/div[2]/div[1]/input"
        CONT_BOX    = "/html/body/div[1]/div/div[2]/div/div[1]/div[3]/div/div[3]/div/span/div[1]/div[1]"
        CONT_POPUP  = "/html/body/div[8]/div/div/div/div[2]"
        SEARCH_BTN  = "/html/body/div[1]/div/div[2]/div/div[1]/div[3]/button"
        POL_DROP    = "/html/body/div[5]/div/div/div/div[2]/div"
        POD_DROP    = "/html/body/div[6]/div/div/div/div[2]/div"
        DATE_PICKER = "/html/body/div[7]/div/div"

        try:
            if not self._find_first_xpath(POL_XPATH, timeout=12, clickable=False):
                print("[WARN] E-Spot: form chưa thấy sau navigate, thử lại sau 3s...")
                time.sleep(3)
                if not self._find_first_xpath(POL_XPATH, timeout=8, clickable=False):
                    raise TimeoutException("Không thấy POL input sau 2 lần th��")
            print("[OK] E-Spot: form sẵn sàng.")
        except TimeoutException:
            print("[ERROR] E-Spot: Timeout chờ form. Dừng nhập liệu E-Spot.")
            return False

        if not self._input_port_popover(POL_XPATH, pol_s, "POL (E-Spot)", POL_DROP, "VIETNAM"):
            print("[ERROR] E-Spot: nhập POL thất bại.")
            return False

        if not self._input_port_popover(POD_XPATH, pod_s, "POD (E-Spot)", POD_DROP, country):
            print("[ERROR] E-Spot: nhập POD thất bại.")
            return False

        self._select_date_espot(DATE_XPATH, DATE_PICKER, days_offset=DATE_OFFSET_DAYS)
        self._set_containers_espot(CONT_BOX, CONT_POPUP)

        print("[INFO] E-Spot: Click Search...")
        try:
            btn = WebDriverWait(self.driver, 6).until(
                EC.element_to_be_clickable((By.XPATH, SEARCH_BTN)))
            safe_click(self.driver, btn, delay=0.2)
            print("[OK] E-Spot: Đã nhấn Search. Chờ kết quả...")
            self._wait_for_results("espot", timeout=12)
            print("[OK] E-Spot: Phase nhập liệu hoàn tất.")
            self._prev_espot_row = (pol, pod)
            return True
        except Exception as e:
            print(f"[ERROR] E-Spot: Click Search thất bại: {e}")
            return False

    def perform_equote_search(self, pol, pod, country=""):
        print(f"\n[INFO] ── E-QUOTE: Nhập liệu {pol} → {pod} ──")

        self._switch_to_tab_by_keyboard(self.equote_tab)

        pol_s = PORT_SEARCH_ALIAS.get(pol, pol)
        pod_s = PORT_SEARCH_ALIAS.get(pod, pod)

        POL_XPATH   = ("/html/body/div[1]/div/div[2]/div[2]/div[1]/form"
                       "/div[1]/div[1]/div[1]/div/span/input")
        POD_XPATH   = ("/html/body/div[1]/div/div[2]/div[2]/div[1]/form"
                       "/div[1]/div[1]/div[2]/div/span/input")
        POL_CLEAR   = ("/html/body/div[1]/div/div[2]/div[2]/div[1]/form"
                       "/div[1]/div[1]/div[1]/div/span/span[2]/span/span")
        POD_CLEAR   = ("/html/body/div[1]/div/div[2]/div[2]/div[1]/form"
                       "/div[1]/div[1]/div[2]/div/span/span[2]/span/span")
        SEARCH_BTN  = ("/html/body/div[1]/div/div[2]/div[2]/div[1]"
                       "/form/div[3]/button[1]")
        POL_DROP    = "/html/body/div[6]/div/div"
        POD_DROP    = "/html/body/div[7]/div/div"

        # ── Smart navigate: chỉ sửa field thay đổi ──
        if self._prev_equote_row is not None:
            prev_pol, prev_pod = self._prev_equote_row
            prev_pol_s = PORT_SEARCH_ALIAS.get(prev_pol, prev_pol)
            prev_pod_s = PORT_SEARCH_ALIAS.get(prev_pod, prev_pod)

            cur_url = self.driver.current_url or ""
            if "my-quotation" in cur_url:
                print(f"[INFO] E-Quote: smart navigate (chỉ sửa field thay đổi).")
                try:
                    pol_changed = (pol_s != prev_pol_s)
                    pod_changed = (pod_s != prev_pod_s)

                    # ── Sửa POL nếu thay đổi ──
                    if pol_changed:
                        print(f"[INFO] E-Quote POL thay đổi: {prev_pol_s} → {pol_s}")
                        try:
                            clear = WebDriverWait(self.driver, 4).until(
                                EC.element_to_be_clickable((By.XPATH, POL_CLEAR)))
                            self.driver.execute_script(
                                "arguments[0].click();", clear)
                            time.sleep(0.4)
                        except Exception:
                            pass
                        self._input_port_popover(
                            POL_XPATH, pol_s, "POL (E-Quote)", POL_DROP, "VIETNAM")
                    else:
                        print(f"[INFO] E-Quote POL không đổi ({pol_s}), bỏ qua.")

                    # ── Sửa POD nếu thay đổi ──
                    if pod_changed:
                        print(f"[INFO] E-Quote POD thay đổi: {prev_pod_s} → {pod_s}")
                        try:
                            clear = WebDriverWait(self.driver, 4).until(
                                EC.element_to_be_clickable((By.XPATH, POD_CLEAR)))
                            self.driver.execute_script(
                                "arguments[0].click();", clear)
                            # Chờ field trống hẳn trước khi nhập mới
                            smart_wait(self.driver,
                                lambda d: (
                                    lambda els: els[0].get_attribute("value") in ("", None)
                                    if els else True
                                )(d.find_elements(By.XPATH, POD_XPATH)),
                                timeout=3, poll=0.05, after=0.1,
                                label="POD field cleared")
                        except Exception:
                            pass
                        self._input_port_popover(
                            POD_XPATH, pod_s, "POD (E-Quote)", POD_DROP, country)
                    else:
                        print(f"[INFO] E-Quote POD không đổi ({pod_s}), bỏ qua.")

                    # ── Click Search ──
                    prev_result_sig = (self._get_equote_result_state().get("signature") or "")
                    btn = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, SEARCH_BTN)))
                    self.driver.execute_script("arguments[0].click();", btn)
                    print("[OK] E-Quote: Smart navigate → đã nhấn Search.")
                    self._wait_for_results(
                        "equote",
                        timeout=OOCL_EQUOTE_RESULT_WAIT_SECONDS,
                        previous_signature=prev_result_sig,
                    )
                    print("[OK] E-Quote: Phase nhập liệu hoàn tất.")
                    self._prev_equote_row = (pol, pod)
                    return True

                except Exception as e:
                    print(f"[WARN] E-Quote smart navigate thất bại: {e} → fallback reload.")

        # ── Fallback: reload trang ──
        print(f"[INFO] E-Quote: navigate về {EQUOTE_NEW_URL}")
        try:
            self.driver.execute_script(f"window.location.href = '{EQUOTE_NEW_URL}';")
            WebDriverWait(self.driver, 10).until(
                lambda d: "my-quotation" in d.current_url)
            time.sleep(0.8)
        except TimeoutException:
            try: self.driver.execute_script("window.stop();")
            except Exception: pass

        self.driver.execute_script("""
            try {
                Object.defineProperty(document,'hidden',
                    {value:false,configurable:true,writable:true});
                Object.defineProperty(document,'visibilityState',
                    {value:'visible',configurable:true,writable:true});
                document.dispatchEvent(
                    new Event('visibilitychange',{bubbles:true}));
                window.dispatchEvent(new Event('focus',{bubbles:true}));
            } catch(e) {}
        """)
        time.sleep(0.3)

        try:
            self.driver.find_element(By.TAG_NAME, "body").click()
        except Exception:
            pass
        time.sleep(0.3)

        try:
            WebDriverWait(self.driver, 12).until(
                EC.presence_of_element_located((By.XPATH, POL_XPATH)))
            print("[OK] E-Quote: form sẵn sàng.")
        except TimeoutException:
            print("[ERROR] E-Quote: Timeout chờ form.")
            return False

        if not self._input_port_popover(
                POL_XPATH, pol_s, "POL (E-Quote)", POL_DROP, "VIETNAM"):
            print("[ERROR] E-Quote: nhập POL thất bại.")
            return False

        if not self._input_port_popover(
                POD_XPATH, pod_s, "POD (E-Quote)", POD_DROP, country):
            print("[ERROR] E-Quote: nhập POD thất bại.")
            return False

        print("[INFO] E-Quote: Click Search...")
        try:
            prev_result_sig = (self._get_equote_result_state().get("signature") or "")
            btn = WebDriverWait(self.driver, 6).until(
                EC.element_to_be_clickable((By.XPATH, SEARCH_BTN)))
            safe_click(self.driver, btn, delay=0.3)
            print("[OK] E-Quote: Đã nhấn Search. Chờ kết quả...")
            self._wait_for_results(
                "equote",
                timeout=OOCL_EQUOTE_RESULT_WAIT_SECONDS,
                previous_signature=prev_result_sig,
            )
            print("[OK] E-Quote: Phase nhập liệu hoàn tất.")
            self._prev_equote_row = (pol, pod)
            return True
        except Exception as e:
            print(f"[ERROR] E-Quote: Click Search thất bại: {e}")
            return False
    # ═══════════════════════════════════════════════════════════
    # E-SPOT PRICE CHECK
    # ═══════════════════════════════════════════════════════════

    def _ensure_inventory_checked(self):
        """Tick 'Available Inventory Only' nếu chưa được tích."""
        XPATH = "/html/body/div[1]/div/div[2]/div[2]/div[1]/div[2]/div[9]"
        try:
            container = WebDriverWait(self.driver, 8).until(
                EC.presence_of_element_located((By.XPATH, XPATH)))
            try:
                cb = container.find_element(
                    By.CSS_SELECTOR, "input.ant-checkbox-input")
                if cb.is_selected():
                    print("[OK] 'Available Inventory Only' đã tick sẵn.")
                    return
            except Exception:
                pass
            label = container.find_element(
                By.CSS_SELECTOR, "label.ant-checkbox-wrapper")
            self.driver.execute_script("arguments[0].click();", label)
            time.sleep(0.6)
            print("[OK] Đã tick 'Available Inventory Only'.")
        except Exception as e:
            print(f"[WARN] _ensure_inventory_checked: {e}")

    def _read_calendar_page(self):
        """
        Đọc 7 date-item của trang lịch hiện tại.
        Returns list of {date: 'YYYY-MM-DD', espot_price: int or None}
        """
        result = []
        try:
            items = self.driver.find_elements(
                By.CSS_SELECTOR, "div.date-selector div.date-item")
            for item in items:
                try:
                    date_str = item.find_element(
                        By.CSS_SELECTOR, "div.date-text").text.strip()
                    price_texts = item.find_elements(
                        By.CSS_SELECTOR, "div.price-text")
                    espot_price = None
                    # price-text[0] = E-Quote, price-text[1] = E-Spot
                    if len(price_texts) >= 2:
                        raw = price_texts[1].text.strip()
                        m = re.search(r'USD\s+([\d,]+(?:\.\d+)?)', raw)
                        if m:
                            espot_price = _ceil_money_value(m.group(1).replace(',', ''))
                    result.append({"date": date_str, "espot_price": espot_price})
                except Exception:
                    pass
        except Exception as e:
            print(f"[WARN] _read_calendar_page: {e}")
        return result

    def _collect_valid_espot_etds(self):
        """
        Quét bảng lịch (scroll thông minh) thu thập ETD hợp lệ:
          - Có giá E-Spot (không phải '-')
          - Cùng mức giá tham chiếu thấp nhất
          - Không cách ETD đầu tiên quá 9 ngày
        Returns list of date objects, đã sort.
        """
        NEXT_BTN = ("/html/body/div[1]/div/div[2]/div[2]/div[2]"
                    "/div/div/div[1]/div[2]/button[2]")

        all_priced = []  # [{date: date_obj, price: int}]

        for page_idx in range(4):   # tối đa 4 trang (28 ngày)
            time.sleep(0.5)
            page_data = self._read_calendar_page()
            if not page_data:
                break

            for item in page_data:
                if item["espot_price"] is not None:
                    try:
                        d = datetime.strptime(item["date"], "%Y-%m-%d").date()
                        all_priced.append({"date": d, "price": item["espot_price"]})
                    except Exception:
                        pass

            # Quyết định dừng scroll khi trang hiện tại đã vượt cutoff
            if all_priced:
                first_d = min(x["date"] for x in all_priced)
                cutoff  = first_d + timedelta(days=9)
                try:
                    last_d = datetime.strptime(
                        page_data[-1]["date"], "%Y-%m-%d").date()
                    if last_d >= cutoff:
                        print(f"[INFO] Lịch trang {page_idx+1}: "
                              f"đã qua cutoff {cutoff} → dừng scroll.")
                        break
                except Exception:
                    break

            # Scroll sang tuần tiếp
            try:
                if not self._move_espot_calendar_week("next"):
                    raise Exception("không chọn được tuần kế tiếp")
            except Exception as e:
                print(f"[WARN] Không scroll lịch tiếp: {e}")
                break

        if not all_priced:
            print("[INFO] E-Spot: Không có ETD nào có giá.")
            return []

        # Lọc theo quy tắc
        min_price   = min(x["price"] for x in all_priced)
        first_min   = min(x["date"] for x in all_priced
                          if x["price"] == min_price)
        cutoff      = first_min + timedelta(days=9)
        valid       = sorted([x["date"] for x in all_priced
                               if x["price"] == min_price
                               and x["date"] <= cutoff
                               and etd_within_max(x["date"])])

        print(f"[INFO] E-Spot valid ETDs: {[str(d) for d in valid]} "
              f"(ref price = {min_price} USD)")
        return valid

    def _click_etd_date(self, date_str):
        """
        Scroll lịch (cả 2 chiều) rồi click vào date-item date_str ('YYYY-MM-DD').
        Returns True nếu click thành công.
        """
        NEXT_BTN = ("/html/body/div[1]/div/div[2]/div[2]/div[2]"
                    "/div/div/div[1]/div[2]/button[2]")
        PREV_BTN = ("/html/body/div[1]/div/div[2]/div[2]/div[2]"
                    "/div/div/div[1]/div[2]/button[1]")
        target = datetime.strptime(date_str, "%Y-%m-%d").date()

        for _ in range(8):
            time.sleep(0.4)
            page_data = self._read_calendar_page()
            dates_on_page = []
            for item in page_data:
                try:
                    dates_on_page.append(
                        datetime.strptime(item["date"], "%Y-%m-%d").date())
                except Exception:
                    pass

            if target in dates_on_page:
                # Tìm và click element
                try:
                    items = self.driver.find_elements(
                        By.CSS_SELECTOR, "div.date-selector div.date-item")
                    for item in items:
                        try:
                            txt = item.find_element(
                                By.CSS_SELECTOR, "div.date-text").text.strip()
                            if txt == date_str:
                                self.driver.execute_script(
                                    "arguments[0].scrollIntoView({block:'center'});",
                                    item)
                                self.driver.execute_script(
                                    "arguments[0].click();", item)
                                time.sleep(1.0)
                                print(f"[OK] Click ETD: {date_str}")
                                return True
                        except Exception:
                            pass
                except Exception as e:
                    print(f"[WARN] _click_etd_date {date_str}: {e}")
                return False

            # Scroll về đúng hướng
            if dates_on_page:
                direction = "next" if target > max(dates_on_page) else "prev"
                try:
                    if not self._move_espot_calendar_week(direction):
                        break
                except Exception:
                    break
            else:
                break

        print(f"[WARN] Không tìm thấy date {date_str} trong lịch.")
        return False

    def _parse_espot_spot_cards(self):
        """
        Đọc các card E-Spot (main cards, bỏ qua sub-packages) sau khi click ETD.
        Returns list of dicts:
          {vessel, etdEta, transitTime, isTransship, p20, p40, p40hq}
        """
        js = """
        const cards = [];
        const sr = document.querySelector('div.search-results');
        if (!sr) return cards;

        // Lấy section E-Spot: section cuối có chứa result-item
        let espotSec = null;
        const secs = sr.children;
        for (let i = secs.length - 1; i >= 0; i--) {
            if (secs[i].querySelector && secs[i].querySelector('.result-item')) {
                espotSec = secs[i];
                break;
            }
        }
        if (!espotSec) return cards;

        for (const ri of espotSec.querySelectorAll('.result-item')) {
            const container = ri.querySelector('.product-card-container');
            if (!container) continue;

            // Tìm main card: child product-card có left-section
            let mainCard = null;
            for (const child of container.children) {
                if (child.classList.contains('product-card')
                        && child.querySelector('.left-section')) {
                    mainCard = child;
                    break;
                }
            }
            if (!mainCard) continue;

            // Vessel
            const vEl = mainCard.querySelector('.bold-text');
            const vessel = vEl ? vEl.textContent.trim() : '';

            // Transshipment: đếm vessel icon trong timeline
            const vIcons = mainCard.querySelectorAll(
                '.timeline .i-icon-vessel');
            const isTransship = vIcons.length >= 2;

            // ETD-ETA và Transit Time
            let etdEta = '', transitTime = '';
            const divs = mainCard.querySelectorAll('div');
            for (let i = 0; i < divs.length; i++) {
                const t = divs[i].textContent.trim();
                if (t === 'ETD \u2013 ETA') {
                    const nxt = divs[i].nextElementSibling;
                    if (nxt) etdEta = nxt.textContent.trim();
                }
                if (t === 'Transit Time') {
                    const nxt = divs[i].nextElementSibling;
                    if (nxt) transitTime = nxt.textContent.trim();
                }
            }

            // Giá 20GP / 40GP / 40HQ từ right-section của main card
            let p20 = null, p40 = null, p40hq = null;
            const pis = mainCard.querySelectorAll(
                '.right-section .price-info');
            for (const pi of pis) {
                const lbl = (pi.querySelector('.moc-body-1') || {}).textContent;
                const raw = (pi.querySelector('.box-price') || {}).textContent;
                if (!lbl || !raw) continue;
                const m = raw.match(/([\\d,]+(?:\\.\\d+)?)/);
                const num = m ? Math.ceil(parseFloat(m[1].replace(/,/g, '')) || 0) : null;
                if (lbl.includes('20GP'))  p20  = num;
                else if (lbl.includes('40GP'))  p40  = num;
                else if (lbl.includes('40HQ'))  p40hq = num;
            }

            if (p20 !== null)
                cards.push({vessel, etdEta, transitTime,
                             isTransship, p20, p40, p40hq});
        }
        return cards;
        """
        try:
            WebDriverWait(self.driver, 8).until(
                lambda d: (
                    len(d.find_elements(
                        By.CSS_SELECTOR,
                        "div.search-results .result-item")) > 0
                    or len(d.find_elements(
                        By.CSS_SELECTOR,
                        "div.search-results .empty-spot")) > 0))
            time.sleep(0.4)
            return self.driver.execute_script(js) or []
        except Exception as e:
            print(f"[WARN] _parse_espot_spot_cards: {e}")
            return []

    def _scrape_espot(self, pod=""):
        """
        Orchestrator E-Spot: sau khi search xong, đọc giá tốt nhất.
        Returns dict dùng cho write_to_excel, hoặc None nếu không có giá.
        """
        print("[INFO] ── E-Spot: Bắt đầu phase đọc giá ──")
        self.driver.switch_to.window(self.espot_tab)
        time.sleep(0.3)

        # 1. Tick 'Available Inventory Only'
        self._ensure_inventory_checked()
        time.sleep(0.5)

        # 2. Thu thập ETD hợp lệ
        valid_etds = self._collect_valid_espot_etds()
        if not valid_etds:
            print("[INFO] E-Spot: Không có ETD → ghi '-'.")
            return None

        # 3. Click từng ETD, parse cards
        all_cards = []
        for etd_date in valid_etds:
            etd_str = etd_date.strftime("%Y-%m-%d")
            if not self._click_etd_date(etd_str):
                continue
            cards = self._parse_espot_spot_cards()
            print(f"[INFO] E-Spot {etd_str}: {len(cards)} card(s).")
            for c in cards:
                c["etd_date"] = etd_date
            all_cards.extend(cards)

        if not all_cards:
            print("[INFO] E-Spot: Không parse được card nào → ghi '-'.")
            return None

        # 4. Chọn card tốt nhất (20GP thấp nhất)
        best    = min(all_cards, key=lambda c: c.get("p20") or 999999)
        best_p20 = best.get("p20", 0)

        # 5. Gom ETD dates tương ứng với best price
        etd_dates_best = sorted(set(
            c["etd_date"] for c in all_cards
            if c.get("p20") == best_p20))

        # Parse transit days
        def _parse_transit(txt):
            m = re.search(r'(\d+)', txt or '')
            return int(m.group(1)) if m else 0

        transit_list = sorted(set(
            _parse_transit(c.get("transitTime", ""))
            for c in all_cards
            if c.get("p20") == best_p20
            and _parse_transit(c.get("transitTime", "")) > 0))

        # Format cho Excel
        etd_text     = format_etd_for_excel(etd_dates_best)
        transit_text = format_transit_for_excel(transit_list)
        valid_for    = calculate_valid_for_espot(etd_dates_best)
        valid_text   = (f"{valid_for.day}-{valid_for.strftime('%b')}"
                        if valid_for else "")

        transship_str = "TRANSSHIP" if best.get("isTransship") else "DIRECT"
        ft_pod        = normalize_port_text(pod)

        result = {
            "price_20":      best_p20,
            "price_40":      best.get("p40") or 0,
            "price_40hq":    best.get("p40hq") or 0,
            "etd_dates":     etd_dates_best,
            "standard_etd_ok": any(
                d >= (datetime.now().date() + timedelta(days=DATE_OFFSET_DAYS))
                for d in etd_dates_best
            ),
            "etd_text":      etd_text,
            "transit_text":  transit_text,
            "valid_text":    valid_text,
            "source":        "E-Spot",
            "remark":        "",
            "ft_pod":        ft_pod,
            "vessel_info":   best.get("vessel", ""),
            "transshipment": transship_str,
        }
        print(f"[OK] E-Spot best → 20GP={best_p20} USD | "
              f"vessel={best.get('vessel')} | ETDs={etd_text} | "
              f"transit={transit_text}")
        return result

    def _scrape_equote(self, pol="", pod="", country="", skip_sailing=False):
        """
        Đọc giá E-Quote.
        - skip_sailing=True  : Chỉ lấy giá + surcharge + remark (nhanh, không lấy sailing)
        - skip_sailing=False : Lấy full (có Place Booking + sailing)
        """
        print("[INFO] ── E-Quote: Bắt đầu phase đọc giá ──")
        self.driver.switch_to.window(self.equote_tab)
        time.sleep(0.3)

        cards = self._collect_equote_cards()
        if not cards:
            print("[INFO] E-Quote: Không có card nào trong bảng Recent Quotations.")
            return None

        valid_cards = self._filter_valid_equote_cards(cards)
        if not valid_cards:
            print("[INFO] E-Quote: Không có card nào còn valid >= 7 ngày.")
            return None

        detail_results = []
        for card in valid_cards[:1]:   # Chỉ xử lý card rẻ nhất
            detail_tab = self._open_equote_detail_from_card(card)
            if not detail_tab:
                continue

            try:
                self.driver.switch_to.window(detail_tab)
                WebDriverWait(self.driver, 12).until(
                    lambda d: "e-spot-detail" in (d.current_url or ""))
                time.sleep(0.8)

                ocean = self._read_equote_ocean_freight()
                ft_pod = self._read_equote_free_time_pod()

                if not self._go_to_equote_surcharges_tab():
                    surcharge = {
                        "add_20": 0, "add_40": 0, "add_40hq": 0,
                        "has_origin_thc": False, "ows_items": []
                    }
                else:
                    surcharge = self._read_equote_surcharges(country=country, pod=pod)

                remark = self._build_equote_remark(
                    has_origin_thc=surcharge.get("has_origin_thc", False),
                    ows_items=surcharge.get("ows_items", []),
                    country=country,
                    pod=pod,
                )

                quote_validity_end = (card.get("validity") or {}).get("to")
                surcharge_variants = self._build_equote_surcharge_variants(surcharge)

                for variant_idx, variant in enumerate(surcharge_variants, start=1):
                    total_20 = ocean.get("20GP", 0) + variant.get("add_20", 0)
                    total_40 = ocean.get("40GP", 0) + variant.get("add_40", 0)
                    total_40hq = ocean.get("40HQ", 0) + variant.get("add_40hq", 0)
                    formula_20 = _excel_formula_from_parts([ocean.get("20GP", 0), variant.get("add_20", 0)])
                    formula_40 = _excel_formula_from_parts([ocean.get("40GP", 0), variant.get("add_40", 0)])
                    formula_40hq = _excel_formula_from_parts([ocean.get("40HQ", 0), variant.get("add_40hq", 0)])

                    selected_tiers = variant.get("selected_tiers") or []
                    tier_validity_starts = [t.get("date_from") for t in selected_tiers if t.get("date_from")]
                    tier_validity_ends = [t.get("date_to") for t in selected_tiers if t.get("date_to")]
                    selected_validity_start = max(tier_validity_starts, default=None)
                    selected_validity_end = min(
                        [d for d in [quote_validity_end, *tier_validity_ends] if d],
                        default=quote_validity_end,
                    )

                    detail_results.append({
                        "ocean": ocean,
                        "surcharge": variant,
                        "ft_pod": ft_pod,
                        "total_20": total_20,
                        "total_40": total_40,
                        "total_40hq": total_40hq,
                        "formula_20": formula_20,
                        "formula_40": formula_40,
                        "formula_40hq": formula_40hq,
                        "remark": remark,
                        "detail_tab": detail_tab,
                        "validity_end": quote_validity_end,
                        "selected_validity_start": selected_validity_start,
                        "selected_validity_end": selected_validity_end,
                        "via_ports": card.get("via_ports", []),
                        "tier_variant_label": variant.get("tier_variant_label", ""),
                    })

                    print(
                        f"[OK] Card #{card.get('index')} tier#{variant_idx}: "
                        f"20={total_20} 40={total_40} 40HQ={total_40hq} | "
                        f"{variant.get('tier_variant_label', '')}"
                    )

            except Exception as e:
                print(f"[WARN] E-Quote detail error: {e}")
                err_msg = str(e).lower()
                if "no such window" in err_msg or "window already closed" in err_msg:
                    # Tab đã bị browser đóng → KHÔNG gọi _close_tab_safe (tránh treo timeout)
                    print("[INFO] Tab detail đã bị đóng → bỏ qua cleanup, switch về E-Quote tab.")
                    try:
                        self.driver.switch_to.window(self.equote_tab)
                    except Exception:
                        pass
                else:
                    _close_tab_safe(self.driver, detail_tab, self.equote_tab)
                continue

        if not detail_results:
            print("[INFO] E-Quote: Không đọc được detail hợp lệ nào.")
            return None

        detail_results.sort(key=lambda x: (
            x["total_20"],
            x.get("total_40") or 0,
            x.get("total_40hq") or 0,
            x.get("selected_validity_start") or date_type.max,
        ))
        best = detail_results[0]
        print(
            f"[INFO] E-Quote cheapest tier: 20GP={best['total_20']} | "
            f"{best.get('tier_variant_label', '')}"
        )
        print(f"[INFO] E-Quote chọn card rẻ nhất: 20GP={best['total_20']}")

        # ====================== TỐI ƯU: Nếu E-Spot rẻ hơn thì bỏ qua sailing ======================
        if skip_sailing:
            print("[OPTIMIZE] skip_sailing=True → Chỉ lấy giá, không lấy sailing")
            result = {
                "price_20": best["total_20"],
                "price_40": best["total_40"],
                "price_40hq": best["total_40hq"],
                "formula_20": best.get("formula_20"),
                "formula_40": best.get("formula_40"),
                "formula_40hq": best.get("formula_40hq"),
                "etd_text": "",
                "transit_text": "",
                "valid_text": f"{best['selected_validity_end'].day}-{best['selected_validity_end'].strftime('%b')}" if best.get("selected_validity_end") else "",
                "source": "E-Quote",
                "remark": best["remark"],
                "ft_pod": best["ft_pod"],
                "vessel_info": "",
                "transshipment": "",
            }
            _close_tab_safe(self.driver, best["detail_tab"], self.equote_tab)
            self.driver.switch_to.window(self.equote_tab)
            return result
        # =====================================================================================

        # ================== Chỉ chạy khi cần lấy sailing ==================
        print("[INFO] E-Quote rẻ hơn hoặc ngang E-Spot → tiếp tục lấy sailing...")

        # Đóng tab detail thừa
        for tab_handle in {item.get("detail_tab") for item in detail_results if item.get("detail_tab")}:
            if tab_handle == best["detail_tab"]:
                continue
            try:
                _close_tab_safe(self.driver, tab_handle, best["detail_tab"])
            except:
                pass

        self.driver.switch_to.window(best["detail_tab"])

        booking_tab = self._click_place_booking_from_detail()
        if not booking_tab:
            print("[WARN] Không mở được booking tab.")
            result = self._build_minimal_equote_result(best)
            _close_tab_safe(self.driver, best["detail_tab"], self.equote_tab)
            return result

        self.driver.switch_to.window(booking_tab)

        try:
            for candidate in detail_results:
                print(
                    f"[INFO] Thu sailing cho tier: 20GP={candidate['total_20']} | "
                    f"{candidate.get('tier_variant_label', '')}"
                )
                sailings = self._collect_equote_sailings_with_date_retries(
                    target_start=candidate.get("selected_validity_start"),
                    target_end=candidate.get("selected_validity_end"),
                )
                if not sailings:
                    print("[INFO] Tier nay khong co sailing raw, thu tier tiep theo.")
                    continue

                ets_tier = (
                    candidate["surcharge"].get("selected_tiers")
                    or candidate["surcharge"].get("ets_tiers")
                    or []
                )
                filtered = self._filter_etd_only(
                    sailings,
                    validity_end=candidate.get("selected_validity_end"),
                    ets_tier=ets_tier,
                )
                filtered = self._dedup_by_etd(filtered)
                filtered = self._select_spaced_sailings(filtered)
                if not filtered:
                    print("[INFO] Tier nay khong con ETD sau khi loc valid, thu tier tiep theo.")
                    continue

                expanded = []
                for s in filtered[:6]:
                    expanded.append(self._expand_sailing_card(s))

                matched = self._filter_sailings_by_via_ports(expanded, candidate.get("via_ports", []))
                chosen_sailings = matched[:3] if matched else expanded[:3]
                etd_dates = [s.get("etd_date") for s in chosen_sailings if s.get("etd_date")]
                if not etd_dates:
                    print("[INFO] Tier nay parse duoc card nhung khong co ETD hop le, thu tier tiep theo.")
                    continue

                transit_list = [s.get("transit_days") for s in chosen_sailings if s.get("transit_days") is not None]
                ts_ports = [s.get("ts_port") for s in chosen_sailings if s.get("ts_port")]

                vessel_lines = []
                for s in chosen_sailings:
                    v = s.get("vessel_name")
                    if v:
                        etd_d = s.get("etd_date")
                        etd_str = f"{etd_d.day}-{etd_d.strftime('%b')}" if etd_d else ""
                        td = s.get("transit_days")
                        td_text = f"{td} Days" if td is not None else "N/A"
                        ts = s.get("ts_port") or "DIRECT"
                        vessel_lines.append(
                            f"{v} / ETD: {etd_str} / Transit time: {td_text} / Transshipment: {ts}"
                        )

                result = {
                    "price_20": candidate["total_20"],
                    "price_40": candidate["total_40"],
                    "price_40hq": candidate["total_40hq"],
                    "formula_20": candidate.get("formula_20"),
                    "formula_40": candidate.get("formula_40"),
                    "formula_40hq": candidate.get("formula_40hq"),
                    "etd_text": format_etd_for_excel(etd_dates),
                    "transit_text": format_transit_for_excel(transit_list),
                    "valid_text": (
                        f"{candidate['selected_validity_end'].day}-{candidate['selected_validity_end'].strftime('%b')}"
                        if candidate.get("selected_validity_end") else ""
                    ),
                    "source": "E-Quote",
                    "remark": candidate["remark"],
                    "ft_pod": candidate["ft_pod"],
                    "vessel_info": "\n".join(vessel_lines),
                    "transshipment": ", ".join(sorted(set([p for p in ts_ports if p]))) or "DIRECT",
                }

                print(f"[OK] E-Quote result: {result}")
                return result

            print("[INFO] Tat ca tier E-Quote deu khong co sailing hop le -> tra gia re nhat khong lich.")
            return self._build_minimal_equote_result(detail_results[0])

            sailings = self._collect_equote_sailings_with_date_retries()
            if not sailings:
                print("[INFO] Không có sailing schedule → trả kết quả không có lịch tàu.")
                result = self._build_minimal_equote_result(best)
                return result
            ets_tier = best["surcharge"].get("selected_tiers") or best["surcharge"].get("ets_tiers") or []
            
            sailings = self._filter_etd_only(sailings, validity_end=best.get("selected_validity_end"), ets_tier=ets_tier)
            sailings = self._dedup_by_etd(sailings)
            sailings = self._select_spaced_sailings(sailings)

            expanded = []
            for s in sailings[:6]:
                expanded.append(self._expand_sailing_card(s))

            # FIX: Khởi tạo an toàn
            matched = self._filter_sailings_by_via_ports(expanded, best.get("via_ports", []))
            chosen_sailings = matched[:3] if matched else expanded[:3]

            # ================== XÂY DỰNG RESULT AN TOÀN ==================
            etd_dates = [s.get("etd_date") for s in chosen_sailings if s.get("etd_date")]
            transit_list = [s.get("transit_days") for s in chosen_sailings if s.get("transit_days") is not None]
            ts_ports = [s.get("ts_port") for s in chosen_sailings if s.get("ts_port")]

            # Format vessel_info giống e-spot: tên tàu / ETD / Transit time / Transshipment
            vessel_lines = []
            for s in chosen_sailings:
                v = s.get("vessel_name")
                if v:
                    etd_d = s.get("etd_date")
                    etd_str = f"{etd_d.day}-{etd_d.strftime('%b')}" if etd_d else ""
                    td = s.get("transit_days")
                    td_text = f"{td} Days" if td is not None else "N/A"
                    ts = s.get("ts_port") or "DIRECT"
                    vessel_lines.append(
                        f"{v} / ETD: {etd_str} / Transit time: {td_text} / Transshipment: {ts}"
                    )

            result = {
                "price_20": best["total_20"],
                "price_40": best["total_40"],
                "price_40hq": best["total_40hq"],
                "formula_20": best.get("formula_20"),
                "formula_40": best.get("formula_40"),
                "formula_40hq": best.get("formula_40hq"),
                "etd_text": format_etd_for_excel(etd_dates),
                "transit_text": format_transit_for_excel(transit_list),
                "valid_text": f"{best['selected_validity_end'].day}-{best['selected_validity_end'].strftime('%b')}" if best.get("selected_validity_end") else "",
                "source": "E-Quote",
                "remark": best["remark"],
                "ft_pod": best["ft_pod"],
                "vessel_info": "\n".join(vessel_lines),
                "transshipment": ", ".join(sorted(set([p for p in ts_ports if p]))) or "DIRECT",
            }

            print(f"[OK] E-Quote result: {result}")
            return result

        finally:
            # Cleanup an toàn
            for tab in [booking_tab, best["detail_tab"]]:
                try:
                    if tab and tab in self.driver.window_handles:
                        _close_tab_safe(self.driver, tab, self.equote_tab)
                except:
                    pass
            try:
                self.driver.switch_to.window(self.equote_tab)
            except:
                pass

    def _choose_best_result(self, espot_result, equote_result):
        candidates = []
        for item in [espot_result, equote_result]:
            if not item:
                continue
            try:
                p20 = int(item.get("price_20") or 0)
                if p20 > 0:
                    candidates.append(item)
            except Exception:
                continue

        if not candidates:
            return None

        espot_standard = [
            x for x in candidates
            if x.get("source") == "E-Spot" and x.get("standard_etd_ok")
        ]
        if espot_standard:
            best_spot = min(espot_standard, key=lambda x: int(x.get("price_20") or 99999999))
            print(
                f"[PRIORITY] Chọn E-Spot trước E-Quote vì E-Spot có ETD chuẩn "
                f">= today+{DATE_OFFSET_DAYS}: {best_spot.get('etd_text')} | 20GP={best_spot.get('price_20')}"
            )
            return best_spot

        scheduled = [
            x for x in candidates
            if str(x.get("etd_text") or "").strip()
            and str(x.get("transit_text") or "").strip()
        ]
        if scheduled:
            candidates = scheduled

        candidates.sort(key=lambda x: int(x.get("price_20") or 99999999))
        best = candidates[0]
        print(f"[INFO] Chọn kết quả tốt nhất: {best.get('source')} | 20GP={best.get('price_20')}")
        return best
    
    def _build_minimal_equote_result(self, best):
        if not best:
            return None
        return {
            "price_20": best.get("total_20", 0),
            "price_40": best.get("total_40", 0),
            "price_40hq": best.get("total_40hq", 0),
            "formula_20": best.get("formula_20"),
            "formula_40": best.get("formula_40"),
            "formula_40hq": best.get("formula_40hq"),
            "etd_text": "",
            "transit_text": "",
            "valid_text": f"{best['selected_validity_end'].day}-{best['selected_validity_end'].strftime('%b')}" if best.get("selected_validity_end") else "",
            "source": "E-Quote",
            "remark": best.get("remark", ""),
            "ft_pod": best.get("ft_pod", ""),
            "vessel_info": "",
            "transshipment": "",
            "schedule_ok": False,
        }

    def _clone_equote_surcharge_variant(self, surcharge, selected_tiers):
        selected_tiers = list(selected_tiers or [])
        base_20 = surcharge.get("base_add_20")
        base_40 = surcharge.get("base_add_40")
        base_40hq = surcharge.get("base_add_40hq")
        if base_20 is None:
            old_tiers = surcharge.get("selected_tiers") or []
            base_20 = (surcharge.get("add_20") or 0) - sum(t.get("amount_20") or 0 for t in old_tiers)
            base_40 = (surcharge.get("add_40") or 0) - sum(t.get("amount_40") or 0 for t in old_tiers)
            base_40hq = (surcharge.get("add_40hq") or 0) - sum(t.get("amount_40hq") or 0 for t in old_tiers)

        variant = dict(surcharge)
        variant["selected_tiers"] = selected_tiers
        variant["ets_tiers"] = selected_tiers
        variant["add_20"] = (base_20 or 0) + sum(t.get("amount_20") or 0 for t in selected_tiers)
        variant["add_40"] = (base_40 or 0) + sum(t.get("amount_40") or 0 for t in selected_tiers)
        variant["add_40hq"] = (base_40hq or 0) + sum(t.get("amount_40hq") or 0 for t in selected_tiers)
        if selected_tiers:
            variant["tier_variant_label"] = " | ".join(
                f"{t.get('code') or t.get('charge_name')}:"
                f"{t.get('date_from')}->{t.get('date_to')}"
                for t in selected_tiers
            )
        else:
            variant["tier_variant_label"] = "no tier"
        return variant

    def _build_equote_surcharge_variants(self, surcharge):
        tier_groups = surcharge.get("tier_groups") or {}
        groups = []
        for group_key, tiers in tier_groups.items():
            cleaned = [t for t in tiers or [] if t.get("date_from") and t.get("date_to")]
            if not cleaned:
                continue
            cleaned.sort(key=lambda x: (
                (x.get("amount_20") or 0) + (x.get("amount_40") or 0) + (x.get("amount_40hq") or 0),
                x.get("date_from") or date_type.max,
            ))
            groups.append(cleaned)

        if not groups:
            return [self._clone_equote_surcharge_variant(surcharge, surcharge.get("selected_tiers") or [])]

        variants = [
            self._clone_equote_surcharge_variant(surcharge, combo)
            for combo in product(*groups)
        ]
        variants.sort(key=lambda x: (
            x.get("add_20") or 0,
            x.get("add_40") or 0,
            x.get("add_40hq") or 0,
            min((t.get("date_from") or date_type.max) for t in x.get("selected_tiers") or [{}]),
        ))
        return variants

    def _parse_money_text(self, text):
        s = (text or "").strip().upper()
        if not s or "---" in s:
            return 0
        nums = re.findall(r'[\d,.]+', s)
        if not nums:
            return 0
        raw = nums[0].replace(",", "")
        try:
            return _ceil_money_value(raw)
        except Exception:
            return 0    

    def _parse_card_validity(self, text):
        try:
            clean = " ".join((text or "").split())
            m = re.search(
                r'(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})\s+to\s+(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})',
                clean
            )
            if not m:
                return {"from": None, "to": None, "raw": text}
            d_from = datetime.strptime(m.group(1), "%d %b %Y").date()
            d_to   = datetime.strptime(m.group(2), "%d %b %Y").date()
            return {"from": d_from, "to": d_to, "raw": text}
        except Exception:
            return {"from": None, "to": None, "raw": text}

    def _parse_via_ports(self, via_text):
        raw = (via_text or "").strip()
        if not raw:
            return []   # [] nghĩa là wildcard: transship ở đâu cũng được
        ports = []
        for p in raw.split(","):
            pp = p.strip().upper()
            if pp and pp not in ports:
                ports.append(pp)
        return ports

    def _collect_equote_cards(self):
        cards = []
        try:
            state = self._get_equote_result_state()
            if state.get("empty") and not state.get("rowCount"):
                print("[INFO] E-Quote: trang báo không có quotation/card.")
                return []

            WebDriverWait(self.driver, OOCL_EQUOTE_COLLECT_WAIT_SECONDS).until(
                lambda d: (
                    len(d.find_elements(
                        By.CSS_SELECTOR,
                        "div.ant-table-body table tbody tr.ant-table-row.ant-table-row-level-0")) > 0
                    or len(d.find_elements(
                        By.CSS_SELECTOR,
                        "div.ant-empty, div[class*='no-data']")) > 0
                    or self._get_equote_result_state().get("empty")
                )
            )

            debug_print(f"[DEBUG] current E-Quote URL: {self.driver.current_url}")
            debug_print(f"[DEBUG] table-container count: {len(self.driver.find_elements(By.CSS_SELECTOR, 'div.table-container'))}")
            debug_print(f"[DEBUG] ant-table-body count: {len(self.driver.find_elements(By.CSS_SELECTOR, 'div.ant-table-body'))}")
            debug_print(f"[DEBUG] all tr.ant-table-row count: {len(self.driver.find_elements(By.CSS_SELECTOR, 'tr.ant-table-row'))}")
            debug_print(f"[DEBUG] scoped row count: {len(self.driver.find_elements(By.CSS_SELECTOR, 'div.ant-table-body table tbody tr.ant-table-row.ant-table-row-level-0'))}")

            rows = self.driver.find_elements(
                By.CSS_SELECTOR,
                "div.ant-table-body table tbody tr.ant-table-row.ant-table-row-level-0"
            )
            rows = [r for r in rows if r.is_displayed()]

            debug_print(f"[DEBUG] visible row count after filter: {len(rows)}")

            for idx, row in enumerate(rows):
                try:
                    tds = row.find_elements(By.CSS_SELECTOR, "td.ant-table-cell")
                    debug_print(f"[DEBUG] row#{idx}: td count = {len(tds)}")

                    if len(tds) < 14:
                        debug_print(f"[DEBUG] row#{idx}: skip vì quá ít td ({len(tds)})")
                        continue

                    type_text = tds[0].text.strip().upper()
                    debug_print(f"[DEBUG] row#{idx}: type='{type_text}'")

                    if "E-QUOTE" not in type_text:
                        continue

                    # Lấy cell theo index thực tế của body row
                    validity_text = tds[4].text.strip() if len(tds) > 4 else ""
                    via_text      = tds[8].text.strip() if len(tds) > 8 else ""
                    p20_text      = tds[11].text.strip() if len(tds) > 11 else ""
                    p40_text      = tds[12].text.strip() if len(tds) > 12 else ""
                    p40hq_text    = tds[13].text.strip() if len(tds) > 13 else ""

                    card = {
                        "index": idx,
                        "row_el": row,
                        "tds": tds,
                        "validity_text": validity_text,
                        "via_text": via_text,
                        "via_ports": self._parse_via_ports(via_text),
                        "table_p20": self._parse_money_text(p20_text),
                        "table_p40": self._parse_money_text(p40_text),
                        "table_p40hq": self._parse_money_text(p40hq_text),
                        "validity": self._parse_card_validity(validity_text),
                    }
                    cards.append(card)

                    debug_print(
                        f"[DEBUG] row#{idx}: valid='{validity_text}' | via='{via_text}' | "
                        f"20={card['table_p20']} 40={card['table_p40']} 40HQ={card['table_p40hq']}"
                    )

                except Exception as e:
                    print(f"[WARN] _collect_equote_cards row#{idx}: {e}")
                    continue

            print(f"[INFO] E-Quote: đọc được {len(cards)} card trong Recent Quotations.")
            return cards

        except Exception as e:
            print(f"[WARN] _collect_equote_cards: {e}")
            return []

    def _filter_valid_equote_cards(self, cards):
        today = datetime.now().date()
        min_valid_to = today + timedelta(days=DATE_OFFSET_DAYS)
        result = []

        for c in cards:
            valid_to = (c.get("validity") or {}).get("to")
            if not valid_to:
                print(f"[INFO] Skip card #{c['index']}: không parse được ngày valid.")
                continue
            if valid_to < min_valid_to:
                print(f"[INFO] Skip card #{c['index']}: valid_to={valid_to} < {min_valid_to}")
                continue
            result.append(c)

        print(f"[INFO] E-Quote valid cards: {len(cards)} → {len(result)}")
        return result

    def _open_equote_detail_from_card(self, card):
        try:
            idx = card.get("index", -1)

            # CARD #N được parse trên E-Quote tab. Khi đang xử lý card trước đó
            # bot có thể đang ở detail tab → element handle thuộc window khác
            # nên Selenium raise NoSuchElement. Switch về E-Quote tab trước,
            # rồi re-fetch row theo index để tránh stale element.
            try:
                self.driver.switch_to.window(self.equote_tab)
            except Exception as e:
                print(f"[WARN] _open_equote_detail_from_card #{idx}: switch về E-Quote tab lỗi: {e}")

            row = card.get("row_el")
            try:
                _probe = row.find_elements(By.CSS_SELECTOR, "td.ant-table-cell") if row else None
                row_ok = bool(_probe)
            except Exception:
                row_ok = False

            if not row_ok:
                try:
                    rows = self.driver.find_elements(
                        By.CSS_SELECTOR,
                        "div.ant-table-body table tbody tr.ant-table-row.ant-table-row-level-0"
                    )
                    rows = [r for r in rows if r.is_displayed()]
                    if 0 <= idx < len(rows):
                        row = rows[idx]
                        card["row_el"] = row
                        debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: re-fetch row OK.")
                    else:
                        print(f"[WARN] _open_equote_detail_from_card #{idx}: re-fetch row fail (index ngoài range).")
                        return None
                except Exception as e:
                    print(f"[WARN] _open_equote_detail_from_card #{idx}: re-fetch row lỗi: {e}")
                    return None

            tds = row.find_elements(By.CSS_SELECTOR, "td.ant-table-cell")
            debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: td count = {len(tds)}")

            if len(tds) < 2:
                print(f"[WARN] _open_equote_detail_from_card #{idx}: row không đủ td.")
                return None

            # Lấy td cuối cùng làm cột Actions
            action_td = tds[-1]
            debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: dùng td cuối cùng làm action cell.")

            buttons = action_td.find_elements(By.CSS_SELECTOR, "button")
            debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: button count trong action td = {len(buttons)}")

            if not buttons:
                # fallback: tìm button trong toàn row
                row_buttons = row.find_elements(By.CSS_SELECTOR, "button")
                debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: fallback row button count = {len(row_buttons)}")
                if not row_buttons:
                    print(f"[WARN] _open_equote_detail_from_card #{idx}: không tìm thấy button action.")
                    return None
                btn = row_buttons[-1]
            else:
                btn = buttons[0]

            before = set(self.driver.window_handles)

            # Scroll đến row trước
            try:
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row)
                time.sleep(0.2)
            except Exception:
                pass

            def _visible_detail_item():
                """Tìm <li> 'View Details' đang visible trong các dropdown đang mở."""
                # Ưu tiên data-menu-id='DETAIL'
                for it in self.driver.find_elements(
                    By.CSS_SELECTOR,
                    "ul.ant-dropdown-menu li[data-menu-id='DETAIL']"
                ):
                    try:
                        if it.is_displayed():
                            return it
                    except Exception:
                        continue
                # Fallback: match theo text
                for it in self.driver.find_elements(By.CSS_SELECTOR, "ul.ant-dropdown-menu li"):
                    try:
                        if not it.is_displayed():
                            continue
                        if "VIEW DETAILS" in (it.text or "").strip().upper():
                            return it
                    except Exception:
                        continue
                return None

            actions = ActionChains(self.driver)

            # Bước 1: dùng ActionChains để hover thật (CDP gửi pointermove +
            # mouseenter/mouseover). Ant Design dropdown nghe React synthetic
            # events nên dispatchEvent('mouseover') JS thường không trigger.
            detail_item = None
            try:
                actions.move_to_element(btn).perform()
            except Exception as e:
                debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: hover via ActionChains lỗi: {e}")

            # Chờ dropdown xuất hiện sau hover (trigger='hover')
            try:
                WebDriverWait(self.driver, 2).until(lambda d: _visible_detail_item() is not None)
                detail_item = _visible_detail_item()
                debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: dropdown hiện sau hover.")
            except TimeoutException:
                detail_item = None

            # Bước 2: nếu hover không mở (trigger='click'), thử click thật nút "..."
            if detail_item is None:
                try:
                    actions.move_to_element(btn).click().perform()
                except Exception:
                    try:
                        btn.click()
                    except Exception:
                        try:
                            self.driver.execute_script("arguments[0].click();", btn)
                        except Exception as e:
                            print(f"[WARN] _open_equote_detail_from_card #{idx}: click button fail: {e}")
                            return None
                try:
                    WebDriverWait(self.driver, 3).until(lambda d: _visible_detail_item() is not None)
                    detail_item = _visible_detail_item()
                    debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: dropdown hiện sau click.")
                except TimeoutException:
                    detail_item = None

            dropdowns = self.driver.find_elements(By.CSS_SELECTOR, "ul.ant-dropdown-menu")
            debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: dropdown count = {len(dropdowns)}")

            if detail_item is None:
                print(f"[WARN] _open_equote_detail_from_card #{idx}: không tìm thấy menu 'View Details'.")
                return None

            debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: click View Details.")

            # Hover qua item rồi click thật — tránh việc rời chuột làm menu đóng
            # giữa lúc đang chuẩn bị click.
            click_ok = False
            try:
                ActionChains(self.driver).move_to_element(detail_item).pause(0.1).click().perform()
                click_ok = True
            except Exception as e1:
                debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: click View Details qua ActionChains lỗi: {e1}")

            if not click_ok:
                try:
                    detail_item.click()
                    click_ok = True
                except Exception as e2:
                    try:
                        self.driver.execute_script("arguments[0].click();", detail_item)
                        click_ok = True
                    except Exception as e3:
                        print(f"[WARN] _open_equote_detail_from_card #{idx}: click View Details fail: {e2} / {e3}")
                        return None

            WebDriverWait(self.driver, 8).until(
                lambda d: len(d.window_handles) > len(before)
            )
            after = set(self.driver.window_handles)
            new_tabs = list(after - before)

            debug_print(f"[DEBUG] _open_equote_detail_from_card #{idx}: new tab count = {len(new_tabs)}")

            if not new_tabs:
                print(f"[WARN] _open_equote_detail_from_card #{idx}: không có tab mới sau click View Details.")
                return None

            new_tab = new_tabs[-1]
            print(f"[OK] Mở detail tab card #{idx}: {new_tab[-6:]}")
            return new_tab

        except Exception as e:
            print(f"[WARN] _open_equote_detail_from_card #{card.get('index')}: {e}")
            return None

    def read_oocl_rows(self, path, sheet):
        rows = []
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet]
            # Port mapping: Excel name → OOCL search name
            OOCL_PORT_MAPPING = {
                "TIANJIN": "XINGANG",
                "INCHEON": "INCHON",
                "VENICE": "VENEZIA",
                "GENOA": "GENOVA",
                "NAPLES": "NAPOLI",
                "TAMATAVE": "TOAMASINA",
                "ANTWERP": "ANTWERPEN",
            }
            eligible_idx = 0
            for r_idx in range(2, ws.max_row + 1):
                if SINGLE_ROW and str(r_idx) != SINGLE_ROW:
                    continue
                carrier = (ws.cell(row=r_idx, column=5).value or "").strip().upper()
                pol     = (ws.cell(row=r_idx, column=3).value or "").strip()
                pod     = (ws.cell(row=r_idx, column=4).value or "").strip()
                country = (ws.cell(row=r_idx, column=2).value or "").strip().upper()
                if carrier == "OOCL" and pol and pod:
                    if FILTER_POL and pol.upper() != FILTER_POL:
                        continue
                    if FILTER_POD and pod.upper() != FILTER_POD:
                        continue
                    if not country:
                        country = os.environ.get("FILTER_COUNTRY", "").strip().upper()
                    if OOCL_WORKER_COUNT > 1 and not SINGLE_ROW:
                        if eligible_idx % OOCL_WORKER_COUNT != OOCL_WORKER_INDEX:
                            eligible_idx += 1
                            continue
                        eligible_idx += 1
                    # Áp dụng port mapping cho search (giữ tên gốc trong Excel output)
                    pol_search = OOCL_PORT_MAPPING.get(pol.upper(), pol)
                    pod_search = OOCL_PORT_MAPPING.get(pod.upper(), pod)
                    rows.append({"row": r_idx, "pol": pol_search, "pod": pod_search, "country": country})
                    print(f"[INFO] Row {r_idx}: POL={pol_search} POD={pod_search} COUNTRY={country}")
        except Exception as e:
            print(f"[ERROR] read_oocl_rows: {e}")
        return rows
    
    def _read_equote_ocean_freight(self):
        result = {"20GP": 0, "40GP": 0, "40HQ": 0}
        # Detail tab is SPA — wait briefly for the section to render
        try:
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//*[normalize-space(text())='Ocean Freight']")
                )
            )
        except Exception:
            pass
        try:
            js = """
            const out = {"20GP":0,"40GP":0,"40HQ":0};

            const titles = Array.from(document.querySelectorAll("div.title span"));
            let title = null;
            for (const t of titles) {
                if ((t.textContent || "").trim() === "Ocean Freight") {
                    title = t;
                    break;
                }
            }
            if (!title) return out;

            const container = title.closest("div.container");
            if (!container) return out;

            // Row 1 of Ocean Freight has a leading td with rowspan="3"
            // (route name spans all 3 container rows), so per/price columns
            // shift between rows. Detect per/price by content rather than
            // by fixed index.
            const rows = container.querySelectorAll("tbody tr");
            for (const tr of rows) {
                const tds = Array.from(tr.querySelectorAll("td"));
                if (tds.length < 2) continue;

                // Find the td whose text matches a known container size.
                let perTd = null;
                let perKey = null;
                for (const td of tds) {
                    const tx = (td.textContent || "").trim().toUpperCase();
                    if (tx === "20GP" || tx === "40GP" || tx === "40HQ") {
                        perTd = td;
                        perKey = tx;
                        break;
                    }
                }
                if (!perTd) continue;

                // Find a td whose text contains a USD price (skip the per td)
                let usdNum = 0;
                for (const td of tds) {
                    if (td === perTd) continue;
                    const tx = (td.textContent || "").trim().toUpperCase();
                    if (!tx.includes("USD")) continue;
                    // First numeric chunk after USD
                    const m = tx.match(/USD\\s*([\\d,\\.]+)/);
                    if (m) {
                            usdNum = Math.ceil(parseFloat(m[1].replace(/,/g, "")) || 0);
                        break;
                    }
                }
                if (usdNum > 0) out[perKey] = usdNum;
            }
            return out;
            """
            result = self.driver.execute_script(js) or result
            print(f"[INFO] Ocean Freight: {result}")
        except Exception as e:
            print(f"[WARN] _read_equote_ocean_freight: {e}")
        return result

    def _read_equote_free_time_pod(self):
        try:
            js = r"""
            const result = {ft_pod: ""};

            const titles = Array.from(document.querySelectorAll("div.title span"));
            let title = null;
            for (const t of titles) {
                if ((t.textContent || "").trim() === "Free Time") {
                    title = t;
                    break;
                }
            }
            if (!title) return result;

            const container = title.closest("div.container");
            if (!container) return result;

            function formatFreeTime(txt) {
                txt = (txt || "").replace(/\s+/g, " ").trim();
                if (!txt) return "";

                const dem = txt.match(/\bDEM(?:URRAGE)?\b\s*(\d+)\s*(?:CD|CALENDAR\s*DAYS?)?/i);
                const det = txt.match(/\bDET(?:ENTION)?\b\s*(\d+)\s*(?:CD|CALENDAR\s*DAYS?)?/i);
                if (dem && det) return `${dem[1]} DEM + ${det[1]} DET`;

                const dd2in1 = txt.match(/(\d+)\s+Calendar\s+days?\s+for\s+DD2in1/i);
                if (dd2in1) return `${dd2in1[1]} COMBINED`;

                const combined = txt.match(/\b(?:COMBINED|DD2IN1)\b.*?(\d+)\s*(?:CD|CALENDAR\s*DAYS?)?/i)
                    || txt.match(/(\d+)\s*(?:CD|CALENDAR\s*DAYS?).*\b(?:COMBINED|DD2IN1)\b/i);
                if (combined) return `${combined[1]} COMBINED`;

                return "";
            }

            const rows = container.querySelectorAll("tbody tr");
            let scope = "";
            let destinationText = "";
            for (const tr of rows) {
                const tds = tr.querySelectorAll("td");
                if (tds.length < 2) continue;

                const item = (tds[0].textContent || "").trim().toUpperCase();
                const val  = (tds[1].textContent || "").trim();
                const rowText = (tr.textContent || "").trim();

                if (item === "DESTINATION") {
                    scope = "DESTINATION";
                    destinationText += " " + val;
                } else if (item === "ORIGIN" || item === "LOADING" || item === "POL") {
                    scope = item;
                } else if (scope === "DESTINATION") {
                    destinationText += " " + rowText;
                }
            }
            const ft = formatFreeTime(destinationText);
            if (ft) result.ft_pod = ft;
            return result;
            """
            res = self.driver.execute_script(js) or {}
            ft = res.get("ft_pod", "")
            print(f"[INFO] Free Time POD: {ft}")
            return ft
        except Exception as e:
            print(f"[WARN] _read_equote_free_time_pod: {e}")
            return ""

    def _go_to_equote_surcharges_tab(self):
        try:
            tab = WebDriverWait(self.driver, 8).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//li[@data-menu-id='Charges']"
                ))
            )
            self.driver.execute_script("arguments[0].click();", tab)

            WebDriverWait(self.driver, 8).until(
                lambda d: len(d.find_elements(
                    By.XPATH,
                    "//span[normalize-space()='Surcharge Items']"
                )) > 0
            )
            print("[OK] Đã mở tab Surcharges Details.")
            time.sleep(0.5)
            return True
        except Exception as e:
            print(f"[WARN] _go_to_equote_surcharges_tab: {e}")
            return False
        
    def _parse_ows_item(self, charge_name, detail, code, pay, curr, v20, v40, v40hq):
        try:
            name_u = (charge_name or "").upper()
            detail_u = (detail or "").upper()
            code_u = (code or "").upper()
            curr_u = (curr or "").upper().strip()

            if curr_u != "USD":
                return None

            if not (
                any(k in name_u for k in EQUOTE_OWS_KEYWORDS)
                or "ONLY IF WEIGHT" in detail_u
                or code_u in ("CWC", "CWX", "OWS")
            ):
                return None

            amt = self._parse_money_text(v20)
            if amt <= 0:
                return None

            threshold = ""
            m = re.search(r'AT OR ABOVE\s+(\d+)\s*NET\s*KG', detail_u)
            if m:
                tons = int(m.group(1)) / 1000.0
                if float(tons).is_integer():
                    tons = int(tons)
                threshold = f"(>{tons}TONS)"

            text = f"OWS ${amt}/20'GP"
            if threshold:
                text += f" {threshold}"

            return {
                "amount_20": amt,
                "code": code_u,
                "threshold": threshold,
                "text": text.strip(),
            }
        except Exception:
            return None
    def _parse_ets_tier(self, charge_name, detail, code, pay, curr, v20, v40, v40hq):
        try:
            name_u = (charge_name or "").upper()
            code_u = (code or "").upper()
            detail_u = (detail or "").upper()

            if not ("ETS" in name_u or code_u == "ETS"):
                return None

            m = re.search(r'VALID FROM (\d{4}-\d{2}-\d{2}) TO (\d{4}-\d{2}-\d{2})', detail_u)
            if not m:
                return None

            return {
                "charge_name": charge_name,
                "code": code_u,
                "date_from": datetime.strptime(m.group(1), "%Y-%m-%d").date(),
                "date_to": datetime.strptime(m.group(2), "%Y-%m-%d").date(),
                "amount_20": self._parse_money_text(v20),
                "amount_40": self._parse_money_text(v40),
                "amount_40hq": self._parse_money_text(v40hq),
            }
        except Exception:
            return None

    def _parse_valid_surcharge_tier(self, charge_name, detail, code, pay, curr, v20, v40, v40hq):
        try:
            if (curr or "").strip().upper() != "USD":
                return None
            if (pay or "").strip().upper() not in ("PREPAID", "ANY"):
                return None

            m = re.search(r'VALID FROM (\d{4}-\d{2}-\d{2}) TO (\d{4}-\d{2}-\d{2})', (detail or "").upper())
            if not m:
                return None

            name_u = (charge_name or "").upper()
            if any(x in name_u for x in EQUOTE_SURCHARGE_HARD_EXCLUDE):
                return None

            code_u = (code or "").strip().upper()
            group_key = code_u or name_u
            return {
                "group_key": group_key,
                "charge_name": charge_name,
                "code": code_u,
                "date_from": datetime.strptime(m.group(1), "%Y-%m-%d").date(),
                "date_to": datetime.strptime(m.group(2), "%Y-%m-%d").date(),
                "amount_20": self._parse_money_text(v20),
                "amount_40": self._parse_money_text(v40),
                "amount_40hq": self._parse_money_text(v40hq),
            }
        except Exception:
            return None

    def _read_equote_surcharges(self, country="", pod=""):
        china_route = is_china_destination(country, pod)
        result = {
            "add_20": 0,
            "add_40": 0,
            "add_40hq": 0,
            "has_origin_thc": False,
            "ows_items": [],
            "ets_tiers": [],
            "tier_groups": {},
            "selected_tiers": [],
            "raw_items": [],
        }

        try:
            rows = self.driver.find_elements(
                By.CSS_SELECTOR,
                "div.container table tbody tr.ant-table-row"
            )

            parent_name = ""

            for idx, tr in enumerate(rows):
                try:
                    tds = tr.find_elements(By.CSS_SELECTOR, "td.ant-table-cell")
                    if len(tds) < 7:
                        continue

                    detail = tds[0].text.strip()
                    code   = tds[1].text.strip().upper()
                    pay    = tds[2].text.strip().upper()
                    curr   = tds[3].text.strip().upper()

                    v20   = tds[4].text.strip()
                    v40   = tds[5].text.strip()
                    v40hq = tds[6].text.strip()

                    if detail and not detail.upper().startswith("VALID FROM"):
                        parent_name = detail

                    charge_name = detail
                    if detail.upper().startswith("VALID FROM") and parent_name:
                        charge_name = parent_name

                    item = {
                        "detail": detail,
                        "charge_name": charge_name,
                        "code": code,
                        "payment": pay,
                        "currency": curr,
                        "v20": self._parse_money_text(v20),
                        "v40": self._parse_money_text(v40),
                        "v40hq": self._parse_money_text(v40hq),
                        "raw_v20": v20,
                        "raw_v40": v40,
                        "raw_v40hq": v40hq,
                    }
                    result["raw_items"].append(item)

                    name_u = (charge_name or "").upper()

                    is_origin_thc = "TERMINAL HANDLING CHARGE AT ORIGIN" in name_u
                    if is_origin_thc:
                        result["has_origin_thc"] = True

                    ows = self._parse_ows_item(charge_name, detail, code, pay, curr, v20, v40, v40hq)
                    if ows:
                        result["ows_items"].append(ows)
                        continue

                    tier = self._parse_valid_surcharge_tier(charge_name, detail, code, pay, curr, v20, v40, v40hq)
                    if tier:
                        result["tier_groups"].setdefault(tier["group_key"], []).append(tier)
                        continue

                    if any(x in name_u for x in EQUOTE_SURCHARGE_HARD_EXCLUDE) and not (china_route and is_origin_thc):
                        continue

                    if pay not in ("PREPAID", "ANY") and not (china_route and is_origin_thc):
                        continue

                    if curr.strip().upper() != "USD" and not (china_route and is_origin_thc):
                        continue

                    if china_route and is_origin_thc:
                        add_20 = charge_amount_to_usd(item["v20"], curr)
                        add_40 = charge_amount_to_usd(item["v40"], curr)
                        add_40hq = charge_amount_to_usd(item["v40hq"], curr)
                        print(
                            f"[INFO] +O.THC CHINA {curr}: "
                            f"20={add_20:.2f} 40={add_40:.2f} 40HQ={add_40hq:.2f} USD"
                        )
                    else:
                        add_20, add_40, add_40hq = item["v20"], item["v40"], item["v40hq"]

                    result["add_20"] += add_20
                    result["add_40"] += add_40
                    result["add_40hq"] += add_40hq

                except Exception as e:
                    print(f"[WARN] _read_equote_surcharges row#{idx}: {e}")
                    continue

            print(f"[INFO] Surcharge cộng thêm: 20={result['add_20']} 40={result['add_40']} 40HQ={result['add_40hq']}")
            result["base_add_20"] = result["add_20"]
            result["base_add_40"] = result["add_40"]
            result["base_add_40hq"] = result["add_40hq"]
            selected_tiers = []
            for group_key, tiers in result["tier_groups"].items():
                if not tiers:
                    continue
                chosen = min(
                    tiers,
                    key=lambda x: (
                        (x.get("amount_20") or 0) + (x.get("amount_40") or 0) + (x.get("amount_40hq") or 0),
                        x.get("date_from") or date_type.max,
                    ),
                )
                selected_tiers.append(chosen)
                result["add_20"] += chosen.get("amount_20") or 0
                result["add_40"] += chosen.get("amount_40") or 0
                result["add_40hq"] += chosen.get("amount_40hq") or 0
                print(
                    f"[INFO] Tier {group_key}: chọn {chosen['date_from']} -> {chosen['date_to']} "
                    f"(20={chosen.get('amount_20')}, 40={chosen.get('amount_40')}, 40HQ={chosen.get('amount_40hq')})"
                )

            result["selected_tiers"] = selected_tiers
            result["ets_tiers"] = selected_tiers

            print(f"[INFO] Surcharge final: 20={result['add_20']} 40={result['add_40']} 40HQ={result['add_40hq']}")
            print(f"[INFO] has_origin_thc={result['has_origin_thc']} | OWS items={len(result['ows_items'])}")
            return result

        except Exception as e:
            print(f"[WARN] _read_equote_surcharges: {e}")
            return result
        
    def _get_manifest_remark_by_country(self, country):
        return get_manifest_code(country or "")

    def _build_equote_remark(self, has_origin_thc, ows_items, country, pod=""):
        manifest = self._get_manifest_remark_by_country(country)

        if has_origin_thc and not is_china_destination(country, pod):
            prefix = "SUBJECT TO THC, BILL, SEAL, TLX"
        else:
            prefix = "INCLUDED O.THC, SUBJECT TO BILL, SEAL, TLX"

        if manifest:
            prefix += f", {manifest}"

        parts = [prefix]

        seen = set()
        for item in (ows_items or []):
            txt = (item or {}).get("text", "").strip()
            if txt and txt not in seen:
                seen.add(txt)
                parts.append(txt)

        return ", ".join(parts)
    
    def _collect_espot_prices_via_popup(self, qualified, country="", pod=""):
        """
        Với từng ETD đủ điều kiện:
        - click ETD
        - parse toàn bộ options hợp lệ
        - chọn best option
        - click Details
        - đọc popup
        - cộng surcharge
        - đóng popup
        Trả về list kết quả từng ETD
        """
        results = []

        for date_str, ref_price, _ in qualified:
            print(f"[INFO] E-Spot: Click ETD {date_str} (ref={ref_price})...")
            try:
                if not self._click_etd_date(date_str):
                    continue
                time.sleep(0.8)
            except Exception as e:
                print(f"[WARN] Click ETD {date_str}: {e}")
                continue

            options = self._parse_espot_cards_current_etd()
            if not options:
                print(f"[INFO]   ETD {date_str}: không có option hợp lệ.")
                continue

            best = self._choose_best_espot_option(options)
            if not best:
                continue

            if not self._click_espot_details(best):
                continue

            overview = self._parse_espot_popup_overview()

            charge_data = {"add_20": 0, "add_40": 0, "add_40hq": 0}
            if self._go_to_espot_charge_breakdown_tab():
                charge_data = self._parse_espot_charge_breakdown(country=country, pod=pod)

            conditional_data = {"has_origin_thc": True, "ows_items": []}
            if self._go_to_espot_conditional_tab():
                conditional_data = self._parse_espot_conditional_charges()

            remark = self._build_espot_remark(country, conditional_data, pod=pod)

            total_20 = int(best.get("price_20") or 0) + int(charge_data.get("add_20") or 0)
            total_40 = int(best.get("price_40") or 0) + int(charge_data.get("add_40") or 0)
            total_40hq = int(best.get("price_40hq") or 0) + int(charge_data.get("add_40hq") or 0)
            formula_20 = _excel_formula_from_parts([best.get("price_20") or 0, charge_data.get("add_20") or 0])
            formula_40 = _excel_formula_from_parts([best.get("price_40") or 0, charge_data.get("add_40") or 0])
            formula_40hq = _excel_formula_from_parts([best.get("price_40hq") or 0, charge_data.get("add_40hq") or 0])

            self._close_espot_popup()

            results.append({
                "date_str": date_str,
                "ref_price": ref_price,
                "price_20": total_20,
                "price_40": total_40,
                "price_40hq": total_40hq,
                "formula_20": formula_20,
                "formula_40": formula_40,
                "formula_40hq": formula_40hq,
                "transit": int(overview.get("transit") or best.get("transit") or 0),
                "vessel": overview.get("vessel_name") or best.get("vessel") or "",
                "ft_pod": overview.get("ft_pod") or "",
                "transshipment": overview.get("transshipment_port") or "",
                "remark": remark,
            })

            print(f"[OK]   ETD {date_str}: TOTAL 20={total_20} 40={total_40} 40HQ={total_40hq} "
                  f"vessel={overview.get('vessel_name') or best.get('vessel')}")

        return results

    def _click_place_booking_from_detail(self):
        try:
            before = set(self.driver.window_handles)

            btn = WebDriverWait(self.driver, 8).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//button[.//span[normalize-space()='Place Booking']]"
                ))
            )
            self.driver.execute_script("arguments[0].click();", btn)

            WebDriverWait(self.driver, 10).until(
                lambda d: len(d.window_handles) > len(before)
            )
            new_tab = list(set(self.driver.window_handles) - before)[-1]
            self.driver.switch_to.window(new_tab)

            if not self._handle_login_after_place_booking():
                return None

            WebDriverWait(self.driver, 20).until(
                lambda d: "booking-request" in (d.current_url or "")
            )
            print(f"[OK] Place Booking mở tab mới: {new_tab[-6:]}")
            return new_tab

        except Exception as e:
            print(f"[WARN] _click_place_booking_from_detail: {e}")
            return None

    def _handle_login_after_place_booking(self):
        try:
            def is_booking_or_login(d):
                url = d.current_url or ""
                if "booking-request" in url:
                    return True
                login_controls = d.find_elements(
                    By.XPATH,
                    "//input[@type='email'] | //input[@name='email'] | "
                    "//input[@type='password'] | "
                    "//button[contains(translate(normalize-space(),"
                    "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')] | "
                    "//button[contains(translate(normalize-space(),"
                    "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]"
                )
                return len(login_controls) > 0

            WebDriverWait(self.driver, 20).until(is_booking_or_login)
            if "booking-request" in (self.driver.current_url or ""):
                return True

            print("[INFO] Place Booking redirected to login, signing in again...")
            login_btn = self._find_first_xpath([
                "//button[contains(translate(normalize-space(),"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
                "//button[contains(translate(normalize-space(),"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
                "//a[contains(translate(normalize-space(),"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
                "//a[contains(translate(normalize-space(),"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
            ], timeout=3, clickable=True)
            if login_btn:
                try:
                    safe_click(self.driver, login_btn, delay=0.2)
                    time.sleep(1.0)
                    if "booking-request" in (self.driver.current_url or ""):
                        return True
                except Exception:
                    pass

            ok = self._do_login(success_url_keywords=OOCL_APP_URL_MARKERS)
            if not ok:
                return False
            if "booking-request" in (self.driver.current_url or ""):
                return True
            WebDriverWait(self.driver, 10).until(
                lambda d: "booking-request" in (d.current_url or ""))
            return True
        except Exception as e:
            print(f"[WARN] _handle_login_after_place_booking: {e}")
            return False

    def _filter_sailings_by_via_ports(self, sailings, via_ports):
        if not via_ports:
            return sailings

        via_set = {x.upper() for x in via_ports if x}
        matched = []
        for s in sailings:
            ts = (s.get("ts_port") or "").upper()
            if ts in via_set:
                matched.append(s)

        print(f"[INFO] Filter theo via ports: {len(sailings)} → {len(matched)}")
        return matched

    def set_sailing_date(self, target_date=None):
        target_date = target_date or (datetime.now().date() + timedelta(days=DATE_OFFSET_DAYS))
        target_str  = target_date.strftime("%Y-%m-%d")
        print(f"[INFO] Chọn ngày sailing: {target_str}")
        try:
            WebDriverWait(self.driver, 20).until(
                lambda d: "booking-request" in d.current_url)

            inp = WebDriverWait(self.driver, 15).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, "div.ant-picker input")))
            safe_click(self.driver, inp)
            r()

            WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "div.ant-picker-panel")))

            for td in self.driver.find_elements(
                    By.CSS_SELECTOR, "td.ant-picker-cell"):
                if td.get_attribute("title") == target_str:
                    safe_click(self.driver, td)
                    print(f"[OK] Đã chọn ngày {target_str}.")
                    r()
                    return True

            print(f"[WARN] Không tìm thấy ngày {target_str}.")
            return False

        except Exception as e:
            print(f"[ERROR] set_sailing_date: {e}")
            return False

    def set_sailing_date_exact(self, target_date):
        target_str = target_date.strftime("%Y-%m-%d")
        print(f"[INFO] Chọn ngày sailing: {target_str}")
        try:
            WebDriverWait(self.driver, 10).until(
                lambda d: "booking-request" in d.current_url)

            inp = WebDriverWait(self.driver, 8).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "div.ant-picker input")))
            safe_click(self.driver, inp)
            r()

            WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.ant-picker-panel")))

            for _ in range(3):
                for td in self.driver.find_elements(By.CSS_SELECTOR, "td.ant-picker-cell"):
                    if td.get_attribute("title") == target_str:
                        safe_click(self.driver, td)
                        print(f"[OK] Đã chọn ngày {target_str}.")
                        r()
                        return True
                next_btns = self.driver.find_elements(
                    By.CSS_SELECTOR,
                    "button.ant-picker-header-next-btn, button.ant-picker-header-super-next-btn"
                )
                if not next_btns:
                    break
                safe_click(self.driver, next_btns[0])
                time.sleep(0.4)

            print(f"[WARN] Không tìm thấy ngày {target_str}.")
            return False

        except TimeoutException:
            print(f"[WARN] set_sailing_date_exact: timeout khi chọn {target_str}.")
            return False
        except Exception as e:
            print(f"[ERROR] set_sailing_date_exact: {e}")
            return False

    def click_find_sailing(self):
        FIND_XPATH = (
            "/html/body/div[1]/div/section/div/div/div/div/main/div"
            "/div/div/div/div[2]/div/div[1]/div[2]/div[2]/div/div[1]"
            "/div/button")
        try:
            try:
                btn = WebDriverWait(self.driver, 8).until(
                    EC.element_to_be_clickable((By.XPATH, FIND_XPATH)))
            except TimeoutException:
                btn = WebDriverWait(self.driver, 8).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//button[.//span[contains(translate(normalize-space(),"
                        "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),"
                        "'find sailing')]]"
                    )))
            safe_click(self.driver, btn)
            print("[OK] Đã click Find Sailing.")
            try:
                WebDriverWait(self.driver, 15).until(
                    lambda d: (
                        len(d.find_elements(
                            By.CSS_SELECTOR, "div.voyage-record-card")) > 0
                        or "no suitable sailing" in (d.find_element(
                            By.TAG_NAME, "body").text or "").lower()
                    ))
            except TimeoutException:
                pass
            # Kiểm tra "No suitable sailing schedule"
            try:
                body_text = self.driver.find_element(By.TAG_NAME, "body").text or ""
                if "no suitable sailing" in body_text.lower():
                    time.sleep(3)
                    if self.driver.find_elements(By.CSS_SELECTOR, "div.voyage-record-card"):
                        return True
                    print("[INFO] No suitable sailing schedule — không có lịch tàu.")
                    return False
            except Exception:
                pass
            r()
            return True
        except Exception as e:
            print(f"[ERROR] click_find_sailing: {e}")
            return False

    def _collect_equote_sailings_with_date_retries(self, target_start=None, target_end=None):
        today = datetime.now().date()
        # Đồng bộ với _filter_etd_only(): lịch quá gần (< today+6) cuối cùng cũng bị loại.
        # Nếu vẫn search từ --date +3 thì OOCL thường tốn 8-12s cho một ngày chắc chắn vô ích.
        base_date = max(
            today + timedelta(days=DATE_OFFSET_DAYS),
            today + timedelta(days=6),
        )
        if target_start:
            base_date = max(base_date, target_start)
        if target_end and base_date > target_end:
            print(f"[INFO] E-Quote tier window {target_start} -> {target_end} khong con ngay search hop le.")
            return []
        for extra_days in (0, 3, 7, 14, 21):
            target_date = base_date + timedelta(days=extra_days)
            if target_end and target_date > target_end:
                continue
            if not self.set_sailing_date_exact(target_date):
                continue
            if not self.click_find_sailing():
                print(f"[INFO] E-Quote không có lịch ở ngày {target_date:%Y-%m-%d}, thử ngày khác...")
                continue
            self.click_more_filter()
            sailings = self.parse_all_sailing_cards()
            if sailings:
                print(f"[OK] E-Quote tìm được {len(sailings)} sailing từ ngày {target_date:%Y-%m-%d}.")
                return sailings
            print(f"[INFO] E-Quote không parse được card ở ngày {target_date:%Y-%m-%d}, thử ngày khác...")
        return []

    def click_more_filter(self):
        """
        Bung filter "More Filter" trong trang booking dể thấy thêm cột/tuỳ chọn
        sailing. Trên trang mới nút này có thể dùng nhãn "More Filter",
        "More Filters" hoặc bị ẩn nếu UI thay đổi — bắt bằng XPath theo text,
        không fail cứng nếu không tìm thấy (đây là bước tùy chọn).
        """
        try:
            btn = None
            try:
                btn = WebDriverWait(self.driver, 4).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//button[.//span[contains(translate(normalize-space(),"
                        "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),"
                        "'more filter')]]"
                    ))
                )
            except TimeoutException:
                # Fallback XPath tuyệt đối (giữ lại từ version cũ)
                MORE_FILTER_XPATH = (
                    "/html/body/div[1]/div/section/div/div/div/div/main/div"
                    "/div/div/div/div[2]/div/div[2]/div[1]/div[2]/div[4]/button")
                try:
                    btn = WebDriverWait(self.driver, 3).until(
                        EC.element_to_be_clickable((By.XPATH, MORE_FILTER_XPATH)))
                except TimeoutException:
                    print("[INFO] click_more_filter: không tìm thấy nút 'More Filter' (bỏ qua).")
                    return
            safe_click(self.driver, btn)
            try:
                WebDriverWait(self.driver, 5).until(
                    lambda d: "hide" in (btn.text or "").lower())
            except Exception:
                pass
            print("[OK] Đã click More Filter."); r()
        except Exception as e:
            print(f"[WARN] click_more_filter: {e}")

    def parse_all_sailing_cards(self):
        cards_el = self.driver.find_elements(
            By.CSS_SELECTOR, "div.voyage-record-card")
        print(f"[INFO] {len(cards_el)} sailing cards.")
        sailings = []
        for idx, card_el in enumerate(cards_el):
            etd_date, etd_text, transit_days = self._read_etd_summary(card_el, idx)
            if not etd_date:
                continue
            print(f"[INFO] Card #{idx} ETD={etd_text} Transit={transit_days}d")
            sailings.append({
                "index": idx,
                "element": card_el,
                "etd_text": etd_text,
                "etd_date": etd_date,
                "transit_days": transit_days,
                "vessel_name": "",
                "ts_port": "",
            })
        return sailings
    
    def _read_etd_summary(self, card_el, idx):
        etd_date = None
        etd_text = ""
        transit_days = None
        try:
            # Tìm ETD
            for item in card_el.find_elements(By.CSS_SELECTOR, "div.ss-item"):
                head = item.find_elements(By.CSS_SELECTOR, "div.ss-item__head")
                if head and any("ETD" in (h.get_attribute("title") or "") for h in head):
                    raw = head[0].text.strip()
                    d = parse_sailing_date(raw)
                    if d:
                        etd_date = d
                        etd_text = f"{d.day}-{d.strftime('%b')}"
                    break

            # Tìm Transit Time
            for item in card_el.find_elements(By.CSS_SELECTOR, "div.ss-item"):
                if "Transit Time" in item.text:
                    strongs = item.find_elements(By.TAG_NAME, "strong")
                    for s in strongs:
                        m = re.search(r'(\d+)', s.text)
                        if m:
                            transit_days = int(m.group(1))
                            break
                    if transit_days is not None:
                        break

            # Fallback: tìm transit time từ toàn bộ text của card
            if transit_days is None:
                full_text = card_el.text
                # Pattern: "Transit Time" theo sau bởi số ngày
                m = re.search(r'Transit\s*Time[:\s]*(\d+)\s*[Dd]ay', full_text, re.IGNORECASE)
                if m:
                    transit_days = int(m.group(1))
                else:
                    # Pattern: số + "Days" đứng riêng
                    m = re.search(r'(\d+)\s*[Dd]ays?', full_text)
                    if m:
                        val = int(m.group(1))
                        if 3 <= val <= 90:
                            transit_days = val

        except Exception as e:
            print(f"[WARN] _read_etd_summary #{idx}: {e}")

        return etd_date, etd_text, transit_days

    def _expand_sailing_card(self, sailing):
        card_el = sailing.get("element")
        idx = sailing.get("index", 0)
        try:
            # Click expand nếu chưa mở
            try:
                expand_btn = card_el.find_element(By.CSS_SELECTOR, "button.ant-btn-link span.anticon-caret-right")
                if expand_btn:
                    expand_btn.click()
                    time.sleep(0.8)
            except:
                pass

            full_text = card_el.text.strip()

            # ====================== LẤY TÀU ĐẦU TIÊN (từ POL) ======================
            vessel = "N/A"

            def _valid_vessel_name(txt):
                t = re.sub(r"\s+", " ", str(txt or "").strip())
                u = t.upper()
                if not t or len(t) < 6:
                    return False
                bad_words = [
                    "TO BE ADVISED", "TBA", "TO BE ANNOUNCED", "TO BE CONFIRMED",
                    "CUT-OFF", "CUTOFF", "CY CUTOFF", "SI CUTOFF",
                    "SERVICE", "VESSEL VOYAGE", "GENERAL", "DAYS",
                ]
                if any(w in u for w in bad_words):
                    return False
                if re.search(r"\d{4}-\d{2}-\d{2}", u):
                    return False
                if re.search(r"\((MON|TUE|WED|THU|FRI|SAT|SUN)\)", u):
                    return False
                if re.fullmatch(r"[A-Z]{2,5}\d{0,2}", u):
                    return False
                return bool(re.search(r"[A-Z]", u) and re.search(r"\d", u))
            
            # Ưu tiên tìm trong deadline-card (chính xác nhất)
            deadline = card_el.find_elements(By.CSS_SELECTOR, "div.deadline-card")
            if deadline:
                bold_spans = deadline[0].find_elements(By.CSS_SELECTOR, "span.tw-font-bold")
                for span in bold_spans:
                    txt = span.text.strip()
                    # Điều kiện lọc tên tàu (không phải ngày, cutoff, service code thuần)
                    if _valid_vessel_name(txt):
                        vessel = txt
                        break   # Chỉ lấy tàu đầu tiên

            # Fallback nếu không tìm thấy
            if vessel == "N/A":
                all_bold = card_el.find_elements(By.CSS_SELECTOR, "span.tw-font-bold")
                for el in all_bold:
                    txt = el.text.strip()
                    if _valid_vessel_name(txt):
                        vessel = txt
                        break

            sailing["vessel_name"] = vessel

            # ====================== Transshipment ======================
            ts_port = ""
            ts_ports_found = []
            for item in card_el.find_elements(By.CSS_SELECTOR, "div.ss-info-item"):
                # Tìm theo label "T/S Port" trong div.desc (chính xác nhất)
                descs = item.find_elements(By.CSS_SELECTOR, "div.desc")
                is_ts = any("T/S" in (d.text or "") for d in descs)
                if not is_ts:
                    # Fallback: tìm keyword trong text
                    txt = item.text.upper()
                    is_ts = "TRANSSHIPMENT" in txt or "T/S PORT" in txt
                if is_ts:
                    name_els = item.find_elements(By.CSS_SELECTOR, "div.name")
                    for n in name_els:
                        p = n.text.strip()
                        if p and len(p) > 2 and p.upper() not in [pp.upper() for pp in ts_ports_found]:
                            ts_ports_found.append(p.upper())
            ts_port = ", ".join(ts_ports_found) if ts_ports_found else ""
            # Fallback: tìm "Transshipment" trong summary line
            if not ts_port:
                for item in card_el.find_elements(By.CSS_SELECTOR, "div.ss-item__line"):
                    line_text = item.text.strip()
                    if "Transshipment" in line_text:
                        # Có transshipment nhưng không tìm được tên port
                        ts_port = "T/S"
                        break

            sailing["ts_port"] = ts_port

            # ETD & Transit
            etd_date, etd_text, transit_days = self._read_etd_summary(card_el, idx)
            sailing["etd_date"] = etd_date
            sailing["etd_text"] = etd_text
            if transit_days:
                sailing["transit_days"] = transit_days

            print(f"[OK] Expanded #{idx}: Vessel='{vessel}' | TS='{ts_port}' | ETD={etd_text}")

        except Exception as e:
            print(f"[ERROR] _expand_sailing_card #{idx}: {e}")
            sailing["vessel_name"] = "N/A"

        return sailing

    def _filter_etd_only(self, sailings, validity_end, ets_tier=None):
        today = datetime.now().date()
        cutoff = today + timedelta(days=6)
        result = []
        if isinstance(ets_tier, (list, tuple)):
            valid_tiers = [t for t in ets_tier if t]
        else:
            valid_tiers = [ets_tier] if ets_tier else []

        for s in sailings:
            d = s.get("etd_date")
            if not d:
                continue
            if d < cutoff:
                print(f"[INFO] Bỏ #{s['index']} ETD={s['etd_text']}: quá gần")
                continue
            if validity_end and d > validity_end:
                print(f"[INFO] Bỏ #{s['index']} ETD={s['etd_text']}: vượt valid end {validity_end}")
                continue
            tier_ok = True
            for tier in valid_tiers:
                if not (tier["date_from"] <= d <= tier["date_to"]):
                    label = tier.get("code") or tier.get("charge_name") or "tier"
                    print(f"[INFO] Bỏ #{s['index']} ETD={s['etd_text']}: ngoài valid tier {label}")
                    tier_ok = False
                    break
            if not tier_ok:
                continue
            result.append(s)

        print(f"[INFO] ETD filter: {len(sailings)} → {len(result)}")
        return result

    def _dedup_by_etd(self, sailings):
        grouped = {}
        for s in sailings:
            d = s.get("etd_date")
            if not d:
                continue
            if d not in grouped:
                grouped[d] = s
            elif (s.get("transit_days") or 9999) < (grouped[d].get("transit_days") or 9999):
                grouped[d] = s

        deduped = sorted(grouped.values(), key=lambda x: x["etd_date"])
        print(f"[INFO] Dedup: {len(sailings)} → {len(deduped)} unique")
        return deduped

    def _select_spaced_sailings(self, sailings, max_count=3, min_gap_days=2, max_window_days=9):
        selected = []
        first_date = None

        for s in sorted(sailings, key=lambda x: x.get("etd_date") or date_type.max):
            d = s.get("etd_date")
            if not d:
                continue
            if first_date and (d - first_date).days > max_window_days:
                break
            if selected and (d - selected[-1]["etd_date"]).days < min_gap_days:
                print(f"[INFO] Bỏ #{s['index']} ETD={s['etd_text']}: sát ETD trước")
                continue
            if not selected:
                first_date = d
            selected.append(s)
            if len(selected) >= max_count:
                break

        print(f"[INFO] Spaced ETD select: {len(sailings)} -> {len(selected)}")
        return selected

    def write_to_excel(self, path, sheet, row, data):
        try:
            from openpyxl.styles import Alignment
            def clean_excel_text(value):
                if value is None:
                    return ""
                if isinstance(value, str):
                    return html.unescape(value).replace("\xa0", " ").strip()
                return value

            wb = openpyxl.load_workbook(path)
            ws = wb[sheet]
            ws.cell(row=row, column=6,  value=_excel_value_with_formula_fallback(data.get("formula_20"), data.get("price_20", 0)))
            ws.cell(row=row, column=7,  value=_excel_value_with_formula_fallback(data.get("formula_40"), data.get("price_40", 0)))
            ws.cell(row=row, column=8,  value=_excel_value_with_formula_fallback(data.get("formula_40hq"), data.get("price_40hq", 0)))
            ws.cell(row=row, column=9,  value=clean_excel_text(data.get("etd_text",    "")))
            ws.cell(row=row, column=10, value=clean_excel_text(data.get("transit_text","")))
            ws.cell(row=row, column=11, value=clean_excel_text(data.get("valid_text",  "")))
            ws.cell(row=row, column=12, value=clean_excel_text(data.get("source",      "")))
            ws.cell(row=row, column=13, value=clean_excel_text(data.get("remark",      "")))
            ws.cell(row=row, column=14, value=clean_excel_text(data.get("ft_pod",      "")))
            ws.cell(row=row, column=15, value=clean_excel_text(data.get("vessel_info", "")))
            ws.cell(row=row, column=16, value=clean_excel_text(data.get("transshipment","")))
            wrap = Alignment(wrap_text=True, vertical="top")
            for col in [13, 15]:
                ws.cell(row=row, column=col).alignment = wrap
            wb.save(path)
            print(f"[OK] Ghi Excel row {row}.")
        except Exception as e:
            print(f"[ERROR] write_to_excel: {e}")

    def _write_no_data(self, path, sheet, row):
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet]
            ws.cell(row=row, column=6, value="-")
            ws.cell(row=row, column=7, value="-")
            ws.cell(row=row, column=8, value="-")
            for col in range(9, 17):
                ws.cell(row=row, column=col, value="")
            wb.save(path)
            print(f"[OK] Ghi '-' row {row}.")
        except Exception as e:
            print(f"[ERROR] _write_no_data: {e}")

    def _process_one_row(self, pol, pod, country, excel_path, sheet_name, row):
        self._abort_if_oocl_blocked(f"process row {row}")
        print(f"[INFO] ── Bắt đầu xử lý: {pol} → {pod} (row {row}) ──")

        # ── Bước 1: Nhập liệu + đọc E-Spot trước ──
        # Tối ưu tốc độ:
        # - Trước đây bot luôn search E-Quote trước khi biết E-Spot có đủ dùng hay chưa.
        # - Với nhiều route, E-Spot đã có giá + ETD chuẩn >= today+DATE_OFFSET_DAYS,
        #   nên search E-Quote bị dư và thường timeout 7s/row.
        # - Vì rule hiện tại ưu tiên E-Spot nếu ETD chuẩn, ta đọc E-Spot trước rồi mới
        #   quyết định có cần đụng E-Quote hay không.
        espot_ok = self.perform_unified_search(pol, pod, country)
        if not espot_ok:
            print(f"[WARN] E-Spot nhập liệu thất bại cho row {row}.")

        espot_result = None
        if espot_ok:
            self.driver.switch_to.window(self.espot_tab)
            espot_result = self._read_espot_prices(
                excel_path, sheet_name, row, country=country, pod=pod
            )

        # Nếu E-Spot đã có giá và ETD chuẩn thì dừng luôn, không search E-Quote nữa.
        # Đây là điểm giảm thời gian lớn nhất cho các tuyến có E-Spot hợp lệ.
        if espot_result and espot_result.get("price_20") and espot_result.get("standard_etd_ok"):
            print(
                f"[PRIORITY] E-Spot có giá + ETD >= today+{DATE_OFFSET_DAYS} "
                f"({espot_result.get('etd_text')}) -> ghi luôn, bỏ qua E-Quote search."
            )
            final_result = self._choose_best_result(espot_result, None)
            self.write_to_excel(excel_path, sheet_name, row, final_result)
            print(f"[OK] ── Hoàn tất row {row}: {pol} → {pod} (nguồn: {final_result.get('source')}) ──")
            return

        # ── Bước 2: Chỉ khi cần mới nhập liệu + đọc E-Quote ──
        equote_result = None
        equote_ok = self.perform_equote_search(pol, pod, country)
        if not equote_ok:
            print(f"[WARN] E-Quote nhập liệu thất bại cho row {row}.")

        if equote_ok:
            self.driver.switch_to.window(self.equote_tab)

            if espot_result and espot_result.get("price_20"):
                espot_p20 = int(espot_result.get("price_20", 999999))
                
                # Chỉ lấy giá E-Quote trước (không lấy sailing)
                equote_temp = self._scrape_equote(
                    pol=pol, pod=pod, country=country, skip_sailing=True
                )
                
                if equote_temp and int(equote_temp.get("price_20", 999999)) > espot_p20:
                    print(f"[OPTIMIZE] E-Spot rẻ hơn ({espot_p20} < {equote_temp['price_20']}) → Bỏ qua lấy lịch tàu E-Quote")
                    equote_result = None
                else:
                    # E-Quote rẻ hơn hoặc bằng → lấy full (có sailing)
                    print(f"[INFO] E-Quote rẻ hơn hoặc ngang → tiếp tục lấy sailing")
                    equote_result = self._scrape_equote(
                        pol=pol, pod=pod, country=country, skip_sailing=False
                    )
            else:
                # Không có E-Spot → lấy full E-Quote
                equote_result = self._scrape_equote(
                    pol=pol, pod=pod, country=country, skip_sailing=False
                )

        # ── Bước 5: Chọn kết quả tốt nhất ──
        final_result = self._choose_best_result(espot_result, equote_result)

        # ── Bước 6: Ghi Excel ──
        if final_result:
            self.write_to_excel(excel_path, sheet_name, row, final_result)
            print(f"[OK] ── Hoàn tất row {row}: {pol} → {pod} (nguồn: {final_result.get('source')}) ──")
        else:
            self._write_no_data(excel_path, sheet_name, row)
            print(f"[OK] ── Hoàn tất row {row}: {pol} → {pod} (không có giá) ──")

    def _authenticated_workspace_tabs_ready(self):
        """Check that both workspace tabs are genuinely signed in, not URL-only."""
        try:
            handles = list(self.driver.window_handles)
            original = self.driver.current_window_handle
        except Exception:
            return False

        has_espot = False
        has_equote = False
        try:
            for handle in handles:
                try:
                    self.driver.switch_to.window(handle)
                    self._abort_if_oocl_blocked("tabs ready scan")
                    if not self._is_oocl_app_ready():
                        continue
                    url = (self.driver.current_url or "").lower()
                    if "my-quotation" in url:
                        has_equote = True
                    elif any(marker in url for marker in (
                        "freightsmart.oocl.com/ui",
                        "freightsmart.oocl.com/digital",
                        "search-result",
                    )):
                        has_espot = True
                except OOCLIpBlockedError:
                    raise
                except Exception:
                    continue
        finally:
            try:
                if original in self.driver.window_handles:
                    self.driver.switch_to.window(original)
            except Exception:
                pass

        print(f"[DEBUG] Tab authenticated check: E-Spot={has_espot}, E-Quote={has_equote}")
        return has_espot and has_equote

    def run(self, excel_path=EXCEL_PATH, sheet_name=SHEET_NAME):
        print("=" * 60)
        print("OOCL COMBINED BOT (E-Spot + E-Quote)")
        print("=" * 60)

        if not self.launch_edge_if_needed():
            raise RuntimeError("Không mở được Edge OOCL")

        if not self.init_browser():
            raise RuntimeError("Không kết nối được Selenium OOCL")
        try:
            self._abort_if_oocl_blocked("run start")
        except OOCLIpBlockedError:
            raise

        try:
            if self._authenticated_workspace_tabs_ready():
                print("[OK] Đã có sẵn 2 tab OOCL đã đăng nhập → bỏ qua login, vào setup luôn.")
            else:
                print("[INFO] Chưa có đủ 2 tab OOCL hợp lệ → kiểm tra session/login.")
                if not self.check_and_login():
                    raise RuntimeError("Đăng nhập OOCL thất bại; không được gộp dữ liệu cũ")
        except OOCLIpBlockedError:
            raise

        # Bước 4: Setup 2 tab cố định
        try:
            self.setup_tabs_with_recovery()
        except OOCLIpBlockedError:
            raise

        oocl_rows = self.read_oocl_rows(excel_path, sheet_name)
        if not oocl_rows:
            print("[ERROR] Không có row OOCL nào trong Excel.")
            return

        max_rows_env = os.getenv("OOCL_MAX_ROWS", "").strip()
        if max_rows_env.isdigit() and int(max_rows_env) > 0:
            limit = int(max_rows_env)
            oocl_rows = oocl_rows[:limit]
            print(f"[INFO] Debug mode: chỉ chạy {limit} row đầu.")

        print(f"[INFO] Tổng cộng {len(oocl_rows)} row OOCL.")

        for item in oocl_rows:
            print(f"\n{'='*60}")
            print(f"[INFO] Row {item['row']}: {item['pol']} → {item['pod']}")
            print(f"{'='*60}")
            try:
                # Health check: driver còn sống không?
                if not self._check_driver_alive(timeout=10):
                    raise RuntimeError(
                        f"msedgedriver OOCL không phản hồi trước row {item['row']}"
                    )
                self._abort_if_oocl_blocked(f"before row {item['row']}")
                self._process_one_row(
                    item["pol"], item["pod"], item["country"],
                    excel_path, sheet_name, item["row"])
            except OOCLIpBlockedError:
                raise
            except Exception as e:
                print(f"[ERROR] Row {item['row']}: {e}")
                # Kiểm tra driver còn sống sau lỗi
                if not self._check_driver_alive(timeout=10):
                    raise RuntimeError(
                        f"msedgedriver OOCL đã chết sau lỗi row {item['row']}"
                    )
                continue

        print("\n" + "=" * 60)
        print("HOÀN THÀNH TẤT CẢ!")
        print("=" * 60)


# ==================== OOCL VALID FALLBACK OVERRIDES ====================
def _oocl_filter_valid_equote_cards_fallback(self, cards):
    today = datetime.now().date()
    min_valid_to = today + timedelta(days=DATE_OFFSET_DAYS)
    result = []
    fallback = []

    for c in cards:
        valid_to = (c.get("validity") or {}).get("to")
        if not valid_to:
            print(f"[INFO] Skip card #{c.get('index')}: cannot parse valid_to.")
            continue
        if valid_to < min_valid_to:
            if valid_to >= today:
                fallback.append(c)
                print(f"[INFO] Fallback candidate card #{c.get('index')}: valid_to={valid_to} < {min_valid_to}")
            else:
                print(f"[INFO] Skip card #{c.get('index')}: valid_to={valid_to} < today {today}")
            continue
        result.append(c)

    if not result and fallback:
        max_valid_to = max((c.get("validity") or {}).get("to") for c in fallback)
        result = [c for c in fallback if (c.get("validity") or {}).get("to") == max_valid_to]
        print(
            f"[INFO] E-Quote fallback: no card valid >= {min_valid_to}; "
            f"use farthest valid_to={max_valid_to} ({len(result)} card)"
        )

    print(f"[INFO] E-Quote valid cards: {len(cards)} -> {len(result)}")
    return result


def _oocl_collect_equote_sailings_with_date_retries_fallback(self, target_start=None, target_end=None):
    today = datetime.now().date()

    # New E-Quote rule:
    # after Place Booking, OOCL may already show available sailings for the quote.
    # Do not force-select today + DATE_OFFSET_DAYS first; parse what is displayed,
    # then _filter_etd_only_fallback will keep ETDs inside valid and choose the
    # farthest fallback ETD if no standard ETD exists.
    try:
        self.click_more_filter()
        current_sailings = self.parse_all_sailing_cards()
        if current_sailings:
            print(f"[OK] E-Quote use current booking sailings: {len(current_sailings)} card(s), no date reset.")
            return current_sailings
        print("[INFO] E-Quote current booking page has no sailing cards; try date fallback.")
    except Exception as e:
        print(f"[WARN] E-Quote current sailing parse failed: {e}; try date fallback.")

    preferred_base = max(
        today + timedelta(days=DATE_OFFSET_DAYS),
        today + timedelta(days=6),
    )
    if target_start:
        preferred_base = max(preferred_base, target_start)

    search_dates = []
    if target_end and preferred_base > target_end:
        fallback_date = target_end
        if target_start:
            fallback_date = max(fallback_date, target_start)
        if fallback_date < today:
            print(f"[INFO] E-Quote tier window {target_start} -> {target_end} expired.")
            return []
        search_dates = [fallback_date]
        print(
            f"[INFO] E-Quote fallback sailing date: no date >= {preferred_base}; "
            f"use farthest date in valid window {fallback_date}"
        )
    else:
        for extra_days in (0, 3, 7, 14, 21):
            target_date = preferred_base + timedelta(days=extra_days)
            if target_end and target_date > target_end:
                continue
            search_dates.append(target_date)

    for target_date in search_dates:
        if not self.set_sailing_date_exact(target_date):
            continue
        if not self.click_find_sailing():
            print(f"[INFO] E-Quote no sailing at {target_date:%Y-%m-%d}, try another date...")
            continue
        self.click_more_filter()
        sailings = self.parse_all_sailing_cards()
        if sailings:
            print(f"[OK] E-Quote found {len(sailings)} sailing from {target_date:%Y-%m-%d}.")
            return sailings
        print(f"[INFO] E-Quote cannot parse sailing card at {target_date:%Y-%m-%d}, try another date...")
    return []


def _oocl_filter_etd_only_fallback(self, sailings, validity_end, ets_tier=None):
    today = datetime.now().date()
    cutoff = today + timedelta(days=6)
    result = []
    fallback = []
    if isinstance(ets_tier, (list, tuple)):
        valid_tiers = [t for t in ets_tier if t]
    else:
        valid_tiers = [ets_tier] if ets_tier else []

    def tier_ok_for(d):
        for tier in valid_tiers:
            if not (tier["date_from"] <= d <= tier["date_to"]):
                return False
        return True

    for s in sailings:
        d = s.get("etd_date")
        if not d:
            continue
        if d < today:
            print(f"[INFO] Skip #{s.get('index')} ETD={s.get('etd_text')}: before today")
            continue
        if validity_end and d > validity_end:
            print(f"[INFO] Skip #{s.get('index')} ETD={s.get('etd_text')}: over valid end {validity_end}")
            continue
        if not etd_within_max(d):
            print(f"[INFO] Skip #{s.get('index')} ETD={s.get('etd_text')}: over max ETD {max_etd_date_only()}")
            continue
        if not tier_ok_for(d):
            print(f"[INFO] Skip #{s.get('index')} ETD={s.get('etd_text')}: outside valid tier")
            continue
        if d < cutoff:
            fallback.append(s)
            print(f"[INFO] Fallback candidate sailing #{s.get('index')}: ETD={s.get('etd_text')} < {cutoff}")
            continue
        result.append(s)

    if not result and fallback:
        max_etd = max(s.get("etd_date") for s in fallback if s.get("etd_date"))
        same_day = [s for s in fallback if s.get("etd_date") == max_etd]
        same_day.sort(key=lambda x: (x.get("transit_days") or 9999, x.get("index") or 9999))
        result = same_day[:1]
        print(f"[INFO] E-Quote ETD fallback: use farthest ETD inside valid window {max_etd}")

    print(f"[INFO] ETD filter: {len(sailings)} -> {len(result)}")
    return result


def _oocl_card_has_red_book_button(self, card_el, idx=None):
    """
    OOCL booking sailing is usable only when the card has a red/enabled Book button.
    Grey Book means not bookable, so the ETD must be ignored even if it is inside valid.
    """
    try:
        info = self.driver.execute_script("""
            const root = arguments[0];
            function visible(el) {
                const st = window.getComputedStyle(el);
                const r = el.getBoundingClientRect();
                return st && st.display !== 'none' && st.visibility !== 'hidden' &&
                       st.opacity !== '0' && r.width > 0 && r.height > 0;
            }
            function rgbRedish(value) {
                const m = String(value || '').match(/rgba?\\((\\d+),\\s*(\\d+),\\s*(\\d+)/i);
                if (!m) return false;
                const r = Number(m[1]), g = Number(m[2]), b = Number(m[3]);
                return r >= 150 && r > g * 1.25 && r > b * 1.25;
            }
            const nodes = Array.from(root.querySelectorAll('button, a, [role="button"]'));
            for (const el of nodes) {
                const text = (el.innerText || el.textContent || el.getAttribute('aria-label') || '').trim();
                if (!/\\bbook\\b/i.test(text)) continue;
                const cls = String(el.className || '').toLowerCase();
                const disabled = !!el.disabled ||
                    el.getAttribute('disabled') !== null ||
                    el.getAttribute('aria-disabled') === 'true' ||
                    cls.includes('disabled') ||
                    cls.includes('disable');
                if (disabled || !visible(el)) {
                    return {hasBook: true, red: false, reason: 'disabled/hidden', text, cls};
                }
                const st = window.getComputedStyle(el);
                const redByColor = rgbRedish(st.backgroundColor) || rgbRedish(st.color) || rgbRedish(st.borderColor);
                const redByClass = /red|danger|error|bookable|primary/.test(cls);
                if (redByColor || redByClass) {
                    return {
                        hasBook: true,
                        red: true,
                        reason: redByColor ? 'red-color' : 'red-class',
                        text,
                        cls,
                        backgroundColor: st.backgroundColor,
                        color: st.color,
                        borderColor: st.borderColor
                    };
                }
                return {
                    hasBook: true,
                    red: false,
                    reason: 'not-red',
                    text,
                    cls,
                    backgroundColor: st.backgroundColor,
                    color: st.color,
                    borderColor: st.borderColor
                };
            }
            return {hasBook: false, red: false, reason: 'no-book-button'};
        """, card_el) or {}
        if info.get("red"):
            return True
        print(
            f"[INFO] Skip sailing card #{idx}: Book not red/bookable "
            f"({info.get('reason')}, text={info.get('text')!r})"
        )
        return False
    except Exception as e:
        print(f"[INFO] Skip sailing card #{idx}: cannot verify red Book button ({e})")
        return False


def _oocl_parse_all_sailing_cards_bookable_only(self):
    cards_el = self.driver.find_elements(By.CSS_SELECTOR, "div.voyage-record-card")
    print(f"[INFO] {len(cards_el)} sailing cards.")
    sailings = []
    for idx, card_el in enumerate(cards_el):
        if not _oocl_card_has_red_book_button(self, card_el, idx):
            continue
        etd_date, etd_text, transit_days = self._read_etd_summary(card_el, idx)
        if not etd_date:
            continue
        print(f"[INFO] Card #{idx} ETD={etd_text} Transit={transit_days}d (red Book)")
        sailings.append({
            "index": idx,
            "element": card_el,
            "etd_text": etd_text,
            "etd_date": etd_date,
            "transit_days": transit_days,
            "vessel_name": "",
            "ts_port": "",
            "bookable": True,
        })
    print(f"[INFO] Bookable red-Book sailing cards: {len(sailings)}")
    return sailings


OOCLCombinedBot._filter_valid_equote_cards = _oocl_filter_valid_equote_cards_fallback
OOCLCombinedBot._collect_equote_sailings_with_date_retries = _oocl_collect_equote_sailings_with_date_retries_fallback
OOCLCombinedBot._filter_etd_only = _oocl_filter_etd_only_fallback
OOCLCombinedBot.parse_all_sailing_cards = _oocl_parse_all_sailing_cards_bookable_only


# ==================== ENTRY POINT ====================
if __name__ == "__main__":
    bot = OOCLCombinedBot()
    try:
        bot.run(
            excel_path=EXCEL_PATH,
            sheet_name=SHEET_NAME,
        )
    finally:
        bot.keep_browser_on_finish()
