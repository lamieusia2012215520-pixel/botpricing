"""
HPL (Hapag-Lloyd) Price Checker — 3 Tab Pipeline (Optimized)
Flow:
  Bước 1: Tab1 search → Tab2 search → Tab3 search
  Bước 2 (vòng lặp):
    Tab1: lấy giá → back → search mới (chỉ nhập field thay đổi)
    Tab2: lấy giá → back → search mới
    Tab3: lấy giá → back → search mới
    lặp lại cho đến hết
Tối ưu:
  - Alt+← back thay vì hash change → giữ form state
  - Cache POL/POD per tab, chỉ nhập lại field khác biệt
  - Sort queue theo POL để tối đa tái sử dụng
  - Chỉ đọc Price Breakdown 1 lần (cùng QQ price = cùng breakdown)
  - Giảm sleep thừa, bỏ human_scroll() khi search
  - Security check nhẹ (check URL thay vì page_source)
Selenium + Edge port 9525
"""

import math
import re
import sys
import time
import random
import os
import requests
from collections import defaultdict
from datetime import datetime, timedelta

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, WebDriverException
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from bot_runtime_utils import switch_to_live_window
from hpl_logic import hpl_selected_port_matches, jwt_is_expired
from remark_rules import build_subject_remark, is_china_destination
import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass



current_folder = os.getcwd()
driver_path    = os.path.join(current_folder, "msedgedriver.exe")
excel_path     = os.environ.get("EXCEL_PATH", os.path.join(current_folder, "input_gia.xlsx"))
FILTER_POL     = os.environ.get("FILTER_POL", "").strip().upper()
FILTER_POD     = os.environ.get("FILTER_POD", "").strip().upper()
HPL_EMAIL      = os.environ.get("HPL_EMAIL", "").strip()
HPL_PASSWORD   = os.environ.get("HPL_PASSWORD", "")
try:
    SINGLE_ROW = int(os.environ.get("SINGLE_ROW", "0") or 0)
except ValueError:
    SINGLE_ROW = 0
DATE_OFFSET_DAYS = parse_date_offset_days()
try:
    HPL_BACK_WAIT_SECONDS = max(2.0, float(os.environ.get("HPL_BACK_WAIT_SECONDS", "5")))
    HPL_BACK_FALLBACK_WAIT_SECONDS = max(4.0, float(os.environ.get("HPL_BACK_FALLBACK_WAIT_SECONDS", "8")))
    HPL_MANUAL_LOGIN_WAIT_SECONDS = max(30.0, float(os.environ.get("HPL_MANUAL_LOGIN_WAIT_SECONDS", "300")))
except ValueError:
    HPL_BACK_WAIT_SECONDS, HPL_BACK_FALLBACK_WAIT_SECONDS = 5.0, 8.0
    HPL_MANUAL_LOGIN_WAIT_SECONDS = 300.0

def _excel_formula_from_parts(parts):
    tokens = []
    for raw in parts:
        try:
            amount = float(raw)
        except (TypeError, ValueError):
            continue
        if abs(amount) < 1e-9:
            continue
        sign = "-" if amount < 0 else "+"
        amount = abs(amount)
        if abs(amount - round(amount)) < 1e-9:
            text = str(int(round(amount)))
        else:
            text = f"{amount:.2f}".rstrip("0").rstrip(".")
        tokens.append((sign if tokens else ("-" if sign == "-" else "")) + text)
    return "=" + "".join(tokens) if tokens else None

def _blank_price_cells(ws, row_i):
    for col in (6, 7, 8):
        ws.cell(row=row_i, column=col).value = ""

import subprocess
import socket
import urllib.request

HPL_DEBUG_PORT = 9525
HPL_PROFILE_DIR = r"C:\edge_hpl"


def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('127.0.0.1', port)) == 0


def _wait_port(port, timeout=8):
    end = time.time() + timeout
    while time.time() < end:
        if is_port_in_use(port):
            return True
        time.sleep(0.2)
    return False


def _edge_debug_ready(port, timeout=2):
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:{port}/json/version", timeout=timeout) as resp:
            text = resp.read().decode("utf-8", errors="ignore")
        return "webSocketDebuggerUrl" in text or "Microsoft Edge" in text
    except Exception:
        return False


def _launch_hpl_edge():
    print("[H? TH?NG] Edge HPL ch?a s?n s?ng. ?ang kh?i ??ng...")
    try:
        subprocess.Popen([
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            f"--remote-debugging-port={HPL_DEBUG_PORT}",
            f"--user-data-dir={HPL_PROFILE_DIR}",
            "--disable-background-timer-throttling",
            "--disable-renderer-backgrounding",
            "--disable-backgrounding-occluded-windows"
        ])
        _wait_port(HPL_DEBUG_PORT, timeout=10)
    except Exception as e:
        print(f"[H? TH?NG] Kh?ng kh?i ??ng ???c Edge HPL: {e}")


def _restart_hpl_edge():
    print("[H? TH?NG] Edge HPL debug port b? treo, restart ri?ng profile HPL...")
    try:
        subprocess.run([
            "powershell", "-NoProfile", "-Command",
            "$procs = Get-CimInstance Win32_Process | "
            "Where-Object { $_.Name -like 'msedge*' -and "
            "($_.CommandLine -match 'remote-debugging-port=9525' -or $_.CommandLine -match 'edge_hpl') }; "
            "foreach ($p in $procs) { Stop-Process -Id $p.ProcessId -Force -ErrorAction SilentlyContinue }"
        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=8)
    except Exception:
        pass
    time.sleep(1.0)
    _launch_hpl_edge()


if not is_port_in_use(HPL_DEBUG_PORT):
    _launch_hpl_edge()
else:
    print("[H? TH?NG] Edge HPL ?? m? s?n. Ki?m tra debug endpoint...")

if not _edge_debug_ready(HPL_DEBUG_PORT, timeout=2):
    _restart_hpl_edge()
else:
    print("[H? TH?NG] Edge HPL debug endpoint OK. Tai su dung phien HPL hien co.")

edge_options = Options()
edge_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{HPL_DEBUG_PORT}")
service = Service(executable_path=driver_path)
try:
    driver = webdriver.Edge(service=service, options=edge_options)
except Exception as e:
    print(f"[H? TH?NG] Attach Edge HPL l?n 1 th?t b?i: {e}")
    _restart_hpl_edge()
    service = Service(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=edge_options)
try:
    driver.maximize_window()
except Exception:
    try:
        info = driver.execute_cdp_cmd("Browser.getWindowForTarget", {})
        driver.execute_cdp_cmd("Browser.setWindowBounds", {
            "windowId": info["windowId"],
            "bounds": {"windowState": "maximized"}
        })
    except Exception:
        pass

driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": """
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(document, 'visibilityState', {get: () => 'visible'});
        Object.defineProperty(document, 'hidden', {get: () => false});
        document.addEventListener('visibilitychange', e => e.stopImmediatePropagation(), true);
    """
})

BASE_URL  = "https://www.hapag-lloyd.com/solutions/new-quote/#/simple"
HPL_GROUP = {"HPL", "HAPAG", "HAPAG-LLOYD", "HAPAG LLOYD"}
HPL_QUOTE_URL_MARKERS = ("new-quote", "quick-quote", "quickquote", "quick-quotes")
HPL_STEALTH_SCRIPT = """
    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
    Object.defineProperty(document, 'visibilityState', {get: () => 'visible'});
    Object.defineProperty(document, 'hidden', {get: () => false});
    const _origAdd = EventTarget.prototype.addEventListener;
    EventTarget.prototype.addEventListener = function(type, listener, options) {
        if (type === 'visibilitychange' || type === 'blur' || type === 'pagehide') return;
        return _origAdd.call(this, type, listener, options);
    };
"""
try:
    HPL_TAB_REPLACE_ATTEMPTS = max(1, int(os.environ.get("HPL_TAB_REPLACE_ATTEMPTS", "3")))
except ValueError:
    HPL_TAB_REPLACE_ATTEMPTS = 3
try:
    HPL_TAB_LOAD_WAIT_SECONDS = max(10.0, float(os.environ.get("HPL_TAB_LOAD_WAIT_SECONDS", "45")))
except ValueError:
    HPL_TAB_LOAD_WAIT_SECONDS = 45.0
HPL_SAFE_SINGLE_TAB = os.environ.get("HPL_SAFE_SINGLE_TAB", "").strip().lower() in {"1", "true", "yes", "y"}
HPL_SAFE_SINGLE_AFTER_MANUAL = os.environ.get("HPL_SAFE_SINGLE_AFTER_MANUAL", "0").strip().lower() in {"1", "true", "yes", "y"}
HPL_OPEN_MISSING_TABS = os.environ.get("HPL_OPEN_MISSING_TABS", "1").strip().lower() in {"1", "true", "yes", "y"}
HPL_ALLOW_RELOAD = os.environ.get("HPL_ALLOW_RELOAD", "0").strip().lower() in {"1", "true", "yes", "y"}


class HPLTabReplaced(RuntimeError):
    """Raised when a captcha/challenge tab was closed and replaced by a fresh HPL tab."""

class HPLPortNoResults(RuntimeError):
    """Raised when HPL autocomplete explicitly shows No results for a port."""

class HPLServiceUnavailable(RuntimeError):
    """Raised when HPL shows 'This service is currently unavailable'."""

class HPLRoutingUnavailable(RuntimeError):
    """Raised when HPL explicitly says the requested routing is unavailable."""

POL_MAP = {
    "HO CHI MINH": "VNSGN",
    "HAI PHONG":   "VNHPH",
    "DA NANG":     "VNDAD",
    "ĐÀ NẴNG":     "VNDAD",
}

POL_ALIASES = {
    "VNSGN": ["VNSGN", "Ho Chi Minh", "Ho Chi Minh City"],
    "VNHPH": ["VNHPH", "HAIPHONG"],
    "VNDAD": ["VNDAD", "Da Nang", "Danang"],
}

EUR_TO_USD = 1.08
HPL_EXCHANGE_RATE_CACHE = {"USD": 1.0}
HPL_EXCHANGE_RATE_FALLBACK = {
    "EUR": EUR_TO_USD,
    "AUD": 0.65,
    "CHF": 1.12,
    "GBP": 1.27,
    "NZD": 0.60,
    "CNY": 0.14,
    "VND": 0.000038,
}

def get_hpl_exchange_rate(currency):
    base = str(currency or "USD").strip().upper()
    if base in HPL_EXCHANGE_RATE_CACHE:
        return HPL_EXCHANGE_RATE_CACHE[base]
    try:
        response = requests.get(
            f"https://api.frankfurter.app/latest?from={base}&to=USD",
            timeout=8,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        response.raise_for_status()
        rate = float(response.json()["rates"]["USD"])
    except Exception:
        rate = HPL_EXCHANGE_RATE_FALLBACK.get(base)
    if rate is not None:
        HPL_EXCHANGE_RATE_CACHE[base] = rate
    return rate

EUROPE_PORTS = {
    "ALGECIRAS", "ANTWERP", "BARCELONA", "FOS SUR MER", "GENOA",
    "HAMBURG", "LE HAVRE", "NAPOLI", "ROTTERDAM", "VALENCIA", "VENEZIA",
    "FELIXSTOWE", "SOUTHAMPTON", "LONDON GATEWAY", "TILBURY",
    "BREMERHAVEN", "BREMEN", "AMSTERDAM", "ZEEBRUGGE", "DUNKIRK",
    "MARSEILLE", "BORDEAUX", "NANTES", "MONTOIR",
    "BILBAO", "VIGO", "SINES", "LISBON", "LEIXOES", "SETUBAL",
    "GIOIA TAURO", "TARANTO", "LA SPEZIA", "LIVORNO", "TRIESTE",
    "VENICE", "RAVENNA", "SALERNO", "CIVITAVECCHIA",
    "GOTHENBURG", "OSLO", "AARHUS", "COPENHAGEN", "HELSINKI",
    "STOCKHOLM", "TALLINN", "RIGA", "KLAIPEDA", "GDANSK", "GDYNIA",
    "SZCZECIN", "ROSTOCK",
    "PIRAEUS", "THESSALONIKI", "CONSTANTA", "VARNA", "BURGAS",
    "SPLIT", "RIJEKA", "KOPER", "DUBROVNIK",
    "TANGER MED",
    "BELFAST", "DUBLIN", "CORK", "LIVERPOOL", "BRISTOL",
    "BASEL", "BASLE",
}

POD_ALIASES = {
    "FOS SUR MER":    ["Fos", "Fos Sur Mer", "Fos-Sur-Mer"],
    "LE HAVRE":       ["Le Havre", "LeHavre"],
    "MARSEILLE":      ["Marseille", "Marseilles"],
    "DUNKIRK":        ["Dunkirk", "Dunkerque"],
    "MONTOIR":        ["Montoir", "Montoir-de-Bretagne", "Saint-Nazaire"],
    "GENOA":          ["Genoa", "Genova", "Genes"],
    "NAPOLI":         ["Napoli", "Naples", "Naple"],
    "LA SPEZIA":      ["La Spezia", "La-Spezia", "Spezia"],
    "LIVORNO":        ["Livorno", "Leghorn"],
    "TRIESTE":        ["Trieste", "Triest"],
    "VENEZIA":        ["Venezia", "Venice", "Venise"],
    "RAVENNA":        ["Ravenna"],
    "GIOIA TAURO":    ["Gioia Tauro", "Gioia-Tauro", "Gioia"],
    "CIVITAVECCHIA":  ["Civitavecchia", "Rome"],
    "SALERNO":        ["Salerno"],
    "TARANTO":        ["Taranto"],
    "ALGECIRAS":      ["Algeciras"],
    "BARCELONA":      ["Barcelona"],
    "VALENCIA":       ["Valencia"],
    "BILBAO":         ["Bilbao"],
    "VIGO":           ["Vigo"],
    "SINES":          ["Sines"],
    "LISBON":         ["Lisbon", "Lisboa"],
    "LEIXOES":        ["Leixoes", "Porto", "Matosinhos"],
    "ROTTERDAM":      ["Rotterdam"],
    "AMSTERDAM":      ["Amsterdam"],
    "ANTWERP":        ["Antwerp", "Antwerpen", "Anvers"],
    "ZEEBRUGGE":      ["Zeebrugge", "Zeebruges"],
    "HAMBURG":        ["Hamburg"],
    "BREMERHAVEN":    ["Bremerhaven", "Bremen"],
    "FELIXSTOWE":     ["Felixstowe"],
    "SOUTHAMPTON":    ["Southampton"],
    "LONDON GATEWAY": ["London Gateway", "London"],
    "TILBURY":        ["Tilbury"],
    "LIVERPOOL":      ["Liverpool"],
    "GOTHENBURG":     ["Gothenburg", "Goteborg", "Göteborg"],
    "OSLO":           ["Oslo"],
    "AARHUS":         ["Aarhus", "Arhus"],
    "COPENHAGEN":     ["Copenhagen", "Kobenhavn"],
    "HELSINKI":       ["Helsinki"],
    "STOCKHOLM":      ["Stockholm"],
    "GDANSK":         ["Gdansk", "Danzig"],
    "GDYNIA":         ["Gdynia"],
    "KLAIPEDA":       ["Klaipeda"],
    "RIGA":           ["Riga"],
    "TALLINN":        ["Tallinn", "Tallin"],
    "PIRAEUS":        ["Piraeus", "Pireaus", "Pireas", "Athens"],
    "THESSALONIKI":   ["Thessaloniki", "Salonika", "Salonica"],
    "CONSTANTA":      ["Constanta", "Constanța"],
    "KOPER":          ["Koper", "Capodistria"],
    "RIJEKA":         ["Rijeka", "Fiume"],
    "TANGER MED":     ["Tanger Med", "Tangier Med", "Tanger", "Tangier"],
    "BASEL":          ["Basle", "Basel"],
    "BASLE":          ["Basle", "Basel"],
}

# Port code ưu tiên khi dropdown có nhiều kết quả cùng tên (vd Rotterdam NL vs Rotterdam US)
PORT_PREFERRED_CODE = {
    "VNSGN":       "VNSGN",
    "Ho Chi Minh": "VNSGN",
    "VNHPH":       "VNHPH",
    "Hai Phong":   "VNHPH",
    "Haiphong":    "VNHPH",
    "Rotterdam":  "NLRTM",
    "Hamburg":    "DEHAM",
    "Antwerp":    "BEANR",
    "Barcelona":  "ESBCN",
    "Valencia":   "ESVLC",
    "Genoa":      "ITGOA",
    "Genova":     "ITGOA",
    "Naples":     "ITNAP",
    "Venice":     "ITVCE",
    "Le Havre":   "FRLEH",
    "Algeciras":  "ESALG",
    "Piraeus":    "GRPIR",
    "Felixstowe": "GBFXT",
    "Southampton":"GBSOU",
    "Gothenburg": "SEGOT",
    "Basle":      "CHBSL",
    "Basel":      "CHBSL",
}

NUM_TABS = 3

# Track last POL/POD per tab để skip nhập lại field không đổi
tab_last_pol = [None] * NUM_TABS  # last POL entered per tab
tab_last_pod = [None] * NUM_TABS  # last POD entered per tab
cookie_dismissed = False          # chỉ dismiss cookie 1 lần

# ===================================================================================
# --- HELPERS ---
# ===================================================================================
def rand_sleep(a=0.3, b=0.8):
    time.sleep(random.uniform(a, b))

def hpl_quote_form_ready():
    try:
        for el in driver.find_elements(By.CSS_SELECTOR, 'input[data-testid="start-input"]'):
            try:
                if el.is_displayed() and el.is_enabled():
                    rect = driver.execute_script("""
                        const r = arguments[0].getBoundingClientRect();
                        return {w:r.width, h:r.height};
                    """, el)
                    if rect and rect.get("w", 0) > 10 and rect.get("h", 0) > 10:
                        return True
            except:
                continue
    except:
        pass
    return False

def hpl_auth_token_status():
    """Return VALID/EXPIRED/MISSING for the quote SPA bearer token."""
    try:
        token = driver.execute_script("return localStorage.getItem('token') || '';") or ""
    except Exception:
        return "MISSING"
    if not token:
        return "MISSING"
    return "EXPIRED" if jwt_is_expired(token, skew_seconds=30) else "VALID"

def hpl_auth_token_expired():
    return hpl_auth_token_status() == "EXPIRED"

def hpl_quote_tab_candidate():
    try:
        url = (driver.current_url or "").lower()
        if "hapag-lloyd.com" not in url:
            return False
        if any(marker in url for marker in HPL_QUOTE_URL_MARKERS):
            return True
        return hpl_quote_form_ready()
    except:
        return False

def hpl_operational_page_present():
    """
    HPL may keep a Cloudflare token in the URL even after the real quote/result UI is usable.
    Treat visible quote/result controls as operational and never as captcha.
    """
    try:
        if hpl_quote_form_ready():
            return True
        if driver.find_elements(By.CSS_SELECTOR, ".q-dialog .q-dialog__content"):
            return True
        if driver.find_elements(By.CSS_SELECTOR, "button.carousel__item"):
            return True
        buttons = driver.find_elements(By.XPATH, (
            "//button[.//span[contains(normalize-space(.),'View Departure Details')] "
            "or .//span[contains(normalize-space(.),'Price Breakdown')] "
            "or .//span[normalize-space(.)='Edit'] "
            "or .//span[normalize-space(.)='Search']]"
        ))
        if buttons:
            return True
        body = (driver.find_element(By.TAG_NAME, "body").text or "").upper()
        operational_markers = (
            "QUICK QUOTES",
            "OFFER SELECTION",
            "DEPARTURES",
            "PRICE BREAKDOWN",
            "VIEW DEPARTURE DETAILS",
            "HELLO MS",
            "PIO LOGISTICS",
        )
        return any(marker in body for marker in operational_markers)
    except:
        return False

def hpl_login_page_present():
    try:
        url = (driver.current_url or "").lower()
        if "identity.hapag-lloyd.com" in url or "/solutions/auth/login" in url:
            return True
        return bool(driver.find_elements(
            By.CSS_SELECTOR,
            "#signInName, input[name='E-mail Address'], #password"
        ))
    except:
        return False

def switch_to_hpl_login_or_quote_tab():
    quote_handle = None
    login_handle = None
    hpl_handle = None
    for handle in list(driver.window_handles):
        try:
            driver.switch_to.window(handle)
            url = (driver.current_url or "").lower()
            if "hapag-lloyd.com" not in url:
                continue
            hpl_handle = hpl_handle or handle
            if "identity.hapag-lloyd.com" in url or "/solutions/auth/login" in url:
                login_handle = handle
            elif any(marker in url for marker in HPL_QUOTE_URL_MARKERS) or hpl_quote_form_ready():
                quote_handle = handle
        except:
            continue
    target = quote_handle or login_handle or hpl_handle
    if target:
        driver.switch_to.window(target)
    return target

def close_hpl_empty_tabs(keep_handle=None):
    for handle in list(driver.window_handles):
        if handle == keep_handle:
            continue
        try:
            driver.switch_to.window(handle)
            url = (driver.current_url or "").lower()
            if (
                url.startswith("edge://newtab")
                or "ntp.msn.com/edge/ntp" in url
                or url in ("about:blank", "")
            ):
                driver.close()
        except:
            continue
    if keep_handle and keep_handle in driver.window_handles:
        driver.switch_to.window(keep_handle)

def hpl_manual_check_present():
    try:
        url = (driver.current_url or "").lower()
        title = (driver.title or "").lower()
        if any(k in title for k in ("security", "challenge", "captcha", "verify", "chờ một chút", "just a moment")):
            return True
        body = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
        keywords = [
            "verify you are human",
            "verify that you are human",
            "xác minh bạn là con người",
            "checking if the site connection is secure",
            "checking your browser",
            "security check",
            "your browser was tested against security",
            "managed challenge",
            "rayid",
            "ray id",
            "technical details",
            "please check following result",
            "captcha",
            "cloudflare",
            "human verification",
            "xác thực",
            "con người",
            "chờ một chút",
            "just a moment",
        ]
        if any(k in body for k in keywords):
            return True
        try:
            if driver.find_elements(By.CSS_SELECTOR, "#challenge-stage, #challenge-running, .cf-turnstile, input[name='cf-turnstile-response']"):
                return True
        except Exception:
            pass
        frames = driver.find_elements(
            By.XPATH,
            "//iframe[contains(@src,'captcha') or contains(@src,'cloudflare') "
            "or contains(@src,'turnstile') or contains(@src,'hcaptcha') or contains(@src,'recaptcha')]"
        )
        if frames:
            return True

        # Cloudflare token can remain in URL after HPL already shows login/quote.
        # Only use URL-based challenge detection after visible login/quote controls
        # fail, otherwise a solved page can be misread as captcha forever.
        if hpl_operational_page_present() or hpl_login_page_present():
            return False
        if any(k in url for k in (
            "security",
            "challenge",
            "captcha",
            "verify",
            "__cf_chl",
            "cf_chl",
            "turnstile",
        )):
            return True
        return False
    except:
        return False

def install_hpl_stealth_script():
    try:
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": HPL_STEALTH_SCRIPT})
    except Exception:
        pass

def wait_hpl_quote_or_block(timeout=None):
    timeout = HPL_TAB_LOAD_WAIT_SECONDS if timeout is None else timeout
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            if hpl_manual_check_present():
                return "BLOCKED"
            if hpl_auth_token_expired():
                return "AUTH_EXPIRED"
            if hpl_quote_form_ready() or hpl_operational_page_present():
                return "READY"
            if hpl_login_page_present():
                return "LOGIN"
        except Exception:
            pass
        time.sleep(0.5)
    if hpl_manual_check_present():
        return "BLOCKED"
    if hpl_auth_token_expired():
        return "AUTH_EXPIRED"
    if hpl_quote_form_ready() or hpl_operational_page_present():
        return "READY"
    if hpl_login_page_present():
        return "LOGIN"
    return "TIMEOUT"

def recover_hpl_expired_auth(reason="token HPL het han"):
    """Refresh exactly once when the visible quote UI has an expired token."""
    if not hpl_auth_token_expired():
        return True

    print(f"   [HPL AUTH] Token het han ({reason}); lam moi session dung 1 lan.")
    try:
        # Remove only the already-expired convenience token.  MSAL refresh
        # state and the signed-in browser profile remain intact.
        driver.execute_script("localStorage.removeItem('token');")
        driver.refresh()
    except Exception as exc:
        print(f"   [HPL AUTH] Khong refresh duoc: {type(exc).__name__}")
        return False

    deadline = time.time() + max(60, HPL_TAB_LOAD_WAIT_SECONDS)
    while time.time() < deadline:
        if hpl_manual_check_present():
            return wait_for_manual_hpl_unlock("captcha sau khi lam moi token HPL")
        if hpl_login_page_present():
            return complete_hpl_login_if_needed("login lai sau khi token HPL het han")
        if hpl_auth_token_status() == "VALID" and hpl_quote_form_ready():
            print("   [HPL AUTH] Session da duoc gia han, form quote san sang.")
            return True
        time.sleep(0.5)

    print("   [HPL AUTH] Het thoi gian cho token moi/form quote.")
    return False

def close_hpl_handle_safely(handle):
    if not handle:
        return
    try:
        handles = list(driver.window_handles)
        if handle not in handles or len(handles) <= 1:
            return
        driver.switch_to.window(handle)
        driver.close()
    except Exception:
        pass
    finally:
        try:
            handles = list(driver.window_handles)
            if handles:
                driver.switch_to.window(handles[-1])
        except Exception:
            pass

def register_replacement_hpl_tab(tab_idx, handle):
    if not tab_idx:
        return
    idx = tab_idx - 1
    try:
        if "tabs" in globals() and 0 <= idx < len(tabs):
            tabs[idx] = handle
        if "tab_last_pol" in globals() and 0 <= idx < len(tab_last_pol):
            tab_last_pol[idx] = None
        if "tab_last_pod" in globals() and 0 <= idx < len(tab_last_pod):
            tab_last_pod[idx] = None
    except Exception:
        pass

def open_fresh_hpl_quote_tab(tab_idx=None, reason="", close_handle=None):
    """
    Open a new HPL quote tab. If it reaches login, complete login there.
    If it reaches Cloudflare/captcha, go silent and wait for user ENTER.
    """
    last_state = ""
    last_error = None
    label = f"Tab{tab_idx}" if tab_idx else "HPL"
    attempts = HPL_TAB_REPLACE_ATTEMPTS if HPL_ALLOW_RELOAD else 1
    for attempt in range(1, attempts + 1):
        new_handle = None
        try:
            # A tab may have been closed by Edge/Cloudflare between discovery
            # and this call.  Always anchor Selenium on a verified live tab.
            switch_to_live_window(driver, preferred_handle=close_handle)
            before_handles = set(driver.window_handles)
            try:
                driver.switch_to.new_window("tab")
            except Exception:
                # Edge occasionally rejects Selenium's new_window even though
                # the attached session is alive.  JS opening is a safe fallback.
                switch_to_live_window(driver, preferred_handle=close_handle)
                driver.execute_script("window.open('about:blank', '_blank');")
                WebDriverWait(driver, 5).until(
                    lambda d: bool(set(d.window_handles) - before_handles)
                )
                opened = list(set(driver.window_handles) - before_handles)
                driver.switch_to.window(opened[-1])
            new_handle = driver.current_window_handle
            install_hpl_stealth_script()
            print(f"   [{label}] 🔁 Mở tab HPL mới ({attempt}/{attempts}) {reason}".rstrip())
            driver.get(BASE_URL)
            state = wait_hpl_quote_or_block()
            last_state = state
            if state in ("LOGIN", "BLOCKED"):
                if complete_hpl_login_if_needed(reason=f"{label} {state} {reason}".strip()):
                    state = "READY"
                    last_state = state
            if state == "READY":
                if close_handle and close_handle != new_handle:
                    close_hpl_handle_safely(close_handle)
                driver.switch_to.window(new_handle)
                register_replacement_hpl_tab(tab_idx, new_handle)
                print(f"   [{label}] ✅ Tab HPL mới sẵn sàng: {new_handle[:8]}...")
                return new_handle

            if not HPL_ALLOW_RELOAD and new_handle:
                print(f"   [{label}] Tab HPL chưa sẵn sàng ({state}); không đóng/mở lại để tránh captcha.")
                if wait_for_manual_hpl_unlock(f"{label} tab mới state={state}"):
                    state = wait_hpl_quote_or_block(timeout=30)
                    if state == "READY":
                        if close_handle and close_handle != new_handle:
                            close_hpl_handle_safely(close_handle)
                        driver.switch_to.window(new_handle)
                        register_replacement_hpl_tab(tab_idx, new_handle)
                        print(f"   [{label}] Tab HPL sẵn sàng sau manual: {new_handle[:8]}...")
                        return new_handle
                raise RuntimeError(f"{label} tab HPL mới chưa sẵn sàng ({state})")

            print(f"   [{label}] ⚠️ Tab mới chưa dùng được: {state} -> đóng và thử tab khác")
        except Exception as exc:
            last_error = exc
            print(f"   [{label}] ⚠️ Mở tab HPL mới lỗi: {type(exc).__name__}")
            if not HPL_ALLOW_RELOAD and new_handle:
                try:
                    driver.switch_to.window(new_handle)
                except Exception:
                    pass
                raise

        if new_handle:
            close_hpl_handle_safely(new_handle)
        try:
            if close_handle and close_handle in driver.window_handles:
                driver.switch_to.window(close_handle)
        except Exception:
            pass
        time.sleep(1.0 + attempt)

    detail = f"state={last_state}" if last_state else f"error={type(last_error).__name__ if last_error else 'unknown'}"
    raise RuntimeError(f"{label} không tạo được tab HPL mới sau {attempts} lần ({detail})")

def replace_hpl_tab_handle(handle=None, tab_idx=None, reason=""):
    try:
        if handle and handle in driver.window_handles:
            driver.switch_to.window(handle)
    except Exception:
        pass
    return open_fresh_hpl_quote_tab(tab_idx=tab_idx, reason=reason, close_handle=handle)

def replace_current_hpl_tab(tab_idx=None, reason=""):
    try:
        old_handle = driver.current_window_handle
    except Exception:
        old_handle = None
    return replace_hpl_tab_handle(old_handle, tab_idx=tab_idx, reason=reason)

def wait_for_manual_hpl_unlock(reason="captcha/login"):
    """
    Stop touching HPL completely while the user solves Cloudflare/CAPTCHA/login.
    Do not poll DOM/title/url here; Cloudflare may loop if Selenium keeps touching it.
    """
    global HPL_SAFE_SINGLE_TAB
    if HPL_SAFE_SINGLE_AFTER_MANUAL:
        HPL_SAFE_SINGLE_TAB = True
    print("")
    print("   [HPL SILENT] HPL dang can xu ly thu cong:", reason)
    print("   [HPL SILENT] Bot se IM LANG, khong doc DOM/URL/title, khong reload, khong click.")
    print("   [HPL SILENT] Hay giai CAPTCHA/login tren Edge HPL. Xong thi quay lai terminal va nhan ENTER.")
    if HPL_SAFE_SINGLE_TAB:
        print("   [HPL SILENT] Sau khi tiep tuc, bot se chay SAFE MODE 1 tab.")
    else:
        print("   [HPL SILENT] Sau khi tiep tuc, bot van chay pipeline 3 tab neu co/mở được tab.")
    try:
        # Khi bot chạy qua main.py, stdout bị pipe theo từng dòng. Prompt của
        # input(...) không có newline nên main có thể không hiển thị, làm tưởng
        # bot bị treo. In prompt bằng print trước, rồi input rỗng.
        print("   [HPL SILENT] Nhan ENTER sau khi HPL da pass captcha/login...")
        input()
    except EOFError:
        print(f"   [HPL SILENT] Khong co stdin, sleep {int(HPL_MANUAL_LOGIN_WAIT_SECONDS)}s.")
        time.sleep(HPL_MANUAL_LOGIN_WAIT_SECONDS)
    except KeyboardInterrupt:
        raise

    state = wait_hpl_quote_or_block(timeout=30)
    if state == "AUTH_EXPIRED":
        return recover_hpl_expired_auth("het han trong luc cho CAPTCHA/login")
    if state in ("READY", "LOGIN"):
        print(f"   [HPL SILENT] Tiep tuc, state={state}.")
        return True
    print(f"   [HPL SILENT] Sau ENTER van chua san sang, state={state}.")
    return False

def submit_hpl_login_if_present():
    if not hpl_login_page_present():
        return False
    print("   [HỆ THỐNG] HPL đang ở trang login, chờ form login load đủ...")
    if HPL_EMAIL and HPL_PASSWORD:
        try:
            email_input = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.ID, "signInName"))
            )
            password_input = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.ID, "password"))
            )
            email_input.send_keys(Keys.CONTROL + "a")
            email_input.send_keys(Keys.DELETE)
            email_input.send_keys(HPL_EMAIL)
            password_input.send_keys(Keys.CONTROL + "a")
            password_input.send_keys(Keys.DELETE)
            password_input.send_keys(HPL_PASSWORD)
            login_btn = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.ID, "next"))
            )
            login_btn.click()
            print("   [HỆ THỐNG] Đã gửi login HPL, chờ form quote...")
            return True
        except Exception as e:
            print(f"   [HỆ THỐNG] Không tự login HPL được: {type(e).__name__}")

    print("   [HỆ THỐNG] Cần login HPL thủ công.")
    return wait_for_manual_hpl_unlock("login HPL")

def complete_hpl_login_if_needed(reason="login HPL"):
    state = wait_hpl_quote_or_block(timeout=HPL_TAB_LOAD_WAIT_SECONDS)
    if state == "AUTH_EXPIRED":
        return recover_hpl_expired_auth(reason)
    if state == "READY":
        return True
    if state == "BLOCKED":
        if not wait_for_manual_hpl_unlock(reason):
            return False
        state = wait_hpl_quote_or_block(timeout=30)
        if state == "READY":
            return True
    if state == "LOGIN":
        if not submit_hpl_login_if_present():
            return False
        state = wait_hpl_quote_or_block(timeout=HPL_MANUAL_LOGIN_WAIT_SECONDS)
        if state == "BLOCKED":
            if not wait_for_manual_hpl_unlock("captcha sau login HPL"):
                return False
            state = wait_hpl_quote_or_block(timeout=30)
        return state == "READY"
    return False

def pause_if_hpl_manual_check(reason="captcha/login", tab_idx=None):
    if hpl_auth_token_expired():
        if not recover_hpl_expired_auth(reason):
            raise RuntimeError(f"HPL token het han va khong gia han duoc: {reason}")
    if hpl_manual_check_present():
        if not wait_for_manual_hpl_unlock(reason):
            raise RuntimeError(f"HPL manual check chua pass: {reason}")
    if not hpl_quote_form_ready():
        state = wait_hpl_quote_or_block(timeout=HPL_TAB_LOAD_WAIT_SECONDS)
        if state == "BLOCKED":
            if not wait_for_manual_hpl_unlock(reason):
                raise RuntimeError(f"HPL manual check chua pass: {reason}")
        elif state == "LOGIN":
            complete_hpl_login_if_needed(reason=reason)
        elif state != "READY":
            raise RuntimeError(f"HPL form chua san sang: {state}")

def activate_tab(handle):
    try:
        driver.execute_script("""
            window.dispatchEvent(new Event('focus'));
            document.dispatchEvent(new Event('focus'));
        """)
        rand_sleep(0.1, 0.2)
    except:
        pass

def human_move_and_click(element):
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    rand_sleep(0.1, 0.2)
    ActionChains(driver).move_to_element_with_offset(
        element, random.randint(-4, 4), random.randint(-3, 3)
    ).pause(random.uniform(0.05, 0.15)).click().perform()
    rand_sleep(0.1, 0.2)

def short_selenium_error(err):
    text = str(err).strip().splitlines()
    return text[0] if text else err.__class__.__name__

def hpl_dropdown_options_for_input(inp):
    try:
        return driver.execute_script("""
            const input = arguments[0];
            const listId = input && input.getAttribute('aria-controls');
            let menu = listId ? document.getElementById(listId) : null;
            const isVisible = el => {
                if (!el) return false;
                const r = el.getBoundingClientRect();
                const s = getComputedStyle(el);
                return r.width > 5 && r.height > 5 && s.display !== 'none' && s.visibility !== 'hidden';
            };
            if (!isVisible(menu)) {
                const visibleMenus = Array.from(document.querySelectorAll(
                    '.q-menu, [role="listbox"], .q-virtual-scroll'
                )).filter(isVisible);
                menu = visibleMenus.length ? visibleMenus[visibleMenus.length - 1] : null;
            }
            const nodes = menu
                ? Array.from(menu.querySelectorAll('.q-item[role="option"], [role="option"], .q-item'))
                : [];
            return nodes.filter(el => {
                return isVisible(el) && (el.innerText || '').trim();
            });
        """, inp) or []
    except Exception:
        return []

def hpl_dropdown_has_no_results(inp):
    try:
        return bool(driver.execute_script("""
            const input = arguments[0];
            const listId = input && input.getAttribute('aria-controls');
            let menu = listId ? document.getElementById(listId) : null;
            const isVisible = el => {
                if (!el) return false;
                const r = el.getBoundingClientRect();
                const s = getComputedStyle(el);
                return r.width > 5 && r.height > 5 && s.display !== 'none' && s.visibility !== 'hidden';
            };
            if (!isVisible(menu)) {
                const visibleMenus = Array.from(document.querySelectorAll(
                    '.q-menu, [role="listbox"], .q-virtual-scroll'
                )).filter(isVisible);
                menu = visibleMenus.length ? visibleMenus[visibleMenus.length - 1] : null;
            }
            if (!menu) return false;
            const hasOptions = menu.querySelector('.q-item[role="option"], [role="option"], .q-item');
            return !hasOptions && /no\\s+results|không\\s+có/i.test(menu.innerText || '');
        """, inp))
    except Exception:
        return False

def go_back():
    """Alt+← (browser back) — nhanh hơn hash change, giữ nguyên form state."""
    try:
        body = driver.find_element(By.TAG_NAME, "body")
        body.send_keys(Keys.ALT + Keys.ARROW_LEFT)
    except:
        driver.execute_script("window.history.back();")
    try:
        WebDriverWait(driver, HPL_BACK_WAIT_SECONDS).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-testid="start-input"]'))
        )
    except:
        driver.execute_script("window.location.hash = '/simple';")
        WebDriverWait(driver, HPL_BACK_FALLBACK_WAIT_SECONDS).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-testid="start-input"]'))
        )
    rand_sleep(0.3, 0.5)

def hpl_soft_return_to_form(timeout=12):
    """
    Return to the quote input form without reloading HPL.
    Prefer in-app Edit/back/hash because driver.get(BASE_URL) tends to trigger Cloudflare.
    """
    if hpl_quote_form_ready():
        return True
    if hpl_manual_check_present():
        return wait_for_manual_hpl_unlock("captcha/security khi quay ve form")

    # Result page normally has an Edit button. Use it instead of reload.
    for xpath in (
        "//button[.//span[normalize-space(.)='Edit']]",
        "//button[contains(normalize-space(.),'Edit')]",
        "//*[self::button or self::a][contains(normalize-space(.),'Edit')]",
    ):
        try:
            btns = driver.find_elements(By.XPATH, xpath)
            for btn in btns:
                if btn.is_displayed() and btn.is_enabled():
                    driver.execute_script("arguments[0].click();", btn)
                    state = wait_hpl_quote_or_block(timeout=timeout)
                    if state == "READY" and hpl_quote_form_ready():
                        return True
        except Exception:
            pass

    # Browser back without reload.
    try:
        body = driver.find_element(By.TAG_NAME, "body")
        body.send_keys(Keys.ALT + Keys.ARROW_LEFT)
        state = wait_hpl_quote_or_block(timeout=timeout)
        if state == "READY" and hpl_quote_form_ready():
            return True
    except Exception:
        pass

    # SPA hash only, still no network reload.
    try:
        driver.execute_script("window.location.hash = '/simple';")
        state = wait_hpl_quote_or_block(timeout=timeout)
        return state == "READY" and hpl_quote_form_ready()
    except Exception:
        return False

def safe_go_back(tab_idx=None):
    """Quay ve form HPL va khong de timeout o buoc back lam chet ca bot."""
    if hpl_manual_check_present():
        return wait_for_manual_hpl_unlock("captcha/security truoc khi back")
    last_error = None
    for attempt in range(3):
        try:
            if hpl_soft_return_to_form(timeout=HPL_BACK_WAIT_SECONDS):
                return True
            raise RuntimeError("soft return failed")
        except Exception as e:
            last_error = e
            print(f"   ⚠️ HPL back chua thay form (lan {attempt + 1}/3): {type(e).__name__}")
            try:
                if hpl_soft_return_to_form(timeout=HPL_BACK_FALLBACK_WAIT_SECONDS):
                    rand_sleep(0.3, 0.5)
                    return True
                if HPL_ALLOW_RELOAD:
                    print("   ⚠️ HPL_ALLOW_RELOAD=1 -> reload fallback.")
                    driver.get(BASE_URL)
                    WebDriverWait(driver, 25).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-testid="start-input"]'))
                    )
                    rand_sleep(0.3, 0.5)
                    return True
            except Exception as nav_error:
                last_error = nav_error
                time.sleep(2)

    return wait_for_manual_hpl_unlock(
        f"go_back khong thay form ({type(last_error).__name__ if last_error else 'unknown'})"
    )

def reload_tab(tab_idx=None):
    if hpl_manual_check_present():
        wait_for_manual_hpl_unlock("captcha/security truoc khi reload")
        return
    if not HPL_ALLOW_RELOAD:
        print("   [HPL] Bỏ reload để tránh captcha; thử quay về form bằng soft navigation.")
        hpl_soft_return_to_form(timeout=HPL_BACK_FALLBACK_WAIT_SECONDS)
        return
    for attempt in range(2):
        try:
            driver.get(BASE_URL)
            state = wait_hpl_quote_or_block(timeout=60)
            if state != "READY":
                raise RuntimeError(f"HPL reload state={state}")
            rand_sleep(1.5, 2.5)
            return
        except Exception as e:
            if attempt == 0:
                print(f"   ⚠️ reload_tab lỗi lần 1: {type(e).__name__}, thử lại...")
                time.sleep(3)
            else:
                wait_for_manual_hpl_unlock(f"reload_tab khong thay form ({type(e).__name__})")
                return

def fmt_date(raw):
    for fmt in ("%Y-%m-%d", "%d %b %Y", "%d-%b-%Y"):
        try:
            dt = datetime.strptime(raw.strip(), fmt)
            return f"{dt.day}-{dt.strftime('%b')}"
        except:
            continue
    return raw.strip()

# ===================================================================================
# --- CHỌN CẢNG ---
# ===================================================================================
def select_port_hpl(selector, port_name, aliases=None):
    names_to_try = [port_name]
    if aliases:
        names_to_try += [a for a in aliases if a != port_name]
    no_result_names = set()

    for name in names_to_try:
        for attempt in range(2):
            try:
                inp = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                )

                # Chỉ giữ nguyên nếu field đã chứa đúng port_name gốc
                current_val = driver.execute_script("return arguments[0].value;", inp)
                field_classes = driver.execute_script("""
                    const field = arguments[0].closest('.q-field');
                    return field ? field.className : '';
                """, inp)
                if hpl_selected_port_matches(
                    current_val,
                    port_name,
                    aliases or [],
                    field_classes,
                ):
                    print(f"        -> Giữ nguyên: {current_val.strip()}")
                    return

                # Luôn fetch inp mới (tránh stale sau bất kỳ DOM re-render nào)
                inp = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
                rand_sleep(0.1, 0.2)
                try:
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    rand_sleep(0.1, 0.2)
                except Exception:
                    pass
                driver.execute_script("arguments[0].click(); arguments[0].focus();", inp)
                rand_sleep(0.2, 0.3)

                # Fetch lại inp ngay trước send_keys (Quasar re-render sau click)
                inp = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
                inp.send_keys(Keys.CONTROL + "a")
                inp.send_keys(Keys.DELETE)
                rand_sleep(0.15, 0.25)
                inp.send_keys(name)
                rand_sleep(0.3, 0.5)

                # Chờ dropdown HOẶC No results
                WebDriverWait(driver, 8).until(lambda d:
                    hpl_dropdown_options_for_input(inp) or hpl_dropdown_has_no_results(inp)
                )

                if hpl_dropdown_has_no_results(inp):
                    print(f"        ⚠️ [{name}] No results → thử tên tiếp...")
                    no_result_names.add(name)
                    break

                suggestions = hpl_dropdown_options_for_input(inp)
                rand_sleep(0.2, 0.3)  # chờ animation dropdown xong
                # Dùng JS check thay vì is_displayed() vì hay fail khi tab nền
                visible = [s for s in suggestions if driver.execute_script(
                    "return arguments[0].offsetParent !== null && arguments[0].offsetHeight > 0;", s
                )]
                if not visible:
                    visible = suggestions  # fallback: thử tất cả nếu JS cũng trống
                if visible:
                    # Ưu tiên chọn item có preferred port code (tránh chọn nhầm cùng tên ở nước khác)
                    preferred_code = PORT_PREFERRED_CODE.get(name)
                    chosen = None
                    if preferred_code:
                        for s in visible:
                            if preferred_code in s.text:
                                chosen = s
                                break
                    if not chosen:
                        chosen = visible[0]
                    chosen_text = chosen.text.strip()
                    try:
                        human_move_and_click(chosen)
                    except (StaleElementReferenceException, WebDriverException):
                        visible = hpl_dropdown_options_for_input(inp)
                        chosen = None
                        if preferred_code:
                            for candidate in visible:
                                try:
                                    if preferred_code in candidate.text:
                                        chosen = candidate
                                        break
                                except Exception:
                                    continue
                        if not chosen:
                            chosen = visible[0] if visible else None
                        if not chosen:
                            raise
                        chosen_text = chosen.text.strip()
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].click();", chosen)
                    print(f"        -> Chốt [{name}]: {chosen_text[:60]}")
                    rand_sleep(0.2, 0.3)
                    return
                else:
                    raise Exception("Dropdown trống")

            except Exception as e:
                print(f"        ⚠️ [{name}] lần {attempt+1}: {short_selenium_error(e)}")
                if hpl_manual_check_present():
                    pause_if_hpl_manual_check(f"Captcha/Security chặn lúc nhập port [{name}]")
                rand_sleep(0.3, 0.5)

    if no_result_names and len(no_result_names) >= len(names_to_try):
        raise HPLPortNoResults(f"HPL port autocomplete No results: {names_to_try}")
    raise Exception(f"Thất bại tất cả aliases: {names_to_try}")

def hpl_field_matches_port(current_value, port_name, aliases=None):
    cur = str(current_value or "").upper().replace(" ", "")
    if not cur:
        return False
    candidates = [port_name] + list(aliases or [])
    for item in candidates:
        token = str(item or "").upper().replace(" ", "")
        if token and token in cur:
            return True
    return False

def hpl_input_has_selected_port(inp, port_name, aliases=None):
    try:
        current_value = driver.execute_script("return arguments[0].value || '';", inp)
        field_classes = driver.execute_script("""
            const field = arguments[0].closest('.q-field');
            return field ? field.className : '';
        """, inp)
        return hpl_selected_port_matches(
            current_value,
            port_name,
            aliases or [],
            field_classes,
        )
    except Exception:
        return False

# ===================================================================================
# --- ĐẶT NGÀY ---
# ===================================================================================
def set_date_hpl():
    target = (datetime.now() + timedelta(days=DATE_OFFSET_DAYS)).strftime("%Y-%m-%d")
    try:
        date_inp = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-testid="validity-input"]'))
        )
        if date_inp.get_attribute("value") == target:
            return
        driver.execute_script("""
            var inp = arguments[0]; var val = arguments[1];
            inp.focus();
            var setter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value').set;


            setter.call(inp, val);
            inp.dispatchEvent(new Event('input', { bubbles: true }));
            inp.dispatchEvent(new Event('change', { bubbles: true }));
        """, date_inp, target)
        rand_sleep(0.2, 0.3)
    except Exception as e:
        print(f"      ⚠️ Lỗi đặt ngày: {e}")
        if hpl_manual_check_present():
            pause_if_hpl_manual_check("Captcha/Security chặn lúc đặt ngày")

# ===================================================================================
# --- KIỂM TRA SECURITY CHECK ---
# ===================================================================================
def _is_security_page():
    """Kiểm tra xem trang hiện tại có phải là trang Security Check không."""
    try:
        url = (driver.current_url or "").lower()
        if "security" in url or "challenge" in url:
            return True
        title = (driver.title or "")
        if "Security" in title and "Check" in title:
            return True
        if hpl_manual_check_present():
            return True
    except:
        pass
    return False

def _do_tab_switch_stealth():
    """No-op: không đổi tab giả lập để tránh làm HPL challenge lặp lại."""
    return

def check_security_block(tab_idx):
    if not _is_security_page():
        return
    print(f"   [Tab{tab_idx}] ⚠️ Phát hiện xác thực HPL/Cloudflare.")
    if not wait_for_manual_hpl_unlock(f"Tab{tab_idx} captcha/security"):
        raise RuntimeError(f"Tab{tab_idx} captcha/security chua pass")

def hpl_no_service_visible():
    """True only when HPL is visibly showing a no-offer/sold-out state."""
    try:
        danger_nodes = driver.find_elements(By.CSS_SELECTOR, ".q-notification--danger")
        for node in danger_nodes:
            try:
                if not node.is_displayed():
                    continue
                text = (node.text or "").strip().upper()
                if not text:
                    continue
                if any(k in text for k in (
                    "NO SERVICE",
                    "SOLD OUT",
                    "NO OFFER",
                    "NO QUICK QUOTE",
                    "NO SCHEDULE",
                    "NO RESULT",
                    "NOT AVAILABLE",
                )):
                    return True
            except Exception:
                continue
        body = (driver.find_element(By.TAG_NAME, "body").text or "").upper()
        return any(k in body for k in (
            "NO SERVICE / SOLD OUT",
            "NO SERVICE",
            "SOLD OUT",
            "NO OFFER AVAILABLE",
        ))
    except Exception:
        return False

def hpl_service_unavailable_visible():
    """Detect HPL transient backend popup: 'This service is currently unavailable'."""
    try:
        selectors = [
            ".q-notification--danger",
            ".q-notification",
            ".q-dialog",
            "[role='alert']",
            "[role='dialog']",
        ]
        nodes = []
        for sel in selectors:
            try:
                nodes.extend(driver.find_elements(By.CSS_SELECTOR, sel))
            except Exception:
                pass
        for node in nodes:
            try:
                if not node.is_displayed():
                    continue
                text = (node.text or "").strip().upper()
                if not text:
                    continue
                if (
                    "THIS SERVICE IS CURRENTLY UNAVAILABLE" in text
                    or ("PLEASE, TRY AGAIN LATER" in text and "GLOBAL TRANSACTION ID" in text)
                    or ("SERVICE IS CURRENTLY UNAVAILABLE" in text and "DISMISS" in text)
                ):
                    return True
            except Exception:
                continue
        body = (driver.find_element(By.TAG_NAME, "body").text or "").upper()
        return (
            "THIS SERVICE IS CURRENTLY UNAVAILABLE" in body
            and "PLEASE, TRY AGAIN LATER" in body
            and "GLOBAL TRANSACTION ID" in body
        )
    except Exception:
        return False

def hpl_routing_unavailable_visible():
    """Detect the red banner saying the requested routing is unavailable."""
    try:
        selectors = [
            ".q-notification--danger",
            ".q-notification",
            ".q-banner",
            "[role='alert']",
            "[role='dialog']",
            "[class*='notification']",
        ]
        nodes = []
        for sel in selectors:
            try:
                nodes.extend(driver.find_elements(By.CSS_SELECTOR, sel))
            except Exception:
                pass
        for node in nodes:
            try:
                if not node.is_displayed():
                    continue
                text = " ".join((node.text or "").upper().split())
                if (
                    "THIS ROUTING IS CURRENTLY UNAVAILABLE" in text
                    or ("ROUTING IS CURRENTLY UNAVAILABLE" in text and "ALTERNATIVE ROUTING" in text)
                ):
                    return True
            except Exception:
                continue
        body = " ".join((driver.find_element(By.TAG_NAME, "body").text or "").upper().split())
        return (
            "THIS ROUTING IS CURRENTLY UNAVAILABLE" in body
            or ("ROUTING IS CURRENTLY UNAVAILABLE" in body and "ALTERNATIVE ROUTING" in body)
        )
    except Exception:
        return False

def hpl_unavailable_banner_visible():
    return hpl_routing_unavailable_visible() or hpl_service_unavailable_visible()

def hpl_dismiss_service_unavailable():
    clicked = False
    xpaths = [
        "//button[.//span[normalize-space()='Dismiss'] or normalize-space()='Dismiss']",
        "//*[self::button or self::a][contains(normalize-space(), 'Dismiss')]",
    ]
    for xp in xpaths:
        try:
            for el in driver.find_elements(By.XPATH, xp):
                try:
                    if el.is_displayed() and el.is_enabled():
                        driver.execute_script("arguments[0].click();", el)
                        clicked = True
                        break
                except Exception:
                    continue
            if clicked:
                break
        except Exception:
            continue
    if not clicked:
        try:
            driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        except Exception:
            pass
    rand_sleep(0.2, 0.4)
    return clicked

def hpl_raise_if_service_unavailable(tab_idx, context=""):
    if hpl_routing_unavailable_visible():
        label = f" ({context})" if context else ""
        print(f"   [Tab{tab_idx}] ⏭️ HPL routing unavailable{label} -> Dismiss và skip row ngay")
        hpl_dismiss_service_unavailable()
        raise HPLRoutingUnavailable("HPL routing currently unavailable")
    if hpl_service_unavailable_visible():
        label = f" ({context})" if context else ""
        print(f"   [Tab{tab_idx}] ⚠️ HPL service currently unavailable{label} -> Dismiss")
        hpl_dismiss_service_unavailable()
        raise HPLServiceUnavailable("HPL service currently unavailable")

def hpl_dismiss_no_service_notification():
    try:
        dismiss = driver.find_element(By.XPATH, "//button[.//span[text()='Dismiss']]")
        driver.execute_script("arguments[0].click();", dismiss)
    except Exception:
        pass

def hpl_wait_search_settled(tab_idx, timeout=45, min_wait=5, empty_confirm=4):
    """
    Wait after Search. HPL sometimes shows stale/early danger notifications while
    results are still rendering. Only return SOLD_OUT when no-service is stable.
    """
    start = time.time()
    empty_since = None
    while time.time() - start < timeout:
        check_security_block(tab_idx)
        if hpl_unavailable_banner_visible():
            return "SERVICE_UNAVAILABLE"
        if driver.find_elements(By.CSS_SELECTOR, "button.carousel__item"):
            return "CARDS"
        if hpl_no_service_visible():
            elapsed = time.time() - start
            if elapsed < min_wait:
                empty_since = None
            else:
                empty_since = empty_since or time.time()
                if time.time() - empty_since >= empty_confirm:
                    return "SOLD_OUT"
        else:
            empty_since = None
        time.sleep(0.5)
    if driver.find_elements(By.CSS_SELECTOR, "button.carousel__item"):
        return "CARDS"
    if hpl_unavailable_banner_visible():
        return "SERVICE_UNAVAILABLE"
    if hpl_no_service_visible():
        return "SOLD_OUT"
    return "TIMEOUT"
     
# ===================================================================================
# --- NHẬP & SEARCH ---
# ===================================================================================
def do_search(pol, pod, tab_idx):
    global cookie_dismissed
    t = tab_idx - 1  # 0-based index

    # YÊU CẦU: Luôn check URL trước khi nhập liệu
    target_url = "https://www.hapag-lloyd.com/solutions/new-quote/#/simple"
    if target_url not in driver.current_url:
        print(f"   [Tab{tab_idx}] 🔄 URL sai, điều hướng về: {target_url}")
        try:
            driver.get(target_url)
            time.sleep(3)
        except Exception as e:
            print(f"   [Tab{tab_idx}] ⚠️ Lỗi khi load URL: {e}")

    print(f"   [Tab{tab_idx}] 📍 Nhập {pol} → {pod}")
    check_security_block(tab_idx)
    pause_if_hpl_manual_check(f"Tab{tab_idx} truoc khi nhap form", tab_idx=tab_idx)
    hpl_raise_if_service_unavailable(tab_idx, "trước khi nhập form")

    # Dismiss cookie/notification banner — chỉ 1 lần
    if not cookie_dismissed:
        try:
            driver.execute_script("""
                ['#onetrust-accept-btn-handler', '.onetrust-close-btn-handler',
                 '.hl-notification__close', '.hl-cookie-banner__close'
                ].forEach(function(s){
                    var el = document.querySelector(s); if (el) el.click();
                });
            """)
            rand_sleep(0.2, 0.3)
            cookie_dismissed = True
        except: pass

    # --- SMART POL: chỉ nhập lại nếu khác row trước ---
    if tab_last_pol[t] == pol:
        # Check xem field có còn giữ giá trị cũ không
        try:
            inp = driver.find_element(By.CSS_SELECTOR, 'input[data-testid="start-input"]')
            cur = driver.execute_script("return arguments[0].value;", inp) or ""
            if hpl_input_has_selected_port(inp, pol, POL_ALIASES.get(pol.upper())):
                print(f"        -> POL giữ nguyên: {cur.strip()}")
            else:
                select_port_hpl('input[data-testid="start-input"]', pol, aliases=POL_ALIASES.get(pol.upper()))
                rand_sleep(0.3, 0.5)
        except:
            select_port_hpl('input[data-testid="start-input"]', pol, aliases=POL_ALIASES.get(pol.upper()))
            rand_sleep(0.3, 0.5)
    else:
        select_port_hpl('input[data-testid="start-input"]', pol, aliases=POL_ALIASES.get(pol.upper()))
        rand_sleep(0.3, 0.5)
    tab_last_pol[t] = pol

    # --- SMART POD: chỉ nhập lại nếu khác row trước ---
    pod_upper   = pod.upper()
    pod_aliases = POD_ALIASES.get(pod_upper, [pod])
    if tab_last_pod[t] == pod:
        try:
            inp = driver.find_element(By.CSS_SELECTOR, 'input[data-testid="end-input"]')
            cur = driver.execute_script("return arguments[0].value;", inp) or ""
            first_alias = pod_aliases[0]
            if hpl_input_has_selected_port(inp, first_alias, pod_aliases):
                print(f"        -> POD giữ nguyên: {cur.strip()}")
            else:
                select_port_hpl('input[data-testid="end-input"]', pod_aliases[0], aliases=pod_aliases)
        except:
            select_port_hpl('input[data-testid="end-input"]', pod_aliases[0], aliases=pod_aliases)
    else:
        select_port_hpl('input[data-testid="end-input"]', pod_aliases[0], aliases=pod_aliases)
    tab_last_pod[t] = pod

    set_date_hpl()

    try:
        search_btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
            (By.XPATH, "//button[.//span[text()='Search']]")))
        human_move_and_click(search_btn)
        print(f"   [Tab{tab_idx}] 🖱️ Đã Search")
    except Exception as e:
        if hpl_manual_check_present():
            pause_if_hpl_manual_check(f"Captcha/Security chặn lúc click Search", tab_idx=tab_idx)
        raise e


# ===================================================================================
# --- ĐỌC VALID DATE TỪ SIDEBAR ---
# ===================================================================================
def read_valid_from_sidebar(tab_idx):
    try:
        el = WebDriverWait(driver, 4).until(
            EC.presence_of_element_located((By.XPATH,
                '//*[@id="hlMain"]/div[2]/div[2]/div/div/div/div[3]/div[2]/div[2]'
                '/div[1]/div[2]/div/div[1]/span/span[2]'
            ))
        )
        raw = el.text.strip()
        if raw:
            return fmt_date(raw)
    except:
        pass
    try:
        spans = driver.find_elements(By.CSS_SELECTOR, ".sidebar-schedule span.text-button-s")
        for sp in spans:
            raw = sp.text.strip()
            if re.match(r'\d{4}-\d{2}-\d{2}', raw) or re.match(r'\d{2} \w+ \d{4}', raw):
                return fmt_date(raw)
    except:
        pass
    return ""

# ===================================================================================
# --- ĐỌC CARDS ---
# FIX: track (etd_str, tt_days) combo để bỏ card trùng qua Next
# ===================================================================================
def read_all_cards(tab_idx):
    print(f"   [Tab{tab_idx}] 📋 Đọc cards...")
    check_security_block(tab_idx)
    hpl_raise_if_service_unavailable(tab_idx, "trước khi đọc cards")
    result     = []
    seen_combo = set()  # track (etd_str, tt_days) để loại trùng

    try:
        WebDriverWait(driver, 30).until(lambda d:
            hpl_unavailable_banner_visible()
            or d.find_elements(By.CSS_SELECTOR, "button.carousel__item")
        )
    except TimeoutException:
        state = hpl_wait_search_settled(tab_idx, timeout=20, min_wait=5, empty_confirm=4)
        if state == "SERVICE_UNAVAILABLE":
            hpl_raise_if_service_unavailable(tab_idx, "chờ cards")
        if state == "SOLD_OUT":
            hpl_dismiss_no_service_notification()
            raise Exception("NO SERVICE / SOLD OUT")
        return []
    hpl_raise_if_service_unavailable(tab_idx, "chờ cards")

    if hpl_no_service_visible() and not driver.find_elements(By.CSS_SELECTOR, "button.carousel__item"):
        state = hpl_wait_search_settled(tab_idx, timeout=12, min_wait=5, empty_confirm=4)
        if state == "SERVICE_UNAVAILABLE":
            hpl_raise_if_service_unavailable(tab_idx, "đọc cards")
        if state == "SOLD_OUT":
            hpl_dismiss_no_service_notification()
            raise Exception("NO SERVICE / SOLD OUT")
        if state != "CARDS":
            return []

    rand_sleep(0.5, 1.0)
    max_pages   = 6
    page_count  = 0

    while page_count < max_pages:
        cards = driver.find_elements(By.CSS_SELECTOR, "button.carousel__item")

        for idx in range(len(cards)):
            try:
                cards = driver.find_elements(By.CSS_SELECTOR, "button.carousel__item")
                card  = cards[idx]

                etd_span = card.find_element(By.CSS_SELECTOR, "div.carousel__date span")
                etd_str  = etd_span.text.strip()
                if not etd_str:
                    continue

                aria = card.get_attribute("aria-label") or ""
                qq_match = re.search(r"Quick Quotes: USD([\d ,]+)", aria)
                if not qq_match:
                    print(f"   [Tab{tab_idx}] ⏭️ Skip {etd_str} (no price)")
                    continue

                base_price = int(qq_match.group(1).strip().replace(" ", "").replace(",", ""))
                etd_dt     = datetime.strptime(etd_str, "%Y-%m-%d")

                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", card)
                rand_sleep(0.05, 0.1)
                driver.execute_script("arguments[0].click();", card)
                rand_sleep(0.05, 0.1)

                try:
                    transit_el = WebDriverWait(driver, 4).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, ".sidebar-schedule__days"))
                    )
                    tt_days = int(re.search(r'\d+', transit_el.text.strip()).group())
                except:
                    tt_days = 999

                # Bỏ combo trùng (etd + tt) — tránh đọc lại card cũ sau Next
                combo = (etd_str, tt_days)
                if combo in seen_combo:
                    continue
                seen_combo.add(combo)

                valid_to = read_valid_from_sidebar(tab_idx)

                print(f"   [Tab{tab_idx}] ✅ ETD={etd_str} TT={tt_days}d Price={base_price} Valid={valid_to}")
                result.append({
                    "etd_dt":     etd_dt,
                    "tt_days":    tt_days,
                    "etd_str":    etd_str,
                    "base_price": base_price,
                    "valid_to":   valid_to,
                })

            except Exception as e:
                print(f"   [Tab{tab_idx}] ⚠️ Lỗi card idx={idx}: {e}")
                continue

        try:
            next_btn = driver.find_element(By.CSS_SELECTOR, "button[aria-label='Next Departures']")
            if next_btn.get_attribute("disabled") is None:
                human_move_and_click(next_btn)
                page_count += 1
                print(f"   [Tab{tab_idx}] ➡️ Next ({page_count}/{max_pages})...")
                rand_sleep(0.5, 0.8)
            else:
                break
        except:
            break

    return result

# ===================================================================================
# --- RESET CAROUSEL ---
# ===================================================================================
def reset_carousel(tab_idx):
    for _ in range(20):
        try:
            prev_btn = driver.find_element(By.CSS_SELECTOR, "button[aria-label='Previous Departures']")
            if prev_btn.get_attribute("disabled") is not None:
                break
            driver.execute_script("arguments[0].click();", prev_btn)
            rand_sleep(0.15, 0.25)
        except:
            break

# ===================================================================================
# --- ĐỌC LỊCH TÀU TỪ POPUP "View Departure Details" ---
# ===================================================================================
def _clean_hpl_via_text(raw):
    if not raw:
        return ""
    text = str(raw).replace("\u2022", "+").replace("•", "+")
    text = re.sub(r"\s*\+\s*\d+\s*", " + ", text)
    text = re.sub(r"\bVIA\b", " ", text, flags=re.I)
    parts = []
    for part in re.split(r"\s*\+\s*|\r?\n|,", text):
        port = re.sub(r"\s+", " ", part).strip(" -:/")
        if not port or port == "-":
            continue
        up = port.upper()
        if up not in parts:
            parts.append(up)
    return " + ".join(parts)

def _extract_hpl_full_via(via_el):
    candidates = []
    try:
        candidates.append(via_el.text)
    except:
        pass
    try:
        candidates.extend(driver.execute_script("""
            const root = arguments[0];
            const out = [];
            function add(t) {
                if (t && String(t).trim()) out.push(String(t).trim());
            }
            add(root.innerText);
            add(root.textContent);
            const attrs = ['title', 'aria-label', 'data-tooltip', 'data-original-title',
                           'data-bs-original-title', 'data-qtip'];
            root.querySelectorAll('*').forEach(n => {
                attrs.forEach(a => add(n.getAttribute(a)));
                add(n.innerText);
            });
            return out;
        """, via_el) or [])
    except:
        pass

    short_text = " ".join(str(x) for x in candidates if x)
    if re.search(r"\+\s*\d+", short_text):
        try:
            ActionChains(driver).move_to_element(via_el).pause(0.35).perform()
            tooltip_texts = driver.execute_script("""
                const sels = ['.q-tooltip', '[role="tooltip"]', '.tooltip', '.v-popper__popper'];
                const out = [];
                sels.forEach(sel => document.querySelectorAll(sel).forEach(el => {
                    const st = window.getComputedStyle(el);
                    if (st && st.display !== 'none' && st.visibility !== 'hidden') {
                        const txt = (el.innerText || el.textContent || '').trim();
                        if (txt) out.push(txt);
                    }
                }));
                return out;
            """) or []
            candidates.extend(tooltip_texts)
        except:
            pass

    cleaned = []
    for text in candidates:
        via = _clean_hpl_via_text(text)
        if via and via != "DIRECT":
            cleaned.append(via)
    if not cleaned:
        return ""

    cleaned.sort(key=lambda s: (s.count("+"), len(s)), reverse=True)
    return cleaned[0]

def _parse_hpl_card_date_text(text):
    raw = re.sub(r"\s+", " ", str(text or "")).strip()
    if not raw:
        return None

    candidates = [raw]
    year = datetime.now().year
    if not re.search(r"\b20\d{2}\b", raw):
        candidates.append(f"{raw} {year}")

    for value in candidates:
        for fmt in (
            "%Y-%m-%d",
            "%d %b %Y",
            "%d-%b-%Y",
            "%d %B %Y",
            "%d-%B-%Y",
            "%b %d %Y",
            "%B %d %Y",
            "%d/%m/%Y",
            "%m/%d/%Y",
        ):
            try:
                return datetime.strptime(value, fmt).replace(hour=0, minute=0, second=0, microsecond=0)
            except Exception:
                pass
    return None

def _hpl_card_date_matches(card_date_text, etd_info):
    target_dt = etd_info.get("etd_dt")
    if target_dt:
        card_dt = _parse_hpl_card_date_text(card_date_text)
        if card_dt and card_dt.date() == target_dt.date():
            return True
    return str(card_date_text or "").strip() == str(etd_info.get("etd_str", "")).strip()

def find_and_click_exact_hpl_card(tab_idx, etd_info):
    etd_target = etd_info.get("etd_str", "")
    target_tt = etd_info.get("tt_days")
    target_vessel = (etd_info.get("vessel") or "").strip().upper()
    target_service = (etd_info.get("service") or "").strip().upper()

    reset_carousel(tab_idx)
    rand_sleep(0.15, 0.25)
    for _ in range(6):
        candidates = []
        for card in driver.find_elements(By.CSS_SELECTOR, "button.carousel__item"):
            try:
                card_date_text = card.find_element(By.CSS_SELECTOR, "div.carousel__date span").text.strip()
                if not _hpl_card_date_matches(card_date_text, etd_info):
                    continue
                card_text = ((card.text or "") + " " + (card.get_attribute("aria-label") or "")).upper()
                score = 10
                if target_vessel and target_vessel[:12] in card_text:
                    score += 4
                if target_service and target_service[:8] in card_text:
                    score += 2
                if target_tt and re.search(rf"\b{int(target_tt)}\s*DAYS?\b", card_text):
                    score += 4
                candidates.append((score, card))
            except:
                continue

        if candidates:
            candidates.sort(key=lambda item: item[0], reverse=True)
            if target_tt:
                for _, card in candidates:
                    try:
                        human_move_and_click(card)
                        rand_sleep(0.18, 0.3)
                        transit_el = driver.find_element(By.CSS_SELECTOR, ".sidebar-schedule__days")
                        m = re.search(r"\d+", transit_el.text or "")
                        current_tt = int(m.group()) if m else None
                        if current_tt == int(target_tt):
                            return True
                    except:
                        continue
            human_move_and_click(candidates[0][1])
            rand_sleep(0.2, 0.3)
            return True

        try:
            next_btn = driver.find_element(By.CSS_SELECTOR, "button[aria-label='Next Departures']")
            if next_btn.get_attribute("disabled") is None:
                driver.execute_script("arguments[0].click();", next_btn)
                rand_sleep(0.3, 0.5)
            else:
                break
        except:
            break
    return False

def find_and_click_any_priced_hpl_card(tab_idx):
    """Last-resort fallback: click the first visible priced departure card."""
    reset_carousel(tab_idx)
    rand_sleep(0.15, 0.25)
    for _ in range(6):
        for card in driver.find_elements(By.CSS_SELECTOR, "button.carousel__item"):
            try:
                card_text = ((card.text or "") + " " + (card.get_attribute("aria-label") or "")).upper()
                if "QUICK QUOTES" not in card_text and "USD" not in card_text:
                    continue
                human_move_and_click(card)
                rand_sleep(0.2, 0.3)
                return True
            except:
                continue
        try:
            next_btn = driver.find_element(By.CSS_SELECTOR, "button[aria-label='Next Departures']")
            if next_btn.get_attribute("disabled") is None:
                driver.execute_script("arguments[0].click();", next_btn)
                rand_sleep(0.3, 0.5)
            else:
                break
        except:
            break
    return False

def read_departure_details(tab_idx):
    """
    Click 'View Departure Details' → đọc tất cả card trong popup.
    Trả về list dict: etd_dt, etd_str, base_price, tt_days,
           vessel, service, transshipment, arrival, valid_to
    Trả về None nếu popup không mở được (caller sẽ fallback sang read_all_cards).
    """
    print(f"   [Tab{tab_idx}] 📋 Mở popup Departure Details...")
    check_security_block(tab_idx)
    hpl_raise_if_service_unavailable(tab_idx, "trước Departure Details")

    # Chờ carousel load xong trước
    try:
        WebDriverWait(driver, 30).until(lambda d:
            hpl_unavailable_banner_visible()
            or d.find_elements(By.CSS_SELECTOR, "button.carousel__item")
        )
    except TimeoutException:
        state = hpl_wait_search_settled(tab_idx, timeout=20, min_wait=5, empty_confirm=4)
        if state == "SERVICE_UNAVAILABLE":
            hpl_raise_if_service_unavailable(tab_idx, "chờ Departure Details")
        if state == "SOLD_OUT":
            hpl_dismiss_no_service_notification()
            raise Exception("NO SERVICE / SOLD OUT")
        return None
    hpl_raise_if_service_unavailable(tab_idx, "chờ carousel")

    if hpl_no_service_visible() and not driver.find_elements(By.CSS_SELECTOR, "button.carousel__item"):
        state = hpl_wait_search_settled(tab_idx, timeout=12, min_wait=5, empty_confirm=4)
        if state == "SERVICE_UNAVAILABLE":
            hpl_raise_if_service_unavailable(tab_idx, "Departure Details")
        if state == "SOLD_OUT":
            hpl_dismiss_no_service_notification()
            raise Exception("NO SERVICE / SOLD OUT")
        if state != "CARDS":
            return None

    rand_sleep(0.5, 1.0)

    # Click "View Departure Details"
    try:
        detail_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
            (By.XPATH, "//button[.//span[text()='View Departure Details']]")))
        human_move_and_click(detail_btn)
        rand_sleep(0.8, 1.2)
    except Exception as e:
        hpl_raise_if_service_unavailable(tab_idx, "click Departure Details")
        print(f"   [Tab{tab_idx}] ⚠️ Không click được 'View Departure Details': {e}")
        return None

    # Chờ dialog hiện
    try:
        WebDriverWait(driver, 15).until(lambda d:
            hpl_unavailable_banner_visible()
            or d.find_elements(By.CSS_SELECTOR, ".q-dialog .q-dialog__content")
        )
    except TimeoutException:
        hpl_raise_if_service_unavailable(tab_idx, "chờ dialog Departure Details")
        print(f"   [Tab{tab_idx}] ⚠️ Popup không hiện")
        return None
    hpl_raise_if_service_unavailable(tab_idx, "mở dialog Departure Details")
    rand_sleep(0.5, 0.8)

    try:
        dialog = driver.find_element(By.CSS_SELECTOR, ".q-dialog")
    except:
        print(f"   [Tab{tab_idx}] ⚠️ Không tìm thấy dialog")
        return None

    # Đảm bảo chip "Routing Details" được chọn
    try:
        chips = dialog.find_elements(By.CSS_SELECTOR, ".q-chip")
        for chip in chips:
            if "Routing" in chip.text and "q-chip--selected" not in (chip.get_attribute("class") or ""):
                chip.click()
                rand_sleep(0.3, 0.5)
                break
    except:
        pass

    result     = []
    seen_combo = set()
    max_pages  = 6
    page_count = 0

    while page_count < max_pages:
        try:
            cards = dialog.find_elements(By.CSS_SELECTOR, "button.carousel__item")
        except:
            break

        for card in cards:
            try:
                # ETD từ div[1]/span
                etd_el  = card.find_element(By.XPATH, "./div[1]/span")
                etd_str = etd_el.text.strip()
                if not etd_str:
                    continue

                # QQ Price từ aria-label
                aria     = card.get_attribute("aria-label") or ""
                qq_match = re.search(r"Quick Quotes:\s*USD\s*([\d.,\s]+)", aria)
                if not qq_match:
                    print(f"   [Tab{tab_idx}] ⏭️ Skip {etd_str} (no price)")
                    continue
                price_str  = qq_match.group(1).strip().replace(" ", "").replace(",", "")
                base_price = int(float(price_str))

                # Transit time - Robust extraction
                try:
                    tt_days = 999
                    card_text = card.text.strip()
                    tt_match = re.search(r'\b(\d+)\s*[Dd]ays?\b', card_text, re.IGNORECASE)
                    if tt_match:
                        tt_days = int(tt_match.group(1))
                    else:
                        for div_idx in (4, 5, 6, 7):
                            try:
                                div_text = card.find_element(By.XPATH, f"./div[{div_idx}]").text.strip()
                                if re.search(r'202\d', div_text) or re.search(r'^[a-zA-Z]+', div_text): 
                                    continue
                                m = re.search(r'^(\d+)$', div_text)
                                if m:
                                    val = int(m.group(1))
                                    if val < 2000:
                                        tt_days = val
                                        break
                            except: pass
                except: tt_days = 999

                print(f"   [Tab{tab_idx}] [DEBUG CARD TEXT]\n{card.text}")
                # Arrival date từ div[4]
                try:
                    arr_div     = card.find_element(By.XPATH, "./div[4]")
                    arrival_str = arr_div.text.strip()
                except:
                    arrival_str = ""

                # Vessel name từ div[6]
                try:
                    vessel_el   = card.find_element(By.XPATH, "./div[6]//div[contains(@class,'ellipsis')]")
                    vessel_name = vessel_el.text.strip()
                except:
                    vessel_name = ""

                # Service từ div[7]
                try:
                    service_el   = card.find_element(By.XPATH, "./div[7]//div[contains(@class,'ellipsis')]")
                    service_name = service_el.text.strip()
                except:
                    service_name = ""

                # Via/Transshipment từ div[11]
                try:
                    via_el   = card.find_element(By.XPATH, "./div[11]")
                    via_text = _extract_hpl_full_via(via_el)
                    if via_text == "-" or not via_text:
                        transshipment = "DIRECT"
                    else:
                        transshipment = via_text.upper()
                except:
                    transshipment = ""

                # Dedup theo (etd, tt, vessel) — tránh đọc trùng sau Next
                combo = (etd_str, tt_days, vessel_name)
                if combo in seen_combo:
                    continue
                seen_combo.add(combo)

                etd_dt = datetime.strptime(etd_str, "%Y-%m-%d")
                result.append({
                    "etd_dt":        etd_dt,
                    "etd_str":       etd_str,
                    "tt_days":       tt_days,
                    "base_price":    base_price,
                    "arrival":       arrival_str,
                    "vessel":        vessel_name,
                    "service":       service_name,
                    "transshipment": transshipment,
                    "valid_to":      "",
                })
                print(f"   [Tab{tab_idx}] 🚢 ETD={etd_str} Price={base_price} TT={tt_days}d "
                      f"Vessel={vessel_name} T/S={transshipment}")

            except Exception as e:
                print(f"   [Tab{tab_idx}] ⚠️ Lỗi đọc popup card: {e}")
                continue

        # Next page trong dialog
        try:
            next_btns = dialog.find_elements(By.CSS_SELECTOR, "button[aria-label='Next Departures']")
            if next_btns and next_btns[0].get_attribute("disabled") is None:
                human_move_and_click(next_btns[0])
                page_count += 1
                print(f"   [Tab{tab_idx}] ➡️ Popup Next ({page_count}/{max_pages})...")
                rand_sleep(0.5, 0.8)
            else:
                break
        except:
            break

    # Đóng dialog
    try:
        close_btn = dialog.find_element(By.CSS_SELECTOR, "button.q-dialog__x")
        driver.execute_script("arguments[0].click();", close_btn)
        rand_sleep(0.3, 0.5)
    except:
        try:
            dialog.send_keys(Keys.ESCAPE)
            rand_sleep(0.3, 0.5)
        except:
            pass

    print(f"   [Tab{tab_idx}] 📋 Popup: tìm thấy {len(result)} departures")
    return result if result else None

# ===================================================================================
# --- 9 QUY TẮC VÀNG ---
# ===================================================================================
def apply_9_golden_rules(danh_sach_chuyen):
    co_gia = [
        c for c in danh_sach_chuyen
        if c.get("base_price") is not None and etd_within_max(c.get("etd_dt"))
    ]
    if not co_gia:
        return [], "N/A", "N/A"

    # --- ĐIỀU KIỆN MỚI: Ưu tiên tháng hiện tại ---
    # Sắp xếp để tìm ngày ETD sớm nhất (tháng hiện tại của view)
    co_gia.sort(key=lambda x: x["etd_dt"])
    first_year, first_month = co_gia[0]["etd_dt"].year, co_gia[0]["etd_dt"].month
    
    # Lấy toàn bộ card thuộc tháng đầu tiên này
    thang_nay = [c for c in co_gia if c["etd_dt"].year == first_year and c["etd_dt"].month == first_month]
    
    # Đếm số lượng ETD khác nhau trong tháng này
    unique_etds_thang_nay = set([c["etd_dt"] for c in thang_nay])
    
    # Nếu tháng này có từ 3 ngày ETD trở lên -> Chốt chỉ chơi với tháng này, mặc kệ giá tháng sau
    if len(unique_etds_thang_nay) >= 3:
        co_gia = thang_nay
    # ---------------------------------------------

    min_price        = min(c["base_price"] for c in co_gia)
    same_price_group = [c for c in co_gia if c["base_price"] == min_price]
    same_price_group.sort(key=lambda x: (x["etd_dt"], x["tt_days"]))

    # Lọc trùng ETD — mỗi ETD chỉ giữ TT ngắn nhất
    seen, list_loc_trung = set(), []
    for c in same_price_group:
        if c["etd_dt"] not in seen:
            list_loc_trung.append(c)
            seen.add(c["etd_dt"])

    etd_dat_chuan = []
    if list_loc_trung:
        first_date = list_loc_trung[0]["etd_dt"]
        for c in list_loc_trung:
            if len(etd_dat_chuan) >= 3:
                break
            if len(etd_dat_chuan) > 0 and (c["etd_dt"] - etd_dat_chuan[-1]["etd_dt"]).days < 2:
                continue
            in_first_window = (c["etd_dt"] - first_date).days <= 9
            if in_first_window:
                etd_dat_chuan.append(c)

        # Nếu có một departure rất xa nhưng transit cực ngắn, đừng để nó làm rỗng nhóm ETD gần nhất.
        if not etd_dat_chuan:
            for c in list_loc_trung:
                if len(etd_dat_chuan) >= 3:
                    break
                if len(etd_dat_chuan) > 0 and (c["etd_dt"] - etd_dat_chuan[-1]["etd_dt"]).days < 2:
                    continue
                etd_dat_chuan.append(c)

    def _fmt_etd(dt):
        return f"{dt.day}-{dt.strftime('%b')}"
    num = len(etd_dat_chuan)
    if num == 0:   str_etd = "N/A"
    elif num == 1: str_etd = _fmt_etd(etd_dat_chuan[0]["etd_dt"])
    elif num == 2: str_etd = (f"{_fmt_etd(etd_dat_chuan[0]['etd_dt'])} & "
                              f"{_fmt_etd(etd_dat_chuan[1]['etd_dt'])}")
    else:
        d1 = str(etd_dat_chuan[0]["etd_dt"].day)
        d2 = str(etd_dat_chuan[1]["etd_dt"].day)
        d3 = _fmt_etd(etd_dat_chuan[2]["etd_dt"])
        str_etd = f"{d1}, {d2}, {d3}"

    all_tt = [c["tt_days"] for c in etd_dat_chuan if c.get("tt_days") is not None]
    if not all_tt:
        return etd_dat_chuan, str_etd, "N/A"
    str_tt = f"{min(all_tt)}" if min(all_tt) == max(all_tt) else f"{min(all_tt)}-{max(all_tt)}"
    return etd_dat_chuan, str_etd, str_tt

# ===================================================================================
# --- PARSE BẢNG GIÁ ---
# ===================================================================================
def _parse_hpl_amount(text):
    cleaned = re.sub(r"[\s,\u00a0\u202f]", "", str(text or ""))
    if not cleaned or cleaned == "-":
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None

def _hpl_ocean_freight_includes_othc(dialog_text):
    upper = re.sub(r"\r\n?", "\n", str(dialog_text or "")).upper()
    marker = "OCEAN FREIGHT INCLUDES"
    start = upper.find(marker)
    if start < 0:
        return False

    end_candidates = []
    for end_marker in (
        "OCEAN FREIGHT IS NOT SUBJECT",
        "\nSURCHARGES",
        "\nFREIGHT SURCHARGES",
        "\nEXPORT SURCHARGES",
        "\nIMPORT SURCHARGES",
        "\nDESTINATION SURCHARGES",
    ):
        pos = upper.find(end_marker, start + len(marker))
        if pos >= 0:
            end_candidates.append(pos)

    end = min(end_candidates) if end_candidates else min(len(upper), start + 600)
    included_block = upper[start:end]
    return (
        "TERMINAL HANDLING CHARGE ORIG" in included_block
        or "TERMINAL HANDLING CHARGE AT ORIGIN" in included_block
        or "ORIGIN TERMINAL HANDLING" in included_block
    )

def _read_hpl_container_total_table(all_tables, tab_idx):
    """
    Use the top Price Breakdown summary table:
    Unit | Curr. | 20STD | 40STD | 40HC.
    The Ctr. row is the authoritative all-in per-container total.
    """
    for table in all_tables:
        try:
            headers = [th.text.strip().upper().replace(" ", "") for th in table.find_elements(By.CSS_SELECTOR, "thead th")]
            if "UNIT" not in headers or not any(h in headers for h in ("20STD", "40STD", "40HC")):
                continue
            unit_idx = headers.index("UNIT")
            curr_idx = headers.index("CURR.") if "CURR." in headers else (headers.index("CURR") if "CURR" in headers else None)
            col_map = {}
            for key in ("20STD", "40STD", "40HC"):
                if key in headers:
                    col_map[headers.index(key)] = key
            if not col_map:
                continue

            for row in table.find_elements(By.CSS_SELECTOR, "tbody tr"):
                tds = row.find_elements(By.CSS_SELECTOR, "td")
                if len(tds) <= unit_idx:
                    continue
                unit_text = re.sub(r"[^A-Z]", "", tds[unit_idx].text.upper())
                if unit_text not in ("CTR", "CNTR", "CONTAINER"):
                    continue

                curr = "USD"
                if curr_idx is not None and curr_idx < len(tds):
                    curr = tds[curr_idx].text.strip().upper()
                rate = 1.0
                if curr == "EUR":
                    rate = EUR_TO_USD
                elif curr and curr != "USD":
                    print(f"   [Tab{tab_idx}] [Total] Skip top total table currency={curr}")
                    return None

                totals = {}
                for col_idx, cont_key in col_map.items():
                    if col_idx >= len(tds):
                        continue
                    amount = _parse_hpl_amount(tds[col_idx].text)
                    if amount is None:
                        totals[cont_key] = "-"
                        totals[f"{cont_key}_FORMULA"] = None
                    else:
                        usd_amount = amount * rate
                        totals[cont_key] = math.ceil(usd_amount)
                        totals[f"{cont_key}_FORMULA"] = _excel_formula_from_parts([usd_amount])
                if totals:
                    print(f"   [Tab{tab_idx}] [Total] Ctr summary table: {totals}")
                    return totals
        except Exception:
            continue
    return None

def parse_hpl_price(tab_idx, pod=""):
    china_route = is_china_destination(pod=pod)
    charges      = {}
    formula_parts = {"20STD": [], "40STD": [], "40HC": []}
    has_thc_orig = False
    ows_entries  = []   # list of (weight_tons, price_20, price_40)
    in_ows       = False
    avail_cols   = set()
    ocean_freight_cols = set()
    top_totals   = None
    china_thc_added = False

    try:
        hpl_raise_if_service_unavailable(tab_idx, "trước Price Breakdown")
        WebDriverWait(driver, 10).until(
            lambda d: (
                hpl_unavailable_banner_visible()
                or d.find_elements(By.CSS_SELECTOR, ".offer-charges")
            )
        )
        hpl_raise_if_service_unavailable(tab_idx, "Price Breakdown")
        # Wait: đếm tổng row 2 lần liên tiếp cách nhau 0.5s phải bằng nhau → DOM ổn định
        def count_all_rows(d):
            tbls = d.find_elements(By.CSS_SELECTOR, ".offer-charges table.q-table")
            return sum(len(t.find_elements(By.CSS_SELECTOR, "tbody tr")) for t in tbls)
        try:
            WebDriverWait(driver, 15).until(lambda d: count_all_rows(d) > 0)
            import time as _t
            _t.sleep(0.2) 
        except:
            rand_sleep(3.0, 4.0)

        all_tables = driver.find_elements(By.CSS_SELECTOR, ".offer-charges table.q-table")
        print(f"   [Tab{tab_idx}] [Tables] {len(all_tables)} tables, {count_all_rows(driver)} rows")
        top_totals = _read_hpl_container_total_table(all_tables, tab_idx)
        dialog_text = ""
        try:
            dialog_text = driver.execute_script("""
                const dlg = document.querySelector('.q-dialog') || document.body;
                return dlg ? dlg.innerText : '';
            """) or ""
        except Exception:
            dialog_text = ""
        othc_included = _hpl_ocean_freight_includes_othc(dialog_text)

        # PASS 1: Scan TẤT CẢ tables để detect THC (kể cả Export Surcharges)
        for tbl in all_tables:
            for row in tbl.find_elements(By.CSS_SELECTOR, "tbody tr"):
                tds = row.find_elements(By.CSS_SELECTOR, "td")
                charge_upper = tds[0].text.upper() if tds else ""
                if (
                    "TERMINAL HANDLING" in charge_upper
                    and ("ORIG" in charge_upper or "ORIGIN" in charge_upper)
                ):
                    has_thc_orig = True
                    print(f"   [Tab{tab_idx}] [THC] Tìm thấy local THC origin")
                    break
            if has_thc_orig:
                break

                        # PASS 2: Cong Freight Charges/Freight Surcharges
        for table in all_tables:
            try:
                all_ths  = table.find_elements(By.CSS_SELECTOR, "thead th")
                th_texts = [th.text.strip() for th in all_ths]
                # Find section by looking at the first non-empty text before 'Unit'
                section = ""
                for t in th_texts:
                    if t.upper() in ("UNIT", "CURR.", "CURR"): break
                    if t: section = t
            except:
                continue
                
            unit_idx = -1
            curr_idx = -1
            for i, txt in enumerate(th_texts):
                txt_up = txt.upper()
                if txt_up == "UNIT": unit_idx = i
                elif txt_up in ("CURR.", "CURR"): curr_idx = i

            charge_idx = unit_idx - 1 if unit_idx > 0 else 0

            col_map = {}
            for i, txt in enumerate(th_texts):
                if txt in ("20STD", "40STD", "40HC"):
                    col_map[i] = txt
                    avail_cols.add(txt)

            for row in table.find_elements(By.CSS_SELECTOR, "tbody tr"):
                try:
                    tds = row.find_elements(By.CSS_SELECTOR, "td")
                    if len(tds) <= max(charge_idx, unit_idx, curr_idx): continue
                    
                    full_charge_name = tds[charge_idx].text.strip()
                    charge_name = full_charge_name.split("\n")[0]
                    unit_text   = tds[unit_idx].text.strip() if unit_idx >= 0 else ""
                    currency    = tds[curr_idx].text.strip() if curr_idx >= 0 else ""
                    
                    charge_scope = f"{section} {full_charge_name} {unit_text}".upper()
                    is_freight_charge = "FREIGHT" in section.upper()
                    is_rail_charge = "RAIL" in charge_scope
                    is_origin_thc_charge = (
                        "TERMINAL HANDLING" in charge_scope
                        and any(marker in charge_scope for marker in ("ORIG", "ORIGIN", "PORT OF LOADING"))
                        and not any(marker in charge_scope for marker in ("DESTINATION", "IMPORT", "DISCHARGE"))
                    )

                    if not is_freight_charge and not is_rail_charge and not (china_route and is_origin_thc_charge):
                        continue

                    # --- OWS: Phát hiện Heavy Lift / Weight Tier → skip cộng giá, thu thập chi tiết ---
                    full_text = full_charge_name
                    is_ows_row = False
                    if "Heavy Lift" in charge_name or "Weight Tier" in charge_name:
                        in_ows = True
                        is_ows_row = True
                    elif in_ows and "Between" in full_text:
                        is_ows_row = True
                    else:
                        if charge_name and "Between" not in full_text:
                            in_ows = False

                    if is_ows_row:
                        # Parse weight threshold từ description
                        wt_match = re.search(r'[Bb]etween\s+([\d.]+)\s+and', full_text)
                        wt_tons = 0
                        if wt_match:
                            wt_tons = int(float(wt_match.group(1)))
                        # Parse giá cho từng container
                        ows_20 = 0
                        ows_40 = 0
                        for col_idx, cont_key in col_map.items():
                            if col_idx < len(tds):
                                val_str = re.sub(r'[\s,\u00a0\u202f]', '', tds[col_idx].text.strip())
                                if val_str and val_str != '-':
                                    try:
                                        v = float(val_str)
                                        if cont_key == '20STD':
                                            ows_20 = int(v)
                                        elif cont_key in ('40STD', '40HC') and ows_40 == 0:
                                            ows_40 = int(v)
                                    except: pass
                        if wt_tons > 0 and (ows_20 > 0 or ows_40 > 0):
                            ows_entries.append((wt_tons, ows_20, ows_40))
                        print(f"   [Tab{tab_idx}] [OWS] {full_text[:60]} → wt={wt_tons} 20'=${ows_20} 40'=${ows_40}")
                        continue

                    if col_map:
                        rate = get_hpl_exchange_rate(currency)
                        if rate is None:
                            print(f"   [Tab{tab_idx}] [FX] Bỏ qua {charge_name}: chưa có tỷ giá {currency}->USD")
                            continue
                        if currency != "USD":
                            print(f"   [Tab{tab_idx}] [FX] {charge_name} ({currency}) → USD x{rate}")

                        for col_idx, cont_key in col_map.items():
                            if col_idx < len(tds):
                                val_str = re.sub(r'[\s,\u00a0\u202f]', '', tds[col_idx].text.strip())
                                if val_str and val_str != "-":
                                    try:
                                        amount = float(val_str) * rate
                                        charges[cont_key] = charges.get(cont_key, 0.0) + amount
                                        formula_parts.setdefault(cont_key, []).append(amount)
                                        if "OCEAN FREIGHT" in charge_name.upper():
                                            ocean_freight_cols.add(cont_key)
                                        if is_origin_thc_charge:
                                            china_thc_added = True
                                    except: pass
                        log = [(k, tds[i].text.strip()) for i, k in col_map.items() if i < len(tds)]
                        charge_kind = "RAIL" if is_rail_charge else "FREIGHT"
                        print(f"   [Tab{tab_idx}] [+{charge_kind}] {section} | {charge_name} "
                              f"[Unit={unit_text}] ({currency}): {log}")

                except: continue

        if False and not has_thc_orig:
            print(f"   [Tab{tab_idx}] ⚠️ Không có THC Orig → Trừ $140/$210")
            if "20STD" in avail_cols:
                charges["20STD"] = charges.get("20STD", 0.0) - 140.0
                formula_parts.setdefault("20STD", []).append(-140.0)
            if "40STD" in avail_cols:
                charges["40STD"] = charges.get("40STD", 0.0) - 210.0
                formula_parts.setdefault("40STD", []).append(-210.0)
            if "40HC"  in avail_cols:
                charges["40HC"]  = charges.get("40HC",  0.0) - 210.0
                formula_parts.setdefault("40HC", []).append(-210.0)

        if othc_included:
            print(f"   [Tab{tab_idx}] [THC] Ocean Freight includes Terminal Handling Charge Orig")
        if china_route and has_thc_orig and not china_thc_added and not othc_included:
            print(f"   [Tab{tab_idx}] [THC CHINA] Cảnh báo: thấy O.THC riêng nhưng không đọc được số tiền")

        # Chỉ ghi INCLUDED khi chính Ocean Freight nói rõ bao gồm THC origin.
        # Không thấy dòng local THC không đồng nghĩa với việc THC đã included.
        base_remark = build_subject_remark(othc_included=(othc_included or china_route), pod=pod)

        # --- GHI REMARK ĐÚNG CHUẨN (bao gồm chi tiết OWS) ---
        if ows_entries:
            ows_parts = []
            for wt, p20, p40 in ows_entries:
                if p20 > 0:
                    ows_parts.append(f"OWS ${p20}/20' (>{wt} TONS)")
                if p40 > 0:
                    ows_parts.append(f"OWS ${p40}/40' (>{wt} TONS)")
            if ows_parts:
                remark = base_remark + ", " + ", ".join(ows_parts)
            else:
                remark = base_remark + ", OWS"
        else:
            remark = base_remark

    except (HPLRoutingUnavailable, HPLServiceUnavailable):
        raise
    except Exception as e:
        print(f"   [Tab{tab_idx}] ⚠️ Lỗi parse giá: {e}")
        remark = build_subject_remark(othc_included=china_route, pod=pod)

    if charges:
        totals = {}
        for key in ("20STD", "40STD", "40HC"):
            has_ocean_freight = key in ocean_freight_cols
            totals[key] = math.ceil(charges[key]) if (has_ocean_freight and charges.get(key, 0) > 0) else "-"
            totals[f"{key}_FORMULA"] = (
                _excel_formula_from_parts(formula_parts.get(key, []))
                if has_ocean_freight else None
            )
    elif top_totals:
        # Fallback only. Top Ctr total is all-in and may contain local destination charges,
        # so it must never override a successfully parsed selective charge calculation.
        print(f"   [Tab{tab_idx}] [Total] Dùng summary table vì không parse được charge chi tiết")
        totals = top_totals
    else:
        totals = {}
        for key in ("20STD", "40STD", "40HC"):
            totals[key] = "-"
            totals[f"{key}_FORMULA"] = None

    return totals, remark

# ===================================================================================
# --- FORMAT ETD/TT ---
# ===================================================================================
def format_etd_tt(etd_list):
    """Rebuild str_etd, str_tt từ danh sách ETD đã lọc."""
    def _fmt(dt):
        return f"{dt.day}-{dt.strftime('%b')}"
    num = len(etd_list)
    if num == 0: return "N/A", "N/A"
    elif num == 1: s = _fmt(etd_list[0]["etd_dt"])
    elif num == 2:
        s = (f"{_fmt(etd_list[0]['etd_dt'])} & "
             f"{_fmt(etd_list[1]['etd_dt'])}")
    else:
        s = (f"{etd_list[0]['etd_dt'].day}, "
             f"{etd_list[1]['etd_dt'].day}, "
             f"{_fmt(etd_list[2]['etd_dt'])}")
    all_tt = [c["tt_days"] for c in etd_list]
    tt = str(min(all_tt)) if min(all_tt)==max(all_tt) else f"{min(all_tt)}-{max(all_tt)}"
    return s, tt

# ===================================================================================
# --- LẤY GIÁ & GHI EXCEL ---
# ===================================================================================
def get_price_and_save(row_i, tab_idx, wb, ws, job_pod=""):
    try:
        state = hpl_wait_search_settled(tab_idx, timeout=45, min_wait=5, empty_confirm=4)
        if state == "SERVICE_UNAVAILABLE":
            hpl_raise_if_service_unavailable(tab_idx, "sau Search")
        if state == "SOLD_OUT":
            hpl_dismiss_no_service_notification()
            raise Exception("NO SERVICE / SOLD OUT")
        if state == "TIMEOUT":
            print(f"   [Tab{tab_idx}] ⚠️ Sau Search chưa thấy card/no-service rõ ràng, vẫn thử đọc kết quả...")

        # Ưu tiên đọc từ popup Departure Details (nhanh hơn + có lịch tàu)
        all_cards = read_departure_details(tab_idx)
        if all_cards is None:
            print(f"   [Tab{tab_idx}] ⚠️ Popup thất bại, dùng carousel cũ")
            all_cards = read_all_cards(tab_idx)
        if not all_cards:
            _blank_price_cells(ws, row_i)
            wb.save(excel_path); return

        etd_chuan, _, _ = apply_9_golden_rules(all_cards)
        if not etd_chuan:
            _blank_price_cells(ws, row_i)
            wb.save(excel_path); return

        # --- Click từng card ETD để đọc valid_to, chỉ mở Price Breakdown 1 lần ---
        def find_and_click_card(etd_target):
            """Tìm card theo ETD string, click vào, return True/False."""
            reset_carousel(tab_idx)
            rand_sleep(0.15, 0.25)
            for _ in range(6):
                for card in driver.find_elements(By.CSS_SELECTOR, "button.carousel__item"):
                    try:
                        if card.find_element(By.CSS_SELECTOR, "div.carousel__date span").text.strip() == etd_target:
                            human_move_and_click(card)
                            rand_sleep(0.2, 0.3)
                            return True
                    except: continue
                try:
                    next_btn = driver.find_element(By.CSS_SELECTOR, "button[aria-label='Next Departures']")
                    if next_btn.get_attribute("disabled") is None:
                        driver.execute_script("arguments[0].click();", next_btn)
                        rand_sleep(0.3, 0.5)
                    else:
                        break
                except:
                    break
            return False

        totals = None
        remark = ""
        valid_to_list = []
        best_etd_list = []

        for idx_etd, etd_info in enumerate(etd_chuan):
            etd_target = etd_info["etd_str"]

            clicked_card = find_and_click_exact_hpl_card(tab_idx, etd_info)
            if not clicked_card:
                print(f"   [Tab{tab_idx}] ⚠️ Không tìm lại card ETD={etd_target}")
                if totals is None:
                    print(f"   [Tab{tab_idx}] Fallback: click card co gia dau tien de doc Price Breakdown")
                    clicked_card = find_and_click_any_priced_hpl_card(tab_idx)
                if not clicked_card:
                    if totals is not None:
                        best_etd_list.append(etd_info)
                    continue

            # Đọc valid_to từ sidebar cho MỖI card
            vt = read_valid_from_sidebar(tab_idx) or etd_info.get("valid_to", "")
            valid_to_list.append(vt)
            best_etd_list.append(etd_info)
            print(f"   [Tab{tab_idx}] 📅 ETD={etd_target} Valid={vt}")

            # Chỉ mở Price Breakdown ở card ĐẦU TIÊN (cùng QQ price = cùng breakdown)
            if totals is None:
                try:
                    price_btn = WebDriverWait(driver, 8).until(EC.element_to_be_clickable(
                        (By.XPATH, "//button[.//span[text()='Price Breakdown']]")))
                    human_move_and_click(price_btn)
                    rand_sleep(0.2, 0.3)
                except Exception as e:
                    print(f"   [Tab{tab_idx}] ⚠️ Không click Price Breakdown: {e}")
                    _blank_price_cells(ws, row_i)
                    wb.save(excel_path); return

                totals, remark = parse_hpl_price(tab_idx, pod=job_pod)
                print(f"   [Tab{tab_idx}] 💰 ETD={etd_target}: 20'={totals.get('20STD')} 40'={totals.get('40STD')} 40HC={totals.get('40HC')}")

                # Đóng Price Breakdown
                try:
                    close_btn = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.q-dialog__x"))
                    )
                    driver.execute_script("arguments[0].click();", close_btn)
                    rand_sleep(0.2, 0.3)
                except:
                    try:
                        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                        rand_sleep(0.2, 0.3)
                    except: pass

        if not best_etd_list or totals is None:
            _blank_price_cells(ws, row_i)
            wb.save(excel_path); return

        # Valid lấy từ card ETD cuối cùng
        valid_to = valid_to_list[-1] if valid_to_list else ""
        str_etd, str_tt = format_etd_tt(best_etd_list)

        print(f"   [Tab{tab_idx}] 🏆 ETD: {str_etd} | T/T: {str_tt}")
        ws.cell(row=row_i, column=6).value  = totals.get("20STD_FORMULA") or totals.get("20STD")
        ws.cell(row=row_i, column=7).value  = totals.get("40STD_FORMULA") or totals.get("40STD")
        ws.cell(row=row_i, column=8).value  = totals.get("40HC_FORMULA") or totals.get("40HC")
        ws.cell(row=row_i, column=9).value  = str_etd
        ws.cell(row=row_i, column=10).value = str_tt
        ws.cell(row=row_i, column=11).value = valid_to
        ws.cell(row=row_i, column=13).value = remark

        # Cột O (15): Tất cả tàu đạt chuẩn — nhiều dòng trong 1 ô
        vessel_lines = []
        for e_info in best_etd_list:
            v_name = e_info.get("vessel", "")
            e_date = f"{e_info['etd_dt'].day}-{e_info['etd_dt'].strftime('%b')}"
            e_tt   = e_info.get("tt_days", "")
            e_ts   = e_info.get("transshipment", "")
            vessel_lines.append(
                f"{v_name} / ETD: {e_date} / Transit time: {e_tt} Days / Transshipment Port: {e_ts}"
            )
        ws.cell(row=row_i, column=15).value = "\n".join(vessel_lines)

        # Cột P (16): Transshipment ports
        ts_per_card = []
        for e_info in best_etd_list:
            ts = e_info.get("transshipment", "")
            if ts and ts != "DIRECT":
                ts_per_card.append(ts)
        if not ts_per_card:
            ts_col = "DIRECT"
        else:
            unique_ts = list(dict.fromkeys(ts_per_card))  # giữ thứ tự
            ts_col = " or\n".join(unique_ts)
        ws.cell(row=row_i, column=16).value = ts_col

        print(f"   [Tab{tab_idx}] 📅 Valid={valid_to} | 📝 {remark}")
        print(f"   [Tab{tab_idx}] 🚢 Vessels: {len(vessel_lines)} | T/S: {ts_col}")
        try:
            wb.save(excel_path)
            print(f"   [Tab{tab_idx}] 💾 Saved dòng {row_i}")
        except PermissionError:
            print(f"   [Tab{tab_idx}] ❌ Tắt Excel đi!")

    except HPLTabReplaced:
        raise
    except (HPLRoutingUnavailable, HPLServiceUnavailable):
        raise
    except Exception as e:
        print(f"   [Tab{tab_idx}] ❌ Lỗi: {e}")
        _blank_price_cells(ws, row_i)
        try: wb.save(excel_path)
        except: pass

# ===================================================================================
# --- MAIN ---
# ===================================================================================

# ===================================================================================
# --- ĐĂNG NHẬP ---
# ===================================================================================
def handle_login():
    print("   [HỆ THỐNG] Kiểm tra trạng thái đăng nhập...")
    target = switch_to_hpl_login_or_quote_tab()
    if not target:
        handles = list(driver.window_handles)
        if handles:
            driver.switch_to.window(handles[0])
        else:
            driver.switch_to.new_window("tab")
        driver.get(BASE_URL)
    else:
        try:
            url = (driver.current_url or "").lower()
            if (
                "identity.hapag-lloyd.com" not in url
                and "/solutions/auth/login" not in url
                and not any(marker in url for marker in HPL_QUOTE_URL_MARKERS)
                and not hpl_quote_form_ready()
                and not hpl_operational_page_present()
            ):
                print("   [HỆ THỐNG] Tab HPL chưa ở quote/login, mở trang quote trước...")
                driver.get(BASE_URL)
        except Exception:
            pass

    # Hard gate: nếu màn hình đang là Cloudflare/CAPTCHA/security thì bot phải
    # dừng hẳn tại đây, không login, không đọc form, không mở tab khác.
    time.sleep(2)
    if hpl_manual_check_present():
        if not wait_for_manual_hpl_unlock("captcha/security ngay sau khi mở HPL"):
            raise RuntimeError(f"HPL captcha/security chưa pass. URL={driver.current_url}")

    if complete_hpl_login_if_needed(reason="khoi dong HPL"):
        print("   [HỆ THỐNG] Đã đăng nhập sẵn / form quote sẵn sàng.")
        return True

    raise RuntimeError(f"HPL không vào được form quote/login. URL={driver.current_url}")

print("""
╔══════════════════════════════════════════════╗
║   HPL Price Checker — 3 Tab Pipeline  🚢     ║
╚══════════════════════════════════════════════╝
""")

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# --- Đọc Free Time từ freetime_hpl.xlsx ---
freetime_map = {}
try:
    ft_path = os.path.join(current_folder, "freetime_hpl.xlsx")
    if os.path.exists(ft_path):
        ft_wb = openpyxl.load_workbook(ft_path, read_only=True)
        ft_ws = ft_wb.active
        for ft_row in ft_ws.iter_rows(min_row=2, values_only=True):
            country = str(ft_row[0] or "").strip().upper()
            ft_val  = str(ft_row[1] or "").strip()
            if country and ft_val:
                freetime_map[country] = ft_val
        ft_wb.close()
        print(f"📋 Đọc được {len(freetime_map)} quốc gia từ freetime_hpl.xlsx")
    else:
        print("⚠️ Không tìm thấy freetime_hpl.xlsx — bỏ qua free time")
except Exception as e:
    print(f"⚠️ Lỗi đọc freetime_hpl.xlsx: {e}")

# --- Điền Free Time vào cột N (14) cho các dòng HAPAG LLOYD ---
if freetime_map:
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        carrier = str(row[4] or "").strip().upper()
        if carrier not in HPL_GROUP:
            continue
        country = str(row[1] or "").strip().upper()   # Cột B = country
        ft = freetime_map.get(country, "")
        if ft:
            ws.cell(row=i, column=14).value = ft
    try:
        wb.save(excel_path)
        print(f"✅ Đã điền free time cho các dòng HPL")
    except:
        pass

row_queue = []
# Port mapping: Excel name → HPL search name (POD)
HPL_POD_MAPPING = {
    "TIANJIN": "XINGANG",
    "PORT KLANG": "PORT KELANG",
    "BASEL": "BASLE",
    }

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if SINGLE_ROW and i != SINGLE_ROW:
        continue
    pol_excel = str(row[2] or "").strip().upper()
    pod       = str(row[3] or "").strip().title()
    carrier   = str(row[4] or "").strip().upper()
    if not pol_excel or not pod: continue
    if carrier not in HPL_GROUP: continue
    if FILTER_POL and pol_excel != FILTER_POL: continue
    if FILTER_POD and pod.upper() != FILTER_POD: continue
    pol_search = POL_MAP.get(pol_excel, pol_excel.title())
    # Áp dụng port mapping cho POD search
    pod = HPL_POD_MAPPING.get(pod.upper(), pod)
    row_queue.append((i, pol_search, pod))

# Sort theo POL → cùng POL liên tiếp nhau → skip nhập lại POL, tiết kiệm thời gian
row_queue.sort(key=lambda x: x[1])
total = len(row_queue)
if SINGLE_ROW:
    print(f"[SINGLE_ROW] Chi chay dong {SINGLE_ROW} theo lenh tu main.py")
if total <= 0:
    print("Khong co dong HPL can chay.")
    try:
        wb.save(excel_path)
        wb.close()
    except:
        pass
    raise SystemExit(0)
if total < NUM_TABS:
    NUM_TABS = total
tab_last_pol = [None] * NUM_TABS
tab_last_pod = [None] * NUM_TABS
print(f"📋 Tổng cộng {total} dòng HPL (đã sort theo POL)")

# Gọi hàm xử lý đăng nhập ở đây
handle_login()
if HPL_SAFE_SINGLE_TAB and NUM_TABS > 1:
    print("   [HPL SAFE] Đã có login/captcha thủ công -> ép chạy 1 tab, không mở/switch nhiều tab.")
    NUM_TABS = 1
    tab_last_pol = [None] * NUM_TABS
    tab_last_pod = [None] * NUM_TABS

# ── CHECK TAB ĐÃ MỞ SẴN WEB HPL ──
STEALTH_SCRIPT = HPL_STEALTH_SCRIPT

tabs = []
print("🔍 Đang kiểm tra các tab đã mở sẵn...")
if HPL_SAFE_SINGLE_TAB and hpl_quote_form_ready() and not hpl_manual_check_present():
    tabs.append(driver.current_window_handle)
    print(f"   [HPL SAFE] Dùng tab hiện tại đã pass: {driver.current_window_handle[:8]}...")
else:
    all_handles = driver.window_handles
    for h in all_handles:
        if len(tabs) >= NUM_TABS:
            break
        try:
            driver.switch_to.window(h)
            url = driver.current_url or ""
            if hpl_quote_tab_candidate():
                # Tab đã mở sẵn trang tìm giá HPL
                if hpl_manual_check_present():
                    print(f"   ⚠️ Tab HPL có sẵn bị captcha/security: {h[:8]}... -> silent chờ bạn giải")
                    wait_for_manual_hpl_unlock(f"tab co san {h[:8]} captcha/security")
                state = wait_hpl_quote_or_block(timeout=5)
                if state in ("LOGIN", "BLOCKED"):
                    complete_hpl_login_if_needed(reason=f"tab co san {h[:8]} state={state}")
                    state = wait_hpl_quote_or_block(timeout=5)
                if state == "READY":
                    tabs.append(h)
                    print(f"   ✅ Tái sử dụng tab có sẵn: {h[:8]}... (URL: {url[:60]})")
                else:
                    print(f"   ⚠️ Tab HPL có sẵn chưa sẵn sàng ({state}), bỏ qua tab: {h[:8]}...")
        except:
            pass

if tabs:
    print(f"♻️  Tìm thấy {len(tabs)} tab HPL có sẵn, tái sử dụng!")

# Mặc định chạy 3-tab pipeline như cũ. Nếu muốn cực kỳ an toàn với captcha,
# đặt HPL_OPEN_MISSING_TABS=0 để chỉ dùng tab HPL đã mở sẵn.
if not tabs and not HPL_OPEN_MISSING_TABS and NUM_TABS > 1:
    print("   [HPL SAFE] Chưa có tab HPL dùng được; chỉ tự mở 1 tab để giảm rủi ro captcha.")
    print("   [HPL SAFE] Muốn chạy 3 tab: bỏ HPL_OPEN_MISSING_TABS=0 hoặc mở sẵn 3 tab HPL quote.")
    NUM_TABS = 1
    tab_last_pol = [None] * NUM_TABS
    tab_last_pod = [None] * NUM_TABS

if tabs and not HPL_OPEN_MISSING_TABS:
    if len(tabs) < NUM_TABS:
        print(f"   [HPL SAFE] Không tự mở thêm tab thiếu ({len(tabs)}/{NUM_TABS}) để tránh captcha/reload.")
        print("   [HPL SAFE] Muốn đủ 3 tab: bỏ HPL_OPEN_MISSING_TABS=0 hoặc mở sẵn 3 tab HPL quote.")
        NUM_TABS = len(tabs)
        tab_last_pol = [None] * NUM_TABS
        tab_last_pod = [None] * NUM_TABS
    need_more = 0
else:
    need_more = NUM_TABS - len(tabs)
if need_more > 0:
    print(f"🌐 Cần mở thêm {need_more} tab mới...")
    for i in range(need_more):
        tab_num = len(tabs) + 1
        try:
            new_h = open_fresh_hpl_quote_tab(tab_idx=tab_num, reason=f"init Tab{tab_num}")
            tabs.append(new_h)
        except Exception as e:
            print(f"   ❌ Không mở được Tab{tab_num} HPL mới: {e}")
            try:
                live_handles = set(driver.window_handles)
                tabs[:] = [handle for handle in tabs if handle in live_handles]
            except Exception:
                tabs[:] = []
            if tabs:
                print(
                    f"   [HPL RECOVERY] Giữ {len(tabs)} tab đang sống và tiếp tục; "
                    "không làm crash toàn bộ bot."
                )
                break
            raise
        rand_sleep(0.8, 1.2)
else:
    print(f"♻️  Đã đủ {NUM_TABS} tab HPL có sẵn, không cần mở thêm. Tiết kiệm RAM!")

if not tabs:
    raise RuntimeError("HPL không còn tab quote nào sử dụng được")
if len(tabs) < NUM_TABS:
    print(f"   [HPL RECOVERY] Pipeline tạm giảm từ {NUM_TABS} xuống {len(tabs)} tab sống.")
    NUM_TABS = len(tabs)
    tab_last_pol = [None] * NUM_TABS
    tab_last_pod = [None] * NUM_TABS

print(f"🌐 {NUM_TABS} tabs sẵn sàng: {[h[:8] for h in tabs]}\n")

# ===================================================================================
# PIPELINE
# ===================================================================================
queue_idx = 0
pending   = [None] * NUM_TABS

def search_job(t, job):
    service_unavailable_retried = False
    for attempt in range(1, HPL_TAB_REPLACE_ATTEMPTS + 2):
        driver.switch_to.window(tabs[t])
        activate_tab(tabs[t])
        print(f"\n{'='*50}")
        print(f"[Tab{t+1}] Dòng {job[0]}: {job[1]} → {job[2]}")
        try:
            do_search(job[1], job[2], t+1)
            return True
        except HPLRoutingUnavailable:
            print(f"[Tab{t+1}] ⏭️ Routing unavailable -> bỏ ngay dòng {job[0]}, không retry")
            _blank_price_cells(ws, job[0])
            try: wb.save(excel_path)
            except: pass
            tab_last_pol[t] = None
            tab_last_pod[t] = None
            return False
        except HPLServiceUnavailable as e:
            if not service_unavailable_retried:
                service_unavailable_retried = True
                print(f"[Tab{t+1}] 🔁 HPL service unavailable -> retry dòng {job[0]} 1 lần")
                try:
                    hpl_soft_return_to_form(timeout=HPL_BACK_FALLBACK_WAIT_SECONDS)
                except Exception:
                    pass
                tab_last_pol[t] = None
                tab_last_pod[t] = None
                continue
            print(f"[Tab{t+1}] ⏭️ HPL service unavailable lần 2 -> skip dòng {job[0]}")
            _blank_price_cells(ws, job[0])
            try: wb.save(excel_path)
            except: pass
            tab_last_pol[t] = None
            tab_last_pod[t] = None
            return False
        except HPLTabReplaced as e:
            print(f"[Tab{t+1}] 🔁 Tab bị captcha đã thay mới, retry lại dòng {job[0]} ({attempt}/{HPL_TAB_REPLACE_ATTEMPTS + 1})")
            tab_last_pol[t] = None
            tab_last_pod[t] = None
            continue
        except HPLPortNoResults as e:
            print(f"[Tab{t+1}] ⏭️ Port không có trong HPL autocomplete: {e}")
            _blank_price_cells(ws, job[0])
            try: wb.save(excel_path)
            except: pass
            tab_last_pod[t] = None
            return False
        except Exception as e:
            msg = str(e)
            print(f"[Tab{t+1}] ❌ Search lỗi: {msg}")
            _blank_price_cells(ws, job[0])
            try: wb.save(excel_path)
            except: pass
            try:
                print(f"[Tab{t+1}] ↩️ Lỗi search: quay về form nhẹ, không reload HPL...")
                hpl_soft_return_to_form(timeout=HPL_BACK_FALLBACK_WAIT_SECONDS)
                # Reset cache sau khi quay form vì UI có thể đã reset
                tab_last_pol[t] = None
                tab_last_pod[t] = None
            except Exception as nav_e:
                print(f"[Tab{t+1}] ⚠️ Không quay form nhẹ được: {type(nav_e).__name__}")
            return False
    print(f"[Tab{t+1}] ❌ Không search được dòng {job[0]} sau khi thay tab nhiều lần")
    return False

# --- Bước 1: Search 3 job đầu ---
import time as _timer
_pipeline_start = _timer.time()
print("🚀 Bước 1: Search 3 tab đầu tiên...")
for t in range(NUM_TABS):
    if queue_idx >= total:
        break
    job = row_queue[queue_idx]; queue_idx += 1
    ok  = search_job(t, job)
    pending[t] = job if ok else None
    while not ok and queue_idx < total:
        job = row_queue[queue_idx]; queue_idx += 1
        ok  = search_job(t, job)
        pending[t] = job if ok else None

# --- Bước 2: Pipeline ---
print(f"\n🔄 Bước 2: Pipeline...")
while any(p is not None for p in pending):
    for t in range(NUM_TABS):
        job = pending[t]
        if job is None:
            continue

        driver.switch_to.window(tabs[t])
        activate_tab(tabs[t])
        print(f"\n[Tab{t+1}] 💰 Lấy giá dòng {job[0]}...")
        try:
            get_price_and_save(job[0], t+1, wb, ws, job_pod=job[2])
        except HPLRoutingUnavailable:
            print(f"[Tab{t+1}] ⏭️ Routing unavailable -> bỏ ngay dòng {job[0]}, không retry/timeout")
            _blank_price_cells(ws, job[0])
            try: wb.save(excel_path)
            except: pass
        except HPLTabReplaced:
            print(f"[Tab{t+1}] 🔁 Tab bị captcha khi lấy giá, search lại dòng {job[0]} trên tab mới")
            tab_last_pol[t] = None
            tab_last_pod[t] = None
            ok = search_job(t, job)
            pending[t] = job if ok else None
            continue
        except HPLServiceUnavailable:
            print(f"[Tab{t+1}] 🔁 HPL service unavailable khi lấy giá -> retry dòng {job[0]} 1 lần")
            try:
                hpl_soft_return_to_form(timeout=HPL_BACK_FALLBACK_WAIT_SECONDS)
            except Exception:
                pass
            tab_last_pol[t] = None
            tab_last_pod[t] = None
            ok = search_job(t, job)
            if not ok:
                pending[t] = None
                continue
            try:
                get_price_and_save(job[0], t+1, wb, ws, job_pod=job[2])
            except HPLRoutingUnavailable:
                print(f"[Tab{t+1}] ⏭️ Routing unavailable trong lần retry -> skip dòng {job[0]}")
                _blank_price_cells(ws, job[0])
                try: wb.save(excel_path)
                except: pass
            except HPLServiceUnavailable:
                print(f"[Tab{t+1}] ⏭️ HPL service unavailable lần 2 khi lấy giá -> skip dòng {job[0]}")
                _blank_price_cells(ws, job[0])
                try: wb.save(excel_path)
                except: pass
            except HPLTabReplaced:
                print(f"[Tab{t+1}] 🔁 Tab bị captcha trong retry lấy giá, search lại dòng {job[0]} trên tab mới")
                tab_last_pol[t] = None
                tab_last_pod[t] = None
                ok = search_job(t, job)
                pending[t] = job if ok else None
                continue

        print(f"[Tab{t+1}] ⬅️ Back...")
        if not safe_go_back(t+1):
            print(f"[Tab{t+1}] ⚠️ Khong quay lai duoc form HPL, bo tab nay de bot khong bi ket.")
            pending[t] = None
            continue

        if queue_idx < total:
            new_job = row_queue[queue_idx]; queue_idx += 1
            ok = search_job(t, new_job)
            pending[t] = new_job if ok else None
            while not ok and queue_idx < total:
                new_job = row_queue[queue_idx]; queue_idx += 1
                ok = search_job(t, new_job)
                pending[t] = new_job if ok else None
        else:
            pending[t] = None
            print(f"[Tab{t+1}] ✅ Hết việc")

elapsed = _timer.time() - _pipeline_start
avg = elapsed / total if total > 0 else 0
print(f"\n✅ Hoàn tất {total} dòng trong {elapsed:.0f}s (trung bình {avg:.1f}s/dòng)")
