"""
BOT WHL (Wan Hai Lines) — Lấy lịch tàu (https://www.wanhai.com/views/skd/SkdByPort.xhtml)
=============================================================================================
Hoạt động:
- Đọc Excel input_gia.xlsx (hoặc env EXCEL_PATH)
- Lọc rows có CARRIER == "WHL"
- Với mỗi row: chọn POL/POD trong các <select> (chained: from_nation → pol; to_nation → pod),
  set dept_date_from = start (theo valid window) và arr_date_to = valid + buffer (60 ngày),
  bấm Search, theo link "ALL SERVICE DETAIL" để có ETD cụ thể, parse rows, áp dụng
  rule tối đa 3 ETD cách nhau ≥ 1 ngày + ETD ≤ valid (cột K).
- Ghi vào Excel cột I (9), J (10), O (15), P (16).
- Captcha: nếu xuất hiện sẽ pause cho user nhập thủ công.

QUY TẮC VALID WINDOW (giống Yang Ming):
  - valid.day ≤ 15 → start = 1 cùng tháng
  - valid.day > 15 → start = 15 cùng tháng
  - end = valid (cho ETD), arr_date_to = valid + 60 ngày để cover transit time

PORT ALIASES (Excel → WHL):
  HO CHI MINH → CAT LAI PORT HOCHIMINH
  HAI PHONG   → HAIPHONG
"""

import os
import re
import sys
import time
import socket
import subprocess
import winsound
from datetime import datetime, timedelta

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = open(sys.stdout.fileno(), mode="w", encoding="utf-8", buffering=1)
    sys.stderr = open(sys.stderr.fileno(), mode="w", encoding="utf-8", buffering=1)

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils.datetime import from_excel

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from bot_runtime_utils import wait_for_terminal_enter

# ===================================================================================
# CONFIG
# ===================================================================================
EXCEL_PATH = os.environ.get("EXCEL_PATH") or os.path.join(os.getcwd(), "input_gia.xlsx")
FILTER_POL = (os.environ.get("FILTER_POL") or "").strip().upper()
FILTER_POD = (os.environ.get("FILTER_POD") or "").strip().upper()
SINGLE_ROW = (os.environ.get("SINGLE_ROW") or "").strip()
WHL_AUTO_CAPTCHA_WAIT = int(os.environ.get("WHL_CAPTCHA_WAIT") or "900")  # tối đa 15 phút
WHL_MANUAL_CAPTCHA_WAIT_PATH = os.environ.get(
    "WHL_MANUAL_CAPTCHA_WAIT_PATH",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), ".whl_manual_captcha_wait"),
)
DATE_OFFSET_DAYS = parse_date_offset_days()

CARRIER_TARGET = "WHL"
DEBUG_PORT     = "9529"
DRIVER_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "msedgedriver.exe")
URL_WHL        = "https://www.wanhai.com/views/skd/SkdByPort.xhtml"
EDGE_EXE       = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
EDGE_PROFILE   = r"C:\edge_whl"

# Mapping POL/POD Excel name → option text trong dropdown WHL
WHL_PORT_ALIASES = {
    "HO CHI MINH":  "CAT LAI PORT HOCHIMINH",
    "HOCHIMINH":    "CAT LAI PORT HOCHIMINH",
    "HCM":          "CAT LAI PORT HOCHIMINH",
    "SAIGON":       "CAT LAI PORT HOCHIMINH",
    "CAT LAI":      "CAT LAI PORT HOCHIMINH",
    "HAI PHONG":    "HAIPHONG",
    "HAIPHONG":     "HAIPHONG",
    "VISAKHAPATNAM": "VIZAG",
    "CAI MEP":      "CAI MEP",
    "TIANJIN":      "XINGANG",
    "TIENTSIN":     "XINGANG",
}

WHL_DIRECT_PRIORITY_ROUTES = {
    ("HO CHI MINH", "CHENNAI"),
    ("HO CHI MINH", "MUNDRA"),
    ("HO CHI MINH", "NHAVA SHEVA"),
    ("HO CHI MINH", "VIZAG"),
}

# Quốc gia POL/POD → giá trị select#from_nation (case-sensitive, all uppercase)
# Có thể mở rộng tuỳ tuyến
COUNTRY_BY_PORT = {
    "CAT LAI PORT HOCHIMINH": "VIETNAM",
    "CAI MEP":                "VIETNAM",
    "HAIPHONG":               "VIETNAM",
    "QUY NHON":               "VIETNAM",
    "DA NANG (PORT)":         "VIETNAM",
    "DA NANG":                "VIETNAM",
    "SHANGHAI上海":            "CHINA",
    "SHEKOU蛇口":              "CHINA",
    "NINGBO寧波":              "CHINA",
    "XIAMEN廈門":              "CHINA",
    "QINGDAO青島":             "CHINA",
    "TIANJIN天津":             "CHINA",
    "DALIAN大連":              "CHINA",
    "FUZHOU福州":              "CHINA",
    "GUANGZHOU廣州":           "CHINA",
    # English-only variants
    "SHANGHAI":              "CHINA",
    "NINGBO":                "CHINA",
    "QINGDAO":               "CHINA",
    "DALIAN":                "CHINA",
    "XIAMEN":                "CHINA",
    "SHEKOU":                "CHINA",
    "TIANJIN":               "CHINA",
    "XINGANG":               "CHINA",
    "GUANGZHOU":             "CHINA",
    "FUZHOU":                "CHINA",
    "HONG KONG":             "HONG KONG",
    "SINGAPORE":             "SINGAPORE",
    "PORT KLANG":            "MALAYSIA",
    "PORT KELANG":           "MALAYSIA",
    "TANJUNG PELEPAS":       "MALAYSIA",
    "JAKARTA":               "INDONESIA",
    "SURABAYA":              "INDONESIA",
    "BANGKOK":               "THAILAND",
    "LAEM CHABANG":          "THAILAND",
    "MANILA":                "PHILIPPINES",
    "BUSAN":                 "KOREA",
    "INCHEON":               "KOREA",
    "TOKYO":                 "JAPAN",
    "YOKOHAMA":              "JAPAN",
    "OSAKA":                 "JAPAN",
    "NAGOYA":                "JAPAN",
    "CHENNAI":               "INDIA",
    "COCHIN":                "INDIA",
    "MUNDRA":                "INDIA",
    "NHAVA SHEVA":           "INDIA",
    "MUMBAI":                "INDIA",
    "KOLKATA":               "INDIA",
    "CALCUTTA":              "INDIA",
    "VIZAG":                 "INDIA",
    "VISAKHAPATNAM":         "INDIA",
    "JEBEL ALI":             "UNITED ARAB EMIRATES",
    "DAMMAM":                "SAUDI ARABIA",
    "JEDDAH":                "SAUDI ARABIA",
    "JEDDAH JEDDAH ISLAMIC PORT":   "SAUDI ARABIA",
    "TAIPEI":                "TAIWAN",
    "KAOHSIUNG":             "TAIWAN",
    "KEELUNG":               "TAIWAN",
    "TAICHUNG":              "TAIWAN",
    "MANZANILLO":            "MEXICO",
    "SALVADOR":              "BRAZIL",
    "DUNKERQUE":             "FRANCE",
    "DUNKIRK":               "FRANCE",
}

TODAY = datetime.now()
MONTHS_VN = {
    "JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,
    "JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12,
}

# ===================================================================================
# DRIVER
# ===================================================================================
def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(("127.0.0.1", int(port))) == 0


def ensure_edge_debug_port():
    if is_port_in_use(DEBUG_PORT):
        print(f"[WHL] Edge debug port {DEBUG_PORT} đã mở sẵn.")
        return

    print(f"[WHL] Edge debug port {DEBUG_PORT} chưa mở, tự khởi động Edge...")
    subprocess.Popen([
        EDGE_EXE,
        f"--remote-debugging-port={DEBUG_PORT}",
        f"--user-data-dir={EDGE_PROFILE}",
        "--disable-background-timer-throttling",
        "--disable-renderer-backgrounding",
        "--disable-backgrounding-occluded-windows",
        "--start-maximized",
    ])
    # FIX: bỏ time.sleep(1) sau khi port đã mở — port mở = Edge ready
    for _ in range(20):
        if is_port_in_use(DEBUG_PORT):
            return
        time.sleep(0.5)
    raise RuntimeError(f"Không mở được Edge debug port {DEBUG_PORT}")


ensure_edge_debug_port()
edge_options = Options()
edge_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{DEBUG_PORT}")
driver = webdriver.Edge(service=Service(executable_path=DRIVER_PATH), options=edge_options)
try:
    driver.maximize_window()
except Exception:
    try:
        info = driver.execute_cdp_cmd("Browser.getWindowForTarget", {})
        driver.execute_cdp_cmd("Browser.setWindowBounds", {
            "windowId": info["windowId"],
            "bounds": {"windowState": "maximized"}
        })
    except Exception:
        pass
WAIT = WebDriverWait(driver, 20)

print(f"[WHL] ✅ Đã attach Edge debug port {DEBUG_PORT}")

# ===================================================================================
# HELPERS
# ===================================================================================
def log(msg):
    print(f"[WHL] {msg}")

def ensure_live_window():
    global driver, WAIT
    try:
        handles = driver.window_handles
        if handles:
            try:
                driver.current_url
            except Exception:
                driver.switch_to.window(handles[0])
            return
        driver.execute_cdp_cmd("Target.createTarget", {"url": "about:blank"})
        time.sleep(0.5)
        driver.switch_to.window(driver.window_handles[-1])
    except Exception:
        driver = webdriver.Edge(service=Service(executable_path=DRIVER_PATH), options=edge_options)
        try:
            driver.maximize_window()
        except Exception:
            pass
        WAIT = WebDriverWait(driver, 20)
        try:
            handles = driver.window_handles
            if handles:
                driver.switch_to.window(handles[0])
        except Exception:
            pass

def safe_click(el):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.15)
        el.click()
    except Exception:
        driver.execute_script("arguments[0].click();", el)

def alias_port(name):
    key = (name or "").strip().upper()
    return WHL_PORT_ALIASES.get(key, name)

def whl_direct_priority_route(pol_excel, pod_excel):
    pol = (pol_excel or "").strip().upper()
    pod = (pod_excel or "").strip().upper()
    pol_key = "HO CHI MINH" if pol in ("HO CHI MINH", "HOCHIMINH", "HCM", "SAIGON", "CAT LAI") else pol
    pod_key = "NHAVA SHEVA" if pod in ("NHAVA SHEVA", "NHAVASHEVA") else pod
    if pod_key in ("VISAKHAPATNAM", "VISAKHAPATNAM (VIZAG)"):
        pod_key = "VIZAG"
    return (pol_key, pod_key) in WHL_DIRECT_PRIORITY_ROUTES

def parse_valid_date(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if hasattr(raw, "year") and hasattr(raw, "month") and hasattr(raw, "day"):
        return datetime(raw.year, raw.month, raw.day)
    s = str(raw).strip()
    if not s:
        return None
    if re.fullmatch(r"\d+(\.\d+)?", s):
        try:
            dt = from_excel(float(s))
            return datetime(dt.year, dt.month, dt.day)
        except Exception:
            pass
    m = re.fullmatch(r"(\d{1,2})-([A-Za-z]{3})", s)
    if m:
        mon = MONTHS_VN.get(m.group(2).upper())
        if mon:
            return datetime(TODAY.year, mon, int(m.group(1)))
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})", s)
    if m:
        return datetime(TODAY.year, int(m.group(2)), int(m.group(1)))
    fmts = ["%d-%b-%Y","%d/%m/%Y","%d/%m/%y","%Y/%m/%d",
            "%d %b %Y","%Y-%m-%d","%d-%m-%Y"]
    for f in fmts:
        try:
            dt = datetime.strptime(s, f)
            if dt.year == 1900:
                dt = dt.replace(year=TODAY.year)
            return dt
        except Exception:
            continue
    m = re.match(r"(\d{1,2})[-/\s]+([A-Za-z]{3})", s)
    if m:
        day = int(m.group(1)); mon = MONTHS_VN.get(m.group(2).upper())
        if mon:
            return datetime(TODAY.year, mon, day)
    log(f"   ⚠️ Không parse được valid='{s}'")
    return None

def compute_valid_window(valid_dt):
    if valid_dt is None:
        return None, None
    start = (TODAY + timedelta(days=DATE_OFFSET_DAYS)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    end = valid_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    return start, end

def fmt_whl_date(dt):
    return dt.strftime("%Y/%m/%d")

# ===================================================================================
# OPEN PAGE + DETECT CAPTCHA
# ===================================================================================
def open_search_page():
    ensure_live_window()
    driver.get(URL_WHL)
    try:
        WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, "select#from_nation")))
    except TimeoutException:
        # Có thể bị captcha gate
        pass
    time.sleep(1.0)
    handle_captcha_if_needed()

def handle_captcha_if_needed():
    """Phát hiện captcha và hú còi báo động cho tới khi giải quyết xong."""
    import winsound
    try:
        body_text = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        body_text = ""
    
    captcha_indicators = ["captcha", "verify code", "verification code", "驗證碼", "验证码", "security check is required", "i am human"]
    has_captcha_text = any(t in body_text for t in captcha_indicators)
    
    captcha_imgs = driver.find_elements(By.XPATH,
        "//img[contains(@src,'captcha') or contains(@src,'CaptchaImg') or contains(@src,'verify') or contains(@src,'imgCode')] | //iframe[contains(@src, 'hcaptcha')]"
    )
    
    if has_captcha_text or captcha_imgs:
        log("   🛑 PHÁT HIỆN CAPTCHA — Vui lòng mở cửa sổ WHL giải quyết Captcha!")
        try:
            winsound.Beep(1500, 500) # Hú còi 1 lần báo hiệu
        except: pass
        log("   [WHL SILENT] Bot sẽ IM LẶNG chờ bạn giải quyết. Xong thì quay lại terminal và nhấn ENTER.")
        try:
            print("   [WHL SILENT] Nhấn ENTER sau khi WHL đã pass captcha...")
            input()
        except EOFError:
            log(f"   [WHL SILENT] Không có stdin, sleep 60s.")
            time.sleep(60)
        except KeyboardInterrupt:
            raise
            
        # Check xem Captcha đã qua chưa (thấy thẻ Select hoặc thẻ Bảng kết quả xuất hiện)
        try:
            if driver.find_elements(By.CSS_SELECTOR, "select#from_nation") or driver.find_elements(By.XPATH, "//table//tr"):
                log("   ✅ Đã qua Captcha, tự động cày tiếp...")
        except:
            pass

# ===================================================================================
# WHL CAPTCHA OVERRIDE
# ===================================================================================
# The older handler above only waited for ENTER and then treated the page as passed
# if a form/table existed. WHL can show an hCaptcha/security page while the form is
# still present behind it, so this override blocks until the security text/widget
# really disappears.
def whl_captcha_present():
    """Return True while WHL security/captcha page is still present."""
    try:
        body_text = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        body_text = ""

    captcha_indicators = [
        "captcha",
        "hcaptcha",
        "i am human",
        "i'm human",
        "not a robot",
        "verify code",
        "verification code",
        "security check is required",
        "additional security check is required",
        "your request is currently unavailable",
        "currently unable to respond",
        "unable to respond to your request",
        "toi la con nguoi",
        "tôi là con người",
        "無法回應",
        "驗證碼",
        "验证码",
    ]
    if any(t in body_text for t in captcha_indicators):
        return True

    try:
        captcha_nodes = driver.find_elements(By.XPATH, """
            //iframe[
                contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'captcha') or
                contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'hcaptcha') or
                contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'recaptcha') or
                contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'turnstile')
            ] |
            //*[
                contains(translate(@id,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'captcha') or
                contains(translate(@id,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'hcaptcha') or
                contains(translate(@class,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'captcha') or
                contains(translate(@class,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'hcaptcha') or
                contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'captcha') or
                contains(translate(@src,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'verify') or
                contains(translate(@alt,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'captcha') or
                contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'captcha') or
                contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'verify')
            ]
        """)
        if captcha_nodes:
            return True
    except Exception:
        pass

    return False

def handle_captcha_if_needed(stage=""):
    """Stop hard on WHL captcha/security page until user really clears it."""
    import winsound

    if not whl_captcha_present():
        return False

    stage_msg = f" ({stage})" if stage else ""
    log(f"   [WHL SILENT] PHAT HIEN CAPTCHA/SECURITY{stage_msg}. Bot se dung han, khong search tiep.")

    while whl_captcha_present():
        try:
            winsound.Beep(1500, 500)
        except Exception:
            pass

        log("   [WHL SILENT] Hay giai captcha tren Edge WHL. Xong thi quay lai terminal va nhan ENTER.")
        try:
            print("   [WHL SILENT] Nhan ENTER sau khi WHL da pass captcha...")
            input()
        except EOFError:
            log(f"   [WHL SILENT] Khong co stdin, sleep {WHL_AUTO_CAPTCHA_WAIT}s roi kiem tra lai.")
            time.sleep(WHL_AUTO_CAPTCHA_WAIT)
        except KeyboardInterrupt:
            raise

        time.sleep(1.0)
        if whl_captcha_present():
            log("   [WHL SILENT] Van con captcha/security page -> tiep tuc dung, chua search lich tau.")

    log("   [WHL SILENT] Captcha da bien mat -> tiep tuc WHL.")
    return True

def open_search_page():
    ensure_live_window()
    driver.get(URL_WHL)
    time.sleep(1.0)
    handle_captcha_if_needed("open search page")
    try:
        WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, "select#from_nation")))
    except TimeoutException:
        handle_captcha_if_needed("wait search form")
        raise
    handle_captcha_if_needed("search form ready")

# ===================================================================================
# SELECT COUNTRY + PORT (chained)
# ===================================================================================
def select_country_and_port(country_select_id, port_select_id, port_target_text, country_hint=None):
    """
    Chọn country (from_nation/to_nation) rồi đợi <select id=pol/pod> load options
    rồi chọn option chứa port_target_text.
    """
    # B1: chọn country
    country_el = driver.find_element(By.ID, country_select_id)
    sel = Select(country_el)
    country = country_hint or COUNTRY_BY_PORT.get(port_target_text.upper())
    if not country:
        # Nếu không biết country, thử mọi country để tìm port
        log(f"   ⚠️ Không biết country của '{port_target_text}', sẽ thử VIETNAM/CHINA/SINGAPORE")
        for c in ["VIETNAM", "CHINA", "SINGAPORE", "KOREA", "TAIWAN", "MALAYSIA", "INDONESIA", "THAILAND", "INDIA", "HONG KONG"]:
            try:
                sel.select_by_visible_text(c)
            except Exception:
                continue
            time.sleep(1.5)
            port_el = driver.find_element(By.ID, port_select_id)
            options = [o.text for o in port_el.find_elements(By.TAG_NAME, "option")]
            if any(port_target_text.upper() in o.upper() for o in options):
                country = c
                log(f"   ✅ Tìm thấy '{port_target_text}' trong country '{c}'")
                break
        if not country:
            return False
    else:
        selected_country = False
        for opt in sel.options:
            if (opt.text or "").strip().upper() == country.upper():
                sel.select_by_visible_text(opt.text)
                selected_country = True
                break
        if not selected_country:
            sel.select_by_visible_text(country)
        time.sleep(1.5)

    # B2: chọn port
    port_el = driver.find_element(By.ID, port_select_id)
    options = port_el.find_elements(By.TAG_NAME, "option")
    chosen_value = None
    chosen_text  = None
    for o in options:
        t = (o.text or "").strip()
        if port_target_text.upper() == t.upper() or port_target_text.upper() in t.upper():
            chosen_value = o.get_attribute("value")
            chosen_text  = t
            break
    if not chosen_value:
        log(f"   ❌ Không tìm thấy port '{port_target_text}' trong dropdown #{port_select_id} (country={country}). Options: {[o.text for o in options]}")
        return False
    Select(port_el).select_by_value(chosen_value)
    log(f"   ✅ Chọn {port_select_id}: '{chosen_text}' (country={country})")
    time.sleep(0.5)
    return True

# ===================================================================================
# SET DATES (input readonly → set via JS)
# ===================================================================================
def set_date(input_id, dt):
    el = driver.find_element(By.ID, input_id)
    driver.execute_script("""
        const el = arguments[0]; const v = arguments[1];
        el.removeAttribute('readonly');
        el.value = v;
        el.setAttribute('readonly', 'readonly');
    """, el, fmt_whl_date(dt))

def set_search_dates(start_dt, end_dt, valid_dt):
    """dept_date_from = start (ETD min), arr_date_to = valid + 60 ngày (cover transit)."""
    set_date("dept_date_from", start_dt)
    arr_to = valid_dt + timedelta(days=60)
    set_date("arr_date_to", arr_to)
    log(f"   📅 dept_date_from={fmt_whl_date(start_dt)}, arr_date_to={fmt_whl_date(arr_to)}")

# ===================================================================================
# CLICK SEARCH → MAIN result → ALL SERVICE DETAIL
# ===================================================================================
def submit_search():
    btn = driver.find_element(By.ID, "subtn")
    safe_click(btn)
    log("   🔍 Đã bấm Search")

def open_all_service_detail():
    """Sau khi click Search, web trỏ tới SkdByPortMain. Bấm 'ALL SERVICE DETAIL' để có ETD cụ thể."""
    try:
        WAIT.until(EC.presence_of_element_located((By.ID, "skd_p2p_detail")))
    except TimeoutException:
        log("   ⚠️ Không thấy link 'ALL SERVICE DETAIL' (có thể không có chuyến)")
        return False
    link = driver.find_element(By.ID, "skd_p2p_detail")
    safe_click(link)
    time.sleep(1.0)
    # Đợi bảng detail load
    end = time.time() + 30
    while time.time() < end:
        rows = driver.find_elements(By.XPATH, "//table//tr")
        if any("20" in (r.text or "") and "/" in (r.text or "") for r in rows):
            return True
        time.sleep(0.5)
    return True  # vẫn cố parse

# ===================================================================================
# PARSE DETAIL TABLE
# ===================================================================================
def parse_detail_rows():
    """
    Bảng detail có mỗi chuyến trên 1 row. Mỗi row dạng innerText:
        CAT LAI PORT HOCHIMINH  2026/06/04
        WAN HAI 291
        N061
                SHEKOU  2026/06/13
        WAN HAI 291
        N061
        10 days
        CHINA-THAILAND-KAMPUCHEA
        Direct
    Trả về list dict { etd_dt, eta_dt, tt_days, pol, pod, vessel, voyage, service, transfer }
    """
    out = []
    # Lấy bảng kết quả: rows chứa pattern \d{4}/\d{2}/\d{2}
    rows = driver.find_elements(By.XPATH, "//table//tr")
    for r in rows:
        try:
            txt = (r.text or "").strip()
            if not txt or not re.search(r"\d{4}/\d{2}/\d{2}", txt):
                continue
            # Mỗi row có nhiều cells → lấy theo cell
            cells = r.find_elements(By.XPATH, "./td")
            if len(cells) < 4:
                continue
            # cell[0] = "POL\n YYYY/MM/DD\n VESSEL\n VOY"
            # cell[1] = "POD\n YYYY/MM/DD\n VESSEL\n VOY"  (hoặc "Feeder")
            # cell[2] = "X days"
            # cell[3] = "Service"
            # cell[4] = "Direct" / "Transfer"
            def cell_text(i):
                return (cells[i].text or "").strip() if i < len(cells) else ""

            if len(cells) >= 7:
                pol = cell_text(0)
                dep_lines = [ln.strip() for ln in cell_text(1).split("\n") if ln.strip()]
                pod = cell_text(2)
                arr_lines = [ln.strip() for ln in cell_text(3).split("\n") if ln.strip()]
                etd_dt = None
                eta_dt = None
                d_vessel = ""
                d_voy = ""
                if dep_lines and re.match(r"\d{4}/\d{2}/\d{2}", dep_lines[0]):
                    etd_dt = datetime.strptime(dep_lines[0], "%Y/%m/%d")
                    d_vessel = dep_lines[1] if len(dep_lines) > 1 else ""
                    d_voy = dep_lines[2] if len(dep_lines) > 2 else ""
                if arr_lines and re.match(r"\d{4}/\d{2}/\d{2}", arr_lines[0]):
                    eta_dt = datetime.strptime(arr_lines[0], "%Y/%m/%d")

                m_tt = re.search(r"(\d+)\s*days?", cell_text(4), re.I)
                tt_days = int(m_tt.group(1)) if m_tt else 0
                service = cell_text(5)
                transfer = cell_text(6).upper()
                if transfer not in ("DIRECT", "TRANSFER"):
                    full = "\n".join(cell_text(i) for i in range(len(cells)))
                    transfer = "TRANSFER" if re.search(r"\bTransfer\b", full, re.I) else "DIRECT"

                if not etd_dt or not pol:
                    continue

                out.append({
                    "etd_dt":   etd_dt,
                    "eta_dt":   eta_dt,
                    "tt_days":  tt_days,
                    "pol":      pol,
                    "pod":      pod,
                    "vessel":   d_vessel or "TBA",
                    "voyage":   d_voy or "",
                    "service":  service,
                    "transfer": transfer,
                })
                continue

            c0 = cell_text(0); c1 = cell_text(1); c2 = cell_text(2); c3 = cell_text(3)
            c4 = cell_text(4) if len(cells) > 4 else ""

            # Parse cell 0 (departure)
            pol = ""
            etd_dt = None
            d_vessel = ""
            d_voy = ""
            for ln in c0.split("\n"):
                ln = ln.strip()
                if not ln:
                    continue
                m = re.match(r"(\d{4}/\d{2}/\d{2})", ln)
                if m:
                    etd_dt = datetime.strptime(m.group(1), "%Y/%m/%d")
                    # POL có thể đứng trước date trong cùng line
                    pre = ln[:m.start()].strip()
                    if pre and not pol:
                        pol = pre
                elif not pol:
                    pol = ln
                elif not d_vessel:
                    d_vessel = ln
                elif not d_voy:
                    d_voy = ln

            # Parse cell 1 (arrival)
            pod = ""
            eta_dt = None
            for ln in c1.split("\n"):
                ln = ln.strip()
                if not ln:
                    continue
                m = re.match(r"(\d{4}/\d{2}/\d{2})", ln)
                if m:
                    eta_dt = datetime.strptime(m.group(1), "%Y/%m/%d")
                    pre = ln[:m.start()].strip()
                    if pre and not pod:
                        pod = pre
                elif not pod:
                    pod = ln

            # Transit time
            m_tt = re.search(r"(\d+)\s*days?", c2, re.I)
            tt_days = int(m_tt.group(1)) if m_tt else 0

            service = c3.strip()
            transfer = (c4 or "").strip().upper()
            if transfer not in ("DIRECT", "TRANSFER"):
                # Đôi khi cell layout khác → tìm trong các cell
                full = "\n".join([c0, c1, c2, c3, c4])
                if re.search(r"\bTransfer\b", full, re.I):
                    transfer = "TRANSFER"
                else:
                    transfer = "DIRECT"

            if not etd_dt or not pol:
                continue

            out.append({
                "etd_dt":   etd_dt,
                "eta_dt":   eta_dt,
                "tt_days":  tt_days,
                "pol":      pol,
                "pod":      pod,
                "vessel":   d_vessel or "TBA",
                "voyage":   d_voy or "",
                "service":  service,
                "transfer": transfer,
            })
        except Exception as e:
            log(f"   ⚠️ Lỗi parse row: {e}")
            continue
    log(f"   📦 Parsed {len(out)} chuyến chi tiết")
    return out

# ===================================================================================
# ETD RULES (giống YM)
# ===================================================================================
def apply_etd_rules(entries, valid_dt):
    min_etd, _ = compute_valid_window(valid_dt)
    future = [
        e for e in entries
        if e["etd_dt"] >= min_etd and etd_within_max(e["etd_dt"]) and (valid_dt is None or e["etd_dt"] <= valid_dt)
    ]
    if not future:
        return []
    future.sort(key=lambda e: e["etd_dt"])
    selected = [future[0]]
    for e in future[1:]:
        if len(selected) >= 3:
            break
        if (e["etd_dt"] - selected[-1]["etd_dt"]).days < 1:
            continue
        selected.append(e)
    return selected

def select_priority_sailings(entries, valid_dt, prefer_direct=False):
    """
    Với các tuyến HCM đi Chennai/Mundra/Nhava Sheva/Vizag:
      - Chỉ xét các ETD thực sự nằm trong vùng ngày hợp lệ.
      - Nếu có ít nhất một chuyến DIRECT hợp lệ, chỉ lấy DIRECT.
      - Chỉ khi không có DIRECT hợp lệ mới lấy chuyến TRANSFER.
    Các tuyến còn lại giữ nguyên logic chọn ETD hiện tại.
    """
    if not prefer_direct:
        return apply_etd_rules(entries, valid_dt), ""

    direct_entries = [
        e for e in entries
        if (e.get("transfer") or "").strip().upper() == "DIRECT"
    ]
    direct_selected = apply_etd_rules(direct_entries, valid_dt)
    if direct_selected:
        return direct_selected, "DIRECT"

    transfer_entries = [
        e for e in entries
        if (e.get("transfer") or "").strip().upper() != "DIRECT"
    ]
    transfer_selected = apply_etd_rules(transfer_entries, valid_dt)
    if transfer_selected:
        return transfer_selected, "TRANSFER"

    return [], ""

# ===================================================================================
# FORMATTERS (giống YM)
# ===================================================================================
def format_etd_text(entries):
    if not entries:
        return ""
    fmt = [(e["etd_dt"].day, e["etd_dt"].strftime("%b")) for e in entries]
    if len(fmt) == 1:
        return f"{fmt[0][0]}-{fmt[0][1]}"
    if len(fmt) == 2:
        return f"{fmt[0][0]}-{fmt[0][1]} & {fmt[1][0]}-{fmt[1][1]}"
    months = [f[1] for f in fmt]
    if months[0] == months[1] == months[2]:
        return f"{fmt[0][0]}, {fmt[1][0]}, {fmt[2][0]}-{fmt[2][1]}"
    if months[0] == months[1]:
        return f"{fmt[0][0]}, {fmt[1][0]}-{fmt[1][1]}, {fmt[2][0]}-{fmt[2][1]}"
    return f"{fmt[0][0]}-{fmt[0][1]}, {fmt[1][0]}-{fmt[1][1]}, {fmt[2][0]}-{fmt[2][1]}"

def format_tt_text(entries):
    if not entries:
        return ""
    tts = [int(e["tt_days"]) for e in entries if e.get("tt_days")]
    if not tts:
        return ""
    if len(set(tts)) == 1:
        return str(tts[0])
    return f"{min(tts)}-{max(tts)}"

def format_vessel_block(entries):
    lines = []
    ts_seen = []
    for e in entries:
        vn = e["vessel"] or "TBA"
        vc = e["voyage"] or ""
        v_str = f"{vn}" + (f" ({vc})" if vc else "")
        # Transfer/Direct
        ts_label = "TRANSIT" if e["transfer"] == "TRANSFER" else "DIRECT"
        lines.append(
            f"{v_str} / ETD: {e['etd_dt'].day}-{e['etd_dt'].strftime('%b')}"
            f" / Transit time: {e['tt_days']} Days / Transshipment: {ts_label}"
        )
        if ts_label not in ts_seen:
            ts_seen.append(ts_label)
    return "\n".join(lines), " or\n".join(ts_seen) if ts_seen else "DIRECT"

# ===================================================================================
# FILTER theo POL (chỉ giữ rows match đúng port đã chọn)
# ===================================================================================
def filter_by_pol(entries, pol_target):
    p = (pol_target or "").strip().upper()

    # WHL groups Ho Chi Minh schedules under more than one physical terminal.
    # A search made with CAT LAI can therefore return valid sailings departing
    # from CAI MEP as well (including the direct India services).  Treat both
    # terminals as the same HCM POL group instead of discarding CAI MEP rows.
    hcm_pol_names = ("CAT LAI", "CAI MEP")
    target_is_hcm = any(name in p for name in hcm_pol_names)

    matched = []
    for entry in entries:
        row_pol = (entry.get("pol") or "").strip().upper()
        if target_is_hcm and any(name in row_pol for name in hcm_pol_names):
            matched.append(entry)
        elif p in row_pol or row_pol in p:
            matched.append(entry)
    return matched

# ===================================================================================
# WRITE EXCEL
# ===================================================================================
def write_excel_row(row_i, etd_text, tt_text, vessel_text, ts_text, error=None):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        if ws.cell(row=row_i, column=6).value is None:
            ws.cell(row=row_i, column=6).value = "-"
        if error:
            ws.cell(row=row_i, column=9).value = error
        else:
            ws.cell(row=row_i, column=9).value  = etd_text
            ws.cell(row=row_i, column=10).value = tt_text
            ws.cell(row=row_i, column=15).value = vessel_text
            ws.cell(row=row_i, column=16).value = ts_text
            try:
                wrap = Alignment(wrap_text=True, vertical="top")
                for col in [15, 16]:
                    ws.cell(row=row_i, column=col).alignment = wrap
            except Exception:
                pass
        wb.save(EXCEL_PATH)
        log(f"   💾 Ghi Excel row {row_i}")
    except PermissionError:
        log("   ❌ Lỗi ghi Excel: TẮT FILE EXCEL ĐI rồi chạy lại!")
    except Exception as e:
        log(f"   ❌ write_excel_row: {e}")

# ===================================================================================
# SEARCH ONE ROUTE
# ===================================================================================
def search_one(pol_excel, pod_excel, valid_dt, country_excel=""):
    pol_target = alias_port(pol_excel)
    pod_target = alias_port(pod_excel)
    country_hint = (country_excel or "").strip().upper()
    log(f"   🔄 Search: '{pol_excel}' → '{pol_target}'  |  '{pod_excel}' → '{pod_target}'")

    start_dt, end_dt = compute_valid_window(valid_dt)
    if not start_dt or not end_dt:
        return {"error": "VALID rỗng/không parse được"}
    if end_dt < start_dt:
        return {"error": f"VALID trước ETD tối thiểu date +{DATE_OFFSET_DAYS}"}

    try:
        open_search_page()
        handle_captcha_if_needed("before selecting route")
        if not driver.find_elements(By.CSS_SELECTOR, "select#from_nation"):
            return {"error": "Form WHL không load (có thể captcha chưa qua)"}
        if not select_country_and_port("from_nation", "pol", pol_target):
            return {"error": "POL không có trong dropdown WHL"}
        if not select_country_and_port("to_nation", "pod", pod_target, country_hint=country_hint or None):
            return {"error": "POD không có trong dropdown WHL"}
        set_search_dates(start_dt, end_dt, valid_dt)
        handle_captcha_if_needed("before search schedule")
        submit_search()
        time.sleep(3.0)
        # Check captcha post-search
        handle_captcha_if_needed("after search schedule")
        if not open_all_service_detail():
            return {"error": "Không có ALL SERVICE DETAIL"}
        handle_captcha_if_needed("after all service detail")
        entries = parse_detail_rows()
        entries = filter_by_pol(entries, pol_target)
        if not entries:
            return {"error": "Không có chuyến phù hợp với POL"}
        prefer_direct = whl_direct_priority_route(pol_excel, pod_excel)
        selected, selected_mode = select_priority_sailings(
            entries,
            valid_dt,
            prefer_direct=prefer_direct,
        )
        if prefer_direct:
            if selected_mode == "DIRECT":
                log(
                    f"   [WHL DIRECT PRIORITY] {pol_excel} -> {pod_excel}: "
                    f"có DIRECT trong valid, chỉ lấy {len(selected)} chuyến DIRECT"
                )
            elif selected_mode == "TRANSFER":
                log(
                    f"   [WHL DIRECT PRIORITY] {pol_excel} -> {pod_excel}: "
                    f"không có DIRECT trong valid, fallback {len(selected)} chuyến TRANSFER"
                )
        if not selected:
            return {"error": "Không có ETD ≤ valid"}
        etd_text   = format_etd_text(selected)
        tt_text    = format_tt_text(selected)
        v_text, ts = format_vessel_block(selected)
        return {
            "etd_text":     etd_text,
            "tt_text":      tt_text,
            "vessel_text":  v_text,
            "ts_text":      ts,
        }
    except Exception as e:
        log(f"   ❌ search_one error: {e}")
        return {"error": f"Exception: {e}"}

# ===================================================================================
# MAIN
# ===================================================================================
def _whl_page_has_real_search_form():
    try:
        required = ["select#from_nation", "select#pol", "select#to_nation", "select#pod"]
        for css in required:
            els = driver.find_elements(By.CSS_SELECTOR, css)
            if not els:
                return False
            if not any(e.is_displayed() for e in els):
                return False
        return True
    except Exception:
        return False

def _whl_security_signals():
    """Read URL/title/body/outerHTML including shadow DOM to detect WHL security gate."""
    try:
        url = (driver.current_url or "").lower()
    except Exception:
        url = ""
    try:
        title = (driver.title or "").lower()
    except Exception:
        title = ""
    try:
        body_text = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        body_text = ""
    try:
        page_blob = driver.execute_script("""
            const out = [];
            function add(v) { if (v) out.push(String(v).toLowerCase()); }
            function walk(root) {
                if (!root) return;
                add(root.innerText);
                add(root.textContent);
                if (root.querySelectorAll) {
                    root.querySelectorAll('*').forEach(el => {
                        add(el.id);
                        add(el.className);
                        add(el.getAttribute && el.getAttribute('src'));
                        add(el.getAttribute && el.getAttribute('href'));
                        add(el.getAttribute && el.getAttribute('alt'));
                        add(el.getAttribute && el.getAttribute('aria-label'));
                        add(el.getAttribute && el.getAttribute('title'));
                        add(el.getAttribute && el.getAttribute('placeholder'));
                        if (el.shadowRoot) walk(el.shadowRoot);
                    });
                }
            }
            walk(document.documentElement);
            return out.join('\\n').slice(0, 300000);
        """) or ""
    except Exception:
        page_blob = ""

    blob = "\n".join([url, title, body_text, page_blob]).lower()
    terms = [
        "captcha",
        "hcaptcha",
        "recaptcha",
        "turnstile",
        "i am human",
        "i'm human",
        "not a robot",
        "human verification",
        "security check",
        "additional security check",
        "your request is currently unavailable",
        "currently unable to respond",
        "unable to respond to your request",
        "powered by wan hai",
        "incapsula",
        "imperva",
        "_incapsula_resource",
        "visid_incap",
        "___utmvc",
        "challenge",
        "verify you are human",
        "tôi là con người",
        "toi la con nguoi",
        "您的請求目前無法回應",
        "目前無法回應",
        "驗證",
        "验证",
    ]
    hits = [t for t in terms if t in blob]
    return hits

def whl_captcha_present():
    hits = _whl_security_signals()
    if hits:
        log(f"   [WHL CAPTCHA DEBUG] Security signals: {hits[:5]}")
        return True
    return False

def handle_captcha_if_needed(stage=""):
    """Hard stop whenever WHL security/captcha is visible."""
    import winsound

    if not whl_captcha_present():
        return False

    stage_msg = f" ({stage})" if stage else ""
    log(f"   [WHL SILENT] PHAT HIEN CAPTCHA/SECURITY{stage_msg}. Bot se dung han, khong search tiep.")

    marker_created = False
    try:
        try:
            with open(WHL_MANUAL_CAPTCHA_WAIT_PATH, "w", encoding="utf-8") as marker:
                marker.write(str(os.getpid()))
            marker_created = True
        except OSError as exc:
            log(f"   [WHL SILENT] Khong tao duoc co terminal: {type(exc).__name__}")

        deadline = time.monotonic() + WHL_AUTO_CAPTCHA_WAIT
        while whl_captcha_present():
            try:
                winsound.Beep(1500, 500)
            except Exception:
                pass
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                raise TimeoutError(
                    f"WHL captcha/security chưa được giải sau {WHL_AUTO_CAPTCHA_WAIT} giây"
                )
            log(
                "   [WHL SILENT] Hay giai captcha tren Edge WHL. Xong thi quay lai terminal "
                f"va nhan ENTER (con toi da {int(remaining)}s)."
            )
            resumed = wait_for_terminal_enter(
                input,
                remaining,
                "   [WHL SILENT] Nhan ENTER sau khi WHL da pass captcha... ",
            )
            if not resumed:
                raise TimeoutError(
                    f"WHL captcha/security chưa được giải sau {WHL_AUTO_CAPTCHA_WAIT} giây"
                )

            time.sleep(1.0)
            if whl_captcha_present():
                log("   [WHL SILENT] Van con captcha/security page -> tiep tuc dung, chua search lich tau.")
    finally:
        if marker_created:
            try:
                os.remove(WHL_MANUAL_CAPTCHA_WAIT_PATH)
            except OSError:
                pass

    log("   [WHL SILENT] Captcha/security da bien mat -> tiep tuc WHL.")
    return True

def open_search_page():
    ensure_live_window()
    driver.get(URL_WHL)
    time.sleep(1.0)
    handle_captcha_if_needed("open search page")

    try:
        WebDriverWait(driver, 20).until(lambda d: whl_captcha_present() or _whl_page_has_real_search_form())
    except TimeoutException:
        handle_captcha_if_needed("wait search form timeout")
        raise

    handle_captcha_if_needed("after wait search form")
    if not _whl_page_has_real_search_form():
        raise TimeoutException("WHL search form not ready after captcha/security check")

def main():
    if not os.path.exists(EXCEL_PATH):
        log(f"❌ Không tìm thấy file Excel: {EXCEL_PATH}")
        sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    target_single_row = None
    if SINGLE_ROW:
        try:
            target_single_row = int(SINGLE_ROW)
            log(f"[SINGLE_ROW] Chi chay dong {target_single_row} theo lenh tu main.py")
        except ValueError:
            target_single_row = None
    target_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if target_single_row is not None and i != target_single_row:
            continue
        country = str(row[1] or "").strip()
        pol     = str(row[2] or "").strip()
        pod     = str(row[3] or "").strip()
        carrier = str(row[4] or "").strip().upper()
        valid_raw = row[10] if len(row) > 10 else None
        if not pol or not pod or carrier != CARRIER_TARGET:
            continue
        if FILTER_POL and pol.upper() != FILTER_POL:
            continue
        if FILTER_POD and pod.upper() != FILTER_POD:
            continue
        target_rows.append((i, country, pol, pod, valid_raw))
    wb.close()

    log(f"📋 Có {len(target_rows)} dòng cần check (carrier = {CARRIER_TARGET})")
    if not target_rows:
        return

    for idx, (row_i, country, pol, pod, valid_raw) in enumerate(target_rows, start=1):
        log(f"\n========== [{idx}/{len(target_rows)}] DÒNG {row_i}: {pol} → {pod} | VALID={valid_raw}")
        valid_dt = parse_valid_date(valid_raw)
        if not valid_dt:
            log("⚠️ Không có ngày Valid trong Excel -> bỏ qua row, không gọi API.")
            continue
        result = search_one(pol, pod, valid_dt, country_excel=country)
        if result.get("error"):
            write_excel_row(row_i, "", "", "", "", error=result["error"])
        else:
            write_excel_row(
                row_i,
                result["etd_text"],
                result["tt_text"],
                result["vessel_text"],
                result["ts_text"],
            )
        time.sleep(1.0)

    log("\n✅ HOÀN TẤT bot WHL")

if __name__ == "__main__":
    main()
