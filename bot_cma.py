from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import openpyxl
import os
import time
import re
import random
import subprocess
import socket
import sys
import io
import calendar
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from cma_logic import (
    classify_cma_card_texts,
    cma_date_input_matches,
    dedupe_cma_card_summaries,
    parse_cma_comparable_price,
)
from remark_rules import charge_amount_to_usd, get_manifest_code, is_china_destination

# ===================================================================================
# ── Timestamp print ──
# ===================================================================================
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
_orig_print = print
def print(*args, **kwargs):
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    _orig_print(f"[{ts}]", *args, **kwargs)
# ===================================================================================
# --- CONFIG ---
# ===================================================================================
def sleep_human(*args):
    # Phớt lờ các tham số thời gian cũ, ép chạy random siêu nhanh từ 0.01s đến 0.04s
    time.sleep(random.uniform(0.01, 0.04))
current_folder     = os.getcwd()
driver_path        = os.path.join(current_folder, "msedgedriver.exe")  # ← THÊM
excel_path         = os.environ.get("EXCEL_PATH", os.path.join(current_folder, "input_gia.xlsx"))
FILTER_POL         = os.environ.get("FILTER_POL", "").strip().upper()
FILTER_POD         = os.environ.get("FILTER_POD", "").strip().upper()
DATE_OFFSET_DAYS = parse_date_offset_days()
try:
    CMA_AVAILABLE_WAIT_SECONDS = max(0.3, float(os.environ.get("CMA_AVAILABLE_WAIT_SECONDS", "0.8")))
except ValueError:
    CMA_AVAILABLE_WAIT_SECONDS = 0.8
try:
    CMA_PRICE_WAIT_SECONDS = max(3.0, float(os.environ.get("CMA_PRICE_WAIT_SECONDS", "8")))
except ValueError:
    CMA_PRICE_WAIT_SECONDS = 8.0
try:
    CMA_PARTIAL_PRICE_STABLE_SECONDS = max(
        0.5,
        float(os.environ.get("CMA_PARTIAL_PRICE_STABLE_SECONDS", "1.5")),
    )
except ValueError:
    CMA_PARTIAL_PRICE_STABLE_SECONDS = 1.5
try:
    CMA_MODIFY_BUTTON_WAIT_SECONDS = max(
        2.0,
        float(os.environ.get("CMA_MODIFY_BUTTON_WAIT_SECONDS", "5")),
    )
    CMA_MODIFY_FORM_WAIT_SECONDS = max(
        3.0,
        float(os.environ.get("CMA_MODIFY_FORM_WAIT_SECONDS", "8")),
    )
except ValueError:
    CMA_MODIFY_BUTTON_WAIT_SECONDS = 5.0
    CMA_MODIFY_FORM_WAIT_SECONDS = 8.0
try:
    CMA_ROW_SLEEP_MIN = max(0.0, float(os.environ.get("CMA_ROW_SLEEP_MIN", "0.8")))
    CMA_ROW_SLEEP_MAX = max(CMA_ROW_SLEEP_MIN, float(os.environ.get("CMA_ROW_SLEEP_MAX", "1.2")))
except ValueError:
    CMA_ROW_SLEEP_MIN, CMA_ROW_SLEEP_MAX = 0.8, 1.2
try:
    CMA_LOGIN_WAIT_SECONDS = max(20, int(os.environ.get("CMA_LOGIN_WAIT_SECONDS", "60")))
except ValueError:
    CMA_LOGIN_WAIT_SECONDS = 60

def _excel_formula_from_parts(parts):
    tokens = []
    for raw in parts:
        try:
            amount = float(raw)
        except (TypeError, ValueError):
            continue
        if abs(amount) < 1e-9:
            continue
        sign = "-" if amount < 0 else "+"
        amount = abs(amount)
        if abs(amount - round(amount)) < 1e-9:
            text = str(int(round(amount)))
        else:
            text = f"{amount:.2f}".rstrip("0").rstrip(".")
        tokens.append((sign if tokens else ("-" if sign == "-" else "")) + text)
    return "=" + "".join(tokens) if tokens else None

EDGE_EXE           = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
EDGE_DEBUG_PORT    = 9524
EDGE_USER_DATA_DIR = r"C:\edge_cma"  # ← profile riêng cho CMA
CMA_URL            = "https://www.cma-cgm.com/ebusiness/pricing/instant-quoting"
LOGIN_URL          = "https://www.cma-cgm.com/myCmaCgm/login"
LOGIN_EMAIL        = os.environ.get("CMA_EMAIL", "celine@pio-logistics.vn")
LOGIN_PASSWORD     = os.environ.get("CMA_PASSWORD", "Xvnt@686868")

# ===================================================================================
# --- AUTO LAUNCH EDGE ---
# ===================================================================================
def is_port_open(port, host="127.0.0.1"):
    try:
        with socket.create_connection((host, port), timeout=1):
            return True
    except (OSError, ConnectionRefusedError):
        return False

def launch_edge_if_needed():
    if is_port_open(EDGE_DEBUG_PORT):
        print(f"[OK] Edge đã mở (port {EDGE_DEBUG_PORT}).")
        return True
    print("[INFO] Khởi động Edge...")
    try:
        subprocess.Popen([
            EDGE_EXE,
            f"--remote-debugging-port={EDGE_DEBUG_PORT}",
            f"--user-data-dir={EDGE_USER_DATA_DIR}",
            "--disable-background-timer-throttling",
            "--disable-renderer-backgrounding",
            "--disable-backgrounding-occluded-windows",
        ])
        for i in range(30):
            time.sleep(0.5)
            if is_port_open(EDGE_DEBUG_PORT):
                print(f"[OK] Edge sẵn sau {(i+1)*0.5:.1f}s.")
                return True
        print("[ERROR] Edge không khởi động được.")
        return False
    except FileNotFoundError:
        print(f"[ERROR] Không thấy Edge: {EDGE_EXE}")
        return False

def init_browser():
    opts = Options()
    opts.use_chromium = True
    opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{EDGE_DEBUG_PORT}")
    try:
        drv = webdriver.Edge(service=Service(driver_path), options=opts)
        drv.set_page_load_timeout(30)
        try:
            drv.maximize_window()
        except Exception:
            try:
                info = drv.execute_cdp_cmd("Browser.getWindowForTarget", {})
                drv.execute_cdp_cmd("Browser.setWindowBounds", {
                    "windowId": info["windowId"],
                    "bounds": {"windowState": "maximized"}
                })
            except Exception:
                pass
        print("[OK] Kết nối Selenium thành công.")
        return drv
    except Exception as e:
        print(f"[ERROR] Không kết nối được Selenium: {e}")
        return None

# ===================================================================================
# --- AUTO LOGIN CMA ---
# ===================================================================================
def _cma_current_url(drv):
    try:
        return (drv.current_url or "").strip()
    except Exception:
        return ""

def _cma_is_login_url(url):
    lower = (url or "").lower()
    return (
        "auth.cma-cgm.com" in lower
        or "login.cma-cgm.com" in lower
        or "mycmacgm/login" in lower
        or "/mycmacgm/login" in lower
        or ("/login" in lower and "cma-cgm.com" in lower)
    )

def _cma_is_spoton_url(url):
    lower = (url or "").lower()
    # Auth URL often contains redirect_uri=.../ebusiness/pricing..., so do not
    # classify it as SpotOn just because the query string mentions pricing.
    if _cma_is_login_url(lower):
        return False
    return "cma-cgm.com/ebusiness/pricing" in lower

def _wait_cma_landing(drv, timeout=10):
    try:
        WebDriverWait(drv, timeout, poll_frequency=0.2).until(
            lambda d: _cma_is_login_url(_cma_current_url(d)) or _cma_is_spoton_url(_cma_current_url(d))
        )
    except Exception:
        pass
    return _cma_current_url(drv)

def _cma_visible(elem):
    try:
        return elem.is_displayed()
    except Exception:
        return False

def _cma_find_visible(drv, selectors, by=By.XPATH, timeout=10, enabled=False):
    items = selectors if isinstance(selectors, (list, tuple)) else [selectors]
    end = time.time() + timeout
    while time.time() < end:
        for sel in items:
            try:
                for elem in drv.find_elements(by, sel):
                    try:
                        if not elem.is_displayed():
                            continue
                        if enabled and not elem.is_enabled():
                            continue
                        return elem
                    except Exception:
                        continue
            except Exception:
                continue
        time.sleep(0.2)
    return None

def _cma_set_input_value(drv, elem, value):
    try:
        drv.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    except Exception:
        pass
    try:
        elem.click()
        time.sleep(0.1)
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.DELETE)
        elem.send_keys(value)
    except Exception:
        pass
    try:
        drv.execute_script("""
            const input = arguments[0];
            const value = arguments[1];
            input.removeAttribute('readonly');
            const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(input, value);
            input.dispatchEvent(new Event('input', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
            input.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: value.slice(-1) || 'a' }));
            input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: value.slice(-1) || 'a' }));
            input.dispatchEvent(new Event('blur', { bubbles: true }));
        """, elem, value)
    except Exception:
        pass
    try:
        return (elem.get_attribute("value") or "").strip() == str(value).strip()
    except Exception:
        return False

def _cma_click_elem(drv, elem):
    try:
        drv.execute_script("arguments[0].scrollIntoView({block:'center'});", elem)
    except Exception:
        pass
    time.sleep(0.15)
    try:
        elem.click()
        return True
    except Exception:
        pass
    try:
        drv.execute_script("arguments[0].click();", elem)
        return True
    except Exception:
        return False

def _cma_click_login_button(drv, timeout=12):
    login_button_xpaths = [
        "//button[not(@disabled) and (@type='submit' or @id='login-submit' or @id='loginButton' or @name='login')]",
        "//button[not(@disabled) and contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
        "//button[not(@disabled) and contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'login')]",
        "//button[not(@disabled) and contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
        "//button[not(@disabled) and contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'continue')]",
        "//button[not(@disabled) and contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'next')]",
        "//input[not(@disabled) and (@type='submit' or @type='button') and contains(translate(@value,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log')]",
        "//input[not(@disabled) and (@type='submit' or @type='button') and contains(translate(@value,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign')]",
        "//a[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'log in')]",
        "//a[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'sign in')]",
    ]
    btn = _cma_find_visible(drv, login_button_xpaths, by=By.XPATH, timeout=timeout, enabled=True)
    if btn:
        text = ""
        try:
            text = (btn.text or btn.get_attribute("value") or btn.get_attribute("aria-label") or "").strip()
        except Exception:
            pass
        print(f"[INFO] CMA login: bấm nút '{text or 'submit'}'")
        return _cma_click_elem(drv, btn)
    return False

def _cma_manual_challenge_present(drv):
    try:
        body = (drv.find_element(By.TAG_NAME, "body").text or "").lower()
    except Exception:
        body = ""
    return any(x in body for x in [
        "captcha",
        "verification code",
        "verify your identity",
        "one-time",
        "one time",
        "mfa",
        "authenticator",
    ])


CMA_PORT_INPUT_XPATH = "//input[@placeholder='Name / Code / Port']"


def _cma_get_quote_port_inputs(drv, visible_only=True):
    """Return the ordered POL/POD inputs when the SpotOn quote form is rendered."""
    try:
        inputs = drv.find_elements(By.XPATH, CMA_PORT_INPUT_XPATH)
    except Exception:
        return []
    if not visible_only:
        return inputs
    visible = []
    for elem in inputs:
        try:
            if elem.is_displayed():
                visible.append(elem)
        except Exception:
            continue
    return visible


def _cma_wait_quote_form(drv, timeout=30):
    """Wait for both POL and POD inputs, not merely a pricing-looking URL."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        if _cma_is_login_url(_cma_current_url(drv)):
            return []
        inputs = _cma_get_quote_port_inputs(drv, visible_only=True)
        if len(inputs) >= 2:
            return inputs
        time.sleep(0.25)
    return []


def _wait_cma_spoton_after_login(drv, timeout=None):
    timeout = timeout or CMA_LOGIN_WAIT_SECONDS
    end = time.time() + timeout
    last_log = time.time()
    while time.time() < end:
        cur_url = _cma_current_url(drv)
        if _cma_is_spoton_url(cur_url):
            inputs = _cma_get_quote_port_inputs(drv, visible_only=True)
            if len(inputs) >= 2:
                print("[OK] Đăng nhập thành công và form Spot-On đã sẵn sàng!")
                return True
        if _cma_manual_challenge_present(drv):
            print("[WARN] CMA đang yêu cầu CAPTCHA/MFA/verification; cần xử lý thủ công trên Edge.")
        if time.time() - last_log > 10:
            print("[INFO] CMA login: đang chờ web chuyển về Spot-On...")
            last_log = time.time()
        time.sleep(0.5)
    return False

def check_and_login(drv):
    """Truy cập Spot-On, nếu bị redirect sang trang auth thì đăng nhập."""
    print("[INFO] Truy cập trang Spot-On...")
    try:
        drv.get(CMA_URL)
    except TimeoutException:
        drv.execute_script("window.stop();")
    
    cur_url = _wait_cma_landing(drv, timeout=8)

    if _cma_is_login_url(cur_url):
        print("[INFO] Bị chuyển hướng sang trang đăng nhập, tiến hành login...")
        try:
            email_xpaths = [
                "//*[@id='login-email']",
                "//input[@name='pf.username']",
                "//input[@type='email']",
                "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'email')]",
                "//input[contains(translate(@aria-label,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'email')]",
                "//input[contains(translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'email')]",
                "//input[contains(translate(@id,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'email')]",
                "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'user')]",
                "//input[contains(translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'username')]",
            ]
            pwd_xpaths = [
                "//*[@id='login-password']",
                "//input[@name='pf.pass']",
                "//input[@type='password']",
                "//input[contains(translate(@name,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'password')]",
                "//input[contains(translate(@id,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'password')]",
            ]

            email_inp = _cma_find_visible(drv, email_xpaths, by=By.XPATH, timeout=12, enabled=True)
            if email_inp:
                if not _cma_set_input_value(drv, email_inp, LOGIN_EMAIL):
                    print("[WARN] CMA login: đã thử bơm email nhưng chưa verify được value, vẫn đi tiếp.")
                else:
                    print("[OK] CMA login: đã điền email.")
                # Một số màn CMA/Ping có bước Continue/Next trước khi hiện password.
                pwd_now = _cma_find_visible(drv, pwd_xpaths, by=By.XPATH, timeout=1.5, enabled=True)
                if not pwd_now:
                    if _cma_click_login_button(drv, timeout=4):
                        time.sleep(1.5)
            else:
                print("[INFO] CMA login: không thấy ô email, có thể đang ở bước password/session sẵn.")

            pwd_inp = _cma_find_visible(drv, pwd_xpaths, by=By.XPATH, timeout=15, enabled=True)
            if pwd_inp:
                if not _cma_set_input_value(drv, pwd_inp, LOGIN_PASSWORD):
                    print("[ERROR] CMA login: không điền được password.")
                    return False
                print("[OK] CMA login: đã điền password.")
            elif _cma_manual_challenge_present(drv):
                print("[WARN] CMA yêu cầu xác minh thủ công trước password.")
            else:
                print("[WARN] CMA login: chưa thấy ô password, thử bấm nút tiếp tục/login nếu có.")

            if not _cma_click_login_button(drv, timeout=12):
                print("[WARN] CMA login: không tìm được nút login rõ ràng, thử ENTER ở password/email.")
                target_inp = pwd_inp or email_inp
                if target_inp:
                    try:
                        target_inp.send_keys(Keys.ENTER)
                    except Exception:
                        pass

            # Chờ web trả về lại trang báo giá
            print("[INFO] Đang chờ web quay lại trang Spot-On...")
            if not _wait_cma_spoton_after_login(drv, timeout=CMA_LOGIN_WAIT_SECONDS):
                print("[WARN] Web không tự chuyển về Spot-On (hoặc bị lag). Tiến hành ép chuyển hướng...")
                drv.get(CMA_URL)
                cur_url = _wait_cma_landing(drv, timeout=15)
                if _cma_is_spoton_url(cur_url) and _cma_wait_quote_form(drv, timeout=30):
                    print("[OK] Đã ép trình duyệt về lại trang Spot-On và form đã sẵn sàng!")
                else:
                    print(f"[ERROR] Sau login vẫn chưa vào được Spot-On. URL hiện tại: {cur_url}")
                    return False

            return True

        except Exception as e:
            print(f"[ERROR] Quá trình nhập liệu login thất bại: {e}")
            print("[INFO] Thử ép chuyển hướng lại trang Spot-On lần cuối...")
            try:
                drv.get(CMA_URL)
                cur_url = _wait_cma_landing(drv, timeout=15)
                return _cma_is_spoton_url(cur_url)
            except:
                return False

    elif _cma_is_spoton_url(cur_url):
        if _cma_wait_quote_form(drv, timeout=30):
            print("[OK] Đã ở trang Spot-On và form quote sẵn sàng, không cần đăng nhập.")
            return True
        print("[WARN] URL Spot-On đúng nhưng form quote chưa render; nạp lại trang quote một lần...")
        try:
            drv.get(CMA_URL)
        except TimeoutException:
            try:
                drv.execute_script("window.stop();")
            except Exception:
                pass
        if _cma_wait_quote_form(drv, timeout=30):
            print("[OK] Form Spot-On đã sẵn sàng sau khi nạp lại.")
            return True
        print(f"[ERROR] Vẫn không thấy form quote. URL={_cma_current_url(drv)}")
        return False
    else:
        print(f"[WARN] URL hiện tại không xác định: {cur_url}")
        return True

def ensure_on_cma_tab(drv):
    print("[DEBUG] BẮT ĐẦU KHÓA MỤC TIÊU VÀO TAB CMA (BỎ QUA TAB RÁC)...")
    try:
        all_tabs = drv.window_handles
        print(f"[DEBUG] Tổng số tab đang mở: {len(all_tabs)} -> Danh sách ID: {all_tabs}")
    except Exception as e:
        print(f"[ERROR] [DEBUG] Không lấy được window_handles: {e}")
        return

    if not all_tabs:
        print("[WARN] [DEBUG] Không có tab nào cả!")
        return

    target = None
    # 1. Quét từng tab để tìm tab CMA
    for h in all_tabs:
        try:
            drv.switch_to.window(h)
            cur_url = drv.current_url
            print(f"[DEBUG] Đang xét tab ID [{h[-6:]}] | URL hiện tại: {cur_url}")
            if "cma-cgm.com" in cur_url:
                target = h
                print(f"   -> [DEBUG] TÌM THẤY tab CMA! Đặt tab [{h[-6:]}] làm Target.")
                break
        except Exception as e:
            print(f"[ERROR] [DEBUG] Lỗi khi đọc tab {h[-6:]}: {e}")

    # 2. Nếu không thấy trang CMA nào, lấy tab đầu tiên làm gốc
    if not target:
        target = all_tabs[0]
        print(f"[WARN] [DEBUG] Không thấy URL CMA nào! Đặt tab đầu tiên [{target[-6:]}] làm Target mặc định.")

    # 3. Trở về tab Target và phớt lờ hoàn toàn các tab khác (Bỏ lệnh close)
    try:
        print(f"[DEBUG] Đang khóa mục tiêu vào tab Target [{target[-6:]}]...")
        drv.switch_to.window(target)
        print(f"[DEBUG] HOÀN TẤT KHÓA MỤC TIÊU! Đang ở URL: {drv.current_url}")
    except Exception as e:
        print(f"[FATAL] [DEBUG] LỖI MẤT SESSION KHI TRỞ VỀ TARGET: {e}")
# ===================================================================================
# --- 1. KHỞI ĐỘNG ---
# ===================================================================================
from selenium.common.exceptions import TimeoutException

if not launch_edge_if_needed():
    print("[ERROR] Không khởi động được Edge. Dừng.")
    exit()

driver = init_browser()
if not driver:
    print("[ERROR] Không kết nối được Selenium. Dừng.")
    exit()

# Bùa tàng hình
try:
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
except:
    pass

ensure_on_cma_tab(driver)

if not check_and_login(driver):
    print("[ERROR] Đăng nhập thất bại. Dừng.")
    sys.exit(1)


print("[OK] CMA Bot sẵn sàng!")

# ===================================================================================
# --- 2. CÁC HÀM XỬ LÝ NHẬP LIỆU ---
# ===================================================================================
def check_and_kill_popup():
    print("   -> Kiểm tra Popup...")

    # 1. Popup cũ
    try:
        remind_btn = WebDriverWait(driver, 1.5).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Remind Me Later')]"))
        )
        print("      [CÓ POPUP CŨ] -> Bấm Remind Me Later!")
        driver.execute_script("arguments[0].click();", remind_btn)
        time.sleep(random.uniform(1.0, 1.5))
    except:
        pass

    # 2. Popup Concierge
    try:
        concierge_close_btn = driver.find_element(By.CSS_SELECTOR, "button.mec-me-popup__close")
        if concierge_close_btn.is_displayed():
            print("      [CÓ POPUP CONCIERGE] -> Đóng popup!")
            driver.execute_script("arguments[0].click();", concierge_close_btn)
            time.sleep(random.uniform(1.0, 1.5))
    except:
        pass

    # 3. Popup lạ
    try:
        generic_close_btn = driver.find_element(
            By.CSS_SELECTOR, "button[aria-label='Close'], button[aria-label='Close dialog']"
        )
        if generic_close_btn.is_displayed():
            print("      [CÓ POPUP LẠ] -> Đóng popup!")
            driver.execute_script("arguments[0].click();", generic_close_btn)
            time.sleep(random.uniform(1.0, 1.5))
    except:
        pass


def ensure_cma_quote_form_inputs(context="", force_navigate=False):
    """Return the two visible POL/POD inputs, recovering the quote page once."""
    label = f" ({context})" if context else ""

    def navigate_to_quote():
        print(f"   -> Nạp trang quote CMA{label}...")
        try:
            driver.get(CMA_URL)
        except TimeoutException:
            try:
                driver.execute_script("window.stop();")
            except Exception:
                pass

    if force_navigate or not _cma_is_spoton_url(_cma_current_url(driver)):
        if _cma_is_login_url(_cma_current_url(driver)):
            if not check_and_login(driver):
                return []
        else:
            navigate_to_quote()

    for attempt in range(2):
        check_and_kill_popup()
        inputs = _cma_wait_quote_form(driver, timeout=20 if attempt == 0 else 30)
        if len(inputs) >= 2:
            if attempt:
                print(f"   [OK] Form quote CMA đã phục hồi{label}.")
            return inputs

        try:
            title = driver.title
        except Exception:
            title = ""
        print(
            f"   [WARN] Chưa thấy đủ ô POL/POD{label}; "
            f"attempt {attempt + 1}/2, URL={_cma_current_url(driver)}, title={title!r}"
        )
        if attempt == 0:
            if _cma_is_login_url(_cma_current_url(driver)):
                if not check_and_login(driver):
                    break
            else:
                navigate_to_quote()

    print(f"   [ERROR] Form quote CMA không sẵn sàng{label}.")
    return []


def perform_modify_search_action():
    print("   -> Bấm Modify Search...")
    started = time.monotonic()
    try:
        clicked = WebDriverWait(
            driver,
            CMA_MODIFY_BUTTON_WAIT_SECONDS,
            poll_frequency=0.2,
        ).until(
            lambda d: d.execute_script("""
                const visible = el => {
                    if (!el) return false;
                    const s = getComputedStyle(el);
                    const r = el.getBoundingClientRect();
                    return s.display !== 'none' && s.visibility !== 'hidden' && r.width > 0 && r.height > 0;
                };
                const nodes = Array.from(document.querySelectorAll('button, a'));
                const target = nodes.find(el =>
                    visible(el) && /modify\\s*search/i.test((el.innerText || el.textContent || '').trim())
                );
                if (!target) return false;
                target.click();
                return true;
            """)
        )
        if not clicked:
            return False

        WebDriverWait(driver, CMA_MODIFY_FORM_WAIT_SECONDS, poll_frequency=0.2).until(
            EC.visibility_of_element_located(
                (By.XPATH, "//button[contains(., 'Get My Quote') or @id='SearchQuote']"))
        )
        check_and_kill_popup()
        print(f"      -> Modify form sẵn sàng sau {time.monotonic() - started:.1f}s")
        return True
    except Exception as exc:
        print(
            f"      [WARN] Modify Search chưa sẵn sàng sau "
            f"{time.monotonic() - started:.1f}s ({type(exc).__name__})"
        )
        return False
    


def is_cma_no_offer_page():
    """Detect the CMA SpotOn no-offer result page without waiting for card timeouts."""
    try:
        cur_url = (driver.current_url or "").lower()
        if "pricing/ignoreresult" in cur_url or "pricing/ignoresresult" in cur_url:
            return True
    except:
        pass

    try:
        page_text = driver.execute_script(
            "return document.body ? document.body.innerText : '';"
        ) or ""
    except:
        page_text = ""

    text = " ".join(page_text.lower().split())
    return (
        "not able to propose spoton offer" in text
        or ("we apologize" in text and "spoton" in text and "offer" in text)
    )


def wait_cma_result_state(timeout=15):
    """Return CARDS, NO_OFFER, NO_ROUTE, or TIMEOUT after Get Quote."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        if is_cma_no_offer_page():
            return "NO_OFFER"

        try:
            alerts = driver.find_elements(By.CSS_SELECTOR, "span.el-alert__title")
            for alert in alerts:
                if alert.is_displayed() and "SpotOn hasn't found possible route" in alert.text:
                    return "NO_ROUTE"
        except:
            pass

        try:
            if driver.find_elements(By.CSS_SELECTOR, "article.card-route-horizontal"):
                return "CARDS"
        except:
            pass

        time.sleep(0.25)
    return "TIMEOUT"


def get_cma_card_text_snapshot():
    """Read all visible CMA card text in one browser round-trip."""
    try:
        return driver.execute_script("""
            return Array.from(document.querySelectorAll('article.card-route-horizontal'))
                .filter(card => {
                    const style = getComputedStyle(card);
                    const rect = card.getBoundingClientRect();
                    return style.display !== 'none' && style.visibility !== 'hidden' && rect.height > 0;
                })
                .slice(0, 8)
                .map(card => card.innerText || card.textContent || '');
        """) or []
    except Exception:
        return []


def wait_for_cma_price_snapshot(timeout=None):
    """Wait for usable card data, accepting a stable partial snapshot early."""
    timeout = CMA_PRICE_WAIT_SECONDS if timeout is None else timeout
    deadline = time.monotonic() + timeout
    last_signature = None
    partial_since = None
    last_state = "EMPTY"

    while time.monotonic() < deadline:
        texts = get_cma_card_text_snapshot()
        state = classify_cma_card_texts(texts)
        last_state = state
        signature = tuple(" ".join(str(text or "").split()) for text in texts)

        if state == "READY":
            return state
        if state == "PARTIAL":
            if signature != last_signature:
                last_signature = signature
                partial_since = time.monotonic()
            elif partial_since is not None and (
                time.monotonic() - partial_since >= CMA_PARTIAL_PRICE_STABLE_SECONDS
            ):
                return state
        else:
            last_signature = signature
            partial_since = None
        time.sleep(0.2)
    return last_state


def collect_cma_card_summaries():
    """Collect summary fields for up to eight cards with one JS call."""
    try:
        return driver.execute_script("""
            return Array.from(document.querySelectorAll('article.card-route-horizontal'))
                .slice(0, 8)
                .map((card, index) => {
                    const dateNode = card.querySelector('span.date, .date');
                    const transitNode = card.querySelector('[class*="transit"]');
                    const tsNode = card.querySelector(
                        '[class*="transit"][class*="transshipment"], [class*="transit"][class*="direct"]'
                    );
                    return {
                        card_idx: index,
                        text: card.innerText || card.textContent || '',
                        date_text: dateNode ? (dateNode.innerText || dateNode.textContent || '') : '',
                        transit_text: transitNode ? (transitNode.innerText || transitNode.textContent || '') : '',
                        ts_text: tsNode ? (tsNode.innerText || tsNode.textContent || '') : '',
                        ts_class: tsNode ? String(tsNode.className || '') : ''
                    };
                });
        """) or []
    except Exception:
        return []


def mark_cma_no_offer(row_index, ws, value="No Offer"):
    ws.cell(row=row_index, column=6).value = value
    try:
        wb.save(excel_path)
    except:
        pass


def mark_cma_form_error(row_index, ws):
    """Do not leave stale price/schedule cells when the quote form is invalid."""
    ws.cell(row=row_index, column=6).value = "Error"
    for column in (7, 8, 9, 10, 11, 13, 14, 15, 16):
        ws.cell(row=row_index, column=column).value = "-"
    try:
        wb.save(excel_path)
    except Exception:
        pass


def smart_update_field(target_type, new_value):
    print(f"   -> Đang sửa {target_type} thành: {new_value}...")
    try:
        wrapper = None
        if target_type == 'POD':
            try:
                dest_div = driver.find_element(By.CSS_SELECTOR, ".destination-search")
                wrapper = dest_div.find_element(By.CSS_SELECTOR, ".selected-value-wrapper")
            except: pass
        else:
            try:
                wrappers = driver.find_elements(By.CSS_SELECTOR, ".selected-value-wrapper")
                if wrappers: wrapper = wrappers[0]
            except: pass

        if wrapper and wrapper.is_displayed():
            driver.execute_script("arguments[0].click();", wrapper)
            try:
                WebDriverWait(driver, 2, poll_frequency=0.05).until(
                    EC.visibility_of_element_located((By.XPATH, "//input[@placeholder='Name / Code / Port']")))
            except: time.sleep(0.3)

        all_inputs = driver.find_elements(By.XPATH, "//input[@placeholder='Name / Code / Port']")
        visible_inputs = [inp for inp in all_inputs if inp.is_displayed()]
        if not visible_inputs: return False
        target_input = visible_inputs[-1] if target_type == 'POD' else visible_inputs[0]

        driver.execute_script("arguments[0].click();", target_input)
        time.sleep(0.15)
        target_input.send_keys(Keys.CONTROL + "a")
        target_input.send_keys(Keys.DELETE)
        if target_input.get_attribute("value"): target_input.clear()

        # Xử lý Logic tách PORT và RAMP
        search_key = new_value
        pick_ramp = False
        pick_port = False
        if new_value == "VNSGN-RAMP":
            search_key = "VNSGN"
            pick_ramp = True
        elif new_value == "VNSGN-PORT":
            search_key = "VNSGN"
            pick_port = True
        elif "VNSGN" in new_value:
            pick_ramp = True # Mặc định tuyến xa lấy RAMP (hoặc Vũng Tàu sau đó)

        target_input.send_keys(search_key)

        try:
            WebDriverWait(driver, 4, poll_frequency=0.05).until(
                EC.visibility_of_any_elements_located((By.CSS_SELECTOR, "li.place-suggestion")))
        except: time.sleep(0.8)

        try:
            suggestions = driver.find_elements(By.CSS_SELECTOR, "li.place-suggestion")
            visible = [s for s in suggestions if s.is_displayed()]
            
            chosen = None
            if search_key == "VNSGN":
                if pick_ramp:
                    chosen = next((s for s in visible if "RAMP" in s.text.upper()), None)
                    if not chosen and len(visible) > 1: chosen = visible[1]
                elif pick_port:
                    chosen = next((s for s in visible if "RAMP" not in s.text.upper()), None)
                    if not chosen and visible: chosen = visible[0]
                else:
                    chosen = visible[1] if len(visible) > 1 else (visible[0] if visible else None)
            else:
                chosen = visible[0] if visible else None

            if chosen: driver.execute_script("arguments[0].click();", chosen)
            else: target_input.send_keys(Keys.ENTER)
        except:
            target_input.send_keys(Keys.ENTER)

        try:
            WebDriverWait(driver, 1.5, poll_frequency=0.05).until_not(
                EC.presence_of_element_located((By.CSS_SELECTOR, "li.place-suggestion")))
        except: time.sleep(0.3)

        return True

    except Exception as e:
        print(f"   -> [ERROR] smart_update_field thất bại: {e}")
        return False


def select_port_full(element, text):
    try:
        driver.execute_script("arguments[0].click();", element)
        element.send_keys(Keys.CONTROL + "a"); time.sleep(0.1)
        element.send_keys(Keys.DELETE)
        
        search_key = text
        pick_ramp = False
        pick_port = False
        if text == "VNSGN-RAMP":
            search_key = "VNSGN"
            pick_ramp = True
        elif text == "VNSGN-PORT":
            search_key = "VNSGN"
            pick_port = True
        elif "VNSGN" in text:
            pick_ramp = True

        element.send_keys(search_key)

        try:
            WebDriverWait(driver, 8).until(EC.presence_of_element_located((By.CSS_SELECTOR, "li.place-suggestion")))
            WebDriverWait(driver, 5).until(EC.visibility_of_any_elements_located((By.CSS_SELECTOR, "li.place-suggestion")))
        except:
            sleep_human(1.5, 2.0)

        try:
            suggestions = driver.find_elements(By.CSS_SELECTOR, "li.place-suggestion")
            visible = [s for s in suggestions if s.is_displayed()]
            
            chosen = None
            if search_key == "VNSGN":
                if pick_ramp:
                    chosen = next((s for s in visible if "RAMP" in s.text.upper()), None)
                    if not chosen and len(visible) > 1: chosen = visible[1]
                elif pick_port:
                    chosen = next((s for s in visible if "RAMP" not in s.text.upper()), None)
                    if not chosen and visible: chosen = visible[0]
                else:
                    chosen = visible[1] if len(visible) > 1 else (visible[0] if visible else None)
            else:
                chosen = visible[0] if visible else None

            if chosen: driver.execute_script("arguments[0].click();", chosen)
            else: element.send_keys(Keys.ENTER)
        except:
            element.send_keys(Keys.ENTER)
    except:
        pass

def select_vung_tau_mandatory():
    print("   -> Đang móc Vũng Tàu (JS Bypass)...")
    try:
        pol_inp = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='Select']"))
        )
        driver.execute_script("arguments[0].click();", pol_inp)

        # Chờ dropdown hiện
        try:
            WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "li.el-select-dropdown__item"))
            )
        except:
            time.sleep(random.uniform(1.0, 1.5))

        js_find_vung_tau = """
        var items = document.querySelectorAll('li.el-select-dropdown__item');
        for(var i=0; i<items.length; i++) {
            var parent = items[i].closest('.el-select-dropdown');
            if(parent && parent.style.display !== 'none') {
                var text = (items[i].textContent || items[i].innerText).toUpperCase();
                if(text.indexOf('VUNG TAU') !== -1 || text.indexOf('VNVUT') !== -1) {
                    items[i].click();
                    return text.trim();
                }
            }
        }
        return null;
        """
        result = driver.execute_script(js_find_vung_tau)

        if not result:
            pol_inp.send_keys("VUNG")
            time.sleep(random.uniform(1.0, 1.5))
            items2 = driver.find_elements(By.CSS_SELECTOR, "li.el-select-dropdown__item")
            for item in items2:
                if item.is_displayed() and (
                    "VUNG TAU" in item.text.upper() or "VNVUT" in item.text.upper()
                ):
                    driver.execute_script("arguments[0].click();", item)
                    break
            else:
                pol_inp.send_keys(Keys.ENTER)
        # ✅ THÊM: Chờ React re-render xong sau khi chọn Vung Tau
        # Đợi dropdown biến mất = form đã stabilize
        try:
            WebDriverWait(driver, 2, poll_frequency=0.05).until_not(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "li.el-select-dropdown__item"))
            )
        except:
            time.sleep(0.4)
    except Exception as e:
        pass


def handle_pod_selection_popup(prefer_port="JEDDAH"):
    """
    Xử lý dropdown 'Select' xuất hiện sau khi chọn port inland (VD: Riyadh).
    Ưu tiên chọn prefer_port (Jeddah), fallback về option đầu tiên.
    """
    try:
        # Chờ input Select xuất hiện tối đa 4s
        dropdown = WebDriverWait(driver, 4).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "input.el-input__inner[placeholder='Select']")
            )
        )
        print(f"   -> Phát hiện dropdown POD → mở lên chọn {prefer_port}...")

        driver.execute_script("arguments[0].click();", dropdown)

        # Chờ options hiện ra
        try:
            WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "li.el-select-dropdown__item"))
            )
        except:
            time.sleep(random.uniform(1.0, 1.5))

        options = driver.find_elements(By.CSS_SELECTOR, "li.el-select-dropdown__item")
        visible = [o for o in options if o.is_displayed()]
        print(f"      Có {len(visible)} options: {[o.text.strip() for o in visible[:5]]}")

        target = next(
            (o for o in visible if prefer_port.upper() in o.text.upper()), None
        )

        if target:
            print(f"      → Chọn: {target.text.strip()}")
            driver.execute_script("arguments[0].click();", target)
        elif visible:
            print(f"      → Không thấy {prefer_port}, chọn đầu tiên: {visible[0].text.strip()}")
            driver.execute_script("arguments[0].click();", visible[0])
        else:
            print(f"      ⚠️ Không có options nào hiện ra")

        sleep_human(0.5, 0.8)

    except Exception:
        pass

def debug_vessel_structure():
    """Debug: In ra cấu trúc HTML của vessel để kiểm tra"""
    try:
        vessel_items = driver.find_elements(By.CSS_SELECTOR, "li.more-infos.vessel")
        if vessel_items:
            first = vessel_items[0]
            html = driver.execute_script("return arguments[0].innerHTML;", first)
            print(f"\n[DEBUG] HTML cấu trúc vessel:\n{html[:500]}...\n")
    except:
        pass

def debug_vessel_html():
    """In ra HTML cấu trúc vessel để kiểm tra"""
    try:
        vessel_items = driver.find_elements(By.CSS_SELECTOR, "li.more-infos.vessel")
        if vessel_items:
            first = vessel_items[0]
            lis = first.find_elements(By.XPATH, ".//ul/li")
            print(f"      [DEBUG] Số <li>: {len(lis)}")
            for i, li in enumerate(lis[:3]):
                html = driver.execute_script("return arguments[0].innerHTML;", li)
                print(f"      [DEBUG] Li #{i}: {html[:200]}...")
    except Exception as e:
        print(f"      [DEBUG] Error: {e}")

def get_cma_commodity_value():
    try:
        return (driver.execute_script("""
            var inputs = document.querySelectorAll('input');
            for (var i = 0; i < inputs.length; i++) {
                var ph = (inputs[i].placeholder || '').toLowerCase();
                var id = (inputs[i].id || '').toLowerCase();
                if (ph.includes('commodity') || id === 'ddlcommodity') {
                    var val = (inputs[i].value || '').trim();
                    if (val && val.toLowerCase() !== 'choose a commodity') return val;
                    var root = inputs[i].closest('.el-select, .commodity, [class*=commodity]');
                    var text = root ? (root.innerText || '').trim() : '';
                    if (/freight all kinds/i.test(text)) return 'Freight All Kinds';
                }
            }
            return '';
        """) or "").strip()
    except Exception:
        return ""


def select_commodity_refresh(max_attempts=3, control_wait_seconds=12):
    print("   -> Chọn lại Commodity...")
    selectors = [
        (By.ID, "DdlCommodity"),
        (By.XPATH, "//div[contains(@class,'commodity')]//input"),
        (By.XPATH, "//*[contains(@placeholder,'commodity') or contains(@placeholder,'Commodity')]"),
        (By.XPATH, "//label[contains(.,'Commodity')]/following::input[1]"),
        (By.XPATH, "//label[contains(.,'Commodity')]/following::div[contains(@class,'el-select')][1]//input"),
    ]
    option_selector = "li.el-select-dropdown__item, [role='option'], .el-select-dropdown__item"

    def find_visible_control():
        for by, sel in selectors:
            try:
                for element in driver.find_elements(by, sel):
                    if element.is_displayed() and element.is_enabled():
                        return element, sel
            except Exception:
                continue
        return None, ""

    for attempt in range(1, max_attempts + 1):
        try:
            comm, used_selector = None, ""
            # Changing POD makes the CMA React form temporarily unmount the
            # Commodity control while it recalculates the port pair. Wait for
            # the control itself instead of doing three short blind probes.
            wait_seconds = control_wait_seconds if attempt == 1 else 4
            control_deadline = time.time() + wait_seconds
            while time.time() < control_deadline:
                comm, used_selector = find_visible_control()
                if comm:
                    break
                time.sleep(0.3)

            if not comm:
                print(
                    f"      ⚠️ Commodity chưa render sau {wait_seconds}s "
                    f"(lần {attempt}/{max_attempts})"
                )
                continue

            print(f"      -> Tìm thấy Commodity bằng selector: {used_selector}")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", comm)
            try:
                comm.click()
            except Exception:
                driver.execute_script("arguments[0].click();", comm)
            if attempt > 1:
                try:
                    comm.send_keys(Keys.ARROW_DOWN)
                except Exception:
                    pass

            deadline = time.time() + 5
            visible = []
            while time.time() < deadline:
                try:
                    visible = [
                        item for item in driver.find_elements(By.CSS_SELECTOR, option_selector)
                        if item.is_displayed() and (item.text or "").strip()
                    ]
                except Exception:
                    visible = []
                if visible:
                    break
                time.sleep(0.2)

            if not visible:
                print(f"      ⚠️ Dropdown chưa hiện options (lần {attempt}/{max_attempts})")
                try:
                    driver.execute_script("arguments[0].click();", comm)
                except Exception:
                    pass
                continue

            chosen = next(
                (item for item in visible if "FREIGHT ALL KINDS" in item.text.upper()),
                None,
            )
            if chosen is None:
                option_names = [(item.text or "").strip() for item in visible[:5]]
                print(f"      ⚠️ Dropdown có options nhưng không có Freight All Kinds: {option_names}")
                continue
            chosen_text = (chosen.text or "").strip()
            driver.execute_script("arguments[0].click();", chosen)

            deadline = time.time() + 3
            while time.time() < deadline:
                value = get_cma_commodity_value()
                if value:
                    print(f"      -> [OK] Đã chọn Commodity: {value or chosen_text}")
                    return True
                time.sleep(0.2)

            print(f"      ⚠️ Đã click '{chosen_text}' nhưng form chưa ghi nhận (lần {attempt}/{max_attempts})")
        except Exception as e:
            print(f"      ⚠️ Lỗi chốt Commodity lần {attempt}/{max_attempts}: {e}")

    print("      ❌ Không chốt được Commodity sau khi retry.")
    return False

def get_main_vessel_and_voyage(search_context=None):
    """
    Lấy tàu chính (Leg 1) từ route detail.
    Format: <TÊN TÀU> (<SERVICE CODE>)
    search_context: card element để tìm kiếm (tránh lấy nhầm từ card khác)
    """
    try:
        # ✅ TÌM TRONG CARD ELEMENT CỤ THỀ (TỪ PARAMETER)
        if search_context is None:
            search_context = driver
        
        vessel_items = search_context.find_elements(By.CSS_SELECTOR, "li.more-infos.vessel")
        
        if not vessel_items:
            print(f"      ⚠️ Không tìm thấy vessel items")
            return "TBA", "N/A"
        
        # ✅ LUÔN LẤYA LI.MORE-INFOS.VESSEL ĐẦU TIÊN TRONG CONTEXT
        first_vessel_li = vessel_items[0]
        
        # ✅ LẤY TÊN TÀU + SERVICE CODE (dùng dt text thay vì index cứng)
        vessel_name = "TBA"
        service_code = "N/A"
        try:
            lis = first_vessel_li.find_elements(By.XPATH, ".//ul/li")
            for li_item in lis:
                dts = li_item.find_elements(By.XPATH, ".//dt")
                dds = li_item.find_elements(By.XPATH, ".//dd")
                for idx_dt, dt in enumerate(dts):
                    dt_text = dt.text.strip().upper()
                    if dt_text == "VESSEL" and idx_dt < len(dds):
                        vessel_name = dds[idx_dt].text.strip()
                    elif dt_text == "SERVICE" and idx_dt < len(dds):
                        service_code = dds[idx_dt].text.strip()
        except:
            pass
        
        result = f"{vessel_name} ({service_code})"
        print(f"      ✅ Tàu chính: {result}")
        return vessel_name, service_code
        
    except Exception as e:
        print(f"      ⚠️ Lỗi bóc vessel tổng quát: {type(e).__name__}")
        return "TBA", "N/A"
              
        
def pick_date_plus_7():
    print(f"   -> Chọn ngày hôm nay +{DATE_OFFSET_DAYS}...")
    try:
        today = datetime.now()
        target = today + timedelta(days=DATE_OFFSET_DAYS)
        current_value = driver.execute_script("""
            const input = document.getElementById('DepartureFrom');
            return input ? (input.value || input.getAttribute('value') || '') : '';
        """) or ""
        if cma_date_input_matches(current_value, target):
            print(f"      -> Ngày {target.strftime('%d-%b-%Y')} đã đúng, không mở calendar lại.")
            return True
        target_day = str(target.day)
        is_next = target.month != today.month

        # Mở lịch
        opened = driver.execute_script("""
            var inp = document.getElementById('DepartureFrom');
            if(!inp) return false;
            var parent = inp.closest('.el-date-editor') || inp.parentElement;
            var icon = parent ? parent.querySelector('i.el-icon, i.el-input__icon') : null;
            if(icon) { icon.click(); return true; }
            inp.click(); inp.dispatchEvent(new MouseEvent('click', {bubbles:true}));
            return true;
        """)

        if not opened:
            print(f"      ⚠️ Không tìm thấy icon calendar")
            return

        # Chờ bảng lịch hiện ra
        try:
            WebDriverWait(driver, 3, poll_frequency=0.05).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "table.el-date-table"))
            )
            time.sleep(0.2)
        except:
            print(f"      ⚠️ Calendar không mở kịp")
            return

        # Nếu là tháng sau, click vào ô chứa class 'next-month'. Cùng tháng thì click 'available'
        js = f"""
        var d='{target_day}';
        var isNext = {'true' if is_next else 'false'};
        var selector = isNext ? 'td.next-month span.el-date-table-cell__text' : 'td.available span.el-date-table-cell__text';
        var s = document.querySelectorAll(selector);
        
        for(var i=0; i<s.length; i++){{
            if(s[i].innerText.trim() === d && s[i].offsetParent !== null){{
                s[i].click();
                return true;
            }}
        }}
        return false;
        """
        result = driver.execute_script(js)

        # Fallback: Đề phòng web không nạp thẻ next-month, ép click nút Next Arrow rồi kiếm lại
        if not result and is_next:
            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, "button.el-picker-panel__icon-btn.arrow-right")
                driver.execute_script("arguments[0].click();", next_btn)
                time.sleep(0.3)
                result = driver.execute_script(js.replace("td.next-month", "td.available"))
            except:
                pass

        print(f"      -> Chọn ngày {target.strftime('%d-%b-%Y')}: {'✅ OK' if result else '⚠️ Không tìm thấy ngày'}")
        return bool(result)

    except Exception as e:
        print(f"      ⚠️ Lỗi pick_date: {e}")  
    

CMA_CONTAINER_20 = "20' Dry Standard"
CMA_CONTAINER_40 = "40' Dry Standard"
CMA_CONTAINER_40HC = "40' Dry High Cube"
CMA_ALL_CONTAINERS = [CMA_CONTAINER_20, CMA_CONTAINER_40, CMA_CONTAINER_40HC]
CMA_CONTAINER_MODES = {
    "20_40HC": [CMA_CONTAINER_20, CMA_CONTAINER_40HC],
    "20_ONLY": [CMA_CONTAINER_20],
}

def remove_cma_unwanted_container(container_name):
    """Gỡ equipment đang lỡ được chọn mà không tick vào card equipment đó."""
    try:
        cards = driver.find_elements(
            By.XPATH,
            f"//ul[contains(@class,'container')]//li[contains(., \"{container_name}\")]"
        )
        for card in cards:
            cls = card.get_attribute("class") or ""
            has_input_value = any((inp.get_attribute("value") or "").strip() for inp in card.find_elements(By.TAG_NAME, "input"))
            if "is-checked" not in cls and not has_input_value:
                continue

            clicked = driver.execute_script("""
                const card = arguments[0];
                const candidates = card.querySelectorAll('button, i, svg, [role="button"], a, span');
                for (const c of candidates) {
                    const cls = (c.className || '').toString().toLowerCase();
                    const aria = (c.getAttribute('aria-label') || '').toLowerCase();
                    const title = (c.getAttribute('title') || '').toLowerCase();
                    const txt = (c.innerText || c.textContent || '').toLowerCase();
                    if (/delete|trash|remove|close|bin/.test(cls + ' ' + aria + ' ' + title + ' ' + txt)) {
                        c.click();
                        return true;
                    }
                }

                // CMA icon thùng rác thường nằm góc phải trên card.
                const rect = card.getBoundingClientRect();
                const target = document.elementFromPoint(rect.right - 18, rect.top + 18);
                if (target && card.contains(target)) {
                    target.click();
                    return true;
                }
                return false;
            """, card)
            if clicked:
                print(f"      -> [{container_name}] gỡ khỏi request (không tick Standard).")
                time.sleep(0.2)
                return True
    except Exception as e:
        print(f"      ⚠️ Không gỡ được [{container_name}]: {type(e).__name__}")
    return False

def configure_cma_containers(mode="20_40HC"):
    wanted = set(CMA_CONTAINER_MODES.get(mode, CMA_CONTAINER_MODES["20_40HC"]))
    print(f"   -> Chọn Container ({mode}): {', '.join(wanted)}")
    try:
        driver.execute_script("window.scrollBy(0, 400);")
        time.sleep(random.uniform(0.3, 0.5))
    except:
        pass

    # Không tick/nhập 40' Dry Standard. Nếu phiên cũ đang giữ equipment này thì gỡ bằng icon thùng rác.
    remove_cma_unwanted_container(CMA_CONTAINER_40)

    for cont in [CMA_CONTAINER_20, CMA_CONTAINER_40HC]:
        try:
            li = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located(
                    (By.XPATH, f"//ul[contains(@class,'container')]//li[contains(., \"{cont}\")]"))
            )
            is_checked = "is-checked" in (li.get_attribute("class") or "")
            should_select = cont in wanted

            if should_select and not is_checked:
                driver.execute_script("arguments[0].click();", li)
                sleep_human(0.2, 0.4)
                is_checked = True
            inputs = li.find_elements(By.TAG_NAME, "input")
            if not inputs:
                continue
            inp = inputs[-1]
            if should_select:
                current_val = (inp.get_attribute("value") or "").strip()
                if current_val != "22222":
                    driver.execute_script("arguments[0].click();", inp)
                    inp.send_keys(Keys.CONTROL + "a")
                    inp.send_keys(Keys.DELETE)
                    inp.send_keys("22222")
                    print(f"      -> [{cont}] ✅ chọn + nhập 22222")
                else:
                    print(f"      -> [{cont}] ✅ đã có 22222")
            elif is_checked:
                remove_cma_unwanted_container(cont)
        except Exception as e:
            print(f"      ⚠️ Không xử lý được [{cont}]: {type(e).__name__}")


def cma_container_configuration_ok(mode="20_40HC"):
    """Fast DOM check used to avoid reopening equipment controls every row."""
    wanted = list(CMA_CONTAINER_MODES.get(mode, CMA_CONTAINER_MODES["20_40HC"]))
    try:
        return bool(driver.execute_script("""
            const wanted = new Set(arguments[0]);
            const names = arguments[1];
            const cards = Array.from(document.querySelectorAll('ul[class*="container"] li'));
            const norm = value => String(value || '').replace(/\\s+/g, ' ').trim();

            for (const name of names) {
                const card = cards.find(item => norm(item.innerText || item.textContent).includes(name));
                if (!card) {
                    if (wanted.has(name)) return false;
                    continue;
                }
                const inputs = Array.from(card.querySelectorAll('input'));
                const values = inputs.map(input => String(input.value || '').trim()).filter(Boolean);
                const selected = String(card.className || '').includes('is-checked') || values.length > 0;
                if (wanted.has(name)) {
                    if (!selected || !values.includes('22222')) return false;
                } else if (selected) {
                    return false;
                }
            }
            return true;
        """, wanted, CMA_ALL_CONTAINERS))
    except Exception:
        return False

def select_containers(mode="20_40HC"):
    configure_cma_containers(mode)

def ensure_containers_filled(mode="20_40HC"):
    if cma_container_configuration_ok(mode):
        print("   -> Container đã đúng cấu hình + 22222kgs, bỏ qua thao tác lại.")
        return True
    print("   -> Container chưa đúng; cấu hình và điền 22222kgs...")
    configure_cma_containers(mode)
    return cma_container_configuration_ok(mode)

# ===================================================================================
# --- 3. BỘ NÃO LỌC ETD & BÓC GIÁ (9 QUY TẮC VÀNG) ---
# ===================================================================================
def parse_cma_date(date_str):
    try:
        return datetime.strptime(date_str, "%d-%b-%Y")
    except:
        return None

def calculate_cma_validity(last_etd):
    """Tính mốc Valid (7, 14, 21, cuối tháng) dựa trên ETD cuối."""
    import calendar
    year, month = last_etd.year, last_etd.month
    last_day = calendar.monthrange(year, month)[1]
    
    milestones = [7, 14, 21, last_day]
    for m in milestones:
        if m >= last_etd.day:
            dt = datetime(year, month, m)
            return f"{dt.day}-{dt.strftime('%b')}"
    return f"{last_day}-{last_etd.strftime('%b')}"

def calculate_cma_validity(last_etd):
    """Tính mốc Valid (7, 14, 21, cuối tháng) dựa trên ETD cuối."""
    import calendar
    year, month = last_etd.year, last_etd.month
    last_day = calendar.monthrange(year, month)[1]
    
    milestones = [7, 14, 21, last_day]
    for m in milestones:
        if m >= last_etd.day:
            dt = datetime(year, month, m)
            return f"{dt.day}-{dt.strftime('%b')}"
    return f"{last_day}-{last_etd.strftime('%b')}"

def scrape_multi_etd_and_save(row_index, ws, pol_text_excel, price_mode="20_40HC"):
    print("8. Đang quét danh sách tàu và bóc tách dữ liệu chi tiết...")
    pod_for_rule = str(ws.cell(row=row_index, column=4).value or "").strip()
    country_for_rule = str(ws.cell(row=row_index, column=2).value or "").strip()
    china_route = is_china_destination(country_for_rule, pod_for_rule)
    if is_cma_no_offer_page():
        print("      ⚠️ CMA báo không có SpotOn offer cho request này. Bỏ qua nhanh.")
        mark_cma_no_offer(row_index, ws)
        return "NO_ROUTE"
    
    # --- THÊM BƯỚC: POLLING JAVASCRIPT ĐỂ CLICK AVAILABLE ---
    try:
        print("      -> Đang Polling tìm và click chọn 'Available solutions only'...")
        
        js_polling = """
        // Tìm trực tiếp input có value là 'OnlyAvailable'
        var input = document.querySelector('input[value="OnlyAvailable"]');
        if (input) {
            var label = input.closest('label');
            if (label) {
                if (!label.className.includes('is-checked')) {
                    label.click();
                    return 'CLICKED';
                }
                return 'ALREADY_CHECKED';
            }
        }
        
        // Fallback: Tìm thẻ label chứa chữ 'Available'
        var labels = document.querySelectorAll('label.el-radio');
        for(var i=0; i<labels.length; i++) {
            if(labels[i].innerText.trim() === 'Available') {
                if(!labels[i].className.includes('is-checked')) {
                    labels[i].click();
                    return 'CLICKED';
                }
                return 'ALREADY_CHECKED';
            }
        }
        // Trả về chuỗi rỗng để WebDriverWait tiếp tục quét
        return ''; 
        """

        # Chạy polling ngắn. Nếu route không có nút Available thì tránh mất 5s mỗi row.
        # Nếu JS trả về 'CLICKED' hoặc 'ALREADY_CHECKED' (Truthy) thì vòng lặp dừng ngay lập tức.
        result = WebDriverWait(driver, CMA_AVAILABLE_WAIT_SECONDS, poll_frequency=0.2).until(
            lambda d: d.execute_script(js_polling)
        )
        
        if result == 'CLICKED':
            print("      -> Đã click chọn bộ lọc 'Available' thành công!")
            # Snapshot waiter ngay bên dưới sẽ theo dõi React; chỉ nhường một
            # nhịp ngắn để DOM bắt đầu cập nhật.
            time.sleep(0.35)
        elif result == 'ALREADY_CHECKED':
            print("      -> Form đã tự động chọn sẵn 'Available', đi tiếp...")
            
    except TimeoutException:
        print(f"      [WARN] Hết {CMA_AVAILABLE_WAIT_SECONDS:g}s Polling không tìm thấy nút Available, tiếp tục kịch bản...")
    except Exception as e:
        print(f"      [WARN] Lỗi không xác định khi Polling Available: {type(e).__name__}")
    # ---------------------------------------------S

    try:
        # Chờ theo trạng thái card. Nếu đã có giá và phần còn lại đứng yên thì
        # đi tiếp sớm, thay vì bắt mọi card phải nhả USD/SOLD OUT.
        print("      -> [WAIT] Chờ snapshot giá CMA ổn định...")
        price_state = wait_for_cma_price_snapshot()
        if price_state in {"READY", "PARTIAL"}:
            print(f"      -> Snapshot giá sẵn sàng ({price_state}).")
        else:
            print(
                f"      [WARN] Hết {CMA_PRICE_WAIT_SECONDS:g}s chờ giá "
                f"(state={price_state}); vẫn đọc các card đã có dữ liệu."
            )

        cards = driver.find_elements(By.CSS_SELECTOR, "article.card-route-horizontal")
        if not cards:
            # 1. Kiểm tra khung thông báo lỗi SpotOn (We apologize, We are currently not able to propose...)
            try:
                no_result = driver.find_elements(By.CSS_SELECTOR, "div.no-result-message")
                if no_result and any(el.is_displayed() for el in no_result):
                    print("      ⚠️ Hãng báo không thể cung cấp SpotOn offer cho request này (No Space / No Route)!")
                    ws.cell(row=row_index, column=6).value = "No Offer"
                    try: wb.save(excel_path)
                    except: pass
                    return "NO_OFFER"
            except:
                pass

            # 2. Kiểm tra fallback thông qua text của trang nếu hãng đổi layout
            page_text_upper = driver.execute_script("return document.body.innerText;").upper()
            if "NO SPACE" in page_text_upper or "FULLY BOOKED" in page_text_upper:
                print("      ⚠️ No Space (tàu đã đầy chỗ)!")
                ws.cell(row=row_index, column=6).value = "No Space"
                try: wb.save(excel_path)
                except: pass
                return "NO_SPACE"
                
            print("      ⚠️ Sold Out toàn bộ (Không có chuyến tàu nào)!")
            ws.cell(row=row_index, column=6).value = "Sold Out"
            try: wb.save(excel_path)
            except: pass
            return "SOLD_OUT"

        # ══════════════════════════════════════════════════════════════════════════════
        # BƯỚC 1: QUÉT TẤT CẢ CARD, LỌC THEO 9 QUY TẮC VÀNG (BẢN GẮN RADAR)
        # ══════════════════════════════════════════════════════════════════════════════
        all_options = []
        skipped_modify_teu = 0
        partial_equipment_no_offer = False
        summaries = collect_cma_card_summaries()
        if not summaries:
            summaries = [{"card_idx": idx, "text": ""} for idx, _ in enumerate(cards[:8])]

        for summary in summaries:
            try:
                idx = int(summary.get("card_idx", 0))
                if idx >= len(cards):
                    continue
                card = cards[idx]
                print(f"      [DEBUG] Đang xét Card #{idx+1}...")
                raw_text = str(summary.get("text") or "")
                if not raw_text:
                    raw_text = driver.execute_script(
                        "return arguments[0].innerText || arguments[0].textContent || '';",
                        card,
                    ) or ""
                raw_text_upper = raw_text.upper()
                if "MODIFY TEU" in raw_text_upper:
                    skipped_modify_teu += 1
                    if "NO OFFER" in raw_text_upper and re.search(r"([\d\s,]+)\s*(?:USD|US\$|\$|EUR)", raw_text_upper):
                        partial_equipment_no_offer = True
                        print("         -> Card thiếu offer cho một số equipment: sẽ Modify và retry 20-only.")
                    else:
                        print("         -> Bỏ qua: Card này chỉ có nút Modify TEU, không phải offer bookable.")
                    continue
                
                # 1. Bóc Date
                try:
                    raw_etd = str(summary.get("date_text") or "").strip()
                    if not raw_etd:
                        raw_etd = card.find_element(By.CSS_SELECTOR, "span.date, .date").text.strip()
                    etd_str = raw_etd.split(",")[1].strip() if "," in raw_etd else raw_etd
                    etd_date = parse_cma_date(etd_str)
                    
                    if not etd_date:
                        clean_str = etd_str.replace("-", " ")
                        try:
                            etd_date = datetime.strptime(clean_str, "%d %b %Y")
                        except: pass
                except Exception as e:
                    print(f"         -> Bỏ qua: Không tìm thấy thẻ Ngày/Tháng ({type(e).__name__})")
                    continue

                if not etd_date:
                    print(f"         -> Bỏ qua: Lỗi format ngày '{raw_etd}'")
                    continue
                if not etd_within_max(etd_date):
                    print(f"         -> Bỏ qua: ETD {etd_date.strftime('%d-%b')} xa quá 21 ngày")
                    continue

                # 2. Bóc Transit
                try:
                    transit_text = str(summary.get("transit_text") or "")
                    if not transit_text:
                        transit_text = card.find_element(
                            By.XPATH,
                            ".//*[contains(@class, 'transit')]",
                        ).text
                    match_tr = re.search(r"(\d+)", transit_text)
                    transit_val = int(match_tr.group(1)) if match_tr else 99
                except:
                    transit_val = 99

                # 3. Bóc Transshipment
                ts_text = "DIRECT"
                try:
                    raw_ts = str(summary.get("ts_text") or "").strip()
                    ts_class = str(summary.get("ts_class") or "").lower()
                    if not raw_ts:
                        ts_div = card.find_element(By.XPATH, ".//div[contains(@class,'transit') and (contains(@class,'transshipment') or contains(@class,'direct'))]")
                        raw_ts = ts_div.text.strip()
                        ts_class = (ts_div.get_attribute("class") or "").lower()
                    if "direct" not in ts_class and "DIRECT" not in raw_ts.upper():
                        raw_ts = re.sub(r"^\s*via\s*", "", raw_ts, flags=re.I).strip()
                        ports  = [p.split(",")[0].strip() for p in raw_ts.split("•")]
                        ts_text = " + ".join(ports)
                except: pass

                # 4. Bóc Giá (Dùng JS ép đọc text)
                price_match = re.search(r"([\d\s,]+)\s*(?:USD|US\$|\$|EUR)", raw_text_upper)
                fallback_match = re.search(r"PER\s*(?:20|40)[A-Z0-9]*\s*([\d\s,]+)", raw_text_upper)
                
                if price_match:
                    current_price = int(re.sub(r"[^\d]", "", price_match.group(1)))
                elif fallback_match:
                    current_price = int(re.sub(r"[^\d]", "", fallback_match.group(1)))
                else:
                    current_price = 0

                # KIỂM TRA ĐIỀU KIỆN LOẠI BỎ
                if "SOLD OUT" in raw_text_upper and current_price == 0:
                    print("         -> Bỏ qua: Card này bị dính chữ SOLD OUT.")
                    continue
                
                if current_price == 0:
                    print("         -> Bỏ qua: Price = 0 (Giá chưa tải hoặc ẩn).")
                    continue

                print(f"         => HỢP LỆ! Ghi nhận: Giá {current_price}, ETD {etd_str}, Transit {transit_val} Ngày")
                
                all_options.append({
                    'date': etd_date, 'price': current_price, 'transit': transit_val,
                    'card_idx': idx, 'ts_port': ts_text
                })
            except Exception as ex:
                print(f"         -> Lỗi văng code không xác định: {ex}")
                continue

        before_dedupe = len(all_options)
        all_options = dedupe_cma_card_summaries(all_options)
        if len(all_options) != before_dedupe:
            print(
                f"      -> Loại {before_dedupe - len(all_options)} card trùng trước khi mở Details."
            )

        if not all_options:
            if partial_equipment_no_offer and price_mode != "20_ONLY":
                print("      ⚠️ CMA báo No offer for some equipments -> cần retry chỉ 20'.")
                return "RETRY_20_ONLY"
            if skipped_modify_teu:
                print("      ⚠️ Không có card bookable sau khi bỏ qua Modify TEU.")
                ws.cell(row=row_index, column=6).value = "No Offer"
                try: wb.save(excel_path)
                except: pass
                return "NO_OFFER"
            print("      ⚠️ Toàn bộ bảng giá đều đã Sold Out hoặc không lấy được dữ liệu!")
            ws.cell(row=row_index, column=6).value = "Sold Out"
            try: wb.save(excel_path)
            except: pass
            return "SOLD_OUT"

        # --- BƯỚC 1.5: TÌM GIÁ ĐÁY VÀ VỨT BỎ GIÁ CAO ---
        min_price = min(opt['price'] for opt in all_options)
        best_price_options = [opt for opt in all_options if opt['price'] == min_price]

        # Lọc cùng ngày (trong nhóm giá rẻ nhất) → giữ transit ngắn nhất
        unique_dates = {}
        for opt in best_price_options:
            d = opt['date']
            if d not in unique_dates:
                unique_dates[d] = opt
            else:
                if opt['transit'] < unique_dates[d]['transit']:
                    unique_dates[d] = opt

        sorted_opts = sorted(unique_dates.values(), key=lambda x: x['date'])
        e_min = sorted_opts[0]['date']

        # Lọc cửa sổ 9 ngày, cách nhau >= 2 ngày, tối đa 3 card
        valid_opts = []
        for opt in sorted_opts:
            if (opt['date'] - e_min).days > 9:
                break
            if not valid_opts:
                valid_opts.append(opt)
            else:
                if (opt['date'] - valid_opts[-1]['date']).days >= 2:
                    valid_opts.append(opt)
            if len(valid_opts) == 3:
                break

        if not valid_opts:
            return

        # ══════════════════════════════════════════════════════════════════════════════
        # BƯỚC 2: BÓC CHI TIẾT TỪNG CARD (VESSEL, FREE TIME)
        #         + LẤY GIÁ TỪ CARD ĐẦU TIÊN
        # ══════════════════════════════════════════════════════════════════════════════
        vessel_entries = []
        ts_entries     = []
        free_time_pod  = "N/A"
        total_20 = total_40 = total_40hc = 0
        formula_20 = formula_40 = formula_40hc = None
        othc_included  = False
        ows_found      = False
        manifest_fee_found = False

        for i, opt in enumerate(valid_opts):
            idx = opt['card_idx']

            fresh_cards = driver.find_elements(By.CSS_SELECTOR, "article.card-route-horizontal")
            if idx >= len(fresh_cards):
                continue
            card_el = fresh_cards[idx]

            print(f"   -> Đang bóc chi tiết Card #{idx + 1}...")

            # Click mở Details
            try:
                # Tăng thời gian chờ từ 2s lên 8s để web có đủ thời gian thở và render nút
                details_btn = WebDriverWait(driver, 8, poll_frequency=0.2).until(
                    lambda d: card_el.find_element(
                        By.XPATH, ".//label[contains(@class,'o-button') and contains(.,'Details')]")
                )
                switches  = card_el.find_elements(By.XPATH, ".//input[@data-role='switch']")
                is_opened = switches and switches[0].is_selected()

                if not is_opened:
                    driver.execute_script(
                        "arguments[0].scrollIntoView({behavior:'instant',block:'center'});"
                        "window.scrollBy(0,-100);", details_btn)
                    time.sleep(0.2)
                    driver.execute_script("arguments[0].click();", details_btn)
                    
                    # Chờ một nhịp để bảng Details thực sự bung ra
                    time.sleep(1.0)
                    
                    # ✅ LẤY FRESH CARD SAU KHI CLICK DETAILS
                    fresh_cards = driver.find_elements(By.CSS_SELECTOR, "article.card-route-horizontal")
                    if idx < len(fresh_cards):
                        card_el = fresh_cards[idx]

            except Exception as e:
                print(f"      ⚠️ Nút Details load chậm ({type(e).__name__}). Đang dùng JS ép click...")
                try:
                    js_force_click = f"""
                    var cards = document.querySelectorAll('article.card-route-horizontal');
                    if(cards.length > {idx}) {{
                        var labels = cards[{idx}].querySelectorAll('label');
                        for(var j=0; j<labels.length; j++) {{
                            if(labels[j].innerText.includes('Details')) {{
                                labels[j].scrollIntoView({{behavior:'instant',block:'center'}});
                                window.scrollBy(0,-100);
                                labels[j].click();
                                return true;
                            }}
                        }}
                    }}
                    return false;
                    """
                    if driver.execute_script(js_force_click):
                        time.sleep(1.5)
                    else:
                        raise Exception("JS không tìm thấy nút Details")
                except Exception as ex:
                    print(f"      ❌ Bất lực với nút Details: {ex}. Bỏ qua card này.")
                    # CÚ PHÁP MỚI CHO FALLBACK
                    vessel_entries.append(
                        f"TBA / ETD: {opt['date'].day}-{opt['date'].strftime('%b')} "
                        f"/ Transit time: {opt['transit']} Days / Transshipment: {opt['ts_port']}")
                    ts_entries.append(opt['ts_port'])
                    continue

            # Lấy lại card tươi sau khi React bung tab
            fresh_cards = driver.find_elements(By.CSS_SELECTOR, "article.card-route-horizontal")
            card_el = fresh_cards[idx]

            # ── GIÁ: Chỉ lấy từ card đầu tiên ──────────────────────
            if i == 0:
                try:
                    print("   -> Đang bóc bảng giá chi tiết...")

                    # Chờ bảng giá xuất hiện
                    try:
                        WebDriverWait(driver, 8, poll_frequency=0.1).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "tr.el-table__row"))
                        )
                        time.sleep(0.3)
                    except:
                        time.sleep(2.0)

                    # Nếu vẫn không thấy → thử click Details lại một lần
                    if not driver.find_elements(By.CSS_SELECTOR, "tr.el-table__row"):
                        driver.execute_script("arguments[0].click();", details_btn)
                        try:
                            WebDriverWait(driver, 6, poll_frequency=0.1).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "tr.el-table__row"))
                            )
                        except:
                            time.sleep(2.0)

                    # Expand các dòng nhóm để đọc đủ phụ phí, gồm export BL fees như AMS.
                    try:
                        expand_rows = driver.find_elements(By.XPATH,
                            "//tr[.//td[contains(.,'Charges payable as per freight')] "
                            "or .//td[contains(.,'Charges payable at export')] "
                            "or .//td[contains(.,'Ocean Freight')]]")
                        for r in expand_rows:
                            try:
                                icon = r.find_element(By.CSS_SELECTOR, ".el-table__expand-icon")
                                if "expanded" not in icon.get_attribute("class"):
                                    driver.execute_script("arguments[0].click();", icon)
                                    time.sleep(0.3)
                            except: pass
                    except: pass

                    # JS quét bảng giá
                    # JS quét bảng giá
                    js_price = """
                    var priceMode = arguments[0] || '20_40HC';
                    var result = {ocean:[0,0,0], surcharge:[0,0,0], ows:[0,0,0], othc_items:[], othc_included:false, manifest_fee_found:false};

                    function amountFromCell(cell) {
                        if(!cell) return 0;
                        return parseInt((cell.textContent || cell.innerText || "").replace(/[^0-9]/g,'')) || 0;
                    }

                    function preciseAmountFromCell(cell) {
                        if(!cell) return 0;
                        var raw = (cell.textContent || cell.innerText || "").replace(/,/g,'').replace(/\\s+/g,' ');
                        var match = raw.match(/-?\\d+(?:\\.\\d+)?/);
                        return match ? (parseFloat(match[0]) || 0) : 0;
                    }

                    function inferColumnMap() {
                        var map = {c20:null, c40:null, c40hc:null};
                        var headers = document.querySelectorAll('.el-table__header-wrapper th');
                        for(var h=0; h<headers.length; h++) {
                            var ht = (headers[h].innerText || headers[h].textContent || '').toUpperCase().replace(/\\s+/g, ' ');
                            if(!ht) continue;
                            if(/20/.test(ht) && (ht.indexOf('ST') !== -1 || ht.indexOf('GP') !== -1 || ht.indexOf("20'") !== -1)) {
                                map.c20 = h;
                            } else if(/40/.test(ht) && (ht.indexOf('HC') !== -1 || ht.indexOf('HQ') !== -1 || ht.indexOf('HIGH CUBE') !== -1)) {
                                map.c40hc = h;
                            } else if(/40/.test(ht) && (ht.indexOf('ST') !== -1 || ht.indexOf('GP') !== -1 || ht.indexOf("40'") !== -1)) {
                                map.c40 = h;
                            }
                        }

                        if(map.c20 === null) map.c20 = 2;
                        if(map.c40 === null && map.c40hc === null) {
                            if(priceMode === '20_ONLY') {
                                map.c40 = null;
                                map.c40hc = null;
                            } else if(priceMode === '20_40HC') {
                                map.c40 = null;
                                map.c40hc = 3;
                            } else {
                                map.c40 = 3;
                                map.c40hc = 4;
                            }
                        }
                        return map;
                    }

                    var colMap = inferColumnMap();

                    var rows = document.querySelectorAll('.el-table__body tbody tr');
                    var manifestRegex = /(ADVANCED?\\s+MANIFEST\\s+DECLARATION\\s+FEE|ADVANCE\\s+MANIFEST\\s+DECLARATION\\s+FEE|MANIFEST\\s+DECLARATION\\s+FEE|ENTRY\\s+SUMMARY\\s+DECLARATION\\s+CHARGE|ENS\\s+SURCHARGE)/;
                    var section = "";
                    for(var i=0; i<rows.length; i++) {
                        var rowText = (rows[i].innerText || rows[i].textContent || "").trim().toUpperCase();
                        if(manifestRegex.test(rowText) && /\\d/.test(rowText)) {
                            result.manifest_fee_found = true;
                        }
                        if(rowText.indexOf('OCEAN FREIGHT') !== -1) {
                            section = "OCEAN";
                        } else if(rowText.indexOf('CHARGES PAYABLE AS PER FREIGHT') !== -1) {
                            section = "FREIGHT";
                        } else if(rowText.indexOf('CHARGES PAYABLE AT EXPORT') !== -1) {
                            section = "EXPORT";
                        } else if(rowText.indexOf('CHARGES PAYABLE AT IMPORT') !== -1) {
                            section = "IMPORT";
                        }

                        if(section === "OCEAN" &&
                           rowText.indexOf('INCLUDED CHARGES') !== -1 &&
                           (rowText.indexOf('TERMINAL HANDLING CHARGE') !== -1 ||
                            rowText.indexOf('(OTHC)') !== -1 ||
                            rowText.indexOf('OTHC') !== -1)) {
                            result.othc_included = true;
                        }

                        var cols = rows[i].querySelectorAll('td');
                        if(cols.length >= 3) {
                            var name  = (cols[1].textContent || cols[1].innerText || "").trim().toUpperCase();
                            var v20   = amountFromCell(cols[colMap.c20]);
                            var v40   = colMap.c40 === null ? 0 : amountFromCell(cols[colMap.c40]);
                            var v40hc = colMap.c40hc === null ? 0 : amountFromCell(cols[colMap.c40hc]);
                            var currencyMatch = rowText.match(/\b(USD|VND|EUR|AUD|CHF|CNY|GBP)\b/);
                            var rowCurrency = currencyMatch ? currencyMatch[1] : 'USD';

                            if(name.indexOf('OCEAN FREIGHT') !== -1) {
                                if(v20>0) result.ocean[0]=v20;
                                if(v40>0) result.ocean[1]=v40;
                                if(v40hc>0) result.ocean[2]=v40hc;
                            }
                            else if(name.indexOf('CHARGES PAYABLE AS PER FREIGHT') !== -1) {
                                if(v20>0) result.surcharge[0]=v20;
                                if(v40>0) result.surcharge[1]=v40;
                                if(v40hc>0) result.surcharge[2]=v40hc;
                            }
                            else if(name.indexOf('OVERWEIGHT SURCHARGE') !== -1) {
                                if(v20>0) result.ows[0]=v20;
                                if(v40>0) result.ows[1]=v40;
                                if(v40hc>0) result.ows[2]=v40hc;
                            }
                            else if(section === 'EXPORT' &&
                                    (name.indexOf('TERMINAL HANDLING') !== -1 ||
                                     name.indexOf('ORIGIN THC') !== -1 ||
                                     name.indexOf('(OTHC)') !== -1 ||
                                     name === 'OTHC' || name.indexOf('THC/L') !== -1) &&
                                    name.indexOf('DESTINATION') === -1 && name.indexOf('IMPORT') === -1) {
                                result.othc_items.push({
                                    currency:rowCurrency,
                                    values:[
                                        preciseAmountFromCell(cols[colMap.c20]),
                                        colMap.c40 === null ? 0 : preciseAmountFromCell(cols[colMap.c40]),
                                        colMap.c40hc === null ? 0 : preciseAmountFromCell(cols[colMap.c40hc])
                                    ],
                                    name:name
                                });
                            }
                        }
                    }
                    return result;
                    """
                    data = driver.execute_script(js_price, price_mode)

                    othc_parts = [[], [], []]
                    if china_route:
                        for thc_item in data.get('othc_items', []):
                            currency = str(thc_item.get('currency') or 'USD').upper()
                            values = thc_item.get('values') or [0, 0, 0]
                            for pos in range(3):
                                raw_amount = values[pos] if pos < len(values) else 0
                                if raw_amount:
                                    usd_amount = charge_amount_to_usd(raw_amount, currency)
                                    othc_parts[pos].append(usd_amount)
                                    print(f"   [+O.THC CHINA] {thc_item.get('name')} {raw_amount} {currency} -> {usd_amount:.2f} USD")

                    total_20   = data['ocean'][0] + data['surcharge'][0] - data['ows'][0] + sum(othc_parts[0])
                    total_40   = data['ocean'][1] + data['surcharge'][1] - data['ows'][1] + sum(othc_parts[1])
                    total_40hc = data['ocean'][2] + data['surcharge'][2] - data['ows'][2] + sum(othc_parts[2])
                    formula_20 = _excel_formula_from_parts([data['ocean'][0], data['surcharge'][0], -data['ows'][0], *othc_parts[0]])
                    formula_40 = _excel_formula_from_parts([data['ocean'][1], data['surcharge'][1], -data['ows'][1], *othc_parts[1]])
                    formula_40hc = _excel_formula_from_parts([data['ocean'][2], data['surcharge'][2], -data['ows'][2], *othc_parts[2]])
                    othc_included = bool(data['othc_included'] or china_route)
                    manifest_fee_found = data['manifest_fee_found']
                    ows_found  = any(data['ows'][j] > 0 for j in range(3))

                    print(f"   => GIÁ CHỐT: 20'={total_20} | 40'={total_40} | 40HC={total_40hc}"
                          f"{' | OTHC included' if othc_included else ''}"
                          f"{' | Manifest/ENS detected' if manifest_fee_found else ''}"
                          f"{' | OWS detected' if ows_found else ''}")

                except Exception as e:
                    print(f"      ⚠️ Lỗi bóc bảng giá: {e}")

            # ── FREE TIME: Chỉ lấy từ card đầu tiên ────────────────
            if i == 0:
                try:
                    print("   -> Đang tìm và mở tab Free Time (D&D)...")
                    tab_dd = None
                    
                    # Ưu tiên tìm tab D&D bằng text
                    dd_tabs = card_el.find_elements(By.XPATH, ".//*[contains(@class,'el-tabs__item') or @role='tab']")
                    for t in dd_tabs:
                        txt = driver.execute_script("return arguments[0].innerText;", t)
                        if txt and any(k in txt.upper() for k in ["D&D", "DND", "DETENTION", "DEMURRAGE", "FREE TIME"]):
                            tab_dd = t
                            break

                    if tab_dd:
                        # Cuộn mượt và click ép bằng JS
                        driver.execute_script(
                            "arguments[0].scrollIntoView({behavior:'instant',block:'center'});"
                            "window.scrollBy(0,-100);", tab_dd)
                        time.sleep(0.3)
                        driver.execute_script("arguments[0].click();", tab_dd)

                        # Chờ bảng dữ liệu hiển thị
                        try:
                            active_dd_pane = WebDriverWait(driver, 4, poll_frequency=0.1).until(
                                lambda d: next((
                                    p for p in card_el.find_elements(By.XPATH, ".//div[contains(@class,'el-tab-pane')]")
                                    if p.is_displayed() and p.find_elements(By.XPATH, ".//table")
                                ), None)
                            )
                        except Exception:
                            print("      [WARN] Click được tab Free Time nhưng không tìm thấy thẻ <table>.")
                            active_dd_pane = None

                        dem, det, merged = "", "", ""
                        debug_headers = []
                        debug_rows = []

                        if active_dd_pane:
                            # TÌM BẢNG IMPORT FREE TIME (Bỏ qua Export)
                            tables = active_dd_pane.find_elements(By.XPATH, ".//table")
                            target_table = None
                            
                            if len(tables) > 1:
                                for tbl in tables:
                                    # Quét text của thẻ cha bọc table để xem có chữ IMPORT không
                                    parent_text = driver.execute_script("return arguments[0].parentElement.innerText;", tbl).upper()
                                    if "IMPORT" in parent_text or "DESTINATION" in parent_text:
                                        target_table = tbl
                                        break
                                # Nếu không thấy chữ Import, mặc định lấy bảng cuối cùng (bên phải)
                                if not target_table:
                                    target_table = tables[-1] 
                            elif tables:
                                target_table = tables[0]

                            if target_table:
                                # THUẬT TOÁN ĐỌC HEADER TRÊN BẢNG IMPORT ĐÃ CHỌN
                                headers = target_table.find_elements(By.XPATH, ".//thead/tr/th")
                                charge_idx, duration_idx = 0, 2 # Mặc định
                                
                                for idx, th in enumerate(headers):
                                    th_text = driver.execute_script("return arguments[0].innerText;", th).upper().strip()
                                    debug_headers.append(th_text)
                                    
                                    if th_text in ["CHARGE", "CHARGE TYPE", "FEE", "DESCRIPTION", "TYPE"]: 
                                        charge_idx = idx
                                    elif th_text in ["DURATION", "PERIOD", "FREE TIME", "LIMIT", "DAYS"]: 
                                        duration_idx = idx

                                # Quét các dòng dữ liệu trong tbody của ĐÚNG BẢNG ĐÓ
                                rows_ft = target_table.find_elements(By.XPATH, ".//tbody/tr")
                                
                                for r in rows_ft:
                                    cols = r.find_elements(By.XPATH, ".//*[self::td or self::th]")
                                    row_texts = [driver.execute_script("return arguments[0].innerText;", c).strip() for c in cols]
                                    if row_texts and any(row_texts):
                                        debug_rows.append(" | ".join(row_texts))
                                    
                                    if len(cols) > max(charge_idx, duration_idx):
                                        charge_type = driver.execute_script("return arguments[0].innerText;", cols[charge_idx]).upper()
                                        duration_val = driver.execute_script("return arguments[0].innerText;", cols[duration_idx])

                                        match = re.search(r"(\d+)", duration_val)
                                        if match:
                                            val = match.group(1)
                                            if any(k in charge_type for k in ["MERGED", "COMBINED", "D&D", "FREE TIME", "D & D"]) and not merged:
                                                merged = val
                                            elif "DEMURRAGE" in charge_type and not dem:
                                                dem = val
                                            elif "DETENTION" in charge_type and not det:
                                                det = val

                        # Tổng hợp và xuất chuỗi format chuẩn chỉnh
                        if merged:
                            free_time_pod = f"{merged} COMBINED"
                        elif dem and det:
                            free_time_pod = f"{dem} DEM + {det} DET"
                        elif dem:
                            free_time_pod = f"{dem} DEM"
                        elif det:
                            free_time_pod = f"{det} DET"
                        else:
                            print("      [WARN] Có bảng nhưng không bóc được số liệu.")
                            print(f"      [DEBUG] Headers: {debug_headers}")
                            print(f"      [DEBUG] Rows: {debug_rows}")
                    else:
                        print("      [WARN] Tuyến này hãng không hiển thị tab Free Time/D&D.")

                except Exception as e:
                    print(f"      [WARN] Lỗi bóc Free Time: {type(e).__name__} - {e}")

            # ── VESSEL & VOYAGE: Tab Route ─────────────────────────
            try:
                tab_route = card_el.find_element(By.XPATH,
                    ".//div[contains(@class,'el-tabs__item') and contains(.,'Route')]")
                driver.execute_script(
                    "arguments[0].scrollIntoView({behavior:'smooth',block:'center'});"
                    "window.scrollBy(0,-150);", tab_route)
                time.sleep(0.1)
                driver.execute_script("arguments[0].click();", tab_route)

                active_pane_xpath = (".//div[contains(@class,'el-tab-pane') "
                                     "and not(contains(@style,'display: none'))]")
                WebDriverWait(driver, 3, poll_frequency=0.05).until(
                    lambda d: card_el.find_elements(By.XPATH,
                        f"{active_pane_xpath}//ul[contains(@class,'list-main') "
                        f"or contains(@class,'route') or contains(@class,'vessel')]")
                )
                time.sleep(0.5)  # ← Tăng từ 0.1 lên 0.5s để React render xong

                # ✅ LẤY FRESH CARD SAU KHI CLICK TAB ROUTE
                fresh_cards = driver.find_elements(By.CSS_SELECTOR, "article.card-route-horizontal")
                if idx < len(fresh_cards):
                    card_el = fresh_cards[idx]

                # ✅ LẤY TÀU CHÍNH (LEG 1) TRỰC TIẾP TRONG CARD ĐANG XỬ LÝ
                v_name = "TBA"
                v_voyage = "N/A"
                try:
                    # Truyền WebElement card_el vào JS để tránh lệch index DOM sau khi React render lại.
                    js_get_vessel = """
                    var card = arguments[0];
                    if(!card) return null;

                    var scope = card.querySelector('.el-tab-pane:not([style*="display: none"])') || card;
                    var lis = scope.querySelectorAll('li.more-infos.vessel, ul.route > li.more-infos.vessel, li[class*="vessel"]');
                    var firstVessel = null;

                    for(var i=0; i<lis.length; i++) {
                        var cls = (lis[i].className || '').toString();
                        if(cls.includes('vessel')) {
                            firstVessel = lis[i];
                            break;
                        }
                    }

                    if(!firstVessel) return null;

                    var result = {service: '', vessel: ''};

                    var dts = firstVessel.querySelectorAll('dt');
                    for(var m=0; m<dts.length; m++) {
                        var dtText = (dts[m].innerText || dts[m].textContent || '').trim().toUpperCase();
                        var dd = dts[m].nextElementSibling;
                        while(dd && dd.tagName && dd.tagName.toUpperCase() !== 'DD') {
                            dd = dd.nextElementSibling;
                        }
                        var val = dd ? (dd.innerText || dd.textContent || '').trim() : '';
                        if(dtText === 'VESSEL' && val && !result.vessel) {
                            result.vessel = val;
                        } else if(dtText === 'SERVICE' && val && !result.service) {
                            result.service = val;
                        }
                    }

                    if(!result.vessel) {
                        var text = (firstVessel.innerText || firstVessel.textContent || '').trim();
                        var mVessel = text.match(/VESSEL\\s*\\n\\s*([^\\n]+)/i);
                        var mService = text.match(/SERVICE\\s*\\n\\s*([^\\n]+)/i);
                        if(mVessel) result.vessel = mVessel[1].trim();
                        if(mService) result.service = mService[1].trim();
                    }

                    return result;
                    """
                    
                    data = driver.execute_script(js_get_vessel, card_el)
                    if data and data.get('vessel'):
                        v_name = data['vessel']
                        v_voyage = data.get('service', 'N/A')
                        print(f"      ✅ Tàu chính: {v_name} ({v_voyage})")
                    else:
                        card_text = driver.execute_script(
                            "return arguments[0].innerText || arguments[0].textContent || '';",
                            card_el
                        )
                        m_vessel = re.search(
                            r"\bVessel\s+(.+?)(?:\s+\d+(?:\.\d+)?\s*CO2|CO2e|TEU|\n|$)",
                            card_text,
                            re.I
                        )
                        if m_vessel:
                            v_name = " ".join(m_vessel.group(1).split())
                            print(f"      ✅ Tàu chính fallback: {v_name} ({v_voyage})")
                        else:
                            print(f"      ⚠️ Không tìm thấy vessel qua JS trong card hiện tại")
                        
                except Exception as e:
                    print(f"      ⚠️ Lỗi bóc vessel JS: {type(e).__name__}")

                vessel_entries.append(
                    f"{v_name} ({v_voyage}) / ETD: {opt['date'].day}-{opt['date'].strftime('%b')}"
                    f" / Transit time: {opt['transit']} Days / Transshipment: {opt['ts_port']}")

            except Exception as e:
                print(f"      ⚠️ Không thể bóc Vessel: {e}")
                vessel_entries.append(
                    f"TBA / ETD: {opt['date'].day}-{opt['date'].strftime('%b')}"
                    f" / Transit time: {opt['transit']} Days / Transshipment: {opt['ts_port']}")

            ts_entries.append(opt['ts_port'])

        # ══════════════════════════════════════════════════════════════════════════════
        # BƯỚC 3: GHI VÀO EXCEL
        # ══════════════════════════════════════════════════════════════════════════════

        # Cột F/G/H: Giá 20GP / 40GP / 40HQ
        ws.cell(row=row_index, column=6).value = formula_20 if total_20 else "Check"
        if price_mode == "20_ONLY":
            ws.cell(row=row_index, column=7).value = "-"
            ws.cell(row=row_index, column=8).value = "-"
        else:
            forty_value = formula_40hc if total_40hc else "Check"
            # CMA chỉ tick 40HQ; giá 40ST lấy bằng 40HQ theo rule vận hành.
            ws.cell(row=row_index, column=7).value = forty_value
            ws.cell(row=row_index, column=8).value = forty_value

        # Cột I: ETD
        def fmt_date(d):
            return f"{d.day}-{d.strftime('%b')}"

        final_dates = [o['date'] for o in valid_opts]
        if len(final_dates) == 1:
            final_etd_str = fmt_date(final_dates[0])
        elif len(final_dates) == 2:
            final_etd_str = f"{fmt_date(final_dates[0])} & {fmt_date(final_dates[1])}"
        elif len(final_dates) >= 3:
            if all(d.month == final_dates[0].month for d in final_dates):
                final_etd_str = (f"{final_dates[0].day}, {final_dates[1].day},"
                                 f" {fmt_date(final_dates[2])}")
            else:
                final_etd_str = " & ".join(fmt_date(d) for d in final_dates)
        else:
            final_etd_str = "N/A"
        ws.cell(row=row_index, column=9).value = final_etd_str

        # Cột J: Transit
        transits = [o['transit'] for o in valid_opts]
        ws.cell(row=row_index, column=10).value = (str(min(transits))
            if min(transits) == max(transits) else f"{min(transits)}-{max(transits)}")

        # Cột K: Valid
        ws.cell(row=row_index, column=11).value = calculate_cma_validity(valid_opts[-1]['date'])

        # Cột M: Remark
        # Lấy POD từ file Excel (cột D) để định vị khu vực
        pod_text = str(ws.cell(row=row_index, column=4).value or "").strip().upper()
        pod_country_text = str(ws.cell(row=row_index, column=2).value or "").strip().upper()
        manifest_acronym = get_manifest_code(pod_country_text, pod_text)

        # Cột M: Remark
        subject_items = ["BILL", "SEAL", "TLX"]
        if not othc_included:
            subject_items.insert(0, "THC")
        remark = "SUBJECT TO " + ", ".join(subject_items)
        if othc_included:
            remark = "INCLUDED O.THC, " + remark
        
        if manifest_fee_found and manifest_acronym:
            remark += f", {manifest_acronym}"
        if ows_found:
            remark += ", OWS"
            
        ws.cell(row=row_index, column=13).value = remark

        # Cột N: Free Time
        ws.cell(row=row_index, column=14).value = free_time_pod

        # Cột O: Vessel (xuống dòng)
        ws.cell(row=row_index, column=15).value = "\n".join(vessel_entries)
        ws.cell(row=row_index, column=15).alignment = openpyxl.styles.Alignment(wrapText=True)

        # Cột P: Transshipment (xuống dòng)
        unique_ts = []
        for ts in ts_entries:
            if ts not in unique_ts: unique_ts.append(ts)
        ws.cell(row=row_index, column=16).value = " or\n".join(unique_ts)
        ws.cell(row=row_index, column=16).alignment = openpyxl.styles.Alignment(wrapText=True)

        try:
            wb.save(excel_path)
        except PermissionError:
            print("      ❌ LỖI GHI FILE: TẮT FILE EXCEL ĐI SẾP ƠI!!!")

        print(f"   [INFO] Đã trích xuất giá dòng {row_index} xong!")
        return "SUCCESS"

    except Exception as e:
        print(f"      ❌ Lỗi Scrape chi tiết tổng quát: {e}")
        ws.cell(row=row_index, column=6).value = "Error"
        try: wb.save(excel_path)
        except: pass
        return "ERROR"

# ===================================================================================
# --- 4. MAIN LOOP ---
# ===================================================================================
# Danh sách nhận diện các cảng/quốc gia thuộc Đông Á và Đông Nam Á
INTRA_ASIA_KEYWORDS = [
    "SHANGHAI", "NINGBO", "QINGDAO", "XINGANG", "TIANJIN", "YANTIAN", "SHEKOU", "NANSHA", "XIAMEN", "DALIAN",
    "HONG KONG", "BUSAN", "INCHEON", "KWANGYANG", "TOKYO", "YOKOHAMA", "NAGOYA", "OSAKA", "KOBE", "MOJI", "HAKATA",
    "KEELUNG", "KAOHSIUNG", "TAICHUNG", "MANILA", "CEBU", "SUBIC", "BATANGAS", "DAVAO",
    "PORT KELANG", "PORT KLANG", "PENANG", "PASIR GUDANG", "TANJUNG PELEPAS", "KUANTAN",
    "JAKARTA", "SURABAYA", "SEMARANG", "BELAWAN", "PANJANG",
    "BANGKOK", "LAEM CHABANG", "LAT KRABANG", "SONGKHLA", "YANGON", "SIHANOUKVILLE", "PHNOM PENH",
    # FIX: bổ sung Myanmar và các alias
    "MYANMAR", "RANGOON", "THILAWA",
    "CHINA", "JAPAN", "KOREA", "TAIWAN", "MALAYSIA", "THAILAND", "INDONESIA", "PHILIPPINES", "SINGAPORE",
    "MYANMAR", "VIETNAM", "CAMBODIA",  # quốc gia gần VN
]

def extract_row_data(ws, r_idx):
    return {c: ws.cell(row=r_idx, column=c).value for c in range(6, 17)}

def write_row_data(ws, r_idx, data):
    for c, val in data.items():
        ws.cell(row=r_idx, column=c).value = val

def execute_single_search(row_index, pol_val, pod, is_riyadh, is_intra_hcm, pol_excel):
    global is_first_run_in_session, previous_pol
    wait = WebDriverWait(driver, 15)

    def do_search_steps(allow_full_route_recovery=True, price_mode="20_40HC"):
        if is_riyadh: handle_pod_selection_popup(prefer_port="JEDDAH")
        # Fix: RAMP luôn cần Vũng Tàu, không phân biệt Intra hay không
        if "RAMP" in pol_val or (not is_intra_hcm and "VNSGN" in pol_val):
            select_vung_tau_mandatory()

        # === CHECK BANNER NO ROUTE ===
        time.sleep(0.5) # Nghỉ nửa nhịp cho web render banner
        try:
            alerts = driver.find_elements(By.CSS_SELECTOR, "span.el-alert__title")
            for alert in alerts:
                if alert.is_displayed() and "SpotOn hasn't found possible route" in alert.text:
                    print("      ⚠️ Phát hiện banner: No Route. Bỏ qua luôn!")
                    ws.cell(row=row_index, column=6).value = "No Route"
                    try: wb.save(excel_path)
                    except: pass
                    return "NO_ROUTE"
        except:
            pass
        # =============================

        pick_date_plus_7()
        if is_first_run_in_session:
            select_containers(price_mode)
        else:
            ensure_containers_filled(price_mode)

        print("   -> Ép React khóa dữ liệu và Recheck form...")
        driver.execute_script("document.body.click();")
        time.sleep(0.3) 
        js_check_form = """
        var txt = "";
        var els = document.querySelectorAll('.selected-value-wrapper, input');
        for(var i=0; i<els.length; i++) txt += (els[i].innerText || els[i].value || "").toUpperCase() + " ";
        return txt;
        """
        form_text = driver.execute_script(js_check_form)
        pod_check = pod.split(",")[0].strip().upper()
        
        if pod_check not in form_text:
            print(f"      ⚠️ Phát hiện form chưa nhận POD [{pod_check}]! Đang bơm lại...")
            smart_update_field('POD', pod)
            if "RAMP" in pol_val or (not is_intra_hcm and "VNSGN" in pol_val): select_vung_tau_mandatory()
            time.sleep(0.5)
            driver.execute_script("document.body.click();")
            time.sleep(0.3)
        print("   -> Kiểm tra Commodity trước khi Get Quote...")
        commodity_val = get_cma_commodity_value()
        if not commodity_val:
            print("      ⚠️ Commodity TRỐNG! Đang chọn lại trước khi Get Quote...")
            if not select_commodity_refresh() or not get_cma_commodity_value():
                print("      ❌ Form chưa nhận Commodity; không bấm Get Quote để tránh ghi nhầm Sold Out.")
                mark_cma_form_error(row_index, ws)
                return "FORM_ERROR"
        else:
            print(f"      -> [OK] Commodity đã có: '{commodity_val}'")
        print("7. Get Quote...")
        btn = driver.find_element(By.XPATH, "//button[contains(., 'Get My Quote') or @id='SearchQuote']")
        driver.execute_script("arguments[0].click();", btn)

        print("   -> Đang chờ web load kết quả...")
        result_state = wait_cma_result_state(timeout=15)
        if result_state == "NO_OFFER":
            print("   ❌ CMA báo không có SpotOn offer -> bỏ qua nhanh, chuyển row tiếp theo.")
            mark_cma_no_offer(row_index, ws)
            return "NO_ROUTE"
        if result_state == "NO_ROUTE":
            print("   ❌ CMA báo không tìm thấy tuyến -> bỏ qua nhanh, chuyển row tiếp theo.")
            mark_cma_no_offer(row_index, ws, "No Route")
            return "NO_ROUTE"
        if result_state != "CARDS":
            print("   ⚠️ Chưa thấy card hoặc màn no-offer sau 15s, chuyển sang bước kiểm tra fallback...")

        page_text = driver.execute_script("return document.body.innerText;").upper()
        if pod_check not in page_text[:3000]:
            print(f"   ⚠️ BÁO ĐỘNG LỆCH ROUTE! Không thấy {pod_check}. Đang sửa lại...")
            if perform_modify_search_action():
                smart_update_field('POD', pod)
                if is_riyadh: handle_pod_selection_popup("JEDDAH")
                if not is_intra_hcm and "VNSGN" in pol_val: select_vung_tau_mandatory()
                select_commodity_refresh()
                configure_cma_containers(price_mode)
                btn = driver.find_element(By.XPATH, "//button[contains(., 'Get My Quote') or @id='SearchQuote']")
                driver.execute_script("arguments[0].click();", btn)
                result_state = wait_cma_result_state(timeout=15)
                if result_state == "NO_OFFER":
                    print("   ❌ CMA báo không có SpotOn offer sau khi sửa route -> bỏ qua nhanh.")
                    mark_cma_no_offer(row_index, ws)
                    return "NO_ROUTE"
                if result_state == "NO_ROUTE":
                    print("   ❌ CMA báo không tìm thấy tuyến sau khi sửa route -> bỏ qua nhanh.")
                    mark_cma_no_offer(row_index, ws, "No Route")
                    return "NO_ROUTE"
                if result_state != "CARDS":
                    print("   ⚠️ Chưa thấy card hoặc màn no-offer sau khi sửa route, để scrape fallback xử lý...")
            else:
                if not allow_full_route_recovery:
                    print("   ❌ Vẫn lệch route sau khi reload form gốc, bỏ qua row.")
                    return False

                print("   ⚠️ Không mở được Modify Search. Reload form gốc và nhập lại toàn bộ row...")
                try:
                    inputs = ensure_cma_quote_form_inputs(
                        context="full route recovery",
                        force_navigate=True,
                    )
                    if not inputs:
                        return False
                    select_port_full(inputs[0], pol_val)
                    sleep_human(0.5, 0.8)
                    inputs = _cma_get_quote_port_inputs(driver, visible_only=False)
                    if len(inputs) < 2:
                        return False
                    select_port_full(inputs[-1], pod)
                    return do_search_steps(allow_full_route_recovery=False, price_mode=price_mode)
                except Exception as exc:
                    print(f"   ❌ Full recovery route thất bại: {type(exc).__name__}: {exc}")
                    return False

        scrape_state = scrape_multi_etd_and_save(row_index, ws, pol_excel, price_mode=price_mode)
        if scrape_state == "RETRY_20_ONLY" and price_mode != "20_ONLY":
            print("   -> CMA thiếu offer 40HQ. Modify Search, bỏ 40HQ và retry chỉ 20GP...")
            if not perform_modify_search_action():
                print("   ❌ Không mở được Modify Search để retry 20-only.")
                mark_cma_no_offer(row_index, ws)
                return "NO_OFFER"

            configure_cma_containers("20_ONLY")
            try:
                commodity_val = driver.execute_script("""
                var inputs = document.querySelectorAll('input');
                for(var i=0; i<inputs.length; i++) {
                    var ph = (inputs[i].placeholder || '').toLowerCase();
                    if(ph.includes('commodity')) return (inputs[i].value || '').trim();
                }
                return '';
                """)
                if not commodity_val:
                    select_commodity_refresh()
            except:
                pass

            btn = driver.find_element(By.XPATH, "//button[contains(., 'Get My Quote') or @id='SearchQuote']")
            driver.execute_script("arguments[0].click();", btn)
            result_state = wait_cma_result_state(timeout=15)
            if result_state == "NO_OFFER":
                print("   ❌ Retry 20-only vẫn No Offer.")
                mark_cma_no_offer(row_index, ws)
                return "NO_OFFER"
            if result_state == "NO_ROUTE":
                print("   ❌ Retry 20-only báo No Route.")
                mark_cma_no_offer(row_index, ws, "No Route")
                return "NO_ROUTE"
            if result_state != "CARDS":
                print("   ⚠️ Retry 20-only chưa thấy card rõ ràng, vẫn thử scrape fallback...")
            return scrape_multi_etd_and_save(row_index, ws, pol_excel, price_mode="20_ONLY")

        return scrape_state

    if is_first_run_in_session:
        print(">> FULL MODE (Lần đầu)...")
        inputs = ensure_cma_quote_form_inputs(context=f"row {row_index}")
        if not inputs:
            ws.cell(row=row_index, column=6).value = "Error"
            try:
                wb.save(excel_path)
            except Exception:
                pass
            return False
        select_port_full(inputs[0], pol_val)
        sleep_human(0.5, 0.8)
        inputs = _cma_get_quote_port_inputs(driver, visible_only=False)
        if len(inputs) < 2:
            ws.cell(row=row_index, column=6).value = "Error"
            try:
                wb.save(excel_path)
            except Exception:
                pass
            return False
        select_port_full(inputs[-1], pod)
        
        status = do_search_steps()
        is_first_run_in_session = False
        previous_pol = pol_val
        if status == "NO_ROUTE": return False
    else:
        print(">> MODIFY MODE...")
        if perform_modify_search_action():
            if pol_val != previous_pol: smart_update_field('POL', pol_val)
            smart_update_field('POD', pod)
            
            status = do_search_steps()
            if status == "NO_ROUTE":
                previous_pol = pol_val
                return False
                
            if status in {"SOLD_OUT", "FORM_ERROR"}:
                reason = "form/Commodity chưa sẵn sàng" if status == "FORM_ERROR" else "Sold Out"
                print(f"   -> ⚠️ Modify Mode bị {reason}; recheck chính row bằng FULL MODE...")
                inputs = ensure_cma_quote_form_inputs(
                    context=f"{status.lower()} retry row {row_index}",
                    force_navigate=True,
                )
                if not inputs:
                    return False
                select_port_full(inputs[0], pol_val)
                sleep_human(0.5, 0.8)
                inputs = _cma_get_quote_port_inputs(driver, visible_only=False)
                if len(inputs) < 2:
                    return False
                select_port_full(inputs[-1], pod)
                
                retry_status = do_search_steps(allow_full_route_recovery=False)
                if retry_status in {"NO_ROUTE", "FORM_ERROR"}:
                    previous_pol = pol_val
                    return False
            previous_pol = pol_val
        else:
            print("   -> Lỗi Modify! Chuyển FULL MODE...")
            inputs = ensure_cma_quote_form_inputs(
                context=f"modify fallback row {row_index}",
                force_navigate=True,
            )
            if not inputs:
                return False
            select_port_full(inputs[0], pol_val)
            sleep_human(0.5, 0.8)
            inputs = _cma_get_quote_port_inputs(driver, visible_only=False)
            if len(inputs) < 2:
                return False
            select_port_full(inputs[-1], pod)
            
            status = do_search_steps()
            previous_pol = pol_val
            if status == "NO_ROUTE": return False

# ✅ LẤY ROW TỪ ARGUMENT TERMINAL
import sys
target_row = None
single_row_env = os.environ.get("SINGLE_ROW", "").strip()
if single_row_env:
    try:
        target_row = int(single_row_env)
        print(f"[INFO] Chạy riêng row {target_row} từ SINGLE_ROW...")
    except ValueError:
        print(f"[WARN] SINGLE_ROW không hợp lệ: {single_row_env}. Chạy tất cả row.")
if len(sys.argv) > 1:
    try:
        target_row = int(sys.argv[1])
        print(f"[INFO] Chạy riêng row {target_row}...")
    except ValueError:
        print(f"[WARN] Argument không hợp lệ: {sys.argv[1]}. Chạy tất cả row.")

try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    is_first_run_in_session = True
    previous_pol = ""

    # Port mapping: Excel name → carrier search name
    CMA_PORT_MAPPING = {
        "TIANJIN": "XINGANG",
        "SOKHNA": "AIN SUKHNA",
        "MANZANILLO": "MANZANILLO, MX",
        "NAPLES": "NAPOLI",
        "VENICE": "VENEZIA",
    }

    total_valid_rows = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if target_row and i != target_row: continue
        pol_excel = str(row[2] or "").strip()
        pod       = str(row[3] or "").strip()
        carrier   = str(row[4] or "").strip().upper()
        if not pol_excel or not pod: continue
        if carrier not in {"CMA", "ANL", "CNC", "APL"}: continue
        if FILTER_POL and pol_excel.upper() != FILTER_POL: continue
        if FILTER_POD and pod.upper() != FILTER_POD: continue
        total_valid_rows += 1
    print(f"Tổng cộng có {total_valid_rows} dòng cần check.")

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # ✅ FILTER THEO TARGET ROW
        if target_row and i != target_row:
            continue
        country   = str(row[1] or "").strip()
        pol_excel = str(row[2] or "").strip()
        pod       = str(row[3] or "").strip()
        carrier   = str(row[4] or "").strip().upper()

        if not pol_excel or not pod: continue
        if carrier not in {"CMA", "ANL", "CNC", "APL"}: continue
        if FILTER_POL and pol_excel.upper() != FILTER_POL: continue
        if FILTER_POD and pod.upper() != FILTER_POD: continue

        # Áp dụng port mapping cho POD (giữ pod_excel_orig để ghi Excel)
        pod = CMA_PORT_MAPPING.get(pod.upper(), pod)

        # Phân loại logic
        is_hcm = "HO CHI MINH" in pol_excel.upper() or "VNSGN" in pol_excel.upper()
        is_intra_asia = any(k in pod.upper() for k in INTRA_ASIA_KEYWORDS)
        is_riyadh = 'RIYADH' in pod.upper()
        
        # Thêm danh sách từ khóa nhận diện các cảng Trung Quốc
        CHINA_KEYWORDS = ["SHANGHAI", "NINGBO", "QINGDAO", "XINGANG", "TIANJIN", "YANTIAN", "SHEKOU", "NANSHA", "XIAMEN", "DALIAN", "CHINA"]
        is_china = is_china_destination(country, pod)

        print(f"\n==========================================")
        print(f"--- DÒNG {i}: {pol_excel} -> {pod} | Hãng: {carrier} ---")

        if "Access Denied" in driver.title:
            print("!!! BỊ CHẶN !!!")
            break

        # Nếu là Intra-Asia từ HCM NHƯNG KHÔNG PHẢI TRUNG QUỐC -> Check cả PORT và RAMP
        if is_hcm and is_intra_asia and not is_china:
            print(f"   [!] TUYẾN INTRA-ASIA TỪ HCM: Cần check cả PORT và RAMP để so giá.")
            
            # CHECK LẦN 1: PORT
            print("\n   >>> CHECK OPTION 1: VNSGN - PORT <<<")
            execute_single_search(i, "VNSGN-PORT", pod, is_riyadh, True, pol_excel)
            data_port = extract_row_data(ws, i)
            
            price_port = parse_cma_comparable_price(data_port.get(6))

            # CHECK LẦN 2: RAMP
            print("\n   >>> CHECK OPTION 2: VNSGN - RAMP <<<")
            execute_single_search(i, "VNSGN-RAMP", pod, is_riyadh, True, pol_excel)
            data_ramp = extract_row_data(ws, i)
            price_ramp = parse_cma_comparable_price(data_ramp.get(6))

            # SO SÁNH GIÁ VÀ GHI LẠI
            print(f"\n   >>> KẾT QUẢ SO SÁNH DÒNG {i}: PORT ({data_port.get(6)}) vs RAMP ({data_ramp.get(6)})")
            if price_port <= price_ramp and price_port != float('inf'):
                print("   => PORT RẺ HƠN HOẶC BẰNG! Ghi đè lại dữ liệu PORT vào Excel.")
                write_row_data(ws, i, data_port)
            elif price_ramp < price_port:
                print("   => RAMP RẺ HƠN! Đã giữ lại dữ liệu RAMP trên Excel.")
            else:
                print("   => CẢ 2 ĐỀU LỖI HOẶC SOLD OUT. Trả về báo cáo PORT.")
                write_row_data(ws, i, data_port)
            
            try: wb.save(excel_path)
            except: pass
            
        else:
            # Các trường hợp chạy 1 lần:
            # 1. Trung Quốc từ HCM -> Bắt buộc check VNSGN-PORT
            # 2. Các cảng như Hải Phòng, hoặc tuyến xa -> Giữ nguyên logic cũ
            if is_hcm:
                pol = "VNSGN-PORT" if is_china else "VNSGN"
            else:
                pol = "VNHPH" if "HAI PHONG" in pol_excel.upper() else pol_excel
                
            if is_hcm and is_china:
                print(f"   [!] TUYẾN TRUNG QUỐC TỪ HCM: Chỉ check VNSGN (PORT).")
                
            execute_single_search(i, pol, pod, is_riyadh, False, pol_excel)

        print(f"   [OK] Đã lưu dữ liệu dòng {i} thành công!")
        print(f"   -> Nghỉ {CMA_ROW_SLEEP_MIN:g}-{CMA_ROW_SLEEP_MAX:g}s để chuyển dòng...")
        sleep_human(CMA_ROW_SLEEP_MIN, CMA_ROW_SLEEP_MAX)

except Exception as e:
    print(f"Lỗi chung vòng lặp: {e}")
    raise
