import requests
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import calendar
import openpyxl
import os
import time
import traceback
import random  
import subprocess
import socket
import sys
import io
import re
import math
from urllib.parse import urlparse
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from bot_runtime_utils import switch_to_live_window
from remark_rules import build_subject_remark, is_china_destination
from cosco_elines_ui import (
    dropdown_has_selected_option,
    is_new_elines_result_card_text,
    is_elines_auth_page,
    is_elines_booking_page,
    is_no_matching_ocean_freight_message,
    parse_new_elines_card_schedule,
    parse_premium_service_row,
    select_preferred_premium_service,
)
from cosco_result_logic import clear_cosco_quote_fields

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')



# ===================================================================================
# --- 1. SETUP HỆ THỐNG & API & BÙA TÀNG HÌNH ---
# ===================================================================================
current_folder = os.getcwd()
driver_path = os.path.join(current_folder, "msedgedriver.exe")
DATE_OFFSET_DAYS = parse_date_offset_days()
COSCO_PORT_INPUT_WAIT_SECONDS = int(os.environ.get("COSCO_PORT_INPUT_WAIT_SECONDS", "15"))
COSCO_PORT_DROPDOWN_WAIT_SECONDS = int(os.environ.get("COSCO_PORT_DROPDOWN_WAIT_SECONDS", "15"))

def _excel_formula_from_parts(parts):
    tokens = []
    for raw in parts:
        try:
            amount = float(raw)
        except (TypeError, ValueError):
            continue
        if abs(amount) < 1e-9:
            continue
        sign = "-" if amount < 0 else "+"
        amount = abs(amount)
        if abs(amount - round(amount)) < 1e-9:
            text = str(int(round(amount)))
        else:
            text = f"{amount:.2f}".rstrip("0").rstrip(".")
        tokens.append((sign if tokens else ("-" if sign == "-" else "")) + text)
    return "=" + "".join(tokens) if tokens else None

def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('127.0.0.1', port)) == 0

def wait_for_port(port, timeout=8):
    """FIX: Thay time.sleep(3) cứng bằng poll port → quay lại ngay khi Edge sẵn sàng."""
    end = time.time() + timeout
    while time.time() < end:
        if is_port_in_use(port):
            return True
        time.sleep(0.2)
    return False

# Chỉ gọi lệnh mở Edge nếu Port 9523 chưa có ai xài
if not is_port_in_use(9523):
    print("[HỆ THỐNG] Edge COSCO chưa mở. Đang tự động khởi động...")
    try:
        subprocess.Popen([
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            "--remote-debugging-port=9523",
            r"--user-data-dir=C:\edge_cosco",
            "--disable-background-timer-throttling",
            "--disable-renderer-backgrounding",
            "--disable-backgrounding-occluded-windows",
            "--start-maximized"
        ])
        wait_for_port(9523, timeout=8)
    except: pass
else:
    print("[HỆ THỐNG] Edge COSCO đã mở sẵn. Bỏ qua lệnh khởi động trình duyệt.")

import threading
from selenium.common.exceptions import TimeoutException

edge_options = Options()
edge_options.add_experimental_option("debuggerAddress", "127.0.0.1:9523")
service = Service(executable_path=driver_path)

# Wrap webdriver.Edge() trong thread timeout + RETRY d? trnh treo v h?n
driver = None
for attempt in range(1, 4):
    _driver_holder = {"driver": None, "error": None}
    def _create_session():
        try:
            _driver_holder["driver"] = webdriver.Edge(service=service, options=edge_options)
        except Exception as ex:
            _driver_holder["error"] = ex

    _t = threading.Thread(target=_create_session, daemon=True)
    _t.start()
    _t.join(timeout=30)
    
    if _t.is_alive() or _driver_holder["driver"] is None:
        print(f"[WARN] webdriver.Edge() b? treo (attempt {attempt}/3)...")
        if attempt == 3:
            if _driver_holder["error"]:
                raise _driver_holder["error"]
            raise TimeoutException("webdriver.Edge() treo qu 30s sau 3 l?n th?")
        continue # Th? l?i
        
    driver = _driver_holder["driver"]
    break # Thnh cng
try:
    driver.maximize_window()
except Exception:
    try:
        info = driver.execute_cdp_cmd("Browser.getWindowForTarget", {})
        driver.execute_cdp_cmd("Browser.setWindowBounds", {
            "windowId": info["windowId"],
            "bounds": {"windowState": "maximized"}
        })
    except Exception:
        pass

# Bùa tàng hình trị WAF
stealth_script = """
Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
window.navigator.chrome = { runtime: {} };
Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] });
"""
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": stealth_script
})

# ===================================================================================
# CẤU HÌNH LOGIN COSCO (UI hiện tại - 2026)
# ===================================================================================
COSCO_EMAIL = os.environ.get("COSCO_EMAIL", "celine@pio-logistics.vn")
# Password lấy từ ENV var COSCO_PASSWORD (không hard-code).
# Nếu không set, bot sẽ chờ user nhập tay password + giải captcha trên trang SSO.
COSCO_PASSWORD = os.environ.get("COSCO_PASSWORD", "")

# Thời gian tối đa chờ user giải slider captcha trên trang Keycloak SSO (giây)
COSCO_CAPTCHA_WAIT = int(os.environ.get("COSCO_CAPTCHA_WAIT", "180"))
try:
    ELINES_RESULT_WAIT_SECONDS = max(30, int(os.environ.get("ELINES_RESULT_WAIT_SECONDS", "75")))
except ValueError:
    ELINES_RESULT_WAIT_SECONDS = 75
try:
    ELINES_EMPTY_CONFIRM_SECONDS = max(5, int(os.environ.get("ELINES_EMPTY_CONFIRM_SECONDS", "20")))
except ValueError:
    ELINES_EMPTY_CONFIRM_SECONDS = 20
ELINES_NO_PRODUCTS = object()
ELINES_NO_PREMIUM_PRICE = object()
SYNCONHUB_NO_SERVICE = object()
COSCO_UNSUPPORTED_PORTS = {"PARADIP"}
ELINES_BOOKING_URL = "https://elines.coscoshipping.com/ebusiness/bookingrequest/"

# New E-Lines booking-request form (inside #aczoneIframe).
ELINES_ORIGIN_XPATH = "/html/body/div[1]/div/div/div/div/div/div[2]/div/div/div[1]/form/div/div[1]/div/div/div[1]/div/div/input"
ELINES_DESTINATION_XPATH = "/html/body/div[1]/div/div/div/div/div/div[2]/div/div/div[1]/form/div/div[3]/div/div/div[1]/div/div/div/input"


class COSCOLoginRequired(RuntimeError):
    """E-Lines redirected to login or its authenticated form disappeared."""



def is_elines_no_products_result(result):
    return result is ELINES_NO_PRODUCTS or result is ELINES_NO_PREMIUM_PRICE


def is_no_service_result(result):
    return (
        result is ELINES_NO_PRODUCTS
        or result is ELINES_NO_PREMIUM_PRICE
        or result is SYNCONHUB_NO_SERVICE
    )


def _cosco_port_key(value):
    return re.sub(r"\s+", " ", str(value or "").strip().upper())


def cosco_has_unsupported_port(*ports):
    return any(_cosco_port_key(port) in COSCO_UNSUPPORTED_PORTS for port in ports)


def _wait_url_leave(driver, host_keyword, timeout):
    """Poll URL hiện tại của tab, return True khi URL KHÔNG còn chứa host_keyword.
    FIX: poll 0.5s thay vì 1s + print tiến trình mỗi 15s để user biết bot không treo."""
    end = time.time() + timeout
    last_log = time.time()
    while time.time() < end:
        try:
            if host_keyword not in (driver.current_url or "").lower():
                return True
        except Exception:
            pass
        if time.time() - last_log > 15:
            remaining = int(end - time.time())
            print(f"   ⏳ Đang chờ rời SSO ({host_keyword}) ... còn ~{remaining}s")
            last_log = time.time()
        time.sleep(0.5)
    return False


def _is_logged_in_elines(driver):
    """True when the authenticated new booking-request page has loaded."""
    try:
        u = (driver.current_url or "").lower()
        return is_elines_booking_page(u) and not is_elines_auth_page(u)
    except Exception:
        return False

def _url_matches_target_path(current_url, target_url):
    try:
        cur = urlparse(current_url or "")
        tgt = urlparse(target_url or "")
        if tgt.netloc and tgt.netloc.lower() not in cur.netloc.lower():
            return False
        target_path = (tgt.path or "").rstrip("/")
        if not target_path:
            return True
        return (cur.path or "").rstrip("/").lower() == target_path.lower()
    except Exception:
        return False

def _switch_to_elines_booking_frame(timeout=20):
    """Find the booking form in either the current document or any live iframe."""

    def form_present():
        selectors = (
            (By.XPATH, ELINES_ORIGIN_XPATH),
            (By.XPATH, "//input[@placeholder='Please input Origin City']"),
            (By.XPATH, "//input[contains(@placeholder,'Origin')]"),
            (By.XPATH, "//button[contains(normalize-space(.),'Search Service')]"),
        )
        for by, value in selectors:
            try:
                if any(_elines_visible(el) for el in driver.find_elements(by, value)):
                    return True
            except Exception:
                continue
        return False

    deadline = time.time() + timeout
    last_frame_count = 0
    while time.time() < deadline:
        driver.switch_to.default_content()
        if is_elines_auth_page(driver.current_url):
            raise COSCOLoginRequired(f"E-Lines cần login: {driver.current_url}")

        if form_present():
            return "TOP"

        frames = driver.find_elements(By.TAG_NAME, "iframe")
        last_frame_count = len(frames)
        preferred = sorted(
            frames,
            key=lambda frame: 0 if (frame.get_attribute("id") or "") == "aczoneIframe" else 1,
        )
        for frame in preferred:
            try:
                driver.switch_to.default_content()
                driver.switch_to.frame(frame)
                if form_present():
                    return "IFRAME"
            except Exception:
                continue
        time.sleep(0.4)

    driver.switch_to.default_content()
    if is_elines_auth_page(driver.current_url):
        raise COSCOLoginRequired(f"E-Lines cần login: {driver.current_url}")
    raise TimeoutException(f"E-Lines booking form chưa render; iframe_count={last_frame_count}")

def _is_elines_booking_form_ready(timeout=3):
    try:
        _switch_to_elines_booking_frame(timeout=timeout)
        return True
    except Exception:
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        return False

def _rebuild_elines_tab(reason=""):
    print(f"      -> [Elines] Rebuild tab booking ({reason})...")
    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    new_tab = _replace_current_tab(driver, ELINES_BOOKING_URL)
    try:
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
    except Exception:
        pass
    return new_tab

def _clear_cosco_session(browser, reason=""):
    print(f"   [COSCO] Logout/clear session do: {reason}")
    clear_urls = [
        "https://elines.coscoshipping.com/ebusiness/",
        "https://synconhub.coscoshipping.com/spot",
        "https://exiamfw.lines.coscoshipping.com/",
    ]
    for url in clear_urls:
        try:
            browser.get(url)
            try:
                WebDriverWait(browser, 6).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
            except Exception:
                pass
            try:
                browser.execute_script("localStorage.clear(); sessionStorage.clear();")
            except Exception:
                pass
            try:
                browser.delete_all_cookies()
            except Exception:
                pass
        except Exception as e:
            print(f"   [COSCO] Không clear được {url}: {type(e).__name__}")
    time.sleep(1)

def _force_cosco_relogin(reason=""):
    print(f"      -> [Elines] Cookie/session có vẻ hỏng ({reason}). Logout rồi đăng nhập lại...")
    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    _clear_cosco_session(driver, reason)
    login_cosco(driver)
    focus_tab_by_url("elines.coscoshipping.com", ELINES_BOOKING_URL)
    return _is_elines_booking_form_ready(timeout=25)


def _is_logged_in_synconhub(driver):
    """True nếu URL trên synconhub VÀ không phải trang home với ?redirect=, không phải SSO."""
    try:
        u = (driver.current_url or "").lower()
        if "synconhub.coscoshipping.com" not in u:
            return False
        if "exiamfw" in u or "/auth/" in u or "login" in u.split("?", 1)[0].split("#", 1)[0]:
            return False
        # Check user menu element (tên user hiển thị ở góc phải)
        try:
            els = driver.find_elements(By.CSS_SELECTOR, ".user-menu, .user-profile, .username")
            for e in els:
                if (e.text or "").strip() and (e.is_displayed() if hasattr(e, "is_displayed") else True):
                    return True
        except Exception:
            pass
        return False
    except Exception:
        return False


def _open_or_focus_tab(driver, domain_keyword, target_url, label, timeout=15):
    domain_keyword = domain_keyword.lower()
    
    # 1. Bắt lỗi mất context window hiện tại (chống NoSuchWindowException)
    try:
        driver.current_url
    except Exception:
        try:
            if driver.window_handles:
                driver.switch_to.window(driver.window_handles[0])
        except Exception:
            pass

    original = None
    try:
        original = driver.current_window_handle
    except Exception:
        pass

    try:
        handles = list(driver.window_handles)
    except Exception:
        handles = []

    for handle in handles:
        try:
            driver.switch_to.window(handle)
            cur_url = driver.current_url or ""
            if domain_keyword in cur_url.lower() and _url_matches_target_path(cur_url, target_url):
                print(f"   ✅ {label}: da thay tab dung trang ({cur_url[:90]})")
                return handle
        except Exception:
            continue

    for handle in handles:
        try:
            driver.switch_to.window(handle)
            if domain_keyword in (driver.current_url or "").lower():
                print(f"   ✅ {label}: đã thấy tab sẵn ({driver.current_url[:90]})")
                return handle
        except Exception:
            continue

    print(f"   🌐 {label}: chưa có tab, Selenium tự mở tab mới...")
    try:
        before = set(driver.window_handles)
    except Exception:
        before = set()

    try:
        switch_to_live_window(driver, preferred_handle=original)
        driver.switch_to.new_window("tab")
    except Exception as e:
        print(f"   ⚠️ {label}: new_window lỗi ({type(e).__name__}), thử JS window.open...")
        try:
            if original:
                driver.switch_to.window(original)
        except Exception:
            try:
                if driver.window_handles:
                    driver.switch_to.window(driver.window_handles[0])
            except Exception:
                pass
                
        try:
            driver.execute_script("window.open('about:blank', '_blank');")
            time.sleep(1)
            new_handles = list(set(driver.window_handles) - before)
            if new_handles:
                driver.switch_to.window(new_handles[-1])
            else:
                driver.switch_to.window(driver.window_handles[-1])
        except Exception:
            pass

    try:
        handle = driver.current_window_handle
        
        # 2. Fix lỗi kẹt 31s: Dùng JS điều hướng ngầm để Selenium không bị đứng chờ
        driver.execute_script("setTimeout(function() { window.location.href = arguments[0]; }, 10);", target_url)
        
        # Chỉ cần thấy URL chuyển hướng là chạy tiếp ngay lập tức
        WebDriverWait(driver, timeout).until(
            lambda d: domain_keyword in (d.current_url or "").lower()
        )
        return handle
    except Exception:
        print(f"   ⚠️ {label}: tab đã mở nhưng chưa vào đúng domain.")
        try:
            return driver.current_window_handle
        except:
            return None


def _ensure_cosco_tab(driver, handle, domain_keyword, target_url, label):
    """Return a live COSCO tab, recreating a stale saved handle when needed."""
    try:
        if handle and handle in driver.window_handles:
            driver.switch_to.window(handle)
            return handle
    except Exception:
        pass

    print(f"   [RECOVERY] {label}: tab cũ đã mất, đang tìm/mở lại...")
    switch_to_live_window(driver)
    recovered = _open_or_focus_tab(
        driver,
        domain_keyword,
        target_url,
        label,
    )
    if not recovered:
        raise RuntimeError(f"COSCO không khôi phục được tab {label}")
    return recovered
        
def _replace_current_tab(driver, target_url="about:blank"):
    """Create a new live tab before closing the current one."""
    old_tab = None
    try:
        old_tab = driver.current_window_handle
    except Exception:
        pass

    try:
        driver.switch_to.new_window("tab")
    except Exception:
        try:
            handles = driver.window_handles
            if handles:
                driver.switch_to.window(handles[0])
        except Exception:
            pass
        driver.execute_script("window.open('about:blank', '_blank');")
        time.sleep(0.5)
        driver.switch_to.window(driver.window_handles[-1])

    new_tab = driver.current_window_handle
    if target_url:
        try:
            driver.get(target_url)
        except Exception:
            pass

    if old_tab and old_tab != new_tab:
        try:
            driver.switch_to.window(old_tab)
            driver.close()
        except Exception:
            pass
        driver.switch_to.window(new_tab)
    return new_tab


def _click_text(driver, tag, text, css_contains=None, timeout=4):
    """Find a visible element of given tag with exact innerText `text`, optionally
    requiring `css_contains` substring in class, then click via JS. Return True on success."""
    end = time.time() + timeout
    while time.time() < end:
        try:
            els = driver.find_elements(By.TAG_NAME, tag)
            for e in els:
                try:
                    t = (e.text or "").strip()
                    if t != text:
                        continue
                    if css_contains and css_contains not in (e.get_attribute("class") or ""):
                        continue
                    if not e.is_displayed():
                        continue
                    driver.execute_script("arguments[0].click();", e)
                    return True
                except Exception:
                    continue
        except Exception:
            pass
        time.sleep(0.3)
    return False


def _fill_input_by_placeholder(driver, placeholder_substr, value, css_class=None, timeout=6):
    """Find a visible <input> whose placeholder contains placeholder_substr (case-insens),
    optionally requiring css_class in class attr, then set value + fire input/change events."""
    end = time.time() + timeout
    needle = placeholder_substr.lower()
    while time.time() < end:
        try:
            ins = driver.find_elements(By.TAG_NAME, "input")
            for i in ins:
                try:
                    if not i.is_displayed():
                        continue
                    ph = (i.get_attribute("placeholder") or "").lower()
                    if needle not in ph:
                        continue
                    if css_class and css_class not in (i.get_attribute("class") or ""):
                        continue
                    driver.execute_script("""
                        const el = arguments[0]; const v = arguments[1];
                        const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                        setter.call(el, v);
                        el.dispatchEvent(new Event('input', {bubbles: true}));
                        el.dispatchEvent(new Event('change', {bubbles: true}));
                    """, i, value)
                    return True
                except Exception:
                    continue
        except Exception:
            pass
        time.sleep(0.3)
    return False


def _do_keycloak_login(driver, label):
    """On Keycloak SSO page (exiamfw.lines.coscoshipping.com): fill username+password,
    then wait for user to solve slider captcha + click Log In. Returns True if URL leaves SSO."""
    try:
        WebDriverWait(driver, 10).until(
            lambda d: "exiamfw" in (d.current_url or "").lower()
        )
    except Exception:
        # Already past SSO (e.g. cookie still valid) → nothing to do
        return True

    print(f"   [SSO] {label}: điền username + password vào Keycloak...")
    try:
        # Username (login_hint thường tự fill, nhưng cứ điền lại cho chắc)
        driver.execute_script("""
            const u = document.getElementById('username');
            const p = document.getElementById('password');
            if (u) {
                const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                setter.call(u, arguments[0]);
                u.dispatchEvent(new Event('input', {bubbles: true}));
            }
            if (p && arguments[1]) {
                const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                setter.call(p, arguments[1]);
                p.dispatchEvent(new Event('input', {bubbles: true}));
            }
        """, COSCO_EMAIL, COSCO_PASSWORD)
    except Exception as e:
        print(f"   [SSO] Không fill được username/password: {e}")

    if not COSCO_PASSWORD:
        print(f"   ⚠️ ENV COSCO_PASSWORD chưa set — anh nhập password tay đi.")

    print(f"   [SSO] {label}: chờ user kéo slider captcha + bấm Log In "
          f"(tối đa {COSCO_CAPTCHA_WAIT}s, thoát sớm khi URL rời SSO)...")
    ok = _wait_url_leave(driver, "exiamfw", COSCO_CAPTCHA_WAIT)
    if ok:
        print(f"   ✅ {label}: SSO xong, URL = {driver.current_url[:90]}")
    else:
        print(f"   ⚠️ {label}: hết {COSCO_CAPTCHA_WAIT}s mà vẫn còn trên SSO. Tiếp tục thử...")
    return ok


def login_cosco(driver):
    print("\n--- BẮT ĐẦU ĐĂNG NHẬP COSCO ---")

    # Mở 2 tab như yêu cầu
    elines_tab = _open_or_focus_tab(
        driver,
        "elines.coscoshipping.com",
        "https://elines.coscoshipping.com/ebusiness/",
        "Elines",
    )
    synconhub_tab = _open_or_focus_tab(
        driver,
        "synconhub.coscoshipping.com",
        "https://synconhub.coscoshipping.com/",
        "Synconhub",
    )

    # ==========================================
    # PHASE 1: ELINES
    # ==========================================
    print("1. Đang xử lý Elines...")
    elines_tab = _ensure_cosco_tab(
        driver,
        elines_tab,
        "elines.coscoshipping.com",
        ELINES_BOOKING_URL,
        "Elines",
    )
    
    current_url_lower = (driver.current_url or "").lower()

    if _is_logged_in_elines(driver):
        print("   ✅ Elines: Đang ở sẵn trang Booking, nhảy thẳng vào việc!")
        
    elif "dashboard" in current_url_lower:
        print("   ✅ Elines: Đang ở Dashboard, bẻ lái sang trang Booking...")
        driver.get(ELINES_BOOKING_URL)
        try:
            WebDriverWait(driver, 8).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except Exception:
            pass
        time.sleep(2)
        
    else:
        # Nếu chưa có thì bắt đầu quay về trang chủ điền form
        driver.get("https://elines.coscoshipping.com/ebusiness/")
        try:
            WebDriverWait(driver, 6).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except Exception:
            pass

        # 1) Cookie "Allow All"
        if _click_text(driver, "button", "Allow All", timeout=3):
            print("   [Elines] đã bấm Allow All.")
            time.sleep(1)

        # 2) Click Login
        clicked = False
        try:
            for a in driver.find_elements(By.CSS_SELECTOR, "a.loginreg_link"):
                if (a.text or "").strip() == "Login" and a.is_displayed():
                    driver.execute_script("arguments[0].click();", a)
                    clicked = True
                    break
        except Exception:
            pass
        if not clicked:
            _click_text(driver, "a", "Login", timeout=3)
        time.sleep(2)

        # 3) Nhập email
        if not _fill_input_by_placeholder(driver, "username", COSCO_EMAIL, css_class="ivu-input", timeout=6):
            _fill_input_by_placeholder(driver, "email", COSCO_EMAIL, timeout=2)
        time.sleep(0.5)

        # 4) Click "Next step"
        if not _click_text(driver, "button", "Next step", timeout=4):
            print("   [Elines] ⚠️ Không tìm được nút 'Next step'.")
        time.sleep(3)

        # 5) Bấm Log In trên trang SSO (nếu có)
        _click_text(driver, "button", "Log In", timeout=3)
        
        # 6) Tạm dừng script chờ bạn giải Captcha
        print("\n" + "="*70)
        captcha_msg = (
            "🛑 DỪNG LẠI: Vui lòng tự giải Captcha trên trình duyệt.\n"
            "👉 Sau khi giải xong và web CHUYỂN SANG TRANG DASHBOARD, hãy nhấn phím ENTER tại đây để bot chạy tiếp..."
        )
        try:
            input(captcha_msg)
        except EOFError:
            print(captcha_msg)
            print(f"   [COSCO] Đang chạy non-interactive, tự chờ rời SSO tối đa {COSCO_CAPTCHA_WAIT}s...")
            if not _wait_url_leave(driver, "exiamfw", COSCO_CAPTCHA_WAIT):
                raise Exception("COSCO login cần giải Captcha thủ công; chưa rời trang SSO.")
        print("="*70 + "\n")

        # 7) Always enter the new booking-request form after SSO succeeds.
        print("   [Elines] Đã nhận lệnh Enter, chuyển sang trang Booking Request...")
        driver.get(ELINES_BOOKING_URL)
        try:
            WebDriverWait(driver, 8).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except Exception:
            pass
        time.sleep(2)

    # ==========================================
    # PHASE 2: SYNCONHUB (Đã fix lỗi đồng bộ SSO)
    # ==========================================
    print("2. Đang xử lý Synconhub...")
    synconhub_tab = _ensure_cosco_tab(
        driver,
        synconhub_tab,
        "synconhub.coscoshipping.com",
        "https://synconhub.coscoshipping.com/spot",
        "Synconhub",
    )
    driver.get("https://synconhub.coscoshipping.com/spot")
    try:
        WebDriverWait(driver, 8).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
    except Exception:
        pass

    spot_loaded = ("synconhub.coscoshipping.com/spot" in (driver.current_url or "").lower()
                   and "redirect=" not in (driver.current_url or ""))
                   
    if spot_loaded:
        print("   ✅ Synconhub: session còn hiệu lực.")
    else:
        if _click_text(driver, "button", "Allow All", timeout=3):
            print("   [Synconhub] đã bấm Allow All.")
            time.sleep(1)

        print("   [Synconhub] Đang kích hoạt đồng bộ SSO từ Elines...")
        
        # Bơm JS để ép click nút Sign In/Up (Chắc chắn dính)
        driver.execute_script("""
            let els = document.querySelectorAll('div, a, span');
            for(let e of els) {
                let txt = (e.innerText || "").trim();
                if((txt === 'Sign In/Up' || txt === 'Login' || txt === 'Sign In') && e.offsetParent !== null) {
                    e.click();
                    break;
                }
            }
        """)
        
        # Chờ hệ thống chuyển hướng hoặc tự văng qua Keycloak
        time.sleep(3)
        
        # (Phòng hờ) Nếu web có form đòi email thay vì tự nhảy
        _fill_input_by_placeholder(driver, "e-mail", COSCO_EMAIL, css_class="el-input__inner", timeout=2)
        _fill_input_by_placeholder(driver, "email", COSCO_EMAIL, timeout=1)
        _click_text(driver, "button", "Next", timeout=2)

        # Chờ web hoàn tất quá trình nhận Cookie SSO (tối đa 15s)
        print("   [Synconhub] Chờ hệ thống xác thực...")
        try:
            WebDriverWait(driver, 15).until(
                lambda d: "synconhub.coscoshipping.com/spot" in (d.current_url or "").lower() and "redirect=" not in (d.current_url or "").lower()
            )
        except Exception:
            pass

        # Ép trình duyệt load lại thẳng vào trang spot để kiểm tra lần cuối
        driver.get("https://synconhub.coscoshipping.com/spot")
        try:
            WebDriverWait(driver, 5).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except Exception:
            pass
            
        final_loaded = ("synconhub.coscoshipping.com/spot" in (driver.current_url or "").lower()
                        and "redirect=" not in (driver.current_url or ""))
                        
        if final_loaded:
            print("   ✅ Synconhub login OK (Đã đồng bộ SSO).")
        else:
            print(f"   ⚠️ Synconhub vẫn kẹt ở URL = {driver.current_url[:100]}")

    print("--- HOÀN TẤT ĐĂNG NHẬP COSCO ---\n")

EXCHANGE_RATE_CACHE = {} # Biến nhớ tỷ giá toàn cục



def get_live_exchange_rate(base_currency, target_currency="USD"):
    base = base_currency.upper()
    if base == target_currency.upper(): return 1.0
    
    # Rút từ Cache (Tốc độ 0.00001s)
    if base in EXCHANGE_RATE_CACHE: return EXCHANGE_RATE_CACHE[base]
    
    try:
        res = requests.get(f"https://api.frankfurter.app/latest?from={base}&to=USD", timeout=2)
        rate = float(res.json()['rates']['USD'])
        EXCHANGE_RATE_CACHE[base] = rate
        return rate
    except:
        fallbacks = {"EUR": 1.16, "AUD": 0.65, "CNY": 0.14, "VND": 0.00004, "THB": 0.028, "CHF": 1.12}
        EXCHANGE_RATE_CACHE[base] = fallbacks.get(base, 1.0)
        return EXCHANGE_RATE_CACHE[base]

def is_arbitrary_charge(name):
    up = str(name or "").upper()
    code = re.sub(r"[^A-Z]", " ", up).split()
    return (
        any(x in up for x in ["ARBITRARY", "ARBITRARIES", "ARBITRAR"])
        or any(x in code for x in ["ARB", "ARD"])
    )
    
# ===================================================================================
# TẢI TRƯỚC TỶ GIÁ VÀO RAM NGAY KHI KHỞI ĐỘNG
# ===================================================================================
print("[HỆ THỐNG] Đang nạp trước tỷ giá ngoại tệ vào bộ nhớ đệm...")
get_live_exchange_rate("EUR", "USD")
get_live_exchange_rate("AUD", "USD")
get_live_exchange_rate("VND", "USD")
get_live_exchange_rate("CHF", "USD")
print("[HỆ THỐNG] Nạp tỷ giá hoàn tất!")

# ===================================================================================
# DANH SÁCH CHẶN ĐỨNG (THC, SEAL, DOC VÀ CÁC PHÍ THEO BILL NHƯ AMS, ENS...)
# ===================================================================================
BLOCKLIST_CHARGES = [
    'THC', 'TERMINAL', 'THD', 'DOC', 'SLF', 'SEAL', 'BILL', 'PBF', 'TLX', 'TELEX', 
    'AMS', 'AFS', 'AFR', 'ENS', 'ISPS', 'ISP', 'PSF', 'PSU', 'DCI', 'CLE', 'EMP'
]
# ===================================================================================
# Quy tắc phí Elines (FIX):
# - Local charge đầu LOADING → luôn bỏ (dù Prepaid hay Prepaid/Collect)
# - Payment term = COLLECT thuần → bỏ (local destination)
# - OWS/Overweight → bỏ, flag has_ows
# - Tất cả phí per-container còn lại → CỘNG VÀO giá
# ===================================================================================
ELINES_LOADING_LOCAL = [
    'THC', 'TERMINAL HANDLING',   # THC đầu load
    'SLF', 'SEAL',                # Seal Fee
    'DOC', 'BILL', 'DOCUMENTATION',  # B/L, doc fee
]
ELINES_OVERWEIGHT = ['OWS', 'OVERWEIGHT', 'HCS', 'HES', 'HEAVY']

def is_overweight_charge(name):
    """OWS/HCS chỉ tạo remark SUBJECT TO OWS, không được cộng vào ocean rate."""
    normalized = re.sub(r"\s+", " ", str(name or "").strip().upper())
    return any(marker in normalized for marker in ELINES_OVERWEIGHT)

def is_synconhub_destination_category(category):
    text = re.sub(r"\s+", " ", str(category or "").strip().upper())
    return any(x in text for x in ["DESTINATION", "DISCHARGE", "IMPORT", "POD"])

def is_synconhub_excluded_pol_charge(charge_name):
    """Synconhub: chỉ loại local POL được chỉ định; các phí POL khác phải cộng."""
    name = re.sub(r"\s+", " ", str(charge_name or "").strip().upper())
    return (
        "THC" in name
        or "TERMINAL HANDLING" in name
        or any(x in name for x in ["BILL", "B/L", "DOCUMENTATION", "DOCUMENT FEE", "DOC FEE"])
        or any(x in name for x in ["SEAL", "SLF"])
        or any(x in name for x in ["TLX", "TELEX"])
    )

def parse_synconhub_charge_value(raw_value):
    text = str(raw_value or "").strip().upper().replace(",", "")
    if not text or text in {"-", "N/A"} or "INCLUDED" in text:
        return None
    currency_match = re.search(r"\b(USD|EUR|CHF|AUD|VND)\b", text)
    amount_match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not amount_match:
        return None
    return (currency_match.group(1) if currency_match else "USD", float(amount_match.group(0)))

def synconhub_thc_is_explicitly_included(values):
    meaningful = [
        str(value or "").strip().upper()
        for value in values
        if str(value or "").strip().upper() not in {"", "-", "N/A"}
    ]
    return bool(meaningful) and all("INCLUDED" in value for value in meaningful)

def focus_tab_by_url(domain_keyword, fallback_url):
    handle = _open_or_focus_tab(driver, domain_keyword, fallback_url, domain_keyword)
    if not handle:
        raise RuntimeError(f"COSCO không focus/mở được tab {domain_keyword}")
    try:
        switch_to_live_window(driver, preferred_handle=handle)
        if not _url_matches_target_path(driver.current_url, fallback_url):
            driver.get(fallback_url)
    except Exception as exc:
        raise RuntimeError(f"COSCO tab {domain_keyword} không còn sử dụng được") from exc
    time.sleep(1)
    return True

def reload_synconhub_base():
    driver.get("https://synconhub.coscoshipping.com/spot")
    time.sleep(1.5)

# ===================================================================================
# ── Tự động thêm Timestamp (Thời gian thực) vào lệnh Print ──
# ===================================================================================
import sys
import io

# Xử lý lỗi Unicode trên Windows Terminal
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Lưu lại hàm print gốc
_orig_print = print

# Định nghĩa hàm print mới "độ" thêm thời gian
def print(*args, **kwargs):
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3] # Định dạng: HH:MM:SS.mili
    _orig_print(f"[{ts}]", *args, **kwargs)

# ===================================================================================
# --- 2. LOGIC LỌC 9 QUY TẮC ETD ---
# ===================================================================================
def calculate_validity(last_etd):
    import calendar
    year, month = last_etd.year, last_etd.month
    last_day = calendar.monthrange(year, month)[1]
    
    milestones = [7, 14, 21, last_day]
    for m in milestones:
        if m >= last_etd.day:
            dt = datetime(year, month, m)
            return f"{dt.day}-{dt.strftime('%b')}"
    return f"{last_day}-{last_etd.strftime('%b')}"

def format_etd_dates_excel(dates):
    if not dates:
        return ""
    ordered = sorted(dates)
    if len(ordered) == 1:
        return f"{ordered[0].day}-{ordered[0].strftime('%b')}"
    if len(ordered) == 2:
        return f"{ordered[0].day}-{ordered[0].strftime('%b')} & {ordered[1].day}-{ordered[1].strftime('%b')}"

    groups = []
    for dt in ordered:
        key = (dt.year, dt.month)
        if not groups or groups[-1][0] != key:
            groups.append((key, [dt]))
        else:
            groups[-1][1].append(dt)

    parts = []
    for _, group_dates in groups:
        month = group_dates[-1].strftime("%b")
        if len(group_dates) == 1:
            parts.append(f"{group_dates[0].day}-{month}")
        else:
            days = ", ".join(str(d.day) for d in group_dates)
            parts.append(f"{days}-{month}")

    if len(parts) == 1:
        return parts[0]
    return f"{', '.join(parts[:-1])} & {parts[-1]}"

def apply_9_golden_rules(danh_sach_chuyen):
    if not danh_sach_chuyen:
        return [], "", ""
    danh_sach_chuyen = [c for c in danh_sach_chuyen if etd_within_max(c.get("etd_dt"))]
    if not danh_sach_chuyen:
        return [], "", ""

    danh_sach_chuyen.sort(key=lambda x: (x["etd_dt"], x["tt_days"]))
    list_loc_trung = []
    seen_dates = set()
    for c in danh_sach_chuyen:
        if c["etd_dt"] not in seen_dates:
            list_loc_trung.append(c)
            seen_dates.add(c["etd_dt"])

    etd_dat_chuan = []
    if list_loc_trung:
        ngan_nhat_global = min(c["tt_days"] for c in list_loc_trung)
        first_date = list_loc_trung[0]["etd_dt"]
        for c in list_loc_trung:
            if len(etd_dat_chuan) >= 3: break 
            if len(etd_dat_chuan) > 0 and (c["etd_dt"] - etd_dat_chuan[-1]["etd_dt"]).days < 2: continue
            if (c["etd_dt"] - first_date).days <= 9 and (c["tt_days"] <= ngan_nhat_global + 10):
                etd_dat_chuan.append(c)

    format_str = ""
    num = len(etd_dat_chuan)
    
    if num >= 1:
        format_str = format_etd_dates_excel([c["etd_dt"] for c in etd_dat_chuan])

    if not etd_dat_chuan:
        # Keep the row usable even when a carrier card contains malformed
        # transit-time data that cannot satisfy the normal filters.
        etd_dat_chuan = list_loc_trung[:1]
        format_str = format_etd_dates_excel([c["etd_dt"] for c in etd_dat_chuan])

    all_tt = [c["tt_days"] for c in etd_dat_chuan]
    tt_min, tt_max = min(all_tt), max(all_tt)
    str_tt = f"{tt_min}" if tt_min == tt_max else f"{tt_min}-{tt_max}"
    
    return etd_dat_chuan, format_str, str_tt

# ===================================================================================
# --- 3. BỘ NÃO NHẬP CẢNG (BẢN TỐI ƯU: KHÔNG XÓA MÙ QUÁNG - ZERO DELAY) ---
# ===================================================================================
COSCO_COUNTRY_ALIASES = {
    "MEXICO": ["MEXICO", "MX", "MEX"],
    "VIETNAM": ["VIETNAM", "VN", "VNM"],
    "CHINA": ["CHINA", "CN", "CHN"],
    "JAPAN": ["JAPAN", "JP", "JPN"],
    "KOREA": ["KOREA", "KR", "KOR"],
    "SOUTH KOREA": ["KOREA", "KR", "KOR"],
    "INDIA": ["INDIA", "IN", "IND"],
    "AUSTRALIA": ["AUSTRALIA", "AU", "AUS"],
    "UNITED STATES": ["UNITED STATES", "USA", "US"],
    "USA": ["UNITED STATES", "USA", "US"],
}

COSCO_PORT_ALIASES = {
    "HAI PHONG": ["HAI PHONG"],
    "HO CHI MINH": ["HO CHI MINH"],
    "NHAVA SHEVA": ["NHAVA SHEVA", "NHAVASHEVA"],
    "PORT KLANG": ["PORT KLANG", "PORTKLANG"],
}
COSCO_EXACT_SPACE_PORTS = {"HAI PHONG", "HO CHI MINH"}

def _cosco_country_tokens(country):
    country_upper = str(country or "").strip().upper()
    if not country_upper:
        return []
    tokens = COSCO_COUNTRY_ALIASES.get(country_upper, [country_upper])
    if country_upper not in tokens:
        tokens = [country_upper] + tokens
    return list(dict.fromkeys([t.strip().upper() for t in tokens if t and t.strip()]))

def _cosco_text_has_country(txt, country):
    text = str(txt or "").upper()
    for token in _cosco_country_tokens(country):
        if len(token) <= 3:
            if re.search(rf"(?<![A-Z0-9]){re.escape(token)}(?![A-Z0-9])", text):
                return True
            if len(token) == 2 and re.search(rf"\({re.escape(token)}[A-Z0-9]*\)", text):
                return True
        elif token in text:
            return True
    return False

def _cosco_compact_text(value):
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())

def _cosco_port_aliases(clean_port_name):
    base = re.sub(r"\s+", " ", str(clean_port_name or "").strip().upper())
    if not base:
        return []
    if base in COSCO_EXACT_SPACE_PORTS:
        return [base]
    aliases = [base]
    compact = _cosco_compact_text(base)
    if compact and compact != base:
        aliases.append(compact)
    aliases.extend(COSCO_PORT_ALIASES.get(base, []))
    return list(dict.fromkeys([a.strip().upper() for a in aliases if a and a.strip()]))

def _cosco_port_query_candidates(clean_port_name, country):
    aliases = _cosco_port_aliases(clean_port_name)
    base = re.sub(r"\s+", " ", str(clean_port_name or "").strip().upper())
    if base in COSCO_EXACT_SPACE_PORTS:
        return aliases
    country_upper = str(country or "").strip().upper()
    candidates = []
    for alias in aliases:
        candidates.append(alias)
        if country_upper:
            candidates.append(f"{alias} {country_upper}".strip())
    return list(dict.fromkeys(candidates))

def _cosco_text_matches_port(txt, clean_port_name):
    text = re.sub(r"\s+", " ", str(txt or "").strip().upper())
    text_compact = _cosco_compact_text(text)
    strict_space_only = re.sub(r"\s+", " ", str(clean_port_name or "").strip().upper()) in COSCO_EXACT_SPACE_PORTS
    aliases = _cosco_port_aliases(clean_port_name)
    if not text or not aliases:
        return False
    for alias in aliases:
        port = re.sub(r"\s+", " ", alias.strip().upper())
        port_compact = _cosco_compact_text(port)
        if text == port:
            return True
        if re.match(rf"^{re.escape(port)}(?=$|[\s,;/()\\-])", text):
            return True
        if re.search(rf"(?<![A-Z0-9]){re.escape(port)}(?![A-Z0-9])", text):
            return True
        if strict_space_only:
            continue
        if port_compact and text_compact.startswith(port_compact):
            return True
    return False

def _cosco_input_has_query(curr_val, query_text):
    query_norm = re.sub(r"\s+", " ", str(query_text or "").strip().upper())
    if query_norm in COSCO_EXACT_SPACE_PORTS:
        curr_norm = re.sub(r"\s+", " ", str(curr_val or "").strip().upper())
        return query_norm in curr_norm
    return _cosco_compact_text(query_text) in _cosco_compact_text(curr_val)

def _cosco_fill_port_input(inp, query_text, remove_readonly=False):
    if remove_readonly:
        driver.execute_script("arguments[0].removeAttribute('readonly');", inp)
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
    try:
        inp.click()
    except Exception:
        driver.execute_script("arguments[0].click();", inp)
    time.sleep(0.1)
    try:
        inp.send_keys(Keys.CONTROL, "a")
        inp.send_keys(Keys.BACKSPACE)
    except Exception:
        pass
    driver.execute_script("""
        const input = arguments[0];
        const nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        nativeInputValueSetter.call(input, '');
        input.dispatchEvent(new Event('input', { bubbles: true }));
        input.dispatchEvent(new Event('change', { bubbles: true }));
    """, inp)
    time.sleep(0.1)
    try:
        inp.send_keys(query_text)
        time.sleep(0.45)
    except Exception:
        pass
    curr = driver.execute_script("return arguments[0].value || '';", inp)
    if not _cosco_input_has_query(curr, query_text):
        driver.execute_script("""
            const input = arguments[0];
            const val = arguments[1];
            const nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            nativeInputValueSetter.call(input, val);
            input.dispatchEvent(new Event('input', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
            input.dispatchEvent(new Event('compositionend', { bubbles: true }));
            input.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: val.slice(-1) || 'a' }));
            input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: val.slice(-1) || 'a' }));
        """, inp, query_text)
        time.sleep(0.45)
    else:
        driver.execute_script("""
            const input = arguments[0];
            const val = arguments[1];
            input.dispatchEvent(new Event('input', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
            input.dispatchEvent(new Event('compositionend', { bubbles: true }));
            input.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: val.slice(-1) || 'a' }));
            input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: val.slice(-1) || 'a' }));
        """, inp, query_text)
        time.sleep(0.25)

def select_port_smart(xpath, port_name, country, sys_name, is_elines=False):
    print(f"      + [{sys_name}] Nạp cảng: {port_name}, {country}")
    
    clean_port_name = port_name.split(',')[0].strip().upper()
    country_upper = str(country or "").upper()
    query_candidates = _cosco_port_aliases(clean_port_name) if is_elines else _cosco_port_query_candidates(clean_port_name, country_upper)

    max_attempts = max(5, len(query_candidates) + 1)
    for attempt in range(max_attempts): 
        try:
            # 1. Tìm ô nhập liệu (Dùng WebDriverWait để không bị rớt do load chậm)
            input_wait = max(COSCO_PORT_INPUT_WAIT_SECONDS, 20 if is_elines else 10)
            inp = WebDriverWait(driver, input_wait).until(lambda d: next((e for e in d.find_elements(By.XPATH, xpath) if e.is_displayed()), None))
            if not inp: raise Exception("Không tìm thấy ô!")
            
            # 2. Đọc giá trị hiện tại
            curr_val = driver.execute_script("return arguments[0].value;", inp).strip().upper()
            
            # KIỂM TRA CHUẨN: 
            if (
                sys_name != "Synconhub"
                and not (clean_port_name == "NANSHA" and "CANTON" in curr_val)
                and (not is_elines or "," in curr_val)
                and (not country_upper or _cosco_text_has_country(curr_val, country_upper))
                and _cosco_text_matches_port(curr_val, clean_port_name)
            ):
                print(f"        -> Đã có sẵn CHUẨN: {curr_val} -> ĐI TIẾP!")
                return

            # 3. NẠP CHỮ (SMART FILL): 
            # Nếu ô đang trắng hoặc chứa chữ tào lao -> Mới xóa gõ lại.
            # Nếu đã có sẵn "ALGECIRAS" -> Giữ nguyên, chỉ kích hoạt sự kiện.
            query_text = query_candidates[min(attempt, len(query_candidates) - 1)] if query_candidates else clean_port_name

            _cosco_fill_port_input(inp, query_text, remove_readonly=(sys_name == "Synconhub"))
            
            # Kích hoạt sự kiện để Dropdown hiện ra (Không xóa chữ cũ)
            driver.execute_script("arguments[0].dispatchEvent(new Event('focus', { bubbles: true }));", inp)
            driver.execute_script("arguments[0].dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: ' ' }));", inp)
            if sys_name == "Synconhub":
                driver.execute_script("""
                    const input = arguments[0];
                    input.dispatchEvent(new Event('input', { bubbles: true }));
                    input.dispatchEvent(new Event('change', { bubbles: true }));
                    input.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: 'ArrowDown' }));
                    input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: 'ArrowDown' }));
                """, inp)
                
            # 4. ĐỢI DROPDOWN VÀ CLICK
            css_sel = "div.el-autocomplete-suggestion:not([style*='display: none']) li" if is_elines else "div.el-select-dropdown:not([style*='display: none']) li"
            
            target_strict = f"{clean_port_name}," 
            target_semi = f"{clean_port_name} ,"
            
            dropdown_wait = max(COSCO_PORT_DROPDOWN_WAIT_SECONDS, 20 if is_elines else 10)
            timeout = time.time() + dropdown_wait
            match = None
            match_score = 999
            debug_options = []
            while time.time() < timeout:
                opts = driver.find_elements(By.CSS_SELECTOR, css_sel)
                if not opts:
                    try:
                        inp.send_keys(Keys.ARROW_DOWN)
                    except Exception:
                        pass
                    time.sleep(0.15)
                    continue
                for o in opts:
                    txt = o.text.strip().upper()
                    if txt and txt not in debug_options:
                        debug_options.append(txt[:120])
                    if country_upper and not _cosco_text_has_country(txt, country_upper):
                        continue
                    
                    # LUẬT LỌC KHẮT KHE: Bắt buộc phần đầu phải khớp hoàn toàn tới dấu phẩy
                    if sys_name == "Synconhub" and clean_port_name == "HO CHI MINH" and "CAT LAI" in txt:
                        continue
                    if clean_port_name == "NANSHA" and "CANTON" in txt:
                        continue
                    is_exact_city = txt.startswith(target_strict) or txt.startswith(target_semi)
                    is_port_match = _cosco_text_matches_port(txt, clean_port_name)
                    if not (is_exact_city or is_port_match):
                        continue
                    score = 0 if is_exact_city else 10
                    if "(" in txt.split(",", 1)[0]:
                        score += 5
                    if score < match_score:
                        match = o
                        match_score = score
                if match and match_score == 0: break
                time.sleep(0.05)
            
            if match:
                final_text = match.text.strip()
                driver.execute_script("arguments[0].click();", match)
                print(f"        -> Đã chốt từ list: {final_text}")
                return 
            
            if debug_options:
                print(f"        -> [DEBUG] Options thấy được: {debug_options[:8]}")
            raise Exception("Không tìm thấy option phù hợp chuẩn xác trong list")
            
        except Exception as e:
            try:
                driver.execute_script("arguments[0].blur();", inp)
            except Exception:
                pass
            time.sleep(0.8 if is_elines else 0.5)
            
    raise Exception(f"Thất bại nạp cảng {port_name} sau {max_attempts} lần thử.")


# ===================================================================================
# --- 4. SYNCONHUB ---
# ===================================================================================
def run_synconhub(pol, pod, pod_country):
    china_route = is_china_destination(pod_country, pod)
    step_tracker = "Khởi động hàm Synconhub"
    try:
        driver.switch_to.default_content()
        
        step_tracker = "Nhập Origin"
        select_port_smart("//div[contains(@class, 'ect-label') and text()='Origin']/following-sibling::div//input", pol, "VIETNAM", "Synconhub")
        
        step_tracker = "Nhập Destination"
        select_port_smart("//div[contains(@class, 'ect-label') and text()='Destination']/following-sibling::div//input", pod, pod_country, "Synconhub")

        step_tracker = "Nhập ETD + 10 Ngày"
        print("      -> [Synconhub] Đang nạp ngày ETD + 10...")
        try:
            etd_inp = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//input[@name='sailing_product_date_picker_start']")))
            dt_str = (datetime.now() + timedelta(days=DATE_OFFSET_DAYS)).strftime("%Y-%m-%d")
            
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", etd_inp)
            driver.execute_script("arguments[0].removeAttribute('readonly');", etd_inp)
            driver.execute_script("arguments[0].value = '';", etd_inp)
            driver.execute_script(f"arguments[0].value = '{dt_str}';", etd_inp)
            driver.execute_script("arguments[0].dispatchEvent(new Event('input')); arguments[0].dispatchEvent(new Event('change'));", etd_inp)
            time.sleep(0.5)
            
            try: etd_inp.send_keys(Keys.ESCAPE)
            except: pass
            driver.execute_script("document.body.click();")
            print(f"      -> [Synconhub] Đã chốt ngày ETD: {dt_str}")
        except Exception as e:
            print(f"      ⚠️ [Synconhub] Kẹt lúc nạp ETD: {e}")

        step_tracker = "Tắt Cont Lạnh (RF/NOR)"
        for lbl in ["RF", "NOR"]:
            try:
                elem = driver.find_element(By.XPATH, f"//label[contains(@class, 'el-checkbox') and .//span[contains(text(), '{lbl}')]]")
                if "is-checked" in elem.get_attribute("class"): driver.execute_script("arguments[0].click();", elem)
            except: pass

        step_tracker = "Bấm nút Search"
        driver.execute_script("arguments[0].click();", driver.find_element(By.XPATH, "//button[contains(@class, 'ect-search-btn')]"))
        print("      -> [Synconhub] Đã bấm Search...")
        
        step_tracker = "Chờ Loading Mask"
        try:
            WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".el-loading-mask")))
            WebDriverWait(driver, 20).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".el-loading-mask")))
        except: pass

        step_tracker = "Kiểm tra kết quả / No Result"
        time.sleep(1) 
        try:
            WebDriverWait(driver, 15).until(
                lambda d: any(e.is_displayed() for e in d.find_elements(By.CSS_SELECTOR, ".ect-search-result-body")) or 
                          any(e.is_displayed() for e in d.find_elements(By.CSS_SELECTOR, ".ect-search-no-result"))
            )
        except: 
            print("      ❌ [Synconhub] 15s trôi qua không thấy thẻ giá!")
            return None
            
        no_res_elements = driver.find_elements(By.CSS_SELECTOR, ".ect-search-no-result")
        if any(e.is_displayed() for e in no_res_elements):
            time.sleep(3)
            cards_after_wait = driver.find_elements(By.CSS_SELECTOR, ".ect-search-result-body")
            if not any(c.is_displayed() for c in cards_after_wait):
                print("      -> [Synconhub] NO SERVICE / SOLD OUT!")
                return SYNCONHUB_NO_SERVICE
            
        step_tracker = "Đọc thẻ giá"
        print("      -> [Synconhub] Thẻ giá hiện, đang đọc...")
        time.sleep(1)
        cards = driver.find_elements(By.CSS_SELECTOR, ".ect-search-result-body")
        list_chuyen = []
        for card in cards:
            if not card.is_displayed(): continue
            try:
                etd_dt = datetime.strptime(card.find_element(By.XPATH, ".//div[text()='ETD']/following-sibling::div").text.strip(), "%Y-%m-%d")
                tt_text = card.find_element(By.CSS_SELECTOR, ".transit-time span").text.replace("days", "").strip()
                prices = [
                    price
                    for i in card.find_elements(By.CSS_SELECTOR, ".container-type-item")
                    if i.find_element(By.CSS_SELECTOR, ".ect-price").text.strip()
                    for price in [
                        float(
                            i.find_element(By.CSS_SELECTOR, ".ect-price")
                            .text.replace('$', '')
                            .replace(',', '')
                            .strip()
                        )
                    ]
                    if math.isfinite(price)
                ]
                if prices: list_chuyen.append({"element": card, "price": min(prices), "etd_dt": etd_dt, "tt_days": int(tt_text) if tt_text.isdigit() else 999})
            except: continue

        if not list_chuyen: 
            print("      ❌ [Synconhub] Không đọc được thẻ giá!")
            return None

        step_tracker = "Lọc 9 quy tắc vàng"
        gia_re_nhat = min(c["price"] for c in list_chuyen)
        list_chuyen = [c for c in list_chuyen if c["price"] == gia_re_nhat]
        etd_chuan, str_etd, str_tt = apply_9_golden_rules(list_chuyen)
        dai_dien = etd_chuan[0]["element"]
        valid_date_str = calculate_validity(etd_chuan[-1]["etd_dt"]) # <-- THÊM DÒNG NÀY
        print(f"      -> [Synconhub] Chọn ETD {str_etd}, TT {str_tt} days")

# --- BẮT ĐẦU THÊM: RÚT THÔNG TIN TÀU VÀ CHUYỂN TẢI (SYNCONHUB) ---
        step_tracker = "Lấy Thông Tin Tàu (Synconhub)"
        print("      -> [Synconhub] Đang soi tên tàu và cảng chuyển tải...")
        vessel_texts = []
        ts_combos = []
        
        for c in etd_chuan:
            card = c["element"]
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", card)
            time.sleep(0.4)
            
            # 1. Lấy Tên Tàu (Cắt trước dấu | và dấu /) BẰNG JAVASCRIPT
            vessel_name = "TBA"
            try:
                # Dùng JS leo ra ngoài thẻ wrapper rồi mới móc xuống footer
                v_raw = driver.execute_script("""
                    let wrapper = arguments[0].closest('.ect-search-result-wrapper');
                    if (wrapper) {
                        let footer = wrapper.querySelector('.ect-search-result-footer');
                        if (footer) {
                            let span = footer.querySelector('span');
                            if (span) return span.textContent.trim();
                        }
                    }
                    return '';
                """, card)
                
                if v_raw:
                    # Tách phần trước dấu | (bỏ phần tàu phụ)
                    v_part = v_raw.split('|')[0].strip()
                    # Tách phần trước dấu / (bỏ mã service)
                    vessel_name = v_part.split('/')[0].strip()
            except Exception as e:
                print(f"      [Debug] Lỗi rút tên tàu: {e}")
                
            # 2. Lấy Cảng Chuyển Tải
            ts_ports = []
            try:
                # Kiểm tra xem có icon T/S không
                ts_el = card.find_elements(By.XPATH, ".//div[contains(@class, 'schedule-tip')]//div[contains(text(), 'T/S')]")
                if ts_el:
                    # Bơm JS kích hoạt hover chuột để ép popup bung ra
                    driver.execute_script("""
                        arguments[0].scrollIntoView({block: 'center', inline: 'center'});
                        arguments[0].dispatchEvent(new MouseEvent('mouseenter', {bubbles: true, cancelable: true, view: window}));
                    """, ts_el[0])
                    time.sleep(1.5) # Đợi popup bung
                    
                    # Rút data từ popup
                    st_names = driver.execute_script("""
                        let pops = document.querySelectorAll('.el-tooltip__popper[x-placement]');
                        for (let i = pops.length - 1; i >= 0; i--) {
                            let p = pops[i];
                            if (window.getComputedStyle(p).display === 'none') continue;
                            let points = p.querySelectorAll('.point');
                            if (points.length > 0) {
                                let names = [];
                                points.forEach(pt => {
                                    let divs = pt.querySelectorAll('div');
                                    // Cột div số 3 (index 2) chứa tên tiếng Anh
                                    if (divs.length >= 3) {
                                        let txt = divs[2].textContent.trim();
                                        if(txt) names.push(txt.toUpperCase());
                                    }
                                });
                                return names;
                            }
                        }
                        return [];
                    """)
                    
                    if st_names and len(st_names) >= 2:
                        origin = st_names[0]
                        dest = st_names[-1]
                        for p in st_names:
                            # Lọc bỏ cảng xếp, cảng dỡ và các trạm trung chuyển bị lặp tên
                            if p != origin and p != dest and p not in ts_ports:
                                ts_ports.append(p)
                                
                    # Đóng popup bằng cách click ra ngoài
                    driver.execute_script("document.body.click();")
                    driver.execute_script("arguments[0].dispatchEvent(new MouseEvent('mouseleave', {bubbles: true, cancelable: true, view: window}));", ts_el[0])
                    time.sleep(0.3)
            except Exception as e:
                print(f"      [Debug] Lỗi rút T/S Synconhub: {e}")
                
            # Tổng hợp chuỗi
            str_ts = " + ".join(ts_ports) if ts_ports else "DIRECT"
            str_etd_fmt = f"{c['etd_dt'].day}-{c['etd_dt'].strftime('%b')}"
            str_tt_fmt = str(c["tt_days"])
            
            info_str = f"{vessel_name} / ETD: {str_etd_fmt} / Transit time: {str_tt_fmt} Days / Transshipment Port: {str_ts}"
            vessel_texts.append(info_str)
            ts_combos.append(str_ts)
            
        unique_ts = []
        for t in ts_combos:
            if t not in unique_ts: unique_ts.append(t)
            
        final_ts_str = " or \n".join(unique_ts)
        final_vessel_str = "\n".join(vessel_texts)
        # --- KẾT THÚC THÊM ---

        step_tracker = "Đọc Base Rate"
        rates = {"20GP": None, "40GP": None, "40HQ": None}
        formula_parts = {"20GP": [], "40GP": [], "40HQ": []}
        for item in dai_dien.find_elements(By.CSS_SELECTOR, ".container-type-item"):
            ctype = item.find_element(By.CSS_SELECTOR, ".cntr-title").text.strip()
            if ctype in rates:
                try:
                    rates[ctype] = float(item.find_element(By.CSS_SELECTOR, ".ect-price").text.replace('$', '').replace(',', '').strip())
                    formula_parts[ctype].append(rates[ctype])
                except: pass
        print(f"      -> [Synconhub] Base rates: {rates}")

        step_tracker = "Mở dấu chấm hỏi Surcharge"
        has_ows = False
        thc_inc = False
        china_thc_row_seen = False
        surcharge_table_ok = False
        manifest_fee_found = False
        try:
            print("      -> [Synconhub] Mở bảng Surcharge...")
            s_btn = dai_dien.find_element(By.CSS_SELECTOR, ".extra-charge-btn")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", s_btn)
            time.sleep(0.5); driver.execute_script("arguments[0].click();", s_btn)
            
            table = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'el-popover') and not(contains(@style, 'display: none'))]//table[contains(@class, 'el-table__body')]")))
            time.sleep(1)
            
            step_tracker = "Xử lý bảng Surcharge Synconhub"
            curr_cat = ""
            for row in table.find_elements(By.TAG_NAME, "tr"):
                tds = row.find_elements(By.TAG_NAME, "td")
                if len(tds) < 6: continue
                if len(tds) == 7: curr_cat = tds[0].text.strip().upper()
                if is_synconhub_destination_category(curr_cat):
                    print(f"        [-POD LOCAL] Bỏ qua nhóm {curr_cat}")
                    continue
                 
                c_name = tds[-6].text.strip().upper()
                 
                if any(x in c_name for x in ["ENS", "AMS", "AFS", "ADVANCED MANIFEST", "ENTRY SUMMARY"]):
                    manifest_fee_found = True
                    print(f"        [FLAG] Đã tóm được {c_name[:25]} -> Ghi chú vào Remark!")

                if is_overweight_charge(c_name):
                    has_ows = True
                    print(f"        [-OWS] Bỏ qua {c_name[:25]} -> chỉ ghi SUBJECT TO OWS")
                    continue

                is_origin_thc = "THC" in c_name or "TERMINAL HANDLING" in c_name
                if is_origin_thc:
                    if china_route:
                        china_thc_row_seen = True
                    thc_vals = [tds[-3+i].text.strip().upper() for i in range(3)]
                    if synconhub_thc_is_explicitly_included(thc_vals):
                        thc_inc = True
                        print("        [THC] Synconhub ghi rõ INCLUDED.")
                    elif china_route:
                        thc_added = False
                        for idx, ctype in enumerate(["20GP", "40GP", "40HQ"]):
                            parsed_value = parse_synconhub_charge_value(tds[-3+idx].text)
                            if not parsed_value:
                                continue
                            currency, amt = parsed_value
                            if currency != "USD":
                                amt *= get_live_exchange_rate(currency, "USD")
                            if rates[ctype] is not None:
                                rates[ctype] += amt
                                formula_parts.setdefault(ctype, []).append(amt)
                                thc_added = True
                                print(f"        [+O.THC CHINA] {amt:.2f} USD -> {ctype}")
                        thc_inc = thc_inc or thc_added
                        if not thc_added:
                            print("        [THC CHINA] Có dòng O.THC riêng nhưng không đọc được số tiền.")
                    else:
                        print("        [-THC] Không cộng THC; không có xác nhận INCLUDED.")
                    continue
                     
                if is_synconhub_excluded_pol_charge(c_name):
                    print(f"        [-POL LOCAL] Bỏ qua {c_name[:25]}")
                    continue
                 
                for idx, ctype in enumerate(["20GP", "40GP", "40HQ"]):
                    parsed_value = parse_synconhub_charge_value(tds[-3+idx].text)
                    if not parsed_value:
                        continue
                    currency, amt = parsed_value
                    if currency == 'EUR':
                        amt *= get_live_exchange_rate("EUR", "USD")
                    elif currency == 'CHF':
                        amt *= get_live_exchange_rate("CHF", "USD")
                    elif currency == 'AUD':
                        amt *= get_live_exchange_rate("AUD", "USD")
                    elif currency == 'VND':
                        amt *= get_live_exchange_rate("VND", "USD")

                    if rates[ctype] is not None:
                        rates[ctype] += amt
                        formula_parts.setdefault(ctype, []).append(amt)
                        print(f"        [+POL] {amt:.2f} USD ({c_name[:20]}) → {ctype}")
                            
            surcharge_table_ok = True
            driver.execute_script("document.body.click();")
            print("      -> [Synconhub] Xong bảng Surcharge!")
        except Exception as e: 
            print(f"      ⚠️ [Synconhub] Lỗi Surcharge: {e}")

        if china_route and not china_thc_row_seen and not thc_inc:
            # No separate O.THC row means Synconhub has already folded it into
            # the China ocean rate.  A separate row is handled above.
            thc_inc = True

        if thc_inc:
            print("      -> [Synconhub] THC ghi rõ INCLUDED -> ghi INCLUDED O.THC vào remark.")
        else:
            print("      -> [Synconhub] Không có xác nhận THC INCLUDED -> remark SUBJECT TO THC.")

        print(f"      -> [Synconhub] XONG! Cước: {rates}")
        # THÊM BIẾN VALID VÀO LÚC RETURN
        return {
            "rates": rates, "etd": str_etd, "tt": str_tt, "ows": has_ows, 
            "thc_inc": thc_inc, "valid": valid_date_str, "manifest_fee": manifest_fee_found,
            "formulas": {ct: _excel_formula_from_parts(formula_parts.get(ct, [])) for ct in ["20GP", "40GP", "40HQ"]},
            "surcharge_error": False, "vessel_info": final_vessel_str, "transshipment": final_ts_str
        }
    
    except Exception as e:
        print("\n" + "!"*60)
        print(f"      🚨 SYNCONHUB CRASH TẠI: [{step_tracker}]")
        print(f"      👉 {e}")
        print(traceback.format_exc())
        print("!"*60 + "\n")
        return None

# ===================================================================================
# --- 5a. HÀM PHỤ: NHẬP CÂN NHANH ELINES (CÓ RADAR BÁO CÁO) ---
# ===================================================================================
def _elines_fill_weight_fast(driver):
    deadline = time.time() + 20
    injected = False
    
    # Biến để theo dõi thời gian in log
    start_time = time.time()
    last_print_time = time.time()
    
    while time.time() < deadline:
        all_inputs = driver.find_elements(By.CSS_SELECTOR,
            "input[type='number']:not([readonly]):not([disabled])")
        visible = [i for i in all_inputs if i.is_displayed()]
        
        if visible:
            # FIX: Chờ 0.5s để UI web bung nốt các ô nhập cân của 20GP và 40GP
            time.sleep(0.5)
            
            # Lấy lại danh sách các ô lần nữa sau khi đã bung đủ
            all_inputs = driver.find_elements(By.CSS_SELECTOR,
                "input[type='number']:not([readonly]):not([disabled])")
            visible = [i for i in all_inputs if i.is_displayed()]
            
            for inp in visible:
                try:
                    driver.execute_script(
                        "arguments[0].value='22222';"
                        "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
                        "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                        inp
                    )
                except:
                    pass
            for inp in driver.find_elements(By.CSS_SELECTOR, ".el-input-number input"):
                try:
                    if inp.is_displayed():
                        driver.execute_script(
                            "arguments[0].value='1';"
                            "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
                            "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
                            inp
                        )
                except:
                    pass
            injected = True
            print(f"      -> [Elines] ✅ Web đã vẽ xong form! Đã nhập 15000kg siêu tốc cho {len(visible)} ô!")
            time.sleep(0.5) # Nghỉ thêm nửa giây cho web nhận đủ số trước khi bấm Calculate
            break
            
        # --- RADAR BÁO CÁO MỖI 1 GIÂY ---
        current_time = time.time()
        if current_time - last_print_time >= 1.0:
            elapsed = int(current_time - start_time)
            print(f"        ⏳ Vẫn đang chờ mạng load form Detail... ({elapsed}s)")
            last_print_time = current_time
            
        time.sleep(0.1)
        
    if not injected:
        print("      ⚠️ [Elines] Chờ 20s mà web COSCO rớt mạng không load nổi form!")


# ===================================================================================
# --- 5b. HÀM PHỤ: ĐỌC GIÁ TỪ UNIT PRICE CELL (FIX) ---
# Dùng find_elements thay vì .//span[1] và .//span[2] để tránh lỗi im lặng
# ===================================================================================
def _parse_price_cell(price_td, rate_eur, rate_aud=None, rate_vnd=None):
    """
    Đọc (currency, amount_usd) từ Unit Price cell của Elines.
    Cell HTML: <div><span class='text-xs'>USD</span><span class='text-sm'>94</span></div>
    Trả về float (USD) hoặc None nếu lỗi.
    """
    try:
        spans = price_td.find_elements(By.TAG_NAME, "span")
        # Lọc span có text thực (bỏ span rỗng)
        valid = [s for s in spans if s.text.strip()]
        if len(valid) < 2:
            return None
        c_str = valid[0].text.strip().upper()
        amt_raw = valid[1].text.strip().replace(',', '')
        if not amt_raw:
            return None
        amt = float(amt_raw)
        if c_str == 'EUR':
            amt *= rate_eur
        elif c_str == 'AUD':
            amt *= (rate_aud or get_live_exchange_rate("AUD", "USD"))
        elif c_str == 'VND':
            amt *= (rate_vnd or get_live_exchange_rate("VND", "USD"))
        elif c_str == 'CHF':
            amt *= get_live_exchange_rate("CHF", "USD")
        # USD: giữ nguyên
        return amt
    except Exception as e:
        return None


def _elines_visible(element):
    try:
        return bool(element and element.is_displayed())
    except Exception:
        return False


def _elines_card_uses_new_ui(card):
    try:
        return is_new_elines_result_card_text(card.text)
    except Exception:
        return False


def _elines_card_has_view_premium(card):
    try:
        return any(_elines_visible(btn) for btn in card.find_elements(
            By.XPATH, ".//button[contains(normalize-space(.), 'View Premium')]"
        ))
    except Exception:
        return False


def _elines_new_result_cards():
    """Return current-UI cards, including cards with Book Now but no premium."""
    cards = []
    seen = set()
    buttons = driver.find_elements(
        By.XPATH,
        "//button[contains(normalize-space(.), 'View Premium') or "
        "contains(normalize-space(.), 'Book Now')]",
    )
    for button in buttons:
        if not _elines_visible(button):
            continue
        try:
            card = driver.execute_script("""
                let node = arguments[0];
                for (let depth = 0; node && depth < 12; depth += 1, node = node.parentElement) {
                    const text = (node.innerText || '').toUpperCase();
                    const vessel = text.includes('VESSEL / VOYAGE') || text.includes('VESSEL/VOYAGE');
                    const action = text.includes('BOOK NOW') || text.includes('VIEW PREMIUM');
                    if (text.includes('CY CUTOFF') && vessel && action) return node;
                }
                return null;
            """, button)
        except Exception:
            card = None
        if not _elines_visible(card):
            continue
        key = getattr(card, "id", None) or str(card)
        if key not in seen:
            seen.add(key)
            cards.append(card)
    return cards


def _elines_visible_result_cards():
    new_cards = _elines_new_result_cards()
    if new_cards:
        return new_cards
    cards = driver.find_elements(By.XPATH,
        "//div[contains(@class,'box-border') and contains(@class,'hover:shadow-lg')]")
    return [card for card in cards if _elines_visible(card)]


def _elines_find_visible_premium_modal():
    for modal in driver.find_elements(By.CSS_SELECTOR, ".el-dialog, [role='dialog']"):
        if not _elines_visible(modal):
            continue
        try:
            text = (modal.text or "").upper()
        except Exception:
            text = ""
        if "PREMIUM SERVICE" in text and "20GP" in text:
            return modal
    return None


def _elines_open_premium_modal(card, timeout=10):
    buttons = [btn for btn in card.find_elements(
        By.XPATH, ".//button[contains(normalize-space(.), 'View Premium')]"
    ) if _elines_visible(btn)]
    if not buttons:
        raise RuntimeError("Không tìm thấy nút View Premium trên card E-Lines mới")
    driver.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].click();", buttons[0])
    return WebDriverWait(driver, timeout).until(lambda d: _elines_find_visible_premium_modal())


def _elines_close_premium_modal(modal=None):
    modal = modal or _elines_find_visible_premium_modal()
    if not modal:
        return
    close_buttons = modal.find_elements(By.XPATH,
        ".//button[normalize-space(.)='Close' or contains(@aria-label, 'Close')]")
    for button in close_buttons:
        if _elines_visible(button):
            driver.execute_script("arguments[0].click();", button)
            time.sleep(0.2)
            return
    header_buttons = modal.find_elements(By.CSS_SELECTOR, ".el-dialog__headerbtn")
    for button in header_buttons:
        if _elines_visible(button):
            driver.execute_script("arguments[0].click();", button)
            time.sleep(0.2)
            return


def _elines_premium_modal_rows(modal):
    """Read each service row without depending on hashed Vue/Tailwind classes."""
    return driver.execute_script("""
        const modal = arguments[0];
        const rows = [];
        const seen = new Set();
        const normalized = value => (value || '').replace(/\\s+/g, ' ').trim();
        for (const marker of modal.querySelectorAll('*')) {
            const markerText = normalized(marker.innerText || marker.textContent);
            let service = '';
            if (/\\bFLASH SALE\\b/i.test(markerText)) service = 'Flash Sale';
            else if (/\\bSTANDARD SERVICE\\b/i.test(markerText)) service = 'Standard Service';
            if (!service) continue;
            let node = marker;
            while (node && node !== modal) {
                const text = normalized(node.innerText || node.textContent).toUpperCase();
                if (text.includes('20GP') && text.includes('40GP') &&
                    text.includes('40HQ') && text.includes('BOOKING')) break;
                node = node.parentElement;
            }
            if (!node || node === modal || seen.has(service)) continue;
            seen.add(service);
            rows.push({ service, text: normalized(node.innerText || node.textContent), element: node });
        }
        return rows;
    """, modal)


def _elines_read_new_premium_for_card(card):
    modal = _elines_open_premium_modal(card)
    try:
        raw_rows = _elines_premium_modal_rows(modal)
        parsed_rows = []
        for raw in raw_rows:
            parsed = parse_premium_service_row(raw.get("text", ""), raw.get("service", ""))
            parsed["element"] = raw.get("element")
            parsed_rows.append(parsed)
        selected = select_preferred_premium_service(parsed_rows)
        if not selected:
            raise RuntimeError("Modal Premium không có giá 20GP/40GP/40HQ hợp lệ")
        return selected
    finally:
        _elines_close_premium_modal(modal)


def _elines_book_new_premium_service(card, preferred_service):
    modal = _elines_open_premium_modal(card)
    raw_rows = _elines_premium_modal_rows(modal)
    preferred = str(preferred_service or "").strip().upper()
    selected = next((row for row in raw_rows if str(row.get("service", "")).upper() == preferred), None)
    if not selected:
        selected = next((row for row in raw_rows if str(row.get("service", "")).upper() == "FLASH SALE"), None)
    if not selected and raw_rows:
        selected = raw_rows[0]
    if not selected:
        _elines_close_premium_modal(modal)
        raise RuntimeError("Không tìm thấy service để Booking trong modal Premium")
    row_element = selected.get("element")
    buttons = row_element.find_elements(By.XPATH, ".//button[contains(normalize-space(.), 'Booking')]")
    button = next((item for item in buttons if _elines_visible(item)), None)
    if not button:
        _elines_close_premium_modal(modal)
        raise RuntimeError("Không tìm thấy nút Booking của service Premium đã chọn")
    driver.execute_script("arguments[0].click();", button)


def _elines_new_card_vessel_and_transshipment(card):
    text = " ".join((card.text or "").split())
    vessel_match = re.search(
        r"VESSEL\s*/\s*VOYAGE\s*:\s*(.+?)(?=\s+(?:PREMIUM SERVICES|BOOK NOW|VIEW PREMIUM)|$)",
        text,
        flags=re.IGNORECASE,
    )
    vessel = vessel_match.group(1).strip() if vessel_match else "TBA"
    transshipment_match = re.search(
        r"TRANS(?:SHIPMENT)?\s*(?:PORT)?\s*[:\-]\s*(.+?)(?=\s+(?:PREMIUM SERVICES|BOOK NOW|VIEW PREMIUM)|$)",
        text,
        flags=re.IGNORECASE,
    )
    ports = []
    if transshipment_match:
        ports = [part.strip().upper() for part in re.split(r"\s*(?:,|\+|/| OR )\s*", transshipment_match.group(1)) if part.strip()]
    return vessel, ports


def _elines_selected_dropdown_labels(inp):
    """Read selected Element Plus values, including multi-select tags."""
    return driver.execute_script("""
        const input = arguments[0];
        const select = input.closest('.el-select');
        const values = new Set();
        const add = value => {
            const text = (value || '').replace(/\\s+/g, ' ').trim();
            if (text) values.add(text);
        };

        add(input.value);
        if (select) {
            for (const tag of select.querySelectorAll('.el-tag, .el-tag__content, .el-select__tags, .el-select-tags-wrapper')) {
                add(tag.innerText || tag.textContent);
            }
        }

        const listId = input.getAttribute('aria-controls');
        const list = listId ? document.getElementById(listId) : null;
        if (list) {
            for (const item of list.querySelectorAll('[role="option"], .el-select-dropdown__item')) {
                if (item.getAttribute('aria-selected') === 'true' || item.classList.contains('is-selected')) {
                    add(item.innerText || item.textContent);
                }
            }
        }
        return [...values];
    """, inp) or []


def _elines_select_new_dropdown(label_text, option_text, prefix=False):
    """Select an Element Plus option using the label rather than unstable ids."""
    label_xpath = f"//label[contains(normalize-space(.), {repr(label_text)})]"
    input_xpath = label_xpath + "/following-sibling::*[1]//input"
    inp = WebDriverWait(driver, 12).until(
        lambda d: next((item for item in d.find_elements(By.XPATH, input_xpath) if _elines_visible(item)), None)
    )
    selected_labels = _elines_selected_dropdown_labels(inp)
    if dropdown_has_selected_option(selected_labels, option_text, prefix=prefix):
        print(f"      -> [Elines] {label_text}: {option_text} đã có sẵn, không toggle lại.")
        return

    # Element Plus multi-select has a tiny search input (often width ~= 0).
    # Clicking that node by JS does not open the list.  Click the visible input
    # wrapper, and do not click again when the list is already expanded because
    # that would close the toggle before the next container is selected.
    expanded = (inp.get_attribute("aria-expanded") or "").lower() == "true"
    if not expanded:
        select_root = inp.find_element(By.XPATH,
            "./ancestor::*[contains(concat(' ',normalize-space(@class),' '),' el-select ')][1]")
        wrapper = select_root.find_element(By.CSS_SELECTOR, ".el-input__wrapper")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", wrapper)
        ActionChains(driver).move_to_element(wrapper).click().perform()

    if prefix:
        option_xpath = (
            "//*[@role='option' and "
            f"starts-with(normalize-space(.), {repr(option_text)})]"
        )
    else:
        option_xpath = (
            "//*[@role='option' and "
            f"normalize-space(.)={repr(option_text)}]"
        )
    option = WebDriverWait(driver, 8).until(
        lambda d: next((item for item in d.find_elements(By.XPATH, option_xpath) if _elines_visible(item)), None)
    )
    option_classes = set((option.get_attribute("class") or "").split())
    already_selected = (
        (option.get_attribute("aria-selected") or "").lower() == "true"
        or "selected" in option_classes
        or "is-selected" in option_classes
    )
    if not already_selected:
        ActionChains(driver).move_to_element(option).click().perform()
        time.sleep(0.2)


def _elines_choose_new_cargo_and_sizes():
    _elines_select_new_dropdown("Cargo Nature", "General")
    WebDriverWait(driver, 10).until(lambda d: any(
        _elines_visible(item) and item.is_enabled()
        for item in d.find_elements(By.XPATH, "//label[contains(normalize-space(.), 'Size Type')]/following-sibling::*[1]//input")
    ))
    for container in ("20GP", "40GP", "40HQ"):
        _elines_select_new_dropdown("Size Type", container, prefix=True)
    try:
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
    except Exception:
        pass


def _elines_filter_flash_sale_if_available():
    inputs = driver.find_elements(By.XPATH,
        "//input[contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'), 'premium services')]")
    selector = next((item for item in inputs if _elines_visible(item)), None)
    if not selector:
        return False
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].click();", selector)
        option = WebDriverWait(driver, 4).until(lambda d: next((item for item in d.find_elements(
            By.XPATH, "//li[contains(@class,'el-select-dropdown__item') and normalize-space(.)='Flash Sale']"
        ) if _elines_visible(item)), None))
        driver.execute_script("arguments[0].click();", option)
        return True
    except Exception:
        return False


def _elines_has_visible_empty_state():
    selectors = [
        (By.CSS_SELECTOR, ".el-empty"),
    ]
    for by, value in selectors:
        for elem in driver.find_elements(by, value):
            try:
                if elem.is_displayed():
                    return True
            except Exception:
                pass
    return False


def _elines_has_visible_no_products_message():
    patterns = [
        "No products matching your criteria were found",
        "No products matching",
        "Sorry, no matching ocean freight products were found",
    ]
    for pattern in patterns:
        try:
            elems = driver.find_elements(
                By.XPATH,
                f"//*[contains(normalize-space(.), {repr(pattern)})]"
            )
        except Exception:
            elems = []
        for elem in elems:
            try:
                if elem.is_displayed() and (
                    is_no_matching_ocean_freight_message(elem.text)
                    or pattern == "No products matching"
                ):
                    return True
            except Exception:
                pass
    return False


def _elines_has_visible_loading_mask():
    for elem in driver.find_elements(By.CSS_SELECTOR, ".el-loading-mask"):
        try:
            if elem.is_displayed():
                return True
        except Exception:
            pass
    return False


def wait_elines_result_state(timeout=None, empty_confirm=None):
    """Chờ Elines thật sự ổn định: có card thì nhận, empty thì phải đứng yên một lúc mới tin."""
    timeout = timeout or ELINES_RESULT_WAIT_SECONDS
    empty_confirm = empty_confirm or ELINES_EMPTY_CONFIRM_SECONDS
    deadline = time.time() + timeout
    empty_since = None
    last_log = time.time()

    while time.time() < deadline:
        if _elines_visible_result_cards():
            return "CARDS"
        if _elines_has_visible_no_products_message():
            return "NO_PRODUCTS"

        if _elines_has_visible_loading_mask():
            empty_since = None
        elif _elines_has_visible_empty_state():
            if empty_since is None:
                empty_since = time.time()
                print(f"      -> [Elines] Web báo trống, chờ thêm {empty_confirm}s để tránh lag...")
            elif time.time() - empty_since >= empty_confirm:
                return "EMPTY"
        else:
            empty_since = None

        if time.time() - last_log > 10:
            remaining = max(0, int(deadline - time.time()))
            print(f"      -> [Elines] Đang kiên trì chờ kết quả... còn ~{remaining}s")
            last_log = time.time()
        time.sleep(0.5)

    return "TIMEOUT"


# ===================================================================================
# --- 5c. ELINES (ĐÃ FIX SURCHARGE LOGIC) ---
# ===================================================================================
def run_elines(pol, pod, pod_country):
    china_route = is_china_destination(pod_country, pod)
    step_tracker = "Khởi động"
    try:
        driver.switch_to.default_content()

        print("      -> [Elines] BƯỚC 1: Tìm form booking (top-level/iframe)...")
        form_context = _switch_to_elines_booking_frame(timeout=25)
        print(f"      -> [Elines] Form sẵn sàng trong {form_context}.")

        is_new_elines_form = any(
            _elines_visible(item)
            for item in driver.find_elements(By.XPATH, ELINES_ORIGIN_XPATH)
        )

        step_tracker = "Nhập Origin"
        select_port_smart(
            ELINES_ORIGIN_XPATH if is_new_elines_form else "//input[@placeholder='Please input Origin City']",
            pol, "VIETNAM", "Elines", True
        )

        step_tracker = "Nhập Destination"
        select_port_smart(
            ELINES_DESTINATION_XPATH if is_new_elines_form else "//input[@placeholder='Please input Destination City']",
            pod, pod_country, "Elines", True
        )

        if is_new_elines_form:
            step_tracker = "Chọn Cargo Nature + Size Type"
            _elines_choose_new_cargo_and_sizes()
            print("      -> [Elines] Đã chọn General | 20GP | 40GP | 40HQ")

        step_tracker = "Nhập ETD"
        try:
            date_inp = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'el-date-editor')]//input")))
            dt_str = (datetime.now() + timedelta(days=DATE_OFFSET_DAYS)).strftime("%Y-%m-%d")
            driver.execute_script("arguments[0].value='';", date_inp)
            driver.execute_script(f"arguments[0].value='{dt_str}';", date_inp)
            driver.execute_script("arguments[0].dispatchEvent(new Event('input')); arguments[0].dispatchEvent(new Event('change'));", date_inp)            
        except: pass

        if not is_new_elines_form:
            step_tracker = "Tick Container"
            for c in ["20GP", "40GP", "40HQ"]:
                try:
                    lbl = driver.find_element(By.XPATH, f"//label[contains(@class,'el-checkbox') and contains(.,'{c}')]")
                    if "is-checked" not in lbl.get_attribute("class"):
                        driver.execute_script("arguments[0].click();", lbl)
                except: pass

        step_tracker = "Tắt Operating Reefer"
        try:
            reefer_lbl = driver.find_element(By.XPATH, "//label[contains(@class,'el-checkbox') and contains(.,'Operating Reefer')]")
            if "is-checked" in reefer_lbl.get_attribute("class"):
                driver.execute_script("arguments[0].click();", reefer_lbl)
                print("      -> [Elines] Đã TẮT tùy chọn Operating Reefer!")
        except: pass

        step_tracker = "Bấm Search Service"
        driver.execute_script("arguments[0].click();", driver.find_element(By.XPATH, "//button[contains(.,'Search Service')]"))
        print("      -> [Elines] BƯỚC 2: Đã bấm Search...")

        time.sleep(2)
        try:
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".el-loading-mask")))
            WebDriverWait(driver, ELINES_RESULT_WAIT_SECONDS).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".el-loading-mask")))
        except: pass
        time.sleep(1)

        step_tracker = "Kiểm tra kết quả"
        result_state = wait_elines_result_state()
        if result_state == "TIMEOUT":
            print(f"      ❌ [Elines] Quá giờ ({ELINES_RESULT_WAIT_SECONDS}s) web không load ra thẻ giá hoặc báo lỗi trống ổn định!")
            return None
        if result_state == "NO_PRODUCTS":
            print("      -> [Elines] NO SERVICE / SOLD OUT! (No products matching -> bỏ qua ngay)")
            return ELINES_NO_PRODUCTS
        if result_state == "EMPTY":
            print(f"      -> [Elines] NO SERVICE / SOLD OUT! (đã chờ xác nhận {ELINES_EMPTY_CONFIRM_SECONDS}s)")
            return None
        
        step_tracker = "Kiểm tra Flash Sale filter"
        print("      -> [Elines] BƯỚC 3: Kiểm tra Flash Sale...")

        is_flash_sale_mode = _elines_filter_flash_sale_if_available() if is_new_elines_form else False
        has_any_flash_sale = bool(driver.find_elements(By.XPATH,
            "//*[contains(normalize-space(.), 'Flash Sale') and "
            "not(contains(normalize-space(.), 'Flash Sale Remaining Stock'))]"))

        if is_flash_sale_mode:
            print("      -> [Elines] Đã filter Flash Sale trên giao diện mới.")
            filter_state = wait_elines_result_state(
                timeout=max(20, ELINES_RESULT_WAIT_SECONDS // 2),
                empty_confirm=ELINES_EMPTY_CONFIRM_SECONDS,
            )
            if filter_state == "NO_PRODUCTS":
                print("      -> [Elines] Flash Sale không có sản phẩm phù hợp (No products matching).")
                return ELINES_NO_PRODUCTS
            if filter_state != "CARDS":
                print(f"      ⚠️ [Elines] Sau khi filter Flash Sale chưa thấy card ổn định ({filter_state}).")
        elif has_any_flash_sale:
            print("      -> [Elines] 🔥 Phát hiện Flash Sale!")
            try:
                select_wrapper = driver.find_element(By.XPATH,
                    "//div[contains(@class,'el-select__tags')]"
                    "[.//input[contains(@class,'el-select__input')]]")
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", select_wrapper)
                time.sleep(0.3)
                driver.execute_script("arguments[0].click();", select_wrapper)

                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "li.el-select-dropdown__item")))
                time.sleep(0.3)

                flash_option = driver.find_element(By.XPATH,
                    "//li[contains(@class,'el-select-dropdown__item')]"
                    "[.//span[text()='Flash Sale']]")
                driver.execute_script("arguments[0].click();", flash_option)
                print("      -> [Elines] Đã chọn Flash Sale!")

                try:
                    WebDriverWait(driver, 3).until(EC.presence_of_element_located(
                        (By.CSS_SELECTOR, ".el-loading-mask")))
                    WebDriverWait(driver, max(20, ELINES_RESULT_WAIT_SECONDS // 2)).until(EC.invisibility_of_element_located(
                        (By.CSS_SELECTOR, ".el-loading-mask")))
                except:
                    pass
                filter_state = wait_elines_result_state(
                    timeout=max(20, ELINES_RESULT_WAIT_SECONDS // 2),
                    empty_confirm=ELINES_EMPTY_CONFIRM_SECONDS,
                )
                if filter_state == "NO_PRODUCTS":
                    print("      -> [Elines] Flash Sale không có sản phẩm phù hợp (No products matching).")
                    return ELINES_NO_PRODUCTS
                if filter_state != "CARDS":
                    print(f"      ⚠️ [Elines] Sau khi filter Flash Sale chưa thấy card ổn định ({filter_state}).")
                time.sleep(1)
                is_flash_sale_mode = True
            except Exception as e:
                print(f"      ⚠️ [Elines] Không filter được Flash Sale: {e}")
        else:
            print("      -> [Elines] Không có Flash Sale, dùng giá thường.")

        step_tracker = "Đọc dữ liệu cards (DEBUG MODE)"
        print("      -> [Elines] Bật chế độ DEBUG soi thẻ card...")
        
        cards = _elines_visible_result_cards()
        list_c = []
        new_ui_card_count = 0
        new_ui_without_premium_count = 0

        if not cards:
            print("      ❌ [DEBUG] Màn hình hiện tại không có thẻ card nào (XPath box-border không khớp)!")
            return None

        # FIX: Biến năm tăng theo `datetime.now().year` thay vì hardcode "2026".
        # COSCO web chỉ hiện thị format `MMMDD` (vd: `JAN05`) không có năm → phải gắn năm thủ công.
        _NOW = datetime.now()
        _Y = _NOW.year

        def _parse_etd_cosco(raw):
            """Parse '2026Jan05' style → datetime. Tự cuốn sang năm sau nếu ngày đã rơi vào quá khứ > 60 ngày."""
            dt = datetime.strptime(f"{_Y}{raw}", "%Y%b%d")
            if (dt - _NOW).days < -60:
                dt = dt.replace(year=_Y + 1)
            return dt

        for idx, card in enumerate(cards):
            if not card.is_displayed():
                continue

            # FIX: Chỉ ghi file debug HTML khi có ENV COSCO_DEBUG_HTML=1.
            # Trước đây ghi mỗi lần đọc card đầu → rác IO + rác file.
            if idx == 0 and os.environ.get("COSCO_DEBUG_HTML") == "1":
                try:
                    html_content = card.get_attribute('outerHTML')
                    with open("debug_card.html", "w", encoding="utf-8") as f:
                        f.write(html_content)
                    print("      ✅ [DEBUG] Đã lưu HTML thẻ 1 vào 'debug_card.html'.")
                except Exception:
                    pass

            try:
                # 1. Đọc Ngày đi (ETD)
                try:
                    if _elines_card_uses_new_ui(card):
                        new_ui_card_count += 1
                        if not _elines_card_has_view_premium(card):
                            new_ui_without_premium_count += 1
                            schedule = parse_new_elines_card_schedule(card.text, now=_NOW)
                            print(
                                f"        Card {idx+1}: {schedule['etd_dt'].strftime('%d-%b')} "
                                "không có View Premium -> không có giá Premium, bỏ ngay."
                            )
                            continue
                        schedule = parse_new_elines_card_schedule(card.text, now=_NOW)
                        premium = _elines_read_new_premium_for_card(card)
                        premium_rates = premium.get("rates", {})
                        prices = [value for value in premium_rates.values() if value and value > 50]
                        if not prices:
                            raise RuntimeError("Modal Premium không trả về giá hợp lệ")
                        list_c.append({
                            "element": card,
                            "price": min(prices),
                            "etd_dt": schedule["etd_dt"],
                            "tt_days": schedule["tt_days"],
                            "space": schedule["space"],
                            "premium_rates": premium_rates,
                            "premium_service": premium.get("service", "Standard Service"),
                            "remaining_stock": premium.get("remaining_stock"),
                        })
                        print(
                            f"        Card {idx+1}: {schedule['etd_dt'].strftime('%d-%b')} "
                            f"TT={schedule['tt_days']}d ${min(prices)} "
                            f"[{schedule['space']}] {premium.get('service', '')}"
                        )
                        continue

                    etd_raw = card.find_element(By.XPATH,
                        ".//div[contains(@class,'bg-[#1890FF]')]"
                        "//p[contains(@class,'bottom-2.5')]//span[1]"
                    ).text.split('(')[0].strip()
                    etd_dt = _parse_etd_cosco(etd_raw)
                except Exception as e_etd:
                    raise Exception(f"Kẹt ở lúc đọc Ngày ETD: {e_etd}")

                # 2. Đọc Transit Time (TT)
                tt_days = 999
                try:
                    tt_el = card.find_element(By.XPATH,
                        ".//p[contains(@class,'space-x-0.5')]//span[1]")
                    tt_text = tt_el.text.strip()
                    if tt_text.isdigit():
                        tt_days = int(tt_text)
                except:
                    print(f"        [DEBUG] Thẻ {idx+1}: Không tìm thấy Transit time bằng chữ, thử trừ từ ETA...")
                    try:
                        eta_raw = card.find_element(By.XPATH,
                            ".//div[contains(@class,'bg-[#CCCCCC]')]"
                            "//p[contains(@class,'bottom-2.5')]//span[1]"
                        ).text.split('(')[0].strip()
                        eta_dt = _parse_etd_cosco(eta_raw)
                        tt_days = (eta_dt - etd_dt).days
                    except Exception as e_eta:
                        print(f"        [DEBUG] Thẻ {idx+1}: ETA cũng móm nốt! Lỗi: {e_eta}")

                # 3. Đọc Giá tiền
                price_spans = card.find_elements(By.XPATH,
                    ".//p[contains(@class,'text-[#F1A104]')]//span[last()]")
                
                if not price_spans:
                    raise Exception("Mảng price_spans bị trống (Không tìm thấy XPath chứa màu chữ vàng #F1A104)")
                    
                prices = []
                for p in price_spans:
                    txt = p.text.replace(',', '').strip()
                    try:
                        val = float(txt)
                        if val > 50:
                            prices.append(val)
                    except:
                        pass

                if not prices:
                    raise Exception(f"Có element giá nhưng không lôi được số ra. Giá trị thô đang lấy được là: {[p.text for p in price_spans]}")

                # 4. Đọc tình trạng chỗ (Space)
                try:
                    space_txt = card.find_element(By.XPATH,
                        ".//p[contains(@class,'text-warning')]//span").text.strip().upper()
                except:
                    space_txt = "TBC"

                list_c.append({
                    "element": card,
                    "price":   min(prices),
                    "etd_dt":  etd_dt,
                    "tt_days": tt_days,
                    "space":   space_txt,
                })
                print(f"        Card {idx+1}: {etd_raw} TT={tt_days}d ${min(prices)} [{space_txt}]")
                
            except Exception as e:
                # IN RA LỖI RÕ RÀNG ĐỂ BẮT BỆNH
                print(f"        ⚠️ [DEBUG] Thẻ số {idx+1} bị rớt đài do: {e}")
                continue

        if not list_c:
            if new_ui_card_count and new_ui_without_premium_count == new_ui_card_count:
                print(
                    "      -> [Elines] Tất cả card chỉ có Book Now, không có View Premium "
                    "-> xác nhận không có giá E-Lines."
                )
                return ELINES_NO_PREMIUM_PRICE
            print("      ❌ [Elines] Không đọc được card nào hoàn chỉnh!")
            return None

        step_tracker = "Lọc 9 quy tắc"
        
        space_priority = {
            "TIGHT": 1,
            "TBC": 2
        }
        
        # --- BẮT ĐẦU SỬA: LỌC TEU VÀ CHỌN GIÁ RẺ NHẤT CÓ ĐỦ CHỖ ---
        import re
        valid_list_re = []
        
        # Lấy danh sách các mức giá từ rẻ đến mắc (Ví dụ: $425 test trước, $525 test sau)
        cac_muc_gia = sorted(list(set(c["price"] for c in list_c)))
        
        for muc_gia in cac_muc_gia:
            # Lấy các card có mức giá này
            list_re_temp = [c for c in list_c if c["price"] == muc_gia]
            
            # Sắp xếp để test card AVAILABLE trước, card TBC sau trong cùng 1 mức giá
            list_re_temp.sort(key=lambda x: space_priority.get(x["space"].upper(), 3), reverse=True)
            
            if is_flash_sale_mode or any(
                str(c.get("premium_service", "")).strip().upper() == "FLASH SALE"
                for c in list_re_temp
            ):
                print(f"      -> [Elines] Xét mức giá ${muc_gia}: Kiểm tra Remaining Stock (yêu cầu >= 5 TEU)...")
                valid_for_this_price = []
                
                for c in list_re_temp:
                    if "premium_rates" in c:
                        if str(c.get("premium_service", "")).strip().upper() == "FLASH SALE":
                            teu_val = c.get("remaining_stock")
                            if teu_val is None:
                                print(f"        ⚠️ Bỏ card ETD {c['etd_dt'].strftime('%d-%b')}: không đọc được Remaining Stock.")
                                continue
                            if teu_val < 5:
                                print(f"        ⚠️ Bỏ card ETD {c['etd_dt'].strftime('%d-%b')} [{c['space']}] vì chỉ còn {teu_val} TEU (< 5)")
                                continue
                            print(f"        ✅ Card ETD {c['etd_dt'].strftime('%d-%b')} [{c['space']}] còn {teu_val} TEU -> Đạt chuẩn!")
                            valid_for_this_price.append(c)
                        elif not is_flash_sale_mode:
                            valid_for_this_price.append(c)
                        continue

                    card = c["element"]
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", card)
                    time.sleep(0.4)
                    
                    # 1. Bơm JS ép click mở panel Flash Sale (nếu có)
                    driver.execute_script("""
                        let card = arguments[0];
                        let tags = card.querySelectorAll('.blue-service, .blue-service-selected');
                        for(let t of tags) {
                            if(t.textContent.includes('Flash Sale')) {
                                t.click();
                            }
                        }
                    """, card)
                    time.sleep(1) # Nghỉ 1s chờ panel bung ra trọn vẹn
                    
                    # 2. Bơm JS đọc xuyên DOM lấy cụm text "Remaining Stock... TEU"
                    teu_text = driver.execute_script("""
                        let card = arguments[0];
                        let h4s = card.querySelectorAll("h4");
                        for (let h of h4s) {
                            if (h.textContent.includes("Remaining Stock")) {
                                return h.parentElement.textContent; 
                            }
                        }
                        return null;
                    """, card)
                    
                    if teu_text:
                        # Dùng regex trích xuất đúng con số nằm trước chữ TEU
                        match = re.search(r'(\d+)\s*TEU', teu_text, re.IGNORECASE)
                        if match:
                            teu_val = int(match.group(1))
                            if teu_val < 5:
                                print(f"        ⚠️ Bỏ qua card ETD {c['etd_dt'].strftime('%d-%b')} [{c['space']}] vì chỉ còn {teu_val} TEU (< 5)")
                                # Đóng bảng Flash Sale của Card bị loại để dọn DOM
                                driver.execute_script("""
                                    let tags = arguments[0].querySelectorAll('.blue-service, .blue-service-selected');
                                    for(let t of tags) {
                                        if(!t.textContent.includes('Flash Sale')) {
                                            t.click();
                                            break;
                                        }
                                    }
                                """, card)
                                time.sleep(0.5)
                                continue
                            else:
                                print(f"        ✅ Card ETD {c['etd_dt'].strftime('%d-%b')} [{c['space']}] còn {teu_val} TEU -> Đạt chuẩn!")
                        else:
                            print(f"        ⚠️ Có chữ Remaining Stock nhưng không trích xuất được số từ: {teu_text}")
                            continue # Lỗi format của hãng tàu -> Không an toàn -> Vứt!
                    else:
                        print(f"        ⚠️ Không tìm thấy dòng Remaining Stock của card ETD {c['etd_dt'].strftime('%d-%b')}. Sẽ bỏ qua thẻ này.")
                        continue # Kỷ luật sắt: Không đọc được cũng vứt luôn!
                    
                    # Nếu chạy đến được dòng này tức là card đã xuất sắc vượt qua test
                    valid_for_this_price.append(c)
                
                if valid_for_this_price:
                    valid_list_re = valid_for_this_price
                    break # Đã tìm thấy các card hợp lệ ở mức giá này, chốt và thoát vòng lặp giá
            else:
                # Nếu giá THƯỜNG (không check được TEU) thì vẫn phải lọc ưu tiên SPACE (Bỏ TIGHT, lấy AVAILABLE)
                max_prio_temp = max(space_priority.get(c["space"].upper(), 3) for c in list_re_temp)
                valid_list_re = [c for c in list_re_temp if space_priority.get(c["space"].upper(), 3) == max_prio_temp]
                break
                
        if not valid_list_re:
            print("      ❌ [Elines] Tất cả các card Flash Sale ở mọi mức giá đều không đủ điều kiện (Dưới 5 TEU hoặc lỗi)!")
            return None
            
        list_re = valid_list_re
        e_chuan, s_etd, s_tt = apply_9_golden_rules(list_re)
        # --- KẾT THÚC SỬA ---       
       
        valid_date_str = calculate_validity(e_chuan[-1]["etd_dt"]) # <-- THÊM DÒNG NÀY
        print(f"      -> [Elines] Chốt ETD: {s_etd} | T/T: {s_tt} days")
        
        # --- BẮT ĐẦU THÊM: LẤY THÔNG TIN TÀU & TRANSSHIPMENT PORT ---
        step_tracker = "Lấy Thông Tin Tàu"
        print("      -> [Elines] Đang soi tên tàu và cảng chuyển tải...")
        vessel_texts = []
        ts_combos = []

        # Import bộ giả lập chuột người thật
        from selenium.webdriver.common.action_chains import ActionChains

        for c in e_chuan:
            card = c["element"]
            # Cuộn trang cho thẻ Card nằm giữa màn hình để dễ rê chuột
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", card)
            time.sleep(0.4)

            if _elines_card_uses_new_ui(card):
                vessel_name, ts_ports = _elines_new_card_vessel_and_transshipment(card)
                str_ts = " + ".join(ts_ports) if ts_ports else "DIRECT"
                str_etd_fmt = f"{c['etd_dt'].day}-{c['etd_dt'].strftime('%b')}"
                info_str = (
                    f"{vessel_name} / ETD: {str_etd_fmt} / "
                    f"Transit time: {c['tt_days']} Days / Transshipment Port: {str_ts}"
                )
                vessel_texts.append(info_str)
                ts_combos.append(str_ts)
                continue
            
            # 1. Lấy Tên Tàu (Cắt bỏ mã Service ở đầu)
            vessel_name = "TBA"
            try:
                v_els = card.find_elements(By.XPATH, ".//span[contains(@class, 'text-dark-black') and contains(@class, 'font-bold')]")
                for v in v_els:
                    t = v.text.strip().replace('\n', ' ')
                    if t and not t.startswith("202"):
                        # Chặt chuỗi ở khoảng trắng đầu tiên (Bỏ HPX2, VTS...)
                        parts = t.split(" ", 1)
                        vessel_name = parts[1] if len(parts) > 1 else t
                        break
            except: pass
                
           
            # 2. KÍCH HOẠT VÀ HÚT DATA POPUP HOÀN TOÀN BẰNG JAVASCRIPT (TRỊ LỖI TEXT TÀNG HÌNH)
            ts_ports = []
            try:
                days_el = card.find_element(By.XPATH, ".//*[contains(translate(text(), 'DAY', 'day'), 'day')]")
                
                # Bơm JS kích hoạt TẤT CẢ các thẻ có khả năng là nút bấm trong khu vực đó
                trigger_ok = driver.execute_script("""
                    let box = arguments[0].closest('.flex-grow') || arguments[0].parentElement.parentElement;
                    // Bắt trọn ổ: class trigger của tooltip, hoặc class cursor-pointer
                    let triggers = box.querySelectorAll('.el-tooltip__trigger, svg.cursor-pointer');
                    if (triggers.length > 0) {
                        triggers[0].scrollIntoView({block: 'center', inline: 'center'});
                        triggers.forEach(t => {
                            // Bắn liên thanh cả hover lẫn click để ép popup phải lòi ra
                            t.dispatchEvent(new MouseEvent('mouseenter', {bubbles: true, cancelable: true, view: window}));
                            t.dispatchEvent(new MouseEvent('click', {bubbles: true, cancelable: true, view: window}));
                        });
                        return true;
                    }
                    return false;
                """, days_el)
                
                if trigger_ok:
                    time.sleep(1.5) # Đợi 1.5s cho popup vẽ xong
                    
                    # CẬP NHẬT: Dò bằng textContent và quét bao quát hơn
                    st_names = driver.execute_script("""
                        let pops = document.querySelectorAll('.el-popper');
                        let names = [];
                        
                        // Quét ngược từ dưới lên (ưu tiên lấy popup vừa mới nặn ra)
                        for (let i = pops.length - 1; i >= 0; i--) {
                            let p = pops[i];
                            
                            // Bỏ qua nếu popup bị ẩn
                            let isHidden = window.getComputedStyle(p).display === 'none' || p.getAttribute('aria-hidden') === 'true';
                            if (isHidden) continue;
                            
                            // Tìm tất cả các trạm
                            let spans = p.querySelectorAll("span.font-bold.text-black");
                            if (spans.length > 0) {
                                spans.forEach(s => {
                                    // Dùng textContent để lấy chữ thô tuyệt đối (Không lo CSS che khuất)
                                    let txt = s.textContent || "";
                                    if(txt.trim()) names.push(txt.trim().toUpperCase());
                                });
                                return names; // Hút xong của popup này là té luôn
                            }
                        }
                        return names;
                    """)
                    
                    if st_names:
                        if len(st_names) > 2:
                            ts_ports = st_names[1:-1] # Cắt bỏ Origin và Dest, lấy phần ruột chuyển tải
                    else:
                        print("      [Debug] JS đã kích hoạt nhưng lúc móc túi thì popup trống rỗng!")
                else:
                    print("      [Debug] JS không tìm thấy cái nút trigger nào!")
                        
                # Dọn dẹp: Bấm ra ngoài và bắn sự kiện leave để đóng popup cũ lại
                driver.execute_script("document.body.click();")
                driver.execute_script("""
                    let box = arguments[0].closest('.flex-grow') || arguments[0].parentElement.parentElement;
                    let triggers = box.querySelectorAll('.el-tooltip__trigger, svg.cursor-pointer');
                    triggers.forEach(t => {
                        t.dispatchEvent(new MouseEvent('mouseleave', {bubbles: true, cancelable: true, view: window}));
                    });
                """, days_el)
                time.sleep(0.3)
                
            except Exception as e:
                print(f"      [Debug] Lỗi rình popup Cảng: {e}")

            str_ts = " + ".join(ts_ports) if ts_ports else "DIRECT"
            str_etd_fmt = f"{c['etd_dt'].day}-{c['etd_dt'].strftime('%b')}"
            str_tt_fmt = str(c["tt_days"])
            
            info_str = f"{vessel_name} / ETD: {str_etd_fmt} / Transit time: {str_tt_fmt} Days / Transshipment Port: {str_ts}"
            vessel_texts.append(info_str)
            ts_combos.append(str_ts)

        unique_ts = []
        for t in ts_combos:
            if t not in unique_ts: unique_ts.append(t)
            
        final_ts_str = " or \n".join(unique_ts)
        final_vessel_str = "\n".join(vessel_texts)
        # --- KẾT THÚC THÊM ---

        selected_candidate = e_chuan[0]
        target_card = selected_candidate["element"]
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target_card)
        time.sleep(0.5)

        step_tracker = "Đọc giá + Booking"
        rates = {"20GP": None, "40GP": None, "40HQ": None}
        formula_parts = {"20GP": [], "40GP": [], "40HQ": []}
        b_btn = None

        if _elines_card_uses_new_ui(target_card):
            premium_rates = selected_candidate.get("premium_rates", {})
            for ctype in rates:
                value = premium_rates.get(ctype)
                if value is not None:
                    rates[ctype] = value
                    formula_parts[ctype].append(value)
                    print(f"        [{selected_candidate.get('premium_service', 'Premium')}] {ctype} = ${value}")
            print(f"      -> [Elines] Bấm Booking {selected_candidate.get('premium_service', 'Premium')}...")
            _elines_book_new_premium_service(target_card, selected_candidate.get("premium_service"))
            time.sleep(0.5)
        elif is_flash_sale_mode:
            try:
                flash_panel = target_card.find_element(By.ID, "spot-booking-premium-service")
                if not flash_panel.is_displayed():
                    raise Exception("panel ẩn")
            except:
                try:
                    fb = target_card.find_element(By.XPATH,
                        ".//div[contains(@class,'blue-service')][contains(text(),'Flash Sale')]")
                    driver.execute_script("arguments[0].click();", fb)
                    WebDriverWait(driver, 8).until(
                        EC.presence_of_element_located((By.ID, "spot-booking-premium-service")))
                    time.sleep(1)
                except Exception as e:
                    print(f"      ⚠️ Không mở Flash Sale panel: {e}")

            try:
                flash_panel = target_card.find_element(By.ID, "spot-booking-premium-service")
                cont_blocks = flash_panel.find_elements(By.XPATH,
                    ".//div[contains(@class,'flex-col') and contains(@class,'items-start')]"
                    "[.//span[contains(@class,'text-[#666666]')]]")
                for block in cont_blocks:
                    try:
                        ctype = block.find_element(By.XPATH,
                            ".//span[contains(@class,'text-[#666666]')]").text.strip()
                        if ctype not in rates: continue
                        val = block.find_element(By.XPATH,
                            ".//div[contains(@class,'text-[#F1A104]')]//span[last()]"
                        ).text.replace(',', '').strip()
                        try:
                            rates[ctype] = float(val)
                            formula_parts[ctype].append(rates[ctype])
                            print(f"        [🔥FLASH] {ctype} = ${rates[ctype]}")
                        except:
                            pass
                            print(f"        [🔥FLASH] {ctype} = ${rates[ctype]}")
                    except:
                        continue
            except Exception as e:
                print(f"      ⚠️ Không đọc giá Flash Sale: {e}")

            print("      -> [Elines] Bấm Booking Flash Sale...")
            b_btn = WebDriverWait(target_card, 5).until(EC.presence_of_element_located((
                By.XPATH, ".//button[starts-with(@id,'bkg2-premium-service-booking-btn-')]")))

        else:
            cont_blocks = target_card.find_elements(By.XPATH,
                ".//div[contains(@class,'flex-col') and contains(@class,'flex-1')]"
                "[.//span[contains(@class,'text-[#666666]')]]")
            for block in cont_blocks:
                try:
                    ctype = block.find_element(By.XPATH,
                        ".//span[contains(@class,'text-[#666666]')]").text.strip()
                    if ctype not in rates: continue
                    val = block.find_element(By.XPATH,
                        ".//p[contains(@class,'text-[#F1A104]')]//span[last()]"
                    ).text.replace(',', '').strip()
                    try:
                        rates[ctype] = float(val)
                        formula_parts[ctype].append(rates[ctype])
                        print(f"        [THƯỜNG] {ctype} = ${rates[ctype]}")
                    except:
                        pass
                        print(f"        [THƯỜNG] {ctype} = ${rates[ctype]}")
                except:
                    continue

            print("      -> [Elines] Bấm Booking thường...")
            b_btn = target_card.find_element(By.XPATH, ".//button[@id='bkg2-spot-booking-btn']")

        if b_btn is not None:
            driver.execute_script("arguments[0].click();", b_btn)
        print("      -> [Elines] ĐÃ BẤM BOOKING! Chờ form load...")

        # ================================================================
        # BƯỚC 7: TRANG DETAIL - NHẬP CÂN
        # ================================================================
        step_tracker = "Trang Detail - Điền cân"
        driver.switch_to.default_content()

        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, "aczoneIframe")))
        except:
            pass

        try:
            iframe_detail = driver.find_element(By.ID, "aczoneIframe")
            driver.switch_to.frame(iframe_detail)
            print("      -> [Elines] Vào iframe Detail!")
        except:
            print("      -> [Elines] Không thấy iframe Detail, tiếp tục...")

        print("      -> [Elines] Bắt đầu dò tìm ô nhập cân (Tốc độ 0.1s)...")

        _elines_fill_weight_fast(driver)

        try:
            WebDriverWait(driver, 15).until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(),'Container Info') or contains(text(),'Cargo Weight')]")))
        except:
            pass

        step_tracker = "Trang Detail - Calculate"
        try:
            calc_btn = driver.find_element(By.XPATH,
                "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'calculate')]")
            driver.execute_script("arguments[0].click();", calc_btn)
            print("      -> [Elines] Đã bấm Calculate!")
        except:
            pass

        # ================================================================
        # BƯỚC 8: HÚT SURCHARGE - LOGIC CHỜ DỮ LIỆU ỔN ĐỊNH 
        # ================================================================
        step_tracker = "Hút Surcharge"
        
        # 1. Chờ vòng xoay (max 2s)
        try:
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".el-loading-mask")))
        except: pass

        # 2. Chờ vòng xoay biến mất (max 10s)
        try:
            WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".el-loading-mask")))
        except: pass

        # 3. POLLING CHỜ ỔN ĐỊNH (Bất chấp web trả về bao nhiêu phí)
        print("      -> [Elines] Đang rình bảng Surcharge (Chờ dữ liệu ngừng nhảy dòng)...")
        timeout_content = time.time() + 15
        
        last_row_count = 0
        stable_time = 0
        
        while time.time() < timeout_content:
            status = driver.execute_script("""
                let rows = document.querySelectorAll('table.el-table__body tr.el-table__row');
                let tableText = document.querySelector('table.el-table__body') ? document.querySelector('table.el-table__body').innerText.toUpperCase() : '';
                return {
                    count: rows.length,
                    has_surcharge: tableText.includes('SURCHARGE'),
                    has_ocean: tableText.includes('OCEAN')
                };
            """)
            
            curr_count = status['count']
            
            # Phải có đủ cả Ocean và Surcharge mới bắt đầu kiểm tra
            if curr_count > 0 and status['has_ocean'] and status['has_surcharge']:
                if curr_count == last_row_count:
                    # Nếu số lượng dòng không đổi trong 1 giây (5 nhịp x 0.2s) -> Web đã bung xong data
                    stable_time += 0.2
                    if stable_time >= 1.0:
                        break
                else:
                    # Đang đếm mà web đẻ thêm dòng mới -> Reset, đếm lại 1 giây từ đầu!
                    last_row_count = curr_count
                    stable_time = 0
            else:
                stable_time = 0
                
            time.sleep(0.2)

        print(f"      -> [Elines] Bảng đã ổn định ({last_row_count} dòng)! Bắt đầu hút...")

        rates_detail = {"20GP": None, "40GP": None, "40HQ": None}
        detail_formula_parts = {"20GP": [], "40GP": [], "40HQ": []}
        has_ows = False
        manifest_fee_found = False
        origin_thc_payable = False
        
        js_code = """
        let data = [];
        let rows = document.querySelectorAll("table.el-table__body tr.el-table__row");
        for (let r of rows) {
            let tds = r.querySelectorAll("td");
            if (tds.length < 8) continue;
            let c_name_idx = tds.length >= 11 ? 1 : 0;
            let c_name = tds[c_name_idx].innerText.trim().toUpperCase();
            let unit = tds[tds.length - 6].innerText.trim().toUpperCase();
            let term = tds[tds.length - 5].innerText.trim().toUpperCase();
            let currency = "";
            let amount = "";
            let price_td = tds[tds.length - 3];
            let spans = price_td.querySelectorAll("span");
            let valid_spans = [];
            for (let s of spans) {
                if (s.innerText.trim() !== "") valid_spans.push(s.innerText.trim());
            }
            if (valid_spans.length >= 2) {
                currency = valid_spans[0].toUpperCase();
                amount = valid_spans[1].replace(/,/g, '');
            }
            data.push({c_name: c_name, unit: unit, term: term, currency: currency, amount: amount});
        }
        return data;
        """
        table_data = driver.execute_script(js_code)
        if os.environ.get("COSCO_DEBUG_ELINES_TABLE", "").strip().lower() in {"1", "true", "yes", "y"}:
            print("      -> [Elines][DEBUG] Raw surcharge rows:")
            for idx, row in enumerate(table_data, start=1):
                print(
                    f"        [ROW {idx:02d}] name={row.get('c_name','')!r} "
                    f"unit={row.get('unit','')!r} term={row.get('term','')!r} "
                    f"currency={row.get('currency','')!r} amount={row.get('amount','')!r}"
                )
      

        # PYTHON XỬ LÝ DỮ LIỆU TỪ JS TRÊN RAM
        for row in table_data:
            c_name = row['c_name']
            unit = row['unit']
            term = row['term']
            curr = row['currency']
            amt_raw = row['amount']

            arbitrary_charge = is_arbitrary_charge(c_name)
            if arbitrary_charge and unit not in ["20GP", "40GP", "40HQ"] and term in ["20GP", "40GP", "40HQ"]:
                print(f"        [FIX-ARB] {c_name[:25]} unit={unit} -> container={term}")
                unit = term
                term = "PREPAID / COLLECT"

            is_terminal_charge = (
                "THC" in c_name
                or "TERMINAL HANDLING" in c_name
                or "TERMINAL" in c_name
            )
            is_destination_terminal = any(x in c_name for x in [
                "DISCHARGE", "DESTINATION", "IMPORT", "THD", "DTHC"
            ])
            is_origin_thc_charge = is_terminal_charge and not is_destination_terminal
            if is_origin_thc_charge and unit in ["20GP", "40GP", "40HQ"] and amt_raw:
                origin_thc_payable = True

            if term == "COLLECT" and not arbitrary_charge and not (china_route and is_origin_thc_charge):
                print(f"        [-COLLECT] Bỏ qua {c_name[:25]} (Do term là COLLECT)")
                continue
            
            if any(x in c_name for x in ["ENS", "AMS", "AFS", "ADVANCED MANIFEST", "ENTRY SUMMARY"]):
                manifest_fee_found = True
                print(f"        [FLAG] Đã tóm được {c_name[:25]} -> Ghi chú vào Remark!")
            
            if any(b in c_name for b in BLOCKLIST_CHARGES) and not (china_route and is_origin_thc_charge):
                print(f"        [-BLOCKLIST] Bỏ qua {c_name[:25]}")
                continue
           
            if "BL" in unit or "B/L" in unit or unit not in ["20GP", "40GP", "40HQ"]:
                print(f"        [-BL] Bỏ qua {c_name[:25]} (Do tính theo B/L hoặc sai đơn vị)")
                continue

            if not amt_raw: 
                print(f"        ⚠️ [BỎ QUA] Phí {c_name[:15]} không có giá tiền (Rỗng).")
                continue
            try:
                amt = float(amt_raw)
            except:
                continue

            # ĐỔI TIỀN LAZY (Chỉ khi nào thấy EUR/AUD/VND mới móc hàm tỷ giá ra tính)
            if curr == 'EUR': 
                amt *= get_live_exchange_rate("EUR", "USD")
            elif curr == 'AUD': 
                amt *= get_live_exchange_rate("AUD", "USD")
            elif curr == 'VND': 
                amt *= get_live_exchange_rate("VND", "USD")
            elif curr == 'CHF':
                amt *= get_live_exchange_rate("CHF", "USD")

            if "OCEAN RATE" in c_name or "OCEAN FREIGHT" in c_name:
                rates_detail[unit] = amt
                detail_formula_parts.setdefault(unit, []).append(amt)
                print(f"        [Ocean] {unit} = ${amt:.2f}")
                continue

            if is_overweight_charge(c_name):
                has_ows = True
                print(f"        [-OWS] Bỏ qua {c_name[:25]} -> chỉ ghi SUBJECT TO OWS")
                continue

            if rates_detail[unit] is not None:
                rates_detail[unit] += amt
                detail_formula_parts.setdefault(unit, []).append(amt)
                print(f"        [+SURCHARGE] +${amt:.2f} ({c_name[:25]}) → {unit}")
            else:
                rates_detail[unit] = amt
                detail_formula_parts.setdefault(unit, []).append(amt)
                print(f"        [+SURCHARGE] +${amt:.2f} ({c_name[:25]}) → {unit}")

        # ─── Merge: ưu tiên Detail, fallback về giá card ───
        final_rates = {}
        for ct in ["20GP", "40GP", "40HQ"]:
            final_rates[ct] = (rates_detail[ct]
                               if rates_detail[ct] is not None
                               else rates.get(ct))
        final_formulas = {}
        for ct in ["20GP", "40GP", "40HQ"]:
            parts = detail_formula_parts.get(ct) if rates_detail.get(ct) is not None else formula_parts.get(ct)
            final_formulas[ct] = _excel_formula_from_parts(parts or [])
        # Kiểm tra xem có lấy được Surcharge nào ngoài Ocean không
        surcharge_error_flag = True
        for row in table_data:
            name = row['c_name'].upper()
            if name and not any(x in name for x in ["OCEAN RATE", "OCEAN FREIGHT"]):
                surcharge_error_flag = False
                break

        src = "🔥Flash Sale" if is_flash_sale_mode else "Thường"
        print(f"      -> [Elines] ✅ XONG [{src}]! Cước cuối: {final_rates}")
        
         # ─── Back ───
        step_tracker = "Back"
        try:
            b_back = driver.find_element(By.XPATH, "//button[contains(.,'Back')]")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", b_back)
            time.sleep(0.4)
            driver.execute_script("arguments[0].click();", b_back)
            time.sleep(1)
        except:
            driver.get(ELINES_BOOKING_URL)

        return {
            "rates": final_rates,
            "etd":   s_etd,
            "tt":    s_tt,
            "ows":   has_ows,
            "thc_inc": bool(china_route or not origin_thc_payable),
            "valid": valid_date_str, 
            "manifest_fee": manifest_fee_found,
            "formulas": final_formulas,
            "surcharge_error": surcharge_error_flag, # <-- CỜ BÁO LỖI
            "vessel_info": final_vessel_str,         # <-- TÊN TÀU
            "transshipment": final_ts_str            # <-- CẢNG CHUYỂN TẢI
        }

    except Exception as e:
        print("\n" + "!"*60)
        print(f"      🚨 ELINES CRASH: [{step_tracker}]")
        print(f"      👉 {e}")
        print(traceback.format_exc())
        print("!"*60 + "\n")
        driver.get(ELINES_BOOKING_URL)
        return None
    

# ===================================================================================
# --- 6. HÀM TẠO REMARK CHUẨN (CÓ BÁO LỖI SURCHARGE) ---
# ===================================================================================
def generate_remark(thc_inc, ows, manifest_fee, pod, source, is_surcharge_error=False):
    rem = build_subject_remark(othc_included=thc_inc, pod=pod, ows=bool(ows))
    
    if is_surcharge_error:
        rem += " (SURCHARGE ERROR)"
        
    return rem

# ===================================================================================
# --- MAIN RUN ---
# ===================================================================================
print("\n>>> BOT COSCO - FIX SURCHARGE v2 <<<")

EXCEL_FILE = os.environ.get("EXCEL_PATH", "input_gia.xlsx")
FILTER_POL = os.environ.get("FILTER_POL", "").strip().upper()
FILTER_POD = os.environ.get("FILTER_POD", "").strip().upper()
wb = openpyxl.load_workbook(EXCEL_FILE)
sheet = wb.active

# --- BẮT ĐẦU THÊM: Đọc dữ liệu từ file FREE TIME.xlsx ---
free_time_dict = {}
try:
    wb_ft = openpyxl.load_workbook("FREE TIME.xlsx", data_only=True)
    sheet_ft = wb_ft.active
    for r in range(2, sheet_ft.max_row + 1):
        key = str(sheet_ft.cell(row=r, column=1).value or "").strip().upper()
        val = str(sheet_ft.cell(row=r, column=2).value or "").strip()
        if key:
            free_time_dict[key] = val
except Exception as e:
    print(f"⚠️ Lỗi khi đọc file FREE TIME.xlsx: {e}")
# --- KẾT THÚC THÊM ---

failed_rows = []  # Mảng lưu lại các dòng bị báo lỗi NO SERVICE để chạy phiên 2

# ---------------------------------------------------------
# PHIÊN CHẠY 1: CHẠY ĐỒNG THỜI SYNCONHUB & ELINES
# ---------------------------------------------------------
print(f"\n{'='*52}")
print("▶️ BẮT ĐẦU PHIÊN CHẠY 1 (SYNCONHUB + ELINES)")
print(f"{'='*52}")

login_cosco(driver) # <--- THÊM DÒNG NÀY VÀO ĐÂY
# --- BẮT ĐẦU THÊM: LOGIC ĐỌC LỆNH TỪ TERMINAL ---
import sys
target_row = None
single_row_env = os.environ.get("SINGLE_ROW", "").strip()
if single_row_env:
    try:
        target_row = int(single_row_env)
        print(f"[SINGLE_ROW] Chi chay dong {target_row} theo lenh tu main.py")
    except Exception:
        print(f"SINGLE_ROW khong hop le: {single_row_env}")
if len(sys.argv) > 1:
    try:
        target_row = int(sys.argv[1])
        print(f"🛠️ [CHẾ ĐỘ TEST] Bỏ qua toàn bộ, CHỈ CHẠY DUY NHẤT DÒNG {target_row}!")
    except: pass

# Nếu có nhập số dòng trên Terminal thì chỉ chạy mảng có 1 dòng đó, nếu không thì chạy từ dòng 2 đến hết
danh_sach_dong = [target_row] if target_row else range(2, sheet.max_row + 1)

total_valid_rows = 0
for row in danh_sach_dong:
    pol = str(sheet.cell(row=row, column=3).value or "").strip()
    pod = str(sheet.cell(row=row, column=4).value or "").strip()
    carrier = str(sheet.cell(row=row, column=5).value or "").strip().upper()
    if not pol or not pod or carrier != "COSCO":
        continue
    if FILTER_POL and pol.upper() != FILTER_POL:
        continue
    if FILTER_POD and pod.upper() != FILTER_POD:
        continue
    total_valid_rows += 1

print(f"Tổng cộng có {total_valid_rows} dòng cần check.")

for row in danh_sach_dong:
    pod_country = str(sheet.cell(row=row, column=2).value or "").strip()
    if not pod_country:
        pod_country = os.environ.get("FILTER_COUNTRY", "").strip()
    pol = str(sheet.cell(row=row, column=3).value or "").strip()
    pod = str(sheet.cell(row=row, column=4).value or "").strip()
    carrier = str(sheet.cell(row=row, column=5).value or "").strip().upper()

    # --- BẮT ĐẦU THÊM: Chuẩn hóa tên cảng đặc biệt COSCO ---
    PORT_MAPPING = {
            "FOS SUR MER": "FOS",
            "GENOA": "GENOVA",
            "VENICE": "VENEZIA",
            "TIANJIN": "XINGANG",
            "DUNKERQUE": "DUNKIRK",
            "MANILA": "MANILA NORTH HARBOUR",
            "COCHIN": "KOCHI"
        }
    
    # Ép kiểu viết hoa, xóa khoảng trắng thừa và tra từ điển
    if pol:
        pol = PORT_MAPPING.get(str(pol).upper().strip(), str(pol).strip())
    if pod:
        pod = PORT_MAPPING.get(str(pod).upper().strip(), str(pod).strip())
    # --- KẾT THÚC THÊM ---

    if not pol or not pod or carrier != "COSCO":
        continue
    if FILTER_POL and str(sheet.cell(row=row, column=3).value or "").strip().upper() != FILTER_POL:
        continue
    if FILTER_POD and str(sheet.cell(row=row, column=4).value or "").strip().upper() != FILTER_POD:
        continue
    if cosco_has_unsupported_port(pol, pod):
        print(f"\n{'='*52}")
        print(f"⏭️ COSCO bỏ qua route unsupported: {pol} -> {pod} (Row {row})")
        print("   Lý do: PARADIP không có trong port list COSCO, bỏ qua cả Synconhub và Elines.")
        clear_cosco_quote_fields(sheet, row)
        sheet.cell(row=row, column=13).value = "NO SERVICE / SOLD OUT"
        wb.save(EXCEL_FILE); print(f"   [OK] Đã lưu dữ liệu dòng {row} thành công!")
        continue

    print(f"\n{'='*52}")
    print(f"🚀 SO SÁNH GIÁ (LẦN 1): {pol} -> {pod} (Row {row})")
    
    # Kiểm tra nếu POD thuộc AUSTRALIA thì bỏ qua Synconhub
    if pod_country and "AUSTRALIA" in pod_country.upper():
        print("   [1] Synconhub... ⏩ BỎ QUA (Tuyến Úc chỉ check Elines)")
        res_s = None
    else:
        print("   [1] Synconhub...")
        focus_tab_by_url("synconhub.coscoshipping.com", "https://synconhub.coscoshipping.com/spot")
        reload_synconhub_base()
        time.sleep(1)
        if "login" in (driver.current_url or "").lower() or "auth" in (driver.current_url or "").lower():
            print("   ⚠️ Phát hiện Session Synconhub hết hạn! Tiến hành đăng nhập lại...")
            login_cosco(driver)
            focus_tab_by_url("synconhub.coscoshipping.com", "https://synconhub.coscoshipping.com/spot")
            reload_synconhub_base()
        res_s = run_synconhub(pol, pod, pod_country)
        if res_s is None:
            print("   [1B] Synconhub fail/lag -> clean reload và retry 1 lần...")
            focus_tab_by_url("synconhub.coscoshipping.com", "https://synconhub.coscoshipping.com/spot")
            reload_synconhub_base()
            time.sleep(1)
            res_s = run_synconhub(pol, pod, pod_country)
        if res_s is None and pod_country and "CHINA" in pod_country.upper():
            print("   [1B] Synconhub retry for China route after clean reload...")
            focus_tab_by_url("synconhub.coscoshipping.com", "https://synconhub.coscoshipping.com/spot")
            reload_synconhub_base()
            res_s = run_synconhub(pol, pod, pod_country)

    print("   [2] Elines...")
    focus_tab_by_url("elines.coscoshipping.com", ELINES_BOOKING_URL)
    time.sleep(1)
    if "login" in (driver.current_url or "").lower() or "auth" in (driver.current_url or "").lower():
        print("   ⚠️ Phát hiện Session Elines hết hạn! Tiến hành đăng nhập lại...")
        login_cosco(driver)
        focus_tab_by_url("elines.coscoshipping.com", ELINES_BOOKING_URL)
    res_e = run_elines(pol, pod, pod_country)
    elines_no_products = is_elines_no_products_result(res_e)

    chot_deal = None
    source = ""

    def get_compare_price(res):
        if res is None or is_no_service_result(res): return float('inf')
        p20 = res['rates']['20GP']
        p40 = res['rates']['40GP']
        if p20 is not None: return p20
        if p40 is not None: return p40
        return float('inf')

    price_s = get_compare_price(res_s)
    price_e = get_compare_price(res_e)

    if price_s == float('inf') and price_e == float('inf'):
        chot_deal = None
    elif price_s <= price_e:
        chot_deal, source = res_s, "SYNCONHUB"
    else:
        chot_deal, source = res_e, "ELINES"

    if chot_deal:
        r = chot_deal['rates']
        print(f"   🏆 CHỐT TỪ: {source}")
        v20 = f"${round(r['20GP'],2)}" if r['20GP'] is not None else "N/A"
        v40 = f"${round(r['40GP'],2)}" if r['40GP'] is not None else "N/A"
        vHQ = f"${round(r['40HQ'],2)}" if r['40HQ'] is not None else "N/A"
        print(f"      20'={v20} | 40'={v40} | 40H={vHQ}")
        
        formulas = chot_deal.get("formulas", {})
        sheet.cell(row=row, column=6).value = formulas.get("20GP") or (round(r['20GP'], 2) if r['20GP'] is not None else "")
        sheet.cell(row=row, column=7).value = formulas.get("40GP") or (round(r['40GP'], 2) if r['40GP'] is not None else "")
        sheet.cell(row=row, column=8).value = formulas.get("40HQ") or (round(r['40HQ'], 2) if r['40HQ'] is not None else "")
        sheet.cell(row=row, column=9).value = chot_deal['etd']
        sheet.cell(row=row, column=10).value = chot_deal['tt']
        sheet.cell(row=row, column=11).value = chot_deal['valid'] # <-- ĐIỀN VALID CỘT K
        
        # BẮT ĐẦU THÊM: Ghi Remark có truyền thêm cờ Surcharge Error
        is_err = chot_deal.get('surcharge_error', False)
        sheet.cell(row=row, column=13).value = generate_remark(chot_deal['thc_inc'], chot_deal['ows'], chot_deal['manifest_fee'], pod, source, is_err)
        
        # Ghi Free Time cột N (14)
        ft_value = free_time_dict.get(pod_country.upper(), "") 
        sheet.cell(row=row, column=14).value = ft_value

        # Ghi Thông tin Tàu (Cột O) và Chuyển tải (Cột P)
        sheet.cell(row=row, column=15).value = chot_deal.get('vessel_info', '')
        sheet.cell(row=row, column=16).value = chot_deal.get('transshipment', '')

        # --- BẮT ĐẦU THÊM: Ghi Free Time cột N (14) tra theo QUỐC GIA ---
        ft_value = free_time_dict.get(pod_country.upper(), "") 
        print(f"      -> [Free Time] Đã tìm thấy cho quốc gia {pod_country.upper()}: '{ft_value}'")
        sheet.cell(row=row, column=14).value = ft_value
        # --- KẾT THÚC THÊM ---
    else:
        print("   ❌ Cả 2 hệ thống đều móm ở Lần 1!")
        clear_cosco_quote_fields(sheet, row)
        sheet.cell(row=row, column=13).value = "NO SERVICE / SOLD OUT"
        if elines_no_products:
            print("   ⏭️ Elines đã xác nhận không có giá Premium -> không retry route này ở Lần 2.")
        else:
            failed_rows.append(row)  # Đưa dòng này vào danh sách đen chờ xử ở Phiên 2

    wb.save(EXCEL_FILE); print(f"   [OK] Đã lưu dữ liệu dòng {row} thành công!")

# ---------------------------------------------------------
# PHIÊN CHẠY 2: CHỈ VÉT LẠI CÁC DÒNG LỖI BẰNG ELINES
# ---------------------------------------------------------
if failed_rows:
    print(f"\n{'='*52}")
    print(f"🔄 BẮT ĐẦU PHIÊN CHẠY 2: VÉT LẠI {len(failed_rows)} TUYẾN BỊ LỖI TRÊN ELINES 🔄")
    print(f"{'='*52}")

    for row in failed_rows:
        pod_country = str(sheet.cell(row=row, column=2).value or "").strip()
        if not pod_country:
            pod_country = os.environ.get("FILTER_COUNTRY", "").strip()
        pol = str(sheet.cell(row=row, column=3).value or "").strip()
        pod = str(sheet.cell(row=row, column=4).value or "").strip()
        # --- BẮT ĐẦU THÊM: Chuẩn hóa tên cảng đặc biệt COSCO ---
        PORT_MAPPING = {
            "FOS SUR MER": "FOS",
            "GENOA": "GENOVA",
            "TIANJIN": "XINGANG",
            "DUNKERQUE": "DUNKIRK",
            "VENICE": "VENEZIA",
            "MANILA": "MANILA NORTH HARBOUR",
            "COCHIN": "KOCHI"
        }
        
        # Ép kiểu viết hoa, xóa khoảng trắng thừa và tra từ điển
        if pol:
            pol = PORT_MAPPING.get(str(pol).upper().strip(), str(pol).strip())
        if pod:
            pod = PORT_MAPPING.get(str(pod).upper().strip(), str(pod).strip())
        # --- KẾT THÚC THÊM ---
        if cosco_has_unsupported_port(pol, pod):
            print(f"\n⏭️ PHIÊN 2 bỏ qua route unsupported COSCO: {pol} -> {pod} (Row {row})")
            sheet.cell(row=row, column=13).value = "NO SERVICE / SOLD OUT"
            wb.save(EXCEL_FILE); print(f"   [OK] Đã lưu dữ liệu dòng {row} thành công!")
            continue
        
        print(f"\n🚀 SO SÁNH GIÁ (LẦN 2): {pol} -> {pod} (Row {row})")
        
        print("   [2] Elines (Retry) - Đang reload lại trang để clear lag...")
        
        # 1. Trỏ vào đúng tab Elines (hoặc mở mới nếu lỡ tắt)
        focus_tab_by_url("elines.coscoshipping.com", ELINES_BOOKING_URL)
        
        # 2. Ép trình duyệt Reload thẳng lại link gốc để dọn sạch rác/state cũ
        driver.get(ELINES_BOOKING_URL)
        time.sleep(4)  # Chờ 4s cho iframe và các component Vue của hãng tàu load xong hoàn toàn

        # A stale booking shell can look authenticated while its API session is
        # already dead.  The clean navigation above then reveals loginPlease.
        # Re-authenticate before spending another full port/dropdown timeout.
        if not _is_logged_in_elines(driver) or is_elines_auth_page(driver.current_url):
            print("   [2] Elines session đã hết hạn -> login lại trước Phiên 2...")
            login_cosco(driver)
            focus_tab_by_url("elines.coscoshipping.com", ELINES_BOOKING_URL)
            if not _is_logged_in_elines(driver):
                driver.get(ELINES_BOOKING_URL)
                time.sleep(2)

        # 3. Bắt đầu chạy lại hàm rút giá
        res_e = run_elines(pol, pod, pod_country)

        if is_elines_no_products_result(res_e):
            print("   ⏭️ PHIÊN 2: Elines xác nhận không có giá Premium -> không xử lý thêm.")
            clear_cosco_quote_fields(sheet, row)
            sheet.cell(row=row, column=13).value = "NO SERVICE / SOLD OUT"
            wb.save(EXCEL_FILE); print(f"   [OK] Đã lưu dữ liệu dòng {row} thành công!")
            continue

        if res_e:
            r = res_e['rates']
            source = "ELINES (RETRY)"
            print(f"   🏆 PHIÊN 2 ĐÃ VỚT ĐƯỢC DEAL TỪ: {source}")
            
            v20 = f"${round(r['20GP'],2)}" if r['20GP'] is not None else "N/A"
            v40 = f"${round(r['40GP'],2)}" if r['40GP'] is not None else "N/A"
            vHQ = f"${round(r['40HQ'],2)}" if r['40HQ'] is not None else "N/A"
            print(f"      20'={v20} | 40'={v40} | 40H={vHQ}")
            
            # Ghi đè lại dữ liệu (xoá cái NO SERVICE cũ)
            formulas = res_e.get("formulas", {})
            sheet.cell(row=row, column=6).value = formulas.get("20GP") or (round(r['20GP'], 2) if r['20GP'] is not None else "")
            sheet.cell(row=row, column=7).value = formulas.get("40GP") or (round(r['40GP'], 2) if r['40GP'] is not None else "")
            sheet.cell(row=row, column=8).value = formulas.get("40HQ") or (round(r['40HQ'], 2) if r['40HQ'] is not None else "")
            sheet.cell(row=row, column=9).value = res_e['etd']
            sheet.cell(row=row, column=10).value = res_e['tt']
            sheet.cell(row=row, column=11).value = res_e['valid'] # <-- ĐIỀN VALID CỘT K
        
            # BẮT ĐẦU THÊM: Ghi Remark có truyền thêm cờ Surcharge Error
            is_err = res_e.get('surcharge_error', False)
            sheet.cell(row=row, column=13).value = generate_remark(res_e['thc_inc'], res_e['ows'], res_e['manifest_fee'], pod, source, is_err)
            
            # Ghi Free Time cột N (14)
            ft_value = free_time_dict.get(pod_country.upper(), "") 
            sheet.cell(row=row, column=14).value = ft_value

            # Ghi Thông tin Tàu (Cột O) và Chuyển tải (Cột P)
            from openpyxl.styles import Alignment
            
            vessel_cell = sheet.cell(row=row, column=15)
            vessel_cell.value = res_e.get('vessel_info', '')
            vessel_cell.alignment = Alignment(wrap_text=True) # Kích hoạt tính năng xuống dòng trong ô Excel
            
            sheet.cell(row=row, column=16).value = res_e.get('transshipment', '')
            # --- BẮT ĐẦU THÊM: Ghi Free Time cột N (14) tra theo QUỐC GIA ---
            ft_value = free_time_dict.get(pod_country.upper(), "") 
            print(f"      -> [Free Time] Đã tìm thấy cho quốc gia {pod_country.upper()}: '{ft_value}'")
            sheet.cell(row=row, column=14).value = ft_value
            # --- KẾT THÚC THÊM ---
        else:
            print("   ❌ Phiên 2 vẫn móm, bỏ qua!")
            clear_cosco_quote_fields(sheet, row)
            sheet.cell(row=row, column=13).value = "NO SERVICE / SOLD OUT"

        wb.save(EXCEL_FILE); print(f"   [OK] Đã lưu dữ liệu dòng {row} thành công!")

print("\n🎉 XONG TOÀN BỘ TIẾN TRÌNH! CHECK FILE EXCEL NHÉ!")
