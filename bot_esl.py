"""ESL (Emirates Shipping Line) point-to-point schedule scraper.

The bot reads ESL rows from the workbook and writes schedule-only fields:
  I: ETD, J: transit time, O: vessel detail, P: transshipment port.

ESL's website identifies Ho Chi Minh / Cai Mep by the location code VNCMT.
"""

import math
import os
import re
import socket
import subprocess
import sys
import time
import unicodedata
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait

from bot_cli import etd_within_max, max_etd_date_only, parse_date_offset_days


try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


EXCEL_PATH = os.environ.get("EXCEL_PATH") or os.path.join(os.getcwd(), "input_gia.xlsx")
FILTER_POL = (os.environ.get("FILTER_POL") or "").strip().upper()
FILTER_POD = (os.environ.get("FILTER_POD") or "").strip().upper()
FILTER_COUNTRY = (os.environ.get("FILTER_COUNTRY") or "").strip().upper()
FILTER_ROW_FROM = (os.environ.get("FILTER_ROW_FROM") or "").strip()
FILTER_ROW_TO = (os.environ.get("FILTER_ROW_TO") or "").strip()
SINGLE_ROW = (os.environ.get("SINGLE_ROW") or "").strip()
DATE_OFFSET_DAYS = parse_date_offset_days()

CARRIER_TARGETS = {"ESL", "EMIRATES LINE", "EMIRATES SHIPPING LINE"}
URL_ESL = "https://esea.emiratesline.com/ecomonline/Welcome.do"
DEBUG_PORT = os.environ.get("ESL_EDGE_PORT", "9535")
DRIVER_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "msedgedriver.exe")
EDGE_PROFILE = os.environ.get("ESL_EDGE_PROFILE", r"C:\edge_esl")
EDGE_EXE_CANDIDATES = (
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
)

TODAY = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
PORT_ALIASES = {
    "HO CHI MINH": "VNCMT",
    "HOCHIMINH": "VNCMT",
    "HCM": "VNCMT",
    "SAIGON": "VNCMT",
    "CAI MEP": "VNCMT",
    # ESL POD aliases used by input_gia.xlsx rows 284-307.
    "JEBEL ALI": "AEJEA",
    "UMM AL QUWAIN": "AEQIW",
    "SHARJAH": "AESCT",
    "KHOR FAKKAN": "AEKLF",
    "AJMAN": "AEAJM",
    "ABU DHABI": "AEAUH",
    "RAS AL KHAIMAH": "AERKT",
    "SHUWAIKH": "KWKWI",
    "SHUAIBA": "KWSAA",
    "HAMAD": "QAHMD",
    "BAHRAIN": "BHBAH",
    "UMM QASR": "IQUQR",
    "DAMMAM": "SADMN",
    "RIYADH": "SARUH",
    "MUNDRA": "INMUN",
    "NHAVA SHEVA": "INNSA",
    "HAZIRA": "INHZA",
    "KARACHI": "PKKHI",
    "SOKHNA": "EGSOK",
    "AQABA": "JOAQJ",
    "JEDDAH": "SAJED",
    "DJIBOUTI": "DJJIB",
    "MANZANILLO": "MXZLO",
    "DAR ES SALAAM": "TZDAR",
}

MONTHS = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

NO_RESULT_PHRASES = (
    "NO SCHEDULE",
    "NO SAILING",
    "NO RECORD",
    "NO RESULT",
    "NO ROUTE",
    "SCHEDULE NOT FOUND",
)


def log(message):
    print(f"[ESL] {message}", flush=True)


def norm(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().upper()


def compact(value):
    return re.sub(r"[^A-Z0-9]", "", norm(value))


def parse_int(value):
    try:
        return int(str(value).strip())
    except Exception:
        return None


def parse_valid_date(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(raw, date):
        return datetime(raw.year, raw.month, raw.day)

    text = str(raw).strip()
    if not text:
        return None

    for fmt in (
        "%d-%b-%Y", "%d-%b-%y", "%d/%m/%Y", "%d/%m/%y",
        "%Y-%m-%d", "%d-%m-%Y", "%d %b %Y",
    ):
        try:
            return datetime.strptime(text, fmt).replace(hour=0, minute=0, second=0, microsecond=0)
        except ValueError:
            pass

    match = re.fullmatch(r"(\d{1,2})[-/\s]+([A-Za-z]{3})", text)
    if match and match.group(2).upper() in MONTHS:
        candidate = datetime(TODAY.year, MONTHS[match.group(2).upper()], int(match.group(1)))
        if candidate < TODAY - timedelta(days=60):
            candidate = candidate.replace(year=candidate.year + 1)
        return candidate

    match = re.fullmatch(r"(\d{1,2})/(\d{1,2})", text)
    if match:
        candidate = datetime(TODAY.year, int(match.group(2)), int(match.group(1)))
        if candidate < TODAY - timedelta(days=60):
            candidate = candidate.replace(year=candidate.year + 1)
        return candidate
    return None


def parse_esl_datetime(raw):
    text = str(raw or "")
    match = re.search(r"\b(\d{1,2}-[A-Za-z]{3}-\d{4})(?:\s+(\d{1,2}:\d{2}))?\b", text)
    if not match:
        return None
    value = match.group(1) + (f" {match.group(2)}" if match.group(2) else "")
    fmt = "%d-%b-%Y %H:%M" if match.group(2) else "%d-%b-%Y"
    try:
        return datetime.strptime(value, fmt)
    except ValueError:
        return None


def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        return sock.connect_ex(("127.0.0.1", int(port))) == 0


def edge_executable():
    for path in EDGE_EXE_CANDIDATES:
        if os.path.exists(path):
            return path
    raise FileNotFoundError("Khong tim thay Microsoft Edge")


def ensure_edge_debug_port():
    if is_port_in_use(DEBUG_PORT):
        log(f"Edge debug port {DEBUG_PORT} da mo san, tai su dung phien ESL.")
        return
    log(f"Khoi dong Edge ESL port {DEBUG_PORT}...")
    subprocess.Popen(
        [
            edge_executable(),
            f"--remote-debugging-port={DEBUG_PORT}",
            f"--user-data-dir={EDGE_PROFILE}",
            "--disable-background-timer-throttling",
            "--disable-renderer-backgrounding",
            "--disable-backgrounding-occluded-windows",
            "--start-maximized",
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    for _ in range(40):
        if is_port_in_use(DEBUG_PORT):
            return
        time.sleep(0.5)
    raise RuntimeError(f"Khong mo duoc Edge ESL port {DEBUG_PORT}")


def connect_edge():
    ensure_edge_debug_port()
    options = Options()
    options.add_experimental_option("debuggerAddress", f"127.0.0.1:{DEBUG_PORT}")
    driver = webdriver.Edge(service=Service(executable_path=DRIVER_PATH), options=options)
    try:
        driver.maximize_window()
    except Exception:
        pass
    driver.set_page_load_timeout(50)
    return driver


def close_esl_edge(driver):
    try:
        if driver:
            driver.quit()
    except Exception:
        pass
    if os.name != "nt":
        return
    profile_pattern = os.path.basename(os.path.normpath(EDGE_PROFILE)) or "edge_esl"
    script = f"""
$procs = Get-CimInstance Win32_Process -Filter "name = 'msedge.exe'"
foreach ($p in $procs) {{
    $cmd = [string]$p.CommandLine
    if ($cmd -like '*{profile_pattern}*' -or $cmd -like '*remote-debugging-port={DEBUG_PORT}*') {{
        Stop-Process -Id $p.ProcessId -Force -ErrorAction SilentlyContinue
    }}
}}
"""
    try:
        kwargs = {}
        if hasattr(subprocess, "CREATE_NO_WINDOW"):
            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        subprocess.run(
            ["powershell", "-NoProfile", "-Command", script],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=15,
            **kwargs,
        )
    except Exception:
        pass


def displayed_elements(driver, selector):
    return [element for element in driver.find_elements(By.CSS_SELECTOR, selector) if element.is_displayed()]


def set_input_value(driver, selector, value):
    element = WebDriverWait(driver, 20).until(lambda d: d.find_element(By.CSS_SELECTOR, selector))
    driver.execute_script(
        "arguments[0].removeAttribute('readonly');"
        "arguments[0].value=arguments[1];"
        "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));"
        "arguments[0].dispatchEvent(new Event('change',{bubbles:true}));",
        element,
        value,
    )


def typeahead_score(text, query, original, country_hint):
    candidate = norm(text)
    score = 0
    if candidate.startswith(norm(query) + " -"):
        score += 1000
    if norm(query) and norm(query) in candidate:
        score += 200
    if compact(original) and compact(original) in compact(candidate):
        score += 100
    if country_hint and norm(country_hint) in candidate:
        score += 80
    if norm(query) == "VNCMT" and candidate.startswith("VNCMT -"):
        score += 2000
    return score


def select_typeahead(driver, input_selector, hidden_selector, query, original, country_hint=""):
    field = WebDriverWait(driver, 20).until(lambda d: d.find_element(By.CSS_SELECTOR, input_selector))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", field)
    field.click()
    field.send_keys(Keys.CONTROL, "a")
    field.send_keys(Keys.BACKSPACE)
    field.send_keys(str(query))

    # Bloodhound/typeahead replaces its suggestion nodes while the menu is
    # open. Holding Selenium WebElements here intermittently raises stale
    # element errors, so read and click a fresh DOM node synchronously.
    def suggestion_texts(drv):
        texts = drv.execute_script(
            """
            return Array.from(document.querySelectorAll('.tt-menu.tt-open .tt-suggestion'))
              .filter(function(node) {
                const style = getComputedStyle(node);
                const rect = node.getBoundingClientRect();
                return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
              })
              .map(function(node) { return (node.innerText || node.textContent || '').trim(); });
            """
        ) or []
        return texts or False

    texts = WebDriverWait(driver, 15, poll_frequency=0.25).until(suggestion_texts)
    chosen_index = max(
        range(len(texts)),
        key=lambda index: typeahead_score(texts[index], query, original, country_hint),
    )
    chosen_text = re.sub(r"\s+", " ", texts[chosen_index]).strip()
    clicked = driver.execute_script(
        """
        const nodes = Array.from(document.querySelectorAll('.tt-menu.tt-open .tt-suggestion'))
          .filter(function(node) {
            const style = getComputedStyle(node);
            const rect = node.getBoundingClientRect();
            return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
          });
        const node = nodes[arguments[0]];
        if (!node) return false;
        node.scrollIntoView({block:'center'});
        node.click();
        return true;
        """,
        chosen_index,
    )
    if not clicked:
        raise RuntimeError(f"ESL typeahead mat suggestion cho '{query}'")

    def selected(drv):
        try:
            visible = drv.find_element(By.CSS_SELECTOR, input_selector).get_attribute("value") or ""
            hidden = drv.find_element(By.CSS_SELECTOR, hidden_selector).get_attribute("value") or ""
            return bool(visible.strip() and hidden.strip())
        except Exception:
            return False

    WebDriverWait(driver, 8).until(selected)
    return chosen_text


def visible_schedule_cards(driver):
    return displayed_elements(driver, ".details_panel")


def page_has_no_results(driver):
    try:
        body = norm(driver.find_element(By.TAG_NAME, "body").text)
    except Exception:
        return False
    return any(phrase in body for phrase in NO_RESULT_PHRASES)


def wait_for_schedule_result(driver):
    def ready(drv):
        if visible_schedule_cards(drv):
            return True
        if page_has_no_results(drv):
            return True
        return False

    try:
        WebDriverWait(driver, 35, poll_frequency=0.4).until(ready)
    except TimeoutException:
        if not visible_schedule_cards(driver):
            raise


def search_schedule_page(driver, pol_excel, pod_excel, country_hint, start_dt, end_dt):
    driver.get(URL_ESL)
    WebDriverWait(driver, 25).until(lambda d: d.find_element(By.CSS_SELECTOR, "#loginschedsearch #pol"))

    pol_query = PORT_ALIASES.get(norm(pol_excel), pol_excel)
    selected_pol = select_typeahead(
        driver, "#loginschedsearch #pol", "#loginschedsearch #typeheadhidden",
        pol_query, pol_excel, "VIETNAM",
    )
    pod_query = PORT_ALIASES.get(norm(pod_excel), pod_excel)
    selected_pod = select_typeahead(
        driver, "#loginschedsearch #pod", "#loginschedsearch #typeheadhidden1",
        pod_query, pod_excel, country_hint,
    )
    log(f"   Port: {selected_pol} -> {selected_pod}")

    set_input_value(driver, "#loginschedsearch #scdshipmentdate", start_dt.strftime("%d-%b-%Y"))
    set_input_value(driver, "#loginschedsearch #todateq", end_dt.strftime("%d-%b-%Y"))
    submit = driver.find_element(By.CSS_SELECTOR, "#loginschedsearch button[type='submit']")
    driver.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].click();", submit)
    wait_for_schedule_result(driver)
    return parse_schedule_cards(driver)


def clean_transshipment_port(raw):
    value = re.sub(r"\s+", " ", str(raw or "")).strip().upper()
    if not value:
        return ""
    return value.split(",", 1)[0].strip()


def parse_schedule_cards(driver):
    raw_cards = driver.execute_script(
        """
        return Array.from(document.querySelectorAll('.details_panel'))
          .filter(function(card) {
            const style = getComputedStyle(card);
            const rect = card.getBoundingClientRect();
            return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
          })
          .map(function(card) {
            const text = function(selector) {
              const node = card.querySelector(selector);
              return node ? node.innerText : '';
            };
            return {
              departure: text('.panel_inner_top_1'),
              arrival: text('.panel_inner_top_2'),
              vessel: text('.panel_inner_top_3'),
              transit: text('.panel_inner_top_4'),
              bottom: text('.panel_inner_bottom'),
              ports: Array.from(card.querySelectorAll('.collapse .port_name')).map(function(node) {
                return (node.textContent || '').trim();
              })
            };
          });
        """
    ) or []

    entries = []
    for raw in raw_cards:
        etd = parse_esl_datetime(raw.get("departure"))
        eta = parse_esl_datetime(raw.get("arrival"))
        if not etd:
            continue

        vessel_lines = [
            re.sub(r"\s+", " ", line).strip()
            for line in str(raw.get("vessel") or "").splitlines()
            if re.sub(r"\s+", " ", line).strip()
        ]
        vessel_lines = [line for line in vessel_lines if norm(line) != "VESSEL"]
        vessel = " ".join(vessel_lines).strip() or "TBA"

        transit_match = re.search(r"(\d+)\s*days?", str(raw.get("transit") or ""), re.I)
        if transit_match:
            transit_days = int(transit_match.group(1))
        elif eta:
            transit_days = max(0, math.ceil((eta - etd).total_seconds() / 86400))
        else:
            transit_days = 0

        direct = "DIRECT SHIPMENT" in norm(raw.get("bottom"))
        ports = [clean_transshipment_port(port) for port in (raw.get("ports") or [])]
        ports = [port for port in ports if port]
        middle_ports = [] if direct or len(ports) < 3 else ports[1:-1]
        unique_middle = []
        for port in middle_ports:
            if port not in unique_middle:
                unique_middle.append(port)

        entries.append(
            {
                "etd_dt": etd,
                "eta_dt": eta,
                "tt_days": transit_days,
                "vessel": vessel,
                "ts_port": " + ".join(unique_middle) if unique_middle else "DIRECT",
                "direct": direct,
            }
        )
    return entries


def choose_one_route_per_etd(entries):
    by_day = {}
    for entry in entries:
        day = entry["etd_dt"].date()
        current = by_day.get(day)
        candidate_score = (int(entry.get("tt_days") or 99999), 0 if entry.get("direct") else 1)
        current_score = (
            int(current.get("tt_days") or 99999),
            0 if current and current.get("direct") else 1,
        ) if current else None
        if current is None or candidate_score < current_score:
            by_day[day] = entry
    return [by_day[key] for key in sorted(by_day)]


def filter_entries(entries, start_dt, end_dt):
    result = []
    for entry in entries:
        etd = entry.get("etd_dt")
        if not etd:
            continue
        day = etd.date()
        if day < start_dt.date() or day > end_dt.date():
            continue
        if not etd_within_max(etd):
            continue
        result.append(entry)
    return choose_one_route_per_etd(result)


def select_normal_entries(entries):
    selected = []
    for entry in entries:
        if len(selected) >= 3:
            break
        if selected and entry["etd_dt"].date() == selected[-1]["etd_dt"].date():
            continue
        selected.append(entry)
    return selected


def run_schedule_search(driver, pol, pod, country, valid_dt):
    hard_end = min(valid_dt.date(), max_etd_date_only())
    if hard_end < TODAY.date():
        return [], "VALID da qua han"

    valid_end = datetime.combine(hard_end, datetime.min.time())
    normal_start = TODAY + timedelta(days=DATE_OFFSET_DAYS)

    if normal_start.date() <= hard_end:
        log(f"   Search {normal_start:%d-%b-%Y} -> {valid_end:%d-%b-%Y}")
        entries = search_schedule_page(driver, pol, pod, country, normal_start, valid_end)
        candidates = filter_entries(entries, normal_start, valid_end)
        selected = select_normal_entries(candidates)
        if selected:
            return selected, ""
        log("   Khong co ETD tu date +N; fallback hom nay -> valid.")

    fallback_start = TODAY
    log(f"   Fallback {fallback_start:%d-%b-%Y} -> {valid_end:%d-%b-%Y}")
    entries = search_schedule_page(driver, pol, pod, country, fallback_start, valid_end)
    candidates = filter_entries(entries, fallback_start, valid_end)
    if not candidates:
        return [], "Khong co lich tau hop le"
    return [candidates[-1]], ""


def fmt_date_short(dt):
    return f"{dt.day}-{dt.strftime('%b')}"


def format_etd_text(entries):
    values = [(entry["etd_dt"].day, entry["etd_dt"].strftime("%b")) for entry in entries]
    if not values:
        return ""
    if len(values) == 1:
        return f"{values[0][0]}-{values[0][1]}"
    if len(values) == 2:
        return f"{values[0][0]}-{values[0][1]} & {values[1][0]}-{values[1][1]}"
    months = [month for _, month in values]
    if months[0] == months[1] == months[2]:
        return f"{values[0][0]}, {values[1][0]}, {values[2][0]}-{values[2][1]}"
    if months[0] == months[1]:
        return f"{values[0][0]}, {values[1][0]}-{values[1][1]} & {values[2][0]}-{values[2][1]}"
    if months[1] == months[2]:
        return f"{values[0][0]}-{values[0][1]}, {values[1][0]} & {values[2][0]}-{values[2][1]}"
    return " & ".join(f"{day}-{month}" for day, month in values)


def format_transit_time(entries):
    values = [int(entry.get("tt_days") or 0) for entry in entries]
    if not values:
        return ""
    if len(set(values)) == 1:
        return str(values[0])
    return f"{min(values)}-{max(values)}"


def format_vessel_details(entries):
    details = []
    transshipment_ports = []
    for entry in entries:
        ts_port = entry.get("ts_port") or "DIRECT"
        details.append(
            f"{entry.get('vessel') or 'TBA'} / ETD: {fmt_date_short(entry['etd_dt'])}"
            f" / Transit time: {int(entry.get('tt_days') or 0)} Days"
            f" / Transshipment Port: {ts_port}"
        )
        if ts_port not in transshipment_ports:
            transshipment_ports.append(ts_port)
    return "\n".join(details), " or\n".join(transshipment_ports)


def write_excel_row(row_number, entries=None, no_schedule=False):
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    worksheet = workbook.active
    if worksheet.cell(row=row_number, column=6).value is None:
        worksheet.cell(row=row_number, column=6).value = "-"

    if no_schedule or not entries:
        for column in (9, 10, 15, 16):
            worksheet.cell(row=row_number, column=column).value = "-"
    else:
        vessel_text, ts_text = format_vessel_details(entries)
        worksheet.cell(row=row_number, column=9).value = format_etd_text(entries)
        worksheet.cell(row=row_number, column=10).value = format_transit_time(entries)
        worksheet.cell(row=row_number, column=15).value = vessel_text
        worksheet.cell(row=row_number, column=16).value = ts_text
        wrap = Alignment(wrap_text=True, vertical="top")
        worksheet.cell(row=row_number, column=15).alignment = wrap
        worksheet.cell(row=row_number, column=16).alignment = wrap

    workbook.save(EXCEL_PATH)
    workbook.close()
    log(f"   Wrote Excel row {row_number}")


def in_requested_row_range(row_number):
    row_from = parse_int(FILTER_ROW_FROM)
    row_to = parse_int(FILTER_ROW_TO)
    if row_from is not None and row_number < row_from:
        return False
    if row_to is not None and row_number > row_to:
        return False
    return True


def load_targets():
    workbook = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=False)
    worksheet = workbook.active
    single_row = parse_int(SINGLE_ROW)
    targets = []

    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, max_col=11, values_only=True), start=2
    ):
        if single_row is not None and row_number != single_row:
            continue
        if not in_requested_row_range(row_number):
            continue
        country = str(values[1] or "").strip()
        pol = str(values[2] or "").strip()
        pod = str(values[3] or "").strip()
        carrier = norm(values[4])
        valid_raw = values[10]
        if carrier not in CARRIER_TARGETS or not pol or not pod:
            continue
        if FILTER_POL and norm(pol) != norm(FILTER_POL):
            continue
        if FILTER_POD and norm(pod) != norm(FILTER_POD):
            continue
        targets.append((row_number, country, pol, pod, valid_raw))

    workbook.close()
    return targets


def main():
    if not os.path.exists(EXCEL_PATH):
        log(f"Khong tim thay Excel: {EXCEL_PATH}")
        return 1

    targets = load_targets()
    log(f"Co {len(targets)} dong can check (carrier = ESL)")
    if not targets:
        return 0

    driver = None
    try:
        driver = connect_edge()
        for index, (row_number, country, pol, pod, valid_raw) in enumerate(targets, start=1):
            log(f"========== [{index}/{len(targets)}] DONG {row_number}: {pol} -> {pod} | VALID={valid_raw}")
            valid_dt = parse_valid_date(valid_raw)
            if not valid_dt:
                log("   Valid rong -> bo qua row, khong search web.")
                continue
            for attempt in range(1, 3):
                try:
                    entries, error = run_schedule_search(
                        driver, pol, pod, country or FILTER_COUNTRY, valid_dt
                    )
                    if error:
                        log(f"   NO SCHEDULE: {error}")
                        write_excel_row(row_number, no_schedule=True)
                    else:
                        vessel_text, ts_text = format_vessel_details(entries)
                        log(f"   ETD={format_etd_text(entries)} | TT={format_transit_time(entries)} | TS={ts_text}")
                        log(f"   Vessel: {vessel_text}")
                        write_excel_row(row_number, entries=entries)
                    break
                except Exception as exc:
                    if attempt < 2:
                        log(
                            f"   Browser error lan {attempt}: {type(exc).__name__}; "
                            "dong phien loi va retry 1 lan."
                        )
                        close_esl_edge(driver)
                        driver = None
                        time.sleep(1)
                        driver = connect_edge()
                        continue
                    log(f"   ERROR row {row_number}: {type(exc).__name__}: {exc}")
                    write_excel_row(row_number, no_schedule=True)
            time.sleep(0.5)
    finally:
        close_esl_edge(driver)

    log("DONE bot ESL")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
