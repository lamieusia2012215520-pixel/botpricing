"""
ZIM schedule checker.

ZIM's public point-to-point page calls APIM endpoints that return schedules, but
the APIM gateway rejects plain Python requests. This bot opens the real ZIM page
once in Edge and executes fetch() inside that browser context.

Excel output matches the schedule-only bots:
  I: ETD, J: transit time, O: vessel detail, P: transshipment port.
"""

import argparse
import json
import math
import os
import re
import socket
import subprocess
import sys
import time
import warnings
from datetime import datetime, timedelta
from urllib.parse import urlencode
import urllib.request

import openpyxl
from bot_cli import etd_within_max, max_etd_date, max_etd_date_only
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
import websocket
try:
    from selenium.webdriver.remote.remote_connection import RemoteConnection
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", DeprecationWarning)
        RemoteConnection.set_timeout(int(os.environ.get("ZIM_SELENIUM_CONNECT_TIMEOUT", "35")))
except Exception:
    pass


try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


CURRENT_FOLDER = os.path.dirname(os.path.abspath(__file__))
DRIVER_PATH = os.path.join(CURRENT_FOLDER, "msedgedriver.exe")
EXCEL_PATH = os.environ.get("EXCEL_PATH") or os.path.join(CURRENT_FOLDER, "input_gia.xlsx")
FILTER_POL = (os.environ.get("FILTER_POL") or "").strip().upper()
FILTER_POD = (os.environ.get("FILTER_POD") or "").strip().upper()
FILTER_COUNTRY = (os.environ.get("FILTER_COUNTRY") or "").strip().upper()
SINGLE_ROW = (os.environ.get("SINGLE_ROW") or "").strip()

CARRIER_TARGETS = {"ZIM", "ZIM LINE", "ZIM LINES"}
ZIM_PORT = int(os.environ.get("ZIM_EDGE_PORT", "9534"))
ZIM_PROFILE = os.environ.get("ZIM_EDGE_PROFILE", r"C:\edge_zim_schedule")
ZIM_URL = "https://www.zim.com/schedules/point-to-point"
APIM = "https://apigw.zim.com"
SUB_KEY = os.environ.get("ZIM_APIM_SUBSCRIPTION_KEY", "9d63cf020a4c4708a7b0ebfe39578300")
TODAY = datetime.now()
ZIM_USE_SELENIUM = (os.environ.get("ZIM_USE_SELENIUM") or "0").strip().lower() in {"1", "true", "yes", "y"}


def parse_cli_args(argv):
    normalized = []
    for arg in argv:
        if arg.lower() == "--date":
            normalized.append("--date")
        elif arg.lower().startswith("--date="):
            normalized.append("--date=" + arg.split("=", 1)[1])
        else:
            normalized.append(arg)

    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--date", default=None)
    parser.add_argument("single_row", nargs="?")
    parsed, _ = parser.parse_known_args(normalized)

    cli_force = parsed.date is not None
    raw_offset = parsed.date if cli_force else os.environ.get("DATE_OFFSET_DAYS", "")
    try:
        offset = int(str(raw_offset or "7").strip().lstrip("+"))
    except Exception:
        offset = 7
    return parsed.single_row, max(0, offset), cli_force


CLI_SINGLE_ROW, DATE_OFFSET_DAYS, CLI_FORCE_DATE_OFFSET = parse_cli_args(sys.argv[1:])
if CLI_SINGLE_ROW and not SINGLE_ROW:
    SINGLE_ROW = CLI_SINGLE_ROW.strip()

FORCE_DATE_OFFSET = (
    CLI_FORCE_DATE_OFFSET
    or os.environ.get("ZIM_FORCE_DATE_OFFSET", "").strip().lower() in {"1", "true", "yes", "y"}
    or os.environ.get("FORCE_DATE_OFFSET", "").strip().lower() in {"1", "true", "yes", "y"}
)

PORT_ALIASES = {
    "HO CHI MINH": ["HO CHI MINH", "HO CHI MINH CITY", "CAT LAI", "SAIGON"],
    "HOCHIMINH": ["HO CHI MINH", "HO CHI MINH CITY", "CAT LAI", "SAIGON"],
    "HCM": ["HO CHI MINH", "HO CHI MINH CITY", "CAT LAI", "SAIGON"],
    "SAIGON": ["HO CHI MINH", "HO CHI MINH CITY", "CAT LAI", "SAIGON"],
    "HAI PHONG": ["HAIPHONG", "HAI PHONG", "HAI PHONG PORT"],
    "HAIPHONG": ["HAIPHONG", "HAI PHONG", "HAI PHONG PORT"],
    "NHAVA SHEVA": ["NHAVA SHEVA", "JAWAHARLAL NEHRU", "JNPT"],
    "NHAVASHEVA": ["NHAVA SHEVA", "JAWAHARLAL NEHRU", "JNPT"],
    "PORT KLANG": ["PORT KLANG", "PORT KELANG"],
    "PORTKLANG": ["PORT KLANG", "PORT KELANG"],
    "JEBEL ALI": ["JEBEL ALI", "JEBELALI"],
    "JEBELALI": ["JEBEL ALI", "JEBELALI"],
    "VISAKHAPATNAM": ["VIZAG", "VISAKHAPATNAM", "VISAKHAPATNAM PORT"],
    "VISAKHAPATNAM PORT": ["VIZAG", "VISAKHAPATNAM", "VISAKHAPATNAM PORT"],
    "VIZAG": ["VIZAG", "VISAKHAPATNAM", "VISAKHAPATNAM PORT"],
    "NAPLES": ["NAPLES", "NAPOLI"],
    "BASLE": ["BASEL", "BASLE"],
    "BASEL": ["BASEL", "BASLE"],
}

COUNTRY_ALIASES = {
    "USA": ["UNITED STATES", "USA", "US"],
    "UNITED STATES": ["UNITED STATES", "USA", "US"],
    "U.S.A": ["UNITED STATES", "USA", "US"],
    "CHINA": ["CHINA", "PEOPLE'S REPUBLIC"],
    "MEXICO": ["MEXICO"],
    "BRAZIL": ["BRAZIL"],
    "INDIA": ["INDIA"],
    "QATAR": ["QATAR"],
    "NIGERIA": ["NIGERIA"],
    "AUSTRALIA": ["AUSTRALIA"],
    "GERMANY": ["GERMANY"],
    "UNITED KINGDOM": ["UNITED KINGDOM", "UK", "GREAT BRITAIN"],
}

MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


def log(msg):
    print(f"[ZIM] {msg}")


class ZimCdpPage:
    def __init__(self, ws_url):
        self.ws_url = ws_url
        self.ws = websocket.create_connection(
            ws_url,
            timeout=20,
            origin="http://127.0.0.1:%s" % ZIM_PORT,
            http_proxy_host=None,
            http_proxy_port=None,
        )
        self._seq = 0
        self.command("Runtime.enable")
        self.command("Page.enable")

    def close(self):
        try:
            self.ws.close()
        except Exception:
            pass

    def command(self, method, params=None, timeout=30):
        self._seq += 1
        msg_id = self._seq
        self.ws.settimeout(timeout)
        self.ws.send(json.dumps({"id": msg_id, "method": method, "params": params or {}}))
        while True:
            msg = json.loads(self.ws.recv())
            if msg.get("id") == msg_id:
                if "error" in msg:
                    raise RuntimeError(f"CDP {method}: {msg['error']}")
                return msg.get("result") or {}

    def eval(self, expression, timeout=30, await_promise=False):
        res = self.command(
            "Runtime.evaluate",
            {
                "expression": expression,
                "awaitPromise": bool(await_promise),
                "returnByValue": True,
            },
            timeout=timeout,
        )
        result = res.get("result") or {}
        if "exceptionDetails" in res:
            raise RuntimeError(f"CDP eval exception: {res['exceptionDetails']}")
        return result.get("value")

    @property
    def current_url(self):
        return self.eval("location.href", timeout=5) or ""

    def get(self, url):
        self.command("Page.navigate", {"url": url}, timeout=10)
        end = time.time() + 30
        while time.time() < end:
            try:
                state = self.eval("document.readyState", timeout=5)
                if state in ("interactive", "complete"):
                    return
            except Exception:
                pass
            time.sleep(0.3)
        raise TimeoutError(f"CDP navigate timeout: {url}")

    def fetch_json(self, url, timeout=60):
        expr = f"""
        (async () => {{
          const ctrl = new AbortController();
          const timer = setTimeout(() => ctrl.abort('timeout'), {int(timeout * 1000)});
          try {{
            const r = await fetch({json.dumps(url)}, {{ credentials: 'include', signal: ctrl.signal }});
            const text = await r.text();
            clearTimeout(timer);
            return JSON.stringify({{ok:r.ok, status:r.status, text}});
          }} catch (e) {{
            clearTimeout(timer);
            return JSON.stringify({{ok:false, status:0, text:String(e && e.message ? e.message : e)}});
          }}
        }})()
        """
        raw = self.eval(expr, timeout=timeout + 5, await_promise=True)
        result = json.loads(raw or "{}")
        if not result.get("ok"):
            raise RuntimeError(f"ZIM fetch failed HTTP {result.get('status')}: {(result.get('text') or '')[:220]}")
        return json.loads(result.get("text") or "{}")


def norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip()).upper()


def parse_valid_date(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if hasattr(raw, "year") and hasattr(raw, "month") and hasattr(raw, "day"):
        return datetime(raw.year, raw.month, raw.day)
    s = str(raw).strip()
    if not s:
        return None
    m = re.fullmatch(r"(\d{1,2})-([A-Za-z]{3})", s)
    if m:
        mon = MONTHS.get(m.group(2).upper())
        if mon:
            return datetime(TODAY.year, mon, int(m.group(1)))
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})", s)
    if m:
        return datetime(TODAY.year, int(m.group(2)), int(m.group(1)))
    for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d", "%d %b %Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            if dt.year == 1900:
                dt = dt.replace(year=TODAY.year)
            return dt
        except Exception:
            pass
    return None


def compute_valid_window(valid_dt):
    if not valid_dt:
        return None, None
    start = (TODAY + timedelta(days=DATE_OFFSET_DAYS)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    end = valid_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    return start, end


def compute_search_window(valid_dt):
    return compute_valid_window(valid_dt)


def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(("127.0.0.1", int(port))) == 0


def wait_port(port, timeout=15):
    end = time.time() + timeout
    while time.time() < end:
        if is_port_in_use(port):
            return True
        time.sleep(0.25)
    return False

def edge_debug_ready(timeout=2):
    try:
        import urllib.request
        with urllib.request.urlopen(f"http://127.0.0.1:{ZIM_PORT}/json/version", timeout=timeout) as resp:
            text = resp.read().decode("utf-8", errors="ignore")
        return "webSocketDebuggerUrl" in text or "Microsoft Edge" in text
    except Exception:
        return False

def cdp_page_ws_url():
    pages = json.loads(
        urllib.request.urlopen(f"http://127.0.0.1:{ZIM_PORT}/json", timeout=5)
        .read()
        .decode("utf-8", errors="ignore")
    )
    for p in pages:
        if p.get("type") == "page" and "zim.com" in (p.get("url") or ""):
            return p.get("webSocketDebuggerUrl")
    for p in pages:
        if p.get("type") == "page" and p.get("webSocketDebuggerUrl"):
            return p.get("webSocketDebuggerUrl")
    raise RuntimeError("Không tìm thấy page websocket ZIM")

def connect_cdp_page():
    last_exc = None
    for attempt in range(1, 3):
        try:
            if not edge_debug_ready(timeout=2):
                restart_zim_edge(f"CDP debug endpoint not ready attempt {attempt}")
            return ZimCdpPage(cdp_page_ws_url())
        except Exception as exc:
            last_exc = exc
            log(f"Connect CDP ZIM lỗi attempt {attempt}/2: {type(exc).__name__}: {exc}")
            restart_zim_edge(f"CDP connect failed attempt {attempt}")
    raise RuntimeError(f"Cannot connect CDP ZIM: {last_exc}")

def restart_zim_edge(reason=""):
    log(f"Restart Edge ZIM để dọn debug port treo. {reason}".strip())
    try:
        ps = (
            "$profile = '" + ZIM_PROFILE.replace("'", "''") + "'; "
            "Get-CimInstance Win32_Process -Filter \"name='msedge.exe'\" | "
            "Where-Object { $_.CommandLine -like ('*' + $profile + '*') } | "
            "ForEach-Object { Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }"
        )
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=8,
        )
    except Exception as exc:
        log(f"   Khong kill duoc Edge ZIM cu: {type(exc).__name__}")
    time.sleep(2)
    start_edge_if_needed(force_start=True)


def edge_exe_path():
    candidates = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return candidates[0]


def start_edge_if_needed(force_start=False):
    if is_port_in_use(ZIM_PORT) and not force_start:
        if edge_debug_ready(timeout=2):
            log(f"Edge debug port {ZIM_PORT} da mo san.")
            return
        log(f"Edge debug port {ZIM_PORT} dang treo/khong response.")
        restart_zim_edge("debug endpoint khong response")
        return
    log(f"Edge ZIM chua mo. Dang khoi dong port {ZIM_PORT}...")
    subprocess.Popen(
        [
            edge_exe_path(),
            f"--remote-debugging-port={ZIM_PORT}",
            "--remote-allow-origins=*",
            f"--user-data-dir={ZIM_PROFILE}",
            "--start-maximized",
            "--window-size=1920,1080",
            "--disable-background-timer-throttling",
            "--disable-renderer-backgrounding",
            "--disable-backgrounding-occluded-windows",
            ZIM_URL,
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    if not wait_port(ZIM_PORT, timeout=18):
        raise RuntimeError(f"Khong khoi dong duoc Edge ZIM port {ZIM_PORT}")


def connect_driver():
    last_exc = None
    for attempt in range(1, 3):
        if not edge_debug_ready(timeout=2):
            restart_zim_edge(f"before connect attempt {attempt}")
        opts = Options()
        opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{ZIM_PORT}")
        try:
            # Always use the verified workspace driver. Selenium Manager is
            # unreliable on this Python/Windows setup (win32 detection error)
            # and can stall before the bot reaches its actual work.
            if not os.path.exists(DRIVER_PATH):
                raise FileNotFoundError(f"Không tìm thấy Edge WebDriver: {DRIVER_PATH}")
            driver = webdriver.Edge(service=Service(executable_path=DRIVER_PATH), options=opts)
            break
        except Exception as exc:
            last_exc = exc
            log(f"Connect Edge ZIM lỗi attempt {attempt}/2: {type(exc).__name__}: {exc}")
            restart_zim_edge(f"connect failed attempt {attempt}")
    else:
        raise RuntimeError(f"Cannot connect Edge ZIM: {last_exc}")
    try:
        driver.maximize_window()
    except Exception:
        try:
            driver.set_window_size(1920, 1080)
        except Exception:
            pass
    return driver


def ensure_zim_page(driver):
    if isinstance(driver, ZimCdpPage):
        try:
            url = driver.current_url or ""
        except Exception:
            url = ""
        if "zim.com" not in url:
            driver.get(ZIM_URL)
        else:
            driver.get(ZIM_URL)
        time.sleep(2.0)
        return
    try:
        url = driver.current_url or ""
    except Exception:
        url = ""
    if "zim.com" not in url:
        driver.get(ZIM_URL)
    else:
        # The APIM CORS grant is page scoped; keep the page fresh at run start.
        driver.get(ZIM_URL)
    WebDriverWait(driver, 30).until(lambda d: d.execute_script("return document.readyState") in ("interactive", "complete"))
    time.sleep(2.0)


def browser_fetch_json(driver, url, timeout=60):
    if isinstance(driver, ZimCdpPage):
        return driver.fetch_json(url, timeout=timeout)
    driver.set_script_timeout(timeout)
    result = driver.execute_async_script(
        """
        const url = arguments[0];
        const done = arguments[arguments.length - 1];
        fetch(url, { credentials: 'include' })
          .then(async r => done({ ok: r.ok, status: r.status, text: await r.text() }))
          .catch(e => done({ ok: false, status: 0, text: String(e && e.message ? e.message : e) }));
        """,
        url,
    )
    if not result or not result.get("ok"):
        raise RuntimeError(f"ZIM fetch failed HTTP {result.get('status') if result else '?'}: {(result or {}).get('text', '')[:220]}")
    return json.loads(result.get("text") or "{}")


def alias_candidates(name):
    key = norm(name)
    candidates = PORT_ALIASES.get(key, [name])
    out = []
    seen = set()
    for item in candidates + [name]:
        item = str(item or "").strip()
        if item and item.upper() not in seen:
            out.append(item)
            seen.add(item.upper())
    return out


def country_tokens(country_hint):
    country = norm(country_hint or FILTER_COUNTRY)
    return [norm(x) for x in COUNTRY_ALIASES.get(country, [country]) if x]


def choose_port(items, original_name, country_hint=""):
    original = norm(original_name)
    country_words = country_tokens(country_hint)

    def score(item):
        name = norm(item.get("name"))
        country = norm(item.get("country"))
        code = norm(item.get("code"))
        loc_type = norm(item.get("locationType"))
        s = 0
        if any(tok and tok in country for tok in country_words):
            s += 100
        if "MARINE" in loc_type:
            s += 10
        if original and name.startswith(original):
            s += 25
        if original and original in name:
            s += 12
        if ";10" in code:
            s += 3
        return s

    return sorted(items, key=score, reverse=True)[0]


def resolve_port(driver, name, country_hint=""):
    last_items = []
    for query in alias_candidates(name):
        qs = urlencode({"searchTerm": query, "subscription-key": SUB_KEY})
        url = f"{APIM}/digitalMasterData/Locations/v1/p2p-route?{qs}"
        data = browser_fetch_json(driver, url, timeout=45)
        items = data.get("results") or []
        if items:
            last_items = items
            chosen = choose_port(items, name, country_hint)
            log(f"   Port '{name}' -> {chosen.get('name')} ({chosen.get('code')})")
            return chosen
    sample = ", ".join(str(x.get("name") or x.get("code") or "") for x in last_items[:5])
    raise ValueError(f"Port not found on ZIM: {name}" + (f" | sample: {sample}" if sample else ""))


def parse_iso_datetime(raw):
    s = str(raw or "").strip()
    if not s:
        return None
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.replace(tzinfo=None)
    except Exception:
        pass
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:19] if "T" in fmt else s[:10], fmt)
        except Exception:
            pass
    return None


def clean_port_name(raw):
    s = re.sub(r"\([^)]*\)", "", str(raw or "")).strip()
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def vessel_from_point(point):
    vessel = str(point.get("vesselName") or "").strip()
    if not vessel:
        full = str(point.get("fullVesselName") or "").strip()
        vessel = re.sub(r"\([^)]*\).*", "", full).strip()
    voyage = str(point.get("voyageNumber") or "").strip()
    leg = str(point.get("leg") or "").strip()
    voy = f"{voyage}{leg}".strip()
    if voy and voy.upper() not in vessel.upper():
        return f"{vessel} {voy}".strip()
    return vessel or "TBA"


def parse_zim_entries(payload):
    routes = payload.get("midPoints") or []
    entries = []
    for points in routes:
        if not isinstance(points, list) or not points:
            continue
        dep_point = next((p for p in points if p.get("departureDate")), None)
        arr_point = next((p for p in reversed(points) if p.get("arrivalDate")), None)
        if not dep_point or not arr_point:
            continue
        etd = parse_iso_datetime(dep_point.get("departureDate"))
        eta = parse_iso_datetime(arr_point.get("arrivalDate"))
        if not etd:
            continue
        if eta:
            tt_days = max(0, (eta.date() - etd.date()).days)
        else:
            tt_days = 0

        origin_code = norm(dep_point.get("portCode"))
        dest_code = norm(arr_point.get("portCode"))
        origin_name = clean_port_name(dep_point.get("portName"))
        dest_name = clean_port_name(arr_point.get("portName"))
        ts_ports = []
        for p in points[1:-1]:
            name = clean_port_name(p.get("portName"))
            code = norm(p.get("portCode"))
            if not name or name in {origin_name, dest_name} or code in {origin_code, dest_code}:
                continue
            if name not in ts_ports:
                ts_ports.append(name)
        entries.append(
            {
                "etd_dt": etd,
                "eta_dt": eta,
                "tt_days": tt_days,
                "vessel": vessel_from_point(dep_point),
                "ts_port": " + ".join(ts_ports) if ts_ports else "DIRECT",
            }
        )
    return entries


def fetch_schedule(driver, pol_code, pod_code, start_dt, end_dt):
    from_dt = max(start_dt, TODAY.replace(hour=0, minute=0, second=0, microsecond=0))
    weeks = max(4, min(12, math.ceil(((end_dt - from_dt).days + 1) / 7) + 2))
    params = {
        "PortCode": pol_code,
        "PortDestinationCode": pod_code,
        "Direction": "true",
        "FromDate": from_dt.strftime("%d-%B-%Y"),
        "WeeksAhead": str(weeks),
        "CountryCode": (pol_code.split(";")[0][:2] or "VN").upper(),
        "subscription-key": SUB_KEY,
    }
    url = f"{APIM}/digitalSchedules/PointToPoint/v2?{urlencode(params)}"
    return browser_fetch_json(driver, url, timeout=60)


def apply_etd_rules(entries, valid_dt):
    min_etd, _ = compute_search_window(valid_dt)
    min_etd_day = min_etd.date()
    valid_day = valid_dt.date() if valid_dt else None
    today_day = TODAY.date()

    # Normal rule: prefer ETD from today + DATE_OFFSET_DAYS, inside valid.
    future = [
        e for e in entries
        if e["etd_dt"].date() >= min_etd_day
        and etd_within_max(e["etd_dt"])
        and (valid_day is None or e["etd_dt"].date() <= valid_day)
    ]

    # Fallback rule: if today + DATE_OFFSET_DAYS already exceeds valid,
    # we cannot wait for a +7 ETD. Take exactly the farthest available ETD
    # that is still inside the valid window.
    if not future and valid_day is not None and min_etd_day > valid_day:
        inside_valid = [
            e for e in entries
            if today_day <= e["etd_dt"].date() <= valid_day
            and etd_within_max(e["etd_dt"])
        ]
        inside_valid.sort(key=lambda e: (e["etd_dt"], -int(e.get("tt_days") or 9999)), reverse=True)
        return inside_valid[:1]

    future.sort(key=lambda e: (e["etd_dt"], int(e.get("tt_days") or 9999)))
    selected = []
    for e in future:
        if len(selected) >= 3:
            break
        if selected and (e["etd_dt"] - selected[-1]["etd_dt"]).days < 1:
            continue
        selected.append(e)
    return selected


def fmt_date_short(dt):
    return f"{dt.day}-{dt.strftime('%b')}"


def format_etd_text(entries):
    if not entries:
        return ""
    fmt = [(e["etd_dt"].day, e["etd_dt"].strftime("%b")) for e in entries]
    if len(fmt) == 1:
        return f"{fmt[0][0]}-{fmt[0][1]}"
    if len(fmt) == 2:
        return f"{fmt[0][0]}-{fmt[0][1]} & {fmt[1][0]}-{fmt[1][1]}"
    months = [m for _, m in fmt]
    if months[0] == months[1] == months[2]:
        return f"{fmt[0][0]}, {fmt[1][0]}, {fmt[2][0]}-{fmt[2][1]}"
    if months[0] == months[1]:
        return f"{fmt[0][0]}, {fmt[1][0]}-{fmt[1][1]} & {fmt[2][0]}-{fmt[2][1]}"
    if months[1] == months[2]:
        return f"{fmt[0][0]}-{fmt[0][1]}, {fmt[1][0]} & {fmt[2][0]}-{fmt[2][1]}"
    return f"{fmt[0][0]}-{fmt[0][1]}, {fmt[1][0]}-{fmt[1][1]} & {fmt[2][0]}-{fmt[2][1]}"


def format_tt_text(entries):
    tts = [int(e.get("tt_days") or 0) for e in entries if e.get("tt_days") is not None]
    if not tts:
        return ""
    if len(set(tts)) == 1:
        return str(tts[0])
    return f"{min(tts)}-{max(tts)}"


def format_vessel_block(entries):
    lines = []
    ts_seen = []
    for e in entries:
        ts = e.get("ts_port") or "DIRECT"
        lines.append(
            f"{e.get('vessel') or 'TBA'} / ETD: {fmt_date_short(e['etd_dt'])}"
            f" / Transit time: {int(e.get('tt_days') or 0)} Days / Transshipment Port: {ts}"
        )
        if ts not in ts_seen:
            ts_seen.append(ts)
    return "\n".join(lines), " or\n".join(ts_seen) if ts_seen else "DIRECT"


def write_excel_row(row_i, etd_text, tt_text, vessel_text, ts_text, error=None, valid_dt=None):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        if ws.cell(row=row_i, column=6).value is None:
            ws.cell(row=row_i, column=6).value = "-"
        if valid_dt and not ws.cell(row=row_i, column=11).value:
            ws.cell(row=row_i, column=11).value = f"{valid_dt.day}-{valid_dt.strftime('%b')}"
        if error:
            ws.cell(row=row_i, column=9).value = "-"
            ws.cell(row=row_i, column=10).value = ""
            ws.cell(row=row_i, column=15).value = ""
            ws.cell(row=row_i, column=16).value = ""
        else:
            ws.cell(row=row_i, column=9).value = etd_text
            ws.cell(row=row_i, column=10).value = tt_text
            ws.cell(row=row_i, column=15).value = vessel_text
            ws.cell(row=row_i, column=16).value = ts_text
            wrap = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_i, column=15).alignment = wrap
            ws.cell(row=row_i, column=16).alignment = wrap
        wb.save(EXCEL_PATH)
        log(f"   Wrote Excel row {row_i}")
    except PermissionError:
        log("   Cannot write Excel: close input_gia.xlsx then run again.")
    except Exception as exc:
        log(f"   write_excel_row error: {exc}")


def search_one(driver, pol_excel, pod_excel, country_hint, valid_dt):
    start_dt, end_dt = compute_search_window(valid_dt)
    if not start_dt or not end_dt:
        return {"error": "VALID rong/khong parse duoc"}
    if end_dt < start_dt:
        if end_dt.date() < TODAY.date():
            return {"error": "VALID da qua ngay hien tai"}
        log(
            f"   VALID {end_dt:%Y-%m-%d} < today+{DATE_OFFSET_DAYS} "
            f"({start_dt:%Y-%m-%d}) -> fallback search today..valid"
        )
        start_dt = TODAY.replace(hour=0, minute=0, second=0, microsecond=0)

    pol = resolve_port(driver, pol_excel, "")
    pod = resolve_port(driver, pod_excel, country_hint)
    log(f"   APIM: {pol.get('code')} -> {pod.get('code')} | {start_dt:%Y-%m-%d}..{end_dt:%Y-%m-%d}")
    payload = fetch_schedule(driver, pol.get("code"), pod.get("code"), start_dt, end_dt)
    entries = parse_zim_entries(payload)
    log(f"   APIM routes={len(payload.get('midPoints') or [])} parsed={len(entries)}")
    selected = apply_etd_rules(entries, end_dt)
    if not selected:
        return {"error": "Khong co lich tau hop le"}
    etd_text = format_etd_text(selected)
    tt_text = format_tt_text(selected)
    vessel_text, ts_text = format_vessel_block(selected)
    return {
        "etd_text": etd_text,
        "tt_text": tt_text,
        "vessel_text": vessel_text,
        "ts_text": ts_text,
    }


def main():
    if not os.path.exists(EXCEL_PATH):
        log(f"Excel not found: {EXCEL_PATH}")
        return

    log(f"DATE_OFFSET_DAYS={DATE_OFFSET_DAYS} (search from today+{DATE_OFFSET_DAYS})")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    target_single_row = None
    if SINGLE_ROW:
        try:
            target_single_row = int(SINGLE_ROW)
            log(f"[SINGLE_ROW] Chi chay dong {target_single_row} theo main.py")
        except Exception:
            log(f"[SINGLE_ROW] Khong hop le: {SINGLE_ROW}")

    targets = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if target_single_row is not None and i != target_single_row:
            continue
        country = str(row[1] or "").strip() if len(row) > 1 else ""
        pol = str(row[2] or "").strip() if len(row) > 2 else ""
        pod = str(row[3] or "").strip() if len(row) > 3 else ""
        carrier = str(row[4] or "").strip().upper() if len(row) > 4 else ""
        valid_raw = row[10] if len(row) > 10 else None
        if not pol or not pod or carrier not in CARRIER_TARGETS:
            continue
        if FILTER_POL and pol.upper() != FILTER_POL:
            continue
        if FILTER_POD and pod.upper() != FILTER_POD:
            continue
        targets.append((i, country, pol, pod, valid_raw))
    wb.close()

    log(f"Co {len(targets)} dong can check (carrier = ZIM)")
    if not targets:
        return

    try:
        start_edge_if_needed()
        if ZIM_USE_SELENIUM:
            driver = connect_driver()
        else:
            try:
                driver = connect_cdp_page()
            except Exception as cdp_exc:
                log(
                    "CDP websocket khong ket noi duoc "
                    f"({type(cdp_exc).__name__}) -> fallback Selenium tren cung Edge."
                )
                driver = connect_driver()
        ensure_zim_page(driver)
    except Exception as exc:
        err = f"ZIM browser connect failed: {type(exc).__name__}: {exc}"
        log(f"ERROR: {err}")
        # Do not write the connection error into every schedule cell and then
        # exit with code 0.  A real failure must reach main.py so its copy is
        # excluded from merge and the previous Excel data stays intact.
        raise RuntimeError(err) from exc

    for idx, (row_i, country, pol, pod, valid_raw) in enumerate(targets, start=1):
        log(f"\n========== [{idx}/{len(targets)}] DONG {row_i}: {pol} -> {pod} | VALID={valid_raw}")
        valid_dt = parse_valid_date(valid_raw)
        if not valid_dt:
            log("   Valid rong -> bo qua row, khong goi API.")
            continue
        try:
            result = search_one(driver, pol, pod, country, valid_dt)
        except Exception as exc:
            result = {"error": f"Exception: {exc}"}
        if result.get("error"):
            log(f"   ERROR: {result['error']}")
            write_excel_row(row_i, "", "", "", "", error=result["error"])
        else:
            log(f"   ETD={result['etd_text']} | TT={result['tt_text']} | TS={result['ts_text']}")
            write_excel_row(
                row_i,
                result["etd_text"],
                result["tt_text"],
                result["vessel_text"],
                result["ts_text"],
            )
        time.sleep(0.8)

    try:
        if isinstance(driver, ZimCdpPage):
            driver.close()
    except Exception:
        pass
    log("\nDONE bot ZIM")


if __name__ == "__main__":
    main()
