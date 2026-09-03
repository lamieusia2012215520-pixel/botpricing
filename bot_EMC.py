"""
EMC Price Checker - API-only mode for the new GreenX site.
Uses direct GreenX HTTP endpoints; Selenium is kept only behind EMC_LEGACY_BROWSER=1.

Cài đặt:
    pip install selenium openpyxl

Chạy:
    python bot_EMC.py
"""

import math
import re
import time
import os
import sys
import json
import calendar
import urllib.request
import requests
from datetime import datetime, timedelta
from html.parser import HTMLParser

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from emc_greenx_logic import (
    build_quote_detail_payload,
    hydrate_quote_with_detail,
    quote_departure_is_on_or_before,
)
from remark_rules import build_subject_remark, is_china_destination

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


current_folder = os.getcwd()
driver_path    = os.path.join(current_folder, "msedgedriver.exe")
excel_path     = os.environ.get("EXCEL_PATH", os.path.join(current_folder, "input_gia.xlsx"))
FILTER_POL     = os.environ.get("FILTER_POL", "").strip().upper()
FILTER_POD     = os.environ.get("FILTER_POD", "").strip().upper()
SINGLE_ROW     = os.environ.get("SINGLE_ROW", "").strip()
DATE_OFFSET_DAYS = parse_date_offset_days()
try:
    EMC_ROW_WAIT_SECONDS = max(0.0, float(os.environ.get("EMC_ROW_WAIT_SECONDS", "1")))
except ValueError:
    EMC_ROW_WAIT_SECONDS = 1.0
try:
    EMC_MAX_RETRIES = max(1, int(os.environ.get("EMC_MAX_RETRIES", "2")))
except ValueError:
    EMC_MAX_RETRIES = 2
# API-only mode: không fallback về Selenium scraping nếu API lỗi.
# Giữ biến này cố định để tránh ai đó set ENV làm bot quay lại Selenium.
EMC_API_MODE = True
EMC_USE_BROWSER = os.environ.get("EMC_LEGACY_BROWSER", "0").strip() == "1"

import subprocess
import socket

def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('127.0.0.1', port)) == 0

def _wait_port(port, timeout=8):
    """FIX: Poll port thay vì sleep cứng."""
    end = time.time() + timeout
    while time.time() < end:
        if is_port_in_use(port):
            return True
        time.sleep(0.2)
    return False

driver = None
if EMC_USE_BROWSER:
    if not is_port_in_use(9521):
        print("[HE THONG] Edge EMC legacy chua mo. Dang khoi dong...")
        try:
            subprocess.Popen([
                r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
                "--headless=new",
                "--remote-debugging-port=9521",
                r"--user-data-dir=C:\edge_emc",
                "--window-size=1920,1080",
                "--disable-gpu",
                "--disable-dev-shm-usage",
                "--no-sandbox",
                "--disable-background-timer-throttling",
                "--disable-renderer-backgrounding",
                "--disable-backgrounding-occluded-windows"
            ])
            _wait_port(9521, timeout=8)
        except Exception:
            pass
    else:
        print("[HE THONG] Edge EMC legacy da mo san.")

    edge_options = Options()
    edge_options.add_argument("--headless=new")
    edge_options.add_argument("--window-size=1920,1080")
    edge_options.add_argument("--disable-gpu")
    edge_options.add_experimental_option("debuggerAddress", "127.0.0.1:9521")
    service = Service(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=edge_options)
    try:
        driver.set_window_size(1920, 1080)
    except Exception:
        pass

    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
else:
    print("[HE THONG] EMC API moi: khong mo Edge.")
def recreate_emc_driver():
    global driver
    print("[HỆ THỐNG] Phiên Edge EMC hỏng, khởi động lại riêng port 9521...")
    try:
        driver.quit()
    except Exception:
        pass
    try:
        subprocess.run([
            "powershell", "-NoProfile", "-Command",
            "Get-CimInstance Win32_Process -Filter \"name='msedge.exe'\" | "
            "Where-Object { $_.CommandLine -like '*--remote-debugging-port=9521*' } | "
            "ForEach-Object { Stop-Process -Id $_.ProcessId -Force }"
        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=10)
    except Exception:
        pass
    time.sleep(1)
    subprocess.Popen([
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        "--headless=new",
        "--remote-debugging-port=9521",
        r"--user-data-dir=C:\edge_emc",
        "--window-size=1920,1080",
        "--disable-gpu",
        "--disable-dev-shm-usage",
        "--no-sandbox",
        "--disable-background-timer-throttling",
        "--disable-renderer-backgrounding",
        "--disable-backgrounding-occluded-windows"
    ])
    _wait_port(9521, timeout=8)
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-gpu")
    opts.add_experimental_option("debuggerAddress", "127.0.0.1:9521")
    driver = webdriver.Edge(service=Service(executable_path=driver_path), options=opts)
    try:
        driver.maximize_window()
    except Exception:
        pass
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    return driver

def ensure_emc_session():
    if not EMC_USE_BROWSER:
        return True
    try:
        driver.current_url
        return True
    except Exception:
        try:
            recreate_emc_driver()
            driver.current_url
            return True
        except Exception as e:
            print(f"[HỆ THỐNG] ❌ Không phục hồi được Edge EMC: {clean_emc_error(e)}")
            return False

BOOKING_FEE = 10
BASE_URL     = "https://portal.greenxtrade.com/quotes"
PORTS_JS_URL = "https://cdn.greenxtrade.com/data/ports.js"
EMC_GROUP    = {"EMC", "EVERGREEN"}

# For EMC India rates, the displayed total is treated as THC/L-inclusive even
# when the website lists THC/L under Included.  PIO's Excel convention stores
# the ocean rate before O.THC, so remove the fixed THC/L amount and mark THC as
# subject instead.
EMC_INDIA_THC_DEDUCTION = {"20": 140, "40": 210, "40hq": 210}
EMC_INDIA_PODS = {
    "CHENNAI", "COCHIN", "KOLKATA", "MUNDRA", "NHAVA SHEVA",
    "NHAVASHEVA", "PARADIP", "PIPAVAV", "TUTICORIN",
    "VISAKHAPATNAM", "VIZAG",
}


def is_emc_india_route(country, pod):
    country_key = re.sub(r"\s+", " ", str(country or "").upper()).strip()
    pod_key = re.sub(r"\s+", " ", str(pod or "").upper()).strip()
    return country_key in {"INDIA", "IN", "IND"} or pod_key in EMC_INDIA_PODS


def force_subject_to_thc(remark, pod=""):
    text = str(remark or "").strip()
    text = re.sub(r"(?i)(?:INCLUDED|INCL(?:UDED)?)\s+O?\.?THC(?:/L)?\s*,?\s*", "", text)
    match = re.search(r"(?i)SUBJECT\s+TO\s+(.+)", text)
    if not match:
        return build_subject_remark(othc_included=False, pod=pod)

    items = [item.strip().upper() for item in match.group(1).split(",") if item.strip()]
    items = [item for item in items if item not in {"THC", "O.THC", "THC/L"}]
    return "SUBJECT TO " + ", ".join(["THC"] + items)


def apply_emc_india_thc_rule(result, country, pod):
    if not result or result.get("error") or not is_emc_india_route(country, pod):
        return result

    for suffix, deduction in EMC_INDIA_THC_DEDUCTION.items():
        price_key = f"price_{suffix}"
        formula_key = f"formula_{suffix}"
        price = result.get(price_key)
        if isinstance(price, (int, float)):
            result[price_key] = price - deduction

        formula = str(result.get(formula_key) or "").strip()
        if formula.startswith("="):
            result[formula_key] = f"{formula}-{deduction}"

    result["remark"] = force_subject_to_thc(result.get("remark"), pod=pod)
    return result
CHINA_PODS = {
    "SHANGHAI", "NINGBO", "QINGDAO", "XINGANG", "TIANJIN", "YANTIAN",
    "SHEKOU", "NANSHA", "XIAMEN", "DALIAN", "HUANGPU", "FUZHOU",
    "ZHONGSHAN", "ZHANJIANG", "CHIWAN", "CHINA"
}

def is_china_pod(port_name):
    return is_china_destination(pod=port_name)

def get_valid_date(etd_dates):
    latest_etd = max(etd_dates)
    day = latest_etd.day
    if day <= 7:
        valid_day = 7
    elif day <= 14:
        valid_day = 14
    elif day <= 21:
        valid_day = 21
    else:
        valid_day = calendar.monthrange(latest_etd.year, latest_etd.month)[1]
    vd = datetime(latest_etd.year, latest_etd.month, valid_day)
    return f"{vd.day}-{vd.strftime('%b')}"

def reload_emc_base(reason=""):
    if not EMC_USE_BROWSER:
        if reason:
            print(f"[HE THONG] EMC API reset do: {reason}")
        greenx_api_reset()
        return
    if reason:
        print(f"[HỆ THỐNG] Reload base URL EMC do: {reason}")
    ensure_emc_session()
    driver.get(BASE_URL)
    WebDriverWait(driver, 20).until(
        lambda d: d.find_elements(By.XPATH, '//input[@aria-label="input for From"]')
               or d.find_elements(By.XPATH, '//input[@aria-label="input for Origin"]')
    )
    time.sleep(1)

# ===================================================================================
# --- CHỌN CẢNG ĐÚNG TỪ DROPDOWN ---
# ===================================================================================
def pick_correct_port(options_text, query):
    query_upper = query.upper().strip()

    # Ưu tiên (All Ports)
    for opt in options_text:
        if "(all ports)" in opt.lower():
            return opt

    # Quy tắc dấu phẩy ngay sau tên
    for opt in options_text:
        if opt.upper().strip().startswith(query_upper + ","):
            return opt

    # Fallback: chứa query
    for opt in options_text:
        if query_upper in opt.upper():
            return opt

    return None


def select_port_emc(xpath_input, port_name):
    print(f"      + Nạp cảng: {port_name}")
    # FIX: tăng số attempt 3→4 và chờ dropdown lâu hơn một chút
    # khi web load chậm — và retry bằng BACKSPACE nếu không khớp.
    for attempt in range(4):
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath_input)))
            inp = next((e for e in driver.find_elements(By.XPATH, xpath_input) if e.is_displayed()), None)
            if not inp:
                raise Exception("Không tìm thấy input!")

            # Check giữ nguyên
            current_val = driver.execute_script("return arguments[0].value;", inp)
            current_upper = str(current_val or "").upper()
            is_selected_option = (
                "(ALL PORTS)" in current_upper
                or re.search(r",\s*[A-Z]{2}\s*,\s*[A-Z0-9]{4,}", current_upper)
            )
            if current_val and port_name.upper() in current_upper and is_selected_option:
                print(f"        -> Đã có sẵn: {current_val} -> GIỮ NGUYÊN!")
                return

            # Vue hack: set value + dispatch input event
            driver.execute_script("""
                var inp = arguments[0];
                var val = arguments[1];
                inp.focus();
                var setter = Object.getOwnPropertyDescriptor(
                    window.HTMLInputElement.prototype, 'value').set;
                setter.call(inp, val);
                inp.dispatchEvent(new Event('input', { bubbles: true }));
            """, inp, port_name)

            # Chờ ul.open xuất hiện
            WebDriverWait(driver, 8).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "ul.open li[data-arrow-control]"))
            )
            time.sleep(0.2)

            opt_spans = driver.find_elements(By.CSS_SELECTOR, "ul.open li[data-arrow-control] span.item-text")
            opts_text = [s.text.strip() for s in opt_spans if s.text.strip()]
            opts      = driver.find_elements(By.CSS_SELECTOR, "ul.open li[data-arrow-control]")
            print(f"        Dropdown: {opts_text[:5]}")

            chosen = pick_correct_port(opts_text, port_name)
            if not chosen:
                # FIX: thử BACKSPACE 1 ký tự để dropdown match rộng hơn
                if attempt < 2 and len(port_name) > 3:
                    try:
                        inp.send_keys(Keys.BACKSPACE)
                        time.sleep(0.3)
                    except Exception:
                        pass
                raise Exception(f"Không khớp option nào cho '{port_name}'")

            for li in opts:
                try:
                    if li.find_element(By.CSS_SELECTOR, "span.item-text").text.strip() == chosen:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", li)
                        time.sleep(0.1)
                        driver.execute_script("arguments[0].click();", li)
                        print(f"        -> Đã chốt: {chosen}")
                        time.sleep(0.25)
                        return
                except: continue

        except Exception as e:
            print(f"        ⚠️ Lần {attempt+1} thất bại: {clean_emc_error(e)}")
            time.sleep(0.4)

    raise Exception(f"Thất bại 4 lần: {port_name}")


# ===================================================================================
# --- 9 QUY TẮC VÀNG (copy từ bot COSCO) ---
# ===================================================================================
def apply_9_golden_rules(danh_sach_chuyen):
    danh_sach_chuyen = [c for c in danh_sach_chuyen if etd_within_max(c.get("etd_dt"))]
    if not danh_sach_chuyen:
        return [], "", ""
    danh_sach_chuyen.sort(key=lambda x: (x["etd_dt"], x["tt_days"]))

    # Lọc trùng ETD (giữ cái TT ngắn nhất)
    seen, list_loc_trung = set(), []
    for c in danh_sach_chuyen:
        if c["etd_dt"] not in seen:
            list_loc_trung.append(c)
            seen.add(c["etd_dt"])

    etd_dat_chuan = []
    if list_loc_trung:
        ngan_nhat = min(c["tt_days"] for c in list_loc_trung)
        first_date = list_loc_trung[0]["etd_dt"]
        for c in list_loc_trung:
            if len(etd_dat_chuan) >= 3: break
            if len(etd_dat_chuan) > 0 and (c["etd_dt"] - etd_dat_chuan[-1]["etd_dt"]).days < 2: continue
            if (c["etd_dt"] - first_date).days <= 9 and c["tt_days"] <= ngan_nhat + 10:
                etd_dat_chuan.append(c)

    # Format ETD string
    def _fmt_etd(dt):
        return f"{dt.day}-{dt.strftime('%b')}"
    num = len(etd_dat_chuan)
    if num == 0:   str_etd = "N/A"
    elif num == 1: str_etd = _fmt_etd(etd_dat_chuan[0]["etd_dt"])
    elif num == 2: str_etd = f"{_fmt_etd(etd_dat_chuan[0]['etd_dt'])} & {_fmt_etd(etd_dat_chuan[1]['etd_dt'])}"
    else:
        d1 = str(etd_dat_chuan[0]["etd_dt"].day)
        d2 = str(etd_dat_chuan[1]["etd_dt"].day)
        d3 = _fmt_etd(etd_dat_chuan[2]["etd_dt"])
        str_etd = f"{d1}, {d2}, {d3}"

    all_tt = [c["tt_days"] for c in etd_dat_chuan]
    str_tt = f"{min(all_tt)}" if min(all_tt) == max(all_tt) else f"{min(all_tt)}-{max(all_tt)}"

    return etd_dat_chuan, str_etd, str_tt


# ===================================================================================
# --- PARSE BẢNG GIÁ TỪ PRICE DETAILS ---
# ===================================================================================
BLOCKLIST = [
    'THC', 'TERMINAL HANDLING', 'DESTINATION', 'D/O', 'DELIVERY ORDER',
    'CFS', 'PORT CONGESTION', 'PSS', 'EBS', 'ERC', 'STF',
]

def parse_usd(text):
    text = re.sub(r'\(.*?\)', '', text)
    match = re.search(r'\$?([\d,]+\.?\d*)', text.replace('USD', '').strip())
    if match:
        return float(match.group(1).replace(",", ""))
    return 0.0

def parse_free_time_text(text):
    text = " ".join((text or "").upper().replace("\xa0", " ").split())
    if not text:
        return ""

    def pick_day(label_patterns):
        for label in label_patterns:
            patterns = [
                rf"{label}[^0-9]{{0,40}}(\d{{1,3}})",
                rf"(\d{{1,3}})[^A-Z0-9]{{0,20}}{label}",
            ]
            for pattern in patterns:
                m = re.search(pattern, text)
                if m:
                    return m.group(1)
        return ""

    combined = pick_day([r"COMBINED", r"D\s*&\s*D", r"DND"])
    if combined:
        return f"{combined} COMBINED"

    dem = pick_day([r"DEMURRAGE", r"\bDEM\b"])
    det = pick_day([r"DETENTION", r"\bDET\b"])
    if dem and det:
        return f"{dem} DEM + {det} DET"
    if dem:
        return f"{dem} DEM"
    if det:
        return f"{det} DET"
    return ""

def parse_destination_free_time_text(text):
    text = (text or "").replace("\xa0", " ")
    if not text.strip():
        return ""

    compact = " ".join(text.upper().split())
    marker = "TARIFF FREE TIME AT DESTINATION"
    marker_pos = compact.find(marker)
    if marker_pos < 0:
        marker_pos = compact.find("FREE TIME AT DESTINATION")
    if marker_pos < 0:
        return ""

    dest_text = compact[marker_pos:]
    next_origin = dest_text.find("TARIFF FREE TIME AT ORIGIN", len(marker))
    if next_origin > 0:
        dest_text = dest_text[:next_origin]

    def pick(label):
        m = re.search(rf"(?:{label})[^0-9]{{0,80}}(\d{{1,3}})\s*CALENDAR\s*DAYS?", dest_text)
        if not m:
            m = re.search(rf"(?:{label})[^0-9]{{0,80}}(\d{{1,3}})", dest_text)
        return m.group(1) if m else ""

    dem = pick(r"CONTAINER\s+DEMURRAGE|DEMURRAGE|\bDEM\b")
    det = pick(r"CONTAINER\s+DETENTION|DETENTION|\bDET\b")
    combined = pick(r"COMBINED|D\s*&\s*D|DND")
    usage = pick(r"CONTAINER\s+USAGE|USAGE")

    if combined:
        return f"{combined} COMBINED"
    if usage:
        return f"{usage} DAYS"
    if dem and det:
        return f"{dem} DEM + {det} DET"
    if dem:
        return f"{dem} DEM"
    if det:
        return f"{det} DET"
    return ""

def scrape_free_time_from_emc_card(card, subrow_el=None):
    """
    EMC đổi UI khá thường xuyên. Chỉ lấy Free Time ở POD:
    block "Tariff Free Time at Destination", bỏ qua Origin.
    """
    scopes = [x for x in [subrow_el, card] if x is not None]
    try:
        free_tabs = card.find_elements(
            By.XPATH,
            ".//*[contains(@class,'quotes-search-results-list-item-tabs-tab')][.//span[normalize-space()='Free Time'] or normalize-space()='Free Time']"
        )
        if not free_tabs:
            free_tabs = card.find_elements(
                By.XPATH,
                ".//*[self::button or @role='tab' or contains(@class,'tab')][.//*[normalize-space()='Free Time'] or normalize-space()='Free Time']"
            )

        for target in free_tabs[:3]:
            try:
                if not target.is_displayed():
                    continue
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
                time.sleep(0.2)
                ActionChains(driver).move_to_element(target).click().perform()
                WebDriverWait(driver, 8).until(
                    lambda d: any(
                        "TARIFF FREE TIME AT DESTINATION" in ((scope.text or "").upper())
                        for scope in scopes
                    ) or "TARIFF FREE TIME AT DESTINATION" in (d.find_element(By.TAG_NAME, "body").text or "").upper()
                )
                time.sleep(0.5)
                parsed = ""
                for scope in scopes:
                    parsed = parse_destination_free_time_text(scope.text)
                    if parsed:
                        break
                if not parsed:
                    parsed = parse_destination_free_time_text(driver.find_element(By.TAG_NAME, "body").text)
                if parsed:
                    return parsed
            except Exception:
                continue
    except Exception as e:
        print(f"      ⚠️ Lỗi click Free Time EMC: {clean_emc_error(e)}")

    for scope in scopes:
        try:
            parsed = parse_destination_free_time_text(scope.text)
            if parsed:
                return parsed
        except Exception:
            continue
    return ""

# ===================================================================================
# --- EXTRACT VESSEL + TRANSSHIPMENT TỪ CARD ---
# ===================================================================================
VESSEL_SELECTORS = [
    ".vessel-name", ".vessel", "[class*='vessel']", "[class*='Vessel']",
    ".voyage-vessel", ".ship-name", "[data-test*='vessel']",
]

def clean_emc_vessel_name(text):
    txt = " ".join((text or "").strip().split())
    txt = re.sub(r"^service\s+vessel\s+voyage\s+cit\s+", "", txt, flags=re.I).strip()
    txt = re.sub(r"^service\s+vessel\s+voyage\s+", "", txt, flags=re.I).strip()
    txt = re.sub(r"^vessel\s*[:\-]?\s*", "", txt, flags=re.I).strip()
    txt = re.sub(r"^(?:CIT|CVM|NE3|CEM|CES|NCI|NSC|JCV|CVS|AIS|AUE|AAS)\s+", "", txt, flags=re.I).strip()
    return txt

def normalize_emc_first_vessel(text):
    txt = clean_emc_vessel_name(text)
    if not txt:
        return ""
    txt = re.sub(r"\b(?:CIT|CVM|NE3|CEM|CES|NCI|NSC|JCV|CVS|AIS|AUE|AAS)\b\s+(?=[A-Z][A-Z\s]+?\s+\d{3,5}[-/]\d{2,4}[A-Z])", " ", txt, flags=re.I).strip()
    m = re.search(r"([A-Z][A-Z ]*?)\s+(\d{3,5}[-/]\d{2,4}[A-Z])", txt)
    if m:
        return f"{m.group(1).strip()} {m.group(2).strip()}"
    lines = [clean_emc_vessel_name(x) for x in (text or "").splitlines() if clean_emc_vessel_name(x)]
    return lines[0] if lines else txt

def is_complete_emc_vessel(text):
    return bool(re.search(r"[A-Z]{2,}\s+\d{3,5}[-/]\d{2,4}[A-Z]", text or ""))

def extract_vessel_from_card(card):
    """Best-effort: trả về tên tàu từ card EMC. Nếu không tìm được → 'TBA'."""
    try:
        full = (card.text or "")
        m = re.search(r"Vessel\s+Voyage\s+(.+?)(?:\n\s*Route Details|\n\s*Price Details|\n\s*Free Time|\n\s*USD|\n\s*Book)", full, re.I | re.S)
        if m:
            first = normalize_emc_first_vessel(m.group(1))
            if first:
                return first
    except Exception:
        pass
    for sel in VESSEL_SELECTORS:
        try:
            elems = card.find_elements(By.CSS_SELECTOR, sel)
            for el in elems:
                txt = normalize_emc_first_vessel(el.text)
                if txt and len(txt) < 80 and not txt.lower().startswith("vessel"):
                    return txt
        except Exception:
            continue
    # Fallback: tìm trong text card pattern "VESSEL ... VOY ..."
    try:
        full = (card.text or "")
        m = re.search(r"Vessel[:\s]+([^\n/]+?)(?:\s+Voy|/|\n)", full, re.I)
        if m:
            return clean_emc_vessel_name(m.group(1))[:60]
    except Exception:
        pass
    return "TBA"

def read_emc_route_details(card):
    """Click Route Details and return first vessel + T/S port from the expanded route tab."""
    vessel = ""
    ts_port = ""
    try:
        route_tab = card.find_element(
            By.XPATH,
            ".//div[contains(@class,'quotes-search-results-list-item-tabs-tab')]"
            "[.//span[normalize-space()='Route Details'] or normalize-space()='Route Details']"
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", route_tab)
        time.sleep(0.15)
        driver.execute_script("arguments[0].click();", route_tab)
        time.sleep(0.6)

        text = card.text or ""
        m = re.search(r"\bT/S\s+PORT\b\s+([A-Za-z][A-Za-z\s,.-]*?)(?:\s+POD\b|\s+PDL\b|\n)", text, re.I)
        if m:
            ts_port = " ".join(m.group(1).replace("\n", " ").split()).strip(" ,-").upper()

        v_match = re.search(r"Vessel\s+Voyage\s+(.+?)(?:\n\s*Route Details|\n\s*Price Details|\n\s*Free Time|\n\s*USD|\n\s*Book)", text, re.I | re.S)
        if v_match:
            vessel = normalize_emc_first_vessel(v_match.group(1))
            if not is_complete_emc_vessel(vessel):
                vessel = ""

        if not vessel:
            route_lines = [x.strip() for x in text.splitlines() if x.strip()]
            for line in route_lines:
                candidate = normalize_emc_first_vessel(line)
                if is_complete_emc_vessel(candidate):
                    vessel = candidate
                    break
    except Exception as e:
        print(f"      [WARN] Khong doc duoc Route Details EMC: {clean_emc_error(e)}")
    return vessel, ts_port

def extract_transshipment_from_card(card):
    """Best-effort: trả về tên cảng transshipment. Mặc định 'DIRECT'."""
    try:
        full = (card.text or "")
    except Exception:
        full = ""
    # Tìm patterns: "T/S: PORT", "Via PORT", "Transshipment: PORT"
    for pat in [
        r"T/?S[:\s]+([A-Z][A-Z\s,]{2,40}?)(?:\n|$|/)",
        r"Via[:\s]+([A-Z][A-Z\s,]{2,40}?)(?:\n|$|/)",
        r"Tranship(?:ment)?[:\s]+([A-Z][A-Z\s,]{2,40}?)(?:\n|$|/)",
    ]:
        m = re.search(pat, full, re.I)
        if m:
            port = m.group(1).strip().rstrip(',').strip()
            if port and port.upper() not in ("DIRECT", "NONE", "N/A"):
                return port.upper()
    # Detect "Direct" / "DIRECT" explicitly
    if re.search(r"\bDirect\b", full, re.I):
        return "DIRECT"
    return "DIRECT"

def clean_emc_error(err):
    text = str(err or "").strip()
    upper = text.upper()
    if not text:
        return "WEB TIMEOUT / LAG"
    if "STACKTRACE" in upper or "GETHANDLEVERIFIER" in upper or "TIMEOUT" in upper:
        return "WEB TIMEOUT / LAG"
    if "NO SERVICE" in upper or "SOLD OUT" in upper or "JOIN WAITLIST" in upper:
        return "NO SERVICE / SOLD OUT"
    return text.splitlines()[0][:120]


def card_has_book(card):
    try:
        buttons = card.find_elements(By.XPATH, ".//*[self::button or @role='button']")
        for btn in buttons:
            text = " ".join((btn.text or "").split()).strip().lower()
            if text == "book" and btn.is_displayed() and btn.is_enabled():
                return True
    except Exception:
        pass
    return False

EMC_PORT_CACHE = None
EMC_PORT_ALIASES = {
    "HO CHI MINH": "VNBDGALL",
    "HOCHIMINH": "VNBDGALL",
    "HCM": "VNBDGALL",
    "SAIGON": "VNBDGALL",
    "HAI PHONG": "VNHPH",
    "HAIPHONG": "VNHPH",
    "SHANGHAI": "CNSHA",
    "NINGBO": "CNNGB",
    "QINGDAO": "CNTAO",
    "XIAMEN": "CNXMN",
    "SHEKOU": "CNSHK",
    "NANSHA": "CNNSA",
    "YANTIAN": "CNYTN",
    "TIANJIN": "CNTSNALL",
    "XINGANG": "CNTXG",
    "HAMBURG": "DEHAM",
}

def _norm_port_name(text):
    return re.sub(r"[^A-Z0-9]+", " ", str(text or "").upper()).strip()

def load_emc_ports():
    global EMC_PORT_CACHE
    if EMC_PORT_CACHE is not None:
        return EMC_PORT_CACHE

    raw = ""
    local_ports = os.path.join(current_folder, "tmp_greenx_js", "ports.js")
    try:
        if os.path.exists(local_ports):
            with open(local_ports, "r", encoding="utf-8") as f:
                raw = f.read()
    except Exception:
        raw = ""

    if not raw:
        try:
            with urllib.request.urlopen(PORTS_JS_URL, timeout=15) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
        except Exception:
            raw = ""

    entries = []
    for m in re.finditer(r'\{c:"([^"]+)",n:"([^"]+)"(?:,l:\[([^\]]*)\])?\}', raw):
        code, name, links_raw = m.group(1), m.group(2), m.group(3) or ""
        links = re.findall(r'"([^"]+)"', links_raw)
        entries.append({"code": code, "name": name, "links": links})

    by_code = {e["code"].upper(): e for e in entries}
    by_norm = {}
    for e in entries:
        name = e["name"]
        by_norm[_norm_port_name(name)] = e["code"]
        by_norm[_norm_port_name(re.sub(r",\s*[A-Z]{2},\s*[A-Z0-9]+$", "", name))] = e["code"]
        by_norm[_norm_port_name(name.split(",")[0])] = e["code"]

    EMC_PORT_CACHE = {"entries": entries, "by_code": by_code, "by_norm": by_norm}
    return EMC_PORT_CACHE

def resolve_emc_port_code(port_name):
    key = _norm_port_name(port_name)
    if key in EMC_PORT_ALIASES:
        return EMC_PORT_ALIASES[key]

    ports = load_emc_ports()
    if key in ports["by_norm"]:
        return ports["by_norm"][key]

    # Fallback partial match, ưu tiên tên bắt đầu bằng query.
    matches = []
    for e in ports["entries"]:
        n = _norm_port_name(e["name"])
        if n.startswith(key) or key in n:
            matches.append(e)
    if matches:
        all_port = next((e for e in matches if "(ALL PORTS)" in e["name"].upper()), None)
        return (all_port or matches[0])["code"]
    return ""

def install_greenx_capture_hook():
    hook = r"""
    (function(){
      window.__GX_CALLS__ = [];
      if (!window.__GX_HOOKED__) {
        window.__GX_HOOKED__ = true;
        const pushCall = (o) => { try { window.__GX_CALLS__.push(o); } catch(e) {} };
        const origFetch = window.fetch;
        if (origFetch) {
          window.fetch = async function(input, init) {
            const url = (typeof input === 'string') ? input : (input && input.url) || '';
            const body = init && init.body;
            const res = await origFetch.apply(this, arguments);
            try {
              const clone = res.clone();
              clone.text().then(txt => pushCall({
                kind:'fetch', url, body:String(body || ''), status:res.status,
                response:txt.slice(0, 500000)
              }));
            } catch(e) {}
            return res;
          };
        }
        const origOpen = XMLHttpRequest.prototype.open;
        const origSend = XMLHttpRequest.prototype.send;
        XMLHttpRequest.prototype.open = function(method, url) {
          this.__gx_method=method; this.__gx_url=url;
          return origOpen.apply(this, arguments);
        };
        XMLHttpRequest.prototype.send = function(body) {
          const xhr=this; const old=xhr.onreadystatechange;
          xhr.onreadystatechange=function(){
            if(xhr.readyState===4){
              pushCall({
                kind:'xhr', method:xhr.__gx_method, url:xhr.__gx_url,
                body:String(body||''), status:xhr.status,
                response:String(xhr.responseText||'').slice(0,500000)
              });
            }
            if(old) return old.apply(xhr, arguments);
          };
          return origSend.apply(this, arguments);
        };
      }
    })();
    """
    try:
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": hook})
    except Exception:
        pass
    driver.execute_script(hook)

def read_quotes2_from_browser(timeout=45):
    end = time.time() + timeout
    last_error = ""
    while time.time() < end:
        try:
            payloads = driver.execute_script("""
                return (window.__GX_CALLS__ || [])
                  .filter(c => String(c.url || '').includes('/graphql'))
                  .map(c => ({body:c.body, response:c.response, status:c.status}));
            """)
            for call in reversed(payloads or []):
                body = call.get("body") or ""
                if '"operationName":"quotes2"' not in body:
                    continue
                resp = json.loads(call.get("response") or "{}")
                quotes2 = (resp.get("data") or {}).get("quotes2")
                if quotes2 is not None:
                    return quotes2
        except Exception as e:
            last_error = clean_emc_error(e)
        time.sleep(0.5)
    raise TimeoutException(f"Khong bat duoc GraphQL quotes2: {last_error or 'timeout'}")

def fmt_emc_dt(ts):
    return datetime.fromtimestamp(int(ts))

def emc_quote_is_bookable(q):
    return (
        not q.get("rateNotAvailable")
        and not q.get("inventoryNotAvailable")
        and not q.get("inventorySoldOut")
        and bool(q.get("pricingInfo"))
        and bool(q.get("originPrepaidChargeItems"))
    )

def emc_quote_vessel(q):
    vessel = ((q.get("vessel") or {}).get("name") or "").strip()
    voyage = ((q.get("voyage") or {}).get("code") or "").strip()
    if not vessel:
        for leg in q.get("legs") or []:
            vessel = ((leg.get("vessel") or {}).get("name") or "").strip()
            voyage = ((leg.get("voyage") or {}).get("code") or "").strip()
            if vessel:
                break
    return clean_emc_vessel_name(f"{vessel} {voyage}".strip()) or "TBA"

def emc_quote_transshipment(q):
    legs = q.get("legs") or []
    pol_code = ((q.get("portOfLoading") or {}).get("code") or "").upper()
    pod_code = ((q.get("portOfDischarge") or {}).get("code") or "").upper()
    ts_ports = []
    for leg in legs:
        leg_type = str(leg.get("type") or "").upper()
        to_loc = leg.get("to") or {}
        to_code = str(to_loc.get("code") or "").upper()
        to_name = str(to_loc.get("name") or "").strip()
        if "TS" in leg_type and to_name:
            ts_ports.append(to_name.upper())
        elif to_name and to_code and to_code not in {pol_code, pod_code} and leg.get("vessel"):
            ts_ports.append(to_name.upper())
    return ts_ports[0] if ts_ports else "DIRECT"

def emc_quote_tt_days(q):
    tt = q.get("transitTime")
    if isinstance(tt, dict):
        start = tt.get("start")
        end = tt.get("end")
        vals = [v for v in [start, end] if isinstance(v, int)]
        if vals:
            return min(vals)
    return 999

def emc_quote_free_time(q):
    pod_ft = ((q.get("freeTime") or {}).get("podFreeTime") or {})
    usage = (pod_ft.get("containerUsage") or {}).get("days")
    dem = (pod_ft.get("demurrage") or {}).get("days")
    det = (pod_ft.get("detention") or {}).get("days")
    if usage:
        return f"{usage} DAYS"
    if dem and det:
        return f"{dem} DEM + {det} DET"
    if dem:
        return f"{dem} DEM"
    if det:
        return f"{det} DET"
    return ""

def emc_prices_from_quote(q, port_to):
    china_pod = is_china_pod(port_to)
    charges = {"20": 0.0, "40": 0.0, "40hq": 0.0}
    formulas = {"20": [], "40": [], "40hq": []}
    type_map = {"CT_20GP": "20", "CT_40GP": "40", "CT_40HQ": "40hq"}
    origin_thc_added = False
    booking_added = False

    for item in q.get("originPrepaidChargeItems") or []:
        charge_type = str(item.get("chargeType") or "").upper()
        charge_item = str(item.get("chargeItem") or "").upper()
        cont_key = type_map.get(item.get("containerType"))
        amt = float(item.get("priceInUsd") or 0)

        if charge_type == "BOOKING_FEE":
            for k in charges:
                charges[k] += amt
                formulas[k].append(amt)
            booking_added = True
            continue

        is_origin_thc = "THC/L" in charge_item or ("THC" in charge_item and "/D" not in charge_item)
        if is_origin_thc and not china_pod:
            continue
        if is_origin_thc and china_pod:
            origin_thc_added = True

        if cont_key and amt > 0:
            charges[cont_key] += amt
            formulas[cont_key].append(amt)

    if not booking_added:
        for k in charges:
            charges[k] += BOOKING_FEE
            formulas[k].append(float(BOOKING_FEE))

    totals = {k: math.ceil(v) if v > 0 else None for k, v in charges.items()}
    price_formulas = {}
    for k, parts in formulas.items():
        expr_parts = [str(int(p)) if float(p).is_integer() else str(p) for p in parts]
        price_formulas[k] = "=" + "+".join(expr_parts) if expr_parts else None

    not_app = [str(x).upper() for x in (q.get("notApplicableChargeItems") or [])]
    remark = build_subject_remark(
        othc_included=china_pod,
        pod=port_to,
        ows=("HWCS" in not_app and not china_pod),
    )
    return totals, remark, price_formulas

def search_one_api(port_from, port_to):
    base = {"from": port_from, "to": port_to,
            "etd": "N/A", "tt": "N/A", "valid": "",
            "price_20": None, "price_40": None, "price_40hq": None,
            "formula_20": None, "formula_40": None, "formula_40hq": None,
            "remark": "", "error": None,
            "free_time": "", "vessel_info": "", "transshipment": ""}
    try:
        from_code = resolve_emc_port_code(port_from)
        to_code = resolve_emc_port_code(port_to)
        if not from_code or not to_code:
            return {**base, "error": f"NO PORT CODE: {port_from}->{from_code}, {port_to}->{to_code}"}

        etd_date = (datetime.now() + timedelta(days=DATE_OFFSET_DAYS)).replace(hour=0, minute=0, second=0, microsecond=0)
        etd_ts = int(etd_date.timestamp())
        url = f"{BASE_URL}/results/{from_code}/{to_code}/1/1/1/{etd_ts}"
        print(f"   [EMC-API] {from_code} -> {to_code} | ETD_FROM={etd_date.strftime('%m/%d/%Y')}")

        ensure_emc_session()
        install_greenx_capture_hook()
        driver.get(url)
        WebDriverWait(driver, 45).until(
            lambda d: d.find_elements(By.CSS_SELECTOR, "li.quotes-search-result")
                   or "NO RESULT" in (d.find_element(By.TAG_NAME, "body").text or "").upper()
                   or (d.execute_script("return (window.__GX_CALLS__ || []).some(c => String(c.body || '').includes('\"operationName\":\"quotes2\"'))") is True)
        )
        quotes2 = read_quotes2_from_browser(timeout=25)
        quotes = quotes2.get("quotes") or []
        bookable = [q for q in quotes if emc_quote_is_bookable(q)]
        print(f"   [EMC-API] Quotes={len(quotes)} | Bookable={len(bookable)}")
        if not bookable:
            return {**base, "error": "NO SERVICE / SOLD OUT"}

        list_chuyen = []
        for q in bookable:
            etd_dt = fmt_emc_dt(q.get("etd"))
            list_chuyen.append({
                "quote": q,
                "etd_dt": etd_dt,
                "tt_days": emc_quote_tt_days(q),
                "vessel_name": emc_quote_vessel(q),
                "ts_port": emc_quote_transshipment(q),
            })
            print(
                f"      ✅ API Book: ETD={etd_dt.strftime('%m/%d/%Y')} "
                f"TT={list_chuyen[-1]['tt_days']}d Vessel={list_chuyen[-1]['vessel_name']} "
                f"TS={list_chuyen[-1]['ts_port']}"
            )

        etd_chuan, str_etd, str_tt = apply_9_golden_rules(list_chuyen)
        valid_text = get_valid_date([c["etd_dt"] for c in etd_chuan])
        representative = etd_chuan[0]["quote"]
        totals, remark, price_formulas = emc_prices_from_quote(representative, port_to)
        if not all(totals.get(k) for k in ("20", "40", "40hq")):
            return {**base, "error": "API co Book nhung khong doc du gia 20/40/40HQ"}

        free_time = emc_quote_free_time(representative)
        vessel_lines = []
        ts_set = []
        for c in etd_chuan:
            ts = c.get("ts_port") or "DIRECT"
            vessel_lines.append(
                f"{c.get('vessel_name') or 'TBA'} / ETD: {c['etd_dt'].day}-{c['etd_dt'].strftime('%b')}"
                f" / Transit time: {c.get('tt_days')} Days / Transshipment Port: {ts}"
            )
            if ts not in ts_set:
                ts_set.append(ts)
        ts_text = " or\n".join(ts_set) if ts_set else "DIRECT"

        print(f"\n   💰 KẾT QUẢ EMC API:")
        print(f"      ETD    : {str_etd}")
        print(f"      T/T    : {str_tt} days")
        print(f"      Valid  : {valid_text}")
        print(f"      20' GP : {price_formulas.get('20') or totals.get('20')}")
        print(f"      40' GP : {price_formulas.get('40') or totals.get('40')}")
        print(f"      40' HQ : {price_formulas.get('40hq') or totals.get('40hq')}")
        if free_time:
            print(f"      Free Time: {free_time}")
        print(f"      📝 Remark: {remark}")

        return {**base,
                "etd": str_etd, "tt": str_tt, "valid": valid_text,
                "price_20": totals.get("20"), "price_40": totals.get("40"), "price_40hq": totals.get("40hq"),
                "formula_20": price_formulas.get("20"), "formula_40": price_formulas.get("40"), "formula_40hq": price_formulas.get("40hq"),
                "remark": remark, "free_time": free_time,
                "vessel_info": "\n".join(vessel_lines), "transshipment": ts_text}
    except Exception as e:
        return {**base, "error": f"EMC API ERROR: {clean_emc_error(e)}"}

def should_skip(charge_name):
    name_upper = charge_name.upper()
    return any(b in name_upper for b in BLOCKLIST)

def parse_price_details(subrow_el, port_to):
    charges = {"20": 0.0, "40": 0.0, "40hq": 0.0}
    formulas = {"20": [], "40": [], "40hq": []}
    remark = ""
    china_pod = is_china_pod(port_to)
    origin_thc_added = False

    CONT_KEY = {
        "20' standard dry": "20",
        "40' standard dry": "40",
        "40' high cube":    "40hq",
    }

    SKIP_KEYWORDS = ["destination", "d/o", "delivery order"]

    try:
        origin_section = subrow_el.find_elements(
            By.CSS_SELECTOR,
            ".quotes-search-results-subrow-price-details-content"
        )
        origin_table = origin_section[0] if origin_section else subrow_el
        rows = origin_table.find_elements(By.CSS_SELECTOR, "table tbody tr")
        if not rows:
            print("      ⚠️ Không tìm thấy section giá")
            return {"20": None, "40": None, "40hq": None}, "", {"20": None, "40": None, "40hq": None}
        booking_fee_added = False

        for row in rows:
            tds = row.find_elements(By.CSS_SELECTOR, "td")
            if len(tds) < 4:
                continue

            charge_name = tds[0].find_element(By.CSS_SELECTOR, "span").text.strip() if tds[0].find_elements(By.CSS_SELECTOR, "span") else ""
            cont_raw    = tds[1].find_element(By.CSS_SELECTOR, "span").text.strip().lower() if tds[1].find_elements(By.CSS_SELECTOR, "span") else ""
            price_spans = tds[3].find_elements(By.CSS_SELECTOR, "span")
            price_text  = price_spans[0].text.strip() if price_spans else ""

            if not charge_name:
                continue

            if "booking fee" in charge_name.lower() and not booking_fee_added:
                fee = parse_usd(price_text) or BOOKING_FEE
                for k in charges:
                    charges[k] += fee
                    formulas[k].append(fee)
                booking_fee_added = True
                print(f"      [+] Booking Fee: ${fee}")
                continue

            charge_lower = charge_name.lower()
            is_origin_thc = (
                "terminal handling" in charge_lower
                or "thc/l" in charge_lower
                or "port of loading" in charge_lower
            ) and "destination" not in charge_lower

            if any(kw in charge_lower for kw in SKIP_KEYWORDS):
                print(f"      [-] Bỏ qua: {charge_name}")
                continue

            if ("terminal handling" in charge_lower or "thc" in charge_lower) and not (china_pod and is_origin_thc):
                print(f"      [-] Bỏ qua: {charge_name}")
                continue

            cont_key = CONT_KEY.get(cont_raw)
            if cont_key:
                amt = parse_usd(price_text)
                charges[cont_key] += amt
                formulas[cont_key].append(amt)
                if china_pod and is_origin_thc and amt > 0:
                    origin_thc_added = True
                print(f"      [+] {charge_name} ({cont_raw}): ${amt}")

        if not booking_fee_added:
            for k in charges:
                charges[k] += BOOKING_FEE
                formulas[k].append(float(BOOKING_FEE))
            print(f"      [+] Booking Fee (default): ${BOOKING_FEE}")

        full_text = subrow_el.text.upper()
        ows_subject = "NON APPLICABLE" in full_text and "HWCS" in full_text
        remark = build_subject_remark(
            othc_included=china_pod,
            pod=port_to,
            ows=(ows_subject and not china_pod),
        )

    except Exception as e:
        print(f"      ⚠️ Lỗi parse Price Details: {clean_emc_error(e)}")

    totals = {k: math.ceil(v) if v > 0 else None for k, v in charges.items()}
    price_formulas = {}
    for k, parts in formulas.items():
        if not parts:
            price_formulas[k] = None
            continue
        expr_parts = []
        for part in parts:
            expr_parts.append(str(int(part)) if float(part).is_integer() else str(part))
        price_formulas[k] = "=" + "+".join(expr_parts)
    return totals, remark, price_formulas


# ===================================================================================
# --- SEARCH 1 CẶP CẢNG ---
# ===================================================================================
def search_one_selenium(port_from, port_to):
    print(f"\n{'='*55}")
    print(f"  🔍 {port_from.upper()} → {port_to.upper()}")
    print(f"{'='*55}")

    base = {"from": port_from, "to": port_to,
            "etd": "N/A", "tt": "N/A", "valid": "",
            "price_20": None, "price_40": None, "price_40hq": None,
            "formula_20": None, "formula_40": None, "formula_40hq": None,
            "remark": "", "error": None,
            "free_time": "", "vessel_info": "", "transshipment": ""}
    try:
        # ── Nhập FROM (try nhiều XPath fallback) ──
        # FIX: EMC đổi UI — thêm fallback theo placeholder + position thứ tự
        print(f"   📍 Nhập FROM: {port_from}")
        from_xpaths = [
            '//input[@aria-label="input for From"]',
            '//input[@aria-label="input for Origin"]',
            '(//input[contains(@placeholder,"From") or contains(@placeholder,"Origin") or contains(@placeholder,"from")])[1]',
            '(//input[@type="text" and not(@disabled)])[1]',
        ]
        _ok = False
        for xp in from_xpaths:
            try:
                select_port_emc(xp, port_from)
                _ok = True
                break
            except Exception as _e:
                print(f"   🔄 FROM xpath '{xp[:60]}...' không ăn, thử fallback...")
        if not _ok:
            raise Exception(f"Không bắt được FROM input cho {port_from}")

        # ── Nhập TO (fallback cũng nhiều XPath) ──
        print(f"   📍 Nhập TO: {port_to}")
        to_xpaths = [
            '//input[@aria-label="input for To"]',
            '//input[@aria-label="input for Destination"]',
            '(//input[contains(@placeholder,"To") or contains(@placeholder,"Destination") or contains(@placeholder,"to")])[1]',
            '(//input[@type="text" and not(@disabled)])[2]',
        ]
        _ok = False
        for xp in to_xpaths:
            try:
                select_port_emc(xp, port_to)
                _ok = True
                break
            except Exception:
                print(f"   🔄 TO xpath '{xp[:60]}...' không ăn, thử fallback...")
        if not _ok:
            raise Exception(f"Không bắt được TO input cho {port_to}")

        # ── Chọn số lượng cont = 1 ──
        print(f"   📦 Kiểm tra số lượng cont...")
        for label in ["20' GP", "40' GP", "40' HQ"]:
            ok = False
            last_error = ""
            for attempt in range(3):
                try:
                    inp_cont = WebDriverWait(driver, 6).until(
                        EC.presence_of_element_located((By.XPATH, f'//input[@aria-label="input for {label}"]'))
                    )
                    driver.execute_script("""
                        const inp = arguments[0];
                        const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                        setter.call(inp, '1');
                        for (const ev of ['input', 'change', 'blur']) {
                            inp.dispatchEvent(new Event(ev, {bubbles: true}));
                        }
                    """, inp_cont)
                    time.sleep(0.25)
                    if inp_cont.get_attribute("value") == "1":
                        ok = True
                        print(f"      -> {label} = 1 ✅")
                        break
                    dropdown = inp_cont.find_element(By.XPATH, "ancestor::div[contains(@class,'dropdown-wrapper')][1]")
                    click_target = dropdown.find_element(By.CSS_SELECTOR, ".bx-dropdown-title-bar")
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", click_target)
                    time.sleep(0.15)
                    driver.execute_script("arguments[0].click();", click_target)
                    WebDriverWait(driver, 6).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "ul.open li[data-arrow-control]")))
                    time.sleep(0.2)
                    for li in driver.find_elements(By.CSS_SELECTOR, "ul.open li[data-arrow-control]"):
                        if li.find_element(By.CSS_SELECTOR, "span.item-text").text.strip() == "1":
                            driver.execute_script("arguments[0].click();", li)
                            time.sleep(0.25)
                            break
                    if inp_cont.get_attribute("value") != "1":
                        driver.execute_script("""
                            const inp = arguments[0];
                            const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                            setter.call(inp, '1');
                            for (const ev of ['input', 'change', 'blur']) {
                                inp.dispatchEvent(new Event(ev, {bubbles: true}));
                            }
                        """, inp_cont)
                        time.sleep(0.25)
                    if inp_cont.get_attribute("value") == "1":
                        ok = True
                        print(f"      -> {label} = 1 ✅")
                        break
                except Exception as e:
                    last_error = str(e).splitlines()[0] if str(e) else repr(e)
                    time.sleep(0.5)
            if not ok:
                raise Exception(f"Không set được container {label}=1: {last_error}")

        # ── Chọn ngày ETD = hôm nay + 7 ──
        print(f"   📅 Kiểm tra ngày ETD...")
        etd_date = datetime.now() + timedelta(days=DATE_OFFSET_DAYS)
        etd_target = etd_date.strftime("%m/%d/%Y")
        try:
            date_input = driver.find_element(By.CSS_SELECTOR, "input.mx-input")
            current_date = date_input.get_attribute("value")
            if current_date == etd_target:
                print(f"      -> Ngày đã đúng: {etd_target}, bỏ qua ✅")
            else:
                ActionChains(driver).move_to_element(date_input).click().perform()
                time.sleep(0.6)
                clicked_date = False
                for fmt in [f"{etd_date.month}/{etd_date.day}/{etd_date.year}", etd_target]:
                    cells = [
                        c for c in driver.find_elements(By.XPATH, f'//td[@title="{fmt}"]')
                        if c.is_displayed() and c.size.get("width", 0) > 0 and c.size.get("height", 0) > 0
                    ]
                    if cells:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cells[0])
                        driver.execute_script("arguments[0].click();", cells[0])
                        print(f"      -> Đã chọn: {fmt} ✅")
                        time.sleep(0.5)
                        clicked_date = True
                        break
                if not clicked_date or date_input.get_attribute("value") != etd_target:
                    driver.execute_script("""
                        const inp = arguments[0], value = arguments[1];
                        const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                        setter.call(inp, value);
                        for (const ev of ['input', 'change', 'blur']) {
                            inp.dispatchEvent(new Event(ev, {bubbles: true}));
                        }
                    """, date_input, etd_target)
                    time.sleep(0.5)
                if date_input.get_attribute("value") != etd_target:
                    raise Exception(f"Không set được ngày ETD {etd_target}, hiện là {date_input.get_attribute('value')}")
        except Exception as e:
            raise Exception(f"Lỗi chọn ngày: {e}")

        # ── Bấm SEARCH ──
        search_btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,
            "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'search')]")))
        driver.execute_script("arguments[0].click();", search_btn)
        print(f"   🖱️ Đã bấm SEARCH, chờ kết quả...")

        # ── Chờ cards load ──
        # FIX: chờ số lượng cards ổn định thay vì sleep cứng
        deadline = time.time() + 45
        while time.time() < deadline:
            cards_now = driver.find_elements(By.CSS_SELECTOR, "li.quotes-search-result")
            if cards_now:
                break
            body_text = ""
            try:
                body_text = driver.find_element(By.TAG_NAME, "body").text.upper()
            except Exception:
                pass
            if any(x in body_text for x in ["NO RESULTS", "NO SAILING", "NO SERVICE", "SOLD OUT"]):
                return {**base, "error": "NO SERVICE / SOLD OUT"}
            time.sleep(0.5)
        else:
            raise TimeoutException("EMC results did not render within 45s")
        # Chờ card count dừng tăng (page render xong)
        prev_count = -1
        for _ in range(8):  # tối đa ~2.4s
            try:
                cnt = len(driver.find_elements(By.CSS_SELECTOR, "li.quotes-search-result"))
                if cnt > 0 and cnt == prev_count:
                    break
                prev_count = cnt
            except Exception:
                pass
            time.sleep(0.3)

    except Exception as e:
        print(f"   ❌ Lỗi: {clean_emc_error(e)}")
        return {**base, "error": str(e)}

    # ── Đọc cards, lọc chỉ lấy card có nút "Book" ──
    print(f"   📋 Đang đọc danh sách kết quả...")
    book_wait_deadline = time.time() + 8
    while time.time() < book_wait_deadline:
        try:
            if any(card_has_book(c) for c in driver.find_elements(By.CSS_SELECTOR, "li.quotes-search-result")):
                break
        except Exception:
            pass
        time.sleep(0.5)

    list_chuyen = []
    idx = 0
    while True:
        cards = driver.find_elements(By.CSS_SELECTOR, "li.quotes-search-result")
        if idx >= len(cards):
            break
        card = cards[idx]
        idx += 1
        try:
            if not card_has_book(card):
                print(f"      ⏭️ Bỏ qua card (không có Book)")
                continue

            etd_str = card.find_element(By.CSS_SELECTOR, ".estimated-dates-left.content span.date").text.strip()
            eta_str = card.find_element(By.CSS_SELECTOR, ".estimated-dates-right.content span.date").text.strip()
            tt_text = card.find_element(By.CSS_SELECTOR, ".estimated-dates-elapsed span").text.strip()
            tt_days = int(re.search(r'\d+', tt_text).group()) if re.search(r'\d+', tt_text) else 999

            etd_dt = datetime.strptime(etd_str, "%m/%d/%Y")

            # Extract vessel name + transshipment (best-effort, defensive)
            vessel_name = extract_vessel_from_card(card)
            ts_port     = extract_transshipment_from_card(card)

            print(f"      ✅ Card hợp lệ: ETD={etd_str} ETA={eta_str} TT={tt_days}d Vessel={vessel_name} TS={ts_port}")
            list_chuyen.append({
                "element":     card,
                "etd_dt":      etd_dt,
                "tt_days":     tt_days,
                "vessel_name": vessel_name,
                "ts_port":     ts_port,
            })

        except Exception as e:
            print(f"      ⚠️ Lỗi đọc card: {clean_emc_error(e)}")
            continue

    if not list_chuyen:
        print(f"   ⚠️ Không có card nào có nút Book (có thể SOLD OUT / JOIN WAITLIST hết)")
        return {**base, "error": "NO SERVICE / SOLD OUT"}

    # ── Áp dụng 9 quy tắc vàng ──
    etd_chuan, str_etd, str_tt = apply_9_golden_rules(list_chuyen)
    valid_text = get_valid_date([c["etd_dt"] for c in etd_chuan])
    print(f"\n   🏆 ETD chọn: {str_etd} | T/T: {str_tt} days")

    # ── Route Details: lấy vessel đầu tiên và T/S PORT thực tế ──
    for c in etd_chuan:
        v_from_route, ts_from_route = read_emc_route_details(c.get("element"))
        if v_from_route:
            c["vessel_name"] = v_from_route
        if ts_from_route:
            c["ts_port"] = ts_from_route

    # ── Build vessel_info + transshipment string từ 3 chuyến đã chọn ──
    vessel_lines = []
    ts_set = []
    for c in etd_chuan:
        v_name = clean_emc_vessel_name(c.get("vessel_name") or "TBA")
        ts     = c.get("ts_port")     or "DIRECT"
        etd_d  = c["etd_dt"]
        td     = c.get("tt_days")
        vessel_lines.append(
            f"{v_name} / ETD: {etd_d.day}-{etd_d.strftime('%b')}"
            f" / Transit time: {td} Days / Transshipment Port: {ts}"
        )
        if ts not in ts_set:
            ts_set.append(ts)
    vessel_info  = "\n".join(vessel_lines)
    ts_text      = " or\n".join(ts_set) if ts_set else "DIRECT"

    # Re-fetch cards để tránh stale element
    time.sleep(1)
    fresh_cards = driver.find_elements(By.CSS_SELECTOR, "li.quotes-search-result")
    etd_chuan_dt = etd_chuan[0]["etd_dt"]

    dai_dien_el = None
    fallback_book_el = None
    used_fallback_book = False
    for fc in fresh_cards:
        try:
            has_book = card_has_book(fc)
            if has_book and fallback_book_el is None:
                fallback_book_el = fc

            etd_str_fc = fc.find_element(By.CSS_SELECTOR, ".estimated-dates-left.content span.date").text.strip()
            if datetime.strptime(etd_str_fc, "%m/%d/%Y") == etd_chuan_dt and has_book:
                dai_dien_el = fc
                break
        except Exception:
            continue

    if not dai_dien_el:
        dai_dien_el = fallback_book_el
        if dai_dien_el:
            used_fallback_book = True
            print("      [WARN] Khong tim lai dung ETD sau re-fetch, dung card Book dau tien de lay Price Details.")

    if not dai_dien_el:
        return {**base, "error": "Khong tim lai duoc card co gia sau re-fetch"}

    if used_fallback_book:
        try:
            etd_str_fb = dai_dien_el.find_element(By.CSS_SELECTOR, ".estimated-dates-left.content span.date").text.strip()
            tt_text_fb = dai_dien_el.find_element(By.CSS_SELECTOR, ".estimated-dates-elapsed span").text.strip()
            etd_dt_fb = datetime.strptime(etd_str_fb, "%m/%d/%Y")
            if not etd_within_max(etd_dt_fb):
                return {**base, "error": "Fallback ETD vuot qua 21 ngay"}
            tt_days_fb = int(re.search(r'\d+', tt_text_fb).group()) if re.search(r'\d+', tt_text_fb) else 999
            v_fb = extract_vessel_from_card(dai_dien_el)
            ts_fb = extract_transshipment_from_card(dai_dien_el)

            v_route_fb, ts_route_fb = read_emc_route_details(dai_dien_el)
            if v_route_fb:
                v_fb = v_route_fb
            if ts_route_fb:
                ts_fb = ts_route_fb

            str_etd = f"{etd_dt_fb.day}-{etd_dt_fb.strftime('%b')}"
            str_tt = str(tt_days_fb)
            valid_text = get_valid_date([etd_dt_fb])
            vessel_info = (
                f"{clean_emc_vessel_name(v_fb or 'TBA')} / ETD: {str_etd}"
                f" / Transit time: {tt_days_fb} Days / Transshipment Port: {ts_fb or 'DIRECT'}"
            )
            ts_text = ts_fb or "DIRECT"
            print(f"      [INFO] Fallback Book card: ETD={str_etd} TT={str_tt} Vessel={v_fb} TS={ts_text}")
        except Exception as e:
            return {**base, "error": f"Khong doc duoc lich cua card co Book: {clean_emc_error(e)}"}

    # ── Bấm Price Details trên card đại diện ──
    try:
        li_el = dai_dien_el

        price_details_tab = dai_dien_el.find_element(By.XPATH,
            ".//*[contains(@class,'quotes-search-results-list-item-tabs-tab') or self::button or @role='tab']"
            "[.//span[normalize-space()='Price Details'] or normalize-space()='Price Details']")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", price_details_tab)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", price_details_tab)
        print(f"   🖱️ Đã bấm Price Details, chờ bảng giá...")

        subrow = li_el.find_element(By.CSS_SELECTOR, ".quotes-search-results-list-item-subrow")
        WebDriverWait(driver, 10).until(
            lambda d: subrow.find_elements(By.CSS_SELECTOR,
                ".quotes-search-results-subrow-price-details-content table tbody tr")
                or subrow.find_elements(By.CSS_SELECTOR, "table tbody tr")
        )
        time.sleep(0.3)

    except Exception as e:
        return {**base, "error": f"Không bấm được Price Details: {e}"}

    # ── Parse bảng giá từ subrow ──
    try:
        print(f"\n   💰 Đang parse bảng giá...")
        totals, remark, price_formulas = parse_price_details(subrow, port_to)
    except Exception as e:
        return {**base, "error": f"Lỗi parse giá: {e}"}

    if not all(totals.get(k) for k in ("20", "40", "40hq")):
        return {**base, "error": "Card co Book nhung khong doc du gia 20/40/40HQ"}

    free_time = scrape_free_time_from_emc_card(dai_dien_el, subrow)
    if free_time:
        print(f"      ⏱️ Free Time: {free_time}")
    else:
        print("      ⚠️ Chưa đọc được Free Time")

    print(f"\n   💰 KẾT QUẢ SAU LỌC:")
    print(f"      ETD    : {str_etd}")
    print(f"      T/T    : {str_tt} days")
    print(f"      Valid  : {valid_text}")
    print(f"      20' GP : {'$' + str(totals['20'])   if totals['20']   else 'N/A'}")
    print(f"      40' GP : {'$' + str(totals['40'])   if totals['40']   else 'N/A'}")
    print(f"      40' HQ : {'$' + str(totals['40hq']) if totals['40hq'] else 'N/A'}")
    if remark:
        print(f"      📝 Remark: {remark}")

    return {**base,
            "etd": str_etd, "tt": str_tt,
            "valid": valid_text,
            "price_20":   totals.get("20"),
            "price_40":   totals.get("40"),
            "price_40hq": totals.get("40hq"),
            "formula_20":   price_formulas.get("20"),
            "formula_40":   price_formulas.get("40"),
            "formula_40hq": price_formulas.get("40hq"),
            "remark":     remark,
            "free_time":   free_time,
            "vessel_info":   vessel_info,
            "transshipment": ts_text}


def search_one(port_from, port_to):
    """API-only: nếu API lỗi thì trả lỗi, KHÔNG fallback qua Selenium."""
    api_result = search_one_api(port_from, port_to)
    if api_result.get("error"):
        print(f"   [EMC-API2] ❌ {api_result.get('error')} → bỏ qua Selenium fallback.")
    return api_result


GREENX_BASE = "https://www.greenxtrade.com"
GREENX_SIGNIN_URL = f"{GREENX_BASE}/_gx/GREENX_SignIn"
GREENX_HOME_URL = f"{GREENX_BASE}/_gx/GREENX_Home"
GREENX_LOGIN_ID = os.environ.get("EMC_LOGIN_ID", "celine@pio-logistics.vn")
GREENX_LOGIN_PASS = os.environ.get("EMC_LOGIN_PASS", "Xvnt20277PioLog")
GREENX_BOOKABLE_STATUS = {"B", "P"}
_GREENX_SESSION = None
_GREENX_TABKEY = None

GREENX_PORT_ALIASES = {
    "HAI PHONG": {
        "zipCode": "35000",
        "locCode": "VNHPG",
        "locName": "HAIPHONG, VIETNAM",
        "fullName": "HAIPHONG, VIETNAM (VNHPH) [ZIP:35000]",
        "stdcCode": "VNHPH",
    },
    "HAIPHONG": {
        "zipCode": "35000",
        "locCode": "VNHPG",
        "locName": "HAIPHONG, VIETNAM",
        "fullName": "HAIPHONG, VIETNAM (VNHPH) [ZIP:35000]",
        "stdcCode": "VNHPH",
    },
}


def greenx_api_reset():
    global _GREENX_SESSION, _GREENX_TABKEY
    _GREENX_SESSION = None
    _GREENX_TABKEY = None


def greenx_api_login(force=False):
    global _GREENX_SESSION, _GREENX_TABKEY
    if _GREENX_SESSION is not None and _GREENX_TABKEY and not force:
        return _GREENX_SESSION, _GREENX_TABKEY

    sess = requests.Session()
    sess.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36 Edg/124.0"
        ),
        "Accept": "application/json, text/javascript, */*; q=0.01",
    })
    sess.get(GREENX_SIGNIN_URL, timeout=30)
    login_resp = sess.post(
        f"{GREENX_BASE}/_gx/ActionDispatcher?xctl=checkpwd&tabkey=null",
        data={"id": GREENX_LOGIN_ID, "pwd": GREENX_LOGIN_PASS},
        headers={
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Referer": GREENX_SIGNIN_URL,
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=30,
    )
    try:
        login_data = login_resp.json()
    except Exception:
        login_data = {}
    if not login_data.get("success"):
        msg = login_data.get("message") or login_data.get("error") or login_resp.text[:120]
        raise Exception(f"GREENX LOGIN FAILED: {msg}")

    home = sess.get(GREENX_HOME_URL, timeout=30).text
    m = re.search(r"GREENX_Quotes\?tabkey=([A-Fa-f0-9]+)", home)
    if not m:
        raise Exception("GREENX LOGIN OK NHUNG KHONG LAY DUOC TABKEY")

    _GREENX_SESSION = sess
    _GREENX_TABKEY = m.group(1)
    return _GREENX_SESSION, _GREENX_TABKEY


def greenx_post(action, payload, referer_action="GREENX_Quotes"):
    last_err = ""
    for attempt in range(2):
        sess, tabkey = greenx_api_login(force=(attempt > 0))
        url = f"{GREENX_BASE}/_gx/{action}?tabkey={tabkey}"
        referer = f"{GREENX_BASE}/_gx/{referer_action}?tabkey={tabkey}"
        resp = sess.post(
            url,
            json=payload,
            headers={
                "Content-Type": "application/json",
                "Referer": referer,
                "X-Requested-With": "XMLHttpRequest",
            },
            timeout=120,
        )
        text = resp.text or ""
        if resp.status_code in (299, 401, 403) or text.lstrip().startswith("<"):
            last_err = f"HTTP {resp.status_code}"
            greenx_api_reset()
            continue
        try:
            return resp.json()
        except Exception as e:
            last_err = f"JSON ERROR: {clean_emc_error(e)}"
            greenx_api_reset()
    raise Exception(last_err or f"{action} FAILED")


def greenx_resolve_port(port_name):
    query = str(port_name or "").strip()
    if not query:
        return None
    q_norm = _norm_port_name(query)
    if q_norm in GREENX_PORT_ALIASES:
        return dict(GREENX_PORT_ALIASES[q_norm])

    query_candidates = [query]
    compact_query = re.sub(r"\s+", "", query)
    if compact_query and compact_query.upper() != query.upper():
        query_candidates.append(compact_query)
    query_candidates = list(dict.fromkeys(query_candidates))

    options = []
    for q in query_candidates:
        data = greenx_post("GREENX_GetLocList", {"loc": q}, referer_action="GREENX_Quotes")
        options = data.get("data") or []
        if data.get("success") and options:
            break
    if not options:
        return None

    q_norms = [_norm_port_name(q) for q in query_candidates]

    def score(opt):
        loc_name = _norm_port_name(opt.get("locName"))
        full_name = _norm_port_name(opt.get("fullName"))
        stdc = _norm_port_name(opt.get("stdcCode"))
        if loc_name in q_norms:
            return 0
        if any(full_name.startswith(q) for q in q_norms):
            return 1
        if any(q in full_name for q in q_norms):
            return 2
        if any(q in loc_name for q in q_norms):
            return 3
        if stdc in q_norms:
            return 4
        return 9

    return sorted(options, key=score)[0]


def greenx_parse_ymd(value):
    return datetime.strptime(str(value), "%Y%m%d")


def greenx_charge_amount(charge):
    for key in ("chrgPrice", "chrgRate", "locPrice"):
        val = charge.get(key)
        if val in (None, ""):
            continue
        try:
            return float(str(val).replace(",", ""))
        except Exception:
            continue
    return 0.0


def greenx_charge_usd_amount(charge):
    cur = str(charge.get("chrgCur") or "USD").upper()
    if cur == "USD":
        return greenx_charge_amount(charge)
    val = charge.get("usdPrice")
    if val in (None, ""):
        return 0.0
    try:
        return float(str(val).replace(",", ""))
    except Exception:
        return 0.0


def greenx_formula_num(value):
    value = float(value)
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def greenx_unit_key(unit):
    return {"2SD": "20", "4SD": "40", "4SH": "40hq"}.get(str(unit or "").upper().strip())


def greenx_is_origin_thc(charge):
    name = str(charge.get("chrgName") or "").upper()
    item = str(charge.get("chrgItem") or "").upper()
    return item == "THC/L" or "PORT OF LOADING" in name or "THC/L" in name


def greenx_prices_from_quote(q, port_to):
    contract = q.get("contract") or {}
    china_pod = is_china_pod(port_to)
    charges = {"20": 0.0, "40": 0.0, "40hq": 0.0}
    formulas = {"20": [], "40": [], "40hq": []}
    origin_thc_added = False
    origin_thc_subject = False

    for item in contract.get("of") or []:
        key = greenx_unit_key(item.get("chrgRevUnit"))
        amt = greenx_charge_amount(item)
        if key and amt > 0:
            charges[key] += amt
            formulas[key].append(amt)

    prepaid = contract.get("prepaid")
    if prepaid is None:
        prepaid = [
            ch for ch in (contract.get("ex") or [])
            if str(ch.get("chrgFtb") or "").upper() == "P"
        ]

    for item in prepaid or []:
        is_thc = greenx_is_origin_thc(item)
        if is_thc and not china_pod:
            origin_thc_subject = True
            continue

        item_code = str(item.get("chrgItem") or "").upper()
        item_name = str(item.get("chrgName") or "").upper()
        is_euis = item_code == "EUIS" or "EU INNOVATION SURCHARGE" in item_name
        amt = greenx_charge_usd_amount(item) if (is_euis or is_thc) else greenx_charge_amount(item)
        cur = str(item.get("chrgCur") or "USD").upper()
        if cur != "USD" and not (is_euis or is_thc):
            continue
        if amt <= 0:
            continue
        if is_thc and china_pod:
            origin_thc_added = True

        key = greenx_unit_key(item.get("chrgRevUnit"))
        unit_desc = str(item.get("chrgRevUnitDesc") or "").upper()
        if key:
            charges[key] += amt
            formulas[key].append(amt)
        elif "B/L" in str(item.get("chrgRevUnit") or "").upper() or "B/L" in unit_desc or "PER B" in unit_desc:
            for k in charges:
                charges[k] += amt
                formulas[k].append(amt)
        else:
            for k in charges:
                charges[k] += amt
                formulas[k].append(amt)

    included = contract.get("in") or []
    origin_thc_included = any(greenx_is_origin_thc(ch) for ch in included)
    na_items = {str(ch.get("chrgItem") or "").upper() for ch in (contract.get("na") or [])}

    totals = {k: math.ceil(v) if v > 0 else None for k, v in charges.items()}
    price_formulas = {
        k: ("=" + "+".join(greenx_formula_num(p) for p in parts)) if parts else None
        for k, parts in formulas.items()
    }

    remark = build_subject_remark(
        othc_included=(china_pod or origin_thc_included or origin_thc_added),
        pod=port_to,
        ows=("HWCS" in na_items and not china_pod),
    )
    return totals, remark, price_formulas


def greenx_quote_is_bookable(q):
    status = str((q.get("inventory") or {}).get("status") or "").upper()
    contract = q.get("contract") or {}
    return status in GREENX_BOOKABLE_STATUS and bool(contract.get("of"))


def greenx_quote_tt_days(q):
    try:
        return int(q.get("tsDay"))
    except Exception:
        try:
            legs = q.get("legInfo") or []
            start = greenx_parse_ymd(legs[0]["rtemp2Depdate"])
            end = greenx_parse_ymd(legs[-1]["rtemp2Arrdate"])
            return max(1, (end - start).days)
        except Exception:
            return 999


def greenx_first_vessel(q):
    for leg in q.get("legInfo") or []:
        vsl_name = str(leg.get("rtemp2VslName") or "").strip()
        voy = str(leg.get("rtemp2Voy") or "").strip()
        vslvoy = str(leg.get("rtemp2Vslvoy") or "").strip()
        if vsl_name and vslvoy and vslvoy != "-":
            return clean_emc_vessel_name(f"{vsl_name} {voy}".strip())
    for txt in q.get("vslvoy") or []:
        txt = clean_emc_vessel_name(txt)
        if txt and txt != "-":
            return txt
    return "TBA"


def greenx_transshipment(q, pod_code=""):
    try:
        if int(q.get("tsLeg") or 1) <= 1:
            return "DIRECT"
    except Exception:
        return "DIRECT"
    pod_code = str(pod_code or "").upper()
    for leg in q.get("legInfo") or []:
        port2 = str(leg.get("rtemp2Port2") or "").upper()
        port2_name = str(leg.get("rtemp2Port2Name") or "").strip()
        if port2_name and port2 and port2 != pod_code:
            return port2_name.upper()
    main = q.get("mainLine") or {}
    port1_name = str(main.get("rtemp2Port1Name") or "").strip()
    return port1_name.upper() if port1_name else "DIRECT"


def greenx_quote_etd(q):
    legs = q.get("legInfo") or []
    if legs and legs[0].get("rtemp2Depdate"):
        return greenx_parse_ymd(legs[0]["rtemp2Depdate"])
    main = q.get("mainLine") or {}
    return greenx_parse_ymd(main.get("rtemp2Depdate"))


def greenx_quote_free_time(q):
    ft = q.get("freeTime") or {}
    dly = ft.get("dly") or []
    if not dly:
        return ""
    dem = det = usage = ""
    for row in dly:
        typ = str(row.get("dmdtType") or "").upper()
        days = str(row.get("dmdtFreedays") or "").strip()
        if not days or days == "-":
            continue
        if typ == "BT":
            usage = days
        elif typ == "DM":
            dem = days
        elif typ == "DT":
            det = days
    if usage:
        return f"{usage} DAYS"
    if dem and det:
        return f"{dem} DEM + {det} DET"
    if dem:
        return f"{dem} DEM"
    if det:
        return f"{det} DET"
    return ""


def greenx_hydrate_quote_summaries(quote_result, search_body):
    """Load price/inventory detail only for schedules inside the 21-day ETD cap.

    The current GreenX list endpoint deliberately returns inventory ``W`` and
    no contract for every card.  A second detail request is therefore required
    before treating a card as sold out or using its price.
    """
    quote_result = quote_result or {}
    summaries = quote_result.get("detail") or []
    max_etd = max_etd_date()
    candidates = [
        item for item in summaries
        if quote_departure_is_on_or_before(item, max_etd)
    ]
    hydrated = []
    failed = 0

    for item in candidates:
        payload = build_quote_detail_payload(quote_result, item, search_body)
        try:
            response = greenx_post(
                "GREENX_GetQuoteResultDetail",
                payload,
                referer_action="GREENX_QuoteResult",
            )
            detail = (response.get("data") or {}) if response.get("success") else {}
            if not detail:
                failed += 1
                continue
            hydrated.append(hydrate_quote_with_detail(item, detail))
        except Exception:
            failed += 1

    return hydrated, len(candidates), failed


def search_one_api(port_from, port_to):
    base = {"from": port_from, "to": port_to,
            "etd": "N/A", "tt": "N/A", "valid": "",
            "price_20": None, "price_40": None, "price_40hq": None,
            "formula_20": None, "formula_40": None, "formula_40hq": None,
            "remark": "", "error": None,
            "free_time": "", "vessel_info": "", "transshipment": ""}
    try:
        from_info = greenx_resolve_port(port_from)
        to_info = greenx_resolve_port(port_to)
        if not from_info or not to_info:
            return {**base, "error": f"NO PORT CODE: {port_from}->{from_info}, {port_to}->{to_info}"}

        from_code = from_info.get("locCode")
        to_code = to_info.get("locCode")
        etd_date = datetime.now() + timedelta(days=DATE_OFFSET_DAYS)
        body = {
            "rct": from_code,
            "dly": to_code,
            "cntr_20sd": "1",
            "cntr_40sd": "1",
            "cntr_40sh": "1",
            "etdDate": etd_date.strftime("%m/%d/%Y"),
        }
        print(f"   [EMC-API2] {from_code} -> {to_code} | ETD_FROM={body['etdDate']}")

        data = greenx_post("GREENX_GetQuoteResult", body, referer_action="GREENX_QuoteResult")
        quote_result = (data.get("data") or {}) if data.get("success") else {}
        summaries = quote_result.get("detail") or []
        details, detail_candidates, detail_failed = greenx_hydrate_quote_summaries(
            quote_result,
            body,
        )
        bookable = [q for q in details if greenx_quote_is_bookable(q)]
        print(
            f"   [EMC-API2] Quotes={len(summaries)} | Detail <= {max_etd_date_only()}="
            f"{detail_candidates} | Loaded={len(details)} | Book/Partial={len(bookable)}"
        )
        if detail_candidates and detail_failed == detail_candidates:
            return {**base, "error": "EMC API DETAIL ERROR: khong tai duoc chi tiet lich tau"}
        if not bookable:
            return {**base, "error": "NO SERVICE / SOLD OUT"}

        list_chuyen = []
        for q in bookable:
            etd_dt = greenx_quote_etd(q)
            tt_days = greenx_quote_tt_days(q)
            vessel_name = greenx_first_vessel(q)
            ts_port = greenx_transshipment(q, to_code)
            status = str((q.get("inventory") or {}).get("status") or "").upper()
            list_chuyen.append({
                "quote": q,
                "etd_dt": etd_dt,
                "tt_days": tt_days,
                "vessel_name": vessel_name,
                "ts_port": ts_port,
                "status": status,
            })
            print(
                f"      OK {status}: ETD={etd_dt.strftime('%m/%d/%Y')} "
                f"TT={tt_days}d Vessel={vessel_name} TS={ts_port}"
            )

        etd_chuan, str_etd, str_tt = apply_9_golden_rules(list_chuyen)
        valid_text = get_valid_date([c["etd_dt"] for c in etd_chuan])
        representative = etd_chuan[0]["quote"]
        totals, remark, price_formulas = greenx_prices_from_quote(representative, port_to)
        if not all(totals.get(k) for k in ("20", "40", "40hq")):
            return {**base, "error": "API co Book nhung khong doc du gia 20/40/40HQ"}

        free_time = greenx_quote_free_time(representative)
        vessel_lines = []
        ts_set = []
        for c in etd_chuan:
            ts = c.get("ts_port") or "DIRECT"
            vessel_lines.append(
                f"{c.get('vessel_name') or 'TBA'} / ETD: {c['etd_dt'].day}-{c['etd_dt'].strftime('%b')}"
                f" / Transit time: {c.get('tt_days')} Days / Transshipment Port: {ts}"
            )
            if ts not in ts_set:
                ts_set.append(ts)
        ts_text = " or\n".join(ts_set) if ts_set else "DIRECT"

        print("\n   KET QUA EMC API MOI:")
        print(f"      ETD    : {str_etd}")
        print(f"      T/T    : {str_tt} days")
        print(f"      Valid  : {valid_text}")
        print(f"      20' GP : {price_formulas.get('20') or totals.get('20')}")
        print(f"      40' GP : {price_formulas.get('40') or totals.get('40')}")
        print(f"      40' HQ : {price_formulas.get('40hq') or totals.get('40hq')}")
        if free_time:
            print(f"      Free Time: {free_time}")
        print(f"      Remark: {remark}")

        return {**base,
                "etd": str_etd, "tt": str_tt, "valid": valid_text,
                "price_20": totals.get("20"), "price_40": totals.get("40"), "price_40hq": totals.get("40hq"),
                "formula_20": price_formulas.get("20"), "formula_40": price_formulas.get("40"), "formula_40hq": price_formulas.get("40hq"),
                "remark": remark, "free_time": free_time,
                "vessel_info": "\n".join(vessel_lines), "transshipment": ts_text}
    except Exception as e:
        greenx_api_reset()
        return {**base, "error": f"EMC API ERROR: {clean_emc_error(e)}"}


# ===================================================================================
# --- MAIN ---
# ===================================================================================
print("""
╔══════════════════════════════════════════════╗
║        EMC Price Checker  🚢                 ║
╚══════════════════════════════════════════════╝
""")

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

print(f"\n🌐 Bắt đầu chạy...")

first_emc_row = None

# Mỗi phiên chạy mới login lại API mới để lấy tabkey sạch.
if EMC_USE_BROWSER:
    print(f"[HỆ THỐNG] Đang khởi tạo legacy browser EMC {BASE_URL} ...")
    try:
        ensure_emc_session()
        driver.get(BASE_URL)
        WebDriverWait(driver, 20).until(
            lambda d: d.find_elements(By.XPATH, '//input[@aria-label="input for From"]')
                   or d.find_elements(By.XPATH, '//input[@aria-label="input for Origin"]')
        )
        print("[HỆ THỐNG] ✅ Trang quotes legacy đã load xong.")
    except Exception as _e:
        print(f"[HỆ THỐNG] ⚠️ Không load được trang legacy EMC: {clean_emc_error(_e)}")
else:
    print(f"[HỆ THỐNG] Đang login GreenX API mới {GREENX_SIGNIN_URL} ...")
    try:
        greenx_api_login(force=True)
        print("[HỆ THỐNG] ✅ GreenX API đã sẵn sàng — bắt đầu vòng lặp.")
    except Exception as _e:
        print(f"[HỆ THỐNG] ❌ Không login được GreenX API: {clean_emc_error(_e)}")

# Port mapping: Excel name → EMC search name
EMC_PORT_MAPPING = {
    "TIANJIN": "XINGANG",
    "VENICE": "VENEZIA",
}

target_single_row = None
if SINGLE_ROW:
    try:
        target_single_row = int(SINGLE_ROW)
        print(f"[SINGLE_ROW] Chỉ chạy dòng {target_single_row} theo lệnh từ main.py")
    except Exception:
        print(f"[SINGLE_ROW] Không hợp lệ: {SINGLE_ROW}")

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if target_single_row is not None and i != target_single_row:
        continue

    country_excel = str(row[1] or "").strip()         # cột B = Country
    pol_excel = str(row[2] or "").strip().title()     # cột C = POL
    pod       = str(row[3] or "").strip().title()     # cột D = POD
    carrier   = str(row[4] or "").strip().upper()     # cột E = Carrier

    if not pol_excel or not pod: continue
    if carrier not in EMC_GROUP: continue
    if FILTER_POL and pol_excel.upper() != FILTER_POL: continue
    if FILTER_POD and pod.upper() != FILTER_POD: continue

    # Áp dụng port mapping cho search (giữ tên gốc trong Excel)
    pol_excel = EMC_PORT_MAPPING.get(pol_excel.upper(), pol_excel)
    pod = EMC_PORT_MAPPING.get(pod.upper(), pod)
    if first_emc_row is None:
        first_emc_row = i

    print(f"\n==========================================")
    print(f"--- DÒNG {i}: {pol_excel} → {pod} | Hãng: {carrier} ---")

    # FIX: Chỉ navigate lần đầu (lần sau dùng lại form đang hiện)
    if EMC_USE_BROWSER and i == first_emc_row:
        try:
            ensure_emc_session()
            WebDriverWait(driver, 10).until(
                lambda d: d.find_elements(By.XPATH, '//input[@aria-label="input for From"]')
                       or d.find_elements(By.XPATH, '//input[@aria-label="input for Origin"]')
            )
        except:
            ensure_emc_session()
            driver.get(BASE_URL)
            WebDriverWait(driver, 15).until(
                lambda d: d.find_elements(By.XPATH, '//input[@aria-label="input for From"]')
                       or d.find_elements(By.XPATH, '//input[@aria-label="input for Origin"]')
            )
        driver.switch_to.window(driver.current_window_handle)
        driver.execute_script("window.focus();")
        time.sleep(1)

    result = None
    for attempt in range(1, EMC_MAX_RETRIES + 1):
        if attempt > 1:
            print(f"   [RETRY] EMC API loi/lag, reload base URL va chay lai dong {i} lan {attempt}/{EMC_MAX_RETRIES}...")
            try:
                reload_emc_base(f"retry row {i} attempt {attempt}")
            except Exception as _reload_e:
                print(f"   [WARN] Reload EMC loi: {clean_emc_error(_reload_e)}")

        result = search_one(pol_excel, pod)
        if not result.get("error"):
            break
        result["error"] = clean_emc_error(result.get("error"))

    if result.get("error"):
        print(f"   [SKIP EXCEL] Dòng {i} lỗi: {clean_emc_error(result['error'])} -> để trống -output")
        for col in [6, 7, 8, 9, 10, 11, 13, 14, 15, 16]:
            ws.cell(row=i, column=col).value = None
    else:
        if is_emc_india_route(country_excel, pod):
            result = apply_emc_india_thc_rule(result, country_excel, pod)
            print(
                "   [EMC INDIA THC] Giá cuối: 20' -140 USD, "
                "40'/40HC -210 USD; remark SUBJECT TO THC."
            )
        ws.cell(row=i, column=6).value  = result.get("formula_20") or result["price_20"]
        ws.cell(row=i, column=7).value  = result.get("formula_40") or result["price_40"]
        ws.cell(row=i, column=8).value  = result.get("formula_40hq") or result["price_40hq"]
        ws.cell(row=i, column=9).value  = result["etd"]
        ws.cell(row=i, column=10).value = result["tt"]
        ws.cell(row=i, column=11).value = result.get("valid", "")
        ws.cell(row=i, column=13).value = result["remark"]
        ws.cell(row=i, column=14).value = result.get("free_time", "")
        ws.cell(row=i, column=15).value = result.get("vessel_info", "")
        ws.cell(row=i, column=16).value = result.get("transshipment", "")
        # Wrap text cho 2 cột multi-line
        try:
            wrap = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
            for col in [13, 15, 16]:
                ws.cell(row=i, column=col).alignment = wrap
        except Exception:
            pass

    try:
        wb.save(excel_path)
        print(f"   💾 Đã lưu dòng {i}")
    except PermissionError:
        print(f"   ❌ LỖI GHI FILE: TẮT FILE EXCEL ĐI!")

    if EMC_ROW_WAIT_SECONDS > 0:
        print(f"   ⏳ Nghỉ {EMC_ROW_WAIT_SECONDS:g}s trước dòng tiếp theo để web ổn định...")
        time.sleep(EMC_ROW_WAIT_SECONDS)

print(f"\n✅ Hoàn tất!")

