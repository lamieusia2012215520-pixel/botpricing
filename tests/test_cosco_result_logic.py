import unittest

from openpyxl import Workbook

from cosco_result_logic import clear_cosco_quote_fields


class CoscoResultLogicTests(unittest.TestCase):
    def test_no_price_clears_stale_rate_schedule_and_vessel_fields(self):
        sheet = Workbook().active
        row = 20
        for column in (6, 7, 8, 9, 10, 11, 15, 16):
            sheet.cell(row=row, column=column).value = f"old-{column}"
        sheet.cell(row=row, column=13).value = "old remark"
        sheet.cell(row=row, column=14).value = "14 COMBINED"

        clear_cosco_quote_fields(sheet, row)

        for column in (6, 7, 8, 9, 10, 11, 15, 16):
            self.assertIsNone(sheet.cell(row=row, column=column).value)
        self.assertEqual("old remark", sheet.cell(row=row, column=13).value)
        self.assertEqual("14 COMBINED", sheet.cell(row=row, column=14).value)


if __name__ == "__main__":
    unittest.main()
