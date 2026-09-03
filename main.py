"""
MAIN ORCHESTRATOR — Điều khiển toàn bộ bot check giá hãng tàu
=========================================================
Cách dùng:
  python main.py                            → chạy FULL Excel, toàn bộ bot theo 2 lượt
  python main.py --skip "YANG MING"         → chạy toàn bộ bot, trừ Yang Ming
  python main.py --skip YM --skip ZIM       → chạy toàn bộ bot, trừ Yang Ming và ZIM
  python main.py "HO CHI MINH" "CHENNAI"   → toàn bộ bot cùng check 1 tuyến
  python main.py 145                        → chỉ chạy bot của carrier ở dòng 145
  python main.py 100-150 --date +12         → chỉ chạy các dòng 100 đến 150

Cách hoạt động:
  - Mỗi bot chạy trong subprocess riêng
  - Mỗi bot ghi vào BẢN COPY Excel riêng (input_gia_CMA.xlsx, input_gia_HPL.xlsx, ...)
  - Chạy bot theo 2 lượt; sau mỗi lượt main.py gộp kết quả từ các bản copy vào file gốc
  - Bot nào xong thì đóng Edge ngay; riêng COSCO, HAPAG LLOYD và OOCL giữ browser
  - Tránh corrupt do nhiều process ghi cùng 1 file xlsx
"""

import sys
import os
import subprocess
import time
import shutil
import threading
import argparse
import atexit
import re
import unicodedata
import json
import html
from datetime import datetime

try:
    import msvcrt
except ImportError:
    msvcrt = None

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    print("[ERROR] Cần cài openpyxl: pip install openpyxl")
    sys.exit(1)

# Cấu hình tự động ghi đè hàm print để xuất log ra file log_main.txt hoàn toàn tự động
_orig_print = print
print_lock = threading.Lock()

def print(*args, **kwargs):
    with print_lock:
        _orig_print(*args, **kwargs)
        try:
            with open("log_main.txt", "a", encoding="utf-8") as f:
                sep = kwargs.get("sep", " ")
                end = kwargs.get("end", "\n")
                f.write(sep.join(str(arg) for arg in args) + end)
        except:
            pass

# ===================================================================================
# CONFIG
# ===================================================================================
current_folder = os.getcwd()
excel_path     = os.path.join(current_folder, "input_gia.xlsx")
PYTHON_EXE     = sys.executable
ERROR_DIR      = os.path.join(current_folder, "errors")
SESSION_LOCK_PATH = os.path.join(current_folder, ".main_session.lock")
OOCL_MANUAL_LOGIN_WAIT_PATH = os.path.join(current_folder, ".oocl_manual_login_wait")
WHL_MANUAL_CAPTCHA_WAIT_PATH = os.path.join(current_folder, ".whl_manual_captcha_wait")
MANUAL_INPUT_WAIT_PATHS = (
    OOCL_MANUAL_LOGIN_WAIT_PATH,
    WHL_MANUAL_CAPTCHA_WAIT_PATH,
)
_session_lock_fd = None
LOG_MODE = (os.environ.get("BOT_LOG_MODE") or os.environ.get("MAIN_LOG_MODE") or "full").strip().lower()
QUIET_SUPPRESSED = {}
PROGRESS_ENABLED = (os.environ.get("MAIN_PROGRESS") or "1").strip().lower() not in {"0", "false", "no", "off"}
PROGRESS_INTERVAL_SECONDS = float(os.environ.get("MAIN_PROGRESS_INTERVAL", "3"))
PROGRESS_DASHBOARD_PATH = os.path.join(current_folder, "progress_dashboard.html")
PROGRESS_TEXT_PATH = os.path.join(current_folder, "progress_status.txt")
PROGRESS_JSON_PATH = os.path.join(current_folder, "progress_status.json")
OPEN_PROGRESS_DASHBOARD = (os.environ.get("OPEN_PROGRESS_DASHBOARD") or "1").strip().lower() in {"1", "true", "yes", "y"}
OOCL_PARALLEL_WORKERS = max(1, int(os.environ.get("OOCL_PARALLEL_WORKERS", "1") or "1"))
BOT_PROGRESS = {}
BOT_PROGRESS_LOCK = threading.Lock()
BOT_CHILD_PIDS = set()
BOT_EDGE_CLEANUP_LOCK = threading.Lock()
BOT_EDGE_CLEANUP_EVENTS = {}
LAST_PROGRESS_PRINT = 0.0

IMPORTANT_LOG_PATTERNS = (
    "==========",
    "dong ",
    "dòng ",
    "row ",
    "route",
    "run ",
    "date_offset",
    "valid",
    "etd=",
    "etd:",
    "tt=",
    "price",
    "gia",
    "giá",
    "wrote excel",
    "da luu",
    "đã lưu",
    "gop",
    "gộp",
    "merge",
    "done",
    "hoan tat",
    "hoàn tất",
    "ok",
    "error",
    "loi",
    "lỗi",
    "exception",
    "traceback",
    "timeout",
    "captcha",
    "security",
    "verify",
    "login",
    "log in",
    "sign in",
    "retry",
    "skip",
    "bo qua",
    "bỏ qua",
    "no service",
    "sold out",
    "no products",
    "no offer",
    "not found",
    "khong co",
    "không có",
    "khong tim",
    "không tìm",
    "failed",
    "exit=",
)

NOISY_LOG_PATTERNS = (
    "-> chốt",
    "-> chot",
    "nhập ",
    "nhap ",
    "click",
    "clicked",
    "đã search",
    "da search",
    "mở popup",
    "mo popup",
    "sleep",
    "waiting",
)

EDGE_KEEP_PATTERNS = (
    "edge_cosco",
    "remote-debugging-port=9523",
    "edge_hpl",
    "remote-debugging-port=9525",
    "edge_oocl",
    "remote-debugging-port=9527",
)

PERSISTENT_EDGE_BOTS = frozenset({
    "bot_cosco.py",
    "bot_hpl.py",
    "bot_oocl.py",
})


def should_keep_bot_edge(bot_file):
    base = os.path.basename(str(bot_file).split("#", 1)[0]).lower()
    return base in PERSISTENT_EDGE_BOTS

EDGE_CLEANUP_PATTERNS = (
    "edge_cosco",
    "remote-debugging-port=9523",
    "edge_cma",
    "remote-debugging-port=9524",
    "edge_hpl",
    "remote-debugging-port=9525",
    "edge_emc",
    "remote-debugging-port=9521",
    "edge_one",
    "remote-debugging-port=9522",
    "edge_kmtc",
    "remote-debugging-port=9526",
    "edge_yangming",
    "remote-debugging-port=9528",
    "edge_whl",
    "remote-debugging-port=9529",
    "edge_msc",
    "remote-debugging-port=9530",
    "edge_hmm",
    "remote-debugging-port=9533",
    "edge_zim_schedule",
    "remote-debugging-port=9534",
    "edge_esl",
    "remote-debugging-port=9535",
    "edge_oocl",
    "remote-debugging-port=9527",
)

def _ps_array(values):
    return "@(" + ",".join("'" + str(v).replace("'", "''") + "'" for v in values) + ")"

def cleanup_bot_edges_after_main():
    if os.name != "nt":
        return
    bot_parent_pids = "@(" + ",".join(str(pid) for pid in sorted(BOT_CHILD_PIDS)) + ")"
    ps = f"""
$killPatterns = {_ps_array(EDGE_CLEANUP_PATTERNS)}
$keepPatterns = {_ps_array(EDGE_KEEP_PATTERNS)}
$botParentPids = {bot_parent_pids}
$killed = 0
$procs = Get-CimInstance Win32_Process -Filter "name = 'msedge.exe'"
foreach ($p in $procs) {{
    $cmd = [string]$p.CommandLine
    if (-not $cmd) {{ continue }}

    $shouldKill = $false
    foreach ($pat in $killPatterns) {{
        if ($cmd -like "*$pat*") {{
            $shouldKill = $true
            break
        }}
    }}
    if (-not $shouldKill) {{ continue }}

    foreach ($pat in $keepPatterns) {{
        if ($cmd -like "*$pat*") {{
            $shouldKill = $false
            break
        }}
    }}
    if ($shouldKill) {{
        Stop-Process -Id $p.ProcessId -Force -ErrorAction SilentlyContinue
        $killed++
    }}
}}
$drivers = Get-CimInstance Win32_Process -Filter "name = 'msedgedriver.exe'"
foreach ($p in $drivers) {{
    if ($botParentPids -contains [int]$p.ParentProcessId) {{
        Stop-Process -Id $p.ProcessId -Force -ErrorAction SilentlyContinue
        $killed++
    }}
}}
Write-Output $killed
"""
    kwargs = {
        "text": True,
        "encoding": "utf-8",
        "errors": "replace",
        "stdout": subprocess.PIPE,
        "stderr": subprocess.DEVNULL,
        "timeout": 20,
    }
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
    try:
        result = subprocess.run(["powershell", "-NoProfile", "-Command", ps], **kwargs)
        killed = (result.stdout or "").strip().splitlines()
        killed_count = killed[-1].strip() if killed else "0"
        print(f"[{ts()}] 🧹 Đã dọn Edge/msedgedriver bot dư: {killed_count} process (giữ COSCO + HAPAG LLOYD + OOCL)")
    except Exception as e:
        print(f"[{ts()}] ⚠️ Không dọn được Edge bot dư: {e}")

def _find_other_main_sessions():
    if os.name != "nt":
        return []
    cmd = [
        "powershell",
        "-NoProfile",
        "-Command",
        (
            "Get-CimInstance Win32_Process -Filter \"name = 'python.exe'\" | "
            f"Where-Object {{ $_.ProcessId -ne {os.getpid()} -and $_.CommandLine -match 'main\\.py' }} | "
            "ForEach-Object { \"$($_.ProcessId)|$($_.CommandLine)\" }"
        ),
    ]
    kwargs = {
        "text": True,
        "encoding": "utf-8",
        "errors": "replace",
        "stderr": subprocess.DEVNULL,
        "timeout": 5,
    }
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
    try:
        out = subprocess.check_output(cmd, **kwargs)
    except Exception:
        return []
    return [line.strip() for line in out.splitlines() if line.strip()]

def release_main_session_lock():
    global _session_lock_fd
    if _session_lock_fd is None:
        return
    fd = _session_lock_fd
    _session_lock_fd = None
    try:
        if msvcrt is not None:
            os.lseek(fd, 0, os.SEEK_SET)
            msvcrt.locking(fd, msvcrt.LK_UNLCK, 1)
    except Exception:
        pass
    try:
        os.close(fd)
    except Exception:
        pass
    try:
        os.remove(SESSION_LOCK_PATH)
    except Exception:
        pass

def acquire_main_session_lock():
    global _session_lock_fd

    if msvcrt is not None:
        fd = os.open(SESSION_LOCK_PATH, os.O_RDWR | os.O_CREAT)
        try:
            os.lseek(fd, 0, os.SEEK_SET)
            msvcrt.locking(fd, msvcrt.LK_NBLCK, 1)
        except OSError:
            try:
                os.lseek(fd, 0, os.SEEK_SET)
                info = os.read(fd, 512).decode("utf-8", errors="replace").strip()
            except Exception:
                info = ""
            os.close(fd)
            print(f"[{ts()}] ❌ Đang có một phiên main.py khác chạy. Không mở thêm phiên mới.")
            if info:
                print(f"[{ts()}]    Phiên đang giữ lock: {info}")
            sys.exit(2)

        _session_lock_fd = fd
        os.ftruncate(fd, 0)
        lock_info = f"pid={os.getpid()} started={datetime.now():%Y-%m-%d %H:%M:%S} args={' '.join(sys.argv)}\n"
        os.write(fd, lock_info.encode("utf-8", errors="replace"))
        os.fsync(fd)
        atexit.register(release_main_session_lock)

    other_sessions = _find_other_main_sessions()
    if other_sessions:
        print(f"[{ts()}] ❌ Đang có một phiên main.py khác chạy. Không mở thêm phiên mới.")
        for session in other_sessions[:5]:
            print(f"[{ts()}]    {session}")
        release_main_session_lock()
        sys.exit(2)

# Carrier → bot file mapping
CARRIER_TO_BOT = {
    "CMA":          "bot_cma.py",
    "ANL":          "bot_cma.py",
    "CNC":          "bot_cma.py",
    "APL":          "bot_cma.py",
    "COSCO":        "bot_COSCO.py",
    "CUL":          "bot_cul.py",
    "CU LINE":      "bot_cul.py",
    "CU LINES":     "bot_cul.py",
    "CULINES":      "bot_cul.py",
    "CHINA UNITED LINES": "bot_cul.py",
    "EMC":          "bot_EMC.py",
    "EVERGREEN":    "bot_EMC.py",
    "ESL":          "bot_esl.py",
    "EMIRATES LINE": "bot_esl.py",
    "EMIRATES SHIPPING LINE": "bot_esl.py",
    "HPL":          "bot_HPL.py",
    "HAPAG":        "bot_HPL.py",
    "HAPAG-LLOYD":  "bot_HPL.py",
    "HAPAG LLOYD":  "bot_HPL.py",
    "HMM":          "bot_hmm.py",
    "HYUNDAI":      "bot_hmm.py",
    "HYUNDAI MERCHANT MARINE": "bot_hmm.py",
    "KMTC":         "bot_KMTC.py",
    "MSC":          "bot_msc.py",
    "ONE":          "bot_one.py",
    "OOCL":         "bot_oocl.py",
    "YANG MING":    "bot_yangming.py",
    "YANGMING":     "bot_yangming.py",
    "YM":           "bot_yangming.py",
    "WHL":          "bot_whl.py",
    "WAN HAI":      "bot_whl.py",
    "WANHAI":       "bot_whl.py",
    "ZIM":          "bot_zim.py",
    "ZIM LINE":     "bot_zim.py",
    "ZIM LINES":    "bot_zim.py",
}

# Bot short names (for display & temp file naming)
BOT_DISPLAY = {
    "bot_cma.py":      "CMA / ANL / CNC / APL",
    "bot_COSCO.py":    "COSCO",
    "bot_cul.py":      "CUL",
    "bot_EMC.py":      "EMC",
    "bot_esl.py":      "ESL",
    "bot_HPL.py":      "HAPAG LLOYD",
    "bot_hmm.py":      "HMM",
    "bot_KMTC.py":     "KMTC",
    "bot_msc.py":      "MSC",
    "bot_one.py":      "ONE",
    "bot_oocl.py":     "OOCL",
    "bot_yangming.py": "YANG MING",
    "bot_whl.py":      "WHL",
    "bot_zim.py":      "ZIM",
}

# Carrier groups — để biết bot nào xử lý carrier nào
BOT_CARRIERS = {
    "bot_cma.py":      {"CMA", "ANL", "CNC", "APL"},
    "bot_COSCO.py":    {"COSCO"},
    "bot_cul.py":      {"CUL", "CU LINE", "CU LINES", "CULINES", "CHINA UNITED LINES"},
    "bot_EMC.py":      {"EMC", "EVERGREEN"},
    "bot_esl.py":      {"ESL", "EMIRATES LINE", "EMIRATES SHIPPING LINE"},
    "bot_HPL.py":      {"HPL", "HAPAG", "HAPAG-LLOYD", "HAPAG LLOYD"},
    "bot_hmm.py":      {"HMM", "HYUNDAI", "HYUNDAI MERCHANT MARINE"},
    "bot_KMTC.py":     {"KMTC"},
    "bot_msc.py":      {"MSC"},
    "bot_one.py":      {"ONE"},
    "bot_oocl.py":     {"OOCL"},
    "bot_yangming.py": {"YANG MING", "YANGMING", "YM"},
    "bot_whl.py":      {"WHL", "WAN HAI", "WANHAI"},
    "bot_zim.py":      {"ZIM", "ZIM LINE", "ZIM LINES"},
}
ALL_BOTS = set(BOT_DISPLAY.keys())
FIRST_ROUND_BOTS = {
    "bot_cma.py",
    "bot_COSCO.py",
    "bot_HPL.py",
    "bot_one.py",
    "bot_oocl.py",
}
MAIN_ROUTE_CARRIERS = [
    "CMA", "COSCO", "CUL", "EMC", "ESL", "HAPAG LLOYD", "HMM",
    "KMTC", "MSC", "ONE", "OOCL", "YANG MING", "WHL", "ZIM",
]

# Columns kết quả mà bot ghi: F(6) G(7) H(8) I(9) J(10) K(11) M(13) N(14) O(15) P(16)
RESULT_COLS = [6, 7, 8, 9, 10, 11, 13, 14, 15, 16]
DEFAULT_DATE_OFFSET_DAYS = 7


# ===================================================================================
# HELPERS
# ===================================================================================
def ts():
    return datetime.now().strftime("%H:%M:%S")

def print_usage():
    print('  python main.py                                      -> chay full Excel')
    print('  python main.py --quiet                              -> chay full, chi in log quan trong')
    print('  python main.py                                      -> tu mo dashboard progress, refresh mac dinh 3 giay')
    print('  python main.py --progress-interval 15               -> doi progress/ETA thanh moi 15 giay')
    print('  python main.py --open-progress                      -> tu mo dashboard progress HTML (mac dinh da bat)')
    print('  python main.py --no-progress                        -> tat progress/ETA')
    print('  python main.py --log important                      -> tuong tu --quiet')
    print('  python main.py --log full                           -> in day du nhu cu')
    print('  python main.py --skip "YANG MING"                   -> chay full, bo qua Yang Ming')
    print('  python main.py --skip YM --skip ZIM                 -> chay full, bo qua Yang Ming va ZIM')
    print('  python main.py --bots HPL,OOCL                      -> chi chay bot HPL va OOCL')
    print('  python main.py HPL OOCL                             -> chi chay bot HPL va OOCL')
    print('  python main.py --date +10                           -> chay full Excel voi ngay +10')
    print('  python main.py ZIM --date +10                       -> chi chay bot ZIM voi ngay +10')
    print('  python main.py ESL --date +7                        -> chi scrape lich tau ESL')
    print('  python main.py "HO CHI MINH" "CHENNAI"              -> check 1 tuyen')
    print('  python main.py "HO CHI MINH" "CHENNAI" --skip YM    -> check 1 tuyen, bo qua Yang Ming')
    print('  python main.py "HO CHI MINH" "OSLO" "NORWAY" --date +10')
    print('  python main.py 145                                  -> check 1 dong')
    print('  python main.py 100-150 --date +12                   -> check dong 100 den 150')

def should_print_bot_line(line):
    mode = (LOG_MODE or "full").strip().lower()
    if mode in {"full", "debug", "verbose", "all"}:
        return True
    if mode not in {"quiet", "important", "minimal", "min"}:
        return True

    text = str(line or "").strip()
    if not text:
        return False

    low = text.lower()
    if any(pat in low for pat in IMPORTANT_LOG_PATTERNS):
        return True

    # In quiet mode, suppress very chatty UI-operation lines unless they also
    # contain an important keyword above.
    if any(pat in low for pat in NOISY_LOG_PATTERNS):
        return False

    return False

def remember_suppressed_line(bot_name):
    QUIET_SUPPRESSED[bot_name] = QUIET_SUPPRESSED.get(bot_name, 0) + 1

def _strip_accents(value):
    text = str(value or "")
    text = text.replace("Đ", "D").replace("đ", "d")
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))

def _fmt_duration(seconds):
    try:
        seconds = int(max(0, float(seconds)))
    except Exception:
        return "?"
    if seconds < 60:
        return f"{seconds}s"
    minutes, sec = divmod(seconds, 60)
    if minutes < 60:
        return f"{minutes}m{sec:02d}s"
    hours, minutes = divmod(minutes, 60)
    return f"{hours}h{minutes:02d}m"

def progress_start_bot(name):
    if not PROGRESS_ENABLED:
        return
    with BOT_PROGRESS_LOCK:
        BOT_PROGRESS[name] = {
            "started": time.time(),
            "last_seen": time.time(),
            "current": 0,
            "total": None,
            "done": False,
            "status": "starting",
            "last_line": "",
        }

def progress_update_from_line(name, line):
    if not PROGRESS_ENABLED:
        return
    text = str(line or "").strip()
    if not text:
        return
    low = _strip_accents(text).lower()
    with BOT_PROGRESS_LOCK:
        p = BOT_PROGRESS.setdefault(name, {
            "started": time.time(),
            "last_seen": time.time(),
            "current": 0,
            "total": None,
            "done": False,
            "status": "running",
            "last_line": "",
        })
        p["last_seen"] = time.time()
        p["last_line"] = text[:120]

        # Match both plain "[9/25]" and labelled forms like "[ONE-API 9/25]".
        m = re.search(r"\[[^\]]*?(\d+)\s*/\s*(\d+)[^\]]*?\]", text)
        if m:
            p["current"] = max(int(m.group(1)), int(p.get("current") or 0))
            p["total"] = max(int(m.group(2)), int(p.get("total") or 0))

        # Keep the physical spreadsheet row separate from the ordinal
        # progress (for example: MSC 3/28 is checking Excel row 150).
        m_excel_row = re.search(
            r"(?:processing|dang check|dang kiem tra|kiem tra|xu ly)\s+"
            r"(?:excel\s+)?(?:row|dong)\s+(\d{1,6})\b",
            low,
        )
        if m_excel_row:
            p["excel_row"] = int(m_excel_row.group(1))

        m_total = re.search(
            r"(?:co|tong cong|total|read|doc duoc|can check|sau khi loc)[^\d]{0,40}(\d{1,5})[^\n]{0,30}(?:dong|row|rows|tuyen)",
            low,
        )
        if m_total and not p.get("total"):
            p["total"] = int(m_total.group(1))

        m_saved_count = re.search(r"(?:da luu|da ghi|saved)\s+(\d{1,5})\s+(?:dong|row|rows)", low)
        if m_saved_count and p.get("total"):
            p["current"] = min(
                max(int(m_saved_count.group(1)), int(p.get("current") or 0)),
                int(p["total"]),
            )

        # OOCL logs spreadsheet row numbers (for example row 149), not an
        # ordinal such as [2/28]. Count each explicit completion once. This
        # covers both priced rows ("Ghi Excel") and no-rate rows ("Ghi '-'").
        is_oocl = _compact_name(name) == "OOCL"
        m_oocl_done = re.search(r"hoan tat row\s+(\d{1,6})\b", low) if is_oocl else None
        if m_oocl_done and p.get("total"):
            completed_rows = p.setdefault("completed_row_ids", set())
            completed_rows.add(int(m_oocl_done.group(1)))
            p["current"] = min(len(completed_rows), int(p["total"]))

        if any(k in low for k in ("captcha", "security", "verify", "login", "log in", "sign in")):
            p["status"] = "waiting auth/captcha"
        elif any(k in low for k in ("error", "exception", "traceback", "timeout", " loi", "loi ")):
            p["status"] = "warning/error"
        elif any(k in low for k in ("wrote excel", "da luu", "da ghi", "ghi excel", "saved row", "saved dong")):
            p["status"] = "saving result"
            if p.get("total") and not is_oocl and not re.search(r"\[[^\]]*?(\d+)\s*/\s*(\d+)[^\]]*?\]", text) and not m_saved_count:
                p["current"] = min(int(p.get("current") or 0) + 1, int(p["total"]))
        elif any(k in low for k in ("dang check", "dang kiem tra", "search", "etd", "price", "gia")):
            p["status"] = "running"

def progress_finish_bot(name, returncode=None):
    if not PROGRESS_ENABLED:
        return
    with BOT_PROGRESS_LOCK:
        p = BOT_PROGRESS.setdefault(name, {"started": time.time()})
        finished_at = time.time()
        p["done"] = True
        p["last_seen"] = finished_at
        # Preserve the first observed finish time so later refreshes cannot change it.
        p.setdefault("completed_at", finished_at)
        p["status"] = "OK" if returncode == 0 else f"exit={returncode}"
        if p.get("total"):
            p["current"] = p["total"]

def _progress_line_for_bot(name, p):
    total = p.get("total")
    current = int(p.get("current") or 0)
    elapsed = time.time() - float(p.get("started") or time.time())
    status = p.get("status") or "running"
    if p.get("done"):
        completed_at = float(p.get("completed_at") or p.get("last_seen") or time.time())
        return f"{name}: DONE {datetime.fromtimestamp(completed_at).strftime('%H:%M')} ({status})"
    if total:
        pct = min(100, int(current * 100 / max(1, int(total))))
        eta = "?"
        if current > 0:
            eta_seconds = (elapsed / current) * max(0, int(total) - current)
            eta = datetime.fromtimestamp(time.time() + eta_seconds).strftime("%H:%M")
        row_text = f" | Excel row {p['excel_row']}" if p.get("excel_row") else ""
        return f"{name}: {current}/{total} {pct}%{row_text} | Expected finish {eta} | {status}"
    return f"{name}: ?/? | {_fmt_duration(elapsed)} | {status}"

def _progress_snapshot_data():
    now = time.time()
    with BOT_PROGRESS_LOCK:
        items = []
        for name, raw in sorted(BOT_PROGRESS.items()):
            p = dict(raw)
            total = p.get("total")
            current = int(p.get("current") or 0)
            started = float(p.get("started") or now)
            elapsed = max(0.0, now - started)
            done = bool(p.get("done"))
            pct = None
            eta_seconds = None
            completed_at_ts = None
            if total:
                total_int = int(total)
                pct = min(100, int(current * 100 / max(1, total_int)))
                if current > 0 and not done:
                    eta_seconds = (elapsed / current) * max(0, total_int - current)
            elif done:
                pct = 100
            if done:
                completed_at_ts = float(p.get("completed_at") or p.get("last_seen") or now)
                finish_time = datetime.fromtimestamp(completed_at_ts).strftime("%H:%M")
            elif eta_seconds is not None:
                finish_time = datetime.fromtimestamp(now + eta_seconds).strftime("%H:%M")
            else:
                finish_time = "?"
            items.append({
                "name": name,
                "current": current,
                "total": int(total) if total else None,
                "excel_row": p.get("excel_row"),
                "pct": pct,
                "eta_seconds": eta_seconds,
                "eta": finish_time,
                "finish_time": finish_time,
                "completed_at": (
                    datetime.fromtimestamp(completed_at_ts).strftime("%Y-%m-%d %H:%M:%S")
                    if completed_at_ts is not None else None
                ),
                "completed_at_seconds": completed_at_ts,
                "elapsed_seconds": elapsed,
                "elapsed": _fmt_duration(elapsed),
                "done": done,
                "status": p.get("status") or "running",
                "last_line": p.get("last_line") or "",
            })

    total_bots = len(items)
    done_bots = sum(1 for x in items if x["done"])
    known = [x for x in items if x["total"]]
    sum_current = sum(x["current"] for x in known)
    sum_total = sum(x["total"] for x in known)
    overall_pct = None
    if sum_total:
        overall_pct = min(100, int(sum_current * 100 / max(1, sum_total)))
    elif total_bots:
        overall_pct = min(100, int(done_bots * 100 / total_bots))

    eta_values = [x["eta_seconds"] for x in items if x["eta_seconds"] is not None and not x["done"]]
    overall_eta = max(eta_values) if eta_values else None
    overall_finish_time = (
        datetime.fromtimestamp(now + overall_eta).strftime("%H:%M")
        if overall_eta is not None else None
    )
    if overall_finish_time is None and items and done_bots == total_bots:
        completed_times = [float(x["completed_at_seconds"] or now) for x in items]
        overall_finish_time = datetime.fromtimestamp(max(completed_times)).strftime("%H:%M")
    elapsed_values = [x["elapsed_seconds"] for x in items]

    return {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "overall": {
            "bots_done": done_bots,
            "bots_total": total_bots,
            "current": sum_current,
            "total": sum_total,
            "pct": overall_pct,
            "eta_seconds": overall_eta,
            "eta": overall_finish_time or "?",
            "finish_time": overall_finish_time or "?",
            "elapsed": _fmt_duration(max(elapsed_values) if elapsed_values else 0),
        },
        "bots": items,
    }

def _atomic_write_text(path, text):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(text)
    os.replace(tmp, path)

def open_progress_dashboard_once():
    if not OPEN_PROGRESS_DASHBOARD:
        return
    if getattr(open_progress_dashboard_once, "_opened", False):
        return
    try:
        if os.path.exists(PROGRESS_DASHBOARD_PATH):
            edge_path = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
            if os.path.exists(edge_path):
                import subprocess
                subprocess.Popen([
                    edge_path,
                    f"--app=file:///{PROGRESS_DASHBOARD_PATH.replace(chr(92), '/')}",
                    "--window-size=450,900",
                    "--window-position=1400,50"
                ])
            else:
                os.startfile(PROGRESS_DASHBOARD_PATH)
            open_progress_dashboard_once._opened = True
    except Exception as e:
        print(f"[{ts()}] Cannot open progress dashboard: {e}")

def write_progress_dashboard():
    if not PROGRESS_ENABLED:
        return
    data = _progress_snapshot_data()
    overall = data["overall"]
    bots = data["bots"]

    txt_lines = [
        f"Updated: {data['updated_at']}",
        f"OVERALL: {overall['bots_done']}/{overall['bots_total']} bots | "
        f"{overall['current']}/{overall['total'] or '?'} rows | "
        f"{overall['pct'] if overall['pct'] is not None else '?'}% | Finish time {overall['eta']}",
        "",
    ]
    rows_html = []
    for b in bots:
        progress_text = f"{b['current']}/{b['total']}" if b["total"] else "?/?"
        excel_row_text = str(b["excel_row"]) if b.get("excel_row") else "-"
        txt_lines.append(
            f"{b['name']}: {progress_text} | Excel row {excel_row_text} | "
            f"{b['pct'] if b['pct'] is not None else '?'}% | Finish time {b['eta']} | "
            f"{b['status']} | {b['last_line']}"
        )
        pct = b["pct"] if b["pct"] is not None else 0
        status_low = str(b["status"]).lower()
        status_class = "done" if b["done"] else ("warn" if "error" in status_low or "captcha" in status_low else "run")
        rows_html.append(f"""
        <tr class="{status_class}">
          <td class="bot">{html.escape(str(b['name']))}</td>
          <td>{html.escape(progress_text)}</td>
          <td>{html.escape(excel_row_text)}</td>
          <td><div class="bar"><div class="fill" style="width:{int(pct)}%"></div></div><span>{html.escape(str(b['pct']) if b['pct'] is not None else '?')}%</span></td>
          <td>{html.escape(str(b['eta']))}</td>
          <td>{html.escape(str(b['elapsed']))}</td>
          <td>{html.escape(str(b['status']))}</td>
          <td class="last">{html.escape(str(b['last_line']))}</td>
        </tr>
        """)

    overall_pct = overall["pct"] if overall["pct"] is not None else 0
    refresh_seconds = max(3, int(PROGRESS_INTERVAL_SECONDS))
    html_text = f"""<!doctype html>
<html lang="vi">
<head>
  <meta charset="utf-8">
  <meta http-equiv="refresh" content="{refresh_seconds}">
  <title>Bot Progress Dashboard</title>
  <style>
    body {{ font-family: Segoe UI, Arial, sans-serif; margin:24px; background:#0f172a; color:#e5e7eb; }}
    h1 {{ margin:0 0 8px; }}
    .muted {{ color:#94a3b8; }}
    .card {{ background:#111827; border:1px solid #334155; border-radius:14px; padding:18px; margin-top:18px; box-shadow:0 8px 24px rgba(0,0,0,.25); }}
    .overall {{ display:grid; grid-template-columns:repeat(4,minmax(120px,1fr)); gap:12px; }}
    .metric {{ background:#0b1220; border-radius:10px; padding:12px; }}
    .metric .label {{ color:#94a3b8; font-size:12px; }}
    .metric .value {{ font-size:24px; font-weight:700; margin-top:4px; }}
    table {{ width:100%; border-collapse:collapse; }}
    th,td {{ padding:10px 12px; border-bottom:1px solid #263244; text-align:left; vertical-align:top; }}
    th {{ color:#cbd5e1; background:#162033; }}
    tr.done td {{ color:#86efac; }}
    tr.warn td {{ color:#fcd34d; }}
    .bot {{ font-weight:700; white-space:nowrap; }}
    .last {{ color:#94a3b8; max-width:520px; overflow:hidden; text-overflow:ellipsis; }}
    .bar {{ display:inline-block; width:180px; height:12px; background:#334155; border-radius:99px; overflow:hidden; margin-right:8px; vertical-align:middle; }}
    .fill {{ height:100%; background:linear-gradient(90deg,#22c55e,#38bdf8); border-radius:99px; }}
  </style>
</head>
<body>
  <h1>Bot Progress Dashboard</h1>
  <div class="muted">Tự refresh mỗi {refresh_seconds} giây · Updated: {html.escape(data['updated_at'])}</div>
  <div class="card overall">
    <div class="metric"><div class="label">Overall</div><div class="value">{overall_pct}%</div></div>
    <div class="metric"><div class="label">Bots</div><div class="value">{overall['bots_done']}/{overall['bots_total']}</div></div>
    <div class="metric"><div class="label">Rows</div><div class="value">{overall['current']}/{overall['total'] or '?'}</div></div>
    <div class="metric"><div class="label">Dự kiến xong</div><div class="value">{html.escape(str(overall['eta']))}</div></div>
  </div>
  <div class="card">
    <table>
      <thead><tr><th>Bot</th><th>Rows</th><th>Excel row</th><th>Progress</th><th>Thời gian xong</th><th>Elapsed</th><th>Status</th><th>Last log</th></tr></thead>
      <tbody>{''.join(rows_html)}</tbody>
    </table>
  </div>
</body>
</html>
"""
    try:
        _atomic_write_text(PROGRESS_TEXT_PATH, "\n".join(txt_lines) + "\n")
        _atomic_write_text(PROGRESS_JSON_PATH, json.dumps(data, ensure_ascii=False, indent=2))
        _atomic_write_text(PROGRESS_DASHBOARD_PATH, html_text)
        open_progress_dashboard_once()
    except Exception as e:
        try:
            _orig_print(f"[{ts()}] [PROGRESS] Cannot write dashboard: {e}")
        except Exception:
            pass

def print_progress_snapshot(force=False):
    global LAST_PROGRESS_PRINT
    if not PROGRESS_ENABLED:
        return
    now = time.time()
    if not force and now - LAST_PROGRESS_PRINT < PROGRESS_INTERVAL_SECONDS:
        return
    with BOT_PROGRESS_LOCK:
        if not BOT_PROGRESS:
            return
        lines = [_progress_line_for_bot(name, p) for name, p in sorted(BOT_PROGRESS.items())]
    LAST_PROGRESS_PRINT = now
    write_progress_dashboard()
    # print(f"[{ts()}] 📊 PROGRESS: " + " || ".join(lines))

def progress_monitor(stop_event):
    while not stop_event.is_set():
        stop_event.wait(PROGRESS_INTERVAL_SECONDS)
        if not stop_event.is_set():
            print_progress_snapshot(force=True)

def _compact_name(value):
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())

def _bot_alias_compacts(bot_file):
    aliases = {
        bot_file,
        bot_file.replace(".py", ""),
        bot_file.replace("bot_", "").replace(".py", ""),
        BOT_DISPLAY.get(bot_file, bot_file),
    }
    aliases.update(BOT_CARRIERS.get(bot_file, set()))
    return {_compact_name(a) for a in aliases if a}

def parse_skip_bots(skip_values):
    """Đổi --skip/--exclude từ carrier/bot name sang set bot_file."""
    skip_bots = set()
    unknown = []
    alias_map = {}
    for bot_file in ALL_BOTS:
        for alias in _bot_alias_compacts(bot_file):
            alias_map[alias] = bot_file

    for raw in skip_values or []:
        for token in str(raw or "").replace(";", ",").split(","):
            token = token.strip()
            if not token:
                continue
            compact = _compact_name(token)
            bot_file = alias_map.get(compact)
            if bot_file:
                skip_bots.add(bot_file)
            else:
                unknown.append(token)

    if unknown:
        print(f"[{ts()}] ⚠️ Không nhận ra --skip: {', '.join(unknown)}")
        print(f"[{ts()}]    Gợi ý: dùng tên carrier/bot như YM, \"YANG MING\", ZIM, CMA, OOCL...")
    return skip_bots

def parse_bot_names(raw_values, warn=True):
    """Doi danh sach carrier/bot name sang set bot_file."""
    bots = set()
    unknown = []
    alias_map = {}
    for bot_file in ALL_BOTS:
        for alias in _bot_alias_compacts(bot_file):
            alias_map[alias] = bot_file

    for raw in raw_values or []:
        for token in str(raw or "").replace(";", ",").split(","):
            token = token.strip()
            if not token:
                continue
            bot_file = alias_map.get(_compact_name(token))
            if bot_file:
                bots.add(bot_file)
            else:
                unknown.append(token)

    if warn and unknown:
        print(f"[{ts()}] Khong nhan ra bot/carrier: {', '.join(unknown)}")
        print(f"[{ts()}] Goi y: HPL, OOCL, YM, ZIM, CMA, MSC...")
    return bots, unknown

def describe_bots(bot_files):
    return ", ".join(BOT_DISPLAY.get(b, b) for b in sorted(bot_files))

def apply_skip_bots(bots, skip_bots, context=""):
    bots = set(bots or [])
    skip_bots = set(skip_bots or [])
    skipped = bots & skip_bots
    if skipped:
        label = f" ({context})" if context else ""
        print(f"[{ts()}] ⏭️ Skip{label}: {describe_bots(skipped)}")
    return bots - skip_bots

def filter_carriers_by_skip(carriers, skip_bots):
    skip_bots = set(skip_bots or [])
    result = []
    skipped = []
    for carrier in carriers:
        bot = CARRIER_TO_BOT.get(str(carrier).strip().upper())
        if bot in skip_bots:
            skipped.append(carrier)
        else:
            result.append(carrier)
    if skipped:
        print(f"[{ts()}] ⏭️ Không tạo/chạy carrier do --skip: {', '.join(skipped)}")
    return result

def parse_cli_args(argv):
    global LOG_MODE, PROGRESS_ENABLED, PROGRESS_INTERVAL_SECONDS, OPEN_PROGRESS_DASHBOARD

    normalized = []
    for arg in argv:
        if arg.lower() == "--date":
            normalized.append("--date")
        elif arg.lower().startswith("--date="):
            normalized.append("--date=" + arg.split("=", 1)[1])
        else:
            normalized.append(arg)

    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--date", default=f"+{DEFAULT_DATE_OFFSET_DAYS}")
    parser.add_argument("--skip", "--exclude", action="append", default=[])
    parser.add_argument("--bots", "--only", action="append", default=[])
    parser.add_argument("--quiet", action="store_true")
    parser.add_argument("--log", "--log-mode", dest="log_mode", default=None)
    parser.add_argument("--progress", action="store_true")
    parser.add_argument("--no-progress", action="store_true")
    parser.add_argument("--open-progress", action="store_true")
    parser.add_argument("--progress-interval", default=None)
    parser.add_argument("positional", nargs="*")

    try:
        parsed = parser.parse_args(normalized)
    except SystemExit:
        print_usage()
        sys.exit(1)

    raw_offset = str(parsed.date).strip()
    try:
        date_offset = int(raw_offset.lstrip("+"))
    except ValueError:
        print(f"[{ts()}] --date khong hop le: '{raw_offset}'. Vi du dung: --date +10")
        sys.exit(1)

    if date_offset < 0:
        print(f"[{ts()}] --date phai la so ngay >= 0. Vi du dung: --date +10")
        sys.exit(1)

    skip_bots = parse_skip_bots(parsed.skip)
    selected_bots, _ = parse_bot_names(parsed.bots, warn=True)
    if parsed.quiet:
        LOG_MODE = "important"
    elif parsed.log_mode:
        LOG_MODE = str(parsed.log_mode).strip().lower()
    if parsed.no_progress:
        PROGRESS_ENABLED = False
    elif parsed.progress:
        PROGRESS_ENABLED = True
    if parsed.open_progress:
        OPEN_PROGRESS_DASHBOARD = True
    if parsed.progress_interval is not None:
        try:
            PROGRESS_INTERVAL_SECONDS = max(3.0, float(parsed.progress_interval))
        except Exception:
            print(f"[{ts()}] --progress-interval khong hop le: {parsed.progress_interval}; dung mac dinh {PROGRESS_INTERVAL_SECONDS}s")
    return parsed.positional, date_offset, skip_bots, selected_bots

def build_date_env(date_offset):
    return {"DATE_OFFSET_DAYS": str(date_offset)}

def parse_row_range_token(value):
    """Parse positional START-END, e.g. 100-150."""
    match = re.fullmatch(r"\s*(\d+)\s*-\s*(\d+)\s*", str(value or ""))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))

def print_header():
    print(f"""
╔═══════════════════════════════════════════════════╗
║   MAIN ORCHESTRATOR — All Bot Price Checker 🚀    ║
╚═══════════════════════════════════════════════════╝
[{ts()}] Thư mục: {current_folder}
[{ts()}] Excel:   {excel_path}
""")

def check_excel():
    if not os.path.exists(excel_path):
        print(f"[{ts()}] ❌ Không tìm thấy {excel_path}")
        sys.exit(1)

def get_all_carriers():
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    carriers = set()
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        c = str(row[4] or "").strip().upper()
        if c:
            carriers.add(c)
    wb.close()
    return carriers

def get_bots_for_carriers(carriers):
    bots = set()
    for c in carriers:
        bot = CARRIER_TO_BOT.get(c)
        if bot:
            bots.add(bot)
        else:
            print(f"[{ts()}] ⚠️ Carrier '{c}' không có bot tương ứng")
    return bots

def make_bot_copy(bot_file, suffix=None):
    """
    Tạo bản copy Excel riêng cho mỗi bot.
    Ví dụ: input_gia_HPL.xlsx
    Bot sẽ đọc/ghi vào bản copy này thay vì file gốc.
    """
    name = BOT_DISPLAY.get(bot_file, bot_file.replace(".py", ""))
    if suffix:
        name = f"{name}_{suffix}"
    import re
    safe_name = re.sub(r'[\\/*?:"<>|]', '-', name)
    copy_path = excel_path.replace(".xlsx", f"_{safe_name}.xlsx")
    shutil.copy2(excel_path, copy_path)
    return copy_path

def restrict_bot_copy_to_row_range(copy_path, extra_env=None):
    """Blank carrier outside FILTER_ROW_FROM/TO while preserving row numbers."""
    env = extra_env or {}
    raw_from = env.get("FILTER_ROW_FROM")
    raw_to = env.get("FILTER_ROW_TO")
    if raw_from in (None, "") or raw_to in (None, ""):
        return copy_path

    try:
        row_from = int(raw_from)
        row_to = int(raw_to)
    except (TypeError, ValueError):
        raise ValueError(f"Khoảng dòng không hợp lệ: {raw_from}-{raw_to}")

    wb = openpyxl.load_workbook(copy_path)
    ws = wb.active
    for row_num in range(2, ws.max_row + 1):
        if row_num < row_from or row_num > row_to:
            ws.cell(row=row_num, column=5).value = None
    wb.save(copy_path)
    wb.close()
    return copy_path

def run_bot(bot_file, excel_override=None, extra_env=None):
    """
    Chạy bot trong subprocess.
    Nếu excel_override, set env EXCEL_PATH để bot dùng file khác.
    """
    bot_path = os.path.join(current_folder, bot_file)
    if not os.path.exists(bot_path):
        print(f"[{ts()}] ❌ Không tìm thấy: {bot_file}")
        return None

    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUNBUFFERED"] = "1"
    env["ERROR_DIR"] = ERROR_DIR
    env["OOCL_MANUAL_LOGIN_WAIT_PATH"] = OOCL_MANUAL_LOGIN_WAIT_PATH
    env["WHL_MANUAL_CAPTCHA_WAIT_PATH"] = WHL_MANUAL_CAPTCHA_WAIT_PATH
    if excel_override:
        env["EXCEL_PATH"] = excel_override
    if extra_env:
        env.update(extra_env)

    proc = subprocess.Popen(
        [PYTHON_EXE, bot_path],
        env=env,
        cwd=current_folder,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    BOT_CHILD_PIDS.add(proc.pid)
    cleanup_key = bot_file
    if str(bot_file).lower() == "bot_oocl.py" and extra_env and "OOCL_WORKER_INDEX" in extra_env:
        try:
            cleanup_key = f"{bot_file}#W{int(extra_env['OOCL_WORKER_INDEX']) + 1}"
        except (TypeError, ValueError):
            pass
    threading.Thread(
        target=watch_bot_completion,
        args=(proc, cleanup_key),
        daemon=True,
    ).start()
    return proc

BOT_EDGE_PATTERNS = {
    "bot_emc.py": "edge_emc",
    "bot_one.py": "edge_one",
    "bot_kmtc.py": "edge_kmtc",
    "bot_yangming.py": "edge_yangming",
    "bot_whl.py": "edge_whl",
    "bot_msc.py": "edge_msc",
    "bot_hmm.py": "edge_hmm",
    "bot_zim.py": "edge_zim_schedule",
    "bot_esl.py": "edge_esl",
    "bot_cma.py": "edge_cma",
    "bot_oocl.py": "edge_oocl",
    "bot_cosco.py": "edge_cosco",
    "bot_hpl.py": "edge_hpl",
}

def kill_specific_bot_edge(bot_file, bot_pid=None):
    import subprocess
    import os
    if os.name != "nt":
        return False
    raw_bot_file = str(bot_file)
    parts = raw_bot_file.split("#", 1)
    base_bot_file = os.path.basename(parts[0]).lower()
    suffix = parts[1].upper() if len(parts) > 1 else ""
    keep_edge = should_keep_bot_edge(raw_bot_file)
    pattern = BOT_EDGE_PATTERNS.get(base_bot_file)
    if base_bot_file == "bot_oocl.py" and re.fullmatch(r"W\d+", suffix):
        pattern = f"edge_oocl_w{int(suffix[1:])}"
    if not pattern and not bot_pid:
        return False

    edge_cleanup_ps = ""
    if pattern and not keep_edge:
        edge_cleanup_ps = f'''
$procs = Get-CimInstance Win32_Process -Filter "name = 'msedge.exe'"
foreach ($p in $procs) {{
    $cmd = [string]$p.CommandLine
    if ($cmd -like "*{pattern}*") {{
        Stop-Process -Id $p.ProcessId -Force -ErrorAction SilentlyContinue
    }}
}}
'''

    ps = f'''
{edge_cleanup_ps}
if ({int(bot_pid or 0)} -gt 0) {{
    Get-CimInstance Win32_Process -Filter "name = 'msedgedriver.exe'" |
        Where-Object {{ $_.ParentProcessId -eq {int(bot_pid or 0)} }} |
        ForEach-Object {{ Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }}
}}
'''
    try:
        kwargs = {}
        if hasattr(subprocess, "CREATE_NO_WINDOW"):
            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        subprocess.run(["powershell", "-NoProfile", "-Command", ps],
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                       timeout=15, **kwargs)
        return True
    except Exception:
        return False

def cleanup_bot_edge_once(bot_file, proc, wait=True):
    """Close one bot's dedicated Edge exactly once when its process exits."""
    pid = int(getattr(proc, "pid", 0) or 0)
    with BOT_EDGE_CLEANUP_LOCK:
        event = BOT_EDGE_CLEANUP_EVENTS.get(pid)
        owner = event is None
        if owner:
            event = threading.Event()
            BOT_EDGE_CLEANUP_EVENTS[pid] = event

    if not owner:
        if wait:
            event.wait(timeout=20)
        return

    try:
        cleaned = kill_specific_bot_edge(bot_file, bot_pid=pid)
        if cleaned:
            base = str(bot_file).split("#", 1)[0]
            display = next(
                (name for file_name, name in BOT_DISPLAY.items()
                 if file_name.lower() == base.lower()),
                base,
            )
            if should_keep_bot_edge(bot_file):
                print(f"[{ts()}] [{display}] 🧹 Bot done → giữ Edge, chỉ dọn msedgedriver.")
            else:
                print(f"[{ts()}] [{display}] 🧹 Bot done → đã tắt Edge ngay.")
    finally:
        event.set()

def watch_bot_completion(proc, bot_file):
    """Independent watcher: browser cleanup must not wait for stdout EOF."""
    try:
        proc.wait()
    finally:
        cleanup_bot_edge_once(bot_file, proc, wait=False)

def stream_output(name, proc, bot_file=None):
    """Đọc stdout của process và in với prefix [BOT_NAME]"""
    try:
        for line in proc.stdout:
            line = line.rstrip("\n\r")
            if line:
                progress_update_from_line(name, line)
                if should_print_bot_line(line):
                    print(f"[{ts()}] [{name}] {line}")
                else:
                    remember_suppressed_line(name)
                print_progress_snapshot()
    except (OSError, ValueError):
        # Browser descendants may retain the inherited stdout handle.
        pass
    # Chờ process exit tối đa 10s sau khi stdout đóng
    try:
        proc.wait(timeout=10)
    except subprocess.TimeoutExpired:
        print(f"[{ts()}] [{name}] ⚠️ Process không exit sau 10s — force kill")
        proc.kill()
        proc.wait()
    rc = proc.returncode
    status = "✅ OK" if rc == 0 else f"❌ exit={rc}"
    progress_finish_bot(name, rc)
    hidden = QUIET_SUPPRESSED.pop(name, 0)
    if hidden and (LOG_MODE or "").lower() in {"quiet", "important", "minimal", "min"}:
        print(f"[{ts()}] [{name}] ... da an {hidden} dong log chi tiet (--log full de xem het)")
    print(f"[{ts()}] [{name}] {status}")

# Không giới hạn timeout — chờ bot tự xong

def run_bots_parallel(bots_to_run, use_copies=True, extra_env=None):
    """
    Chạy nhiều bot song song. Mỗi bot ghi vào bản copy xlsx riêng.
    Trả về dict {bot_file: copy_path}
    extra_env: dict env vars bổ sung (VD: FILTER_POL, FILTER_POD)
    """
    procs = {}   # name → (proc, thread, bot_file)
    copies = {}  # bot_file → copy_path

    for bot_file in sorted(bots_to_run):
        name = BOT_DISPLAY.get(bot_file, bot_file)

        if bot_file == "bot_oocl.py" and use_copies and OOCL_PARALLEL_WORKERS > 1:
            print(f"[{ts()}] 🚀 OOCL chạy {OOCL_PARALLEL_WORKERS} Edge/worker song song")
            for worker_idx in range(OOCL_PARALLEL_WORKERS):
                worker_no = worker_idx + 1
                worker_name = f"OOCL-W{worker_no}"
                copy_key = f"{bot_file}#W{worker_no}"
                copy_path = make_bot_copy(bot_file, suffix=f"W{worker_no}")
                restrict_bot_copy_to_row_range(copy_path, extra_env)
                copies[copy_key] = copy_path
                worker_env = dict(extra_env or {})
                worker_env.update({
                    "OOCL_WORKER_INDEX": str(worker_idx),
                    "OOCL_WORKER_COUNT": str(OOCL_PARALLEL_WORKERS),
                    "OOCL_EDGE_PORT": str(9540 + worker_idx),
                    "OOCL_EDGE_PROFILE": rf"C:\edge_oocl_w{worker_no}",
                })
                print(f"[{ts()}] 🚀 {worker_name} → {os.path.basename(copy_path)} | port {9540 + worker_idx}")
                proc = run_bot(bot_file, excel_override=copy_path, extra_env=worker_env)
                if proc:
                    progress_start_bot(worker_name)
                    t = threading.Thread(target=stream_output, args=(worker_name, proc, copy_key), daemon=True)
                    t.start()
                    procs[worker_name] = (proc, t, copy_key)
                    time.sleep(2)
            continue

        if use_copies:
            copy_path = make_bot_copy(bot_file)
            restrict_bot_copy_to_row_range(copy_path, extra_env)
            copies[bot_file] = copy_path
            print(f"[{ts()}] 🚀 {name} → {os.path.basename(copy_path)}")
            proc = run_bot(bot_file, excel_override=copy_path, extra_env=extra_env)
        else:
            print(f"[{ts()}] 🚀 {name}")
            proc = run_bot(bot_file, extra_env=extra_env)

        if proc:
            progress_start_bot(name)
            t = threading.Thread(target=stream_output, args=(name, proc, bot_file), daemon=True)
            t.start()
            procs[name] = (proc, t, bot_file)
            time.sleep(2)  # stagger 2s giữa các bot

    # Chờ tất cả kết thúc — không giới hạn thời gian
    print(f"\n[{ts()}] ⏳ Đang chờ {len(procs)} bot hoàn tất...\n")
    progress_stop = threading.Event()
    progress_thread = None
    if PROGRESS_ENABLED and procs:
        print(f"[{ts()}] Dashboard progress: {PROGRESS_DASHBOARD_PATH}")
        print(f"[{ts()}] Text progress:      {PROGRESS_TEXT_PATH}")
        print_progress_snapshot(force=True)
        progress_thread = threading.Thread(target=progress_monitor, args=(progress_stop,), daemon=True)
        progress_thread.start()
    for name, (proc, thread, bot_file) in procs.items():
        proc.wait()
        # Give a normally closed pipe one second. Only invoke the process
        # cleanup fallback when Edge/msedgedriver still owns that pipe.
        # Never close proc.stdout from this thread: on Windows that can
        # deadlock with stream_output() reading it.
        thread.join(timeout=1)
        if thread.is_alive():
            cleanup_bot_edge_once(bot_file, proc)
            thread.join(timeout=5)
        if thread.is_alive():
            print(f"[{ts()}] [{name}] Bot đã exit; bỏ reader pipe còn bị Edge giữ, không chặn main.")
        if proc.returncode != 0:
            failed_copy = copies.pop(bot_file, None)
            if failed_copy and os.path.exists(failed_copy):
                try:
                    os.remove(failed_copy)
                except:
                    pass
            print(f"[{ts()}] ⚠️ [{name}] lỗi nên bỏ qua file copy, không gộp kết quả vào Excel gốc")

    if progress_thread:
        progress_stop.set()
        progress_thread.join(timeout=2)
        print_progress_snapshot(force=True)

    return copies

def launch_bots_async(bots_to_run, procs, copies, extra_env=None, use_copies=True, phase_label=""):
    """
    Launch bot subprocesses without waiting.
    procs: shared dict name -> (proc, thread, bot_key)
    copies: shared dict bot_key -> copy_path
    Returns list of launched process display names.
    """
    launched = []
    for bot_file in sorted(bots_to_run):
        name = BOT_DISPLAY.get(bot_file, bot_file)
        phase_prefix = f"{phase_label} " if phase_label else ""

        if bot_file == "bot_oocl.py" and use_copies and OOCL_PARALLEL_WORKERS > 1:
            print(f"[{ts()}] 🚀 {phase_prefix}OOCL chạy {OOCL_PARALLEL_WORKERS} Edge/worker song song")
            for worker_idx in range(OOCL_PARALLEL_WORKERS):
                worker_no = worker_idx + 1
                worker_name = f"OOCL-W{worker_no}"
                copy_key = f"{bot_file}#W{worker_no}"
                copy_path = make_bot_copy(bot_file, suffix=f"W{worker_no}")
                restrict_bot_copy_to_row_range(copy_path, extra_env)
                copies[copy_key] = copy_path
                worker_env = dict(extra_env or {})
                worker_env.update({
                    "OOCL_WORKER_INDEX": str(worker_idx),
                    "OOCL_WORKER_COUNT": str(OOCL_PARALLEL_WORKERS),
                    "OOCL_EDGE_PORT": str(9540 + worker_idx),
                    "OOCL_EDGE_PROFILE": rf"C:\edge_oocl_w{worker_no}",
                })
                print(f"[{ts()}] 🚀 {phase_prefix}{worker_name} → {os.path.basename(copy_path)} | port {9540 + worker_idx}")
                proc = run_bot(bot_file, excel_override=copy_path, extra_env=worker_env)
                if proc:
                    progress_start_bot(worker_name)
                    t = threading.Thread(target=stream_output, args=(worker_name, proc, copy_key), daemon=True)
                    t.start()
                    procs[worker_name] = (proc, t, copy_key)
                    launched.append(worker_name)
                    time.sleep(2)
            continue

        if use_copies:
            copy_path = make_bot_copy(bot_file)
            restrict_bot_copy_to_row_range(copy_path, extra_env)
            copies[bot_file] = copy_path
            print(f"[{ts()}] 🚀 {phase_prefix}{name} → {os.path.basename(copy_path)}")
            proc = run_bot(bot_file, excel_override=copy_path, extra_env=extra_env)
        else:
            print(f"[{ts()}] 🚀 {phase_prefix}{name}")
            proc = run_bot(bot_file, extra_env=extra_env)

        if proc:
            progress_start_bot(name)
            t = threading.Thread(target=stream_output, args=(name, proc, bot_file), daemon=True)
            t.start()
            procs[name] = (proc, t, bot_file)
            launched.append(name)
            time.sleep(2)
    return launched

def finalize_procs_and_filter_copies(procs, copies):
    """Wait/join all launched bot processes and remove failed copies."""
    for name, (proc, thread, bot_file) in list(procs.items()):
        proc.wait()
        thread.join(timeout=1)
        if thread.is_alive():
            cleanup_bot_edge_once(bot_file, proc)
            thread.join(timeout=5)
        if thread.is_alive():
            print(f"[{ts()}] [{name}] Bot đã exit; bỏ reader pipe còn bị Edge giữ, không chặn main.")
        if proc.returncode != 0:
            failed_copy = copies.pop(bot_file, None)
            if failed_copy and os.path.exists(failed_copy):
                try:
                    os.remove(failed_copy)
                except:
                    pass
            print(f"[{ts()}] ⚠️ [{name}] lỗi nên bỏ qua file copy, không gộp kết quả vào Excel gốc")
    return copies

def split_bots_two_rounds(bots_to_run):
    """Chia bot thành 2 lượt để giảm tải Edge/web."""
    bots = set(bots_to_run or [])
    first = bots & FIRST_ROUND_BOTS
    second = bots - first
    rounds = []
    if first:
        rounds.append(("1", first))
    if second:
        rounds.append(("2", second))
    return rounds

def run_bots_two_rounds_and_merge(bots_to_run, extra_env=None, merge_kwargs=None):
    """
    Chạy main theo 2 lượt như flow cũ: mỗi lượt chạy xong thì merge ngay.
    Trả về tổng số bot-copy đã được tạo/chạy thành công đủ điều kiện merge.
    """
    merge_kwargs = merge_kwargs or {}
    rounds = split_bots_two_rounds(bots_to_run)
    total_rounds = len(rounds)
    total_copies = 0

    for idx, (round_name, round_bots) in enumerate(rounds, start=1):
        print(f"\n{'='*55}")
        print(f"[{ts()}] 🚀 BẮT ĐẦU CHẠY LƯỢT {idx}/{total_rounds}: {describe_bots(round_bots)}")
        print(f"{'='*55}")

        copies = run_bots_parallel(round_bots, extra_env=extra_env)
        total_copies += len(copies)

        print()
        merge_copies_to_main(copies, **merge_kwargs)

        if idx < total_rounds:
            print(f"\n[{ts()}] ✅ Xong lượt {idx}/{total_rounds}. Nghỉ 10s rồi chạy lượt tiếp theo...")
            time.sleep(10)

    return total_copies

def run_bots_two_rounds_and_merge(bots_to_run, extra_env=None, merge_kwargs=None):
    """
    Overlap phase mode:
    - Start phase 1 first.
    - When phase 1 has completed 2 bots, start phase 2 immediately.
    - Merge all copies only after every bot in both phases has finished.
    """
    merge_kwargs = merge_kwargs or {}
    rounds = split_bots_two_rounds(bots_to_run)
    if not rounds:
        return 0

    if len(rounds) == 1:
        _, only_bots = rounds[0]
        print(f"\n{'='*55}")
        print(f"[{ts()}] 🚀 BẮT ĐẦU CHẠY LƯỢT 1/1: {describe_bots(only_bots)}")
        print(f"{'='*55}")
        copies = run_bots_parallel(only_bots, extra_env=extra_env)
        print()
        merge_copies_to_main(copies, **merge_kwargs)
        return len(copies)

    first_bots = rounds[0][1]
    second_bots = rounds[1][1]
    phase1_trigger = min(2, len(first_bots))

    procs = {}
    copies = {}
    phase1_names = set()
    phase2_started = False
    progress_stop = threading.Event()
    progress_thread = None

    print(f"\n{'='*55}")
    print(f"[{ts()}] 🚀 BẮT ĐẦU PHASE 1: {describe_bots(first_bots)}")
    print(f"[{ts()}] ⏩ Phase 2 sẽ tự mở khi phase 1 xong {phase1_trigger}/{len(first_bots)} bot")
    print(f"{'='*55}")

    phase1_names.update(
        launch_bots_async(first_bots, procs, copies, extra_env=extra_env, phase_label="P1")
    )

    if PROGRESS_ENABLED and procs:
        print(f"[{ts()}] Dashboard progress: {PROGRESS_DASHBOARD_PATH}")
        print(f"[{ts()}] Text progress:      {PROGRESS_TEXT_PATH}")
        print_progress_snapshot(force=True)
        progress_thread = threading.Thread(target=progress_monitor, args=(progress_stop,), daemon=True)
        progress_thread.start()

    while True:
        phase1_done = sum(
            1 for name in phase1_names
            if name in procs and procs[name][0].poll() is not None
        )

        if (not phase2_started) and phase1_done >= phase1_trigger:
            phase2_started = True
            print(f"\n{'='*55}")
            print(f"[{ts()}] ✅ Phase 1 đã xong {phase1_done}/{len(phase1_names)} bot → MỞ PHASE 2 NGAY")
            print(f"[{ts()}] 🚀 BẮT ĐẦU PHASE 2: {describe_bots(second_bots)}")
            print(f"{'='*55}")
            launch_bots_async(second_bots, procs, copies, extra_env=extra_env, phase_label="P2")

        if (not phase2_started) and not phase1_names:
            phase2_started = True
            launch_bots_async(second_bots, procs, copies, extra_env=extra_env, phase_label="P2")

        if phase2_started and procs and all(proc.poll() is not None for proc, _, _ in procs.values()):
            break

        time.sleep(2)

    print(f"\n[{ts()}] ✅ Tất cả bot phase 1 + phase 2 đã chạy xong. Đang cleanup và gộp Excel...")
    finalize_procs_and_filter_copies(procs, copies)

    if progress_thread:
        progress_stop.set()
        progress_thread.join(timeout=2)
        print_progress_snapshot(force=True)

    print()
    merge_copies_to_main(copies, **merge_kwargs)
    return len(copies)

def merge_copies_to_main(
    copies,
    filter_pol=None,
    filter_pod=None,
    filter_row=None,
    filter_row_from=None,
    filter_row_to=None,
):
    """
    Đọc kết quả từ các bản copy, ghi vào file gốc.
    Chỉ ghi các cột kết quả (F-P) của dòng mà bot xử lý (dựa vào carrier).

    filter_pol/filter_pod: chỉ gộp dòng có POL/POD khớp (cho Mode 2)
    filter_row: chỉ gộp đúng 1 dòng (cho Mode 3)
    filter_row_from/filter_row_to: chỉ gộp khoảng dòng inclusive
    """
    if not copies:
        return

    if filter_row:
        print(f"[{ts()}] 📊 Gộp kết quả dòng {filter_row} vào file gốc...")
    elif filter_row_from is not None and filter_row_to is not None:
        print(f"[{ts()}] 📊 Gộp kết quả dòng {filter_row_from}-{filter_row_to} vào file gốc...")
    elif filter_pol and filter_pod:
        print(f"[{ts()}] 📊 Gộp kết quả tuyến {filter_pol} → {filter_pod} vào file gốc...")
    else:
        print(f"[{ts()}] 📊 Gộp kết quả từ {len(copies)} bản copy vào file gốc...")

    wb_main = openpyxl.load_workbook(excel_path)
    ws_main = wb_main.active

    merge_started = time.time()
    total_merged = 0
    for bot_file, copy_path in copies.items():
        if not os.path.exists(copy_path):
            continue

        base_bot_file = str(bot_file).split("#", 1)[0]
        suffix = str(bot_file).split("#", 1)[1] if "#" in str(bot_file) else ""
        name = BOT_DISPLAY.get(base_bot_file, base_bot_file)
        if suffix:
            name = f"{name}-{suffix}"
        carrier_set = BOT_CARRIERS.get(base_bot_file, set())

        try:
            copy_started = time.time()
            wb_copy = openpyxl.load_workbook(copy_path, read_only=True, data_only=False)
            ws_copy = wb_copy.active

            count = 0
            max_col = max(max(RESULT_COLS), 5)
            if filter_row:
                start_row = end_row = int(filter_row)
            elif filter_row_from is not None and filter_row_to is not None:
                start_row = int(filter_row_from)
                end_row = min(int(filter_row_to), ws_copy.max_row)
            else:
                start_row = 2
                end_row = ws_copy.max_row

            for r, row_values in enumerate(
                ws_copy.iter_rows(
                    min_row=start_row,
                    max_row=end_row,
                    max_col=max_col,
                    values_only=True,
                ),
                start=start_row,
            ):
                # read_only workbooks are stream-based; iter_rows avoids very slow
                # random ws.cell(row, col) access when merging many bot copies.
                def value_at(col):
                    idx = col - 1
                    return row_values[idx] if idx < len(row_values) else None

                carrier = str(value_at(5) or "").strip().upper()
                if carrier not in carrier_set:
                    continue

                # Filter theo POL/POD (Mode 2)
                if filter_pol and filter_pod:
                    r_pol = str(value_at(3) or "").strip().upper()
                    r_pod = str(value_at(4) or "").strip().upper()
                    if r_pol != filter_pol or r_pod != filter_pod:
                        continue

                # Check xem có kết quả mới không (cột F không trống)
                val_f = value_at(6)
                if val_f is None:
                    continue

                # Copy các cột kết quả
                for col in RESULT_COLS:
                    val = value_at(col)
                    if val is not None:
                        ws_main.cell(row=r, column=col).value = val

                count += 1

            wb_copy.close()
            total_merged += count
            print(f"[{ts()}]    [{name}] Gộp {count} dòng ({time.time() - copy_started:.1f}s)")

        except Exception as e:
            print(f"[{ts()}]    [{name}] ❌ Lỗi đọc copy: {e}")

        # Xóa bản copy
        try:
            os.remove(copy_path)
        except:
            pass

    try:
        wb_main.save(excel_path)
        print(f"[{ts()}] ✅ Đã gộp {total_merged} dòng vào {os.path.basename(excel_path)} ({time.time() - merge_started:.1f}s)")
    except PermissionError:
        print(f"[{ts()}] ❌ Không ghi được — đóng file Excel rồi thử lại!")
        # Lưu tạm
        fallback = excel_path.replace(".xlsx", "_merged.xlsx")
        wb_main.save(fallback)
        print(f"[{ts()}] 💾 Đã lưu tạm: {os.path.basename(fallback)}")
    wb_main.close()

# ===================================================================================
# MODE 1: python main.py → chạy FULL Excel
# ===================================================================================
def mode_full(date_offset=DEFAULT_DATE_OFFSET_DAYS, skip_bots=None):
    print(f"[{ts()}] DATE_OFFSET_DAYS={date_offset}")
    print(f"[{ts()}] 📋 Chế độ: FULL — chạy toàn bộ bot trong 2 lượt")
    check_excel()

    carriers = get_all_carriers()
    bots = apply_skip_bots(ALL_BOTS, skip_bots, context="FULL")

    if not bots:
        print(f"[{ts()}] ⚠️ Không còn bot nào để chạy sau khi áp dụng --skip")
        return

    if carriers:
        print(f"[{ts()}] 📊 Carriers đang có trong Excel: {', '.join(sorted(carriers))}")
    print(f"[{ts()}] 🤖 Bots sẽ chạy: {describe_bots(bots)}")

    print(f"\n{'='*55}")
    print(f"[{ts()}] 🚀 BẮT ĐẦU CHẠY TOÀN BỘ BOT TRONG 2 LƯỢT")
    print(f"{'='*55}")
    run_bots_two_rounds_and_merge(
        bots,
        extra_env=build_date_env(date_offset),
    )

    print()
    print(f"\n[{ts()}] 🎉 HOÀN TẤT!")

# ===================================================================================
# MODE ROW RANGE: python main.py 100-150 --date +12
# ===================================================================================
def mode_row_range(row_from, row_to, date_offset=DEFAULT_DATE_OFFSET_DAYS, skip_bots=None):
    print(f"[{ts()}] DATE_OFFSET_DAYS={date_offset}")
    print(f"[{ts()}] 🎯 Chế độ: ROW RANGE — dòng {row_from}-{row_to} (inclusive)")
    check_excel()

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    max_row = ws.max_row

    if row_from < 2 or row_from > row_to or row_to > max_row:
        wb.close()
        print(f"[{ts()}] ❌ Khoảng dòng {row_from}-{row_to} không hợp lệ; phạm vi cho phép là 2-{max_row}")
        return

    carriers = set()
    populated_rows = 0
    for row in ws.iter_rows(
        min_row=row_from,
        max_row=row_to,
        min_col=5,
        max_col=5,
        values_only=True,
    ):
        carrier = str(row[0] or "").strip().upper()
        if carrier:
            populated_rows += 1
            carriers.add(carrier)
    wb.close()

    bots = get_bots_for_carriers(carriers)
    bots = apply_skip_bots(bots, skip_bots, context="ROW RANGE")
    if not bots:
        print(f"[{ts()}] ⚠️ Không có bot nào cần chạy trong dòng {row_from}-{row_to}")
        return

    print(f"[{ts()}] 📋 Có {populated_rows} dòng có carrier: {', '.join(sorted(carriers))}")
    print(f"[{ts()}] 🤖 Bots sẽ chạy: {describe_bots(bots)}")

    range_env = {
        "FILTER_ROW_FROM": str(row_from),
        "FILTER_ROW_TO": str(row_to),
        **build_date_env(date_offset),
    }
    if "bot_zim.py" in bots:
        range_env["ZIM_FORCE_DATE_OFFSET"] = "1"

    run_bots_two_rounds_and_merge(
        bots,
        extra_env=range_env,
        merge_kwargs={
            "filter_row_from": row_from,
            "filter_row_to": row_to,
        },
    )
    print(f"\n[{ts()}] 🎉 HOÀN TẤT dòng {row_from}-{row_to}!")

# ===================================================================================
# MODE 2: python main.py "HO CHI MINH" "CHENNAI" ["ĐẤT NƯỚC"] → toàn bộ bot check 1 tuyến
# ===================================================================================
def mode_route(pol, pod, country=None, date_offset=DEFAULT_DATE_OFFSET_DAYS, skip_bots=None):
    pol_upper = pol.strip().upper()
    pod_upper = pod.strip().upper()
    country_upper = country.strip().upper() if country else ""

    if country_upper:
        print(f"[{ts()}] 🗺️  Chế độ: ROUTE — {pol_upper} → {pod_upper} ({country_upper})")
    else:
        print(f"[{ts()}] 🗺️  Chế độ: ROUTE — {pol_upper} → {pod_upper}")
    print(f"[{ts()}] DATE_OFFSET_DAYS={date_offset}")
    check_excel()

    # Đảm bảo mỗi carrier chính có 1 row cho tuyến này
    main_carriers = filter_carriers_by_skip(MAIN_ROUTE_CARRIERS, skip_bots)
    if not main_carriers:
        print(f"[{ts()}] ⚠️ Không còn carrier nào để chạy sau khi áp dụng --skip")
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Scan existing rows + cập nhật country nếu cần
    existing_carriers = set()
    country_updated = 0
    for r in range(2, ws.max_row + 1):
        r_pol     = str(ws.cell(row=r, column=3).value or "").strip().upper()
        r_pod     = str(ws.cell(row=r, column=4).value or "").strip().upper()
        r_carrier = str(ws.cell(row=r, column=5).value or "").strip().upper()
        if r_pol == pol_upper and r_pod == pod_upper and r_carrier:
            existing_carriers.add(r_carrier)
            # Nếu có country mà cột B trống → điền country
            if country_upper:
                existing_country = str(ws.cell(row=r, column=2).value or "").strip()
                if not existing_country:
                    ws.cell(row=r, column=2).value = country_upper
                    country_updated += 1
    if country_updated > 0:
        print(f"[{ts()}] 📝 Đã cập nhật country '{country_upper}' cho {country_updated} dòng có sẵn")

    # Thêm row mới cho carrier chưa có
    next_row = ws.max_row + 1
    added = 0
    today_text = datetime.now().strftime("%d-%b%y")
    for carrier in main_carriers:
        if carrier not in existing_carriers:
            ws.cell(row=next_row, column=1).value = today_text
            if country_upper:
                ws.cell(row=next_row, column=2).value = country_upper
            ws.cell(row=next_row, column=3).value = pol_upper
            ws.cell(row=next_row, column=4).value = pod_upper
            ws.cell(row=next_row, column=5).value = carrier
            print(f"[{ts()}] ➕ Thêm dòng {next_row}: {pol_upper} → {pod_upper} [{carrier}]")
            next_row += 1
            added += 1

    if added > 0 or country_updated > 0:
        try:
            wb.save(excel_path)
            if added > 0:
                print(f"[{ts()}] 💾 Đã thêm {added} dòng mới")
            if added == 0:
                print(f"[{ts()}] ✅ Tất cả carrier đã có row cho tuyến này")
        except PermissionError:
            print(f"[{ts()}] ❌ Đóng file Excel trước khi chạy!")
            sys.exit(1)
    else:
        print(f"[{ts()}] ✅ Tất cả carrier đã có row cho tuyến này")
    wb.close()

    # Truyền filter POL/POD/COUNTRY cho các bot → bot chỉ check tuyến này
    route_env = {"FILTER_POL": pol_upper, "FILTER_POD": pod_upper, **build_date_env(date_offset)}
    if country_upper:
        route_env["FILTER_COUNTRY"] = country_upper

    bots = apply_skip_bots({CARRIER_TO_BOT[c] for c in main_carriers}, skip_bots, context="ROUTE")
    if not bots:
        print(f"[{ts()}] ⚠️ Không còn bot nào để chạy sau khi áp dụng --skip")
        return

    print(f"\n[{ts()}] 🚀 BẮT ĐẦU CHẠY 2 LƯỢT: {describe_bots(bots)}")
    run_bots_two_rounds_and_merge(
        bots,
        extra_env=route_env,
        merge_kwargs={"filter_pol": pol_upper, "filter_pod": pod_upper},
    )
    print(f"\n[{ts()}] 🎉 HOÀN TẤT tuyến {pol_upper} → {pod_upper}!")

# ===================================================================================
# MODE 3: python main.py 145 → chỉ chạy bot của carrier ở dòng 145
# ===================================================================================
def mode_single_row(row_num, date_offset=DEFAULT_DATE_OFFSET_DAYS, skip_bots=None):
    print(f"[{ts()}] DATE_OFFSET_DAYS={date_offset}")
    print(f"[{ts()}] 🎯 Chế độ: SINGLE ROW — dòng {row_num}")

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active
    max_row = ws.max_row

    if row_num < 2 or row_num > max_row:
        print(f"[{ts()}] ❌ Dòng {row_num} ngoài phạm vi (2-{max_row})")
        wb.close()
        return

    pol     = str(ws.cell(row=row_num, column=3).value or "").strip()
    pod     = str(ws.cell(row=row_num, column=4).value or "").strip()
    carrier = str(ws.cell(row=row_num, column=5).value or "").strip().upper()
    wb.close()

    if not carrier:
        print(f"[{ts()}] ❌ Dòng {row_num} không có carrier (cột E trống)")
        return
    if not pol or not pod:
        print(f"[{ts()}] ❌ Dòng {row_num} thiếu POL hoặc POD")
        return

    bot_file = CARRIER_TO_BOT.get(carrier)
    if not bot_file:
        print(f"[{ts()}] ❌ Carrier '{carrier}' không có bot")
        return
    if bot_file in set(skip_bots or []):
        print(f"[{ts()}] ⏭️ Dòng {row_num} thuộc {BOT_DISPLAY.get(bot_file, bot_file)} nhưng bot này đang bị --skip, bỏ qua.")
        return

    name = BOT_DISPLAY.get(bot_file, bot_file)
    print(f"[{ts()}] 📋 Dòng {row_num}: {pol} → {pod} [{carrier}] → {name}")

    # Dùng bản copy để an toàn, sau đó chỉ gộp đúng dòng row_num
    copy_path = make_bot_copy(bot_file)
    print(f"[{ts()}] 🚀 {name} → {os.path.basename(copy_path)}")
    proc = run_bot(bot_file, excel_override=copy_path, extra_env={"SINGLE_ROW": str(row_num), **build_date_env(date_offset)})
    if not proc:
        return

    print(f"\n[{ts()}] ⏳ Đang chờ {name}...\n")

    # Stream output
    progress_start_bot(name)
    if PROGRESS_ENABLED:
        # print(f"[{ts()}] Dashboard progress: {PROGRESS_DASHBOARD_PATH}")
        print_progress_snapshot(force=True)
    t = threading.Thread(target=stream_output, args=(name, proc, bot_file), daemon=True)
    t.start()
    proc.wait()
    t.join(timeout=1)
    if t.is_alive():
        cleanup_bot_edge_once(bot_file, proc)
        t.join(timeout=5)
    if t.is_alive():
        print(f"[{ts()}] [{name}] Bot đã exit; bỏ reader pipe còn bị Edge giữ, không chặn main.")
    if proc.returncode != 0:
        try:
            os.remove(copy_path)
        except:
            pass
        print(f"[{ts()}] ❌ {name} lỗi, không gộp kết quả dòng {row_num}")
        return

    # Gộp chỉ dòng row_num
    print()
    merge_copies_to_main({bot_file: copy_path}, filter_row=row_num)
    print(f"\n[{ts()}] 🎉 HOÀN TẤT dòng {row_num}!")

def mode_carrier(carrier, date_offset=DEFAULT_DATE_OFFSET_DAYS, skip_bots=None):
    carrier_upper = carrier.strip().upper()
    print(f"[{ts()}] DATE_OFFSET_DAYS={date_offset}")
    print(f"[{ts()}] Mode: SINGLE CARRIER - {carrier_upper}")
    check_excel()

    bot_file = CARRIER_TO_BOT.get(carrier_upper)
    if not bot_file:
        print(f"[{ts()}] Carrier '{carrier_upper}' khong co bot")
        return
    if bot_file in set(skip_bots or []):
        print(f"[{ts()}] ⏭️ Carrier {carrier_upper} thuộc {BOT_DISPLAY.get(bot_file, bot_file)} nhưng bot này đang bị --skip, bỏ qua.")
        return

    name = BOT_DISPLAY.get(bot_file, bot_file)
    if bot_file == "bot_oocl.py" and OOCL_PARALLEL_WORKERS > 1:
        extra_env = build_date_env(date_offset)
        print(f"[{ts()}] RUN {name} bằng {OOCL_PARALLEL_WORKERS} worker song song")
        copies = run_bots_parallel({bot_file}, extra_env=extra_env)
        print()
        if copies:
            merge_copies_to_main(copies)
        print(f"\n[{ts()}] HOAN TAT carrier {carrier_upper}!")
        return

    copy_path = make_bot_copy(bot_file)
    print(f"[{ts()}] RUN {name} -> {os.path.basename(copy_path)}")

    extra_env = build_date_env(date_offset)
    if bot_file == "bot_zim.py":
        extra_env["ZIM_FORCE_DATE_OFFSET"] = "1"

    proc = run_bot(bot_file, excel_override=copy_path, extra_env=extra_env)
    if not proc:
        return

    print(f"\n[{ts()}] Dang cho {name}...\n")
    progress_start_bot(name)
    if PROGRESS_ENABLED:
        # print(f"[{ts()}] Dashboard progress: {PROGRESS_DASHBOARD_PATH}")
        print_progress_snapshot(force=True)
    t = threading.Thread(target=stream_output, args=(name, proc, bot_file), daemon=True)
    t.start()
    proc.wait()
    t.join(timeout=1)
    if t.is_alive():
        cleanup_bot_edge_once(bot_file, proc)
        t.join(timeout=5)
    if t.is_alive():
        print(f"[{ts()}] [{name}] Bot đã exit; bỏ reader pipe còn bị Edge giữ, không chặn main.")

    if proc.returncode != 0:
        try:
            os.remove(copy_path)
        except:
            pass
        print(f"[{ts()}] {name} loi, khong gop ket qua")
        return

    print()
    merge_copies_to_main({bot_file: copy_path})
    print(f"\n[{ts()}] HOAN TAT carrier {carrier_upper}!")

def mode_bots(bot_files, date_offset=DEFAULT_DATE_OFFSET_DAYS, skip_bots=None):
    bot_files = set(bot_files or [])
    bot_files = apply_skip_bots(bot_files, skip_bots, context="multi-bot")
    print(f"[{ts()}] DATE_OFFSET_DAYS={date_offset}")
    print(f"[{ts()}] Mode: MULTI BOT - {describe_bots(bot_files)}")
    check_excel()

    if not bot_files:
        print(f"[{ts()}] Khong co bot nao de chay.")
        return

    extra_env = build_date_env(date_offset)
    if "bot_zim.py" in bot_files:
        extra_env["ZIM_FORCE_DATE_OFFSET"] = "1"

    print(f"\n[{ts()}] RUN MULTI BOT 1 LUOT: {describe_bots(bot_files)}")
    copies = run_bots_parallel(bot_files, extra_env=extra_env)
    total = len(copies)
    print()
    if copies:
        merge_copies_to_main(copies)
    print(f"\n[{ts()}] HOAN TAT multi-bot: {describe_bots(bot_files)} ({total} file copy)")



import psutil
import msvcrt

_is_paused = False
def _pause_listener_thread():
    global _is_paused
    while True:
        try:
            # OOCL's child process owns console input while waiting for OTP.
            # Do not consume ENTER (or any other key) from the parent process.
            if any(os.path.exists(path) for path in MANUAL_INPUT_WAIT_PATHS):
                time.sleep(0.1)
                continue
            if msvcrt.kbhit():
                key = msvcrt.getch()
                if key.lower() == b'p':
                    _is_paused = not _is_paused
                    try:
                        current_proc = psutil.Process(os.getpid())
                        children = current_proc.children(recursive=True)
                        if _is_paused:
                            print("\n[PAUSE] === DÃ TẠM DỪNG CÁC BOT PYTHON === (Trình duyệt vẫn hoạt động. Bấm 'P' để tiếp tục)")
                            for child in children:
                                try:
                                    if "python" in child.name().lower():
                                        child.suspend()
                                except: pass
                        else:
                            print("\n[RESUME] === DÃ TIẾP TỤC CHẠY CÁC BOT ===")
                            for child in children:
                                try:
                                    if "python" in child.name().lower():
                                        child.resume()
                                except: pass
                    except Exception as e:
                        print(f"\n[PAUSE ERROR] Lỗi khi pause/resume: {e}")
        except Exception:
            pass
        time.sleep(0.1)

# ===================================================================================
# ENTRY POINT
# ===================================================================================
if __name__ == "__main__":
    threading.Thread(target=_pause_listener_thread, daemon=True).start()

    acquire_main_session_lock()
    print_header()

    args, date_offset, skip_bots, selected_bots = parse_cli_args(sys.argv[1:])
    if skip_bots:
        print(f"[{ts()}] ⏭️ Danh sách bot bị skip: {describe_bots(skip_bots)}")
    ran_work = False

    try:
        positional_bots = set()
        positional_unknown = []
        if len(args) >= 2:
            positional_bots, positional_unknown = parse_bot_names(args, warn=False)

        if selected_bots:
            if args:
                print(f"[{ts()}] --bots/--only se chay bot tren file Excel hien tai; bo qua positional: {' '.join(args)}")
            ran_work = True
            mode_bots(selected_bots, date_offset=date_offset, skip_bots=skip_bots)

        elif len(args) >= 2 and positional_bots and not positional_unknown:
            ran_work = True
            mode_bots(positional_bots, date_offset=date_offset, skip_bots=skip_bots)

        elif len(args) == 0:
            ran_work = True
            mode_full(date_offset=date_offset, skip_bots=skip_bots)

        elif len(args) == 1:
            row_range = parse_row_range_token(args[0])
            if row_range:
                ran_work = True
                mode_row_range(
                    row_range[0],
                    row_range[1],
                    date_offset=date_offset,
                    skip_bots=skip_bots,
                )
            else:
                try:
                    row_num = int(args[0])
                    ran_work = True
                    mode_single_row(row_num, date_offset=date_offset, skip_bots=skip_bots)
                except ValueError:
                    carrier = args[0].strip().upper()
                    if CARRIER_TO_BOT.get(carrier):
                        ran_work = True
                        mode_carrier(carrier, date_offset=date_offset, skip_bots=skip_bots)
                        sys.exit(0)
                    print(f"[{ts()}] ❌ Tham số không hợp lệ: '{args[0]}'")
                    print(f"  Dùng: python main.py 145           (chạy 1 dòng)")
                    print(f"  Dùng: python main.py 100-150       (chạy một khoảng dòng)")
                    print(f"  Hoặc: python main.py \"POL\" \"POD\"   (chạy 1 tuyến)")
                    sys.exit(1)

        elif len(args) == 2:
            pol, pod = args[0], args[1]
            ran_work = True
            mode_route(pol, pod, date_offset=date_offset, skip_bots=skip_bots)

        elif len(args) == 3:
            pol, pod, country = args[0], args[1], args[2]
            ran_work = True
            mode_route(pol, pod, country=country, date_offset=date_offset, skip_bots=skip_bots)

        else:
            print(f"[{ts()}] ❌ Sai cú pháp!")
            print(f"  python main.py                                    → chạy full Excel")
            print(f"  python main.py \"HO CHI MINH\" \"CHENNAI\"            → check 1 tuyến")
            print(f"  python main.py \"HO CHI MINH\" \"HAMBURG\" \"GERMANY\"  → check 1 tuyến + quốc gia")
            print(f"  python main.py 145                                → check 1 dòng")
            print(f"  python main.py 100-150 --date +12                 → check một khoảng dòng")
            sys.exit(1)
    finally:
        if ran_work:
            cleanup_bot_edges_after_main()
            print(f"[{ts()}] ✅ Main kết thúc; đã đóng browser các bot xong, giữ COSCO + HAPAG LLOYD + OOCL.")
