"""
BOT YANG MING — Lấy lịch tàu YANG MING (https://www.yangming.com/esolution/schedule/point_to_point_search)
====================================================================================================
Hoạt động:
- Đọc Excel input_gia.xlsx (hoặc env EXCEL_PATH)
- Lọc rows có CARRIER == "YANG MING"
- Với mỗi row: nhập POL/POD (qua react-select), set Start/End Date theo valid (cột K),
  bấm Search, parse các card kết quả, áp dụng quy tắc tối đa 3 ETD cách nhau ≥ 1 ngày
  (cùng cách OOCL chọn), rồi ghi vào Excel cột I (9), J (10), O (15), P (16).
- Lịch chỉ giữ ETD ≤ valid date.

QUY TẮC VALID WINDOW:
  - Nếu valid ≤ ngày 15 trong tháng → start = ngày 1 cùng tháng, end = valid
  - Nếu valid > ngày 15 → start = ngày 15 cùng tháng, end = valid

PORT ALIASES (Excel POL → Yang Ming search keyword):
  HO CHI MINH → CAT LAI
  HAI PHONG   → HAIPHONG
  AQABA       → AL'AQABAH
"""

import os
import re
import sys
import time
import json
import socket
import subprocess
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils.datetime import from_excel

from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only

# ===================================================================================
# CONFIG
# ===================================================================================
EXCEL_PATH = os.environ.get("EXCEL_PATH") or os.path.join(os.getcwd(), "input_gia.xlsx")
FILTER_POL = (os.environ.get("FILTER_POL") or "").strip().upper()
FILTER_POD = (os.environ.get("FILTER_POD") or "").strip().upper()
SINGLE_ROW = (os.environ.get("SINGLE_ROW") or "").strip()
DATE_OFFSET_DAYS = parse_date_offset_days()

CARRIER_TARGETS = {"YANG MING", "YANGMING", "YM"}
DEBUG_PORT     = "9528"
DRIVER_PATH    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "msedgedriver.exe")
URL_YM         = "https://www.yangming.com/esolution/schedule/point_to_point_search"
EDGE_EXE       = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
EDGE_PROFILE   = r"C:\edge_yangming"

# Excel POL/POD → từ khoá search trên Yang Ming
YM_PORT_ALIASES = {
    "HO CHI MINH":  "CAT LAI",
    "HOCHIMINH":    "CAT LAI",
    "HCM":          "CAT LAI",
    "SAIGON":       "CAT LAI",
    "HAI PHONG":    "HAIPHONG",
    "HAIPHONG":     "HAIPHONG",
    "AQABA":        "AL'AQABAH",
    "TIANJIN":      "XINGANG",
    "ANTWERP":     "ANTWERPEN",
    "NAPLES":       "NAPOLI",
    "TIENTSIN":     "XINGANG",
}

YM_HCM_AU_ALLOWED_TS_PORTS = {"TWKHH", "SGSIN"}
YM_AUSTRALIA_PODS = {
    "ADELAIDE", "BRISBANE", "FREMANTLE", "MELBOURNE", "SYDNEY",
}

# Năm hiện tại (dùng khi parse valid date không có năm)
TODAY = datetime.now()

# ===================================================================================
# DRIVER
# ===================================================================================
def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(("127.0.0.1", int(port))) == 0


def ensure_edge_debug_port():
    if is_port_in_use(DEBUG_PORT):
        print(f"[YM] Edge debug port {DEBUG_PORT} đã mở sẵn.")
        return

    print(f"[YM] Edge debug port {DEBUG_PORT} chưa mở, tự khởi động Edge...")
    subprocess.Popen([
        EDGE_EXE,
        f"--remote-debugging-port={DEBUG_PORT}",
        f"--user-data-dir={EDGE_PROFILE}",
        "--disable-background-timer-throttling",
        "--disable-renderer-backgrounding",
        "--disable-backgrounding-occluded-windows",
        "--start-maximized",
    ])
    # FIX: bỏ time.sleep(1) sau khi port đã mở — port mở = Edge ready
    for _ in range(20):
        if is_port_in_use(DEBUG_PORT):
            return
        time.sleep(0.5)
    raise RuntimeError(f"Không mở được Edge debug port {DEBUG_PORT}")


ensure_edge_debug_port()
edge_options = Options()
edge_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{DEBUG_PORT}")
driver = webdriver.Edge(service=Service(executable_path=DRIVER_PATH), options=edge_options)
try:
    driver.maximize_window()
except Exception:
    try:
        info = driver.execute_cdp_cmd("Browser.getWindowForTarget", {})
        driver.execute_cdp_cmd("Browser.setWindowBounds", {
            "windowId": info["windowId"],
            "bounds": {"windowState": "maximized"}
        })
    except Exception:
        pass
WAIT = WebDriverWait(driver, 15)

print(f"[YM] ✅ Đã attach Edge debug port {DEBUG_PORT}")

# ===================================================================================
# HELPERS
# ===================================================================================
def log(msg):
    print(f"[YM] {msg}")

def ensure_live_window():
    global driver, WAIT
    try:
        handles = driver.window_handles
        if handles:
            try:
                driver.current_url
            except Exception:
                driver.switch_to.window(handles[0])
            return
        driver.execute_cdp_cmd("Target.createTarget", {"url": "about:blank"})
        time.sleep(0.5)
        driver.switch_to.window(driver.window_handles[-1])
    except Exception:
        driver = webdriver.Edge(service=Service(executable_path=DRIVER_PATH), options=edge_options)
        try:
            driver.maximize_window()
        except Exception:
            pass
        WAIT = WebDriverWait(driver, 15)
        try:
            handles = driver.window_handles
            if handles:
                driver.switch_to.window(handles[0])
        except Exception:
            pass

def safe_click(el):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.15)
        el.click()
    except Exception:
        driver.execute_script("arguments[0].click();", el)

def alias_port(name):
    key = (name or "").strip().upper()
    return YM_PORT_ALIASES.get(key, name)

# ===================================================================================
# VALID DATE PARSER
# ===================================================================================
MONTHS_VN = {
    "JAN":1, "FEB":2, "MAR":3, "APR":4, "MAY":5, "JUN":6,
    "JUL":7, "AUG":8, "SEP":9, "OCT":10, "NOV":11, "DEC":12,
}

def parse_valid_date(raw):
    """
    Nhận giá trị valid từ cột K:
      - datetime/date → trả lại date
      - "15-Jun" / "15/06" / "15/06/2026" / "15-Jun-2026" → datetime
    Trả về datetime (year mặc định = năm hiện tại nếu thiếu).
    """
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if hasattr(raw, "year") and hasattr(raw, "month") and hasattr(raw, "day"):
        return datetime(raw.year, raw.month, raw.day)
    s = str(raw).strip()
    if not s:
        return None
    if re.fullmatch(r"\d+(\.\d+)?", s):
        try:
            dt = from_excel(float(s))
            return datetime(dt.year, dt.month, dt.day)
        except Exception:
            pass
    m = re.fullmatch(r"(\d{1,2})-([A-Za-z]{3})", s)
    if m:
        mon = MONTHS_VN.get(m.group(2).upper())
        if mon:
            return datetime(TODAY.year, mon, int(m.group(1)))
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})", s)
    if m:
        return datetime(TODAY.year, int(m.group(2)), int(m.group(1)))
    # Try formats
    fmts = ["%d-%b-%Y", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d",
            "%d %b %Y", "%Y-%m-%d", "%d-%m-%Y"]
    for f in fmts:
        try:
            dt = datetime.strptime(s, f)
            if dt.year == 1900:
                dt = dt.replace(year=TODAY.year)
            return dt
        except Exception:
            continue
    # Regex fallback "DD-Mon"
    m = re.match(r"(\d{1,2})[-/\s]+([A-Za-z]{3})", s)
    if m:
        day = int(m.group(1))
        mon = MONTHS_VN.get(m.group(2).upper())
        if mon:
            return datetime(TODAY.year, mon, day)
    log(f"   ⚠️ Không parse được valid='{s}'")
    return None

def compute_valid_window(valid_dt):
    """
    ETD luôn bắt đầu từ hôm nay + DATE_OFFSET_DAYS.
    Valid trong Excel chỉ đóng vai trò ngày ETD tối đa.
    Trả về (start_dt, end_dt) ở format datetime.
    """
    if valid_dt is None:
        return None, None
    start = (TODAY + timedelta(days=DATE_OFFSET_DAYS)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    end = valid_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    return start, end

def compute_fallback_valid_window(valid_dt):
    """
    Fallback giống MSC/OOCL:
    nếu không có ETD đạt mốc today + DATE_OFFSET_DAYS thì search lại từ hôm nay
    tới ngày valid, rồi lấy ETD xa nhất còn nằm trong vùng valid.
    """
    if valid_dt is None:
        return None, None
    start = TODAY.replace(hour=0, minute=0, second=0, microsecond=0)
    end = valid_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    return start, end

def fmt_ym_date(dt):
    """Yang Ming nhận format YYYY/MM/DD"""
    return dt.strftime("%Y/%m/%d")

# ===================================================================================
# OPEN PAGE + ACCEPT COOKIES
# ===================================================================================
def open_search_page():
    ensure_live_window()
    if "schedule/point_to_point_search" not in driver.current_url:
        driver.get(URL_YM)
    else:
        # Reset form bằng cách reload
        driver.get(URL_YM)
    WAIT.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[id^=react-select][id$=-input]")))
    time.sleep(1.0)

    # Bấm "Agree" cookie nếu xuất hiện
    try:
        agree = driver.find_elements(By.XPATH, "//button[normalize-space()='Agree']")
        for b in agree:
            if b.is_displayed():
                safe_click(b)
                time.sleep(0.5)
                break
    except Exception:
        pass

    # Đảm bảo đang ở tab Point-to-Point Search (tab đầu)
    try:
        tab = driver.find_element(By.XPATH, "//button[@role='tab' and normalize-space()='Point-to-Point Search']")
        if "selected" not in (tab.get_attribute("aria-selected") or "").lower():
            safe_click(tab)
            time.sleep(0.5)
    except Exception:
        pass

# ===================================================================================
# CHỌN POL / POD QUA REACT-SELECT
# ===================================================================================
def select_port(index, keyword):
    """index 0 = From, index 1 = To. Trả về True nếu chọn được."""
    inputs = driver.find_elements(By.CSS_SELECTOR, "input[id^=react-select][id$=-input]")
    if len(inputs) < 2:
        raise Exception("Không tìm thấy 2 react-select cho POL/POD")
    inp = inputs[index]
    safe_click(inp)
    inp.send_keys(Keys.CONTROL + "a")
    inp.send_keys(Keys.DELETE)
    time.sleep(0.2)
    inp.send_keys(keyword)
    time.sleep(1.5)
    try:
        WebDriverWait(driver, 6).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='option']"))
        )
    except TimeoutException:
        log(f"   ⚠️ Không có option cho '{keyword}'")
        return False
    options = driver.find_elements(By.CSS_SELECTOR, "div[role='option']")
    if not options:
        return False
    # Ưu tiên option có chứa keyword
    chosen = None
    for o in options:
        if keyword.upper() in (o.text or "").upper():
            chosen = o
            break
    chosen = chosen or options[0]
    log(f"   ✅ Port [{index}]: '{keyword}' → '{(chosen.text or '').strip()}'")
    safe_click(chosen)
    time.sleep(0.4)
    return True

# ===================================================================================
# SET START/END DATE (input.w-24)
# ===================================================================================
def set_dates(start_dt, end_dt):
    """Set 2 input date (start + end) theo format YYYY/MM/DD."""
    date_inputs = driver.find_elements(By.CSS_SELECTOR, "input.w-24")
    if len(date_inputs) < 2:
        raise Exception("Không tìm thấy 2 input ngày start/end")
    for idx, dt in [(0, start_dt), (1, end_dt)]:
        el = date_inputs[idx]
        safe_click(el)
        # Set value qua JS (an toàn vì react điều khiển)
        driver.execute_script("""
            const el = arguments[0]; const v = arguments[1];
            const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(el, v);
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
            el.dispatchEvent(new Event('blur', {bubbles: true}));
        """, el, fmt_ym_date(dt))
        time.sleep(0.3)
        # Đóng datepicker nếu nó mở
        try:
            driver.find_element(By.TAG_NAME, "body").click()
        except Exception:
            pass

    # Đảm bảo radio "Departure Date" được chọn (cột Period)
    try:
        dep_radio_label = driver.find_element(By.XPATH, "//label[contains(., 'Departure Date')]")
        safe_click(dep_radio_label)
        time.sleep(0.2)
    except Exception:
        pass

    log(f"   📅 Set dates: {fmt_ym_date(start_dt)} → {fmt_ym_date(end_dt)}")

# ===================================================================================
# BẤM SEARCH + ĐỢI KẾT QUẢ
# ===================================================================================
def set_end_date_only(end_dt):
    """YM: only set To/End Date = valid; clear From/Start Date."""
    date_inputs = driver.find_elements(By.CSS_SELECTOR, "input.w-24")
    if len(date_inputs) < 2:
        raise Exception("Khong tim thay 2 input ngay start/end")

    try:
        start_el = date_inputs[0]
        safe_click(start_el)
        driver.execute_script("""
            const el = arguments[0];
            const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(el, '');
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
            el.dispatchEvent(new Event('blur', {bubbles: true}));
        """, start_el)
        time.sleep(0.2)
    except Exception:
        pass

    end_el = date_inputs[1]
    safe_click(end_el)
    driver.execute_script("""
        const el = arguments[0]; const v = arguments[1];
        const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        setter.call(el, v);
        el.dispatchEvent(new Event('input', {bubbles: true}));
        el.dispatchEvent(new Event('change', {bubbles: true}));
        el.dispatchEvent(new Event('blur', {bubbles: true}));
    """, end_el, fmt_ym_date(end_dt))
    time.sleep(0.3)

    try:
        driver.find_element(By.TAG_NAME, "body").click()
    except Exception:
        pass

    try:
        dep_radio_label = driver.find_element(By.XPATH, "//label[contains(., 'Departure Date')]")
        safe_click(dep_radio_label)
        time.sleep(0.2)
    except Exception:
        pass

    log(f"   [YM] Set End Date only: To={fmt_ym_date(end_dt)} | From blank")

def click_search():
    """Bấm Search ở dưới form. Có nhiều nút 'Search' (tab Vessel Schedule),
    nên ưu tiên nút trong section Point-to-Point Search."""
    btns = driver.find_elements(By.XPATH, "//button[normalize-space()='Search']")
    # Lọc nút visible
    btn = None
    for b in btns:
        try:
            if b.is_displayed():
                btn = b
                break
        except Exception:
            continue
    if not btn:
        raise Exception("Không tìm thấy nút Search")
    safe_click(btn)
    log("   🔍 Đã bấm Search")

def wait_results(timeout=25):
    """Đợi kết quả có (có 'Days' card) hoặc thông báo 'No matching'."""
    end = time.time() + timeout
    while time.time() < end:
        try:
            body = driver.find_element(By.TAG_NAME, "body").text or ""
            if re.search(r"\d+\s+Days", body) and "VNCLI" in body.upper() + " " or re.search(r"VOYAGE CODE", body, re.I):
                return True
            if re.search(r"(no\s+matching|no\s+result|no\s+schedule)", body, re.I):
                return False
        except Exception:
            pass
        time.sleep(0.5)
    # Best-effort fallback
    return True

# ===================================================================================
# PARSE RESULT CARDS
# ===================================================================================
def parse_result_cards():
    """
    Mỗi card YM có dạng (innerText):
        7 Days
        CAT LAI
        VNCLI (CY)
        2026/06/06
        TWKHH               ← (optional) transshipment port code
        SHANGHAI
        CNSHA (CY)
        2026/06/13
        Voyage Code
        THX2621B
        Vsl Name - Com.voy
        EVER OMNI-056B
    Trả về list dict { etd_dt, eta_dt, tt_days, pol, pod, ts_port, voyage, vessel }
    """
    cards = driver.find_elements(By.CSS_SELECTOR, "div.my-5")
    results = []
    for c in cards:
        try:
            txt = (c.text or "").strip()
            if not txt or "Days" not in txt:
                continue
            lines = [ln.strip() for ln in txt.split("\n") if ln.strip()]
            if len(lines) < 8:
                continue
            m_tt = re.match(r"(\d+)\s+Days?", lines[0])
            if not m_tt:
                continue
            tt_days = int(m_tt.group(1))
            # POL block: lines[1] = port name, lines[2] = code (CY), lines[3] = etd
            pol = lines[1]
            etd_str = None
            ts_port = "DIRECT"
            etd_idx = None
            for i, ln in enumerate(lines[2:6], start=2):
                if re.match(r"\d{4}/\d{2}/\d{2}", ln):
                    etd_str = ln
                    etd_idx = i
                    break
            if not etd_str:
                continue
            etd_dt = datetime.strptime(etd_str, "%Y/%m/%d")

            # Sau ETD co the la TS port code/list (VD: TWKHH hoac SGSIN,ITSPE)
            # hoac truc tiep POD. Neu khong skip list nay thi filter POD se rot.
            next_idx = etd_idx + 1
            if next_idx < len(lines) and re.match(r"^[A-Z0-9]{5}(?:\s*,\s*[A-Z0-9]{5})*$", lines[next_idx]):
                ts_port = " + ".join(part.strip() for part in lines[next_idx].split(",") if part.strip())
                next_idx += 1

            pod = lines[next_idx] if next_idx < len(lines) else ""
            # tìm ETA: dòng đầu tiên sau pod khớp date pattern
            eta_dt = None
            for ln in lines[next_idx:]:
                if re.match(r"\d{4}/\d{2}/\d{2}", ln):
                    eta_dt = datetime.strptime(ln, "%Y/%m/%d")
                    break

            # Voyage code và vessel name
            voyage = ""
            vessel = ""
            for i, ln in enumerate(lines):
                if ln.lower().startswith("voyage code") and i + 1 < len(lines):
                    voyage = lines[i + 1]
                if ln.lower().startswith("vsl name") and i + 1 < len(lines):
                    vessel = lines[i + 1]

            results.append({
                "etd_dt":  etd_dt,
                "eta_dt":  eta_dt,
                "tt_days": tt_days,
                "pol":     pol,
                "pod":     pod,
                "ts_port": ts_port,
                "voyage":  voyage,
                "vessel":  vessel or "TBA",
            })
        except Exception as e:
            log(f"   ⚠️ Lỗi parse card: {e}")
            continue
    log(f"   📦 Parsed {len(results)} cards")
    return results

# ===================================================================================
# ETD RULES — tối đa 3 ETD, cách nhau ≥ 1 ngày, ETD ≤ valid
# ===================================================================================
def apply_etd_rules(entries, valid_dt):
    min_etd, _ = compute_valid_window(valid_dt)
    future = [
        e for e in entries
        if e["etd_dt"] >= min_etd and etd_within_max(e["etd_dt"]) and (valid_dt is None or e["etd_dt"] <= valid_dt)
    ]
    if not future:
        return []
    future.sort(key=lambda e: e["etd_dt"])
    selected = [future[0]]
    for e in future[1:]:
        if len(selected) >= 3:
            break
        if (e["etd_dt"] - selected[-1]["etd_dt"]).days < 1:
            continue
        selected.append(e)
    return selected

def apply_etd_fallback_rules(entries, valid_dt):
    """
    Khi không có ETD chuẩn >= today + DATE_OFFSET_DAYS:
    lấy đúng 1 ETD xa nhất trong khoảng today..valid.
    """
    min_etd, _ = compute_fallback_valid_window(valid_dt)
    future = [
        e for e in entries
        if e["etd_dt"] >= min_etd and etd_within_max(e["etd_dt"]) and (valid_dt is None or e["etd_dt"] <= valid_dt)
    ]
    if not future:
        return []
    future.sort(key=lambda e: (e["etd_dt"], -(e.get("tt_days") or 9999)), reverse=True)
    return [future[0]]

# ===================================================================================
# FORMAT ETD / TT / VESSEL / TS
# ===================================================================================
def format_etd_text(entries):
    if not entries:
        return ""
    fmt = [(e["etd_dt"].day, e["etd_dt"].strftime("%b")) for e in entries]
    if len(fmt) == 1:
        return f"{fmt[0][0]}-{fmt[0][1]}"
    if len(fmt) == 2:
        return f"{fmt[0][0]}-{fmt[0][1]} & {fmt[1][0]}-{fmt[1][1]}"
    months = [f[1] for f in fmt]
    if months[0] == months[1] == months[2]:
        return f"{fmt[0][0]}, {fmt[1][0]}, {fmt[2][0]}-{fmt[2][1]}"
    if months[0] == months[1]:
        return f"{fmt[0][0]}, {fmt[1][0]}-{fmt[1][1]}, {fmt[2][0]}-{fmt[2][1]}"
    return f"{fmt[0][0]}-{fmt[0][1]}, {fmt[1][0]}-{fmt[1][1]}, {fmt[2][0]}-{fmt[2][1]}"

def format_tt_text(entries):
    if not entries:
        return ""
    tts = [e["tt_days"] for e in entries]
    if len(set(tts)) == 1:
        return f"{tts[0]}"
    return f"{min(tts)}-{max(tts)}"

def format_vessel_block(entries):
    lines = []
    ts_seen = []
    for e in entries:
        vn = e["vessel"] or "TBA"
        vc = e["voyage"] or ""
        v_str = f"{vn}" + (f" ({vc})" if vc else "")
        ts = e["ts_port"] or "DIRECT"
        lines.append(
            f"{v_str} / ETD: {e['etd_dt'].day}-{e['etd_dt'].strftime('%b')}"
            f" / Transit time: {e['tt_days']} Days / Transshipment Port: {ts}"
        )
        if ts not in ts_seen:
            ts_seen.append(ts)
    return "\n".join(lines), " or\n".join(ts_seen) if ts_seen else "DIRECT"

# ===================================================================================
# FILTER kết quả theo POL/POD đã chọn
# ===================================================================================
def filter_by_route(entries, pol_keyword, pod_keyword):
    """Chỉ giữ chuyến mà pol/pod match keyword đã search."""
    p1 = (pol_keyword or "").upper()
    p2 = (pod_keyword or "").upper()
    out = []
    for e in entries:
        ok_pol = (not p1) or (p1 in (e["pol"] or "").upper())
        ok_pod = (not p2) or (p2 in (e["pod"] or "").upper())
        if ok_pol and ok_pod:
            out.append(e)
    return out


def is_hcm_to_australia(pol_excel, pod_excel, country_excel=""):
    pol = re.sub(r"\s+", " ", str(pol_excel or "").upper()).strip()
    pod = re.sub(r"\s+", " ", str(pod_excel or "").upper()).strip()
    country = re.sub(r"\s+", " ", str(country_excel or "").upper()).strip()
    is_hcm = pol in {"HO CHI MINH", "HOCHIMINH", "HCM", "SAIGON", "CAT LAI"}
    is_australia = country in {"AUSTRALIA", "AU", "AUS"} or pod in YM_AUSTRALIA_PODS
    return is_hcm and is_australia


def filter_hcm_australia_transshipment(entries, pol_excel, pod_excel, country_excel=""):
    """
    HCM -> Australia:
      - keep DIRECT sailings;
      - for transshipment sailings, every TS code must be TWKHH or SGSIN.
    """
    if not is_hcm_to_australia(pol_excel, pod_excel, country_excel):
        return entries, []

    kept = []
    rejected = []
    for entry in entries:
        raw_ts = str(entry.get("ts_port") or "DIRECT").strip().upper()
        if not raw_ts or raw_ts == "DIRECT":
            kept.append(entry)
            continue

        ts_codes = set(re.findall(r"\b[A-Z]{2}[A-Z0-9]{3}\b", raw_ts))
        if ts_codes and ts_codes.issubset(YM_HCM_AU_ALLOWED_TS_PORTS):
            kept.append(entry)
        else:
            rejected.append(entry)
    return kept, rejected

# ===================================================================================
# WRITE EXCEL ROW
# ===================================================================================
def write_excel_row(row_i, etd_text, tt_text, vessel_text, ts_text, error=None, valid_dt=None):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        if ws.cell(row=row_i, column=6).value is None:
            ws.cell(row=row_i, column=6).value = "-"
        if valid_dt and not ws.cell(row=row_i, column=11).value:
            ws.cell(row=row_i, column=11).value = f"{valid_dt.day}-{valid_dt.strftime('%b')}"
        if error:
            ws.cell(row=row_i, column=9).value = error
        else:
            ws.cell(row=row_i, column=9).value  = etd_text
            ws.cell(row=row_i, column=10).value = tt_text
            ws.cell(row=row_i, column=15).value = vessel_text
            ws.cell(row=row_i, column=16).value = ts_text
            try:
                wrap = Alignment(wrap_text=True, vertical="top")
                for col in [15, 16]:
                    ws.cell(row=row_i, column=col).alignment = wrap
            except Exception:
                pass
        wb.save(EXCEL_PATH)
        log(f"   💾 Ghi Excel row {row_i}")
    except PermissionError:
        log("   ❌ Lỗi ghi Excel: TẮT FILE EXCEL ĐI rồi chạy lại!")
    except Exception as e:
        log(f"   ❌ write_excel_row: {e}")

# ===================================================================================
# MAIN LOOP
# ===================================================================================
def search_one(pol_excel, pod_excel, valid_dt):
    pol_search = alias_port(pol_excel)
    pod_search = alias_port(pod_excel)
    log(f"   🔄 Search: '{pol_excel}' → '{pol_search}'  |  '{pod_excel}' → '{pod_search}'")

    start_dt, end_dt = compute_valid_window(valid_dt)
    if not start_dt or not end_dt:
        return {"error": "VALID rỗng/không parse được"}
    if end_dt < start_dt:
        return {"error": f"VALID trước ETD tối thiểu date +{DATE_OFFSET_DAYS}"}

    try:
        open_search_page()
        if not select_port(0, pol_search):
            return {"error": "POL không tìm thấy trong YM"}
        if not select_port(1, pod_search):
            return {"error": "POD không tìm thấy trong YM"}
        set_dates(start_dt, end_dt)
        click_search()
        wait_results(timeout=30)
        time.sleep(2.0)
        entries = parse_result_cards()
        entries = filter_by_route(entries, pol_search, pod_search)
        if not entries:
            return {"error": "Không có chuyến phù hợp"}
        selected = apply_etd_rules(entries, valid_dt)
        if not selected:
            return {"error": "Không có ETD hợp lệ (≤ valid)"}
        etd_text   = format_etd_text(selected)
        tt_text    = format_tt_text(selected)
        v_text, ts = format_vessel_block(selected)
        return {
            "etd_text":     etd_text,
            "tt_text":      tt_text,
            "vessel_text":  v_text,
            "ts_text":      ts,
        }
    except Exception as e:
        log(f"   ❌ search_one error: {e}")
        return {"error": f"Exception: {e}"}

def search_one(pol_excel, pod_excel, valid_dt, country_excel=""):
    pol_search = alias_port(pol_excel)
    pod_search = alias_port(pod_excel)
    log(f"   [YM] Search: '{pol_excel}' -> '{pol_search}' | '{pod_excel}' -> '{pod_search}'")

    start_dt, end_dt = compute_valid_window(valid_dt)
    if not start_dt or not end_dt:
        return {"error": "VALID empty/unparseable"}
    if end_dt < TODAY.replace(hour=0, minute=0, second=0, microsecond=0):
        return {"error": "VALID expired"}

    try:
        open_search_page()
        if not select_port(0, pol_search):
            return {"error": "POL not found in YM"}
        if not select_port(1, pod_search):
            return {"error": "POD not found in YM"}

        def run_schedule_search(e_dt, label):
            set_end_date_only(e_dt)
            log(f"   [YM] Search window {label}: To={fmt_ym_date(e_dt)}")
            click_search()
            wait_results(timeout=30)
            time.sleep(2.0)
            raw_entries = parse_result_cards()
            return filter_by_route(raw_entries, pol_search, pod_search)

        entries = run_schedule_search(end_dt, "end-date-only")
        if is_hcm_to_australia(pol_excel, pod_excel, country_excel):
            before_ts_filter = len(entries)
            entries, rejected_ts = filter_hcm_australia_transshipment(
                entries, pol_excel, pod_excel, country_excel
            )
            rejected_codes = sorted({
                str(e.get("ts_port") or "").strip().upper()
                for e in rejected_ts
                if str(e.get("ts_port") or "").strip()
            })
            log(
                "   [YM HCM-AU TS] Chỉ giữ DIRECT/TWKHH/SGSIN: "
                f"{len(entries)}/{before_ts_filter} chuyến"
                + (f" | loại: {', '.join(rejected_codes)}" if rejected_codes else "")
            )
        selected = apply_etd_rules(entries, valid_dt)

        if not selected:
            log(f"   [YM] No ETD >= date +{DATE_OFFSET_DAYS}; fallback choose farthest ETD inside valid")
            selected = apply_etd_fallback_rules(entries, valid_dt)

        if not entries:
            if is_hcm_to_australia(pol_excel, pod_excel, country_excel):
                return {"error": "No schedule via DIRECT/TWKHH/SGSIN for HCM-Australia"}
            return {"error": "No matching route"}
        if not selected:
            return {"error": "No ETD in valid window"}

        etd_text = format_etd_text(selected)
        tt_text = format_tt_text(selected)
        v_text, ts = format_vessel_block(selected)
        return {
            "etd_text": etd_text,
            "tt_text": tt_text,
            "vessel_text": v_text,
            "ts_text": ts,
        }
    except Exception as e:
        log(f"   [YM] search_one error: {e}")
        return {"error": f"Exception: {e}"}

def main():
    if not os.path.exists(EXCEL_PATH):
        log(f"❌ Không tìm thấy file Excel: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    target_single_row = None
    if SINGLE_ROW:
        try:
            target_single_row = int(SINGLE_ROW)
            log(f"[SINGLE_ROW] Chỉ chạy dòng {target_single_row} theo lệnh từ main.py")
        except Exception:
            log(f"[SINGLE_ROW] Không hợp lệ: {SINGLE_ROW}")

    target_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if target_single_row is not None and i != target_single_row:
            continue
        country = str(row[1] or "").strip()
        pol     = str(row[2] or "").strip()
        pod     = str(row[3] or "").strip()
        carrier = str(row[4] or "").strip().upper()
        valid_raw = row[10] if len(row) > 10 else None
        if not pol or not pod or carrier not in CARRIER_TARGETS:
            continue
        if FILTER_POL and pol.upper() != FILTER_POL:
            continue
        if FILTER_POD and pod.upper() != FILTER_POD:
            continue
        target_rows.append((i, country, pol, pod, valid_raw))
    wb.close()

    log(f"📋 Có {len(target_rows)} dòng cần check (carrier = {', '.join(sorted(CARRIER_TARGETS))})")
    if not target_rows:
        log("   Không có gì để làm, exit.")
        return

    for idx, (row_i, country, pol, pod, valid_raw) in enumerate(target_rows, start=1):
        log(f"\n========== [{idx}/{len(target_rows)}] DÒNG {row_i}: {pol} → {pod} | VALID={valid_raw}")
        valid_dt = parse_valid_date(valid_raw)
        if not valid_dt:
            log("⚠️ Không có ngày Valid trong Excel -> bỏ qua row, không gọi API.")
            continue
        result = search_one(pol, pod, valid_dt, country_excel=country)
        if result.get("error"):
            write_excel_row(row_i, "", "", "", "", error=result["error"])
        else:
            write_excel_row(
                row_i,
                result["etd_text"],
                result["tt_text"],
                result["vessel_text"],
                result["ts_text"],
            )
        time.sleep(1.0)

    log("\n✅ HOÀN TẤT bot YANG MING")

if __name__ == "__main__":
    main()
