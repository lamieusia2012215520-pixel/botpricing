"""
Bot Maersk — check giá tự động
URL: https://www.maersk.com/book/
"""
import sys
import zlib
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
import openpyxl
import os, time, re, random, math, json, subprocess, urllib.request
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from remark_rules import build_subject_remark, charge_amount_to_usd, is_china_destination

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

print("\n" + "🔥"*20)
print(">>> ĐANG CHẠY BẢN XỊN NHẤT CỦA SẾP PIO <<<")
print("🔥"*20 + "\n")
from datetime import datetime


# ===================================================================================
# GLOBAL VARIABLES & EXCEPTIONS (Dùng cho Human-in-the-loop)
# ===================================================================================
CURRENT_POL = ""
CURRENT_POD = ""
STDIN_DISABLED = False
STDIN_POLLING_ANNOUNCED = False

class ManualHandoverComplete(Exception):
    pass
class WebLagException(Exception):
    pass
class NoSailingsException(Exception):
    pass
# ===================================================================================
# CẤU HÌNH
# ===================================================================================
current_folder = os.getcwd()
excel_path     = os.environ.get("EXCEL_PATH", os.path.join(current_folder, "input_gia.xlsx"))
FILTER_POL     = os.environ.get("FILTER_POL", "").strip().upper()
FILTER_POD     = os.environ.get("FILTER_POD", "").strip().upper()
FILTER_COUNTRY = os.environ.get("FILTER_COUNTRY", "").strip().upper()
SINGLE_ROW     = os.environ.get("SINGLE_ROW", "").strip()
MAERSK_SIMPLE_MODE = os.environ.get("MAERSK_SIMPLE", "1").strip().lower() not in {"0", "false", "no"}
DATE_OFFSET_DAYS = parse_date_offset_days()

def _excel_formula_from_parts(parts):
    tokens = []
    for raw in parts:
        try:
            amount = float(raw)
        except (TypeError, ValueError):
            continue
        if abs(amount) < 1e-9:
            continue
        sign = "-" if amount < 0 else "+"
        amount = abs(amount)
        if abs(amount - round(amount)) < 1e-9:
            text = str(int(round(amount)))
        else:
            text = f"{amount:.2f}".rstrip("0").rstrip(".")
        tokens.append((sign if tokens else ("-" if sign == "-" else "")) + text)
    return "=" + "".join(tokens) if tokens else None

MAERSK_BROWSER_BACKEND = os.environ.get("MAERSK_BROWSER_BACKEND", "edge_debug" if MAERSK_SIMPLE_MODE else "cloak").strip().lower()
MAERSK_ALLOW_FALLBACK  = os.environ.get("MAERSK_ALLOW_UNDETECTED_FALLBACK", "1").strip().lower() not in {"0", "false", "no"}
MAERSK_PROFILE_DIR     = os.environ.get("MAERSK_PROFILE_DIR", r"C:\ChromeProfile_Maersk")
MAERSK_EDGE_BINARY     = os.environ.get("MAERSK_EDGE_BINARY", r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe")
MAERSK_DEBUG_PORT      = int(os.environ.get("MAERSK_DEBUG_PORT", "9521"))
MAERSK_HEADLESS        = os.environ.get("MAERSK_HEADLESS", "").strip().lower() in {"1", "true", "yes"}
MAERSK_API_CAPTURE     = os.environ.get("MAERSK_API_CAPTURE", "1").strip().lower() in {"1", "true", "yes"}
MAERSK_CACHE_ENABLED   = os.environ.get("MAERSK_CACHE", "1" if MAERSK_SIMPLE_MODE else "0").strip().lower() not in {"0", "false", "no"}
MAERSK_CACHE_PATH      = os.environ.get("MAERSK_CACHE_PATH", os.path.join(current_folder, "maersk_cache.json"))
MAERSK_SIMPLE_RETRY_LIMIT = max(0, int(os.environ.get("MAERSK_SIMPLE_RETRY_LIMIT", "1")))
try:
    MAERSK_PRE_SEARCH_MIN_SECONDS = max(0.0, float(os.environ.get("MAERSK_PRE_SEARCH_MIN_SECONDS", "2" if MAERSK_SIMPLE_MODE else "18")))
except ValueError:
    MAERSK_PRE_SEARCH_MIN_SECONDS = 2.0 if MAERSK_SIMPLE_MODE else 18.0
try:
    MAERSK_PRE_SEARCH_MAX_SECONDS = max(MAERSK_PRE_SEARCH_MIN_SECONDS, float(os.environ.get("MAERSK_PRE_SEARCH_MAX_SECONDS", "5" if MAERSK_SIMPLE_MODE else "35")))
except ValueError:
    MAERSK_PRE_SEARCH_MAX_SECONDS = 5.0 if MAERSK_SIMPLE_MODE else 35.0
try:
    MAERSK_ROW_BREAK_MIN_SECONDS = max(0.0, float(os.environ.get("MAERSK_ROW_BREAK_MIN_SECONDS", "4" if MAERSK_SIMPLE_MODE else "25")))
except ValueError:
    MAERSK_ROW_BREAK_MIN_SECONDS = 4.0 if MAERSK_SIMPLE_MODE else 25.0
try:
    MAERSK_ROW_BREAK_MAX_SECONDS = max(MAERSK_ROW_BREAK_MIN_SECONDS, float(os.environ.get("MAERSK_ROW_BREAK_MAX_SECONDS", "8" if MAERSK_SIMPLE_MODE else "45")))
except ValueError:
    MAERSK_ROW_BREAK_MAX_SECONDS = 8.0 if MAERSK_SIMPLE_MODE else 45.0
try:
    MAERSK_AFTER_CAPTCHA_SECONDS = max(0.0, float(os.environ.get("MAERSK_AFTER_CAPTCHA_SECONDS", "12")))
except ValueError:
    MAERSK_AFTER_CAPTCHA_SECONDS = 12.0

BASE_URL     = "https://www.maersk.com/book/"
MAERSK_GROUP = {"MAERSK"}

HAIPHONG_NEARBY_COUNTRIES = {
    "CHINA", "THAILAND", "JAPAN", "PHILIPPINES", "INDONESIA",
    "MALAYSIA", "KOREA", "KOREA, SOUTH", "CAMBODIA", "TAIWAN",
    "SINGAPORE", "MYANMAR", "TIMOR-LESTE", "EAST TIMOR", "BRUNEI",
    "HONG KONG", "HONG KONG CHINA", "MACAU", "VIETNAM",
}

MAERSK_PORT_ALIASES = {
    "HO CHI MINH":         "HO CHI MINH",
    "HAI PHONG":           "haiphong",
    "HAIPHONG_LACH_HUYEN": "haiphong",
    "HAIPHONG_NORMAL":     "haiphong",
    "NHAVA SHEVA":         "Jawaharlal Nehru",
    "JNPT":                "Jawaharlal Nehru",
    "CHENNAI":             "Kattupalli",
    "MADRAS":              "Kattupalli",
    "INCHEON":             "Inchon",
    "INCHON":              "Inchon",
    "FOS SUR MER":         "Fos",
    "LE HAVRE":            "Le Havre",
    "GENOA":               "Genoa",
    "NAPLES":              "Naples",
    "VENICE":              "Venice",
    "GDANSK":              "Gdansk",
    "UMM AL QUWAIN":       "Umm Al Quwain",
    "UMM AL QAIWAIN":      "Umm Al Quwain",
    "UMM QASR":            "Umm Qasr",
}

MAERSK_EXACT_OPTION = {
    "HO CHI MINH":         "Ho Chi Minh City (Ho Chi Minh), Vietnam",
    "HAIPHONG_LACH_HUYEN": "Haiphong - Lach Huyen (Hai Phong), Vietnam",
    "HAIPHONG_NORMAL":     "Haiphong (Hai Phong), Vietnam",
}

COL_20  = 6
COL_40  = 7
COL_40H = 8

UNAVAILABLE_STATUSES = [
    'container not available',
    'vessel sold out',
    'vessel is not open',
    'vessel is full',
]

driver        = None
row_i_current = 0

# ===================================================================================
# KẾT NỐI — DÙNG UNDETECTED CHROME + CDP STEALTH
# ===================================================================================
def _connect_legacy_inline_driver_unused():
    global driver
    import undetected_chromedriver as uc
    try:
        options = uc.ChromeOptions()
        options.add_argument("--no-first-run")
        options.add_argument("--no-default-browser-check")
        options.add_argument("--disable-infobars")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-background-timer-throttling")
        options.add_argument("--disable-backgrounding-occluded-windows")
        options.add_argument("--disable-renderer-backgrounding")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--lang=vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7")
        # Profile riêng — giữ cookie/session giữa các lần chạy
        options.add_argument(r"--user-data-dir=C:\ChromeProfile_Maersk")

        driver = uc.Chrome(options=options, version_main=146)
        try:
            driver.maximize_window()
        except Exception:
            try:
                info = driver.execute_cdp_cmd("Browser.getWindowForTarget", {})
                driver.execute_cdp_cmd("Browser.setWindowBounds", {
                    "windowId": info["windowId"],
                    "bounds": {"windowState": "maximized"}
                })
            except Exception:
                pass

        # ===== CDP STEALTH PATCHES =====
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                // FIX 1: WebDriver — xóa hoàn toàn khỏi prototype
                try {
                    const newProto = navigator.__proto__;
                    delete newProto.webdriver;
                    navigator.__proto__ = newProto;
                } catch(e) {}
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined,
                    configurable: true
                });

                // FIX 2: Plugins — dùng PluginArray thật từ iframe
                try {
                    const iframe = document.createElement('iframe');
                    iframe.style.display = 'none';
                    document.documentElement.appendChild(iframe);
                    const iframeNavigator = iframe.contentWindow.navigator;
                    Object.defineProperty(navigator, 'plugins', {
                        get: () => iframeNavigator.plugins,
                        configurable: true
                    });
                    Object.defineProperty(navigator, 'mimeTypes', {
                        get: () => iframeNavigator.mimeTypes,
                        configurable: true
                    });
                    document.documentElement.removeChild(iframe);
                } catch(e) {}

                // FIX 3: WebGL — spoof GPU thật của máy
                try {
                    const getParameter = WebGLRenderingContext.prototype.getParameter;
                    WebGLRenderingContext.prototype.getParameter = function(parameter) {
                        if (parameter === 37445) return 'Google Inc. (Intel)';
                        if (parameter === 37446) return 'ANGLE (Intel, Intel(R) UHD Graphics 620 (0x00003EA0) Direct3D11 vs_5_0 ps_5_0, D3D11)';
                        return getParameter.call(this, parameter);
                    };
                    const getParameter2 = WebGL2RenderingContext.prototype.getParameter;
                    WebGL2RenderingContext.prototype.getParameter = function(parameter) {
                        if (parameter === 37445) return 'Google Inc. (Intel)';
                        if (parameter === 37446) return 'ANGLE (Intel, Intel(R) UHD Graphics 620 (0x00003EA0) Direct3D11 vs_5_0 ps_5_0, D3D11)';
                        return getParameter2.call(this, parameter);
                    };
                } catch(e) {}

                // Languages và chrome object
                Object.defineProperty(navigator, 'languages', {
                    get: () => ['vi-VN', 'vi', 'en-US', 'en']
                });
                window.chrome = {
                    runtime: {},
                    loadTimes: function() {},
                    csi: function() {},
                    app: {}
                };
                const originalQuery = window.navigator.permissions.query;
                window.navigator.permissions.query = (parameters) => (
                    parameters.name === 'notifications'
                        ? Promise.resolve({ state: Notification.permission })
                        : originalQuery(parameters)
                );
            """
        })

        print("✅ Đã khởi động Chrome undetected + CDP stealth!")
        return True
    except Exception as e:
        print(f"❌ Lỗi khởi động Chrome: {e}")
        return False

def _stable_fingerprint_seed():
    raw_seed = os.environ.get("MAERSK_FINGERPRINT_SEED", "").strip()
    if raw_seed and raw_seed.lower() != "random":
        try:
            return str(max(1, int(raw_seed)))
        except ValueError:
            print(f"⚠️ MAERSK_FINGERPRINT_SEED không hợp lệ: {raw_seed}. Dùng seed tự động.")
    if raw_seed.lower() == "random":
        return None
    return str(10000 + (zlib.crc32(MAERSK_PROFILE_DIR.encode("utf-8")) % 90000))


def _add_common_chrome_args(options):
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-infobars")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    options.add_argument("--lang=vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7")
    options.add_argument(f"--user-data-dir={MAERSK_PROFILE_DIR}")
    if MAERSK_HEADLESS:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
    proxy = os.environ.get("MAERSK_PROXY", "").strip()
    if proxy:
        options.add_argument(f"--proxy-server={proxy}")
    try:
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
    except Exception:
        pass
    return options


def _maximize_driver_window(active_driver):
    try:
        active_driver.maximize_window()
    except Exception:
        try:
            info = active_driver.execute_cdp_cmd("Browser.getWindowForTarget", {})
            active_driver.execute_cdp_cmd("Browser.setWindowBounds", {
                "windowId": info["windowId"],
                "bounds": {"windowState": "maximized"}
            })
        except Exception:
            pass


def _apply_legacy_cdp_stealth_patch(active_driver):
    try:
        active_driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                try {
                    const newProto = navigator.__proto__;
                    delete newProto.webdriver;
                    navigator.__proto__ = newProto;
                } catch(e) {}
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined,
                    configurable: true
                });
                Object.defineProperty(navigator, 'languages', {
                    get: () => ['vi-VN', 'vi', 'en-US', 'en']
                });
            """
        })
    except Exception as e:
        print(f"⚠️ Không inject được legacy CDP stealth: {e}")


def _connect_cloak_driver():
    global driver
    from cloakbrowser.download import ensure_binary
    from cloakbrowser.config import get_default_stealth_args

    os.environ.setdefault("CLOAKBROWSER_AUTO_UPDATE", "false")

    binary_path = ensure_binary()
    options = ChromeOptions()
    options.binary_location = binary_path
    _add_common_chrome_args(options)

    stealth_args = get_default_stealth_args()
    seed = _stable_fingerprint_seed()
    if seed:
        stealth_args = [arg for arg in stealth_args if not arg.startswith("--fingerprint=")]
        stealth_args.append(f"--fingerprint={seed}")
    stealth_args.extend([
        "--fingerprint-locale=vi-VN",
        "--fingerprint-timezone=Asia/Ho_Chi_Minh",
    ])

    seen_args = set()
    for arg in stealth_args:
        if arg not in seen_args:
            options.add_argument(arg)
            seen_args.add(arg)

    driver = webdriver.Chrome(options=options)
    _maximize_driver_window(driver)
    print(f"✅ Đã khởi động Maersk bằng CloakBrowser: {binary_path}")
    return True


def _connect_undetected_driver():
    global driver
    import undetected_chromedriver as uc

    options = uc.ChromeOptions()
    _add_common_chrome_args(options)
    options.add_argument("--disable-blink-features=AutomationControlled")

    driver = uc.Chrome(options=options, version_main=146)
    _maximize_driver_window(driver)
    _apply_legacy_cdp_stealth_patch(driver)
    print("✅ Đã khởi động Chrome undetected + legacy CDP stealth!")
    return True


def _connect_edge_driver():
    global driver

    options = EdgeOptions()
    if os.path.exists(MAERSK_EDGE_BINARY):
        options.binary_location = MAERSK_EDGE_BINARY
    _add_common_chrome_args(options)

    driver_path = os.path.join(current_folder, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        raise FileNotFoundError(f"Không tìm thấy Edge WebDriver: {driver_path}")
    driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)

    _maximize_driver_window(driver)
    print(f"✅ Đã khởi động Maersk bằng Microsoft Edge profile: {MAERSK_PROFILE_DIR}")
    return True


def _edge_debug_endpoint_ok():
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:{MAERSK_DEBUG_PORT}/json/version", timeout=2) as resp:
            text = resp.read().decode("utf-8", errors="ignore")
        return "webSocketDebuggerUrl" in text or "Microsoft Edge" in text
    except Exception:
        return False


def _start_edge_debug():
    if _edge_debug_endpoint_ok():
        print(f"✅ Edge Maersk debug port {MAERSK_DEBUG_PORT} đã mở sẵn.")
        return True

    print(f"🌐 Mở Edge Maersk debug port {MAERSK_DEBUG_PORT}...")
    args = [
        MAERSK_EDGE_BINARY,
        f"--remote-debugging-port={MAERSK_DEBUG_PORT}",
        f"--user-data-dir={MAERSK_PROFILE_DIR}",
        "--start-maximized",
        "--no-first-run",
        "--no-default-browser-check",
    ]
    try:
        subprocess.Popen(args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as e:
        print(f"❌ Không mở được Edge Maersk: {e}")
        return False

    for _ in range(30):
        if _edge_debug_endpoint_ok():
            return True
        time.sleep(0.5)
    return False


def _connect_edge_debug_driver():
    global driver

    if not _start_edge_debug():
        raise RuntimeError(f"Edge debug port {MAERSK_DEBUG_PORT} chưa sẵn sàng")

    options = EdgeOptions()
    options.add_experimental_option("debuggerAddress", f"127.0.0.1:{MAERSK_DEBUG_PORT}")

    driver_path = os.path.join(current_folder, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        raise FileNotFoundError(f"Không tìm thấy Edge WebDriver: {driver_path}")
    driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)

    try:
        handles = driver.window_handles
        if handles:
            driver.switch_to.window(handles[-1])
    except Exception:
        pass
    _maximize_driver_window(driver)
    print(f"✅ Đã attach Edge Maersk port {MAERSK_DEBUG_PORT} | profile: {MAERSK_PROFILE_DIR}")
    return True


def connect_driver():
    backend = MAERSK_BROWSER_BACKEND or "cloak"
    if backend not in {"cloak", "undetected", "edge", "edge_debug"}:
        print(f"⚠️ MAERSK_BROWSER_BACKEND không hợp lệ: {backend}. Dùng cloak.")
        backend = "cloak"

    if backend == "edge_debug":
        try:
            return _connect_edge_debug_driver()
        except Exception as e:
            print(f"❌ Lỗi attach Microsoft Edge debug: {e}")
            return False

    if backend == "edge":
        try:
            return _connect_edge_driver()
        except Exception as e:
            print(f"❌ Lỗi khởi động Microsoft Edge: {e}")
            return False

    if backend == "undetected":
        try:
            return _connect_undetected_driver()
        except Exception as e:
            print(f"❌ Lỗi khởi động undetected Chrome: {e}")
            return False

    try:
        return _connect_cloak_driver()
    except Exception as e:
        print(f"❌ Lỗi khởi động CloakBrowser: {e}")
        if not MAERSK_ALLOW_FALLBACK:
            return False
        print("⚠️ Fallback sang undetected_chromedriver để không chặn luồng chạy.")
        try:
            return _connect_undetected_driver()
        except Exception as fallback_error:
            print(f"❌ Fallback undetected cũng lỗi: {fallback_error}")
            return False


if not connect_driver():
    print("❌ Không khởi động được Chrome. Thoát.")
    sys.exit(1)

if driver is None:
    print("❌ driver vẫn là None. Thoát.")
    sys.exit(1)

# ===================================================================================
# HELPERS
# ===================================================================================
def rand_sleep(a=0.05, b=0.15):
    time.sleep(random.uniform(a, b))

def human_sleep(a, b):
    time.sleep(random.uniform(a * 0.5, b * 0.5))

def micro_pause():
    time.sleep(random.uniform(0.02, 0.08))

def _stdin_is_interactive():
    global STDIN_DISABLED
    if STDIN_DISABLED:
        return False
    try:
        return bool(sys.stdin and sys.stdin.isatty())
    except Exception:
        return False

def manual_pause(prompt, poll_seconds=3):
    global STDIN_DISABLED, STDIN_POLLING_ANNOUNCED
    if _stdin_is_interactive():
        try:
            input(prompt)
            return True
        except EOFError:
            STDIN_DISABLED = True
            print("   ⚠️ Không đọc được Enter từ stdin. Chuyển sang chế độ tự polling trình duyệt.")
    else:
        if not STDIN_POLLING_ANNOUNCED:
            print(prompt)
            print(f"   ℹ️ Không có stdin tương tác; bot sẽ kiểm tra trình duyệt mỗi {poll_seconds}s.")
            STDIN_POLLING_ANNOUNCED = True
    time.sleep(poll_seconds)
    return False

def wait_for_manual_condition(prompt, condition_fn, timeout_seconds=None, poll_seconds=3):
    global STDIN_DISABLED
    started = time.time()
    announced_polling = False

    while True:
        if _stdin_is_interactive():
            try:
                input(prompt)
            except EOFError:
                STDIN_DISABLED = True
                print("   ⚠️ Không đọc được Enter từ stdin. Chuyển sang chế độ tự polling trình duyệt.")
            if condition_fn():
                return True
            print("   ⚠️ Trạng thái trình duyệt chưa đúng, kiểm tra lại sau khi thao tác xong.")
        else:
            if not announced_polling:
                print(prompt)
                print(f"   ℹ️ Không có stdin tương tác; bot sẽ tự kiểm tra trình duyệt mỗi {poll_seconds}s.")
                announced_polling = True
            if condition_fn():
                return True

        if timeout_seconds is not None and time.time() - started >= timeout_seconds:
            return False
        time.sleep(poll_seconds)

def is_maersk_book_ready():
    try:
        current_url = (driver.current_url or "").lower()
    except Exception:
        return False
    if "login" in current_url or "signin" in current_url:
        return False
    return "maersk.com/book/" in current_url or "/sailings" in current_url

def wait_until_book_ready(timeout_seconds=25):
    end = time.time() + timeout_seconds
    while time.time() < end:
        if is_maersk_book_ready():
            return True
        time.sleep(1)
    return is_maersk_book_ready()

def human_scroll():
    steps     = random.randint(2, 4)
    scroll_px = random.randint(80, 200)
    per_step  = scroll_px // steps
    for _ in range(steps):
        driver.execute_script(f"window.scrollBy(0, {per_step});")
        time.sleep(random.uniform(0.03, 0.08))
    human_sleep(0.2, 0.5)
    for _ in range(steps):
        driver.execute_script(f"window.scrollBy(0, -{per_step});")
        time.sleep(random.uniform(0.03, 0.07))
    human_sleep(0.1, 0.3)

def bezier_point(p0, p1, p2, p3, t):
    """Tính điểm trên đường cong Bezier bậc 3"""
    x = (1-t)**3*p0[0] + 3*(1-t)**2*t*p1[0] + 3*(1-t)*t**2*p2[0] + t**3*p3[0]
    y = (1-t)**3*p0[1] + 3*(1-t)**2*t*p1[1] + 3*(1-t)*t**2*p2[1] + t**3*p3[1]
    return (x, y)

def human_move_to(element):
    """Di chuột theo đường cong Bezier — giống tay người thật"""
    try:
        rect = driver.execute_script("""
            var r = arguments[0].getBoundingClientRect();
            return {x: r.left + r.width/2, y: r.top + r.height/2};
        """, element)

        target_x = rect['x'] + random.randint(-4, 4)
        target_y = rect['y'] + random.randint(-4, 4)

        start_x = random.randint(200, 800)
        start_y = random.randint(200, 600)

        ctrl1 = (
            start_x + (target_x - start_x) * random.uniform(0.2, 0.4) + random.randint(-60, 60),
            start_y + (target_y - start_y) * random.uniform(0.1, 0.3) + random.randint(-60, 60)
        )
        ctrl2 = (
            start_x + (target_x - start_x) * random.uniform(0.6, 0.8) + random.randint(-40, 40),
            start_y + (target_y - start_y) * random.uniform(0.6, 0.8) + random.randint(-40, 40)
        )

        steps = random.randint(20, 35)
        action = ActionChains(driver)
        action.move_by_offset(start_x - 400, start_y - 300)

        prev_x, prev_y = start_x, start_y
        for i in range(1, steps + 1):
            t = i / steps
            t_eased = t * t * (3 - 2 * t)
            cur_x, cur_y = bezier_point(
                (start_x, start_y), ctrl1, ctrl2, (target_x, target_y), t_eased
            )
            dx = int(cur_x - prev_x)
            dy = int(cur_y - prev_y)
            if dx != 0 or dy != 0:
                action.move_by_offset(dx, dy)
                pause = random.uniform(0.005, 0.015) if 0.2 < t < 0.8 else random.uniform(0.015, 0.04)
                action.pause(pause)
            prev_x, prev_y = cur_x, cur_y

        action.perform()
        time.sleep(random.uniform(0.08, 0.18))
    except:
        pass

def typo_then_correct(inp, correct_text):
    if random.random() < 0.2:
        typo_len   = random.randint(1, 2)
        typo_chars = random.choices('abcdefghijklmnopqrstuvwxyz', k=typo_len)
        for ch in typo_chars:
            inp.send_keys(ch)
            time.sleep(random.uniform(0.05, 0.12))
        human_sleep(0.1, 0.3)
        for _ in range(typo_len):
            inp.send_keys(Keys.BACK_SPACE)
            time.sleep(random.uniform(0.03, 0.08))
        human_sleep(0.05, 0.15)
    for ch in correct_text:
        inp.send_keys(ch)
        time.sleep(max(0.03, random.gauss(0.06, 0.02)))

def row_break():
    t = random.uniform(MAERSK_ROW_BREAK_MIN_SECONDS, MAERSK_ROW_BREAK_MAX_SECONDS)
    print(f"   ⏳ Nghỉ {t:.1f}s trước row tiếp theo...")
    time.sleep(t)

def pre_interaction_idle():
    if MAERSK_SIMPLE_MODE:
        time.sleep(random.uniform(0.4, 1.0))
        return
    """Giả lập người thật vừa vào trang: nhìn, di chuột lơ đãng rồi mới điền form"""
    idle_duration = random.uniform(2.0, 4.5)
    print(f"   👁️ Pre-idle {idle_duration:.1f}s (giả lập đọc trang)...")
    steps = random.randint(3, 6)
    for _ in range(steps):
        time.sleep(idle_duration / steps)
        if random.random() < 0.45:
            px = random.randint(-60, 120)
            driver.execute_script(f"window.scrollBy(0, {px});")
        if random.random() < 0.5:
            try:
                action = ActionChains(driver)
                action.move_by_offset(
                    random.randint(-80, 80),
                    random.randint(-50, 50)
                ).pause(random.uniform(0.1, 0.3)).perform()
            except:
                pass

def is_india_route(country):
    return country.strip().upper() == "INDIA"

def resolve_maersk_alias(port_name, pod_country=""):
    upper = port_name.strip().upper()
    search = MAERSK_PORT_ALIASES.get(upper, port_name.strip())
    exact  = MAERSK_EXACT_OPTION.get(upper, None)
    return search, exact

# ===================================================================================
# EVENT LOGGER (DEBUG — xóa sau khi xong)
# ===================================================================================
def inject_event_logger():
    if MAERSK_SIMPLE_MODE:
        return
    driver.execute_script("""
        window._botLog = [];
        ['mousedown','mouseup','click','mousemove','mouseover',
         'focus','blur','keydown','keyup','keypress','input','change',
         'pointerdown','pointerup','pointermove']
        .forEach(function(evt) {
            document.addEventListener(evt, function(e) {
                window._botLog.push({
                    type: e.type,
                    target: e.target.tagName + (e.target.type ? '['+e.target.type+']' : ''),
                    isTrusted: e.isTrusted,
                    t: Date.now()
                });
            }, true);
        });
    """)

def dump_event_log(label=""):
    if MAERSK_SIMPLE_MODE:
        return
    logs = driver.execute_script("return window._botLog || [];")
    print(f"\n{'='*40}")
    print(f"EVENT LOG {label} — {len(logs)} events")
    for e in logs[-30:]:
        trusted = "✅HUMAN" if e['isTrusted'] else "❌BOT"
        print(f"  {trusted} | {e['type']:15} | {e['target']}")
    print(f"{'='*40}\n")
    driver.execute_script("window._botLog = [];")

def inject_api_recorder():
    if not MAERSK_API_CAPTURE:
        return
    hook = """
    (function(){
      if (window.__MAERSK_API_RECORDER__) return;
      window.__MAERSK_API_RECORDER__ = true;
      window.__MAERSK_API_CALLS__ = [];
      const keep = (url) => {
        url = String(url || '').toLowerCase();
        return url.includes('maersk') || url.includes('hcaptcha') ||
               url.includes('captcha') || url.includes('booking') ||
               url.includes('sailing') || url.includes('offer') ||
               url.includes('price') || url.includes('graphql');
      };
      const push = (o) => {
        try {
          if (!keep(o.url)) return;
          o.t = Date.now();
          window.__MAERSK_API_CALLS__.push(o);
          if (window.__MAERSK_API_CALLS__.length > 300) window.__MAERSK_API_CALLS__.shift();
        } catch(e) {}
      };
      const origFetch = window.fetch;
      if (origFetch) {
        window.fetch = async function(input, init) {
          const url = (typeof input === 'string') ? input : (input && input.url);
          const body = init && init.body;
          const res = await origFetch.apply(this, arguments);
          try {
            const clone = res.clone();
            clone.text().then(txt => push({
              kind:'fetch', method:(init && init.method) || 'GET',
              url:url, status:res.status, body:String(body || '').slice(0,3000),
              response:String(txt || '').slice(0,20000)
            })).catch(()=>{});
          } catch(e) {}
          return res;
        };
      }
      const origOpen = XMLHttpRequest.prototype.open;
      const origSend = XMLHttpRequest.prototype.send;
      XMLHttpRequest.prototype.open = function(method, url) {
        this.__maersk_method = method;
        this.__maersk_url = url;
        return origOpen.apply(this, arguments);
      };
      XMLHttpRequest.prototype.send = function(body) {
        const xhr = this;
        const old = xhr.onreadystatechange;
        xhr.onreadystatechange = function(){
          if (xhr.readyState === 4) {
            push({
              kind:'xhr', method:xhr.__maersk_method, url:xhr.__maersk_url,
              status:xhr.status, body:String(body || '').slice(0,3000),
              response:String(xhr.responseText || '').slice(0,20000)
            });
          }
          if (old) return old.apply(xhr, arguments);
        };
        return origSend.apply(this, arguments);
      };
    })();
    """
    try:
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": hook})
    except Exception:
        pass
    try:
        driver.execute_script(hook)
    except Exception:
        pass

def dump_api_calls(label=""):
    if not MAERSK_API_CAPTURE:
        return []
    try:
        calls = driver.execute_script("return window.__MAERSK_API_CALLS__ || [];") or []
    except Exception as e:
        print(f"   [MAERSK-API] Không đọc được API calls: {e}")
        return []
    print(f"\\n{'='*40}")
    print(f"MAERSK API CALLS {label} - {len(calls)} calls")
    shown_calls = [c for c in calls if "productoffer" in str(c.get("url") or "").lower()] if MAERSK_SIMPLE_MODE else calls[-40:]
    for call in shown_calls[-10:]:
        url = re.sub(r'([?&])(token|code|id_token|access_token|nonce|code_challenge|client_secret)=[^&]+', r'\\1\\2=REDACTED', str(call.get('url') or ''))
        body = str(call.get('body') or '')
        resp = str(call.get('response') or '')
        print(f"  [{call.get('kind')}] {call.get('method')} {call.get('status')} {url[:220]}")
        if body and not MAERSK_SIMPLE_MODE:
            print(f"    body: {body[:300]}")
        if resp:
            print(f"    resp: {resp[:500].replace(chr(10), ' ')}")
    print(f"{'='*40}\\n")
    try:
        dump_path = os.environ.get("MAERSK_API_DUMP_PATH", os.path.join(current_folder, "maersk_api_calls_latest.json"))
        with open(dump_path, "w", encoding="utf-8") as f:
            json.dump(calls, f, ensure_ascii=False, indent=2)
        print(f"   [MAERSK-API] Đã lưu full API calls: {dump_path}")
    except Exception as e:
        print(f"   [MAERSK-API] Không lưu được API dump: {e}")
    return calls

def latest_productoffer_api_error():
    if not MAERSK_API_CAPTURE:
        return ""
    try:
        calls = driver.execute_script("return window.__MAERSK_API_CALLS__ || [];") or []
    except Exception:
        return ""
    for call in reversed(calls):
        url = str(call.get("url") or "").lower()
        if "productoffer" not in url:
            continue
        status = int(call.get("status") or 0)
        if status < 400:
            return ""
        resp = str(call.get("response") or "")
        msg = resp[:500].replace("\n", " ")
        if status == 429:
            return f"MAERSK RATE LIMIT 429: {msg}"
        return f"MAERSK API ERROR {status}: {msg}"
    return ""

# ===================================================================================
# SHADOW DOM HELPERS
# ===================================================================================
_PIERCE_JS = """
function pierce(root, sel) {
    if (!root) return null;
    var el = root.querySelector ? root.querySelector(sel) : null;
    if (el) return el;
    var all = root.querySelectorAll ? root.querySelectorAll('*') : [];
    for (var i = 0; i < all.length; i++) {
        if (all[i].shadowRoot) {
            var f = pierce(all[i].shadowRoot, sel);
            if (f) return f;
        }
    }
    return null;
}
function pierceAll(root, tag) {
    var found = [];
    if (!root || !root.querySelectorAll) return found;
    var all = root.querySelectorAll('*');
    for (var i = 0; i < all.length; i++) {
        if (all[i].tagName && all[i].tagName.toLowerCase() === tag) found.push(all[i]);
        if (all[i].shadowRoot) found = found.concat(pierceAll(all[i].shadowRoot, tag));
    }
    return found;
}
"""

# ===================================================================================
# CAPTCHA HANDLER (Human-in-the-loop)
# ===================================================================================
def handle_captcha_if_present(**kwargs):
    global CURRENT_POL, CURRENT_POD

    captcha_detected = False

    for _ in range(12):
        time.sleep(1)

        captcha_detected = driver.execute_script("""
            if (window.location.href.includes('validate.maersk.com') || window.location.href.includes('captcha')) return true;
            var frames = document.querySelectorAll('iframe[src*="hcaptcha"]');
            for(var i=0; i<frames.length; i++) {
                var node = frames[i];
                var isHidden = false;
                while(node && node !== document.body) {
                    var style = window.getComputedStyle(node);
                    if (style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') {
                        isHidden = true;
                        break;
                    }
                    node = node.parentElement;
                }
                if (!isHidden) {
                    var rect = frames[i].getBoundingClientRect();
                    if(rect.width > 10 && rect.height > 10) return true;
                }
            }
            return false;
        """)

        if captcha_detected:
            break

        error_msg = check_booking_error_notification()
        if error_msg and "no sailings" in error_msg.lower():
            print(f"   🚨 Bot phát hiện banner: Tuyến này không có chuyến! Rút lui sang dòng tiếp theo...")
            raise NoSailingsException()

        if "/sailings" in driver.current_url:
            return 'ok'

    if captcha_detected:
        dump_api_calls("CAPTCHA")
        api_error = latest_productoffer_api_error()
        if api_error:
            print(f"   🚫 {api_error}")
            return 'api_error'

        print(f"\n   🛑 SẾP PIO ƠI, DÍNH CHỐT hCAPTCHA RỒI!")
        print(f"   👉 Bot sẽ NGỦ ĐÔNG hoàn toàn để ẩn mình khỏi radar Cloudflare.")
        print(f"   👉 Sếp hãy tự tay giải mồi, sau đó đợi web xoay xong.")
        print(f"   🎯 Nhiệm vụ hiện tại: {CURRENT_POL} -> {CURRENT_POD}")

        while True:
            manual_pause(f"\n   🟢 [HÀNH ĐỘNG]: Giải xong Captcha, ĐỢI WEB HẾT XOAY rồi BẤM ENTER...", poll_seconds=5)

            current_url = driver.current_url

            if "/sailings" in current_url:
                page_text = driver.execute_script("return document.body.innerText;").lower()

                pol_alias, pol_exact = resolve_maersk_alias(CURRENT_POL)
                pod_alias, pod_exact = resolve_maersk_alias(CURRENT_POD)

                pol_match = (CURRENT_POL.lower() in page_text) or (pol_alias.lower() in page_text) or (pol_exact and pol_exact.lower() in page_text)
                pod_match = (CURRENT_POD.lower() in page_text) or (pod_alias.lower() in page_text) or (pod_exact and pod_exact.lower() in page_text)

                if pol_match and pod_match:
                    print(f"   ✅ Dữ liệu /sailings ĐÃ KHỚP! Bot xin lại quyền điều khiển...")
                    if MAERSK_AFTER_CAPTCHA_SECONDS > 0:
                        print(f"   ⏳ Nghỉ {MAERSK_AFTER_CAPTCHA_SECONDS:.0f}s sau captcha trước khi đọc kết quả...")
                        time.sleep(MAERSK_AFTER_CAPTCHA_SECONDS)
                    time.sleep(1)
                    raise ManualHandoverComplete()
                else:
                    print(f"\n   ⚠️ Ê Sếp! Đang ở /sailings nhưng dữ liệu KHÔNG KHỚP với {CURRENT_POL} -> {CURRENT_POD}!")
                    print(f"   👉 Sếp vui lòng nhập lại cho chuẩn, rồi bấm Enter lại nhé.")

            else:
                error_msg = None
                for _ in range(5):
                    time.sleep(1)
                    error_msg = check_booking_error_notification()
                    if error_msg:
                        break

                if error_msg and "no sailings" in error_msg.lower():
                    print(f"   🚨 Bot phát hiện banner: Tuyến này không có chuyến! Rút lui sang dòng tiếp theo...")
                    raise NoSailingsException()

                print(f"   ⚠️ Sếp chưa vào tới trang /sailings mà cũng không thấy banner đỏ? URL hiện tại: {current_url[:50]}")

    return 'ok'


# ===================================================================================
# CLICK RADIO (ActionChains — isTrusted=true)
# ===================================================================================
def ensure_radio_checked():
    radio_element = driver.execute_script(_PIERCE_JS + """
        var radios = pierceAll(document, 'mc-radio');
        for (var i = 0; i < radios.length; i++) {
            var r = radios[i];
            var txt = (r.textContent || '').trim();
            if (!txt && r.shadowRoot) txt = (r.shadowRoot.textContent || '').trim();

            if (txt.toLowerCase().indexOf('i am the price owner') !== -1) {
                var inp = r.querySelector('input[type="radio"]');
                if (!inp && r.shadowRoot) inp = r.shadowRoot.querySelector('input[type="radio"]');
                if (inp && inp.checked) return null;
                if (r.hasAttribute('checked') || r.getAttribute('aria-checked') === 'true') return null;

                var lbl = r.querySelector('label');
                if (!lbl && r.shadowRoot) lbl = r.shadowRoot.querySelector('label');
                return lbl || inp || r;
            }
        }
        return null;
    """)

    if radio_element is None:
        print(f"   🔍 Radio check: already checked")
        return 'already checked'

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", radio_element)
        time.sleep(random.uniform(0.1, 0.2))
        ActionChains(driver).move_to_element(radio_element).pause(
            random.uniform(0.1, 0.25)
        ).click().perform()
        print(f"   🔍 Radio check: clicked (isTrusted)")
        return 'clicked'
    except Exception as e:
        print(f"   🔍 Radio check fallback: {e}")
        driver.execute_script("arguments[0].click();", radio_element)
        return 'clicked fallback'

# ===================================================================================
# SELECT DATE (ActionChains — isTrusted=true)
# ===================================================================================
def _open_date_calendar(date_inp):
    """
    Mở calendar picker bằng cách click vào HOST element (mc-input-date),
    không phải inner input — tránh ActionChains lệch vào DIV overlay.
    Fallback: focus + gõ phím ArrowDown để force-open.
    """
    # Lấy host element (mc-input-date) để click — an toàn hơn inner input
    host_el = driver.execute_script("""
        var mids = document.querySelectorAll('mc-input-date');
        if (mids.length > 0) return mids[0];
        // Thử tìm trong shadow DOM nếu bị wrap
        function findHost(root) {
            var all = root.querySelectorAll ? root.querySelectorAll('*') : [];
            for (var i = 0; i < all.length; i++) {
                if (all[i].tagName && all[i].tagName.toLowerCase() === 'mc-input-date') return all[i];
                if (all[i].shadowRoot) { var f = findHost(all[i].shadowRoot); if (f) return f; }
            }
            return null;
        }
        return findHost(document);
    """)

    opened = False

    if host_el:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", host_el)
            time.sleep(random.uniform(0.2, 0.35))
            ActionChains(driver).move_to_element(host_el).pause(
                random.uniform(0.08, 0.15)
            ).click().perform()
            time.sleep(random.uniform(0.5, 0.9))
            opened = True
            print("   📅 Date: click host element OK")
        except Exception as e:
            print(f"   ⚠️ Date: click host lỗi ({e}), thử fallback...")

    if not opened:
        # Fallback 1: focus inner input rồi dùng ActionChains click
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", date_inp)
            time.sleep(random.uniform(0.15, 0.25))
            ActionChains(driver).move_to_element(date_inp).pause(
                random.uniform(0.08, 0.15)
            ).click().perform()
            time.sleep(random.uniform(0.4, 0.7))
            opened = True
            print("   📅 Date: click inner input OK")
        except:
            pass

    if not opened:
        # Fallback 2: JS focus
        driver.execute_script("arguments[0].focus();", date_inp)
        time.sleep(random.uniform(0.3, 0.5))

    # Nếu calendar vẫn chưa mở (data-date rỗng), thử phím ArrowDown để force-open
    calendar_visible = driver.execute_script("""
        function scanDates(root, found) {
            var all = root.querySelectorAll ? root.querySelectorAll('[data-date]') : [];
            for (var i = 0; i < all.length; i++) found.push(all[i].getAttribute('data-date'));
            var allEls = root.querySelectorAll ? root.querySelectorAll('*') : [];
            for (var j = 0; j < allEls.length; j++) {
                if (allEls[j].shadowRoot) scanDates(allEls[j].shadowRoot, found);
            }
        }
        var found = [];
        scanDates(document, found);
        return found.length > 0;
    """)

    if not calendar_visible:
        print("   📅 Date: calendar chưa mở, thử gõ ArrowDown...")
        try:
            ActionChains(driver).send_keys_to_element(date_inp, Keys.ARROW_DOWN).perform()
            time.sleep(random.uniform(0.5, 0.8))
        except:
            driver.execute_script("arguments[0].dispatchEvent(new KeyboardEvent('keydown', {key:'ArrowDown', bubbles:true}));", date_inp)
            time.sleep(random.uniform(0.4, 0.6))


def _find_date_btn(target_date):
    """Tìm nút calendar theo data-date (primary) hoặc aria-label / text content (fallback)."""
    return driver.execute_script("""
        var target = arguments[0];
        // Parse target thành ngày để so sánh text
        var parts = target.split('-');
        var targetDay = parseInt(parts[2], 10);
        var targetMonth = parseInt(parts[1], 10) - 1; // 0-indexed
        var targetYear = parseInt(parts[0], 10);

        function findBtn(root) {
            if (!root) return null;

            // --- Tìm theo data-date ---
            var byDate = root.querySelectorAll ? root.querySelectorAll('[data-date="' + target + '"]') : [];
            for (var i = 0; i < byDate.length; i++) {
                var el = byDate[i];
                if (el.hasAttribute('disabled')) continue;
                var inner = el.shadowRoot ? el.shadowRoot.querySelector('button:not([disabled])') : null;
                return inner || el;
            }

            // --- Tìm theo aria-label chứa ngày đúng ---
            var allEls = root.querySelectorAll ? root.querySelectorAll('button, [role="button"], [role="gridcell"]') : [];
            for (var j = 0; j < allEls.length; j++) {
                var el2 = allEls[j];
                if (el2.hasAttribute('disabled')) continue;
                var label = (el2.getAttribute('aria-label') || '').toLowerCase();
                // aria-label thường kiểu "14 April 2026" hoặc "April 14, 2026"
                if (label.indexOf(String(targetDay)) !== -1
                    && label.indexOf(String(targetYear)) !== -1) {
                    return el2;
                }
                // Nếu text content đúng là số ngày và không có ngày khác trùng
                var txt = (el2.textContent || '').trim();
                if (txt === String(targetDay)) {
                    // Kiểm tra thêm nếu có data-date gần đúng trên ancestor
                    var p = el2.parentElement;
                    for (var k = 0; k < 4 && p; k++, p = p.parentElement) {
                        var dd = p.getAttribute('data-date') || '';
                        if (dd === target) return el2;
                    }
                }
            }

            // --- Đệ quy Shadow DOM ---
            var shadow = root.querySelectorAll ? root.querySelectorAll('*') : [];
            for (var s = 0; s < shadow.length; s++) {
                if (shadow[s].shadowRoot) {
                    var f = findBtn(shadow[s].shadowRoot);
                    if (f) return f;
                }
            }
            return null;
        }
        return findBtn(document);
    """, target_date)


def select_date_plus7():
    from datetime import datetime, timedelta
    target = (datetime.now() + timedelta(days=DATE_OFFSET_DAYS)).strftime("%Y-%m-%d")

    # Lấy inner input để scrollIntoView và fallback focus
    date_inp = driver.execute_script(_PIERCE_JS + """
        var mids = pierceAll(document, 'mc-input-date');
        if (mids.length === 0) return null;
        return pierce(mids[0].shadowRoot || mids[0], 'input[type="text"]');
    """)
    if not date_inp:
        print("   ⚠️ Date: Không tìm thấy mc-input-date")
        return False

    # Mở calendar (click host element — tránh lệch vào DIV overlay)
    _open_date_calendar(date_inp)
    human_sleep(0.6, 1.0)

    for attempt in range(5):
        btn_element = _find_date_btn(target)

        if btn_element:
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_element)
                time.sleep(random.uniform(0.1, 0.2))
                ActionChains(driver).move_to_element(btn_element).pause(
                    random.uniform(0.08, 0.18)
                ).click().perform()
                print(f"   ✅ Date: {target} (today+{DATE_OFFSET_DAYS})")
                return True
            except Exception as e:
                print(f"   ⚠️ Date click lỗi lần {attempt+1}: {e}")

        # DEBUG lần đầu
        if attempt == 0:
            debug = driver.execute_script("""
                function scanDates(root, found) {
                    if (!root) return;
                    var all = root.querySelectorAll ? root.querySelectorAll('[data-date]') : [];
                    for (var i = 0; i < all.length; i++) found.push(all[i].getAttribute('data-date'));
                    var allEls = root.querySelectorAll ? root.querySelectorAll('*') : [];
                    for (var j = 0; j < allEls.length; j++) {
                        if (allEls[j].shadowRoot) scanDates(allEls[j].shadowRoot, found);
                    }
                }
                var found = [];
                scanDates(document, found);
                return found;
            """)
            print(f"   🔍 data-date tìm được trong DOM: {debug[:10]}")
            if not debug:
                # Calendar chưa mở — thử mở lại
                print("   🔄 Calendar chưa render, thử mở lại...")
                _open_date_calendar(date_inp)

        print(f"   🔄 Date: Chưa thấy nút {target}, chờ thêm... (lần {attempt+1})")
        time.sleep(0.8)

    # Fallback: thử Next month
    print(f"   ⚠️ Date: Không tìm được {target}, thử click Next month...")
    next_btn = driver.execute_script("""
        function findNext(root) {
            if (!root) return null;
            var all = root.querySelectorAll ? root.querySelectorAll('*') : [];
            for (var i = 0; i < all.length; i++) {
                var aria = (all[i].getAttribute('aria-label') || '').toLowerCase();
                if (aria.includes('next')) return all[i];
                if (all[i].shadowRoot) {
                    var f = findNext(all[i].shadowRoot);
                    if (f) return f;
                }
            }
            return null;
        }
        return findNext(document);
    """)
    if next_btn:
        try:
            ActionChains(driver).move_to_element(next_btn).pause(0.1).click().perform()
        except:
            driver.execute_script("arguments[0].click();", next_btn)
        time.sleep(0.8)
        # Tìm lại sau khi qua tháng (không đệ quy để tránh vòng lặp vô tận)
        btn_element = _find_date_btn(target)
        if btn_element:
            try:
                ActionChains(driver).move_to_element(btn_element).pause(0.1).click().perform()
                print(f"   ✅ Date: {target} (next month fallback)")
                return True
            except:
                driver.execute_script("arguments[0].click();", btn_element)
                return True

    print(f"   ❌ Date: Bó tay hoàn toàn")
    return False

# ===================================================================================
# CLICK SEARCH
# ===================================================================================
def check_booking_error_notification():
    return driver.execute_script(_PIERCE_JS + """
        function getShadowText(el) {
            var txt = (el.textContent || '') + ' ';
            if (el.shadowRoot) {
                var children = el.shadowRoot.childNodes;
                for (var i=0; i<children.length; i++) {
                    txt += getShadowText(children[i]);
                }
            }
            return txt.toLowerCase();
        }
        function findNotif(root) {
            var els = root.querySelectorAll('*');
            for (var i = 0; i < els.length; i++) {
                var el = els[i];
                if (el.tagName && el.tagName.toLowerCase() === 'mc-notification') {
                    var text = getShadowText(el);
                    if(text.includes('error')
                       || text.includes('system is busy')
                       || text.includes('no sailings')
                       || text.includes('there are no sailings')
                       || text.includes('no service')) {
                        return text;
                    }
                }
                if (el.shadowRoot) {
                    var res = findNotif(el.shadowRoot);
                    if (res) return res;
                }
            }
            return null;
        }
        return findNotif(document);
    """)


def smart_click_search():
    max_attempts = 5
    for attempt in range(max_attempts):
        current_url = driver.current_url

        if "/book/" in current_url and "/sailings" not in current_url:
            # Scroll tự nhiên thay vì JS scroll thẳng xuống đáy
            human_scroll()

            # Chờ ngẫu nhiên — giống người đọc lại form
            wait_time = random.uniform(MAERSK_PRE_SEARCH_MIN_SECONDS, MAERSK_PRE_SEARCH_MAX_SECONDS)
            print(f"   ⏳ Đọc lại form {wait_time:.1f}s trước khi Search...")
            steps = random.randint(2, 4)
            for _ in range(steps):
                time.sleep(wait_time / steps)
                if random.random() < 0.6:
                    try:
                        action = ActionChains(driver)
                        action.move_by_offset(
                            random.randint(-120, 120),
                            random.randint(-80, 80)
                        ).pause(random.uniform(0.1, 0.3)).perform()
                    except:
                        pass

            btn_element = driver.execute_script(_PIERCE_JS + """
                var form = document.querySelector('form');
                if (!form) return null;
                var btns = form.querySelectorAll('mc-button');
                if (btns.length === 0) return null;

                for(var i = 0; i < btns.length; i++) {
                    var txt = btns[i].textContent.toLowerCase().trim();
                    if(txt === 'search' || txt === 'continue to book' || txt === 'continue') {
                        var innerBtn = btns[i].shadowRoot ? btns[i].shadowRoot.querySelector('button') : btns[i].querySelector('button');
                        if(innerBtn && !innerBtn.hasAttribute('disabled')) return innerBtn;
                        if(!btns[i].hasAttribute('disabled')) return btns[i];
                    }
                }
                var lastBtn = btns[btns.length - 1];
                var innerLast = lastBtn.shadowRoot ? lastBtn.shadowRoot.querySelector('button') : lastBtn.querySelector('button');
                if(innerLast && !innerLast.hasAttribute('disabled')) return innerLast;
                if(!lastBtn.hasAttribute('disabled')) return lastBtn;

                return null;
            """)

            if not btn_element:
                print("   ⚠️ Không tìm thấy nút Search/Continue, web lag mất layout!")
                return 'ERROR_RETRY'

            print(f"   🔘 Đang rà chuột vật lý vào nút Search (Lần {attempt+1})...")
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_element)
                time.sleep(random.uniform(0.2, 0.4))
                ActionChains(driver).move_to_element(btn_element).pause(
                    random.uniform(0.15, 0.3)
                ).click().perform()
            except Exception as e:
                driver.execute_script("arguments[0].click();", btn_element)

            print("   ⏳ Đang rình xem hCaptcha có nhảy ra không...")
            captcha_status = handle_captcha_if_present()
            if captcha_status == 'api_error':
                return 'API_ERROR'

            error_msg = None
            for _ in range(5):
                time.sleep(1)
                error_msg = check_booking_error_notification()
                if error_msg: break

            if error_msg:
                error_msg = error_msg.lower()
                if "no sailings" in error_msg:
                    print(f"   🚨 Bị chặn cửa: There are no sailings for your search. -> GHI 'NO SERVICE'")
                    return 'NO_SAILINGS'
                elif "system is busy" in error_msg or "error" in error_msg:
                    print(f"   🚨 Server Busy. Nghỉ 20s rồi đập nút lại...")
                    time.sleep(20)
                    continue
                else:
                    print(f"   🚨 Lỗi lạ banner: {error_msg}. -> RESET LẠI TRANG CHỦ!")
                    return 'ERROR_RETRY'

            time.sleep(4)

        elif "/sailings" in current_url:
            print("   🚀 Đã vào trang /sailings thành công!")
            return 'SUCCESS'

    return 'ERROR_RETRY'

# ===================================================================================
# INPUT HELPERS
# ===================================================================================
def set_input_value(inp, value):
    driver.execute_script("""
        var inp = arguments[0], val = arguments[1];
        inp.click(); inp.focus();
        var setter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value').set;
        setter.call(inp, val);
        inp.dispatchEvent(new Event('input', {bubbles:true}));
        inp.dispatchEvent(new Event('change', {bubbles:true}));
    """, inp, value)

def get_mc_options_text():
    return driver.execute_script(_PIERCE_JS + """
        var opts = pierceAll(document, 'mc-option');
        var results = [];
        for (var i = 0; i < opts.length; i++) {
            var clone = opts[i].cloneNode(true);
            clone.querySelectorAll('template').forEach(function(t){ t.remove(); });
            var txt = clone.textContent.trim();
            if (txt && txt.length > 2) results.push(txt);
        }
        return [...new Set(results)];
    """)

def click_mc_option_text(text):
    return driver.execute_script(_PIERCE_JS + """
        var target = arguments[0];
        var opts = pierceAll(document, 'mc-option');
        for (var i = 0; i < opts.length; i++) {
            var clone = opts[i].cloneNode(true);
            clone.querySelectorAll('template').forEach(function(t){ t.remove(); });
            if (clone.textContent.trim() === target) {
                var btn = opts[i].shadowRoot
                    ? opts[i].shadowRoot.querySelector('button')
                    : opts[i].querySelector('button');
                if (btn) { btn.click(); return true; }
                opts[i].click(); return true;
            }
        }
        return false;
    """, text)

def has_no_suggestions():
    return driver.execute_script("""
        var walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
        var node;
        while (node = walker.nextNode()) {
            if (node.textContent.trim() === 'No suggestions found') return true;
        }
        return false;
    """)

def normalize_port(raw):
    return re.sub(r'\s*\([^)]*\)', '', raw).strip().lower()

def match_port(search_name, country, options):
    country_lower = country.strip().lower()
    port_lower    = search_name.strip().lower()
    filtered = [o for o in options if country_lower in o.lower()] or options
    for opt in filtered:
        opt_port = normalize_port(opt).split(',')[0].strip()
        if opt_port == port_lower:
            return opt
    for opt in filtered:
        if port_lower in normalize_port(opt):
            return opt
    return None

def get_port_options(nth):
    return driver.execute_script(_PIERCE_JS + """
        var sms = pierceAll(document, 'mc-c-location-servicemode');
        var sm = sms[arguments[0]];
        if(!sm) return [];
        var opts = pierceAll(sm.shadowRoot || sm, 'mc-option');
        var res = [];
        for(var i=0; i<opts.length; i++){
            var clone = opts[i].cloneNode(true);
            clone.querySelectorAll('template').forEach(function(t){ t.remove(); });
            var txt = clone.textContent.trim();
            if(txt.length > 2) res.push(txt);
        }
        return res;
    """, nth)

def wait_for_matching_port_option(nth, search_name, country, exact_option=None, timeout=9):
    end = time.time() + timeout
    last_options = []
    while time.time() < end:
        options = get_port_options(nth)
        if options:
            last_options = options
            if exact_option and exact_option in options:
                return exact_option, options
            best_opt = match_port(search_name, country, options)
            if best_opt:
                return best_opt, options
        time.sleep(0.5)
    return None, last_options

# ===================================================================================
# TYPE_PORT
# ===================================================================================
def find_port_input(nth):
    sm_idx = nth + 1
    for _ in range(3):
        el = driver.execute_script(_PIERCE_JS + """
            var smIdx = arguments[0];
            var od = pierce(document, 'mc-c-origin-destination');
            if (!od) return null;
            var sms = pierceAll(od.shadowRoot || od, 'mc-c-location-servicemode');
            var sm = sms[smIdx - 1];
            if (!sm) return null;
            return pierce(sm.shadowRoot || sm, 'input[type="text"]');
        """, sm_idx)
        if el:
            return el
        time.sleep(0.5)
    return None

def type_port(nth, port_name, country, label="PORT", pod_country=""):
    search_name, exact_option = resolve_maersk_alias(port_name, pod_country)

    inp = find_port_input(nth)
    if not inp:
        print(f"   ⚠️ Không tìm thấy ô nhập liệu {label}")
        return False

    current_val = driver.execute_script("return arguments[0].value || '';", inp).strip()
    if current_val:
        if exact_option:
            if exact_option.lower() == current_val.lower():
                print(f"   ✅ {label} giữ nguyên: {current_val[:60]}")
                return True
        else:
            if search_name.lower() in current_val.lower():
                print(f"   ✅ {label} giữ nguyên: {current_val[:60]}")
                return True

    for attempt in range(3):
        inp = find_port_input(nth)
        if not inp: continue

        # Xóa cảng cũ
        driver.execute_script(_PIERCE_JS + """
            var od = pierce(document, 'mc-c-origin-destination');
            if(!od) return false;
            var sms = pierceAll(od.shadowRoot || od, 'mc-c-location-servicemode');
            var sm = sms[arguments[0]];
            if(!sm) return false;
            var ta = pierce(sm.shadowRoot || sm, 'mc-typeahead');

            var clearBtn = pierce(sm.shadowRoot || sm, 'mc-button') ||
                           (ta ? pierce(ta.shadowRoot || ta, 'mc-button') : null) ||
                           (ta ? pierce(ta.shadowRoot || ta, 'mc-icon-button') : null);

            if(clearBtn) {
                var btn = clearBtn.shadowRoot ? clearBtn.shadowRoot.querySelector('button') : clearBtn.querySelector('button');
                if(btn) btn.click(); else clearBtn.click();
            }

            if(ta) {
                var inps = pierceAll(ta.shadowRoot || ta, 'input[type="text"]');
                if(inps.length > 0) {
                    inps[0].value = '';
                    inps[0].dispatchEvent(new Event('input', { bubbles: true }));
                    inps[0].dispatchEvent(new Event('change', { bubbles: true }));
                }
            }
            return true;
        """, nth)

        human_move_to(inp)
        try:
            ActionChains(driver).click(inp).perform()
        except:
            driver.execute_script("arguments[0].focus(); arguments[0].click();", inp)
        inp.send_keys(Keys.CONTROL + "a")
        inp.send_keys(Keys.BACK_SPACE)
        inp.send_keys(Keys.DELETE)
        micro_pause()

        val_after = driver.execute_script("return arguments[0].value || '';", inp)
        if val_after != "":
            print(f"   ⚠️ Lần {attempt+1}: Thất bại khi xóa field {label}. Thử lại...")
            continue

        # Gõ cảng mới
        inp.send_keys(search_name)
        human_sleep(1.2, 2.0)
        typed_val = driver.execute_script("return arguments[0].value || '';", inp).strip()
        if search_name.lower() not in typed_val.lower():
            print(f"   ⚠️ {label}: typed value lệch '{typed_val}' != '{search_name}', thử lại...")
            driver.execute_script("arguments[0].focus(); arguments[0].click();", inp)
            inp.send_keys(Keys.CONTROL + "a")
            inp.send_keys(Keys.BACK_SPACE)
            inp.send_keys(search_name)
            human_sleep(1.2, 2.0)

        best_opt, options = wait_for_matching_port_option(
            nth, search_name, country, exact_option=exact_option, timeout=9
        )

        if not options:
            print(f"   ⚠️ {label}: Không tìm thấy popup options!")
        else:
            if best_opt:
                # Lấy element ra — ActionChains click (isTrusted=true)
                opt_element = driver.execute_script(_PIERCE_JS + """
                    var sms = pierceAll(document, 'mc-c-location-servicemode');
                    var sm = sms[arguments[0]];
                    var opts = pierceAll(sm.shadowRoot || sm, 'mc-option');
                    for(var i=0; i<opts.length; i++){
                        var clone = opts[i].cloneNode(true);
                        clone.querySelectorAll('template').forEach(function(t){ t.remove(); });
                        if(clone.textContent.trim() === arguments[1]){
                            var btn = opts[i].shadowRoot ? opts[i].shadowRoot.querySelector('button') : opts[i].querySelector('button');
                            return btn || opts[i];
                        }
                    }
                    return null;
                """, nth, best_opt)

                if opt_element:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", opt_element)
                        time.sleep(random.uniform(0.1, 0.2))
                        ActionChains(driver).move_to_element(opt_element).pause(
                            random.uniform(0.08, 0.18)
                        ).click().perform()
                        human_sleep(0.3, 0.6)
                        print(f"   ✅ {label}: {best_opt[:60]}")
                        return True
                    except Exception as e:
                        print(f"   ⚠️ ActionChains click lỗi: {e} → fallback JS")
                        driver.execute_script("arguments[0].click();", opt_element)
                        human_sleep(0.3, 0.6)
                        print(f"   ✅ {label} (fallback): {best_opt[:60]}")
                        return True
            else:
                print(f"   ⚠️ {label}: Tên không khớp '{search_name}' trong {options[:3]}")

        print(f"   🔄 Lần {attempt+1} web không đổ popup cho {label}. Thử lại...")
        time.sleep(0.5)

    print(f"   ❌ LỖI TRẦM TRỌNG: Đã thử 3 lần nhưng không nhập được {label}!")
    return False

# ===================================================================================
# TYPE_COMMODITY (ActionChains — isTrusted=true)
# ===================================================================================
def type_commodity():
    SEARCH_TEXT   = "garments"
    TARGET_OPTION = "Garments, apparel, new"

    comm = driver.execute_script(_PIERCE_JS + "return pierce(document, 'mc-c-commodity');")
    if not comm:
        print("   ⚠️ Commodity: Không tìm thấy mc-c-commodity")
        return False

    inp = driver.execute_script(_PIERCE_JS + """
        var sr = arguments[0].shadowRoot || arguments[0];
        return pierce(sr, 'input[type="text"]');
    """, comm)
    if not inp:
        print("   ⚠️ Commodity: Không tìm thấy input")
        return False

    current = driver.execute_script("return arguments[0].value || '';", inp)
    if current and TARGET_OPTION.lower() in current.lower():
        print(f"   ✅ Commodity giữ nguyên: {current.strip()[:50]}")
        return True

    human_move_to(inp)
    human_sleep(0.1, 0.2)
    try:
        ActionChains(driver).click(inp).perform()
    except:
        driver.execute_script("arguments[0].click(); arguments[0].focus();", inp)
    micro_pause()
    inp.send_keys(Keys.CONTROL + "a")
    inp.send_keys(Keys.DELETE)
    human_sleep(0.1, 0.2)
    typo_then_correct(inp, SEARCH_TEXT)
    human_sleep(0.6, 1.0)

    try:
        WebDriverWait(driver, 5).until(
            lambda d: driver.execute_script(_PIERCE_JS + """
                var comm = arguments[0];
                var sr = comm.shadowRoot || comm;
                return pierceAll(sr, 'mc-option').length > 0;
            """, comm)
        )
    except:
        pass

    human_sleep(0.2, 0.4)

    opt_element = driver.execute_script(_PIERCE_JS + """
        var comm   = arguments[0];
        var target = arguments[1];
        var sr     = comm.shadowRoot || comm;
        var opts   = pierceAll(sr, 'mc-option');
        for (var i = 0; i < opts.length; i++) {
            var clone = opts[i].cloneNode(true);
            clone.querySelectorAll('template').forEach(function(t){ t.remove(); });
            if (clone.textContent.trim().toLowerCase().indexOf(target.toLowerCase()) !== -1) {
                var btn = opts[i].shadowRoot
                    ? opts[i].shadowRoot.querySelector('button')
                    : opts[i].querySelector('button');
                return btn || opts[i];
            }
        }
        if (opts.length > 0) {
            var btn0 = opts[0].shadowRoot
                ? opts[0].shadowRoot.querySelector('button')
                : opts[0].querySelector('button');
            return btn0 || opts[0];
        }
        return null;
    """, comm, TARGET_OPTION)

    if opt_element:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", opt_element)
            time.sleep(random.uniform(0.1, 0.2))
            ActionChains(driver).move_to_element(opt_element).pause(
                random.uniform(0.08, 0.18)
            ).click().perform()
            human_sleep(0.3, 0.6)
            print(f"   ✅ Commodity: {TARGET_OPTION}")
            return True
        except Exception as e:
            print(f"   ⚠️ Commodity ActionChains lỗi: {e} → fallback JS click")
            driver.execute_script("arguments[0].click();", opt_element)
            human_sleep(0.3, 0.6)
            print(f"   ✅ Commodity (fallback): {TARGET_OPTION}")
            return True

    print(f"   ⚠️ Commodity: Không tìm thấy option nào")
    return False

# ===================================================================================
# SELECT CONTAINERS
# ===================================================================================
def set_container_slot(slot_idx, label):
    cs = driver.execute_script(_PIERCE_JS + "return pierce(document, 'mc-c-container-select');")
    if not cs: return False

    si = driver.execute_script(_PIERCE_JS + """
        var sis = pierceAll(arguments[0].shadowRoot || arguments[0], 'mc-c-container-selection-input');
        return sis[arguments[1] - 1] || null;
    """, cs, slot_idx)
    if not si: return False

    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", si)
    time.sleep(0.3)
    driver.execute_script(_PIERCE_JS + """
        var ta = pierce(arguments[0].shadowRoot || arguments[0], 'mc-typeahead');
        if(ta) {
            var inp = pierce(ta.shadowRoot || ta, 'input[type="text"]');
            if(inp) { inp.focus(); inp.click(); }
        }
    """, si)

    if "20" in label:
        keyword = "20"
    elif "High" in label:
        keyword = "High"
    else:
        keyword = "40 Dry Standard"

    is_selected = False
    for _ in range(3):
        time.sleep(0.8)
        is_selected = driver.execute_script(_PIERCE_JS + """
            var si = arguments[0];
            var kw = arguments[1];
            var opts = pierceAll(si.shadowRoot || si, 'mc-option');
            for(var i=0; i<opts.length; i++) {
                if(opts[i].textContent.includes(kw)) {
                    var btn = null;
                    if (opts[i].shadowRoot) {
                        btn = opts[i].shadowRoot.querySelector('button');
                    }
                    if (btn) {
                        btn.click();
                    } else {
                        opts[i].click();
                    }
                    return true;
                }
            }
            return false;
        """, si, keyword)

        if is_selected: break
        driver.execute_script("arguments[0].click();", si)

    if not is_selected:
        print(f"   ❌ Slot {slot_idx}: Không chọn được loại cont {label}")
        return False

    print(f"   ✅ Slot {slot_idx} đã chọn: {label}")
    time.sleep(0.5)

    try:
        weight_input = driver.execute_script(_PIERCE_JS + """
            var si = arguments[0];
            var mci = pierce(si.shadowRoot || si, 'mc-input');
            if(!mci) return null;
            return pierce(mci.shadowRoot || mci, 'input');
        """, si)

        if weight_input:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", weight_input)
            weight_input.click()
            weight_input.send_keys(Keys.CONTROL + "a")
            weight_input.send_keys(Keys.BACK_SPACE)
            time.sleep(0.2)
            for char in "22222":
                weight_input.send_keys(char)
                time.sleep(0.04)
            print(f"      -> Đã nhập cân nặng: 22222")
            return True
        else:
            print(f"   ⚠️ Slot {slot_idx}: Không tìm thấy ô nhập cân nặng!")
            return False

    except Exception as e:
        print(f"   ⚠️ Slot {slot_idx}: Lỗi khi gõ cân nặng")
        return False

def select_containers(country=""):
    TARGETS = [
        {"label": "20 Dry Standard"},
        {"label": "40 Dry High"},
    ]

    current_labels = driver.execute_script(_PIERCE_JS + """
        var cs = pierce(document, 'mc-c-container-select');
        if (!cs) return [];
        var sis = pierceAll(cs.shadowRoot || cs, 'mc-c-container-selection-input');
        var current = [];
        for (var i = 0; i < sis.length; i++) {
            var ta = pierce(sis[i].shadowRoot || sis[i], 'mc-typeahead');
            if (ta) {
                var inp = pierce(ta.shadowRoot || ta, 'input[type="text"]');
                if (inp) current.push(inp.value.trim());
            }
        }
        return current;
    """)
    print(f"   🔍 Quét web đang có sẵn cont: {current_labels}")

    cs = driver.execute_script(_PIERCE_JS + "return pierce(document, 'mc-c-container-select');")

    for i, target in enumerate(TARGETS):
        slot_idx = i + 1
        label = target["label"]

        if i < len(current_labels):
            if label.lower() in current_labels[i].lower() and current_labels[i] != "":
                print(f"   ✅ Slot {slot_idx} chuẩn khỏi chỉnh: {label}")
            else:
                print(f"   🔄 Đổi Slot {slot_idx}: đang là '{current_labels[i]}' → thành {label}")
                if not set_container_slot(slot_idx, label): return False
        else:
            add_btn = driver.execute_script(_PIERCE_JS + "var mcbtn = pierce(arguments[0].shadowRoot || arguments[0], 'section mc-button'); if(!mcbtn) return null; return pierce(mcbtn.shadowRoot || mcbtn, 'button') || mcbtn;", cs)
            if add_btn:
                print(f"   🔘 Bấm Add → chờ tạo slot {slot_idx}")
                driver.execute_script("arguments[0].click();", add_btn)
                human_sleep(0.5, 0.8)
                if not set_container_slot(slot_idx, label): return False
            else:
                print(f"   ⚠️ Lỗi: Không tìm thấy nút Add cont!"); return False

    if len(current_labels) > len(TARGETS):
        for idx in range(len(current_labels), len(TARGETS), -1):
            delete_container_slot(idx)

    return True

# ===================================================================================
# RESULTS PAGE
# ===================================================================================
def wait_for_sailings_page(timeout=40):
    start_time = time.time()
    print("   ⏳ Đang đợi kết quả sailings ổn định...")

    while time.time() - start_time < timeout:
        curr_url = driver.current_url

        if "/book/" in curr_url and "/sailings" not in curr_url:
            print("   ⚠️ Phát hiện bị điều hướng ngược về trang nhập liệu! Hủy đợi giá.")
            return False

        articles = driver.execute_script("return document.querySelectorAll('article').length;")
        if articles > 0:
            time.sleep(1.5)
            return True

        time.sleep(1.5)
    return False

def read_price_divs():
    results = driver.execute_script("""
        var articles = document.querySelectorAll('article');
        if (articles.length === 0) return ['NO ARTICLES'];
        var out = [];
        for (var i = 0; i < articles.length; i++) {
            var span = articles[i].querySelector('span.status-label');
            if (span && span.innerText.trim()) {
                out.push(span.innerText.trim());
            } else {
                var priceEl = articles[i].querySelector('[data-test="sailing-price"] p');
                var price = priceEl ? priceEl.innerText.trim() : '';
                if (price && price.includes('USD')) {
                    out.push('HAS_PRICE: ' + price);
                }
            }
        }
        if (out.length === 0) return ['NO STATUS LABELS'];
        return out;
    """)
    return results or []

def go_back():
    clicked = driver.execute_script("""
        var el = document.querySelector(
            'mc-step-indicator mc-step-indicator-item:first-child span div'
        );
        if (el) { el.click(); return 'clicked step indicator'; }
        return false;
    """)
    if not clicked:
        driver.execute_script(_PIERCE_JS + """
            var stepEl = document.querySelector('mc-step-indicator');
            if (!stepEl) return 'no mc-step-indicator';
            var items = stepEl.querySelectorAll('mc-step-indicator-item');
            if (!items || items.length === 0) {
                items = stepEl.shadowRoot
                    ? stepEl.shadowRoot.querySelectorAll('mc-step-indicator-item')
                    : [];
            }
            if (!items || items.length === 0) return 'no items';
            var first = items[0];
            var span = first.querySelector('span');
            if (span) { span.click(); return 'clicked span'; }
            first.click(); return 'clicked item';
        """)
    print(f"   🔙 go_back: {clicked}")
    human_sleep(1.5, 2.5)

# ===================================================================================
# ĐỌC GIÁ CHI TIẾT
# ===================================================================================
def get_priced_articles():
    return driver.execute_script("""
        var articles = document.querySelectorAll('article');
        var UNAVAIL  = arguments[0];
        var out = [];
        for (var i = 0; i < articles.length; i++) {
            var span = articles[i].querySelector('span.status-label');
            if (span) {
                var t = span.innerText.trim().toLowerCase();
                var bad = false;
                for (var j = 0; j < UNAVAIL.length; j++) {
                    if (t.includes(UNAVAIL[j])) { bad = true; break; }
                }
                if (bad) continue;
            }
            var priceEl = articles[i].querySelector('[data-test="sailing-price"] p');
            var price   = priceEl ? priceEl.innerText.trim() : '';

            var etd = "N/A";
            var tt = "N/A";

            var etdNode = articles[i].querySelector('header dl div:nth-child(1) dd time');
            if(etdNode) etd = etdNode.innerText.trim();

            var ttNode = articles[i].querySelector('header dl div:nth-child(4) dd mc-c-duration-display');
            if(!ttNode) ttNode = articles[i].querySelector('header dl div:nth-child(4) dd time');
            if(ttNode) tt = ttNode.innerText.trim();

            if (price && price.includes('USD')) {
                out.push({idx: i, price: price, etd: etd, tt: tt});
            }
        }
        return out;
    """, UNAVAILABLE_STATUSES) or []

def click_price_breakdown(article_idx):
    driver.execute_script("""
        var pd = document.querySelector('[data-test="price-details"]');
        if (pd && pd.offsetHeight > 0) {
            var spans = document.querySelectorAll('span.hyperlink-button');
            for (var i = 0; i < spans.length; i++) {
                if (spans[i].innerText.toLowerCase().includes('price breakdown')) {
                    spans[i].click();
                    break;
                }
            }
        }
    """)
    time.sleep(0.4)

    result = driver.execute_script("""
        var articles = document.querySelectorAll('article');
        var art = articles[arguments[0]];
        if (!art) return 'no article';
        var spans = art.querySelectorAll('span.hyperlink-button');
        for (var i = 0; i < spans.length; i++) {
            if (spans[i].innerText.toLowerCase().includes('price breakdown')) {
                spans[i].click();
                return 'clicked';
            }
        }
        return 'not found';
    """, article_idx)
    print(f"   💰 Click price breakdown: {result}")
    return 'clicked' in str(result)

def read_freight_charges():
    human_sleep(1.5, 2.5)

    for _ in range(10):
        has_table = driver.execute_script("""
            var pd = document.querySelector('[data-test="price-details"]');
            if (!pd) return false;
            var t = pd.querySelector('mc-c-table');
            return t && t.shadowRoot && t.shadowRoot.querySelector('tbody tr') ? true : false;
        """)
        if has_table:
            break
        print(f"   ⏳ Chờ bảng giá render...")
        time.sleep(1)

    charges = driver.execute_script("""
        var priceDetail = document.querySelector('[data-test="price-details"]');
        if (!priceDetail) return {freight: [], origin_thc: [], has_ths_origin: false, debug: 'no price-details'};

        var mcTable = priceDetail.querySelector('mc-c-table');
        if (!mcTable || !mcTable.shadowRoot)
            return {freight: [], origin_thc: [], has_ths_origin: false, debug: 'no table/shadowRoot'};

        var rows = mcTable.shadowRoot.querySelectorAll('tbody tr');
        if (rows.length === 0)
            return {freight: [], origin_thc: [], has_ths_origin: false, debug: 'no rows'};

        var freightCharges = [];
        var originTHSCharges = [];
        var inFreight    = true;
        var inOrigin     = false;
        var hasTHSOrigin = false;

        for (var i = 0; i < rows.length; i++) {
            var tds = rows[i].querySelectorAll('td');
            if (tds.length < 6) continue;

            var name      = tds[0].innerText.trim();
            var basis     = tds[1].innerText.trim();
            var subtext   = tds[1].querySelector('.mds-table__subtext');
            var contType  = subtext ? subtext.innerText.trim() : '';
            var currency  = tds[3].innerText.trim();
            var unitPrice = tds[4].innerText.trim();

            if (name === 'Origin charges')     { inFreight = false; inOrigin = true;  continue; }
            if (name === 'Destination charges' || name === 'Other charges') {
                inFreight = false; inOrigin = false; continue;
            }
            if (basis === 'Basis') continue;

            if (inOrigin && name.toLowerCase().includes('terminal handling service - origin')) {
                hasTHSOrigin = true;
                var originPrice = parseFloat(unitPrice.replace(/[^0-9.]/g, '')) || 0;
                if (basis !== 'Bill of Lading' && currency && originPrice) {
                    originTHSCharges.push({
                        name: name, cont_type: contType, currency: currency,
                        unit_price: originPrice, unit_price_raw: unitPrice
                    });
                }
            }

            if (inFreight && basis !== 'Bill of Lading' && name && currency && unitPrice) {
                var price = parseFloat(unitPrice.replace(/[^0-9.]/g, '')) || 0;
                freightCharges.push({
                    name:           name,
                    cont_type:      contType,
                    currency:       currency,
                    unit_price:     price,
                    unit_price_raw: unitPrice
                });
            }
        }

        return {
            freight:        freightCharges,
            origin_thc:     originTHSCharges,
            has_ths_origin: hasTHSOrigin,
            debug:          'ok, rows=' + rows.length + ', freight=' + freightCharges.length
        };
    """)

    print(f"   🔍 Table debug: {charges.get('debug', '?')}")
    return charges or {'freight': [], 'origin_thc': [], 'has_ths_origin': False}

def calculate_of(charges_data, cont_label, include_origin_thc=False):
    freight  = charges_data.get('freight', [])
    has_ths  = charges_data.get('has_ths_origin', False)

    formula_parts = []
    for c in freight:
        if cont_label.lower() in c.get('cont_type', '').lower():
            formula_parts.append(charge_amount_to_usd(c['unit_price'], c.get('currency') or 'USD'))
    if include_origin_thc:
        for c in charges_data.get('origin_thc', []):
            if cont_label.lower() in c.get('cont_type', '').lower():
                usd_amount = charge_amount_to_usd(c['unit_price'], c.get('currency') or 'USD')
                formula_parts.append(usd_amount)
                print(f"      + O.THC CHINA [{c.get('cont_type')}]: {usd_amount:.2f} USD")
    total = sum(formula_parts)

    for c in freight:
        if cont_label.lower() in c.get('cont_type', '').lower():
            print(f"      + {c['name']} [{c['cont_type']}]: {c['unit_price_raw']} {c['currency']}")

    if not has_ths:
        print(f"   ⚠️ Không có THS-Origin → GIỮ NGUYÊN GIÁ (Sẽ ghi INCLUDED O.THC)")

    print(f"   💵 O/F: USD {total:.2f}")
    return total, has_ths, _excel_formula_from_parts(formula_parts)

ROW_OVERWRITE_STATE = {}

# ===================================================================================
# LỌC 9 QUY TẮC ETD
# ===================================================================================
def apply_9_golden_rules(danh_sach_chuyen):
    if not danh_sach_chuyen:
        return [], "", ""
    danh_sach_chuyen = [c for c in danh_sach_chuyen if etd_within_max(c.get("etd_dt"))]
    if not danh_sach_chuyen:
        return [], "", ""

    danh_sach_chuyen.sort(key=lambda x: (x["etd_dt"], x["tt_days"]))
    list_loc_trung = []
    seen_dates = set()
    for c in danh_sach_chuyen:
        if c["etd_dt"] not in seen_dates:
            list_loc_trung.append(c)
            seen_dates.add(c["etd_dt"])

    etd_dat_chuan = []
    if list_loc_trung:
        ngan_nhat_global = min(c["tt_days"] for c in list_loc_trung)
        first_date = list_loc_trung[0]["etd_dt"]
        for c in list_loc_trung:
            if len(etd_dat_chuan) >= 3: break
            if len(etd_dat_chuan) > 0 and (c["etd_dt"] - etd_dat_chuan[-1]["etd_dt"]).days < 2: continue
            if (c["etd_dt"] - first_date).days <= 9 and (c["tt_days"] <= ngan_nhat_global + 10):
                etd_dat_chuan.append(c)

    format_str = ""
    num = len(etd_dat_chuan)

    if num == 1:
        format_str = etd_dat_chuan[0]["etd_dt"].strftime("%d-%b")
    elif num == 2:
        format_str = f"{etd_dat_chuan[0]['etd_dt'].strftime('%d-%b')} & {etd_dat_chuan[1]['etd_dt'].strftime('%d-%b')}"
    elif num >= 3:
        day1 = etd_dat_chuan[0]["etd_dt"].strftime("%d")
        day2 = etd_dat_chuan[1]["etd_dt"].strftime("%d")
        day3_month = etd_dat_chuan[2]["etd_dt"].strftime("%d-%b")
        format_str = f"{day1}, {day2}, {day3_month}"

    all_tt = [c["tt_days"] for c in etd_dat_chuan]
    tt_min, tt_max = min(all_tt), max(all_tt)
    str_tt = f"{tt_min}" if tt_min == tt_max else f"{tt_min}-{tt_max}"

    return etd_dat_chuan, format_str, str_tt

# ===================================================================================
# GHI KẾT QUẢ VÀO EXCEL
# ===================================================================================
def write_result_to_excel(row_index, cont_label, value, note="", etd="", tt="", remark="", *args, **kwargs):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        col_map = {
            "20 Dry Standard": 6,
            "40 Dry Standard": 7,
            "40 Dry High": 8
        }

        col_idx = col_map.get(cont_label)
        if col_idx:
            ws.cell(row=row_index, column=col_idx).value = value

        if etd:
            ws.cell(row=row_index, column=9).value = etd
        if tt:
            ws.cell(row=row_index, column=10).value = tt
        if note:
            ws.cell(row=row_index, column=11).value = note
        if remark:
            ws.cell(row=row_index, column=12).value = remark

        wb.save(excel_path)
    except Exception as e:
        print(f"   ⚠️ Lỗi ghi Excel: {e}")


def _maersk_target_date_key():
    from datetime import timedelta
    return (datetime.now() + timedelta(days=DATE_OFFSET_DAYS)).strftime("%Y-%m-%d")


def _maersk_cache_key(country, pol, pod):
    return "|".join([
        datetime.now().strftime("%Y-%m-%d"),
        _maersk_target_date_key(),
        str(country or "").strip().upper(),
        str(pol or "").strip().upper(),
        str(pod or "").strip().upper(),
    ])


def _load_maersk_cache():
    if not MAERSK_CACHE_ENABLED:
        return {}
    try:
        with open(MAERSK_CACHE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _save_maersk_cache(data):
    if not MAERSK_CACHE_ENABLED:
        return
    try:
        with open(MAERSK_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"   ?? Kh?ng l?u ???c Maersk cache: {e}")


def _read_excel_row_result(row_index):
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb.active
    values = {str(col): ws.cell(row=row_index, column=col).value for col in range(6, 13)}
    wb.close()
    return values


def _is_cacheable_maersk_result(values):
    text = " ".join(str(v or "").upper() for v in values.values())
    if not any(values.get(str(col)) not in (None, "") for col in (6, 7, 8)):
        return False
    bad_markers = ("RATE LIMIT", "NO SERVICE", "ERROR", "L?I", "LOI WEB", "N/A")
    return not any(marker in text for marker in bad_markers)


def apply_maersk_cache(row_index, country, pol, pod):
    cache = _load_maersk_cache()
    item = cache.get(_maersk_cache_key(country, pol, pod))
    if not item:
        return False
    values = item.get("values") or {}
    if not _is_cacheable_maersk_result(values):
        return False
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    for col_s, value in values.items():
        ws.cell(row=row_index, column=int(col_s)).value = value
    wb.save(excel_path)
    print(f"   ?? D?ng Maersk cache, kh?ng g?i web/API cho d?ng {row_index}.")
    return True


def save_maersk_cache(row_index, country, pol, pod):
    if not MAERSK_CACHE_ENABLED:
        return
    values = _read_excel_row_result(row_index)
    if not _is_cacheable_maersk_result(values):
        return
    cache = _load_maersk_cache()
    cache[_maersk_cache_key(country, pol, pod)] = {
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "values": values,
    }
    _save_maersk_cache(cache)
    print("   ?? ?? l?u Maersk cache cho route n?y.")

EUROPE_COUNTRIES = {"ALBANIA", "ANDORRA", "AUSTRIA", "BELARUS", "BELGIUM", "BOSNIA AND HERZEGOVINA", "BULGARIA", "CROATIA", "CYPRUS", "CZECH REPUBLIC", "DENMARK", "ESTONIA", "FINLAND", "FRANCE", "GERMANY", "GREECE", "HUNGARY", "ICELAND", "IRELAND", "ITALY", "LATVIA", "LITHUANIA", "LUXEMBOURG", "MALTA", "MOLDOVA", "MONACO", "MONTENEGRO", "NETHERLANDS", "NORWAY", "POLAND", "PORTUGAL", "ROMANIA", "SERBIA", "SLOVAKIA", "SLOVENIA", "SPAIN", "SWEDEN", "SWITZERLAND", "UKRAINE", "UNITED KINGDOM", "UK"}

def process_priced_articles(priced_articles, cont_labels, country="", pod=""):
    if not isinstance(cont_labels, list):
        cont_labels = [cont_labels]

    if not priced_articles:
        print(f"   ⚠️ Không có giá → Bỏ trống")
        for lbl in cont_labels:
            write_result_to_excel(row_i_current, lbl, "")
        return

    min_price_val = min(float(x['price'].replace('USD', '').replace(',', '').strip()) for x in priced_articles)
    cheapest_articles = [x for x in priced_articles if float(x['price'].replace('USD', '').replace(',', '').strip()) == min_price_val]

    list_chuyen = []
    from datetime import datetime
    for x in cheapest_articles:
        try:
            etd_str = x['etd'].split(',')[0].strip()
            etd_dt = datetime.strptime(etd_str, "%d %b %Y")
            tt_str = x['tt'].split('days')[0].strip()
            tt_days = int(tt_str) if tt_str.isdigit() else 999

            list_chuyen.append({
                "original_data": x,
                "etd_dt": etd_dt,
                "tt_days": tt_days,
                "price": min_price_val
            })
        except Exception as e:
            continue

    if list_chuyen:
        etd_chuan, str_etd, str_tt = apply_9_golden_rules(list_chuyen)
        best = etd_chuan[0]["original_data"]
        print(f"   ✅ Áp 9 Quy Tắc -> ETD: {str_etd} | T/T: {str_tt} days")
    else:
        best = cheapest_articles[0]
        str_etd, str_tt = best['etd'], best['tt']
        print(f"   ✅ Chọn article rẻ nhất {best['idx']+1}: {best['price']} | ETD: {str_etd} | T/T: {str_tt}")

    ok_pb = click_price_breakdown(best['idx'])
    if not ok_pb:
        print(f"   ⚠️ Không click được price breakdown")
        return

    charges = read_freight_charges()
    china_route = is_china_destination(country=country, pod=pod)

    for cont_label in cont_labels:
        print(f"   >>> Đang tính giá cho: {cont_label}")
        of_val, has_ths, of_formula = calculate_of(charges, cont_label, include_origin_thc=china_route)

        remark = build_subject_remark(othc_included=(china_route or not has_ths), country=country)

        print(f"   📝 Đã tạo Remark cho {cont_label}: {remark}")
        write_result_to_excel(row_i_current, cont_label, of_formula or (of_val if of_val > 0 else ""), etd=str_etd, tt=str_tt, remark=remark)


def handle_priced_results(country="", pod=""):
    priced = get_priced_articles()
    print(f"   🔍 Articles có giá: {len(priced)}")
    process_priced_articles(priced, ["20 Dry Standard", "40 Dry High"], country, pod)
    print(f"   ⏭️ Bỏ 40 Dry Standard → Ghi dấu trừ")
    write_result_to_excel(row_i_current, "40 Dry Standard", "-")

def check_sailings_and_handle(country="", pol_search="", pod_search=""):
    print(f"   ⏳ Đang chờ TẤT CẢ các thẻ bung đủ giá/status (tối đa 20s)...")
    price_divs = []
    for i in range(20):
        handle_captcha_if_present()
        api_error = latest_productoffer_api_error()
        if api_error:
            print(f"   🚫 {api_error}")
            dump_api_calls("PRODUCTOFFER ERROR")
            return False
        price_divs = read_price_divs()
        if price_divs and price_divs != ['NO ARTICLES'] and price_divs != ['NO STATUS LABELS']:
            if 'UNKNOWN STATUS' not in price_divs:
                print(f"   ⚡ XONG! Đã load đủ ruột cho {len(price_divs)} thẻ ở giây thứ {i+1}!")
                break
        time.sleep(1)

    print(f"   📋 Sailings status: {len(price_divs)} kết quả")
    for i, txt in enumerate(price_divs):
        print(f"      [{i+1}]: {txt}")

    if not price_divs or price_divs == ['NO ARTICLES']:
        print(f"   ⚠️ Web không load được thẻ bài nào (Lag) -> Gửi tín hiệu Reset!")
        return False

    if price_divs == ['NO STATUS LABELS']:
        print(f"   ⚠️ Lỗi load status (Lag) -> Gửi tín hiệu Reset!")
        return False

    if all('HAS_PRICE' in t for t in price_divs):
        print(f"   ✅ Tất cả articles có giá!")
        handle_priced_results(country=country, pod=pod_search)
        return True

    has_container_not_available = any('container not available' in t.lower() for t in price_divs)

    if has_container_not_available:
        print(f"   ⚠️ Có Container not available → back + check từng cont...")
        go_back()
        check_individual_containers(pol_search=pol_search, pod_search=pod_search, country=country)
    else:
        handle_priced_results(country=country, pod=pod_search)

    return True

# ===================================================================================
# XÓA CONTAINER SLOT
# ===================================================================================
def delete_container_slot(slot_idx):
    result = driver.execute_script(_PIERCE_JS + """
        var cs = pierce(document, 'mc-c-container-select');
        if (!cs) return 'no container-select';
        var sis = pierceAll(cs.shadowRoot || cs, 'mc-c-container-selection-input');
        var si  = sis[arguments[0] - 1];
        if (!si) return 'no slot ' + arguments[0];
        var mcbtns = pierceAll(si.shadowRoot || si, 'mc-button');
        for (var i = mcbtns.length - 1; i >= 0; i--) {
            var sr  = mcbtns[i].shadowRoot || mcbtns[i];
            var btn = sr.querySelector('button:not([disabled])');
            if (btn) { btn.click(); return 'ok slot ' + arguments[0]; }
        }
        return 'no button';
    """, slot_idx)
    print(f"   🗑️ Delete slot {slot_idx}: {result}")
    human_sleep(0.3, 0.5)
    return 'ok' in str(result)

# ===================================================================================
# ĐỔI LOẠI CONTAINER Ở SLOT 1
# ===================================================================================
def change_slot1_to(opt_idx, label):
    cs = driver.execute_script(_PIERCE_JS + "return pierce(document, 'mc-c-container-select');")
    if not cs:
        print("   ⚠️ Không tìm thấy mc-c-container-select")
        return False

    si = driver.execute_script(_PIERCE_JS + """
        var sis = pierceAll(arguments[0].shadowRoot || arguments[0], 'mc-c-container-selection-input');
        return sis[0] || null;
    """, cs)
    if not si:
        print("   ⚠️ Không tìm thấy slot 1")
        return False

    cleared = driver.execute_script(_PIERCE_JS + """
        var si = arguments[0];
        var ta = pierce(si.shadowRoot || si, 'mc-typeahead');
        if (!ta) return 'no typeahead';
        var sr = ta.shadowRoot || ta;
        var clearMcBtn = pierce(sr, 'mc-icon-button');
        if (clearMcBtn) {
            var btn = clearMcBtn.shadowRoot
                ? clearMcBtn.shadowRoot.querySelector('button')
                : clearMcBtn.querySelector('button');
            if (btn) { btn.click(); return 'cleared via mc-icon-button'; }
            clearMcBtn.click(); return 'cleared via mc-icon-button (fallback)';
        }
        var inp = pierce(sr, 'input[type="text"]');
        if (inp) {
            inp.click(); inp.focus();
            var setter = Object.getOwnPropertyDescriptor(
                window.HTMLInputElement.prototype, 'value').set;
            setter.call(inp, '');
            inp.dispatchEvent(new Event('input',  {bubbles: true}));
            inp.dispatchEvent(new Event('change', {bubbles: true}));
            inp.click();
            return 'cleared via input reset';
        }
        return 'cannot clear';
    """, si)
    print(f"   🔄 Clear slot 1: {cleared}")
    human_sleep(0.3, 0.5)

    options_visible = driver.execute_script(_PIERCE_JS + """
        return pierceAll(arguments[0].shadowRoot || arguments[0], 'mc-option').length > 0;
    """, si)
    if not options_visible:
        driver.execute_script(_PIERCE_JS + """
            var si  = arguments[0];
            var ta  = pierce(si.shadowRoot || si, 'mc-typeahead');
            if (!ta) return;
            var inp = pierce(ta.shadowRoot || ta, 'input[type="text"]');
            if (inp) { inp.click(); inp.focus(); }
        """, si)
        human_sleep(0.3, 0.5)

    try:
        WebDriverWait(driver, 4).until(lambda d: driver.execute_script(
            _PIERCE_JS + """
            return pierceAll(arguments[0].shadowRoot || arguments[0], 'mc-option').length > 0;
            """, si))
    except:
        print(f"   ⚠️ Timeout chờ dropdown options")

    human_sleep(0.1, 0.3)

    clicked = driver.execute_script(_PIERCE_JS + """
        var si   = arguments[0];
        var opts = pierceAll(si.shadowRoot || si, 'mc-option');
        var opt  = opts[arguments[1] - 1];
        if (!opt) return false;
        var btn = opt.shadowRoot
            ? opt.shadowRoot.querySelector('button')
            : opt.querySelector('button');
        if (btn) { btn.click(); return true; }
        opt.click(); return true;
    """, si, opt_idx)

    if not clicked:
        print(f"   ⚠️ Không chọn được option {opt_idx} ({label})")
        return False

    human_sleep(0.2, 0.4)

    weight_inp = driver.execute_script(_PIERCE_JS + """
        var sr  = arguments[0].shadowRoot || arguments[0];
        var mci = pierce(sr, 'mc-input');
        if (!mci) return null;
        return pierce(mci.shadowRoot || mci, 'input');
    """, si)
    if weight_inp:
        human_move_to(weight_inp)
        driver.execute_script("arguments[0].click(); arguments[0].focus();", weight_inp)
        weight_inp.send_keys(Keys.CONTROL + "a")
        weight_inp.send_keys(Keys.DELETE)
        weight_inp.send_keys(Keys.BACK_SPACE * 10)
        human_sleep(0.05, 0.1)
        for ch in "22222":
            weight_inp.send_keys(ch)
            time.sleep(random.uniform(0.04, 0.08))
        human_sleep(0.1, 0.2)
        print(f"   ✅ Slot 1 → {label} (weight=22222)")
    else:
        print(f"   ⚠️ Slot 1 → {label} (không nhập được weight)")

    return True

# ===================================================================================
# CHECK TỪNG CONT
# ===================================================================================
INDIVIDUAL_CONT = [
    {"label": "20 Dry Standard", "opt_idx": 1},
    {"label": "40 Dry High",     "opt_idx": 3},
]

def check_individual_containers(pol_search="", pod_search="", country=""):
    for idx, cont in enumerate(INDIVIDUAL_CONT):
        label   = cont["label"]
        opt_idx = cont["opt_idx"]

        print(f"\n   📦 Individual check: {label}")

        if idx == 0:
            delete_container_slot(2)
        else:
            ok = change_slot1_to(opt_idx, label)
            if not ok:
                print(f"   ❌ Không đổi được sang {label}")
                continue

        human_sleep(0.2, 0.4)
        human_scroll()

        ok_s = smart_click_search()
        if not ok_s:
            print(f"   ⚠️ Không bấm được Search ({label})")
            continue

        human_sleep(1.0, 1.5)
        if check_booking_error_notification():
            print(f"   🚨 Lỗi notification đỏ cho {label} → bỏ qua!")
            write_result_to_excel(row_i_current, label, "N/A")
            remaining = [c for c in INDIVIDUAL_CONT[idx+1:]]
            if remaining:
                go_back()
            continue

        captcha_result = handle_captcha_if_present(
            pol_search=pol_search, pod_search=pod_search)
        if captcha_result == 'timeout':
            continue
        elif captcha_result == 'reloaded':
            print(f"   ❌ Trang reload sau captcha → bỏ qua {label}")
            continue

        loaded = wait_for_sailings_page()
        if not loaded:
            continue

        print(f"   ⏳ Đang chờ TẤT CẢ các thẻ bung đủ giá/status (tối đa 20s)...")
        price_divs = []
        for i in range(20):
            handle_captcha_if_present()
            price_divs = read_price_divs()
            if price_divs and price_divs != ['NO ARTICLES'] and price_divs != ['NO STATUS LABELS']:
                if 'UNKNOWN STATUS' not in price_divs:
                    print(f"   ⚡ XONG! Đã load đủ ruột cho {len(price_divs)} thẻ ở giây thứ {i+1}!")
                    break
            time.sleep(1)

        print(f"   📋 [{label}] status: {len(price_divs)} kết quả")
        for i, txt in enumerate(price_divs):
            print(f"      [{i+1}]: {txt}")

        priced = get_priced_articles()
        print(f"   🔍 Articles có giá: {len(priced)}")
        process_priced_articles(priced, label)

        if label == "20 Dry Standard":
            print(f"   ⏭️ Bỏ 40 Dry Standard → ghi '-'")
            write_result_to_excel(row_i_current, "40 Dry Standard", "-")

        remaining = [c for c in INDIVIDUAL_CONT[idx+1:]]
        if remaining:
            print(f"   🔙 Back về trang nhập liệu...")
            go_back()

# ===================================================================================
# ĐỌC EXCEL (CÓ MODULE NHÂN BẢN TUYẾN HẢI PHÒNG)
# ===================================================================================
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

target_single_row = None
if SINGLE_ROW:
    try:
        target_single_row = int(SINGLE_ROW)
        print(f"[SINGLE_ROW] Chỉ chạy dòng {target_single_row} theo lệnh từ main.py")
    except ValueError:
        print(f"[SINGLE_ROW] Không hợp lệ: {SINGLE_ROW}. Chạy tất cả dòng.")

queue = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    country   = str(row[1] or "").strip().upper()
    pol_excel = str(row[2] or "").strip().upper()
    pod_excel = str(row[3] or "").strip().upper()
    carrier   = str(row[4] or "").strip().upper()

    if target_single_row and i != target_single_row: continue
    if not pol_excel or not pod_excel: continue
    if carrier not in MAERSK_GROUP: continue
    if FILTER_POL and pol_excel != FILTER_POL: continue
    if FILTER_POD and pod_excel != FILTER_POD: continue
    if FILTER_COUNTRY and country != FILTER_COUNTRY: continue

    if pol_excel in ("HAI PHONG", "HAIPHONG"):
        if country in HAIPHONG_NEARBY_COUNTRIES:
            queue.append((i, country, "HAIPHONG_LACH_HUYEN", pod_excel))
            queue.append((i, country, "HAIPHONG_NORMAL", pod_excel))
        else:
            queue.append((i, country, "HAIPHONG_LACH_HUYEN", pod_excel))
    else:
        queue.append((i, country, pol_excel, pod_excel))

wb.close()
print(f"📋 Tổng cộng {len(queue)} lượt chạy (đã tự động nhân bản Hải Phòng tuyến gần)")
if not queue:
    print("✅ Không có dòng MAERSK phù hợp filter. Thoát.")
    try:
        driver.quit()
    except Exception:
        pass
    sys.exit(0)

# ===================================================================================
# MAIN LOOP
# ===================================================================================
driver.switch_to.window(driver.window_handles[-1])

try:
    if BASE_URL not in driver.current_url:
        print(f"🌐 Điều hướng đến {BASE_URL}")
        driver.get(BASE_URL)
        wait_until_book_ready(timeout_seconds=25)

    current = driver.current_url
    if "login" in current or "signin" in current or "maersk.com/book/" not in current:
        print("\n" + "="*50)
        print("👤 CHƯA ĐĂNG NHẬP!")
        print("👉 Hãy đăng nhập Maersk thủ công trên Chrome.")
        print("👉 Sau khi vào được trang maersk.com/book/ thì quay lại đây bấm Enter.")
        print("="*50)
        if not wait_for_manual_condition(
            "\n   🟢 Bấm Enter khi đã đăng nhập xong và đang ở trang /book/...",
            is_maersk_book_ready,
            poll_seconds=5,
        ):
            raise WebLagException("Chưa đăng nhập được Maersk")
except Exception as e:
    print(f"⚠️ Lỗi điều hướng ban đầu: {e}")
    sys.exit(1)

row_idx = 0
row_retry_count = 0

while row_idx < len(queue):
    row_i, country, pol, pod = queue[row_idx]
    row_i_current = row_i

    CURRENT_POL = pol
    CURRENT_POD = pod

    pol_search_str = resolve_maersk_alias(pol, country)[0].lower()
    pod_search_str = resolve_maersk_alias(pod, country)[0].lower()

    print(f"\n{'='*50}")
    print(f"DÒNG {row_i}: {pol} -> {pod}")

    try:
        current_url = driver.current_url

        if "sailings" in current_url and row_retry_count == 0:
            print(f"   🔙 Đã check xong giá. Bấm lùi về trang nhập liệu để giữ nguyên Form...")
            go_back()
            time.sleep(2.5)

        elif "maersk.com/book/" not in current_url or row_retry_count > 0:
            print(f"   🔄 [CẢM BIẾN] Web lệch hướng nặng hoặc đang cứu hộ. Bắt buộc F5...")
            driver.get(BASE_URL)
            wait_until_book_ready(timeout_seconds=25)

        # Inject event logger (debug — xóa sau khi không cần nữa)
        inject_event_logger()
        inject_api_recorder()

        # Idle trước khi bắt đầu điền form
        pre_interaction_idle()

        if not type_port(0, pol, country, "POL", pod_country=country):
            raise WebLagException("Không gõ được POL")
        dump_event_log("SAU POL")

        time.sleep(random.uniform(0.8, 1.8))

        if not type_port(1, pod, country, "POD"):
            raise WebLagException("Không gõ được POD")
        dump_event_log("SAU POD")

        time.sleep(random.uniform(1.2, 2.5))

        if not type_commodity():
            raise WebLagException("Không gõ được Hàng hóa")
        dump_event_log("SAU COMMODITY")

        time.sleep(random.uniform(0.6, 1.4))

        if not select_containers(country=country):
            raise WebLagException("Không chọn được Container")
        dump_event_log("SAU CONTAINER")

        ensure_radio_checked()
        dump_event_log("SAU RADIO")

        select_date_plus7()
        dump_event_log("SAU DATE")

        time.sleep(random.uniform(1.5, 3.0))

        search_status = smart_click_search()
        dump_event_log("SAU SEARCH CLICK")
        api_error = latest_productoffer_api_error()
        if api_error:
            print(f"   🚫 {api_error}")
            dump_api_calls("PRODUCTOFFER ERROR")
            write_result_to_excel(row_i_current, "20 Dry Standard", api_error[:120])
            row_break()
            row_idx += 1
            row_retry_count = 0
            continue

        if search_status == 'NO_SAILINGS':
            print(f"   ❌ Tuyến này không có chuyến. (No sailings)")
            write_result_to_excel(row_i_current, "20 Dry Standard", "NO SERVICE")
            row_break()
            row_idx += 1
            row_retry_count = 0
            continue

        elif search_status == 'ERROR_RETRY':
            raise WebLagException("Web lag làm mất nút Search hoặc có banner lạ")

        if not wait_for_sailings_page():
            raise WebLagException("Quá thời gian tải trang Sailings")

        if not check_sailings_and_handle(country=country, pol_search=pol_search_str, pod_search=pod_search_str):
            raise WebLagException("Trang sailings trống rỗng, không có data")

        # Xong xuôi — nghỉ rồi chuyển dòng tiếp
        row_break()
        row_idx += 1
        row_retry_count = 0

    except ManualHandoverComplete:
        check_sailings_and_handle(country=country, pol_search=pol_search_str, pod_search=pod_search_str)
        row_break()
        row_idx += 1
        row_retry_count = 0

    except NoSailingsException:
        print(f"   ❌ Tuyến này không có chuyến. (Xác nhận sau khi giải Captcha)")
        write_result_to_excel(row_i_current, "20 Dry Standard", "NO SERVICE")
        row_break()
        row_idx += 1
        row_retry_count = 0

    except WebLagException as e:
        print(f"   💥 [BÁO ĐỘNG] {e}. Kích hoạt cơ chế chống kẹt!")
        row_retry_count += 1
        if row_retry_count <= 3:
            print(f"   🔄 Thử lại dòng này (Lần {row_retry_count}/3)...")
            continue
        else:
            print(f"   ❌ Hết 3 lượt cứu hộ! Chịu thua dòng {row_i}.")
            write_result_to_excel(row_i_current, "20 Dry Standard", "LỖI WEB")
            row_break()
            row_idx += 1
            row_retry_count = 0
            continue

    except Exception as e:
        print(f"   💥 Lỗi hệ thống sâu: {e}")
        time.sleep(5)
        manual_pause("   [!] Đã dừng lại do lỗi. Hãy đưa web về trang chủ rồi bấm Enter để Bot thử lại...", poll_seconds=5)

print("\n✅ DONE")
