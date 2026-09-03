"""
HMM Price Checker

The HMM site exposes the quotation result and rate-detail endpoints after login.
This bot uses Selenium only to keep a logged-in Edge session, then calls those
endpoints from the browser context with fetch().
"""

import calendar
import math
import os
import re
import socket
import subprocess
import sys
import time
from datetime import datetime, timedelta

import openpyxl
import requests
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait

from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from remark_rules import apply_manifest_rule, get_manifest_code, is_china_destination, normalize_remark_text


try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


CURRENT_FOLDER = os.getcwd()
DRIVER_PATH = os.path.join(CURRENT_FOLDER, "msedgedriver.exe")
EXCEL_PATH = os.environ.get("EXCEL_PATH", os.path.join(CURRENT_FOLDER, "input_gia.xlsx"))
FILTER_POL = os.environ.get("FILTER_POL", "").strip().upper()
FILTER_POD = os.environ.get("FILTER_POD", "").strip().upper()
SINGLE_ROW = os.environ.get("SINGLE_ROW", "").strip()

DATE_OFFSET_DAYS = parse_date_offset_days()

HMM_PORT = int(os.environ.get("HMM_EDGE_PORT", "9533"))
HMM_PROFILE = os.environ.get("HMM_EDGE_PROFILE", r"C:\edge_hmm")
HMM_URL = "https://www.hmm21.com/e-service/hiquote/quotationNew.do"
HMM_LOGIN_URL = "https://www.hmm21.com/e-service/auth/login.do"
HMM_BASE_URL = "https://www.hmm21.com"
HMM_USER = os.environ.get("HMM_USER", "PIOLOG")
HMM_PASS = os.environ.get("HMM_PASS", "hmm7980762")
HMM_NAV_RETRY = int(os.environ.get("HMM_NAV_RETRY", "4"))
HMM_PAGELOAD_TIMEOUT = int(os.environ.get("HMM_PAGELOAD_TIMEOUT", "45"))
HMM_SCRIPT_TIMEOUT = int(os.environ.get("HMM_SCRIPT_TIMEOUT", "25"))
HMM_AJAX_TIMEOUT_MS = int(os.environ.get("HMM_AJAX_TIMEOUT_MS", "18000"))
HMM_REQUEST_TIMEOUT = max(5, int(os.environ.get("HMM_REQUEST_TIMEOUT", "10")))
HMM_FETCH_TIMEOUT_MS = int(os.environ.get("HMM_FETCH_TIMEOUT_MS", str(max(12000, HMM_REQUEST_TIMEOUT * 1000))))
HMM_API_ENGINE = os.environ.get("HMM_API_ENGINE", "browser").strip().lower()  # browser|requests
HMM_REQUEST_SESSION = None


CONTAINERS = {
    "20": "DC20",
    "40": "DC40",
    "40HC": "DC4H",
}


PORT_CODES = {
    "HO CHI MINH": "VNSGN",
    "HOCHIMINH": "VNSGN",
    "HO CHI MINH CITY": "VNSGN",
    "HCM": "VNSGN",
    "SAIGON": "VNSGN",
    "CAT LAI": "VNSGN",
    "VNSGN": "VNSGN",
    "HAI PHONG": "VNHPH",
    "HAIPHONG": "VNHPH",
    "HPH": "VNHPH",
    "VNHPH": "VNHPH",
    "MUNDRA": "INMUN",
    "INMUN": "INMUN",
    "NHAVA SHEVA": "INNSA",
    "NHAVASHEVA": "INNSA",
    "JNPT": "INNSA",
    "JAWAHARLAL NEHRU": "INNSA",
    "INNSA": "INNSA",
    "HAMBURG": "DEHAM",
    "DEHAM": "DEHAM",
    "BREMERHAVEN": "DEBRV",
    "DEBRV": "DEBRV",
    "FELIXSTOWE": "GBFXT",
    "GBFXT": "GBFXT",
    "SOUTHAMPTON": "GBSOU",
    "GBSOU": "GBSOU",
    "LIVERPOOL": "GBLIV",
    "GBLIV": "GBLIV",
    "ROTTERDAM": "NLRTM",
    "NLRTM": "NLRTM",
    "ANTWERP": "BEANR",
    "BEANR": "BEANR",
    "LE HAVRE": "FRLEH",
    "LEHAVRE": "FRLEH",
    "FRLEH": "FRLEH",
    "FOS": "FRFOS",
    "FOS-SUR-MER": "FRFOS",
    "FOS SUR MER": "FRFOS",
    "FRFOS": "FRFOS",
    "GENOA": "ITGOA",
    "GENOVA": "ITGOA",
    "ITGOA": "ITGOA",
    "LA SPEZIA": "ITSPE",
    "LASPEZIA": "ITSPE",
    "ITSPE": "ITSPE",
    "BARCELONA": "ESBCN",
    "ESBCN": "ESBCN",
    "VALENCIA": "ESVLC",
    "ESVLC": "ESVLC",
    "ALGECIRAS": "ESALG",
    "ESALG": "ESALG",
    "GDANSK": "PLGDN",
    "PLGDN": "PLGDN",
    "KOPER": "SIKOP",
    "SIKOP": "SIKOP",
    "PIRAEUS": "GRPIR",
    "GRPIR": "GRPIR",
    "BUSAN": "KRPUS",
    "PUSAN": "KRPUS",
    "KRPUS": "KRPUS",
    "SHANGHAI": "CNSHA",
    "CNSHA": "CNSHA",
    "NINGBO": "CNNGB",
    "CNNGB": "CNNGB",
    "QINGDAO": "CNTAO",
    "CNTAO": "CNTAO",
    "XIAMEN": "CNXMN",
    "CNXMN": "CNXMN",
    "NANSHA": "CNNSA",
    "CNNSA": "CNNSA",
    "SHEKOU": "CNSHK",
    "CNSHK": "CNSHK",
    "YANTIAN": "CNYTN",
    "CNYTN": "CNYTN",
    "HONG KONG": "HKHKG",
    "HONGKONG": "HKHKG",
    "HKHKG": "HKHKG",
    "TOKYO": "JPTYO",
    "JPTYO": "JPTYO",
    "YOKOHAMA": "JPYOK",
    "JPYOK": "JPYOK",
    "KOBE": "JPKOB",
    "JPKOB": "JPKOB",
    "OSAKA": "JPOSA",
    "JPOSA": "JPOSA",
    "SINGAPORE": "SGSIN",
    "SGSIN": "SGSIN",
    "PORT KLANG": "MYPKG",
    "PORTKLANG": "MYPKG",
    "MYPKG": "MYPKG",
    "LAEM CHABANG": "THLCH",
    "LAEMCHABANG": "THLCH",
    "THLCH": "THLCH",
    "BANGKOK": "THBKK",
    "THBKK": "THBKK",
    "JAKARTA": "IDJKT",
    "IDJKT": "IDJKT",
    "SURABAYA": "IDSUB",
    "IDSUB": "IDSUB",
    "MANILA": "PHMNL",
    "PHMNL": "PHMNL",
    "CEBU": "PHCEB",
    "PHCEB": "PHCEB",
    "JEDDAH": "SAJED",
    "SAJED": "SAJED",
    "JEBEL ALI": "AEJEA",
    "JEBELALI": "AEJEA",
    "AEJEA": "AEJEA",
    "DAMMAM": "SADMM",
    "SADMM": "SADMM",
    "HAMAD": "QAHMD",
    "QAHMD": "QAHMD",
    "KARACHI": "PKKHI",
    "PKKHI": "PKKHI",
    "SANTOS": "BRSSZ",
    "BRSSZ": "BRSSZ",
    "NAVEGANTES": "BRNVT",
    "BRNVT": "BRNVT",
    "SALVADOR": "BRSSA",
    "BRSSA": "BRSSA",
    "MANZANILLO": "MXZLO",
    "MXZLO": "MXZLO",
    "SYDNEY": "AUSYD",
    "AUSYD": "AUSYD",
    "MELBOURNE": "AUMEL",
    "AUMEL": "AUMEL",
}


EXCHANGE_RATE_CACHE = {}


def log(msg):
    print(f"[HMM] {msg}")


def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(("127.0.0.1", port)) == 0


def wait_port(port, timeout=10):
    end = time.time() + timeout
    while time.time() < end:
        if is_port_in_use(port):
            return True
        time.sleep(0.25)
    return False


def start_edge_if_needed():
    if is_port_in_use(HMM_PORT):
        log(f"Edge debug port {HMM_PORT} da mo san.")
        return
    log(f"Edge HMM chua mo. Dang khoi dong port {HMM_PORT}...")
    edge = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    if not os.path.exists(edge):
        edge = r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"
    subprocess.Popen([
        edge,
        f"--remote-debugging-port={HMM_PORT}",
        f"--user-data-dir={HMM_PROFILE}",
        "--start-maximized",
        "--window-size=1920,1080",
        "--disable-background-timer-throttling",
        "--disable-renderer-backgrounding",
        "--disable-backgrounding-occluded-windows",
        HMM_URL,
    ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    if not wait_port(HMM_PORT, timeout=15):
        raise RuntimeError(f"Khong khoi dong duoc Edge HMM port {HMM_PORT}")


def connect_driver():
    opts = Options()
    opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{HMM_PORT}")
    service = Service(executable_path=DRIVER_PATH)
    drv = webdriver.Edge(service=service, options=opts)
    try:
        drv.set_page_load_timeout(HMM_PAGELOAD_TIMEOUT)
        drv.set_script_timeout(HMM_SCRIPT_TIMEOUT)
    except Exception:
        pass
    try:
        drv.maximize_window()
    except Exception:
        try:
            drv.set_window_size(1920, 1080)
        except Exception:
            pass
    return drv


def body_text(driver):
    try:
        return driver.find_element(By.TAG_NAME, "body").text
    except Exception:
        return ""


def hmm_quote_ready(driver):
    text = body_text(driver)
    if "Get a Quote" in text or "MS CELINE" in text:
        return True
    try:
        return bool(driver.find_elements(By.ID, "srchPointFrom"))
    except Exception:
        return False


def hmm_hi_quote_inactive(driver):
    text = body_text(driver).lower()
    return "hi quote service is not currently activated" in text


def hmm_safe_get(driver, url, attempts=None, wait_after=1.5):
    attempts = attempts or HMM_NAV_RETRY
    transient_tokens = (
        "ERR_INTERNET_DISCONNECTED",
        "ERR_NETWORK_CHANGED",
        "ERR_CONNECTION",
        "ERR_TIMED_OUT",
        "ERR_NAME_NOT_RESOLVED",
        "timeout",
    )
    last_error = ""
    for attempt in range(1, attempts + 1):
        try:
            driver.get(url)
            time.sleep(wait_after)
            return True
        except TimeoutException as e:
            last_error = str(e).splitlines()[0] if str(e) else "page load timeout"
            try:
                driver.execute_script("window.stop();")
            except Exception:
                pass
            if hmm_quote_ready(driver):
                log(f"Mo HMM bi timeout nhung form da len -> di tiep ({url})")
                return True
        except WebDriverException as e:
            msg = str(e)
            last_error = msg.splitlines()[0] if msg else type(e).__name__
            if not any(token.lower() in msg.lower() for token in transient_tokens):
                raise

        if attempt < attempts:
            log(f"Mo HMM loi tam thoi: {last_error} -> thu lai {attempt + 1}/{attempts}")
            time.sleep(min(3 * attempt, 10))
            continue
        raise RuntimeError(f"HMM khong mo duoc URL sau {attempts} lan: {url} | {last_error}")


def ensure_login(driver):
    hmm_safe_get(driver, HMM_URL)
    if hmm_quote_ready(driver):
        log("Da co session HMM.")
        return

    log("Chua login HMM -> dang login...")
    for attempt in range(1, 3):
        hmm_safe_get(driver, HMM_LOGIN_URL)
        WebDriverWait(driver, 25).until(lambda d: hmm_quote_ready(d) or d.find_elements(By.ID, "userId"))
        if hmm_quote_ready(driver):
            log("Login HMM xong.")
            return

        driver.execute_script(
            """
            const user = document.getElementById('userId');
            const pwd = document.getElementById('userPwd');
            if (!user || !pwd) {
                throw new Error('HMM login form not found');
            }
            user.value = arguments[0];
            pwd.value = arguments[1];
            user.dispatchEvent(new Event('input', {bubbles:true}));
            pwd.dispatchEvent(new Event('input', {bubbles:true}));
            if (typeof esvcLogin === 'function') {
                esvcLogin();
            } else {
                const btn = document.querySelector('button[type="submit"], input[type="submit"], a[onclick*="esvcLogin"]');
                if (btn) btn.click();
                else document.querySelector('form').submit();
            }
            """,
            HMM_USER,
            HMM_PASS,
        )

        try:
            WebDriverWait(driver, 35).until(lambda d: hmm_quote_ready(d) or hmm_hi_quote_inactive(d))
        except TimeoutException:
            if attempt < 2:
                log("Login HMM chua thay form quote -> thu login lai...")
                continue
            raise

        if hmm_hi_quote_inactive(driver):
            raise RuntimeError(
                "HMM login duoc nhung account chua active Hi Quote. "
                "Can vao My Info va bat Activate Hi Quote = Yes."
            )

        hmm_safe_get(driver, HMM_URL)
        WebDriverWait(driver, 25).until(lambda d: hmm_quote_ready(d) or hmm_hi_quote_inactive(d))
        if hmm_hi_quote_inactive(driver):
            raise RuntimeError(
                "HMM login duoc nhung account chua active Hi Quote. "
                "Can vao My Info va bat Activate Hi Quote = Yes."
            )
        log("Login HMM xong.")
        return


def parse_num(value):
    if value is None:
        return 0.0
    s = str(value).strip()
    if not s:
        return 0.0
    s = re.sub(r"[^\d.,\-]", "", s)
    if not s:
        return 0.0
    if "," in s and "." in s:
        s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def fmt_num(value):
    try:
        n = float(value)
    except Exception:
        return ""
    if abs(n - round(n)) < 1e-9:
        return str(int(round(n)))
    return f"{n:.2f}".rstrip("0").rstrip(".")


def excel_formula(parts):
    tokens = []
    for raw in parts:
        try:
            amount = float(raw)
        except Exception:
            continue
        if abs(amount) < 1e-9:
            continue
        sign = "-" if amount < 0 else "+"
        text = fmt_num(abs(amount))
        tokens.append((sign if tokens else ("-" if sign == "-" else "")) + text)
    return "=" + "".join(tokens) if tokens else None


def get_live_exchange_rate(base_currency, target_currency="USD"):
    base = (base_currency or "").upper().strip()
    target = (target_currency or "USD").upper().strip()
    if not base or base == target:
        return 1.0
    key = (base, target)
    if key in EXCHANGE_RATE_CACHE:
        return EXCHANGE_RATE_CACHE[key]
    env_key = f"HMM_{base}_TO_{target}"
    try:
        rate = float(os.environ.get(env_key, "").strip())
        EXCHANGE_RATE_CACHE[key] = rate
        return rate
    except Exception:
        pass
    try:
        res = requests.get(f"https://api.frankfurter.app/latest?from={base}&to={target}", timeout=3)
        rate = float(res.json()["rates"][target])
        EXCHANGE_RATE_CACHE[key] = rate
        return rate
    except Exception as e:
        fallbacks = {
            ("EUR", "USD"): 1.08,
            ("CHF", "USD"): 1.12,
            ("AUD", "USD"): 0.66,
            ("JPY", "USD"): 0.0068,
            ("CNY", "USD"): 0.14,
            ("VND", "USD"): 0.00004,
        }
        rate = fallbacks.get(key, 1.0)
        EXCHANGE_RATE_CACHE[key] = rate
        log(f"WARN FX {base}->{target} loi ({e}); fallback={rate}")
        return rate


def to_usd(amount, currency):
    return float(amount) * get_live_exchange_rate(currency, "USD")


def clean_port_name(value):
    s = re.sub(r"\([^)]*\)", "", str(value or "")).strip()
    if "," in s:
        s = s.split(",", 1)[0].strip()
    return re.sub(r"\s+", " ", s).upper() or "DIRECT"


def clean_vessel(value):
    s = re.sub(r"\s+", " ", str(value or "")).strip()
    s = re.sub(r"\s*\([A-Z0-9]{2,5}\)\s*$", "", s).strip()
    return s or "TBA"


def date_from_any(value):
    s = str(value or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y", "%d-%b"):
        try:
            dt = datetime.strptime(s, fmt)
            if fmt == "%d-%b":
                dt = dt.replace(year=datetime.now().year)
                if dt.date() < datetime.now().date() - timedelta(days=60):
                    dt = dt.replace(year=dt.year + 1)
            return dt.date()
        except Exception:
            pass
    return None


def get_valid_date(etd_dates):
    if not etd_dates:
        return ""
    latest = max(etd_dates)
    day = latest.day
    if day <= 7:
        valid_day = 7
    elif day <= 14:
        valid_day = 14
    elif day <= 21:
        valid_day = 21
    else:
        valid_day = calendar.monthrange(latest.year, latest.month)[1]
    return f"{valid_day}-{latest.strftime('%b')}"


def hmm_csrf_token(driver):
    try:
        return driver.execute_script(
            """
            return (
                document.querySelector('input[name="X-CSRF-TOKEN"]')?.value
                || document.querySelector('input[name="_csrf"]')?.value
                || document.querySelector('meta[name="_csrf"]')?.content
                || document.querySelector('meta[name="csrf-token"]')?.content
                || ''
            );
            """
        ) or ""
    except Exception:
        return ""


def hmm_requests_session(driver):
    global HMM_REQUEST_SESSION
    if HMM_REQUEST_SESSION is None:
        HMM_REQUEST_SESSION = requests.Session()
    sess = HMM_REQUEST_SESSION
    for c in driver.get_cookies():
        try:
            sess.cookies.set(c["name"], c["value"], domain=c.get("domain"), path=c.get("path", "/"))
        except Exception:
            sess.cookies.set(c.get("name"), c.get("value"))
    return sess


def hmm_post(driver, url, payload, expect_json=True):
    full_url = url if str(url).lower().startswith("http") else HMM_BASE_URL + url
    use_form = not expect_json
    if HMM_API_ENGINE != "requests":
        try:
            return hmm_browser_post(driver, full_url, payload, expect_json=expect_json, use_form=use_form)
        except Exception as exc:
            msg = str(exc).splitlines()[0] if str(exc) else type(exc).__name__
            log(f"      HMM browser-fetch lỗi -> fallback requests: {msg[:120]}")

    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json, text/javascript, */*; q=0.01" if expect_json else "text/html, */*; q=0.01",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8" if use_form else "application/json; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": HMM_URL,
    }
    csrf = hmm_csrf_token(driver)
    if csrf:
        headers["X-CSRF-TOKEN"] = csrf
    resp = hmm_requests_session(driver).post(
        full_url,
        data=_form_encode(payload or {}) if use_form else json_dumps(payload or {}),
        headers=headers,
        timeout=HMM_REQUEST_TIMEOUT,
    )
    text = resp.text or ""
    if not resp.ok:
        return {"ok": False, "status": resp.status_code, "error": text[:300], "text": text}
    if expect_json:
        try:
            data = resp.json()
        except Exception as exc:
            return {"ok": False, "status": resp.status_code, "error": f"JSON parse error: {exc}", "text": text[:300]}
        return {"ok": True, "status": resp.status_code, "data": data, "text": text}
    return {"ok": True, "status": resp.status_code, "text": text}


def _form_encode(payload):
    from urllib.parse import urlencode
    pairs = []
    for key, value in (payload or {}).items():
        if isinstance(value, (list, tuple)):
            for item in value:
                pairs.append((key, "" if item is None else str(item)))
        else:
            pairs.append((key, "" if value is None else str(value)))
    return urlencode(pairs)


def hmm_browser_post(driver, full_url, payload, expect_json=True, use_form=False):
    """
    Gọi API ngay trong tab HMM bằng CDP Runtime.evaluate.
    Cách này dùng đúng browser session/cookie và không phụ thuộc execute_async_script
    (execute_async_script từng bị treo ở HMM).
    """
    body_expr = json_dumps(payload or {})
    expression = f"""
    (async () => {{
      const url = {json_dumps(full_url)};
      const payload = {body_expr};
      const useForm = {str(bool(use_form)).lower()};
      const ctrl = new AbortController();
      const timer = setTimeout(() => ctrl.abort('timeout'), {int(HMM_FETCH_TIMEOUT_MS)});
      const headers = {{
        'Accept': {json_dumps("application/json, text/javascript, */*; q=0.01" if expect_json else "text/html, */*; q=0.01")},
        'X-Requested-With': 'XMLHttpRequest'
      }};
      const csrfHeader = window.globalHeader || document.querySelector('input[name="csrfHeader"]')?.value || 'X-CSRF-TOKEN';
      const csrfToken = window.globalToken
        || document.querySelector('input[name="_csrf"]')?.value
        || document.querySelector('meta[name="_csrf"]')?.content
        || document.querySelector('meta[name="csrf-token"]')?.content
        || '';
      if (csrfHeader && csrfToken) headers[csrfHeader] = csrfToken;
      let body;
      if (useForm) {{
        headers['Content-Type'] = 'application/x-www-form-urlencoded; charset=UTF-8';
        const params = new URLSearchParams();
        Object.entries(payload || {{}}).forEach(([k, v]) => {{
          if (Array.isArray(v)) v.forEach(x => params.append(k, x == null ? '' : String(x)));
          else params.append(k, v == null ? '' : String(v));
        }});
        body = params.toString();
      }} else {{
        headers['Content-Type'] = 'application/json; charset=UTF-8';
        body = JSON.stringify(payload || {{}});
      }}
      try {{
        const resp = await fetch(url, {{
          method: 'POST',
          credentials: 'include',
          headers,
          body,
          signal: ctrl.signal
        }});
        const text = await resp.text();
        return {{ ok: resp.ok, status: resp.status, text }};
      }} catch (e) {{
        return {{ ok: false, status: 0, error: String(e && (e.message || e)) }};
      }} finally {{
        clearTimeout(timer);
      }}
    }})()
    """
    res = driver.execute_cdp_cmd("Runtime.evaluate", {
        "expression": expression,
        "awaitPromise": True,
        "returnByValue": True,
        "timeout": HMM_FETCH_TIMEOUT_MS + 5000,
    })
    if res.get("exceptionDetails"):
        raise RuntimeError(str(res.get("exceptionDetails"))[:300])
    value = (res.get("result") or {}).get("value") or {}
    text = value.get("text") or ""
    if not value.get("ok"):
        return {"ok": False, "status": value.get("status", 0), "error": value.get("error") or text[:300], "text": text}
    if expect_json:
        try:
            import json
            data = json.loads(text or "{}")
        except Exception as exc:
            return {"ok": False, "status": value.get("status", 0), "error": f"JSON parse error: {exc}", "text": text[:300]}
        return {"ok": True, "status": value.get("status", 200), "data": data, "text": text}
    return {"ok": True, "status": value.get("status", 200), "text": text}


def json_dumps(obj):
    import json
    return json.dumps(obj, ensure_ascii=False)


def resolve_port_code(name):
    raw = str(name or "").strip()
    up = re.sub(r"\s+", " ", raw.upper())
    if not up:
        return ""
    if up in PORT_CODES:
        return PORT_CODES[up]
    compact = up.replace(" ", "")
    if compact in PORT_CODES:
        return PORT_CODES[compact]
    if re.fullmatch(r"[A-Z]{5}", up):
        return up
    return ""


def hmm_json_post(driver, url, payload):
    return hmm_post(driver, url, payload, expect_json=True)


def hmm_form_post(driver, url, payload, as_json=False):
    return hmm_post(driver, url, payload, expect_json=as_json)


def parse_cards_html(driver, html):
    return driver.execute_script(
        r"""
        const html = arguments[0] || '';
        const doc = new DOMParser().parseFromString(html, 'text/html');
        const clean = s => (s || '').replace(/\s+/g, ' ').trim();
        const cleanPort = s => {
            s = clean(s).replace(/\([^)]*\)/g, '').trim();
            if (s.includes(',')) s = s.split(',')[0].trim();
            return s.toUpperCase();
        };
        const parseObj = (ul) => {
            const btn = ul.querySelector("button[onclick*='switchAdded']");
            if (!btn) return {};
            const onclick = btn.getAttribute('onclick') || '';
            const m = onclick.match(/switchAdded\s*\(\s*this\s*,\s*(\{[\s\S]*\})\s*\)/);
            if (!m) return {};
            try { return JSON.parse(m[1]); } catch (e) { return {}; }
        };
        const quotNo = doc.querySelector('.quotNo')?.value || '';
        const cards = Array.from(doc.querySelectorAll('.result-area > ul')).map((ul, idx) => {
            const obj = parseObj(ul);
            const text = clean(ul.innerText);
            const seq = String(obj.quotNoSeq || ul.querySelector('.quotNoSeq')?.value || '');
            const spc = String(obj.spcChkRsltCd || ul.querySelector('.spcChkRsltCd')?.value || ul.className || '').toUpperCase();
            const transit = clean(obj.tsTp || ul.querySelector('.transit')?.textContent || '');
            const period = clean(obj.days || ul.querySelector('.period')?.textContent || '');
            const tt = parseInt((period.match(/(\d+)/) || [])[1] || '0', 10);
            let etd = clean(obj.polDpartDt || '');
            if (!etd) etd = clean((text.match(/Departure Date\s*:\s*(\d{4}-\d{2}-\d{2})/) || [])[1] || '');
            let vessel = clean(obj.mthVvdNm || obj.vslDtlNm || ul.querySelector('.vessel')?.textContent || '');
            vessel = vessel.replace(/\s*\([A-Z0-9]{2,5}\)\s*$/, '').trim();
            let ts = clean(obj.tsNm || '');
            if (!ts) {
                const routes = Array.from(ul.querySelectorAll('.route'));
                const tsRoute = routes.find(r => /Transship/i.test(r.textContent || ''));
                ts = clean(tsRoute?.querySelector('.text')?.textContent || '');
            }
            if (!/TRANSSHIP/i.test(transit) || !ts) ts = 'DIRECT';
            else ts = cleanPort(ts);
            let freeTime = '';
            const dem = obj.spclDemFtDys, det = obj.spclDetFtDys, thru = obj.spclThruFtDys;
            if (thru) freeTime = String(thru) + ' COMBINED';
            else if (dem && det) freeTime = String(dem) + ' DEM + ' + String(det) + ' DET';
            return {
                idx, quotNo: obj.quotNo || quotNo, quotNoSeq: seq, spc, transit, tt,
                etd, vessel: vessel || 'TBA', ts: ts || 'DIRECT',
                origin: obj.polCd || obj.polNm || '', destination: obj.podCd || obj.podNm || '',
                freeTime, rawText: text.slice(0, 600)
            };
        });
        return cards;
        """,
        html,
    )


def parse_rate_detail_html(driver, html):
    return driver.execute_script(
        r"""
        const html = arguments[0] || '';
        const doc = new DOMParser().parseFromString(html, 'text/html');
        const clean = s => (s || '').replace(/\s+/g, ' ').trim();
        const rows = [];
        const sectionNames = ['Freight', 'Origin Charges', 'Destination Charges'];
        Array.from(doc.querySelectorAll('tr')).forEach(tr => {
            const direct = Array.from(tr.children).map(td => clean(td.textContent));
            const section = direct[0] || '';
            if (!sectionNames.includes(section)) return;
            const detail = tr.nextElementSibling;
            if (!detail || !detail.classList.contains('result-detail')) return;
            detail.querySelectorAll('table tbody tr').forEach(r => {
                const cells = Array.from(r.children).map(td => clean(td.textContent));
                if (cells.length >= 6) {
                    rows.push({
                        section,
                        item: cells[0],
                        basis: cells[1],
                        qty: cells[2],
                        currency: cells[3],
                        rate: cells[4],
                        amount: cells[5]
                    });
                }
            });
        });
        return rows;
        """,
        html,
    )


def extract_grm_no(data):
    def walk(obj):
        if isinstance(obj, dict):
            for key in ("GrmNo", "grmNo", "GRM_NO"):
                if obj.get(key):
                    return obj.get(key)
            for value in obj.values():
                found = walk(value)
                if found:
                    return found
        elif isinstance(obj, list):
            for value in obj:
                found = walk(value)
                if found:
                    return found
        return ""

    found = walk(data)
    if found:
        return found
    text = str(data)
    m = re.search(r"GrmNo['\"]?\s*[:=]\s*['\"]([^'\"]+)", text)
    return m.group(1) if m else ""


def fetch_container_cards(driver, pol_code, pod_code, sail_date, cntr_code, priority="A"):
    priority = (priority or "A").upper()
    if priority not in {"A", "D", "T"}:
        priority = "A"
    sail_ymd = sail_date.strftime("%Y%m%d")
    sail_iso = sail_date.strftime("%Y-%m-%d")
    first_payload = {
        "srchPointFromCd": pol_code,
        "srchCityFrom": "CY",
        "srchPointToCd": pod_code,
        "srchCityTo": "CY",
        "srchSailDate": sail_ymd,
        "srchSelWeeks": "8",
        "srchPorFcltyCd": "",
        "srchPvyFcltyCd": "",
        "srchSelPriority": priority,
        "srchSelSortBy": "D",
    }
    first = hmm_json_post(driver, "/e-service/hiquote/quotation/apiPointToPointList.do", first_payload)
    first = first or {}
    if not first.get("ok"):
        raise RuntimeError(first.get("error") or f"apiPointToPointList HTTP {first.get('status')}")
    grm_no = extract_grm_no(first.get("data"))
    if not grm_no:
        return []

    list_payload = {
        "duration": "2",
        "srchViewType": "L",
        "itemGrmNo": grm_no,
        "srchGrmNo": grm_no,
        "srchSelPriority": priority,
        "srchSelSortBy": "D",
        "orderby": "DEPART_DT",
        "itemWebPri": "D",
        "bisCdId": "",
        "originCode": pol_code,
        "destinationCode": pod_code,
        "originCodeNm": pol_code,
        "destinationCodeNm": pod_code,
        "serviceTypeFrom": "CY",
        "serviceTypeTo": "CY",
        "cmdtCd": "FAK",
        "cmdtNm": "FAK",
        "date": sail_iso,
        "sort": "D",
        "cntrTpSzCdList": [cntr_code],
        "bkgCntrQtyList": ["1"],
        "esptPrmtnNo": "",
        "quotNo": "",
        "quotNoSeq": "",
        "esptDeclCustTpCd": "",
        "webUsrDocuSeq": "",
        "pageStart": "1",
        "pageEnd": "20",
    }
    second = hmm_form_post(driver, "/e-service/hiquote/quotation/selectPointToPointListNew.do", list_payload)
    second = second or {}
    if not second.get("ok"):
        raise RuntimeError(second.get("error") or f"selectPointToPointListNew HTTP {second.get('status')}")

    cards = parse_cards_html(driver, second.get("text") or "")
    out = []
    for c in cards:
        if c.get("spc") not in {"NM", "SS"}:
            continue
        if not c.get("quotNo") or not c.get("quotNoSeq"):
            continue
        price = hmm_form_post(
            driver,
            "/e-service/hiquote/quotation/getDspPricingInfo.do",
            {"quotNo": c["quotNo"], "quotNoSeq": c["quotNoSeq"], "spaceChk": c.get("spc") or "NM"},
            as_json=True,
        ) or {}
        pdata = (((price.get("data") or {}).get("RTN_DATA") or {}) if price.get("ok") else {})
        prepaid = parse_num(pdata.get("totPpdUsdAmt") or pdata.get("totUsdAmt"))
        if prepaid <= 0:
            continue
        etd = date_from_any(c.get("etd"))
        if not etd:
            continue
        c.update({"cntr": cntr_code, "prepaid": prepaid, "etd_date": etd})
        out.append(c)
    return out


def fetch_rate_detail(driver, card, cntr_code, pol_code, pod_code):
    payload = {
        "popupId": "popup_rateDetails",
        "quotNo": card["quotNo"],
        "quotNoSeq": card["quotNoSeq"],
        "ctrTp": cntr_code,
        "ctrQty": "1",
        "destination": pod_code,
        "origin": pol_code,
        "oogCheck": "",
    }
    res = hmm_form_post(driver, "/e-service/hiquote/quotation/openPopUp2.do", payload)
    if not res.get("ok"):
        raise RuntimeError(res.get("error") or f"openPopUp2 HTTP {res.get('status')}")
    return parse_rate_detail_html(driver, res.get("text") or "")


def is_origin_local_charge(item):
    s = re.sub(r"[^A-Z0-9]+", " ", str(item or "").upper()).strip()
    if not s:
        return False
    local_patterns = [
        "THC",
        "TERMINAL HANDLING",
        "DOC FEE",
        "DOCUMENT",
        "BILL",
        "B L",
        "SEAL",
        "TELEX",
        "SURRENDER",
        "ENS",
        "AMS",
        "AFS",
        "AFR",
        "MANIFEST",
        "ADVANCED MANIFEST",
        "ENTRY SUMMARY",
        "HEAVY",
        "OVERWEIGHT",
        "OWS",
    ]
    return any(p in s for p in local_patterns)


def summarize_detail(rows, include_origin_thc=False):
    parts = []
    flags = {
        "origin_thc": False,
        "manifest": False,
        "ows": False,
    }
    included_names = []
    for row in rows:
        section = row.get("section") or ""
        item = row.get("item") or ""
        item_up = item.upper()
        amount = parse_num(row.get("amount"))
        cur = (row.get("currency") or "USD").upper()
        is_origin_thc = section == "Origin Charges" and (
            "THC" in item_up or "TERMINAL HANDLING" in item_up
        )
        if is_origin_thc:
            flags["origin_thc"] = True
        if any(x in item_up for x in ["ENS", "AMS", "AFS", "AFR", "MANIFEST", "ENTRY SUMMARY"]):
            flags["manifest"] = True
        if any(x in item_up for x in ["HEAVY", "OVERWEIGHT", "OWS"]):
            flags["ows"] = True
        if section == "Destination Charges":
            continue
        if section == "Origin Charges" and is_origin_local_charge(item) and not (include_origin_thc and is_origin_thc):
            continue
        usd = to_usd(amount, cur)
        if usd:
            parts.append(usd)
            included_names.append(item)
    return {
        "parts": parts,
        "formula": excel_formula(parts),
        "total": math.ceil(sum(parts)) if parts else None,
        "flags": flags,
        "included_names": included_names,
    }


def build_remark(detail_summaries, country, pod):
    origin_thc = any(s["flags"].get("origin_thc") for s in detail_summaries if s)
    has_manifest_fee = any(s["flags"].get("manifest") for s in detail_summaries if s)
    has_ows = any(s["flags"].get("ows") for s in detail_summaries if s)
    china_route = is_china_destination(country, pod)
    remark = "SUBJECT TO THC, BILL, SEAL" if (origin_thc and not china_route) else "INCLUDED O.THC, SUBJECT TO BILL, SEAL"
    manifest_code = get_manifest_code(country, pod)
    remark = apply_manifest_rule(remark, country, pod)
    if has_manifest_fee and not manifest_code:
        remark += ", AMS"
    if has_ows:
        remark += ", OWS"
    return normalize_remark_text(remark)


def pick_schedule(candidates):
    min_etd = datetime.now().date() + timedelta(days=DATE_OFFSET_DAYS)
    future = [
        c for c in candidates
        if c.get("etd_date") and c["etd_date"] >= min_etd and etd_within_max(c["etd_date"])
    ]
    if not future:
        return []
    min_price = min(c["prepaid"] for c in future)
    near = [c for c in future if c["prepaid"] <= min_price + 40]
    by_date = {}
    for c in near:
        k = c["etd_date"]
        if k not in by_date or int(c.get("tt") or 9999) < int(by_date[k].get("tt") or 9999):
            by_date[k] = c
    selected = []
    for c in sorted(by_date.values(), key=lambda x: (x["etd_date"], int(x.get("tt") or 9999))):
        if not selected:
            selected.append(c)
        elif (c["etd_date"] - selected[-1]["etd_date"]).days >= 2 and (c["etd_date"] - selected[0]["etd_date"]).days <= 14:
            selected.append(c)
        if len(selected) >= 3:
            break
    return selected


def fmt_etd(selected):
    if not selected:
        return ""
    dates = [c["etd_date"] for c in selected]
    chunks = [f"{d.day}-{d.strftime('%b')}" for d in dates]
    if len(chunks) == 1:
        return chunks[0]
    if len(chunks) == 2:
        return f"{chunks[0]} & {chunks[1]}"
    if len({d.strftime('%b') for d in dates}) == 1:
        return f"{dates[0].day}, {dates[1].day}, {dates[2].day}-{dates[0].strftime('%b')}"
    return ", ".join(chunks)


def fmt_tt(selected):
    tts = [int(c.get("tt") or 0) for c in selected if int(c.get("tt") or 0) > 0]
    if not tts:
        return ""
    if len(set(tts)) == 1:
        return f"{tts[0]} days"
    return f"{min(tts)}-{max(tts)} days"


def format_vessel(selected):
    lines = []
    ts_seen = []
    for c in selected:
        ts = c.get("ts") or "DIRECT"
        vessel = clean_vessel(c.get("vessel"))
        lines.append(
            f"{vessel} / ETD: {c['etd_date'].day}-{c['etd_date'].strftime('%b')}"
            f" / Transit time: {int(c.get('tt') or 0)} Days / Transshipment Port: {ts}"
        )
        if ts not in ts_seen:
            ts_seen.append(ts)
    return "\n".join(lines), " or\n".join(ts_seen) if ts_seen else "DIRECT"


def find_matching_card(cards, selected_card):
    if not cards:
        return None
    same_date = [c for c in cards if c.get("etd_date") == selected_card.get("etd_date")]
    if same_date:
        return sorted(same_date, key=lambda x: (x.get("prepaid", 999999), int(x.get("tt") or 9999)))[0]
    return sorted(cards, key=lambda x: (x.get("prepaid", 999999), x.get("etd_date")))[0]


def fetch_by_priority(driver, pol_code, pod_code, sail_date, priority, label_name):
    by_container = {}
    log(f"   Thu {label_name} ({priority})...")
    for label, code in CONTAINERS.items():
        try:
            cards = fetch_container_cards(driver, pol_code, pod_code, sail_date, code, priority=priority)
        except Exception as exc:
            msg = str(exc).splitlines()[0] if str(exc) else type(exc).__name__
            log(f"      {label}: loi API/timeout -> coi nhu 0 card ({msg[:100]})")
            cards = []
            if "timeout" in msg.lower() or "timed out" in msg.lower():
                by_container[label] = cards
                log(f"      {label}: 0 bookable cards")
                log(f"      HMM API timeout ở {label}; bỏ qua container còn lại của {label_name}.")
                break
        by_container[label] = cards
        log(f"      {label}: {len(cards)} bookable cards")
    base_cards = by_container.get("20") or by_container.get("40") or by_container.get("40HC") or []
    return by_container, base_cards


def search_one(driver, country, pol_name, pod_name):
    pol_code = resolve_port_code(pol_name)
    pod_code = resolve_port_code(pod_name)
    if not pol_code or not pod_code:
        return {"error": f"UNKNOWN PORT CODE: {pol_name}->{pod_name}"}

    sail_date = datetime.now().date() + timedelta(days=DATE_OFFSET_DAYS)
    log(f"API {pol_code}->{pod_code} | ETD_FROM={sail_date.strftime('%Y-%m-%d')}")

    by_container, base_cards = fetch_by_priority(driver, pol_code, pod_code, sail_date, "D", "Direct")
    route_mode = "DIRECT"
    if not base_cards:
        log("   Direct khong co gia/lich bookable -> thu Indirect...")
        by_container, base_cards = fetch_by_priority(driver, pol_code, pod_code, sail_date, "T", "Indirect")
        route_mode = "INDIRECT"

    selected = pick_schedule(base_cards)
    if not selected:
        return {"error": "NO BOOKABLE PRICE"}

    detail_summaries = {}
    china_route = is_china_destination(country, pod_name)
    for label, code in CONTAINERS.items():
        match = find_matching_card(by_container.get(label) or [], selected[0])
        if not match:
            detail_summaries[label] = None
            continue
        rows = fetch_rate_detail(driver, match, code, pol_code, pod_code)
        detail_summaries[label] = summarize_detail(rows, include_origin_thc=china_route)

    if not any(s and s.get("formula") for s in detail_summaries.values()):
        return {"error": "NO PRICE DETAIL"}

    vessel_text, ts_text = format_vessel(selected)
    remark = build_remark([x for x in detail_summaries.values() if x], country, pod_name)
    free_time = selected[0].get("freeTime") or ""
    valid = get_valid_date([c["etd_date"] for c in selected])

    result = {
        "formula_20": (detail_summaries.get("20") or {}).get("formula"),
        "formula_40": (detail_summaries.get("40") or {}).get("formula"),
        "formula_40hc": (detail_summaries.get("40HC") or {}).get("formula"),
        "price_20": (detail_summaries.get("20") or {}).get("total"),
        "price_40": (detail_summaries.get("40") or {}).get("total"),
        "price_40hc": (detail_summaries.get("40HC") or {}).get("total"),
        "etd": fmt_etd(selected),
        "tt": fmt_tt(selected),
        "valid": valid,
        "remark": remark,
        "free_time": free_time,
        "vessel_info": vessel_text,
        "transshipment": ts_text,
        "route_mode": route_mode,
    }
    return result


def clear_result_cells(ws, row_i):
    for col in [6, 7, 8, 9, 10, 11, 13, 14, 15, 16]:
        ws.cell(row=row_i, column=col).value = None


def write_result(ws, row_i, result):
    ws.cell(row=row_i, column=6).value = result.get("formula_20") or result.get("price_20") or "-"
    ws.cell(row=row_i, column=7).value = result.get("formula_40") or result.get("price_40") or ""
    ws.cell(row=row_i, column=8).value = result.get("formula_40hc") or result.get("price_40hc") or ""
    ws.cell(row=row_i, column=9).value = result.get("etd", "")
    ws.cell(row=row_i, column=10).value = result.get("tt", "")
    ws.cell(row=row_i, column=11).value = result.get("valid", "")
    ws.cell(row=row_i, column=13).value = result.get("remark", "")
    ws.cell(row=row_i, column=14).value = result.get("free_time", "")
    ws.cell(row=row_i, column=15).value = result.get("vessel_info", "")
    ws.cell(row=row_i, column=16).value = result.get("transshipment", "")
    wrap = Alignment(wrap_text=True, vertical="top")
    for col in [13, 15, 16]:
        ws.cell(row=row_i, column=col).alignment = wrap


def collect_rows(ws):
    rows = []
    target_row = None
    if SINGLE_ROW:
        try:
            target_row = int(SINGLE_ROW)
            log(f"[SINGLE_ROW] Chi chay dong {target_row} theo lenh tu main.py")
        except Exception:
            log(f"[WARN] SINGLE_ROW khong hop le: {SINGLE_ROW}")
    for i in range(2, ws.max_row + 1):
        if target_row and i != target_row:
            continue
        country = str(ws.cell(i, 2).value or "").strip()
        pol = str(ws.cell(i, 3).value or "").strip()
        pod = str(ws.cell(i, 4).value or "").strip()
        carrier = str(ws.cell(i, 5).value or "").strip().upper()
        if carrier not in {"HMM", "HYUNDAI", "HYUNDAI MERCHANT MARINE"}:
            continue
        if FILTER_POL and pol.upper() != FILTER_POL:
            continue
        if FILTER_POD and pod.upper() != FILTER_POD:
            continue
        rows.append((i, country, pol, pod))
    return rows


def main():
    print("\n" + "=" * 50)
    print("   HMM Price Checker")
    print("=" * 50 + "\n")
    start_edge_if_needed()
    driver = connect_driver()
    ensure_login(driver)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    rows = collect_rows(ws)
    log(f"Co {len(rows)} dong can check (carrier = HMM)")

    for idx, (row_i, country, pol, pod) in enumerate(rows, start=1):
        print(f"\n========== [{idx}/{len(rows)}] DONG {row_i}: {pol} -> {pod} | HMM")
        try:
            result = search_one(driver, country, pol, pod)
            if result.get("error"):
                log(f"   SKIP: {result['error']} -> de trong")
                clear_result_cells(ws, row_i)
            else:
                log(f"   ETD={result['etd']} | TT={result['tt']} | Valid={result['valid']}")
                log(f"   20={result.get('formula_20') or result.get('price_20')}")
                log(f"   40={result.get('formula_40') or result.get('price_40')}")
                log(f"   HC={result.get('formula_40hc') or result.get('price_40hc')}")
                log(f"   Remark: {result['remark']}")
                write_result(ws, row_i, result)
        except Exception as e:
            msg = str(e).splitlines()[0] if str(e) else type(e).__name__
            log(f"   LOI row {row_i}: {msg} -> de trong")
            clear_result_cells(ws, row_i)
        try:
            wb.save(EXCEL_PATH)
            log(f"   Da luu row {row_i}")
        except PermissionError:
            log("   LOI GHI FILE: dong Excel truoc khi chay!")
            raise
        time.sleep(1)

    wb.close()
    log("HOAN TAT bot HMM")


if __name__ == "__main__":
    main()
