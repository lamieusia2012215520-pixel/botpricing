"""
KMTC Price Checker
- Login tự động
- Freight Inquiry: nhập POL/POD → search → đọc giá
- Nếu POD dropdown có nhiều option → tạo thêm hàng Excel, check tất cả
- Không tạo hàng trùng nếu đã có trong Excel
- Không nhập POD 2 lần — giữ dropdown từ lần get_options để chọn luôn
- Port aliases: Pusan=Busan, Inchon=Incheon, ...
- Clean POD name: 'Chennai, India (MAA)' → 'CHENNAI'
Selenium + Edge port 9526
"""
import math, re, time, random, os, json, calendar, sys, io, threading
import urllib.request, urllib.parse, urllib.error
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException, WebDriverException
import openpyxl
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from remark_rules import build_subject_remark, charge_amount_to_usd

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

_orig_print = print
def print(*args, **kwargs):
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    _orig_print(f"[{ts}]", *args, **kwargs)

current_folder = os.getcwd()
driver_path    = os.path.join(current_folder, "msedgedriver.exe")
excel_path     = os.environ.get("EXCEL_PATH", os.path.join(current_folder, "input_gia.xlsx"))
FILTER_POL     = os.environ.get("FILTER_POL", "").strip().upper()
FILTER_POD     = os.environ.get("FILTER_POD", "").strip().upper()
SINGLE_ROW     = os.environ.get("SINGLE_ROW", "").strip()
DATE_OFFSET_DAYS = parse_date_offset_days()
try:
    KMTC_REGISTERED_FREIGHT_WAIT_SECONDS = max(2.0, float(os.environ.get("KMTC_REGISTERED_FREIGHT_WAIT_SECONDS", "4")))
except ValueError:
    KMTC_REGISTERED_FREIGHT_WAIT_SECONDS = 4.0

def _excel_formula_from_parts(parts):
    tokens = []
    for raw in parts:
        try:
            amount = float(raw)
        except (TypeError, ValueError):
            continue
        if abs(amount) < 1e-9:
            continue
        sign = "-" if amount < 0 else "+"
        amount = abs(amount)
        if abs(amount - round(amount)) < 1e-9:
            text = str(int(round(amount)))
        else:
            text = f"{amount:.2f}".rstrip("0").rstrip(".")
        tokens.append((sign if tokens else ("-" if sign == "-" else "")) + text)
    return "=" + "".join(tokens) if tokens else None

try:
    KMTC_SLEEP_SCALE = float(os.environ.get("KMTC_SLEEP_SCALE", "0.45"))
except ValueError:
    KMTC_SLEEP_SCALE = 0.45
try:
    KMTC_DROPDOWN_TIMEOUT = float(os.environ.get("KMTC_DROPDOWN_TIMEOUT", "2.5"))
except ValueError:
    KMTC_DROPDOWN_TIMEOUT = 2.5

import subprocess
import socket

def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('127.0.0.1', port)) == 0

def _wait_port(port, timeout=8):
    """FIX: Poll port thay vì sleep cứng."""
    end = time.time() + timeout
    while time.time() < end:
        if is_port_in_use(port):
            return True
        time.sleep(0.2)
    return False

def restart_kmtc_edge():
    global driver
    print("[HỆ THỐNG] Restart Edge KMTC để dọn sạch tab dư...")
    try:
        driver.quit()
    except Exception:
        pass
    try:
        subprocess.run([
            "powershell", "-NoProfile", "-Command",
            "Get-CimInstance Win32_Process -Filter \"name='msedge.exe'\" | "
            "Where-Object { $_.CommandLine -like '*--remote-debugging-port=9526*' } | "
            "ForEach-Object { Stop-Process -Id $_.ProcessId -Force }"
        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=10)
    except Exception:
        pass
    time.sleep(1)
    subprocess.Popen([
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        "--remote-debugging-port=9526",
        r"--user-data-dir=C:\edge_kmtc",
        "--disable-background-timer-throttling",
        "--disable-renderer-backgrounding",
        "--disable-backgrounding-occluded-windows"
    ], stdin=subprocess.DEVNULL, stdout=subprocess.DEVNULL,
       stderr=subprocess.DEVNULL)
    _wait_port(9526, timeout=8)
    driver = webdriver.Edge(service=service, options=edge_options)

if not is_port_in_use(9526):
    print("[HỆ THỐNG] Edge KMTC chưa mở. Đang tự động khởi động...")
    try:
        subprocess.Popen([
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            "--remote-debugging-port=9526",
            r"--user-data-dir=C:\edge_kmtc",
            "--disable-background-timer-throttling",
            "--disable-renderer-backgrounding",
            "--disable-backgrounding-occluded-windows"
        ], stdin=subprocess.DEVNULL, stdout=subprocess.DEVNULL,
           stderr=subprocess.DEVNULL)
        _wait_port(9526, timeout=8)
    except: pass
else:
    print("[HỆ THỐNG] Edge KMTC đã mở sẵn. Bỏ qua lệnh khởi động trình duyệt.")

edge_options = Options()
edge_options.add_experimental_option("debuggerAddress", "127.0.0.1:9526")
service = Service(executable_path=driver_path)
driver  = webdriver.Edge(service=service, options=edge_options)
try:
    driver.maximize_window()
except Exception:
    try:
        info = driver.execute_cdp_cmd("Browser.getWindowForTarget", {})
        driver.execute_cdp_cmd("Browser.setWindowBounds", {
            "windowId": info["windowId"],
            "bounds": {"windowState": "maximized"}
        })
    except Exception:
        pass

_KMTC_TABS_CLEANED = False

def close_all_kmtc_tabs():
    """Close every tab in the dedicated KMTC Edge profile when the bot exits."""
    global _KMTC_TABS_CLEANED
    if _KMTC_TABS_CLEANED:
        return
    _KMTC_TABS_CLEANED = True
    drv = globals().get("driver")

    result = {"closed": 0, "error": None}

    def _close_tabs():
        try:
            handles = list(drv.window_handles)
            for handle in reversed(handles):
                try:
                    if handle not in drv.window_handles:
                        continue
                    drv.switch_to.window(handle)
                    drv.close()
                    result["closed"] += 1
                except Exception:
                    continue
            try:
                drv.quit()
            except Exception:
                pass
        except Exception as exc:
            result["error"] = exc

    worker = None
    if drv is not None:
        worker = threading.Thread(target=_close_tabs, daemon=True)
        worker.start()
        worker.join(timeout=5)

    # Fallback cứng cho trường hợp Selenium close/quit bị treo. Chỉ tắt Edge
    # profile KMTC và msedgedriver do chính process bot này sinh ra.
    ps = f'''
Get-CimInstance Win32_Process -Filter "name = 'msedge.exe'" |
    Where-Object {{
        $cmd = ([string]$_.CommandLine).ToLower()
        $cmd.Contains('c:\\edge_kmtc') -or $cmd.Contains('--remote-debugging-port=9526')
    }} |
    ForEach-Object {{ Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }}
Get-CimInstance Win32_Process -Filter "name = 'msedgedriver.exe'" |
    Where-Object {{ $_.ParentProcessId -eq {os.getpid()} }} |
    ForEach-Object {{ Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }}
'''
    kwargs = {
        "stdin": subprocess.DEVNULL,
        "stdout": subprocess.DEVNULL,
        "stderr": subprocess.DEVNULL,
        "timeout": 12,
    }
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
    needs_force_close = bool(worker is not None and worker.is_alive()) or is_port_in_use(9526)
    if needs_force_close:
        try:
            subprocess.run(["powershell", "-NoProfile", "-Command", ps], **kwargs)
        except Exception as exc:
            print(f"[HỆ THỐNG] ⚠️ KMTC force-close Edge lỗi: {type(exc).__name__}")

    if worker is not None and worker.is_alive():
        worker.join(timeout=2)
    print(f"[HỆ THỐNG] ✅ KMTC đã đóng browser sau khi hoàn tất ({result['closed']} tab qua Selenium).")

_KMTC_PREV_EXCEPTHOOK = sys.excepthook

def _kmtc_excepthook(exc_type, exc_value, exc_traceback):
    close_all_kmtc_tabs()
    _KMTC_PREV_EXCEPTHOOK(exc_type, exc_value, exc_traceback)

sys.excepthook = _kmtc_excepthook

BASE_URL   = "https://www.ekmtc.com/index.html#/main"
KMTC_GROUP = {"KMTC"}

EXCLUDE_EXPORT = {"SEAL CHARGE"}

# POD cần dùng CMP làm POL schedule (khi POL gốc là HO CHI MINH)
CMP_PODS = {"PORT KLANG", "JEBEL ALI", "DAMMAM", "MUNDRA", "NHAVA SHEVA"}
CHINA_PODS = {
    "SHANGHAI", "SHEKOU", "XINGANG", "TIANJIN", "XIAMEN", "NINGBO",
    "QINGDAO", "YANTIAN", "NANSHA", "DALIAN", "GUANGZHOU", "GUANG ZHOU",
    "HONG KONG", "KAOHSIUNG",
}

POL_SCHEDULE_MAP = {
    "HO CHI MINH": "hochiminh",
    "HAI PHONG":   "haiphong",
}

# ===================================================================================
# --- PORT ALIASES: Excel input → KMTC search term ---
# ===================================================================================
PORT_ALIASES = {
    # ---- VIETNAM ----
    "HO CHI MINH":                  "HOCHIMINH",
    "HOCHIMINH CITY":               "HOCHIMINH",
    "HCM":                          "HOCHIMINH",
    "SAIGON":                       "HOCHIMINH",
    "SGN":                          "HOCHIMINH",
    "HAI PHONG":                    "HAIPHONG",
    "HPH":                          "HAIPHONG",
    "CAI MEP":                      "CAI MEP",
    "CMP":                          "CAI MEP",
    "CAT LAI":                      "CAT LAI GIANG NAM",
    "TAN CANG HIEP PHUOC":          "TAN CANG HIEP PHUOC PORT",

    # ---- MALAYSIA ----
    "PORT KLANG":                   "PORT KELANG",
    "PORT KLANG (NORTHPORT)":       "PORT KELANG",
    "PORT KLANG (WESTPORT)":        "PORT KELANG",
    "KELANG":                       "PORT KELANG",
    "KLANG":                        "PORT KELANG",
    "PORT KELANG":                  "PORT KELANG",
    "TANJUNG PELEPAS":              "TANJUNG PELAPAS",
    "PORT TANJUNG PELEPAS":         "TANJUNG PELAPAS",
    "PTP":                          "TANJUNG PELAPAS",
    "TANJUNG PELAPAS":              "TANJUNG PELAPAS",

    # ---- INDONESIA ----
    "TANJUNG PRIOK":                "JAKARTA",
    "TG PRIOK":                     "JAKARTA",
    "TG. PRIOK":                    "JAKARTA",
    "SURABAYA":                     "SURABAYA",
    "UJUNG PANDANG":                "UJUNG PANDANG",
    "MAKASSAR":                     "MAKASSAR",

    # ---- INDIA ----
    "JNPT":                         "NHAVA SHEVA",
    "BOMBAY":                       "MUMBAI",
    "CALCUTTA":                     "KOLKATA",
    "MADRAS":                       "CHENNAI",
    "VIZAG":                        "VISAKHAPATNAM(VIZAG)",
    "VISHAKHAPATNAM":               "VISAKHAPATNAM(VIZAG)",
    "VISAKHAPATNAM":                "VISAKHAPATNAM(VIZAG)",
    "COCHIN":                       "COCHIN",
    "NHAVA SHEVA":                  "NHAVA SHEVA",
    "CHENNAI (J. MATADEE FTZ)":     "J. MATADEE CHENNAI FREE TRADE ZONE",

    # ---- CHINA ----
    "CANTON":                       "GUANG ZHOU",
    "GUANGZHOU":                    "GUANG ZHOU",
    "XI AN":                        "XI AN",
    "XIAN":                         "XI AN",
    "XINGANG":                      "XINGANG",
    "TIANJIN":                      "XINGANG",
    "TIENTSIN":                     "XINGANG",
    "PEKING":                       "BEIJING",

    # ---- UAE ----
    "DUBAI":                        "JEBEL ALI",
    "UMM AL QUWAIN":                "UMM AL QAIWAIN",
    "UMM AL QAIWIN":                "UMM AL QAIWAIN",
    "UMM AL QUIWAIN":               "UMM AL QAIWAIN",
    "RAS AL KHAIMAH":               "RAS AL KHAIMAH",

    # ---- KOREA ----
    "PUSAN":                        "BUSAN",
    "INCHON":                       "INCHEON",
    "INCH'ON":                      "INCHEON",
    "KWANGYANG":                    "KWANGYANG",
    "CHEJU":                        "JEJU",
    "MASAN":                        "CHANGWON",

    # ---- THAILAND ----
    "SRIHRACHA":                    "LAEM CHABANG",
    "SIHRACHA":                     "LAEM CHABANG",
    "SI RACHA":                     "LAEM CHABANG",
    "LAEM CHABANG":                 "LAEM CHABANG",
    "BANGKOK":                      "BANGKOK",

    # ---- JAPAN ----
    "KOBE":                         "KOBE",
    "OSAKA":                        "OSAKA",
    "TOKYO":                        "TOKYO",
    "YOKOHAMA":                     "YOKOHAMA",
    "NAGOYA":                       "NAGOYA",

    # ---- SAUDI ARABIA ----
    "JEDDAH":                       "JEDDAH",
    "DAMMAM":                       "DAMMAM",
    "JUBAIL":                       "JUBAIL",
    "RIYADH":                       "RIYADH",
    "YANBU":                        "YANBU AL-BAHR",

    # ---- OTHER ----
    "COLOMBO":                      "COLOMBO",
    "SINGAPORE":                    "SINGAPORE",
    "HONG KONG":                    "HONG KONG",
    "HKG":                          "HONG KONG",
    "KAOHSIUNG":                    "KAOHSIUNG",
    "KHH":                          "KAOHSIUNG",
    "BUSAN":                        "BUSAN",
    "INCHEON":                      "INCHEON",
    "SOKHNA":                       "SOKHNA",
    "PORT SAID":                    "SOKHNA",
    "MOMBASA":                      "MOMBASA",
    "DAR ES SALAAM":                "DAR ES SALAAM",
    "VLADIVOSTOK":                  "VLADIVOSTOK",
    "YANGON":                       "YANGON",
    "RANGOON":                      "YANGON",
    "KARACHI":                      "KARACHI",
    "CHITTAGONG":                   "CHATTOGRAM",
    "DHAKA":                        "DHAKA",
    "PHNOM PENH":                   "PHNOM PENH",
    "SIHANOUKVILLE":                "SIHANOUKVILLE",
    "MELBOURNE":                    "MELBOURNE",
    "SYDNEY":                       "SYDNEY",
    "LONG BEACH":                   "LONG BEACH",
    "LOS ANGELES":                  "LONG BEACH",
    "MANZANILLO":                   "MANZANILLO",
    "DOHA":                         "DOHA",
    "HAMAD":                        "HAMAD",
    "AQABA":                        "AQABA",
    "SOHAR":                        "SOHAR",
    "MUSCAT":                       "SOHAR",
    "BUSHEHR":                      "BUSHEHR / PERSIAN GULF",
    "BANDAR ABBAS":                 "BANDAR-ABBAS",
    "BANDAR IMAM":                  "BANDAR IMAM KHOMEINI",
    "KHOR FAKKAN":                  "KHOR FAKKAN",
    "FUJAIRAH":                     "FUJAIRAH",
    "SHARJAH":                      "SHARJAH",
    "ABU DHABI":                    "ABU DHABI",
    "AJMAN":                        "AJMAN",
    "SHUWAIKH":                     "SHUWAIKH",
    "SHUAIBA":                      "SHUAIBA",
    "KUWAIT":                       "SHUWAIKH",
}

def resolve_alias(port_name):
    upper = port_name.strip().upper()
    return PORT_ALIASES.get(upper, port_name.strip())

def clean_pod_name(raw):
    """Lấy phần tên port trước dấu phẩy"""
    if not raw:
        return raw
    return raw.strip().split(',')[0].strip()

def pick_best_pod_option(pod_texts, pod_search):
    """
    FIX: Chọn option POD chuẩn nhất khi dropdown có nhiều kết quả.
    Ưu tiên theo thứ tự:
      1. Option chứa '(ALL PORTS)'
      2. Option khớp chính xác tên cảng (không có ICD / ICO / CONTAINER YARD)
      3. Option có CY hoặc PORT
      4. Option đầu tiên (fallback)
    """
    if not pod_texts:
        return None

    pod_up = pod_search.upper()

    # Ưu tiên 1: ALL PORTS
    for opt in pod_texts:
        if "(ALL PORTS)" in opt.upper():
            print(f"   🎯 Chọn option '(ALL PORTS)': {opt}")
            return opt

    # Ưu tiên 2: khớp chính xác phần tên trước dấu phẩy, không có ICD/ICO/INLAND
    for opt in pod_texts:
        name_part = clean_pod_name(opt).upper()
        if name_part == pod_up and not any(x in opt.upper() for x in ("ICD", "ICO", "INLAND", "DRY PORT")):
            print(f"   🎯 Chọn option khớp chính xác: {opt}")
            return opt

    # Ưu tiên 3: chứa tên cảng + (CY) hoặc PORT
    for opt in pod_texts:
        opt_up = opt.upper()
        if pod_up in opt_up and any(x in opt_up for x in ("CY", " PORT", "(PORT)")):
            print(f"   🎯 Chọn option CY/PORT: {opt}")
            return opt

    # Ưu tiên 4: chứa tên cảng, không phải ICD
    for opt in pod_texts:
        if pod_up in opt.upper() and not any(x in opt.upper() for x in ("ICD", "ICO", "INLAND")):
            print(f"   🎯 Chọn option non-ICD: {opt}")
            return opt

    # Fallback: option đầu tiên
    print(f"   ⚠️ Không tìm được option tối ưu, lấy đầu tiên: {pod_texts[0]}")
    return pod_texts[0]

# ===================================================================================
# --- KMTC_TO_EXCEL: KMTC dropdown result → tên chuẩn ghi vào Excel ---
# ===================================================================================
KMTC_TO_EXCEL = {
    # ---- VIETNAM ----
    "HOCHIMINH":                            "HO CHI MINH",
    "HAIPHONG":                             "HAI PHONG",
    "CAI MEP":                              "CAI MEP",
    "CAT LAI GIANG NAM":                    "CAT LAI GIANG NAM",
    "TAN CANG CAI MEP":                     "TAN CANG CAI MEP",
    "TAN CANG HIEP PHUOC PORT":             "TAN CANG HIEP PHUOC",
    "TAN CANG NHON TRACH":                  "TAN CANG NHON TRACH",
    "KHANH HOI":                            "KHANH HOI",
    "SPITC":                                "SPITC",
    "VICT":                                 "VICT",
    "ICD GIA LAM":                          "ICD GIA LAM",
    "ICD MY DINH":                          "ICD MY DINH",
    "ICD PHUOC LONG":                       "ICD PHUOC LONG",
    "ICD SONG THAN":                        "ICD SONG THAN",
    "ICD SOTRANS":                          "ICD SOTRANS",
    "ICD TRANSIMEX":                        "ICD TRANSIMEX",
    "MY DINH":                              "MY DINH",

    # ---- MALAYSIA ----
    "PORT KELANG NORTHPORT":                "PORT KLANG (NORTHPORT)",
    "PORT KELANG WESTPORT":                 "PORT KLANG (WESTPORT)",
    "TANJUNG PELAPAS":                      "TANJUNG PELEPAS",
    "PASIR GUDANG":                         "PASIR GUDANG",
    "PENANG":                               "PENANG",
    "KOTA KINABALU":                        "KOTA KINABALU",
    "KUCHING":                              "KUCHING",
    "BINTULU":                              "BINTULU",
    "IPOH/PERAK":                           "IPOH",

    # ---- INDIA ----
    "J. MATADEE CHENNAI FREE TRADE ZONE":   "CHENNAI (J. MATADEE FTZ)",
    "NHAVA SHEVA":                          "NHAVA SHEVA",
    "VISAKHAPATNAM(VIZAG)":                 "VISAKHAPATNAM",
    "KOLKATA":                              "KOLKATA",
    "MUMBAI":                               "MUMBAI",
    "CHENNAI":                              "CHENNAI",
    "COCHIN":                               "COCHIN",
    "HALDIA":                               "HALDIA",
    "HAZIRA":                               "HAZIRA",
    "MUNDRA":                               "MUNDRA",
    "KANDLA":                               "KANDLA",
    "TUTICORIN":                            "TUTICORIN",
    "KATTUPALLI":                           "KATTUPALLI",

    # ---- CHINA ----
    "GUANG ZHOU":                           "GUANGZHOU",
    "XI AN":                                "XI AN",
    "XI'AN":                                "XI AN",
    "XINGANG":                              "XINGANG",

    # ---- UAE ----
    "UMM AL QAIWAIN":                       "UMM AL QAIWAIN",
    "JEBEL ALI":                            "JEBEL ALI",
    "ABU DHABI":                            "ABU DHABI",
    "SHARJAH":                              "SHARJAH",
    "SHARJAH ICD":                          "SHARJAH ICD",
    "FUJAIRAH":                             "FUJAIRAH",
    "KHOR FAKKAN":                          "KHOR FAKKAN",
    "RAS AL KHAIMAH":                       "RAS AL KHAIMAH",
    "AJMAN":                                "AJMAN",

    # ---- TAIWAN ----
    "KAOHSIUNG":                            "KAOHSIUNG",
    "KEELUNG":                              "KEELUNG",
    "TAICHUNG":                             "TAICHUNG",
    "TAOYUAN":                              "TAOYUAN",

    # ---- SAUDI ARABIA ----
    "YANBU AL-BAHR":                        "YANBU",

    # ---- MYANMAR ----
    "YANGON":                               "YANGON",

    # ---- EGYPT ----
    "SOKHNA":                               "SOKHNA",
}

# ===================================================================================
# --- HELPERS ---
# ===================================================================================
def rand_sleep(a=0.45, b=1.05):
    time.sleep(max(0.05, random.uniform(a, b) * KMTC_SLEEP_SCALE))

def activate_tab(handle, label=""):
    """Switch Selenium + ask Chromium to bring the tab to the foreground."""
    driver.switch_to.window(handle)
    try:
        driver.execute_cdp_cmd("Target.activateTarget", {"targetId": handle})
    except Exception:
        pass
    try:
        driver.execute_script("window.focus(); document.hasFocus && document.hasFocus();")
    except Exception:
        pass
    if label:
        print(f"   🧭 Active tab: {label}")

def human_click(element):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
    rand_sleep(0.15, 0.3)
    ActionChains(driver).move_to_element_with_offset(
        element, random.randint(-2, 2), random.randint(-2, 2)
    ).pause(random.uniform(0.1, 0.2)).click().perform()
    rand_sleep(0.15, 0.3)

def clear_and_type(inp, text):
    driver.execute_script("arguments[0].click(); arguments[0].focus();", inp)
    rand_sleep(0.1, 0.2)
    inp.send_keys(Keys.CONTROL + "a")
    inp.send_keys(Keys.DELETE)
    rand_sleep(0.1, 0.15)
    inp.send_keys(text)
    rand_sleep(0.3, 0.5)

def find_visible(xpath_list, timeout=6):
    end = time.time() + timeout
    while time.time() < end:
        for xp in xpath_list:
            try:
                for el in driver.find_elements(By.XPATH, xp):
                    if el.is_displayed():
                        return el
            except Exception:
                pass
        time.sleep(0.1)
    raise Exception(f"Không tìm thấy element: {xpath_list[0]}")

def visible_dropdown_items(root_id):
    items = driver.find_elements(By.CSS_SELECTOR, f"#{root_id} .list-group-item")
    return [it for it in items if it.is_displayed()]

def dismiss_simplert():
    try:
        WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.simplert--shown"))
        )
        btns = driver.find_elements(By.CSS_SELECTOR,
            "div.simplert--shown .simplert__btn, "
            "div.simplert--shown a, "
            "div.simplert--shown button"
        )
        for btn in btns:
            if btn.is_displayed():
                txt = btn.text.strip().lower()
                if txt in ("no", "cancel", "닫기", "close", "확인", "ok"):
                    driver.execute_script("arguments[0].click();", btn)
                    rand_sleep(0.3, 0.5)
                    WebDriverWait(driver, 5).until_not(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div.simplert--shown"))
                    )
                    return
        if btns:
            driver.execute_script("arguments[0].click();", btns[-1])
            rand_sleep(0.3, 0.5)
            WebDriverWait(driver, 5).until_not(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.simplert--shown"))
            )
    except:
        pass

# ===================================================================================
# --- LOGIN ---
# ===================================================================================
def ensure_logged_in():
    try:
        login_link = WebDriverWait(driver, 3).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a.login.cursor_pointer"))
        )
        print("   🔐 Phát hiện nút login → bấm...")
        human_click(login_link)
        rand_sleep(0.5, 0.8)
        login_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a.button.blue.sm"))
        )
        human_click(login_btn)
        print("   🔐 Đã bấm Login → chờ reload...")
        rand_sleep(2.0, 3.0)
        WebDriverWait(driver, 15).until(
            lambda d: d.find_elements(By.XPATH, "//a[contains(@href,'shipping-cost')]") or
                      d.find_elements(By.XPATH, "//*[contains(text(),'로그아웃') or contains(text(),'Logout')]")
        )
        print("   ✅ Login xong")
    except:
        pass

# ===================================================================================
# --- MỞ FREIGHT INQUIRY ---
# ===================================================================================
def open_freight_inquiry():
    driver.get(BASE_URL)
    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))
    try:
        find_visible(["//*[@id='polPol2']//input"], timeout=1)
    except Exception:
        fare_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a.link_icon.fare"))
        )
        driver.execute_script("arguments[0].click();", fare_link)
    find_visible(["//*[@id='polPol2']//input"], timeout=12)
    rand_sleep(0.1, 0.2)

def ensure_freight_ready():
    if "#/main" not in (driver.current_url or "").lower():
        open_freight_inquiry()
        return
    try:
        find_visible(["//*[@id='polPol2']//input"], timeout=1)
    except Exception:
        open_freight_inquiry()

# ===================================================================================
# --- NHẬP POL (skip nếu đã đúng) ---
# ===================================================================================
def type_pol(pol):
    pol_inp = find_visible([
        "//*[@id='polPol2']//input",
        "//*[@id='polPol']//input",
        "//input[@placeholder='Enter Origin']",
    ], timeout=8)
    current = pol_inp.get_attribute("value") or ""
    if pol.upper() in current.upper() and current.strip():
        print(f"   ✅ POL giữ nguyên: {current.strip()[:50]}")
        return True

    driver.execute_script("arguments[0].value = '';", pol_inp)
    driver.execute_script("arguments[0].click(); arguments[0].focus();", pol_inp)
    rand_sleep(0.1, 0.2)
    pol_inp.send_keys(Keys.CONTROL + "a")
    pol_inp.send_keys(Keys.DELETE)
    rand_sleep(0.1, 0.15)
    pol_inp.send_keys(pol)
    rand_sleep(0.3, 0.5)
    try:
        WebDriverWait(driver, KMTC_DROPDOWN_TIMEOUT).until(
            lambda d: visible_dropdown_items("polPol2") or visible_dropdown_items("polPol")
        )
        visible = visible_dropdown_items("polPol2") or visible_dropdown_items("polPol")
        if visible:
            txt = visible[0].text.strip()
            driver.execute_script("arguments[0].click();", visible[0])
            rand_sleep(0.1, 0.15)
            driver.execute_script("arguments[0].click();", visible[0])
            print(f"   ✅ POL: {txt[:50]}")
            try:
                WebDriverWait(driver, 3).until_not(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "#polPol2 .list-group-item"))
                )
            except:
                pass
            return True
    except:
        pass
    pol_inp.send_keys(Keys.RETURN)
    rand_sleep(0.2, 0.3)
    return True

# ===================================================================================
# --- NHẬP POD VÀ LẤY DANH SÁCH OPTIONS (skip nếu đã đúng, giữ dropdown để chọn luôn) ---
# Trả về (texts, elements) — elements để chọn trực tiếp không nhập lại
# ===================================================================================
def type_pod_get_options(pod):
    pod_inp = find_visible([
        "//*[@id='podPod2']//input",
        "//*[@id='podPod']//input",
        "//input[@placeholder='Enter Destination']",
    ], timeout=8)
    current = pod_inp.get_attribute("value") or ""

    # Nếu POD hiện tại đã chứa keyword → click vào input để mở lại dropdown
    if pod.upper() in current.upper() and current.strip():
        print(f"   ✅ POD giữ nguyên: {current.strip()[:50]}")
        driver.execute_script("arguments[0].click(); arguments[0].focus();", pod_inp)
        rand_sleep(0.2, 0.3)
    else:
        clear_and_type(pod_inp, pod)

    try:
        WebDriverWait(driver, KMTC_DROPDOWN_TIMEOUT).until(
            lambda d: visible_dropdown_items("podPod2") or visible_dropdown_items("podPod")
        )
        visible = visible_dropdown_items("podPod2") or visible_dropdown_items("podPod")
        texts = [it.text.strip() for it in visible]
        return texts, visible
    except:
        return [], []

def select_pod_by_element(target_text, pod_elements):
    """Chọn POD từ elements đã có sẵn — không nhập lại"""
    for it in pod_elements:
        try:
            if it.text.strip() == target_text:
                driver.execute_script("arguments[0].click();", it)
                print(f"   ✅ POD: {clean_pod_name(target_text)}")
                return True
        except:
            continue
    # Fallback: chọn item đầu tiên
    for it in pod_elements:
        try:
            if it.is_displayed():
                driver.execute_script("arguments[0].click();", it)
                print(f"   ✅ POD (fallback): {clean_pod_name(it.text.strip())}")
                return True
        except:
            continue
    return False

def select_pod_by_typing(pod_search, target_text):
    """Dùng cho lần search thứ 2 trở đi trong process_one_route — cần nhập lại"""
    pod_inp = find_visible([
        "//*[@id='podPod2']//input",
        "//*[@id='podPod']//input",
        "//input[@placeholder='Enter Destination']",
    ], timeout=8)
    clear_and_type(pod_inp, pod_search)
    try:
        WebDriverWait(driver, KMTC_DROPDOWN_TIMEOUT).until(
            lambda d: visible_dropdown_items("podPod2") or visible_dropdown_items("podPod")
        )
        items = visible_dropdown_items("podPod2") or visible_dropdown_items("podPod")
        for it in items:
            if it.is_displayed() and it.text.strip() == target_text:
                driver.execute_script("arguments[0].click();", it)
                print(f"   ✅ POD: {clean_pod_name(target_text)}")
                return True
        visible = [it for it in items if it.is_displayed()]
        if visible:
            driver.execute_script("arguments[0].click();", visible[0])
            print(f"   ✅ POD (fallback): {clean_pod_name(visible[0].text.strip())}")
            return True
    except:
        pass
    return False

def ensure_container_type_dry():
    try:
        changed = driver.execute_script("""
            const selects = Array.from(document.querySelectorAll('select'));
            for (const sel of selects) {
                const st = window.getComputedStyle(sel);
                const rect = sel.getBoundingClientRect();
                if (st.display === 'none' || st.visibility === 'hidden' || rect.width === 0 || rect.height === 0) continue;
                const options = Array.from(sel.options || []);
                const dryOpt = options.find(o => /dry/i.test((o.textContent || '').trim()) || String(o.value || '').toUpperCase() === 'GP');
                if (!dryOpt) continue;
                if (sel.value !== dryOpt.value) {
                    sel.value = dryOpt.value;
                    sel.dispatchEvent(new Event('input', {bubbles:true}));
                    sel.dispatchEvent(new Event('change', {bubbles:true}));
                    return (dryOpt.textContent || dryOpt.value || '').trim();
                }
                return 'already';
            }
            return '';
        """)
        if changed and changed != "already":
            print(f"   ✅ Container Type: {changed}")
            rand_sleep(0.2, 0.3)
        return bool(changed)
    except Exception:
        return False

def click_search_btn():
    ensure_container_type_dry()
    try:
        search_btn = driver.find_element(By.XPATH, '//*[@id="frm_main"]/div/div[3]/div/div/a')
    except Exception:
        search_btn = find_visible([
            "//a[contains(@class,'button') and contains(.,'Check Rates')]",
            "//button[contains(.,'Check Rates')]",
        ], timeout=5)
    driver.execute_script("arguments[0].click();", search_btn)
    print(f"   🖱️ Đã Search")

def select_popup_container_dry(max_attempts=4):
    """KMTC alert popup sometimes re-renders the select, making the element stale."""
    last_err = ""
    for attempt in range(1, max_attempts + 1):
        try:
            sel_el = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.ID, "popCntrTypCd"))
            )
            Select(sel_el).select_by_value("GP")
            rand_sleep(0.15, 0.25)
            ok_btn = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.ID, "e-alert-set2-btn1"))
            )
            human_click(ok_btn)
            print(f"   ✅ Đã chọn Dry → OK")
            return True
        except (StaleElementReferenceException, WebDriverException) as e:
            last_err = type(e).__name__
            print(f"   ⚠️ Popup container stale/lag ({last_err}) -> thử lại {attempt}/{max_attempts}")
            rand_sleep(0.2, 0.4)
        except Exception as e:
            last_err = type(e).__name__
            print(f"   ⚠️ Chọn Dry popup lỗi ({last_err}) -> thử lại {attempt}/{max_attempts}")
            rand_sleep(0.2, 0.4)
    print(f"   ❌ Không chọn được Dry trong popup container ({last_err})")
    return False

# ===================================================================================
# --- XỬ LÝ KẾT QUẢ SAU SEARCH ---
# ===================================================================================
def handle_after_search(retried_after_type=False):
    WebDriverWait(driver, 15).until(lambda d:
        d.find_elements(By.CSS_SELECTOR, "div.content_box") or
        d.find_elements(By.CSS_SELECTOR, "div.simplert--shown")
    )
    rand_sleep(0.2, 0.4)

    if driver.find_elements(By.CSS_SELECTOR, "div.simplert--shown"):
        alert_text = ""
        try:
            alert_text = driver.find_element(By.CSS_SELECTOR,
                "div.simplert--shown .simplert__content, #e-alert-message"
            ).text.strip()
            print(f"   ⚠️ Alert: {alert_text[:80]}")
        except:
            pass

        if driver.find_elements(By.ID, "popCntrTypCd"):
            if not select_popup_container_dry():
                dismiss_simplert()
                return False
            try:
                WebDriverWait(driver, 5).until_not(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.simplert--shown"))
                )
            except:
                dismiss_simplert()
            rand_sleep(0.3, 0.5)
            if driver.find_elements(By.CSS_SELECTOR, "div.content_box"):
                return True
            if not retried_after_type:
                print("   🔁 Đã chọn Dry, bấm Search lại...")
                click_search_btn()
                return handle_after_search(retried_after_type=True)
            return False
        elif "registered freight" in alert_text.lower():
            dismiss_simplert()
            try:
                WebDriverWait(driver, KMTC_REGISTERED_FREIGHT_WAIT_SECONDS).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div.content_box"))
                )
                return True
            except:
                return False
        else:
            dismiss_simplert()
            return False

    try:
        WebDriverWait(driver, 3).until_not(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.simplert--shown"))
        )
    except:
        pass

    if driver.find_elements(By.CSS_SELECTOR, "div.content_box"):
        return True
    return False

# ===================================================================================
# --- SCHEDULE (API-based) ---
# ===================================================================================

# Port code mapping: tên port phổ biến → KMTC port code + country code
KMTC_PORT_CODES = {
    # VIETNAM
    "HOCHIMINH":      ("SGN", "VN"),
    "HO CHI MINH":    ("SGN", "VN"),
    "HCM":            ("SGN", "VN"),
    "SGN":            ("SGN", "VN"),
    "HAIPHONG":       ("HPH", "VN"),
    "HAI PHONG":      ("HPH", "VN"),
    "CAI MEP":        ("CMP", "VN"),
    "CMP":            ("CMP", "VN"),
    # INDIA
    "CHENNAI":        ("MAA", "IN"),
    "COCHIN":         ("COK", "IN"),
    "KOLKATA":        ("CCU", "IN"),
    "MUNDRA":         ("MUN", "IN"),
    "NHAVA SHEVA":    ("NSA", "IN"),
    "MUMBAI":         ("BOM", "IN"),
    "CALCUTTA":       ("CCU", "IN"),
    "JNPT":           ("NSA", "IN"),
    "KANDLA":         ("ICD", "IN"),
    "TUTICORIN":      ("TUT", "IN"),
    "VIZAG":          ("VTZ", "IN"),
    "VISAKHAPATNAM":  ("VTZ", "IN"),
    "MANGALORE":      ("ICD", "IN"),
    "PIPAVAV":        ("PIP", "IN"),
    # MIDDLE EAST
    "JEBEL ALI":      ("JEA", "AE"),
    "DAMMAM":         ("DMM", "SA"),
    "JEDDAH":         ("JED", "SA"),
    "SOHAR":          ("SOH", "OM"),
    "MUSCAT":         ("MCT", "OM"),
    # SOUTH EAST ASIA
    "PORT KLANG":     ("PKL", "MY"),
    "PORT KELANG":    ("PKL", "MY"),
    "TANJUNG PELEPAS": ("TPP", "MY"),
    "SINGAPORE":      ("SIN", "SG"),
    "BANGKOK":        ("BKK", "TH"),
    "LAEM CHABANG":   ("LCB", "TH"),
    "JAKARTA":        ("JKT", "ID"),
    "SURABAYA":       ("SUB", "ID"),
    "MANILA":         ("MNL", "PH"),
    # EAST ASIA
    "BUSAN":          ("PUS", "KR"),
    "INCHEON":        ("ICN", "KR"),
    "SHANGHAI":       ("SHA", "CN"),
    "QINGDAO":        ("TAO", "CN"),
    "NINGBO":         ("NBO", "CN"),
    "SHEKOU":         ("SHK", "CN"),
    "TIANJIN":        ("TSN", "CN"),
    "TIENTSIN":       ("TSN", "CN"),
    "XINGANG":        ("XGG", "CN"),
    "XIAMEN":         ("XMN", "CN"),
    "TOKYO":          ("TYO", "JP"),
    "YOKOHAMA":       ("YOK", "JP"),
    "OSAKA":          ("OSA", "JP"),
    "NAGOYA":         ("NGO", "JP"),
}

def get_kmtc_jwt_token():
    """Lấy JWT token từ session browser Selenium.
    FIX: Dò 5 key trong 1 lần execute_script (thay vì 10 round-trip
    sessionStorage + localStorage) — nhanh hơn ~5–10x.
    """
    candidate_keys = ["access_token", "accessToken", "token", "Authorization", "authorization"]
    # 1+2. session/localStorage trong 1 JS call
    try:
        tok = driver.execute_script("""
            const keys = arguments[0];
            for (const k of keys) {
              const v = sessionStorage.getItem(k);
              if (v) return v;
            }
            for (const k of keys) {
              const v = localStorage.getItem(k);
              if (v) return v;
            }
            return null;
        """, candidate_keys)
        if tok:
            return str(tok).replace("Bearer ", "").strip()
    except Exception:
        pass
    # 3. cookies (fallback)
    try:
        for cookie in driver.get_cookies():
            if cookie['name'] in candidate_keys:
                return cookie['value']
    except Exception:
        pass
    return None

def resolve_port_code(port_name):
    """Tra port code từ tên port. Trả về (plcCd, ctrCd) hoặc None"""
    key = port_name.strip().upper()
    if key in KMTC_PORT_CODES:
        return KMTC_PORT_CODES[key]

    # Thử tìm partial match
    for k, v in KMTC_PORT_CODES.items():
        if k in key or key in k:
            return v

    # Thử dùng browser JS để tra cứu qua autocomplete (execute_async_script vì có setTimeout)
    try:
        js = """
        const keyword = arguments[0];
        const callback = arguments[arguments.length - 1];
        const origOpen = XMLHttpRequest.prototype.open;
        const origSend = XMLHttpRequest.prototype.send;
        let found = null;

        XMLHttpRequest.prototype.open = function(method, url, ...args) {
            this._captUrl = url;
            return origOpen.apply(this, [method, url, ...args]);
        };
        XMLHttpRequest.prototype.send = function(body) {
            const self = this;
            const origORL = this.onreadystatechange;
            this.onreadystatechange = function() {
                if (self.readyState === 4 && self._captUrl &&
                    (self._captUrl.includes('port') || self._captUrl.includes('Port'))) {
                    try {
                        const data = JSON.parse(self.responseText);
                        if (Array.isArray(data) && data.length > 0) {
                            found = {plcCd: data[0].plcCd, ctrCd: data[0].ctrCd};
                        }
                    } catch(e) {}
                }
                if (origORL) origORL.apply(this, arguments);
            };
            return origSend.apply(this, arguments);
        };

        const inputs = document.querySelectorAll('input#autocomplete-form-input');
        if (inputs.length >= 2) {
            inputs[1].value = keyword;
            inputs[1].dispatchEvent(new Event('input', {bubbles: true}));
        }

        setTimeout(() => {
            XMLHttpRequest.prototype.open = origOpen;
            XMLHttpRequest.prototype.send = origSend;
            callback(found);
        }, 3000);
        """
        try:
            driver.set_script_timeout(10)
        except Exception:
            pass
        result = driver.execute_async_script(js, port_name.lower())
        if result and isinstance(result, dict) and result.get('plcCd'):
            code = (result['plcCd'], result['ctrCd'])
            KMTC_PORT_CODES[key] = code
            return code
    except Exception:
        pass

    return None

def parse_tt_to_days(raw):
    """'17Day 10Hour 30Min' → 18 (làm tròn lên)"""
    days  = int(re.search(r'(\d+)\s*Day',  raw, re.I).group(1)) if re.search(r'(\d+)\s*Day',  raw, re.I) else 0
    hours = int(re.search(r'(\d+)\s*Hour', raw, re.I).group(1)) if re.search(r'(\d+)\s*Hour', raw, re.I) else 0
    mins  = int(re.search(r'(\d+)\s*Min',  raw, re.I).group(1)) if re.search(r'(\d+)\s*Min',  raw, re.I) else 0
    total_mins = days * 24 * 60 + hours * 60 + mins
    return math.ceil(total_mins / (24 * 60))

def kmtc_api_search_schedule(pol_code, pol_ctr, pod_code, pod_ctr):
    """Gọi API KMTC để lấy lịch tàu. Trả về list dict từ JSON response."""
    token = get_kmtc_jwt_token()
    if not token:
        print(f"   [Sched-API] ⚠️ Không lấy được JWT token")
        return []

    now = datetime.now()
    results = []

    # FIX: tăng range 2→3 tháng để không bỏ lọt schedule xa hơn,
    # đặc biệt khi valid_to_str cũng xa (60-90 ngày).
    for month_offset in range(3):
        target = now + timedelta(days=30 * month_offset)
        year = target.strftime("%Y")
        month = target.strftime("%m")
        # FIX: cho tháng hiện tại (offset 0), startDt = hôm nay thay vì ngày 1
        # — tránh gọi API với range past ññ (KMTC API từ chối)
        if month_offset == 0:
            start_dt = now.strftime("%Y%m%d")
        else:
            start_dt = f"{year}{month}01"
        # Tính ngày cuối tháng chính xác
        last_day = calendar.monthrange(int(year), int(month))[1]
        end_dt = f"{year}{month}{last_day}"

        params = {
            "startPlcCd": pol_code,
            "searchMonth": month,
            "pointChangeYN": "",
            "bound": "O",
            "filterPolCd": "",
            "pointLength": "",
            "startPlcName": "",
            "destPlcCd": pod_code,
            "destPlcName": "",
            "searchYear": year,
            "startCtrCd": pol_ctr,
            "destCtrCd": pod_ctr,
            "calendarOrList": "L",
            "searchOpt": "",
            "filterDirect": "Y",
            "filterTS": "Y",
            "eiCatCd": "O",
            "kmtcSpotYn": "N",
            "polCtrCd": pol_ctr,
            "polPortCd": pol_code,
            "podCtrCd": pod_ctr,
            "podPortCd": pod_code,
            "startDt": start_dt,
            "endDt": end_dt,
            "searchYN": "Y",
            "rfYN": "",
            "filterRteNm": "",
            "filterVslNm": "",
        }

        url = "https://api.ekmtc.com/schedule/schedule/leg/search-schedule?" + urllib.parse.urlencode(params)

        # FIX: thêm retry 1 lần nếu API trả ERROR — KMTC API hay flaky
        last_resp = None
        for retry_i in range(2):
            try:
                # FIX: execute_script không await Promise → dùng execute_async_script
                js = f"""
                const callback = arguments[arguments.length - 1];
                const candidates = ['access_token','accessToken','token','Authorization','authorization'];
                let token = '';
                for (const k of candidates) {{
                    const t = sessionStorage.getItem(k) || localStorage.getItem(k);
                    if (t) {{ token = String(t).replace('Bearer ', '').trim(); break; }}
                }}
                const xhr = new XMLHttpRequest();
                xhr.open('GET', '{url}', true);
                if (token) xhr.setRequestHeader('Authorization', 'Bearer ' + token);
                xhr.timeout = 25000;
                xhr.onload = function() {{ callback(xhr.responseText); }};
                xhr.onerror = function() {{ callback('ERROR:' + xhr.status + ' ' + xhr.statusText); }};
                xhr.ontimeout = function() {{ callback('ERROR:timeout'); }};
                xhr.send();
                """
                try:
                    driver.set_script_timeout(30)
                except Exception:
                    pass
                resp_text = driver.execute_async_script(js)
                last_resp = resp_text

                if resp_text and isinstance(resp_text, str) and not resp_text.startswith("ERROR:"):
                    data = json.loads(resp_text)
                    schedule_list = data.get("listSchedule", [])
                    results.extend(schedule_list)
                    print(f"   [Sched-API] Tháng {month}/{year}: {len(schedule_list)} chuyến")
                    break  # success → không retry
                else:
                    if retry_i == 0:
                        print(f"   [Sched-API] ⚠️ Retry tháng {month}/{year} (lỗi: {str(resp_text)[:100]})")
                        time.sleep(1.0)
                    else:
                        print(f"   [Sched-API] ⚠️ Lỗi API tháng {month}/{year}: {str(resp_text)[:200]}")
            except Exception as e:
                if retry_i == 0:
                    print(f"   [Sched-API] ⚠️ Exception, retry tháng {month}/{year}: {e}")
                    time.sleep(1.0)
                else:
                    print(f"   [Sched-API] ❌ Exception tháng {month}/{year}: {e}")

    return results

def build_schedule_entries_from_api(api_results, pol_code, valid_to_str=""):
    """
    Từ kết quả API, build danh sách entries giống format cũ + thêm vessel details.
    Mỗi entry: {etd_dt, etd_str, tt_days, vessel_name, voyage, route_code, ts_ports, legs}
    """
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    min_etd = today + timedelta(days=DATE_OFFSET_DAYS)
    use_cmp = pol_code.upper() == "CMP"

    valid_dt = None
    if valid_to_str:
        try:
            valid_dt = datetime.strptime(valid_to_str + f"-{datetime.now().year}", "%d-%b-%Y")
            if valid_dt < today:
                valid_dt = valid_dt.replace(year=datetime.now().year + 1)
        except:
            pass

    entries = []
    for item in api_results:
        try:
            etd_raw = item.get("etd", "")
            if not etd_raw or len(etd_raw) < 8:
                continue
            etd_dt = datetime.strptime(etd_raw[:8], "%Y%m%d")

            tt_raw = item.get("transitTime", "")
            if not tt_raw:
                continue
            tt_days = parse_tt_to_days(tt_raw)

            if use_cmp:
                # Lấy ETD của leg đầu tiên từ CMP (leg 2 nếu có rail leg 1)
                for leg_i in range(1, 5):
                    pol_leg = item.get(f"pol{leg_i}", "")
                    vsl_leg = item.get(f"vslNm{leg_i}", "")
                    if pol_leg.upper() == "CMP" or (vsl_leg and vsl_leg.upper() != "RAIL"):
                        cmp_etd_raw = item.get(f"polEtd{leg_i}", "")
                        if cmp_etd_raw and len(cmp_etd_raw) >= 8:
                            cmp_dt = datetime.strptime(cmp_etd_raw[:8], "%Y%m%d")
                            etd_dt = cmp_dt - timedelta(days=2)
                            tt_days = tt_days + 2
                        break

            if etd_dt < min_etd:
                continue
            if not etd_within_max(etd_dt):
                continue
            if valid_dt and etd_dt > valid_dt:
                continue

            # Xác định vessel chính (leg đầu tiên không phải Rail)
            vessel_name = item.get("vslNm", "TBA")
            voyage = item.get("voyNo", "")
            route_code = item.get("rteCd", "")

            # Tìm vessel chính (bỏ qua Rail feeder)
            for leg_i in range(1, 5):
                vsl = item.get(f"vslNm{leg_i}", "")
                if vsl and vsl.upper() != "RAIL":
                    vessel_name = vsl
                    voyage = item.get(f"voyNo{leg_i}", item.get("voyNo", ""))
                    if not route_code:
                        route_code = item.get(f"rteCd{leg_i}", "")
                    break

            # Xác định transshipment ports
            ts_ports = []
            ts_flag = item.get("ts", "N")
            if ts_flag == "Y":
                ts_degree = int(item.get("tsDegree", "1") or "1")
                for leg_i in range(1, ts_degree):
                    pod_leg = item.get(f"pod{leg_i}", "")
                    pod_nm = item.get(f"pod{leg_i}Nm", "")
                    if pod_leg and pod_nm:
                        port_name = pod_nm.split(",")[0].strip()
                        if port_name and port_name.upper() not in ("HOCHIMINH", "HO CHI MINH", "CAI MEP", "CAT LAI"):
                            ts_ports.append(port_name)
            ts_text = " + ".join(ts_ports) if ts_ports else "DIRECT"

            etd_str = f"{etd_dt.day}-{etd_dt.strftime('%b')}"
            entries.append({
                "etd_dt": etd_dt,
                "etd_str": etd_str,
                "tt_days": tt_days,
                "vessel_name": vessel_name,
                "voyage": voyage,
                "route_code": route_code,
                "ts_text": ts_text,
            })
        except Exception as e:
            continue

    # Sắp xếp theo ETD
    entries.sort(key=lambda x: x["etd_dt"])
    return entries

def do_schedule_search(pol_sched, pod_search, tab_sched, valid_to_str=""):
    """
    Tìm lịch tàu qua API KMTC.
    Trả về danh sách entries: [{etd_dt, etd_str, tt_days, vessel_name, voyage, route_code, ts_text}, ...]
    """
    # Resolve port codes
    pol_info = resolve_port_code(pol_sched)
    pod_info = resolve_port_code(pod_search)

    if not pol_info:
        print(f"   [Sched] ⚠️ Không tìm được port code cho POL: {pol_sched}")
        return []
    if not pod_info:
        print(f"   [Sched] ⚠️ Không tìm được port code cho POD: {pod_search}")
        return []

    pol_code, pol_ctr = pol_info
    pod_code, pod_ctr = pod_info

    print(f"   [Sched-API] Tìm kiếm: {pol_sched}({pol_code}/{pol_ctr}) → {pod_search}({pod_code}/{pod_ctr})")

    # Đảm bảo đang ở tab schedule (để JS có thể lấy token)
    activate_tab(tab_sched)
    rand_sleep(0.1, 0.2)

    api_results = kmtc_api_search_schedule(pol_code, pol_ctr, pod_code, pod_ctr)
    if not api_results:
        print(f"   [Sched-API] ⚠️ Không có kết quả từ API")
        return []

    entries = build_schedule_entries_from_api(api_results, pol_code, valid_to_str)
    print(f"   [Sched-API] 📋 Tổng {len(entries)} ETD hợp lệ")
    return entries

def apply_etd_rules_kmtc(schedule_list):
    """Tối đa 3 ETD, cách nhau >=2 ngày, trong vòng 9 ngày. Trả về (str_etd, str_tt, selected)"""
    today  = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    future = sorted(
        [s for s in schedule_list if s["etd_dt"] >= today and etd_within_max(s["etd_dt"])],
        key=lambda x: x["etd_dt"]
    )
    if not future:
        past = [s for s in schedule_list if s["etd_dt"] < today]
        if past:
            print(f"   [Sched] ⚠️ Có {len(past)} ETD nhưng tất cả đã qua ngày hôm nay")
        else:
            print(f"   [Sched] ⚠️ Bị lọc hết bởi valid_to hoặc không có ETD tương lai")
        return "N/A", "N/A", []

    selected   = [future[0]]
    first_date = future[0]["etd_dt"]

    for s in future[1:]:
        if len(selected) >= 3:
            break
        if (s["etd_dt"] - first_date).days > 9:
            break
        if (s["etd_dt"] - selected[-1]["etd_dt"]).days < 2:
            continue
        selected.append(s)

    num = len(selected)
    if num == 0:   str_etd = "N/A"
    elif num == 1: str_etd = selected[0]["etd_str"]
    elif num == 2: str_etd = f"{selected[0]['etd_str']} & {selected[1]['etd_str']}"
    else:
        if all(s["etd_dt"].month == selected[0]["etd_dt"].month for s in selected):
            d1 = str(selected[0]["etd_dt"].day)
            d2 = str(selected[1]["etd_dt"].day)
            d3 = f"{selected[2]['etd_dt'].day}-{selected[2]['etd_dt'].strftime('%b')}"
            str_etd = f"{d1}, {d2}, {d3}"
        else:
            str_etd = " & ".join(s["etd_str"] for s in selected)

    all_tt = [s["tt_days"] for s in selected]
    str_tt = str(min(all_tt)) if min(all_tt) == max(all_tt) else f"{min(all_tt)}-{max(all_tt)}"
    return str_etd, str_tt, selected

# ===================================================================================
# --- PARSE BẢNG GIÁ ---
# ===================================================================================
def parse_freight_popup(is_china_pod=False):
    charges  = {"20'": 0.0, "40'": 0.0, "HC": 0.0}
    formula_parts = {"20'": [], "40'": [], "HC": []}
    valid_to = ""
    free_time = ""
    export_thc_found = False

    popup = driver.find_element(By.CSS_SELECTOR, "div.content_box")

    def extract_combined_free_time():
        try:
            dem_days = []
            det_days = []
            combined_days = []

            def read_discharge_days(cells):
                # KMTC free time table has Loading GP/HC then Discharging GP/HC.
                # Excel only needs Discharging; ignore Loading values completely.
                discharge_cells = cells[-2:] if len(cells) >= 5 else cells[1:]
                values = []
                for cell in discharge_cells:
                    cell_txt = " ".join((cell.text or "").split())
                    if not cell_txt or cell_txt == "-":
                        continue
                    m = re.search(r"\b(\d{1,3})\b", cell_txt)
                    if m:
                        values.append(int(m.group(1)))
                return values

            def fmt_days(values):
                if not values:
                    return ""
                unique_days = sorted(set(values))
                if len(unique_days) == 1:
                    return str(unique_days[0])
                return f"{unique_days[0]}-{unique_days[-1]}"

            for row in popup.find_elements(By.CSS_SELECTOR, "tr"):
                try:
                    if not row.is_displayed():
                        continue
                    cells = row.find_elements(By.CSS_SELECTOR, "th, td")
                    if not cells:
                        continue
                    kind = " ".join((cells[0].text or "").split()).upper()
                    days = read_discharge_days(cells)
                    if not days:
                        continue

                    if kind == "DEM":
                        dem_days.extend(days)
                    elif kind == "DET":
                        det_days.extend(days)
                    elif "COMBIN" in kind:
                        combined_days.extend(days)
                except:
                    continue

            dem_txt = fmt_days(dem_days)
            det_txt = fmt_days(det_days)
            if dem_txt and det_txt:
                return f"{dem_txt} DEM + {det_txt} DET"
            if dem_txt:
                return f"{dem_txt} DEM"
            if det_txt:
                return f"{det_txt} DET"
            if combined_days:
                combined_txt = fmt_days(combined_days)
                return f"{combined_txt} COMBINED"

            texts = []
            for row in popup.find_elements(By.CSS_SELECTOR, "tr, p, div, span"):
                try:
                    if not row.is_displayed():
                        continue
                    txt = " ".join((row.text or "").split())
                    if txt and "COMBIN" in txt.upper():
                        texts.append(txt)
                except:
                    continue

            for txt in texts:
                upper_txt = txt.upper()
                patterns = [
                    r"\b(\d{1,3})\s*(?:DAYS?|DAY)?\s*COMBIN(?:ED|E)?\b",
                    r"\bCOMBIN(?:ED|E)?\b\s*(?:FREE\s*TIME|TIME)?\s*[:\-]?\s*(\d{1,3})\b",
                ]
                for pattern in patterns:
                    m = re.search(pattern, upper_txt)
                    if m:
                        return f"{int(m.group(1))} COMBINED"
        except:
            pass
        return ""

    def format_valid_date(raw_date):
        raw_date = (raw_date or "").strip()
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%d-%b-%Y", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(raw_date, fmt)
                return f"{dt.day}-{dt.strftime('%b')}"
            except:
                pass
        return raw_date

    def extract_valid_to():
        try:
            text_parts = []
            for el in popup.find_elements(By.CSS_SELECTOR, "p, div, span, tr"):
                try:
                    if not el.is_displayed():
                        continue
                    txt = " ".join((el.text or "").split())
                    if txt:
                        text_parts.append(txt)
                except:
                    continue

            full_text = "\n".join(text_parts)
            date_pat = r"\d{4}[-./]\d{1,2}[-./]\d{1,2}"
            patterns = [
                rf"(?:FREIGHT\s*)?VALIDITY\s*PERIOD\s*:?\s*({date_pat})\s*[~\-]\s*({date_pat})",
                rf"(?:VALID\s*TO|EXPIRY\s*DATE|EXPIRE\s*DATE|VALID\s*UNTIL)\s*:?\s*({date_pat})",
            ]
            for pattern in patterns:
                m = re.search(pattern, full_text, re.IGNORECASE)
                if not m:
                    continue
                return format_valid_date(m.group(2) if len(m.groups()) >= 2 else m.group(1))
        except:
            pass
        return ""

    free_time = extract_combined_free_time()
    valid_to = extract_valid_to()

    tables = popup.find_elements(By.CSS_SELECTOR, "table.tbl_col")
    for tbl in tables:
        try:
            header = tbl.find_element(By.CSS_SELECTOR, "thead th").text.strip()
        except:
            continue

        is_freight = "Freight Charges" in header
        is_export  = "Export Charges"  in header
        if not is_freight and not is_export:
            continue

        rows = tbl.find_elements(By.CSS_SELECTOR, "tbody tr, tr")
        for row in rows:
            if "display: none" in (row.get_attribute("style") or ""):
                continue
            tds = row.find_elements(By.CSS_SELECTOR, "td")
            if len(tds) < 7:
                continue
            if "display: none" in (tds[0].get_attribute("style") or ""):
                continue

            spans = tds[0].find_elements(By.CSS_SELECTOR, "span")
            charge_name = spans[0].text.strip() if spans else tds[0].text.strip()
            currency    = tds[1].text.strip()
            cont_type   = tds[2].text.strip()
            cargo       = tds[3].text.strip()

            if cont_type != "GP": continue
            if cargo != "":       continue
            charge_key = " ".join((charge_name or "").split()).upper()
            is_china_origin_thc = bool(is_china_pod and is_export and charge_key == "T.H.C.")
            if currency != "USD" and not is_china_origin_thc:
                continue
            if is_export and charge_key in EXCLUDE_EXPORT:
                continue
            if is_export and charge_key == "T.H.C." and not is_china_pod:
                continue

            def get_val(td):
                divs = td.find_elements(By.CSS_SELECTOR, "div")
                txt  = divs[0].text.strip() if divs else td.text.strip()
                try: return float(txt.replace(",", "").replace(" ", ""))
                except: return 0.0

            v20, v40, v40h = get_val(tds[4]), get_val(tds[5]), get_val(tds[6])
            if is_china_origin_thc and currency != "USD":
                fx = charge_amount_to_usd(1, currency)
                v20, v40, v40h = v20 * fx, v40 * fx, v40h * fx
            if v20 or v40 or v40h:
                if is_export and charge_key == "T.H.C.":
                    export_thc_found = True
                    print(f"   [+] Export O.THC detected, cộng vào cước: 20'={v20} 40'={v40} HC={v40h}")
                elif is_export:
                    print(f"   [+] {header} | {charge_name}: 20'={v20} 40'={v40} HC={v40h}")
                elif is_freight:
                    print(f"   [+] {header} | {charge_name}: 20'={v20} 40'={v40} HC={v40h}")
                charges["20'"] += v20
                charges["40'"] += v40
                charges["HC"]  += v40h
                if v20:
                    formula_parts["20'"].append(v20)
                if v40:
                    formula_parts["40'"].append(v40)
                if v40h:
                    formula_parts["HC"].append(v40h)

    totals = {
        "20'": math.ceil(charges["20'"]) if charges["20'"] > 0 else "-",
        "40'": math.ceil(charges["40'"]) if charges["40'"] > 0 else "-",
        "HC":  math.ceil(charges["HC"])  if charges["HC"]  > 0 else "-",
        "_formula_20": _excel_formula_from_parts(formula_parts["20'"]),
        "_formula_40": _excel_formula_from_parts(formula_parts["40'"]),
        "_formula_hc": _excel_formula_from_parts(formula_parts["HC"]),
    }
    return totals, valid_to, free_time, {"othc_included": bool(is_china_pod), "export_thc_found": export_thc_found}

# ===================================================================================
# --- ĐÓNG POPUP GIÁ ---
# ===================================================================================
def close_popup():
    try:
        close_btn = driver.find_element(By.XPATH,
            "/html/body/div/div[1]/div[2]/body/div/div[1]/div/button")
        driver.execute_script("arguments[0].click();", close_btn)
        rand_sleep(0.3, 0.5)
        return
    except: pass
    try:
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        rand_sleep(0.3, 0.5)
    except: pass

# ===================================================================================
# --- GHI EXCEL ---
# ===================================================================================
def write_row(ws, row_i, pol_excel, pod_display, country, totals, valid_to, free_time="", freight_meta=None):
    if country:
        ws.cell(row=row_i, column=2).value = country
    ws.cell(row=row_i, column=3).value  = pol_excel
    ws.cell(row=row_i, column=4).value  = pod_display
    ws.cell(row=row_i, column=5).value  = "KMTC"
    ws.cell(row=row_i, column=6).value  = totals.get("_formula_20") or totals["20'"]
    ws.cell(row=row_i, column=7).value  = totals.get("_formula_40") or totals["40'"]
    ws.cell(row=row_i, column=8).value  = totals.get("_formula_hc") or totals["HC"]
    ws.cell(row=row_i, column=11).value = valid_to
    ws.cell(row=row_i, column=14).value = free_time
    freight_meta = freight_meta or {}
    othc_included = bool(freight_meta.get("othc_included"))
    remark = build_subject_remark(othc_included=othc_included, country=country, pod=pod_display)
    ws.cell(row=row_i, column=13).value = remark

def write_no_price_row(ws, row_i, pol_excel, pod_display, country):
    write_row(ws, row_i, pol_excel, pod_display, country,
              {"20'": "-", "40'": "-", "HC": "-"}, "", "")
    for col in (9, 10, 11, 13, 14, 15, 16):
        ws.cell(row=row_i, column=col).value = None

def has_real_price(totals):
    return any(str(totals.get(k, "")).strip() not in ("", "-") for k in ("20'", "40'", "HC"))

def write_schedule_to_row(ws, row_i, selected_entries):
    """
    Ghi vessel details (col 15) và transshipment (col 16) theo format CMA.
    Format mỗi dòng: {VESSEL_NAME} ({ROUTE_CODE}) / ETD: {DAY}-{MON} / Transit time: {TT} Days / Transshipment: {TS_PORT}
    """
    if not selected_entries:
        return

    vessel_entries = []
    ts_entries = []

    for entry in selected_entries:
        v_name = entry.get("vessel_name", "TBA")
        v_route = entry.get("route_code", "N/A")
        etd_dt = entry["etd_dt"]
        tt_days = entry["tt_days"]
        ts_text = entry.get("ts_text", "DIRECT")

        vessel_line = (
            f"{v_name} ({v_route}) / ETD: {etd_dt.day}-{etd_dt.strftime('%b')}"
            f" / Transit time: {tt_days} Days / Transshipment: {ts_text}"
        )
        vessel_entries.append(vessel_line)
        ts_entries.append(ts_text)

    # Cột O (15): Vessel details (xuống dòng)
    ws.cell(row=row_i, column=15).value = "\n".join(vessel_entries)
    ws.cell(row=row_i, column=15).alignment = openpyxl.styles.Alignment(wrapText=True)

    # Cột P (16): Transshipment (unique, xuống dòng với " or\n")
    unique_ts = []
    for ts in ts_entries:
        if ts not in unique_ts:
            unique_ts.append(ts)
    ws.cell(row=row_i, column=16).value = " or\n".join(unique_ts)
    ws.cell(row=row_i, column=16).alignment = openpyxl.styles.Alignment(wrapText=True)

# ===================================================================================
# --- XỬ LÝ 1 ROUTE (dùng select_pod_by_typing vì cần nhập lại) ---
# ===================================================================================
def process_one_route(pol_search, pol_excel, pod_search, pod_option_text, row_i, ws, wb, country, pod_display=None):
    pod_clean = clean_pod_name(pod_option_text)
    pod_display = (pod_display or pod_clean).strip().upper()
    route_has_price = False
    try:
        type_pol(pol_search)
        rand_sleep(0.15, 0.25)

        ok = select_pod_by_typing(pod_search, pod_option_text)
        if not ok:
            print(f"   ⚠️ Không chọn được POD: {pod_option_text}")
            write_no_price_row(ws, row_i, pol_excel, pod_display, country)
            try: wb.save(excel_path)
            except PermissionError: print("   ❌ Tắt Excel đi!")
            return False

        click_search_btn()
        has_price = handle_after_search()

        if has_price:
            is_china_pod = "CHINA" in str(country or "").upper() or pod_clean.upper() in CHINA_PODS
            totals, valid_to, free_time, freight_meta = parse_freight_popup(is_china_pod=is_china_pod)
            t20, t40, thc = totals["20'"], totals["40'"], totals["HC"]
            print(f"   💰 20'={t20} 40'={t40} HC={thc} Valid={valid_to} FreeTime={free_time}")
            if not has_real_price(totals):
                write_no_price_row(ws, row_i, pol_excel, pod_display, country)
            else:
                write_row(ws, row_i, pol_excel, pod_display, country, totals, valid_to, free_time, freight_meta)
            close_popup()
            route_has_price = has_real_price(totals)
        else:
            print(f"   ❌ NO SERVICE")
            write_no_price_row(ws, row_i, pol_excel, pod_display, country)

    except Exception as e:
        print(f"   ❌ Lỗi: {e}")
        write_row(ws, row_i, pol_excel, pod_display, country,
                  {"20'": f"LỖI: {e}", "40'": "", "HC": ""}, "")
        dismiss_simplert()
        try: close_popup()
        except: pass

    try:
        wb.save(excel_path)
        print(f"   💾 Saved dòng {row_i}")
    except PermissionError:
        print(f"   ❌ Tắt Excel đi!")
    return route_has_price

# ===================================================================================
# --- MAIN ---
# ===================================================================================
print("""
╔══════════════════════════════════════════════╗
║   KMTC Price Checker  🚢                     ║
╚══════════════════════════════════════════════╝
""")

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

initial_queue = []
single_row_num = None
if SINGLE_ROW:
    try:
        single_row_num = int(SINGLE_ROW)
        print(f"[SINGLE_ROW] Chỉ chạy dòng {single_row_num} theo lệnh từ main.py")
    except ValueError:
        print(f"[WARN] SINGLE_ROW không hợp lệ: {SINGLE_ROW}")
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if single_row_num and i != single_row_num:
        continue
    country   = str(row[1] or "").strip()
    pol_excel = str(row[2] or "").strip().upper()
    pod       = str(row[3] or "").strip()
    carrier   = str(row[4] or "").strip().upper()
    if not pol_excel or not pod: continue
    if carrier not in KMTC_GROUP: continue
    if FILTER_POL and pol_excel != FILTER_POL: continue
    if FILTER_POD and pod.upper() != FILTER_POD: continue
    initial_queue.append((i, pol_excel, pod, country))

total_orig = len(initial_queue)
print(f"📋 Tổng cộng {total_orig} dòng KMTC ban đầu")

SCHED_URL = "https://www.ekmtc.com/index.html#/schedule/leg"

print("[HỆ THỐNG] Kiểm tra và dọn tab KMTC cũ...")
try:
    handles = list(driver.window_handles)
except Exception:
    handles = []

if not handles:
    driver.switch_to.new_window("tab")
    handles = [driver.current_window_handle]

tab_freight = None
tab_sched = None
for h in list(handles):
    try:
        driver.switch_to.window(h)
        url = (driver.current_url or "").lower()
        if "#/schedule" in url and tab_sched is None:
            tab_sched = h
        elif "ekmtc.com" in url and tab_freight is None:
            tab_freight = h
    except Exception:
        continue

if tab_freight is None:
    tab_freight = next((h for h in handles if h != tab_sched), None)
    if tab_freight is None:
        driver.switch_to.new_window("tab")
        tab_freight = driver.current_window_handle
        handles.append(tab_freight)

if tab_sched is None:
    tab_sched = next((h for h in handles if h != tab_freight), None)
    if tab_sched is None:
        activate_tab(tab_freight, "KMTC base")
        driver.switch_to.new_window("tab")
        tab_sched = driver.current_window_handle
        handles.append(tab_sched)

keep_handles = [tab_freight, tab_sched]
extra_handles = [h for h in handles if h not in keep_handles]
if extra_handles:
    print(f"[HỆ THỐNG] Đóng {len(extra_handles)} tab KMTC dư để giảm RAM...")
    close_failed = False
    for h in extra_handles:
        try:
            driver.switch_to.window(h)
            driver.close()
            time.sleep(0.2)
        except Exception as e:
            print(f"   ⚠️ Không đóng được tab dư: {type(e).__name__}")
            close_failed = True
    if close_failed:
        restart_kmtc_edge()
        handles = list(driver.window_handles)
        while len(handles) < 2:
            activate_tab(handles[0], "KMTC base")
            driver.switch_to.new_window("tab")
            handles.append(driver.current_window_handle)
        keep_handles = handles[:2]
        tab_freight, tab_sched = keep_handles

activate_tab(tab_freight, "Freight")
driver.get(BASE_URL)
print(f"   🌐 Tab Freight load KMTC...")
WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))
rand_sleep(0.8, 1.2)

ensure_logged_in()
open_freight_inquiry()

activate_tab(tab_sched, "Schedule")
driver.get(SCHED_URL)
print(f"   🌐 Tab Schedule load KMTC...")
WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))
rand_sleep(0.3, 0.6)

print(f"✅ 2 tabs sẵn sàng (Freight + Schedule)\n")
activate_tab(tab_freight, "Freight")

# ===================================================================================
# MAIN LOOP
# ===================================================================================
row_offset = 0

for orig_row_i, pol_excel, pod_orig, country in initial_queue:
    row_started_at = time.perf_counter()
    actual_row_i = orig_row_i + row_offset

    # Đảm bảo đang ở tab freight trước khi làm gì
    activate_tab(tab_freight)
    ensure_freight_ready()

    pol_search = resolve_alias(pol_excel)
    pod_search = resolve_alias(pod_orig)

    print(f"\n{'='*50}")
    print(f"Dòng {actual_row_i}: {pol_excel} → {pod_orig}" +
          (f" (alias: {pod_search})" if pod_search.upper() != pod_orig.upper() else ""))

    # Nhập POL
    try:
        type_pol(pol_search)
        rand_sleep(0.15, 0.25)
    except Exception as e:
        print(f"   ❌ Lỗi nhập POL: {e}")
        write_row(ws, actual_row_i, pol_excel, pod_orig.upper(), country,
                  {"20'": f"LỖI POL: {e}", "40'": "", "HC": ""}, "")
        try: wb.save(excel_path)
        except: pass
        activate_tab(tab_freight)
        print(f"   ⏱️ Dòng {actual_row_i} xong trong {time.perf_counter() - row_started_at:.1f}s")
        continue

    # Nhập POD và lấy options — giữ dropdown
    pod_texts, pod_elements = type_pod_get_options(pod_search)

    if not pod_texts:
        print(f"   ⚠️ Không tìm thấy POD '{pod_search}' → NO SERVICE")
        write_no_price_row(ws, actual_row_i, pol_excel, pod_orig.upper(), country)
        try: wb.save(excel_path)
        except: pass
        activate_tab(tab_freight)
        print(f"   ⏱️ Dòng {actual_row_i} xong trong {time.perf_counter() - row_started_at:.1f}s")
        continue

    print(f"   📋 {len(pod_texts)} option(s): {[clean_pod_name(o) for o in pod_texts]}")

    route_has_price = False
    if len(pod_texts) == 1:
        # Chọn luôn từ element đang hiện — không nhập lại
        select_pod_by_element(pod_texts[0], pod_elements)
        click_search_btn()
        has_price = handle_after_search()
        if has_price:
            pod_clean_single = clean_pod_name(pod_texts[0])
            is_china_pod = "CHINA" in str(country or "").upper() or pod_clean_single.upper() in CHINA_PODS
            totals, valid_to, free_time, freight_meta = parse_freight_popup(is_china_pod=is_china_pod)
            t20, t40, thc = totals["20'"], totals["40'"], totals["HC"]
            print(f"   💰 20'={t20} 40'={t40} HC={thc} Valid={valid_to} FreeTime={free_time}")
            if not has_real_price(totals):
                write_no_price_row(ws, actual_row_i, pol_excel, pod_orig.upper(), country)
            else:
                write_row(ws, actual_row_i, pol_excel, pod_orig.upper(),
                          country, totals, valid_to, free_time, freight_meta)
            close_popup()
            route_has_price = has_real_price(totals)
        else:
            print(f"   ❌ NO SERVICE")
            write_no_price_row(ws, actual_row_i, pol_excel, pod_orig.upper(), country)
        try: wb.save(excel_path)
        except PermissionError: print("   ❌ Tắt Excel đi!")

    else:
        # FIX: Nhiều options → chọn option chuẩn nhất, KHÔNG thêm dòng mới
        try:
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            rand_sleep(0.2, 0.3)
        except: pass

        # Dùng hàm pick_best_pod_option thay vì luôn lấy pod_texts[0]
        best_opt = pick_best_pod_option(pod_texts, pod_search)
        best_idx = pod_texts.index(best_opt) if best_opt in pod_texts else 0
        first_opt = best_opt
        print(f"\n   --- Option 1/{len(pod_texts)}: {clean_pod_name(first_opt)} → dòng {actual_row_i} ---")
        route_has_price = process_one_route(pol_search, pol_excel, pod_search, first_opt,
                                            actual_row_i, ws, wb, country, pod_orig.upper())

        # Các option còn lại → BỎ QUA (không thêm dòng mới để tránh phình Excel)
        # FIX: tắt tính năng tự tạo thêm dòng theo yêu cầu
        if len(pod_texts) > 1:
            print(f"   ℹ️ Có {len(pod_texts)} option POD, chỉ lấy option đầu tiên (đã tắt auto-thêm dòng).")

    if not route_has_price:
        print(f"   [Sched] Bỏ qua check lịch vì tuyến này không có giá.")
        activate_tab(tab_freight)
        print(f"   ⏱️ Dòng {actual_row_i} xong trong {time.perf_counter() - row_started_at:.1f}s")
        continue

    # --- Check Schedule cho dòng hiện tại ---
    valid_to_cell = ws.cell(row=actual_row_i, column=11).value or ""
    pod_upper     = pod_orig.upper()
    pol_sched     = POL_SCHEDULE_MAP.get(pol_excel, pol_excel.lower())
    if pol_excel == "HO CHI MINH" and pod_upper in CMP_PODS:
        pol_sched = "CMP"

    print(f"\n   [Sched] 🗓️ Check lịch: {pol_sched} → {pod_search}")
    try:
        sched_list = do_schedule_search(pol_sched, pod_search, tab_sched, str(valid_to_cell))
        if sched_list:
            str_etd, str_tt, selected = apply_etd_rules_kmtc(sched_list)
            print(f"   [Sched] 🏆 ETD: {str_etd} | T/T: {str_tt}")
            ws.cell(row=actual_row_i, column=9).value  = str_etd
            ws.cell(row=actual_row_i, column=10).value = str_tt
            # Ghi vessel details + transshipment theo format CMA
            if selected:
                write_schedule_to_row(ws, actual_row_i, selected)
            try: wb.save(excel_path)
            except PermissionError: print("   ❌ Tắt Excel đi!")
        else:
            print(f"   [Sched] ⚠️ Không có lịch tàu")
    except Exception as e:
        print(f"   [Sched] ❌ Lỗi: {e}")

    # Switch về tab freight cho dòng tiếp theo
    activate_tab(tab_freight)
    print(f"   ⏱️ Dòng {actual_row_i} xong trong {time.perf_counter() - row_started_at:.1f}s")

print(f"\n✅ Hoàn tất! ({total_orig} dòng ban đầu → {total_orig + row_offset} dòng sau expand)")
close_all_kmtc_tabs()
