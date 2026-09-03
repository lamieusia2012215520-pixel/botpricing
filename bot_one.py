import re
import calendar
from datetime import datetime, timedelta
import pandas as pd
import openpyxl  # <-- Đã thêm thư viện openpyxl
import math
import time
import json
import uuid
from urllib.parse import urlencode
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchWindowException
from selenium.webdriver.common.action_chains import ActionChains
import subprocess
import os
from pathlib import Path
import requests
import warnings
from selenium.webdriver.remote.remote_connection import RemoteConnection
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only
from remark_rules import apply_manifest_rule, get_manifest_code, is_china_destination

DATE_OFFSET_DAYS = parse_date_offset_days()
DRIVER_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "msedgedriver.exe")

def _excel_formula_from_parts(parts):
    tokens = []
    for raw in parts:
        try:
            amount = float(raw)
        except (TypeError, ValueError):
            continue
        if abs(amount) < 1e-9:
            continue
        sign = "-" if amount < 0 else "+"
        amount = abs(amount)
        if abs(amount - round(amount)) < 1e-9:
            text = str(int(round(amount)))
        else:
            text = f"{amount:.2f}".rstrip("0").rstrip(".")
        tokens.append((sign if tokens else ("-" if sign == "-" else "")) + text)
    return "=" + "".join(tokens) if tokens else None

EXCHANGE_RATE_CACHE = {}

def get_live_exchange_rate(base_currency, target_currency="USD"):
    base = (base_currency or "").upper().strip()
    target = (target_currency or "USD").upper().strip()
    if not base or base == target:
        return 1.0
    cache_key = (base, target)
    if cache_key in EXCHANGE_RATE_CACHE:
        return EXCHANGE_RATE_CACHE[cache_key]
    env_key = f"ONE_{base}_TO_{target}"
    try:
        rate = float(os.environ.get(env_key, "").strip())
        EXCHANGE_RATE_CACHE[cache_key] = rate
        return rate
    except Exception:
        pass
    try:
        res = requests.get(f"https://api.frankfurter.app/latest?from={base}&to={target}", timeout=3)
        rate = float(res.json()["rates"][target])
        EXCHANGE_RATE_CACHE[cache_key] = rate
        return rate
    except Exception as e:
        fallbacks = {
            ("EUR", "USD"): 1.08,
            ("CHF", "USD"): 1.12,
            ("VND", "USD"): 0.00004,
        }
        rate = fallbacks.get(cache_key, 1.0)
        EXCHANGE_RATE_CACHE[cache_key] = rate
        print(f"[WARN] FX {base}->{target} failed ({e}); fallback={rate}")
        return rate

# ── Timestamp print ──
import sys, io
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
_orig_print = print
def print(*args, **kwargs):
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    _orig_print(f"[{ts}]", *args, **kwargs)

# ── Flag bật/tắt timing chi tiết ──
ENABLE_TIMING = True  # Đổi thành False để tắt khi không cần nữa

# ==========================================
# TỰ ĐỘNG MỞ EDGE 9522 (nếu chưa mở)
# ==========================================
import socket

ONE_DEBUG_PORT = 9522
ONE_EDGE_PROFILE = r"C:\edge_one"
ONE_EDGE_START_ATTEMPTS = 2
ONE_EDGE_ATTACH_ATTEMPTS = 3
ONE_EDGE_READY_TIMEOUT = 30

try:
    # Do not let one bad debugger session block the entire bot for 120 seconds.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", DeprecationWarning)
        RemoteConnection.set_timeout(int(os.environ.get("ONE_SELENIUM_CONNECT_TIMEOUT", "35")))
except Exception:
    pass

def is_port_in_use(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('127.0.0.1', port)) == 0

def _wait_port(port, timeout=8):
    """FIX: Poll port thay vì sleep cứng → quay lại ngay khi Edge sẵn sàng."""
    end = time.time() + timeout
    while time.time() < end:
        if is_port_in_use(port):
            return True
        time.sleep(0.2)
    return False

def _legacy_open_edge_9522():
    """Mở Edge với remote debugging port 9522 nếu chưa có"""
    if is_port_in_use(9522):
        print("[HỆ THỐNG] Edge ONE đã mở sẵn (port 9522). Bỏ qua lệnh khởi động trình duyệt.")
        return
    print("[HỆ THỐNG] Edge ONE chưa mở. Đang tự động khởi động...")
    edge_path = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    args = [
        edge_path,
        "--remote-debugging-port=9522",
        r"--user-data-dir=C:\edge_one",
        "--disable-background-timer-throttling",
        "--disable-renderer-backgrounding",
        "--disable-backgrounding-occluded-windows"
    ]
    try:
        subprocess.Popen(args)
        print("[OK] Đã mở Edge 9522 thành công!")
        _wait_port(9522, timeout=8)
    except FileNotFoundError:
        print("[ERROR] Không tìm thấy Edge tại đường dẫn trên. Kiểm tra lại đường dẫn!")

def edge_debug_ready(timeout=2):
    """Require a healthy CDP endpoint, not merely a listening TCP port."""
    try:
        response = requests.get(
            f"http://127.0.0.1:{ONE_DEBUG_PORT}/json/version",
            timeout=timeout,
        )
        if response.status_code != 200:
            return False
        return bool(response.json().get("webSocketDebuggerUrl"))
    except Exception:
        return False


def _wait_edge_debug(timeout=ONE_EDGE_READY_TIMEOUT):
    end = time.time() + timeout
    while time.time() < end:
        if edge_debug_ready(timeout=1.5):
            return True
        time.sleep(0.4)
    return False


def stop_edge_one(reason=""):
    """Stop only Edge processes belonging to ONE's dedicated profile/port."""
    if reason:
        print(f"[HỆ THỐNG] Restart riêng Edge ONE: {reason}")
    profile = ONE_EDGE_PROFILE.replace("'", "''")
    script = (
        f"$profile='{profile}'; $port='remote-debugging-port={ONE_DEBUG_PORT}'; "
        "Get-CimInstance Win32_Process -Filter \"name='msedge.exe'\" | "
        "Where-Object { $_.CommandLine -like ('*' + $profile + '*') -or "
        "$_.CommandLine -like ('*' + $port + '*') } | "
        "ForEach-Object { Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }"
    )
    try:
        kwargs = {
            "stdout": subprocess.DEVNULL,
            "stderr": subprocess.DEVNULL,
            "timeout": 12,
        }
        if os.name == "nt":
            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            **kwargs,
        )
    except Exception as exc:
        print(f"[HỆ THỐNG] Không dừng được Edge ONE cũ: {type(exc).__name__}")

    end = time.time() + 8
    while time.time() < end and is_port_in_use(ONE_DEBUG_PORT):
        time.sleep(0.25)


def open_edge_9522(force_restart=False):
    """Open ONE Edge and return only after /json/version is ready."""
    if not force_restart and edge_debug_ready(timeout=2):
        print("[HỆ THỐNG] Edge ONE debug endpoint đã sẵn sàng; tái sử dụng phiên hiện có.")
        return

    if is_port_in_use(ONE_DEBUG_PORT) or force_restart:
        stop_edge_one("debug endpoint không phản hồi" if not force_restart else "retry attach")

    edge_path = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    last_error = None
    for attempt in range(1, ONE_EDGE_START_ATTEMPTS + 1):
        print(f"[HỆ THỐNG] Đang mở Edge ONE port {ONE_DEBUG_PORT} ({attempt}/{ONE_EDGE_START_ATTEMPTS})...")
        args = [
            edge_path,
            f"--remote-debugging-port={ONE_DEBUG_PORT}",
            f"--user-data-dir={ONE_EDGE_PROFILE}",
            "--remote-allow-origins=*",
            "--disable-background-timer-throttling",
            "--disable-renderer-backgrounding",
            "--disable-backgrounding-occluded-windows",
        ]
        try:
            subprocess.Popen(args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            if _wait_edge_debug():
                print(f"[OK] Edge ONE {ONE_DEBUG_PORT} đã sẵn sàng (CDP OK).")
                return
            last_error = TimeoutError("CDP /json/version chưa sẵn sàng")
        except FileNotFoundError as exc:
            raise RuntimeError(f"Không tìm thấy Edge: {edge_path}") from exc
        except Exception as exc:
            last_error = exc

        stop_edge_one(f"khởi động lần {attempt} chưa sẵn sàng")

    raise RuntimeError(f"Không mở được Edge ONE sau {ONE_EDGE_START_ATTEMPTS} lần: {last_error}")


open_edge_9522()

# ==========================================
# CẤU HÌNH TRÌNH DUYỆT & HẰNG SỐ
# ==========================================
edge_options = Options()
edge_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{ONE_DEBUG_PORT}")


def connect_edge_one():
    last_error = None
    for attempt in range(1, ONE_EDGE_ATTACH_ATTEMPTS + 1):
        try:
            if attempt > 1:
                open_edge_9522(force_restart=True)
            print(f"[HỆ THỐNG] Attach Selenium vào Edge ONE ({attempt}/{ONE_EDGE_ATTACH_ATTEMPTS})...")
            attached = webdriver.Edge(service=Service(executable_path=DRIVER_PATH), options=edge_options)
            _ = attached.current_url
            print("[OK] Selenium đã attach Edge ONE thành công.")
            return attached
        except Exception as exc:
            last_error = exc
            print(f"[HỆ THỐNG] Attach Edge ONE lỗi: {type(exc).__name__}: {exc}")
            stop_edge_one(f"attach thất bại lần {attempt}")

    raise RuntimeError(
        f"Không attach được Selenium vào Edge ONE sau {ONE_EDGE_ATTACH_ATTEMPTS} lần: {last_error}"
    )


driver = connect_edge_one()
try:
    driver.maximize_window()
except Exception:
    try:
        info = driver.execute_cdp_cmd("Browser.getWindowForTarget", {})
        driver.execute_cdp_cmd("Browser.setWindowBounds", {
            "windowId": info["windowId"],
            "bounds": {"windowState": "maximized"}
        })
    except Exception:
        pass
wait   = WebDriverWait(driver, 15)

SPEED      = 0.1
ONE_URL    = "https://ecomm.one-line.com/one-ecom/prices/one-quote-booking"
FORM_XPATH = "(//input[@placeholder='Please search location'])[1]"

ONE_PORT_SEARCH_MAX_TRIMS = int(os.environ.get("ONE_PORT_SEARCH_MAX_TRIMS", "3"))
ONE_PORT_SEARCH_MIN_LEN = int(os.environ.get("ONE_PORT_SEARCH_MIN_LEN", "3"))

PORT_ALIASES = {
    "BUSAN":       "PUSAN",
    "HO CHI MINH": "HO CHI MINH",
    "ANTWERP":     "ANTWERP"
}
# Danh sách POD theo khu vực để xác định remark phụ
def one_port_search_terms(port_name, allow_trim=True):
    """
    Bình thường chỉ trả tên gốc/alias.
    Chỉ khi ONE đã báo No port pair mới bật allow_trim=True để thử rút ngắn:
    SYDNEY -> SYDNE -> SYDN...
    """
    original = str(port_name or "").strip()
    aliased = PORT_ALIASES.get(original.upper(), original)
    terms = []
    seen = set()
    for base in (aliased, original):
        base = str(base or "").strip()
        if not base:
            continue
        max_trim = min(ONE_PORT_SEARCH_MAX_TRIMS, max(0, len(base) - ONE_PORT_SEARCH_MIN_LEN)) if allow_trim else 0
        for trim in range(0, max_trim + 1):
            term = base[:-trim] if trim else base
            term = term.strip()
            key = term.upper()
            if term and key not in seen:
                terms.append(term)
                seen.add(key)
    return terms

CHINA_PORTS    = ["CHINA", "CN", "SHANGHAI", "NINGBO", "GUANGZHOU", "SHENZHEN",
                  "TIANJIN", "QINGDAO", "XIAMEN", "DALIAN", "BEIJING"]
JAPAN_PORTS    = ["JAPAN", "JP", "TOKYO", "OSAKA", "NAGOYA", "YOKOHAMA",
                  "KOBE", "HAKATA", "MOJI"]
EUROPE_PORTS   = ["GERMANY", "DE", "FRANCE", "FR", "NETHERLANDS", "NL",
                  "BELGIUM", "BE", "SPAIN", "ES", "ITALY", "IT", "POLAND", "PL",
                  "SWEDEN", "SE", "DENMARK", "DK", "FINLAND", "FI", "NORWAY", "NO",
                  "PORTUGAL", "PT", "GREECE", "GR", "TURKEY", "TR", "UK", "GB",
                  "ROTTERDAM", "HAMBURG", "ANTWERP", "BARCELONA", "GDANSK",
                  "FELIXSTOWE", "SOUTHAMPTON", "LE HAVRE", "MARSEILLE",
                  "GENOA", "VALENCIA", "BREMERHAVEN", "PIRAEUS"]

def parse_one_free_time_text(raw_text):
    text = " ".join((raw_text or "").upper().replace("\xa0", " ").split())
    if not text:
        return ""

    def pick_day(labels):
        for label in labels:
            patterns = [
                rf"{label}[^0-9]{{0,50}}(\d{{1,3}})",
                rf"(\d{{1,3}})[^A-Z0-9]{{0,20}}{label}",
            ]
            for pattern in patterns:
                m = re.search(pattern, text)
                if m:
                    return m.group(1)
        return ""

    combined = pick_day([r"COMBINED", r"D\s*&\s*D", r"DND"])
    if combined:
        return f"{combined} COMBINED"

    dem = pick_day([r"DEMURRAGE", r"\bDEM\b"])
    det = pick_day([r"DETENTION", r"\bDET\b"])
    if dem and det:
        return f"{dem} DEM + {det} DET"
    if dem:
        return f"{dem} DEM"
    if det:
        return f"{det} DET"
    return ""

# ==========================================
# CÁC HÀM HỖ TRỢ CƠ BẢN
# ==========================================

SPEED = 0.01  # Bạn có thể đổi thành 0.1, 0.2 tùy ý
ERROR_DIR = Path(os.environ.get("ERROR_DIR", os.path.join(os.getcwd(), "errors")))
ONE_TAB_PROGRESS = {}
ONE_COMMODITY_STEP_WAIT = float(os.environ.get("ONE_COMMODITY_STEP_WAIT", "2.5"))
ONE_DATE_INPUT_STEP_WAIT = float(os.environ.get("ONE_DATE_INPUT_STEP_WAIT", "4"))
ONE_DATE_RENDER_STEP_WAIT = float(os.environ.get("ONE_DATE_RENDER_STEP_WAIT", "6"))
ONE_PIPELINE_MAX_RETRIES = int(os.environ.get("ONE_PIPELINE_MAX_RETRIES", "12"))
ONE_API_TIMEOUT_DEFAULT = float(os.environ.get("ONE_API_TIMEOUT_DEFAULT", "35"))
ONE_FREE_TIME_API_TIMEOUT = float(os.environ.get("ONE_FREE_TIME_API_TIMEOUT", "12"))

def wait_and_speed(xpath, timeout=10):
    """Đợi phần tử sẵn sàng bằng WebDriverWait + nghỉ thêm SPEED"""
    element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    time.sleep(SPEED) # Nghỉ theo ý bạn sau khi web đã load xong
    return element

def capture_error_artifacts(prefix):
    try:
        ERROR_DIR.mkdir(parents=True, exist_ok=True)
        safe_prefix = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(prefix))[:120]
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = ERROR_DIR / f"ONE_{safe_prefix}_{stamp}"
        driver.save_screenshot(str(base) + ".png")
        with open(str(base) + ".html", "w", encoding="utf-8") as f:
            f.write(driver.page_source or "")
        print(f"  📸 Đã lưu lỗi ONE: {base}.png/.html")
    except Exception as e:
        print(f"  ⚠️ Không lưu được screenshot/html lỗi ONE: {e}")

def retry_action(label, fn, retries=3, wait_seconds=1.5):
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            return fn()
        except Exception as e:
            last_err = e
            print(f"  ⚠️ {label} lỗi lần {attempt}/{retries}: {type(e).__name__}")
            if attempt < retries:
                time.sleep(wait_seconds)
    raise last_err

def click_icon_add(timeout=10):
    """
    Đợi nút icon-add xuất hiện rồi click. Tránh IndexError khi React
    chưa render xong sau add_equipment.
    """
    last_err = None
    for try_i in range(3):
        try:
            WebDriverWait(driver, timeout).until(
                EC.presence_of_all_elements_located(
                    (By.XPATH, "//img[@alt='icon-add']")
                )
            )
            icons = driver.find_elements(By.XPATH, "//img[@alt='icon-add']")
            if icons:
                # Scroll đến nút cuối cùng để chắc chắn nó visible
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", icons[-1]
                )
                time.sleep(0.2)
                js_click(icons[-1])
                return True
        except Exception as e:
            last_err = e
        # Chờ React render thêm
        time.sleep(0.6)
    raise Exception(f"Không tìm thấy nút icon-add sau {3*timeout}s (last err: {last_err})")

def ensure_equipment_slots(target_count, row_label=""):
    xpath = "//input[@placeholder='Select an Equipment Type']"

    def count_slots():
        return len(driver.find_elements(By.XPATH, xpath))

    for _ in range(3):
        current = count_slots()
        if current >= target_count:
            return True
        print(f"  [INFO] {row_label} cần {target_count} ô container, hiện có {current}. Bấm Add...")
        retry_action("bấm Add container", lambda: click_icon_add(timeout=8), retries=2, wait_seconds=1)
        try:
            WebDriverWait(driver, 10).until(lambda d: len(d.find_elements(By.XPATH, xpath)) >= target_count)
            return True
        except Exception:
            pass

    capture_error_artifacts(f"{row_label}_equipment_slots_{target_count}")
    raise Exception(f"WEB_LAG_RETRY: Không tạo đủ {target_count} ô container")

def get_valid_date(etd_dates):
    latest_etd = max(etd_dates)
    day = latest_etd.day
    if day <= 7:    valid_day = 7
    elif day <= 14: valid_day = 14
    elif day <= 21: valid_day = 21
    else:           valid_day = calendar.monthrange(latest_etd.year, latest_etd.month)[1]
    vd = datetime(latest_etd.year, latest_etd.month, valid_day)
    return f"{vd.day}-{vd.strftime('%b')}"

def js_click(element):
    driver.execute_script("arguments[0].click();", element)

def set_input_value(element, value):
    driver.execute_script("""
        const el = arguments[0];
        const value = arguments[1];
        el.scrollIntoView({block: 'center'});
        el.focus();
        const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        setter.call(el, value);
        el.dispatchEvent(new Event('input', {bubbles: true}));
        el.dispatchEvent(new Event('change', {bubbles: true}));
    """, element, str(value))

def select_port(input_index, port_name, country_name, allow_trim=False):
    actual_port = PORT_ALIASES.get(str(port_name).upper(), str(port_name))
    original_actual_port = actual_port
    if allow_trim:
        trim_terms = one_port_search_terms(original_actual_port, allow_trim=True)
        if len(trim_terms) > 1:
            actual_port = trim_terms[1]

    # Lọc country_str
    country_raw = str(country_name).strip()
    if (country_raw.upper() == "NAN"
            or not country_raw
            or re.match(r'\d{4}-\d{2}-\d{2}', country_raw)
            or re.match(r'\d{2}/\d{2}/\d{4}', country_raw)):
        country_str = ""
    else:
        country_str = country_raw.upper()

    port_input_xpath = f"(//input[@placeholder='Please search location'])[{input_index}]"
    port_input = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, port_input_xpath))
    )

    # Đọc giá trị hiện tại của ô input
    current_val = port_input.get_attribute("value") or ""
    if allow_trim:
        current_val = ""
    if actual_port.upper() in current_val.upper():
        print(f"  ✅ Cảng {actual_port} đã có sẵn, bỏ qua bước nhập.")
        return # Thoát hàm luôn, không cần tìm kiếm

    print(f"  👉 Nhập cảng: {actual_port} ({country_str if country_str else 'no country filter'})")

    driver.execute_script("""
        const el = arguments[0];
        el.scrollIntoView({block: 'center'});
        window.scrollBy(0, -120);
        el.focus();
        el.click();
    """, port_input)
    driver.execute_script("""
        var el = arguments[0];
        var nativeSetter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value').set;
        nativeSetter.call(el, '');
        el.dispatchEvent(new Event('input', {bubbles:true}));
    """, port_input)
    port_input.send_keys(actual_port)


    # Đợi dropdown xuất hiện
    try:
        WebDriverWait(driver, 4).until(
            EC.presence_of_element_located((By.XPATH, "//li[@role='option']"))
        )
    except:
        time.sleep(0.5)
    all_options = driver.find_elements(By.XPATH, "//li[@role='option']")
    # ── RETRY: nếu không có option khớp tên cảng → xóa 1 ký tự cuối rồi thử lại ──
    def has_matching_option(opts, name):
        return any(name.upper() in opt.text.strip().upper() for opt in opts)

    if allow_trim and not has_matching_option(all_options, actual_port):
        print(f"  🔄 Không thấy option '{actual_port}', thử xóa 1 ký tự cuối...")
        trimmed = actual_port[:-1]
        port_input.send_keys(Keys.BACKSPACE)
        try:
            WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.XPATH, "//li[@role='option']"))
            )
        except:
            time.sleep(0.5)
        all_options = driver.find_elements(By.XPATH, "//li[@role='option']")
        if not has_matching_option(all_options, trimmed):
            print(f"  ❌ Vẫn không thấy option sau khi xóa ký tự → bỏ qua")
        else:
            print(f"  ✅ Tìm thấy option sau khi trim thành '{trimmed}' (dùng làm key tìm kiếm tiếp)")
            # FIX: dùng trimmed làm key cho VÒNG 1/2/3 phía dưới, nếu không các vòng đó
            # vẫn tìm theo 'PARADIP'/'VISAKHAPATNAM' nguyên gốc và không match được.
            actual_port = trimmed

    # ── HÀM NỘI BỘ: Đọc badge CY/DOOR trực tiếp từ element con ──
    if allow_trim and not has_matching_option(all_options, actual_port):
        print(f"  🔁 ONE retry autocomplete mở rộng cho '{original_actual_port}'...")
        retry_terms = one_port_search_terms(original_actual_port, allow_trim=True)
        if allow_trim and len(retry_terms) > 1:
            retry_terms = retry_terms[1:]
        for term in retry_terms:
            if term.upper() == actual_port.upper():
                continue
            try:
                driver.execute_script("""
                    const el = arguments[0];
                    el.scrollIntoView({block: 'center'});
                    window.scrollBy(0, -120);
                    el.focus();
                    el.click();
                    const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                    setter.call(el, '');
                    el.dispatchEvent(new Event('input', {bubbles:true}));
                    el.dispatchEvent(new Event('change', {bubbles:true}));
                """, port_input)
                port_input.send_keys(term)
                try:
                    WebDriverWait(driver, 4).until(
                        EC.presence_of_element_located((By.XPATH, "//li[@role='option']"))
                    )
                except Exception:
                    time.sleep(0.5)
                all_options = driver.find_elements(By.XPATH, "//li[@role='option']")
                if has_matching_option(all_options, term):
                    print(f"  ✅ ONE autocomplete dùng query '{term}' cho cảng '{original_actual_port}'")
                    actual_port = term
                    break
                print(f"  🔄 Query '{term}' vẫn chưa ra option hợp lệ")
            except Exception as e:
                print(f"  ⚠️ Retry query '{term}' lỗi: {type(e).__name__}")

    def get_badge_text(opt):
        """
        Đọc text của badge (CY / DOOR / CFS...) bên trong option.
        Badge là element con nhỏ nhất, không phải toàn bộ text của li.
        """
        # Tìm tất cả element con có background/color khác biệt (badge)
        candidates = opt.find_elements(By.XPATH,
            ".//*[string-length(normalize-space(text())) <= 6 "   # badge ngắn: CY, DOOR, CFS
            "and normalize-space(text()) != '']"
        )
        for el in candidates:
            t = el.text.strip().upper()
            if t in ("CY", "DOOR", "CFS", "RAMP"):
                return t
        return ""

    def get_port_name_text(opt):
        """Lấy tên cảng từ option, bỏ badge ra"""
        full = opt.text.strip().upper()
        for badge in ("CY", "DOOR", "CFS", "RAMP"):
            full = full.replace(badge, "").strip()
        return full

    selected = False
    # ── VÒNG 1: khớp tên cảng (+ country nếu có) + badge = CY ──
    for opt in all_options:
        try:
            port_text = get_port_name_text(opt)
            if actual_port.upper() not in port_text:
                continue
            if country_str and country_str not in port_text:
                continue
            badge = get_badge_text(opt)
            if badge == "CY":
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", opt)
                js_click(opt)
                selected = True
                print(f"  ✅ CY (với country): {opt.text.strip()}")
                break
        except:
            continue

    # ── VÒNG 2: bỏ lọc country, vẫn cần badge = CY ──
    if not selected:
        for opt in all_options:
            try:
                port_text = get_port_name_text(opt)
                if actual_port.upper() not in port_text:
                    continue
                badge = get_badge_text(opt)
                if badge == "CY":
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", opt)
                    js_click(opt)
                    selected = True
                    print(f"  ✅ CY (no country filter): {opt.text.strip()}")
                    break
            except:
                continue

    # ── VÒNG 3: fallback - chọn đầu tiên có tên cảng (dù badge gì) ──
    if not selected:
        for opt in all_options:
            try:
                if actual_port.upper() in opt.text.strip().upper():
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", opt)
                    js_click(opt)
                    selected = True
                    print(f"  ⚠️ Không có CY, chọn đầu tiên: {opt.text.strip()}")
                    break
            except:
                continue

    if not selected:
        print(f"  ❌ Không tìm thấy option nào cho {actual_port}. Dùng phím mũi tên...")
        capture_error_artifacts(f"select_port_{actual_port}")
        raise Exception(f"Không chọn được cảng {actual_port}")

    time.sleep(0.2)

def wait_for_loading_popup():
    """Học từ Bot April: Đợi vòng xoay biến mất hoàn toàn"""
    loader_xpath = "//div[contains(@class, 'CarouselLoadingPopup_progress-container') or contains(@class, 'ajax-progress')]"
    try:
        # Đợi nó xuất hiện (trong tối đa 1.5s)
        WebDriverWait(driver, 1.5).until(EC.presence_of_element_located((By.XPATH, loader_xpath)))
        # Đợi nó biến mất (trong tối đa 30s)
        WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.XPATH, loader_xpath)))
        time.sleep(SPEED) # Web xong rồi, nghỉ SPEED theo ý bạn
    except TimeoutException:
        pass # Nếu không có popup thì thôi chạy tiếp

def add_equipment(index, equip_type, weight, qty_clicks=0):
    equip_input_xpath = f"(//input[@placeholder='Select an Equipment Type'])[{index}]"
    equip_input = wait.until(EC.presence_of_element_located((By.XPATH, equip_input_xpath)))
    
    # Check Loại Container
    current_type = equip_input.get_attribute("value") or ""
    if current_type.strip() != equip_type:
        print(f"  📦 Container {index}: Đang chọn {equip_type}...")
        driver.execute_script("arguments[0].scrollIntoView({behavior:'smooth',block:'center'});", equip_input)
        time.sleep(SPEED)
        js_click(equip_input)
        time.sleep(0.3)
        equip_options = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//li[@role='option']")))
        for opt in equip_options:
            if opt.text.strip() == equip_type:
                js_click(opt)
                break
        time.sleep(SPEED)
    
    # Check Số lượng (Qty)
    inc_btn_xpath = f"(//button[@data-action='increment'])[{index}]"
    inc_btn = wait.until(EC.presence_of_element_located((By.XPATH, inc_btn_xpath)))
    qty_input = inc_btn.find_element(By.XPATH, "./parent::div//input")
    if qty_input.get_attribute("value") != "1":
        try:
            js_click(qty_input)
            time.sleep(SPEED)
            qty_input.send_keys(Keys.CONTROL + "a")
            qty_input.send_keys(Keys.BACKSPACE)
            qty_input.send_keys("1")
        except Exception:
            set_input_value(qty_input, "1")
        time.sleep(SPEED)

    # Check Khối lượng (Weight)
    weight_input_xpath = f"(//input[@placeholder='0' and @inputmode='numeric'])[{index}]"
    weight_input = wait.until(EC.presence_of_element_located((By.XPATH, weight_input_xpath)))
    current_weight = weight_input.get_attribute("value") or ""
    if current_weight.replace(",", "") != weight:
        print(f"  ⚖️ Cập nhật khối lượng thành {weight}kg")
        try:
            js_click(weight_input)
            time.sleep(SPEED)
            weight_input.send_keys(Keys.CONTROL + "a")
            weight_input.send_keys(Keys.BACKSPACE)
            weight_input.send_keys(weight)
        except Exception:
            set_input_value(weight_input, weight)
        time.sleep(SPEED)
    else:
        print(f"  ✅ Container {index} ({equip_type} - {weight}kg) đã cấu hình chuẩn.")

# ==========================================
# HÀM ĐĂNG NHẬP (dùng lại ở nhiều nơi)
# ==========================================
def do_login():
    """Đăng nhập nhanh bằng JavaScript inject giá trị trực tiếp"""
    print("  🔐 Đang đăng nhập (nhanh)...")
    
    user_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//input[@type='text' or @name='username' or @id='username']"))
    )
    pass_input = WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, "//input[@type='password']"))
    )

    # Nhập username bằng JS (nhanh nhất)
    driver.execute_script("arguments[0].value = arguments[1];", user_input, "PIOLOG")
    driver.execute_script(
        "arguments[0].dispatchEvent(new Event('input', {bubbles:true})); "
        "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));",
        user_input
    )

    # Nhập password bằng JS
    driver.execute_script("arguments[0].value = arguments[1];", pass_input, "Vankiep@21")
    driver.execute_script(
        "arguments[0].dispatchEvent(new Event('input', {bubbles:true})); "
        "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));",
        pass_input
    )

    time.sleep(0.3)  # Đợi React cập nhật state

    login_btn = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//button[@type='submit'] | //button[@id='btn-login'] | //button[contains(text(), 'Log in')] | //button[contains(text(), 'Login')]"))
    )
    js_click(login_btn)
    print("  ⏳ Đã bấm Login, đợi form nhập liệu...")
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, FORM_XPATH)))
    print("  ✅ Đăng nhập thành công!")

# ==========================================
# MỞ TAB MỚI + TỰ ĐỘNG ĐĂNG NHẬP NẾU CẦN
# ==========================================
ONE_URL = "https://ecomm.one-line.com/one-ecom/prices/one-quote-booking"
LOGIN_URL_PREFIX = "https://auth.one-line.com/login"

def ensure_current_tab_ready(timeout=15):
    """Ensure current tab is on the ONE quote form; login/reload when needed.
    FIX: thay time.sleep(2) sau driver.get bằng đợi document.readyState."""
    try:
        cur = driver.current_url or ""
    except Exception:
        cur = ""

    def _wait_ready():
        try:
            WebDriverWait(driver, 8).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
        except Exception:
            pass

    if ONE_URL not in cur:
        driver.get(ONE_URL)
        _wait_ready()

    for attempt in range(2):
        cur = driver.current_url or ""
        if LOGIN_URL_PREFIX in cur or "auth.one-line.com" in cur or "login" in cur.lower():
            print("  ⚠️ Tab ONE đang ở login, đăng nhập lại...")
            do_login()
            driver.get(ONE_URL)
            _wait_ready()

        try:
            WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, FORM_XPATH))
            )
            print("  ✅ ONE quote form đã sẵn sàng.")
            return True
        except TimeoutException:
            if attempt == 0:
                print("  ⚠️ Chưa thấy form, reload lại ONE quote page...")
                driver.get(ONE_URL)
                _wait_ready()

    print(f"  ❌ Không vào được ONE quote form. URL hiện tại: {(driver.current_url or '')[:120]}")
    return False

def open_tab_and_ensure_ready():
    driver.switch_to.new_window('tab')
    handle = driver.current_window_handle
    ensure_current_tab_ready()
    return handle

def one_handle_alive(handle):
    try:
        return bool(handle) and handle in driver.window_handles
    except Exception:
        return False

def switch_to_one_handle(handle):
    if not one_handle_alive(handle):
        raise NoSuchWindowException(f"ONE tab handle no longer exists: {str(handle)[-6:]}")
    driver.switch_to.window(handle)
    return True

def replace_dead_one_tab(index, row=None, reason="dead tab"):
    print(f"  🔁 ONE tab {index+1} mất handle ({reason}) -> mở tab mới sạch")
    handle = open_tab_and_ensure_ready()
    tabs[index] = handle
    if row is not None:
        fill_tab_ports(row)
    return handle

def open_tab_and_ensure_ready_legacy():
    """
    Mở tab mới → vào ONE_URL.
    Check URL hiện tại:
      - URL = ONE_URL  → session còn, vào luôn
      - URL chứa auth.one-line.com/login → đăng nhập rồi vào form
    """
    driver.switch_to.new_window('tab')
    handle = driver.current_window_handle
    driver.get(ONE_URL)

    # Đợi trang load xong (tối đa 10s)
    time.sleep(2)

    # CHECK URL để quyết định hành động
    current_url = driver.current_url

    if ONE_URL in current_url:
        # Đang ở trang check giá → vào luôn
        try:
            WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.XPATH, FORM_XPATH))
            )
            print("  ✅ Session còn hiệu lực! URL đúng trang check giá.")
        except:
            print("  ⚠️ URL đúng nhưng form chưa hiện, đợi thêm...")
            time.sleep(3)
        return handle

    elif LOGIN_URL_PREFIX in current_url:
        # Đang ở trang đăng nhập
        print("  ⚠️ Bị redirect trang login. Tiến hành đăng nhập...")
        try:
            do_login()
        except Exception as e:
            print(f"  ❌ Lỗi đăng nhập: {e}")
        return handle

    else:
        # URL lạ khác → thử đăng nhập phòng hờ
        print(f"  ⚠️ URL không xác định: {current_url}")
        print("  🔄 Thử navigate lại trang check giá...")
        driver.get(ONE_URL)
        time.sleep(2)
        current_url = driver.current_url
        if LOGIN_URL_PREFIX in current_url:
            try:
                do_login()
            except Exception as e:
                print(f"  ❌ Lỗi đăng nhập: {e}")
        return handle

# ==========================================
# NHẬP POL/POD SAU RELOAD (có check login)
# ==========================================
def fill_tab_ports(row):
    """
    Đợi form trống sau reload rồi nhập POL/POD.
    Nếu thay vào đó thấy trang login (session hết) → tự đăng nhập trước.
    """
    country = str(row[0]).strip()
    pol     = str(row[2]).strip()
    pod     = str(row[3]).strip()

    # Đợi form nhập liệu
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, FORM_XPATH)))
    except TimeoutException:
        # Session có thể đã hết sau reload
        print("  ⚠️ Sau reload không thấy form → kiểm tra đăng nhập lại...")
        try:
            do_login()
        except Exception as e:
            print(f"  ❌ Không thể đăng nhập lại: {e}")
            return

    select_port(1, pol, country)
    select_port(2, pod, country)
    time.sleep(0.6)
    if has_no_port_pair_error():
        print(f"  🔁 Web báo No port pair sau khi nhập full -> retry POD bằng query rút ngắn")
        select_port(2, pod, country, allow_trim=True)


# ==========================================
# SWITCH TAB TRICK → ĐỢI HẾT POPUP LOADING
# ==========================================
def switch_tab_trick_until_clear(tabs, rounds=3, pause=0.15):
    # FIX: rounds 5→3, pause 0.2→0.15 — cắt ~50% thời gian đảo tab
    # mà vẫn đủ để React triệu hồi popup.
    print("🔄 Switch tab ép React render + đợi hết popup loading...")
    loader_xpath = "//div[contains(@class, 'CarouselLoadingPopup_progress-container') or contains(@class, 'ajax-progress')]"

    # Phase 0: Trigger focus/visibility events trên mỗi tab
    for tab in tabs:
        try:
            switch_to_one_handle(tab)
        except Exception:
            continue
        try:
            driver.execute_script("""
                window.dispatchEvent(new Event('focus'));
                document.dispatchEvent(new Event('focus'));
                document.dispatchEvent(new Event('visibilitychange'));
            """)
        except:
            pass
        time.sleep(0.05)

    # Phase 1: đảo nhanh để React kịp văng popup lên màn hình
    for _ in range(rounds):
        for tab in tabs:
            try:
                switch_to_one_handle(tab)
            except Exception:
                continue
            time.sleep(pause)

    # Phase 2: từng tab, đợi popup biến mất hẳn
    print("⏳ Đang chờ popup tắt trên từng tab...")
    for i, tab in enumerate(tabs):
        try:
            switch_to_one_handle(tab)
        except Exception:
            print(f"  ⚠️ Tab {i+1}: handle đã chết, bỏ qua wait popup")
            continue
        try:
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, loader_xpath)))
            WebDriverWait(driver, 45).until(EC.invisibility_of_element_located((By.XPATH, loader_xpath)))
            print(f"  ✅ Tab {i+1}: popup đã tắt")
        except TimeoutException:
            pass  # Không có popup hoặc đã tắt rồi → OK

    # FIX Phase 3: Xác nhận từng tab đã load xong bằng cách đợi form equipment HOẶC thông báo lỗi
    # → tránh tình trạng popup dùng class khác mà Phase 2 không phát hiện được
    print("🔎 Xác nhận trang đã load xong (chờ form equipment hoặc error msg)...")
    PAGE_READY_XPATH = (
        "//input[@placeholder='Select an Equipment Type']"
        " | //div[contains(@class,'RouteInput_wrap-error-loading')]"
        " | //p[contains(text(),'No port pair')]"
        " | //input[@type='password']"
    )
    for i, tab in enumerate(tabs):
        try:
            switch_to_one_handle(tab)
        except Exception:
            print(f"  ⚠️ Tab {i+1}: handle đã chết, bỏ qua xác nhận form")
            continue
        # Kiểm tra nếu trang đang ở login page → skip chờ form
        try:
            cur_url = driver.current_url or ""
            if "login" in cur_url.lower() or "auth" in cur_url.lower() or "sso" in cur_url.lower():
                print(f"  ⚠️ Tab {i+1}: đang ở trang login, thử đăng nhập lại")
                ensure_current_tab_ready()
        except:
            pass
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, PAGE_READY_XPATH))
            )
            print(f"  ✅ Tab {i+1}: form sẵn sàng")
        except TimeoutException:
            # Thử thêm 1 lần switch tab cuối để kick React
            for other_tab in tabs:
                if other_tab != tab:
                    try:
                        switch_to_one_handle(other_tab)
                    except Exception:
                        continue
                    time.sleep(0.3)
                    break
            try:
                switch_to_one_handle(tab)
            except Exception:
                continue
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, PAGE_READY_XPATH))
                )
                print(f"  ✅ Tab {i+1}: form sẵn sàng (sau retry)")
            except TimeoutException:
                print(f"  ⚠️ Tab {i+1}: timeout — tiếp tục (sẽ handle trong scrape_tab)")

    print("✅ Tất cả tab sạch popup!")


# ==========================================
# RELOAD ĐỒNG LOẠT (KHÔNG CHỜ NHAU)
# ==========================================
def reload_all_tabs_simultaneously(tabs):
    print(f"🔄 Gửi lệnh Reload đồng loạt {len(tabs)} tab...")
    for i, tab in enumerate(tabs):
        try:
            switch_to_one_handle(tab)
            driver.refresh()  # Không đợi, sang tab kế ngay → tất cả reload song song
        except Exception as e:
            print(f"  ⚠️ Tab {i+1}: không reload được ({type(e).__name__})")
        print(f"  ↩️  Tab {i+1}: đã gửi lệnh reload")
    print("✅ Đã gửi reload cho toàn bộ tabs!")


# ==========================================
# BÓC GIÁ MỘT TAB (driver đã switch sẵn)
# ==========================================
def has_commodity_mismatch_error():
    try:
        body_text = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
        return (
            "couldn't find a matching commodity" in body_text
            or "could not find a matching commodity" in body_text
            or ("please select again" in body_text and "commodity" in body_text)
        )
    except Exception:
        return False

def has_no_port_pair_error():
    xpaths = [
        "//div[contains(@class,'RouteInput_wrap-error-loading')]",
        "//div[contains(@class,'popover') or contains(@class,'tooltip') or contains(@class,'dropdown') or contains(@class,'error')][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'no port pair available')]",
        "//div[contains(@class,'popover') or contains(@class,'tooltip') or contains(@class,'dropdown') or contains(@class,'error')][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'request the missing port pair')]",
        "//input[contains(@class,'error') or contains(@aria-invalid,'true')]/ancestor::*[self::div or self::section][1][contains(translate(normalize-space(.), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'no port pair available')]",
    ]
    for xp in xpaths:
        try:
            elems = driver.find_elements(By.XPATH, xp)
            if any(e.is_displayed() for e in elems):
                return True
        except Exception:
            continue
    return False

def one_no_port_pair_result(row_data):
    return {
        "POL": row_data[2], "POD": row_data[3], "Status": "No port pair",
        "20 DRY": "-", "40 DRY": "-", "40 HC": "-",
        "ETD": "-", "Transit Time": "-", "Valid": "-",
        "Remark": "-", "Free Time": "-",
        "Vessel": "-", "Transshipment": "-"
    }

def reset_current_tab_for_row(row):
    try:
        ONE_TAB_PROGRESS.pop(driver.current_window_handle, None)
    except Exception:
        pass
    driver.get(ONE_URL)
    if not ensure_current_tab_ready(timeout=20):
        raise Exception("Không reset được tab ONE về form sạch")
    fill_tab_ports(row)

ONE_USE_API = os.environ.get("ONE_USE_API", "1").strip() != "0"
ONE_API_FALLBACK_SELENIUM = os.environ.get("ONE_API_FALLBACK_SELENIUM", "1").strip() != "0"
ONE_API_STRICT_FREE_TIME = os.environ.get("ONE_API_STRICT_FREE_TIME", "1").strip() != "0"
ONE_API_BASE = "https://ecomm.one-line.com/api"

ONE_EQUIPMENT_TARGETS = {
    "DRY 20": {"iso": "22G1", "fallback_size": "20", "fallback_one": "D2"},
    "DRY 40": {"iso": "42G1", "fallback_size": "40", "fallback_one": "D4"},
    "DRY 40H": {"iso": "45G1", "fallback_size": "40H", "fallback_one": "D5"},
}

ONE_ORIGIN_LOCAL_EXCLUDED_KEYWORDS = (
    "terminal handling charge (l)",
    "terminal handling charge at origin",
    "terminal handling charge at port of loading",
    "thc/l",
    "thc at origin",
    "doc fee (origin)",
    "document fee",
    "bill of lading",
    "b/l fee",
    "bl fee",
    "seal fee",
    "seal charge",
    "entry summary declaration surcharge",
    "advanced manifest",
    "manifest declaration",
    "manifest fee",
    "customs manifest submission fee",
    "ams",
    "ens",
    "afs",
    "afr",
    "heavy surcharge",
    "heavy lift",
    "overweight",
    "ows",
)

ONE_ORIGIN_LOCAL_EXCLUDED_CODES = {
    "THC", "THCL", "OTHC", "DOC", "BLF", "BLC", "SEAL", "SLF",
    "AMS", "ENS", "AFS", "AFR", "EST", "OWS", "HWC", "HLC", "OOG",
}

def one_fee_name(charge_or_text):
    if isinstance(charge_or_text, dict):
        return str(charge_or_text.get("chargeName") or "").strip().lower()
    return str(charge_or_text or "").strip().lower()

def one_fee_code(charge):
    if isinstance(charge, dict):
        return str(charge.get("chargeCode") or "").strip().upper()
    return ""

def one_is_origin_local_charge(fee_name, charge_code=""):
    fee = one_fee_name(fee_name)
    code = str(charge_code or "").strip().upper()
    if any(k in fee for k in ONE_ORIGIN_LOCAL_EXCLUDED_KEYWORDS):
        return True
    return bool(code and code in ONE_ORIGIN_LOCAL_EXCLUDED_CODES)

def one_is_origin_thc_charge(fee_name, charge_code=""):
    fee = one_fee_name(fee_name)
    code = str(charge_code or "").strip().upper()
    return (
        code in {"THC", "THCL", "OTHC"}
        or "terminal handling charge (l)" in fee
        or "terminal handling charge at origin" in fee
        or "terminal handling charge at port of loading" in fee
        or "thc/l" in fee
        or "thc at origin" in fee
    )

def one_is_ows_charge(fee_name, charge_code=""):
    """OWS/Heavy chỉ dùng để tạo remark, không được cộng vào giá."""
    fee = one_fee_name(fee_name)
    code = str(charge_code or "").strip().upper()
    return (
        code in {"OWS", "HWC", "HLC"}
        or "heavy surcharge" in fee
        or "heavy lift" in fee
        or "overweight" in fee
        or re.search(r"\bows\b", fee) is not None
    )

def one_should_include_charge(fee_name, group="", charge_code="", include_origin_thc=False):
    fee = one_fee_name(fee_name)
    group = str(group or "").strip()
    code = str(charge_code or "").strip().upper()
    if not fee:
        return False
    # OWS có thể nằm trong premiumCharges/freightCharges chứ không chỉ
    # originCharges. Luôn loại khỏi tổng, nhưng caller vẫn bật cờ remark.
    if one_is_ows_charge(fee, code):
        return False
    if group == "basicOceanFreightCharges":
        return True
    if group == "originCharges":
        if include_origin_thc and one_is_origin_thc_charge(fee, code):
            return True
        return not one_is_origin_local_charge(fee, code)
    if group == "destinationCharges":
        return False
    if not group:
        if include_origin_thc and one_is_origin_thc_charge(fee, code):
            return True
        return not one_is_origin_local_charge(fee, code)
    return True

def one_selenium_charge_group(ul):
    try:
        group = driver.execute_script("""
            const ul = arguments[0];
            const findGroup = (text) => {
                text = String(text || '').toLowerCase();
                if (text.includes('destination charge')) return 'destinationCharges';
                if (text.includes('origin charge')) return 'originCharges';
                if (text.includes('basic ocean freight')) return 'basicOceanFreightCharges';
                if (text.includes('freight charge')) return 'freightCharges';
                return '';
            };
            let node = ul;
            for (let depth = 0; node && depth < 10; depth++, node = node.parentElement) {
                let sib = node.previousElementSibling;
                for (let hop = 0; sib && hop < 12; hop++, sib = sib.previousElementSibling) {
                    const found = findGroup(sib.innerText || sib.textContent || '');
                    if (found) return found;
                }
            }
            return '';
        """, ul)
        return str(group or "")
    except Exception:
        return ""

def one_api_token_exists():
    try:
        return bool(driver.execute_script("return document.cookie.includes('accessToken=')"))
    except Exception:
        return False

def one_api_request(method, path, payload=None, timeout=None):
    if timeout is None:
        timeout = ONE_API_TIMEOUT_DEFAULT
    url = path if str(path).startswith("http") else ONE_API_BASE + path
    result = driver.execute_async_script("""
        const method = arguments[0], url = arguments[1], payload = arguments[2], timeout = arguments[3];
        const done = arguments[arguments.length - 1];
        const tokenPair = document.cookie.split('; ').find(x => x.startsWith('accessToken='));
        const token = tokenPair ? decodeURIComponent(tokenPair.split('=').slice(1).join('=')) : '';
        const headers = {'Accept': 'application/json'};
        if (token) headers['Authorization'] = 'Bearer ' + token;
        if (payload !== null && payload !== undefined) headers['Content-Type'] = 'application/json';
        let finished = false;
        const finish = (value) => { if (!finished) { finished = true; done(value); } };
        const timer = setTimeout(() => finish({ok:false, status:0, error:'ONE API timeout'}), timeout * 1000);
        fetch(url, {method, headers, credentials:'include',
            body: (payload !== null && payload !== undefined) ? JSON.stringify(payload) : undefined
        }).then(async r => {
            clearTimeout(timer);
            const text = await r.text();
            let data = null;
            try { data = text ? JSON.parse(text) : null; } catch (e) { data = text; }
            finish({ok:r.ok, status:r.status, data, text});
        }).catch(e => { clearTimeout(timer); finish({ok:false, status:0, error:String(e)}); });
    """, method.upper(), url, payload, timeout)
    if not isinstance(result, dict) or not result.get("ok"):
        status = result.get("status") if isinstance(result, dict) else "?"
        err = result.get("error") or result.get("text") or result.get("data") if isinstance(result, dict) else result
        raise Exception(f"ONE_API_HTTP_{status}: {err}")
    return result.get("data")

def one_api_get(path, params=None, timeout=None):
    if params:
        path = path + ("&" if "?" in path else "?") + urlencode(params, doseq=True)
    return one_api_request("GET", path, timeout=timeout)

def one_api_post(path, payload, timeout=None):
    return one_api_request("POST", path, payload, timeout=timeout)

def one_api_location_candidates(port_name, org_dest, country_name="", allow_trim=False, trim_only=False):
    rows_out = []
    seen = set()
    country = str(country_name or "").strip().upper()
    search_terms = one_port_search_terms(port_name, allow_trim=allow_trim)
    if trim_only and len(search_terms) > 1:
        search_terms = search_terms[1:]
    for name in search_terms:
        try:
            data = one_api_get("/v2/quotation/locations", {"location": name, "orgDest": org_dest, "searchFrom": "mdm"})
        except Exception as e:
            print(f"      [ONE-API] location query '{name}' lỗi: {type(e).__name__}")
            continue
        rows = (data or {}).get("data") or []
        if not rows:
            print(f"      [ONE-API] location query '{name}' không có data")
            continue
        print(f"      [ONE-API] location query '{name}' -> {len(rows)} option")
        for r in rows:
            key = (
                r.get("locationCode") or r.get("UNLocationCode") or "",
                r.get("locationType") or "",
                r.get("displayedName") or r.get("locationName") or "",
            )
            key = tuple(str(x).upper() for x in key)
            if key in seen:
                continue
            seen.add(key)
            rows_out.append(r)

    def score(r):
        hay = f"{r.get('displayedName','')} {r.get('locationName','')} {r.get('countryName','')} {r.get('countryCode','')}".upper()
        s = 0
        if str(r.get("locationType", "")).upper() == "CY":
            s += 100
        if country and org_dest == "destination" and country in hay:
            s += 60
        port_key = str(port_name or "").strip().upper()
        if port_key and port_key in hay:
            s += 25
        return s

    return sorted(rows_out, key=score, reverse=True)

def one_api_pick_location(port_name, org_dest, country_name=""):
    rows = one_api_location_candidates(port_name, org_dest, country_name)
    if not rows:
        return None
    cy_rows = [r for r in rows if str(r.get("locationType", "")).upper() == "CY"] or rows
    country = str(country_name or "").strip().upper()
    if country and org_dest == "destination":
        for r in cy_rows:
            hay = f"{r.get('countryName','')} {r.get('countryCode','')} {r.get('displayedName','')}".upper()
            if country in hay:
                return r
    return cy_rows[0]

def one_extract_codes(values):
    codes = []
    for item in values or []:
        code = item.get("commodityCode") if isinstance(item, dict) else item
        if code is not None:
            codes.append(str(code))
    return codes

def one_api_build_containers(schedule_data):
    equipment_types = (schedule_data or {}).get("equipmentTypes") or []
    containers = []
    for display_name, meta in ONE_EQUIPMENT_TARGETS.items():
        found = None
        for item in equipment_types:
            if item.get("equipmentDisplayName") == display_name or item.get("equipmentIsoCode") == meta["iso"]:
                found = dict(item)
                break
        if not found:
            found = {"cargoType": "DR", "equipmentIsoCode": meta["iso"], "equipmentDisplayName": display_name,
                     "equipmentName": display_name, "equipmentSize": meta["fallback_size"],
                     "equipmentONECntrTpSz": meta["fallback_one"], "isSoc": False, "commodityGroups": []}
        found.update({"cargoType": found.get("cargoType") or "DR", "equipmentName": found.get("equipmentName") or display_name,
                      "quantity": 1, "unit": "KGS", "isSelected": False, "isFocus": False,
                      "isWeightError": False, "uuid": str(uuid.uuid4()), "isFoodGrade": False,
                      "isInland": False, "cargoWeight": "22222"})
        containers.append(found)
    return containers

def one_api_charge_amount_usd(charge):
    try:
        if charge.get("totalAmountInUSD") not in (None, ""):
            return float(str(charge.get("totalAmountInUSD")).replace(",", ""))
    except Exception:
        pass
    try:
        amount = float(str(charge.get("totalAmount") or charge.get("chargeAmount") or 0).replace(",", ""))
    except Exception:
        amount = 0.0
    return amount * get_live_exchange_rate(str(charge.get("chargeCurrency") or "USD").upper(), "USD")

def one_api_charge_equipment_key(charge):
    iso = str(charge.get("equipmentIsoCode") or "").upper()
    if iso in {"22G1", "42G1", "45G1"}:
        return {"22G1": "DRY 20", "42G1": "DRY 40", "45G1": "DRY 40H"}[iso]
    text = " ".join(
        str(charge.get(k) or "")
        for k in (
            "equipmentDisplayName", "equipmentName", "equipmentType",
            "equipmentSize", "equipmentONECntrTpSz", "containerType",
            "cntrTpSz", "chargeUnit", "unit", "description"
        )
    ).upper()
    if any(x in text for x in ["45G1", "DRY 40H", "40H", "40 HC", "40HQ", "HIGH CUBE"]):
        return "DRY 40H"
    if any(x in text for x in ["42G1", "DRY 40", "40GP", "40 GP", "40'"]):
        return "DRY 40"
    if any(x in text for x in ["22G1", "DRY 20", "20GP", "20 GP", "20'"]):
        return "DRY 20"
    return ""

def one_api_add_charge(final_prices, formula_parts, charge):
    key = one_api_charge_equipment_key(charge)
    amount = one_api_charge_amount_usd(charge)
    if abs(amount) < 1e-9:
        return

    # Charges such as EST / Customs Manifest Submission Fee are per
    # shipment/B/L and therefore have no equipmentIsoCode.  They still belong
    # in the sell rate: apply the common amount once to each standalone
    # equipment quote instead of silently dropping it.
    target_keys = [key] if key else list(final_prices.keys())
    for target_key in target_keys:
        final_prices[target_key] += amount
        formula_parts[target_key].append(amount)

def one_api_status_bookable(status):
    text = str(status or "").strip().lower()
    return not any(w in text for w in ["sold", "wait", "unavailable", "no space", "expired"])

def one_api_clean_port_name(value):
    text = re.sub(r"\s*\([^)]*\)", "", str(value or "")).strip()
    return re.split(r"\s*,\s*", text)[0].strip().upper()

def _one_api_day_value(value):
    try:
        if value in (None, ""):
            return ""
        number = float(str(value).replace(",", "").strip())
        if number <= 0:
            return ""
        return str(int(number)) if abs(number - int(number)) < 1e-9 else str(number).rstrip("0").rstrip(".")
    except Exception:
        return ""

def _one_api_parse_dd_section(section):
    if not isinstance(section, dict):
        return ""
    dem = section.get("demurrage") or {}
    det = section.get("detention") or {}
    dem_days = _one_api_day_value(dem.get("ftDys") or dem.get("freeDays"))
    det_days = _one_api_day_value(det.get("ftDys") or det.get("freeDays"))
    combined_codes = {"CTOC", "CTIC"}
    dem_code = str(dem.get("dmdtTrfCd") or "").upper()
    det_code = str(det.get("dmdtTrfCd") or "").upper()
    if dem_code in combined_codes or det_code in combined_codes:
        days = dem_days or det_days
        return f"{days} COMBINED" if days else ""
    if dem_days and det_days:
        return f"{dem_days} DEM + {det_days} DET"
    if dem_days:
        return f"{dem_days} DEM"
    if det_days:
        return f"{det_days} DET"
    return ""

def one_api_parse_free_time(freight_info, date_item=None):
    for source in (freight_info, date_item):
        if not isinstance(source, dict):
            continue
        dd_info = (
            source.get("detentionDemurrageInfo")
            or source.get("demurrageDetentionInfo")
            or source.get("dndInfo")
        )
        if isinstance(dd_info, dict):
            parsed = _one_api_parse_dd_section(dd_info.get("destination"))
            if parsed:
                return parsed
        dd = source.get("demerageDetention") or source.get("demurrageDetention")
        if isinstance(dd, dict) and isinstance(dd.get("destination"), dict):
            parsed = _one_api_parse_dd_section(dd.get("destination"))
            if parsed:
                return parsed
    return "N/A"

def one_api_charge_currency(fi):
    if not isinstance(fi, dict):
        return "USD"
    for group in ("freightCharges", "basicOceanFreightCharges", "premiumCharges", "originCharges", "additionalCharges"):
        for charge in fi.get(group) or []:
            currency = str(charge.get("chargeCurrency") or "").strip().upper()
            if currency:
                return currency
    return "USD"

def one_api_destination_effective_date(fi, date_item):
    if not isinstance(fi, dict):
        fi = {}
    if not isinstance(date_item, dict):
        date_item = {}
    candidates = [
        (fi.get("arrivalPod") or {}).get("arrivalDateEstimated"),
        (fi.get("arrival") or {}).get("arrivalDateEstimated"),
        fi.get("arrivalDateEstimated"),
        date_item.get("arrivalDateEstimated"),
    ]
    departures = fi.get("departures") or []
    if departures:
        candidates.extend([
            departures[-1].get("arrivalDateEstimated"),
            departures[-1].get("arrivalDate"),
        ])
    for value in candidates:
        if value:
            return value
    return date_item.get("departureDateEstimated") or fi.get("departureDateEstimated") or datetime.now().strftime("%Y-%m-%d")

def one_api_enrich_free_time(candidate, origin, dest, containers, commodity):
    fi = candidate.get("freight_info") or {}
    if one_api_parse_free_time(fi, candidate.get("date_item")) != "N/A":
        return
    departures = fi.get("departures") or []
    payload = {
        "originUNLocationCode": origin.get("locationCode") or origin.get("UNLocationCode") or "",
        "destinationUNLocationCode": dest.get("locationCode") or dest.get("UNLocationCode") or "",
        "effectiveDateTime": one_api_destination_effective_date(fi, candidate.get("date_item")),
        "containerCargos": [
            {
                "equipmentIsoCode": c.get("equipmentIsoCode"),
                "equipmentONECntrTpSz": c.get("equipmentONECntrTpSz"),
                "quantity": c.get("quantity") or 1,
            }
            for c in containers
        ],
        "additionalDays": 1,
        "oftCurrency": one_api_charge_currency(fi),
        "commodityCode": (commodity or {}).get("commodityCode") or "621101",
        "porYardCode": departures[0].get("departureYardCode") if departures else "",
        "delYardCode": departures[-1].get("arrivalYardCode") if departures else "",
        "serviceScope": fi.get("serviceScope") or fi.get("serviceScopeCode") or "",
    }
    try:
        data = one_api_post("/v2/quotation/demurrage-detention-info", payload, timeout=ONE_FREE_TIME_API_TIMEOUT) or {}
        fi["detentionDemurrageInfo"] = {
            "origin": data.get("origin"),
            "destination": data.get("destination"),
        }
        parsed = one_api_parse_free_time(fi, candidate.get("date_item"))
        print(f"      [ONE-API] Free time API ETD={candidate.get('etd')} -> {parsed}")
    except Exception as e:
        print(f"      [ONE-API] Không đọc được free time API ETD={candidate.get('etd')}: {e}")

def one_api_parse_vessel(fi, etd_date, pol_name):
    departures = fi.get("departures") or []
    ocean_segments = [d for d in departures if "ocean" in str(d.get("transportType", "")).lower()]
    first_ocean = ocean_segments[0] if ocean_segments else (departures[0] if departures else {})
    vessel = " ".join(str(x or "").strip() for x in [
        first_ocean.get("transportName") or fi.get("transportName") or "TBA",
        first_ocean.get("conveyanceNumber") or fi.get("conveyanceNumber") or "",
    ]).strip()
    final_arrival = (fi.get("arrival") or {}).get("arrivalLoc")
    ts_ports = []
    for seg in departures:
        arr_loc = seg.get("arrivalLoc")
        if not arr_loc or arr_loc == final_arrival:
            continue
        pname = one_api_clean_port_name(seg.get("arrivalTerminal") or arr_loc)
        if ("HO CHI MINH" in str(pol_name).upper() or "CAI MEP" in str(pol_name).upper()) and ("CAI MEP" in pname or "VUNG TAU" in pname):
            continue
        if pname and pname not in ts_ports:
            ts_ports.append(pname)
    if not ts_ports and str(fi.get("routeType") or "").upper() == "TRANSIT":
        ts_ports = ["TRANSIT"]
    ts_excel = " + ".join(ts_ports) if ts_ports else "DIRECT"
    return f"{vessel} / ETD: {etd_date.day}-{etd_date.strftime('%b')} / Transit time: {int(fi.get('duration') or 0)} Days / Transshipment Port: {ts_excel}", ts_excel

def one_api_format_result(row_data, candidates):
    if not candidates:
        return {"POL": row_data[2], "POD": row_data[3], "Status": "No Schedule",
                "20 DRY": "-", "40 DRY": "-", "40 HC": "-", "ETD": "-",
                "Transit Time": "-", "Valid": "-", "Remark": "-", "Free Time": "-",
                "Vessel": "-", "Transshipment": "-"}
    min_price = min(c["price"] for c in candidates)
    unique_etd = {}
    for c in [x for x in candidates if x["price"] <= min_price + 40]:
        if c["etd"] not in unique_etd or c["transit"] < unique_etd[c["etd"]]["transit"]:
            unique_etd[c["etd"]] = c
    final_sels, min_etd = [], datetime.today().date() + timedelta(days=DATE_OFFSET_DAYS)
    for c in sorted(unique_etd.values(), key=lambda x: x["etd"]):
        if c["etd"] < min_etd:
            continue
        if not etd_within_max(c["etd"]):
            continue
        if not final_sels:
            final_sels.append(c)
        elif (c["etd"] - final_sels[-1]["etd"]).days >= 2 and (c["etd"] - final_sels[0]["etd"]).days <= 9:
            final_sels.append(c)
        if len(final_sels) == 3:
            break
    if not final_sels:
        return {"POL": row_data[2], "POD": row_data[3], "Status": "No Valid ETD"}

    etd_strs = [f"{c['etd'].day}-{c['etd'].strftime('%b')}" for c in final_sels]
    if len(etd_strs) == 1:
        etd_excel = etd_strs[0]
    elif len(etd_strs) == 2:
        etd_excel = f"{etd_strs[0]} & {etd_strs[1]}"
    else:
        etd_excel = f"{', '.join(str(c['etd'].day) for c in final_sels[:-1])}, {final_sels[-1]['etd'].day}-{final_sels[0]['etd'].strftime('%b')}"
    transits = [c["transit"] for c in final_sels]
    transit_excel = str(transits[0]) if len(set(transits)) == 1 else f"{transits[0]}-{transits[-1]}"
    valid_excel = get_valid_date([c["etd"] for c in final_sels])
    target, fi = final_sels[0], final_sels[0]["freight_info"]
    debug_charges = os.environ.get("ONE_DEBUG_CHARGES", "0").strip().lower() in {"1", "true", "yes", "y"}
    if debug_charges:
        print(f"      [ONE-CHARGE] freightInfo keys={sorted(fi.keys())}")
        for debug_key, debug_value in fi.items():
            try:
                debug_text = json.dumps(debug_value, ensure_ascii=False, default=str)
            except Exception:
                debug_text = str(debug_value)
            if "premium" in str(debug_key).lower() or "premium" in debug_text.lower():
                print(f"      [ONE-PREMIUM-RAW] {debug_key}={debug_text[:5000]}")
    final_prices = {"DRY 20": 0.0, "DRY 40": 0.0, "DRY 40H": 0.0}
    formula_parts = {"DRY 20": [], "DRY 40": [], "DRY 40H": []}
    country_upper = str(row_data[0]).strip().upper()
    pod_upper = str(row_data[3]).strip().upper()
    china_route = is_china_destination(country_upper, pod_upper)
    has_thc = has_ens_ams = has_ows = False
    for group in ["basicOceanFreightCharges", "premiumCharges", "originCharges", "freightCharges", "destinationCharges", "additionalCharges", "inlandCharges"]:
        for charge in fi.get(group) or []:
            fee_name = str(charge.get("chargeName") or "").strip().lower()
            if one_is_origin_thc_charge(fee_name, one_fee_code(charge)): has_thc = True
            if "entry summary declaration surcharge" in fee_name: has_ens_ams = True
            if one_is_ows_charge(fee_name, one_fee_code(charge)): has_ows = True
            include_charge = one_should_include_charge(
                fee_name, group, one_fee_code(charge), include_origin_thc=china_route
            )
            if debug_charges:
                print(
                    "      [ONE-CHARGE] "
                    f"group={group} code={one_fee_code(charge) or '-'} "
                    f"equipment={one_api_charge_equipment_key(charge) or '-'} "
                    f"unit={charge.get('chargeUnit') or charge.get('unit') or '-'} "
                    f"amountUSD={one_api_charge_amount_usd(charge):g} "
                    f"include={include_charge} name={charge.get('chargeName') or '-'}"
                )
            if not include_charge:
                continue
            one_api_add_charge(final_prices, formula_parts, charge)

    if debug_charges:
        print(
            f"      [ONE-CHARGE] displayedPrice={target.get('price')} "
            f"calculated={final_prices} formulas={formula_parts}"
        )

    remark_str = "SUBJECT TO THC, BILL, SEAL" if (has_thc and not china_route) else "INCLUDED O.THC, SUBJECT TO BILL, SEAL"
    manifest_code = get_manifest_code(country_upper, pod_upper)
    remark_str = apply_manifest_rule(remark_str, country_upper, pod_upper)
    if has_ens_ams and not manifest_code:
        remark_str += ", AMS"
    if has_ows: remark_str += ", OWS"

    free_time = one_api_parse_free_time(target["freight_info"], target["date_item"])
    if free_time == "N/A" and ONE_API_STRICT_FREE_TIME:
        raise Exception("ONE_API_FREE_TIME_UNAVAILABLE")

    vessel_excel, ts_excel = one_api_parse_vessel(fi, target["etd"], row_data[2])
    print(f"   [ONE-API] Chọn ETD={etd_excel} | TT={transit_excel} | Valid={valid_excel}")
    return {"POL": row_data[2], "POD": row_data[3],
            "20 DRY": final_prices["DRY 20"], "40 DRY": final_prices["DRY 40"], "40 HC": final_prices["DRY 40H"],
            "20 FORMULA": _excel_formula_from_parts(formula_parts["DRY 20"]),
            "40 FORMULA": _excel_formula_from_parts(formula_parts["DRY 40"]),
            "40HC FORMULA": _excel_formula_from_parts(formula_parts["DRY 40H"]),
            "ETD": etd_excel, "Transit Time": transit_excel, "Valid": valid_excel,
            "Remark": remark_str, "Free Time": free_time,
            "Vessel": vessel_excel, "Transshipment": ts_excel}

def scrape_one_api(row_data):
    if not one_api_token_exists():
        driver.get(ONE_URL)
        ensure_current_tab_ready(timeout=20)
    if not one_api_token_exists():
        raise Exception("ONE_API_NO_ACCESS_TOKEN")
    country, pol_name, pod_name = str(row_data[0]).strip(), str(row_data[2]).strip(), str(row_data[3]).strip()
    print(f"   [ONE-API] {pol_name} -> {pod_name}")
    origin_candidates = one_api_location_candidates(pol_name, "origin", country)
    dest_candidates = one_api_location_candidates(pod_name, "destination", country)
    if not origin_candidates or not dest_candidates:
        return one_no_port_pair_result(row_data)

    def probe_one_pair(origin_list, dest_list):
        for o in origin_list[:5]:
            for d in dest_list[:8]:
                origin_code = o.get("locationCode") or o.get("UNLocationCode")
                dest_code = d.get("locationCode") or d.get("UNLocationCode")
                if not origin_code or not dest_code:
                    continue
                try:
                    invalid = one_api_get("/v2/quotation/invalid-pp/exists", {"originLocationCode": origin_code, "destinationLocationCode": dest_code})
                    if (invalid or {}).get("exists") is True:
                        print(f"      [ONE-API] pair invalid: {origin_code}->{dest_code}, thử candidate khác")
                        continue
                    probe_trips = one_api_get("/v2/quotation/trips", {"fromCode": origin_code, "toCode": dest_code,
                        "originLocationType": o.get("locationType") or "CY", "destinationLocationType": d.get("locationType") or "CY", "searchFrom": "all"})
                    probe_rows = (probe_trips or {}).get("data") or []
                    if not probe_rows:
                        print(f"      [ONE-API] pair no trips: {origin_code}->{dest_code}, thử candidate khác")
                        continue
                    print(f"      [ONE-API] chọn pair: {origin_code}->{dest_code} ({len(probe_rows)} trips)")
                    return o, d, probe_trips, probe_rows
                except Exception as e:
                    print(f"      [ONE-API] pair probe lỗi {origin_code}->{dest_code}: {type(e).__name__}")
                    continue
        return None, None, None, []

    origin, dest, trips, trip_rows = probe_one_pair(origin_candidates, dest_candidates)

    if not origin or not dest or not trip_rows:
        print("      [ONE-API] full query bị No port pair/no trips -> retry bằng query rút ngắn")
        origin_candidates = one_api_location_candidates(pol_name, "origin", country, allow_trim=True, trim_only=True)
        dest_candidates = one_api_location_candidates(pod_name, "destination", country, allow_trim=True, trim_only=True)
        origin, dest, trips, trip_rows = probe_one_pair(origin_candidates, dest_candidates)

    if not origin or not dest or not trip_rows:
        return one_no_port_pair_result(row_data)

    origin_code, dest_code = origin.get("locationCode") or origin.get("UNLocationCode"), dest.get("locationCode") or dest.get("UNLocationCode")
    include_scopes = []
    for trip in trip_rows:
        code = trip.get("serviceScopeCode")
        if code and code not in include_scopes:
            include_scopes.append(code)
    session_id = str(uuid.uuid4())
    schedule = one_api_get("/v2/quotation/schedules/schedule-rates-information", {"originLoc": origin_code, "destinationLoc": dest_code,
        "originLocationType": origin.get("locationType") or "CY", "destinationLocationType": dest.get("locationType") or "CY",
        "porRhqCode": origin.get("regionHeadQuarter") or "", "includeSvcScopeCd": ",".join(include_scopes), "sessionId": session_id})
    containers = one_api_build_containers((schedule or {}).get("data") or {})
    cg_payload = {"originLoc": origin_code, "destinationLoc": dest_code, "originLocationType": origin.get("locationType") or "CY",
        "destinationLocationType": dest.get("locationType") or "CY", "containers": containers, "isSoc": False,
        "isFoodGrade": False, "includeSvcScopeCd": include_scopes, "sessionId": session_id}
    commodity_groups = (one_api_post("/v2/quotation/commodity-groups", cg_payload) or {}).get("data") or []
    cg_source = (containers[0].get("commodityGroups") or commodity_groups or [{}])[0]
    commodity_payload = {"commodityName": "6211", "commodityCode": "", "page": "1", "limit": "50",
        "include": ",".join(one_extract_codes(cg_source.get("include"))), "exclude": ",".join(one_extract_codes(cg_source.get("exclude"))),
        "includeChilled": "", "includeFrozen": "", "excludeChilled": "", "excludeFrozen": "", "cargoType": "DR", "reeferType": None}
    commodity_rows = (one_api_post("/v2/quotation/commodity", commodity_payload) or {}).get("data") or []
    commodity = next((x for x in commodity_rows if str(x.get("commodityCode", "")).startswith("6211") and "TRACK SUITS" in str(x.get("commodityName", "")).upper()), None)
    commodity = commodity or (commodity_rows[0] if commodity_rows else {"commodityCode": "621101", "reeferType": ""})
    vessel_payload = {"originLoc": origin_code, "originLocationType": origin.get("locationType") or "CY", "destinationLoc": dest_code,
        "destinationLocationType": dest.get("locationType") or "CY", "containers": containers, "commodityCode": commodity.get("commodityCode") or "621101",
        "reeferType": commodity.get("reeferType") or "", "isSoc": False, "isFoodGrade": False, "commodityGroups": commodity_groups,
        "polLocs": ["all"], "podLocs": ["all"], "serviceScopeCode": include_scopes[0], "includeSvcScopeCd": include_scopes,
        "sessionId": session_id, "isInlandProcess": bool((trips or {}).get("isInland"))}
    vessel_data = one_api_post("/v2/quotation/schedules/vessel-dates-booking", vessel_payload)
    candidates = []
    for date_item in (vessel_data or {}).get("data") or []:
        try:
            etd = datetime.strptime(str(date_item.get("departureDateEstimated"))[:10], "%Y-%m-%d").date()
        except Exception:
            continue
        if not etd_within_max(etd):
            continue
        for fi in date_item.get("freightInfos") or []:
            if not one_api_status_bookable(fi.get("status")):
                continue
            try:
                price = float(fi.get("price") or date_item.get("totalPrice") or 0)
            except Exception:
                price = 0.0
            if price <= 0:
                continue
            duration = int(float(fi.get("duration") or 0))
            candidate = {"price": price, "etd": etd, "transit": duration, "freight_info": fi, "date_item": date_item}
            one_api_enrich_free_time(candidate, origin, dest, containers, commodity)
            candidates.append(candidate)
            print(f"      [ONE-API] Bookable: ETD={etd} TT={duration}d Vessel={fi.get('transportName','')} {fi.get('conveyanceNumber','')} Status={fi.get('status','')}")
    return one_api_format_result(row_data, candidates)

def scrape_tab(row_data):
    tab_handle = driver.current_window_handle
    progress = ONE_TAB_PROGRESS.setdefault(tab_handle, {})
    pol_name = str(row_data[2]).strip().upper()
    try:
        cur_url = driver.current_url or ""
        if "login" in cur_url.lower() or "auth.one-line.com" in cur_url:
            ensure_current_tab_ready()
    except Exception:
        pass
    
    # ── KIỂM TRA "No port pair" BẰNG CLASS + TEXT CỦA WEB ──
    # Case này bỏ qua row luôn, không retry container/commodity/date.
    no_pair_deadline = time.time() + 2.5
    while time.time() < no_pair_deadline:
        if has_no_port_pair_error():
            print(f"  ⚠️ Web báo 'No port pair available' → bỏ qua row này")
            ONE_TAB_PROGRESS.pop(tab_handle, None)
            return one_no_port_pair_result(row_data)
        time.sleep(0.25)

    # ── Kiểm tra thêm: ô chọn container có hiện không ──
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located(
                (By.XPATH, "//input[@placeholder='Select an Equipment Type']")
            )
        )
    except TimeoutException:
        print(f"  ⚠️ Không thấy ô chọn container → trả về '-'")
        return {
            "POL": row_data[2], "POD": row_data[3], "Status": "No container form",
            "20 DRY": "-", "40 DRY": "-", "40 HC": "-",
            "ETD": "-", "Transit Time": "-", "Valid": "-",
            "Remark": "-", "Free Time": "-",
            "Vessel": "-", "Transshipment": "-"
        }

    # ── Nhập container bình thường ──
    if not progress.get("containers_done"):
        add_equipment(1, "DRY 20", "22222")
        row_label = f"row_{row_data[-1] + 2}_{row_data[2]}_{row_data[3]}"
        ensure_equipment_slots(2, row_label)
        
        # Chỉ bấm Add nếu ô thứ 2 chưa tồn tại (giải quyết triệt để lỗi khi tab được Retry)
        if len(driver.find_elements(By.XPATH, "(//input[@placeholder='Select an Equipment Type'])")) < 2:
            try: click_icon_add(); time.sleep(SPEED)
            except Exception: pass
        add_equipment(2, "DRY 40", "22222")
        ensure_equipment_slots(3, row_label)
        
        # Chỉ bấm Add nếu ô thứ 3 chưa tồn tại
        if len(driver.find_elements(By.XPATH, "(//input[@placeholder='Select an Equipment Type'])")) < 3:
            try: click_icon_add(); time.sleep(SPEED)
            except Exception: pass
        add_equipment(3, "DRY 40H", "22222")    

        # ── CHỐT CHẶN: ÉP CHẮC CHẮN 3 Ô SỐ LƯỢNG LÀ 1 TRƯỚC KHI ĐI TIẾP ──
        driver.execute_script("""
            let qty_inputs = document.querySelectorAll('button[data-action="increment"]');
            qty_inputs.forEach(btn => {
                let input = btn.parentElement.querySelector('input');
                if(input) {
                    // Ép form React nhận giá trị 1 một cách triệt để
                    let nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, "value").set;
                    nativeInputValueSetter.call(input, '1');
                    input.dispatchEvent(new Event('input', { bubbles: true }));
                    input.dispatchEvent(new Event('change', { bubbles: true }));
                }
            });
        """)
        progress["containers_done"] = True
        progress["step"] = "containers_done"
        progress["updated_at"] = time.time()
        print("  ✅ Container đã xong. Nếu commodity chưa render, bot sẽ chuyển tab khác rồi quay lại.")
        time.sleep(0.3)
        raise Exception("WEB_LAG_RETRY: Container đã xong, chuyển tab khác trong lúc commodity render")
    else:
        print("  ✅ Container đã cấu hình từ lượt trước, bỏ qua bước container.")

# ── Nhập Commodity (Check trước khi nhập) ──
    # Tăng timeout từ 1.5s → 15s vì web ONE đôi khi load commodity rất chậm
    if has_commodity_mismatch_error():
        raise Exception("ONE_COMMODITY_MISMATCH_RETRY: Commodity không match sau khi chọn container")

    COMMODITY_TIMEOUT = ONE_COMMODITY_STEP_WAIT
    commodity_wait = WebDriverWait(driver, COMMODITY_TIMEOUT)
    if not progress.get("commodity_done"):
        try:
            commodity_input = commodity_wait.until(EC.presence_of_element_located((By.XPATH, "//input[@name='searchCommodityByName']")))
            current_commodity = commodity_input.get_attribute("value") or ""

            # Nếu chưa có mã 6211 hoặc chữ TRACK SUITS thì mới nhập
            if "6211" not in current_commodity and "TRACK SUITS" not in current_commodity.upper():
                driver.execute_script("arguments[0].scrollIntoView({behavior:'smooth',block:'center'});", commodity_input)
                time.sleep(SPEED)
                commodity_wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@name='searchCommodityByName']"))).click()
                time.sleep(SPEED)
                commodity_input.send_keys(Keys.CONTROL + "a"); commodity_input.send_keys(Keys.BACKSPACE)
                commodity_input.send_keys("6211")
                # Dropdown option có thể chậm load → nếu chưa có thì chuyển tab khác rồi quay lại
                js_click(commodity_wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@role='option' and contains(., 'TRACK SUITS')]"))))
                time.sleep(SPEED)
            else:
                print("  ✅ Commodity đã có sẵn, bỏ qua bước nhập.")
            progress["commodity_done"] = True
            progress["step"] = "commodity_done"
            progress["updated_at"] = time.time()
        except TimeoutException:
            raise Exception(f"WEB_LAG_RETRY: Commodity chưa render, chuyển tab khác rồi quay lại (> {COMMODITY_TIMEOUT}s)")
    else:
        print("  ✅ Commodity đã chọn từ lượt trước, bỏ qua bước commodity.")

    # Chọn ngày tím đầu tiên còn chỗ (bỏ qua Sold out)
    if has_commodity_mismatch_error():
        raise Exception("ONE_COMMODITY_MISMATCH_RETRY: Commodity không match sau khi chọn container")

    calendar_opened = False

    # Danh sách XPath ưu tiên để tìm ô lịch tàu (tách riêng, không dùng | trong WebDriverWait)
    CALENDAR_XPATHS = [
        "/html/body/div[1]/main/div[2]/div[2]/main/div[2]/div[2]/div/div[2]/div/div[2]/div[4]/div[3]/div/div/div/div/div/input",
        "//input[@placeholder='Please select vessel departure date at origin']",
        "//div[contains(@class,'date-picker')]//input",
    ]

    for attempt in range(2):
        try:
            date_input = None
            for xp in CALENDAR_XPATHS:
                try:
                    date_input = WebDriverWait(driver, ONE_DATE_INPUT_STEP_WAIT).until(
                        EC.presence_of_element_located((By.XPATH, xp))
                    )
                    if date_input:
                        print(f"  📅 Tìm thấy ô lịch tàu bằng XPath: {xp[:60]}...")
                        break
                except TimeoutException:
                    continue

            if date_input is None:
                raise Exception(f"WEB_LAG_RETRY: Chưa tìm được ô input lịch tàu, chuyển tab khác rồi quay lại (> {ONE_DATE_INPUT_STEP_WAIT}s/xpath)")

            driver.execute_script("arguments[0].scrollIntoView({behavior:'smooth',block:'center'});", date_input)
            time.sleep(0.5)

            try:
                ActionChains(driver).move_to_element(date_input).click().perform()
            except Exception:
                driver.execute_script("arguments[0].click();", date_input)

            CALENDAR_READY_XPATH = (
                "//div[contains(@class,'date-picker-date-highlight') and @role='option' and @aria-disabled='false']"
            )
            try:
                WebDriverWait(driver, ONE_DATE_RENDER_STEP_WAIT).until(
                    EC.presence_of_element_located((By.XPATH, CALENDAR_READY_XPATH))
                )
                print(f"  ✅ Calendar đã render xong ngày có giá!")
            except TimeoutException:
                # Nếu lag, văng lỗi ngay lập tức để chuyển tab
                raise Exception(f"WEB_LAG_RETRY: Lịch tàu chưa hiện ngày tím, chuyển tab khác rồi quay lại (> {ONE_DATE_RENDER_STEP_WAIT}s)")

            # ── DEBUG: dump TẤT CẢ date cells để xem class thực tế ──
            all_date_cells = driver.find_elements(By.XPATH,
                "//div[@role='option' and @aria-disabled='false' and not(contains(@class,'outside-month'))]"
            )
            print(f"  [DEBUG] Tổng date cells (enabled, in-month): {len(all_date_cells)}")
            for i, dc in enumerate(all_date_cells[:15]):
                try:
                    cls = dc.get_attribute("class") or ""
                    lbl = (dc.get_attribute("aria-label") or "")[:60]
                    txt = dc.text.strip().replace('\n', ' ')[:40]
                    print(f"  [DEBUG]  [{i}] class='{cls}' | label='{lbl}' | text='{txt}'")
                except Exception:
                    pass

            # Tìm ngày có giá: ưu tiên highlight, fallback sang tất cả enabled date cells có giá
            HIGHLIGHT_XPATH = "//div[contains(@class,'date-picker-date-highlight') and @role='option' and @aria-disabled='false' and not(contains(@class,'outside-month'))]"
            highlight_dates = driver.find_elements(By.XPATH, HIGHLIGHT_XPATH)
            print(f"  🔍 Số ngày 'highlight': {len(highlight_dates)}")

            # Nếu ít highlight → thử tìm ngày có giá bằng cách rộng hơn
            if len(highlight_dates) <= 1:
                # Thử tìm ngày có text giá (chứa "K" như 8.6K, 9.5K) hoặc có class khác
                broader_xpaths = [
                    "//div[@role='option' and @aria-disabled='false' and not(contains(@class,'outside-month')) and (contains(@class,'highlight') or contains(@class,'in-range') or contains(@class,'available') or contains(@class,'selectable'))]",
                    "//div[@role='option' and @aria-disabled='false' and not(contains(@class,'outside-month')) and not(contains(@class,'disabled'))]",
                ]
                for bxp in broader_xpaths:
                    broader_dates = driver.find_elements(By.XPATH, bxp)
                    # Lọc chỉ những ngày có text chứa giá (K = nghìn)
                    dates_with_price = [d for d in broader_dates if re.search(r'\d+\.?\d*K', d.text or "")]
                    if dates_with_price:
                        print(f"  🔍 Tìm thêm {len(dates_with_price)} ngày có giá (broader search)")
                        highlight_dates = dates_with_price
                        break

            print(f"  🔍 Tổng ngày có giá: {len(highlight_dates)}")

            if not highlight_dates:
                # FIX: cho thêm 2s rồi đọc lại 1 lần nữa, nhiều khi ngày chưa kịp populate
                print(f"  ⏳ Chưa thấy ngày có giá, chờ thêm 2s rồi đọc lại...")
                time.sleep(2)
                highlight_dates = driver.find_elements(By.XPATH, HIGHLIGHT_XPATH)
                if len(highlight_dates) <= 1:
                    all_enabled = driver.find_elements(By.XPATH,
                        "//div[@role='option' and @aria-disabled='false' and not(contains(@class,'outside-month')) and not(contains(@class,'disabled'))]"
                    )
                    dates_with_price = [d for d in all_enabled if re.search(r'\d+\.?\d*K', d.text or "")]
                    if dates_with_price:
                        highlight_dates = dates_with_price
                print(f"  🔍 Lần đọc lại: {len(highlight_dates)} ngày có giá")

            if not highlight_dates:
                print(f"  ℹ️ Calendar đã mở nhưng không có ngày có giá → thử attempt khác")
                try:
                    date_input.send_keys(Keys.ESCAPE)
                except Exception:
                    pass
                time.sleep(1)
                calendar_opened = False
                continue  # FIX: retry thay vì break luôn

            # ── Lưu aria-label thay vì element để tránh stale ──
            chosen_aria_label = None
            chosen_date_str = ""
            min_etd_date = datetime.today().date() + timedelta(days=DATE_OFFSET_DAYS)

            for hd in highlight_dates:
                try:
                    aria_label = hd.get_attribute("aria-label") or ""
                    if "sold out" in aria_label.lower():
                        continue
                    m = re.search(r'Choose \w+,\s+(\w+)\s+(\d+)\w*,\s+(\d+)', aria_label)
                    if not m:
                        continue
                    month_str = m.group(1)
                    day_num   = int(m.group(2))
                    year_num  = int(m.group(3))
                    etd_date  = datetime.strptime(
                        f"{day_num} {month_str} {year_num}", "%d %B %Y"
                    ).date()
                    if etd_date < min_etd_date:
                        print(f"  ⏭️ Bỏ qua {etd_date} (< today+{DATE_OFFSET_DAYS}={min_etd_date})")
                        continue
                    # ── Chỉ lưu aria-label, KHÔNG lưu element ──
                    chosen_aria_label = aria_label
                    chosen_date_str = f"{etd_date.day}-{etd_date.strftime('%b')}"
                    print(f"  ✅ Chọn ngày còn chỗ: {chosen_date_str}")
                    break
                except Exception:
                    continue

            if chosen_aria_label is None:
                # Thử chờ thêm 3s cho tháng sau render giá (async loading)
                print(f"  ⏳ Chưa có ngày hợp lệ, chờ thêm 3s cho tháng sau load giá...")
                time.sleep(3)

                # Tìm lại với broader search
                highlight_dates_2 = driver.find_elements(By.XPATH, HIGHLIGHT_XPATH)
                if len(highlight_dates_2) <= 1:
                    all_en = driver.find_elements(By.XPATH,
                        "//div[@role='option' and @aria-disabled='false' and not(contains(@class,'outside-month')) and not(contains(@class,'disabled'))]"
                    )
                    dp = [d for d in all_en if re.search(r'\d+\.?\d*K', d.text or "")]
                    if dp:
                        highlight_dates_2 = dp
                        print(f"  🔍 Sau 3s: {len(highlight_dates_2)} ngày có giá (broader)")
                    else:
                        print(f"  🔍 Sau 3s: {len(highlight_dates_2)} ngày highlight")
                else:
                    print(f"  🔍 Sau 3s: {len(highlight_dates_2)} ngày highlight")

                for hd in highlight_dates_2:
                    try:
                        aria_label = hd.get_attribute("aria-label") or ""
                        if "sold out" in aria_label.lower():
                            continue
                        m = re.search(r'Choose \w+,\s+(\w+)\s+(\d+)\w*,\s+(\d+)', aria_label)
                        if not m: continue
                        etd_date = datetime.strptime(
                            f"{int(m.group(2))} {m.group(1)} {int(m.group(3))}", "%d %B %Y"
                        ).date()
                        if etd_date < min_etd_date: continue
                        chosen_aria_label = aria_label
                        chosen_date_str = f"{etd_date.day}-{etd_date.strftime('%b')}"
                        print(f"  ✅ Chọn ngày còn chỗ (sau wait): {chosen_date_str}")
                        break
                    except Exception: continue

            # Nếu vẫn chưa có → thử navigate sang tháng sau
            if chosen_aria_label is None:
                print(f"  ⏳ Thử chuyển sang tháng sau trên calendar...")
                next_month_clicked = False
                for next_xp in [
                    "//button[contains(@class,'next') and contains(@class,'month')]",
                    "//button[@aria-label='Next month']",
                    "//span[contains(@class,'next-icon')]/ancestor::button",
                    "//div[contains(@class,'date-picker')]//button[last()]",
                ]:
                    try:
                        next_btn = driver.find_element(By.XPATH, next_xp)
                        next_btn.click()
                        next_month_clicked = True
                        print(f"  ✅ Đã click next month")
                        break
                    except Exception: continue

                if next_month_clicked:
                    time.sleep(3)
                    # DEBUG: dump cells sau next month
                    all_cells_nm = driver.find_elements(By.XPATH,
                        "//div[@role='option' and @aria-disabled='false' and not(contains(@class,'outside-month'))]"
                    )
                    print(f"  [DEBUG] Sau next month: {len(all_cells_nm)} date cells")
                    for i, dc in enumerate(all_cells_nm[:10]):
                        try:
                            cls = dc.get_attribute("class") or ""
                            txt = dc.text.strip().replace('\n', ' ')[:40]
                            print(f"  [DEBUG]  [{i}] class='{cls}' | text='{txt}'")
                        except Exception: pass

                    highlight_dates_3 = driver.find_elements(By.XPATH, HIGHLIGHT_XPATH)
                    if len(highlight_dates_3) <= 1:
                        all_en = driver.find_elements(By.XPATH,
                            "//div[@role='option' and @aria-disabled='false' and not(contains(@class,'outside-month')) and not(contains(@class,'disabled'))]"
                        )
                        dp = [d for d in all_en if re.search(r'\d+\.?\d*K', d.text or "")]
                        if dp:
                            highlight_dates_3 = dp
                            print(f"  🔍 Tháng sau: {len(highlight_dates_3)} ngày có giá (broader)")
                        else:
                            print(f"  🔍 Tháng sau: {len(highlight_dates_3)} ngày highlight")
                    else:
                        print(f"  🔍 Tháng sau: {len(highlight_dates_3)} ngày highlight")

                    for hd in highlight_dates_3:
                        try:
                            aria_label = hd.get_attribute("aria-label") or ""
                            if "sold out" in aria_label.lower(): continue
                            m = re.search(r'Choose \w+,\s+(\w+)\s+(\d+)\w*,\s+(\d+)', aria_label)
                            if not m: continue
                            etd_date = datetime.strptime(
                                f"{int(m.group(2))} {m.group(1)} {int(m.group(3))}", "%d %B %Y"
                            ).date()
                            if etd_date < min_etd_date: continue
                            chosen_aria_label = aria_label
                            chosen_date_str = f"{etd_date.day}-{etd_date.strftime('%b')}"
                            print(f"  ✅ Chọn ngày còn chỗ (tháng sau): {chosen_date_str}")
                            break
                        except Exception: continue

            if chosen_aria_label is None:
                print(f"  ℹ️ Tất cả ngày tím đều Sold out hoặc quá sớm → không có lịch khả dụng")
                try:
                    date_input.send_keys(Keys.ESCAPE)
                except Exception:
                    pass
                calendar_opened = False
                if attempt == 0:
                    print("  🔁 No Schedule lần 1 → click lại ô schedule thêm 1 lần...")
                    time.sleep(1)
                    continue
                break

            # ── Tìm lại element TƯƠI bằng aria-label rồi mới click ──
            escaped_label = chosen_aria_label.replace('"', '\\"')
            fresh_day = driver.find_elements(By.XPATH,
                f'//div[@role="option" and @aria-label="{escaped_label}" and not(contains(@class,"outside-month"))]'
            )
            if not fresh_day:
                print(f"  ⚠️ Không tìm lại được ngày {chosen_date_str}")
                try:
                    date_input.send_keys(Keys.ESCAPE)
                except Exception:
                    pass
                calendar_opened = False
                if attempt == 0:
                    print("  🔁 Schedule stale lần 1 → click lại ô schedule thêm 1 lần...")
                    time.sleep(1)
                    continue
                break

            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", fresh_day[0]
            )
            time.sleep(0.3)
            
            # Dòng Debug: Báo cáo ngày chọn gần nhất trước khi click
            print(f"  🐛 [DEBUG] Chuẩn bị click chọn ngày: {chosen_date_str} (aria-label: {chosen_aria_label})")
            
            # Ưu tiên dùng Selenium click chuẩn để React nhận sự kiện, nếu lỗi mới ép JS
            try:
                fresh_day[0].click()
            except:
                driver.execute_script("arguments[0].click();", fresh_day[0])
            
            # BẮT BUỘC CHỜ 1 giây để React xử lý sự kiện và điền ngày vào ô input
            time.sleep(1)

            # Đợi nút View Quote xuất hiện (nếu có) rồi click
            try:
                view_quote_btn = WebDriverWait(driver, 2).until(
                    EC.element_to_be_clickable((By.XPATH,
                        "//button[contains(text(),'View Quote') or contains(text(),'ViewQuote')]"
                    ))
                )
                driver.execute_script("arguments[0].scrollIntoView({behavior:'smooth',block:'center'});", view_quote_btn)
                time.sleep(0.3)
                driver.execute_script("arguments[0].click();", view_quote_btn)
                print(f"  ✅ Đã click View Quote cho ngày {chosen_date_str}!")
            except TimeoutException:
                print(f"  ℹ️ Không thấy nút View Quote → tiếp tục bình thường")            

            time.sleep(SPEED)
            calendar_opened = True
            progress["date_selected"] = True
            progress["step"] = "date_selected"
            progress["chosen_date"] = chosen_date_str
            progress["updated_at"] = time.time()
            break  # Thành công thì thoát vòng lặp ngay

        except TimeoutException:
            print(f"  ⚠️ Web lag, chưa mở được lịch (thử lại lần {attempt + 1}), chờ 1s...")
            time.sleep(1)

    # Nếu sau 3 lần ráng sức mà vẫn không có -> Trả về No Schedule
    if not calendar_opened:
        print("  ⚠️ Không có lịch tàu khả dụng (không có ngày màu tím)!")
        return {
            "POL": row_data[2], "POD": row_data[3], "Status": "No Schedule",
            "20 DRY": "-", "40 DRY": "-", "40 HC": "-",
            "ETD": "-", "Transit Time": "-", "Valid": "-",
            "Remark": "-", "Free Time": "-",
            "Vessel": "-", "Transshipment": "-"
        }

    # Get Quote
    get_quote_btn = wait.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), 'GetQuote') or contains(text(), 'Get Quote')]")))
    driver.execute_script("arguments[0].scrollIntoView({behavior:'smooth',block:'center'});", get_quote_btn)
    time.sleep(SPEED); js_click(get_quote_btn)
    print("  🎉 Đã bấm Get Quote. Đang đợi card giá...")
    progress["quote_clicked"] = True
    progress["step"] = "quote_clicked"
    progress["updated_at"] = time.time()
    time.sleep(3)

    # Parse cards - CHỈ LƯU DATA THUẦN, KHÔNG LƯU ELEMENT
    cards = driver.find_elements(By.XPATH, "//div[contains(@class, 'NewQuoteSummary_summary-card__')]")
    parsed_cards = []
    for idx, card in enumerate(cards):
        try:
            accept_btn = card.find_element(By.XPATH, ".//button[contains(@class, 'NewQuoteSummary_accept-btn__')]")
            if "Accept" not in accept_btn.text:
                continue
            price_text  = card.find_element(By.XPATH, ".//p[contains(@class, 'FormatMoneyV2_total-price__')]").text
            price_clean = re.sub(r'[^\d.]', '', price_text.replace('\n', '').strip())
            if '.' in price_clean:
                price_clean = price_clean.split('.')[0]
            clean_price = float(price_clean) if price_clean else 0.0

            etd_text     = card.find_element(By.XPATH, ".//div[contains(@class, 'NewQuoteSummary_route-info__')]/div[1]/p[2]").text
            etd_date     = datetime.strptime(etd_text, "%Y-%m-%d").date()
            transit_text = card.find_element(By.XPATH, ".//div[contains(@class, 'NewQuoteSummary_estimate-item__')]/p[1]").text
            transit_time = int(re.search(r'\d+', transit_text).group())

            # Chỉ lưu INDEX của card, không lưu element
            parsed_cards.append({
                "card_index": idx,
                "price": clean_price,
                "etd": etd_date,
                "transit": transit_time
            })
        except:
            continue

    if not parsed_cards:
        print("  ⚠️ Không có thẻ giá Accept nào!")
        return {"POL": row_data[2], "POD": row_data[3], "Status": "No Accept Card"}

    # Lọc ETD
    min_price  = min(c["price"] for c in parsed_cards)
    price_filt = [c for c in parsed_cards if c["price"] <= min_price + 40]
    unique_etd = {}
    for c in price_filt:
        d = c["etd"]
        if d not in unique_etd or c["transit"] < unique_etd[d]["transit"]:
            unique_etd[d] = c
    sorted_cands = sorted(unique_etd.values(), key=lambda x: x["etd"])
    min_etd      = datetime.today().date() + timedelta(days=DATE_OFFSET_DAYS)
    final_sels   = []
    for c in sorted_cands:
        if c["etd"] < min_etd:
            continue
        if not etd_within_max(c["etd"]):
            continue
        if not final_sels:
            final_sels.append(c)
        else:
            if (c["etd"] - final_sels[-1]["etd"]).days >= 2 and \
               (c["etd"] - final_sels[0]["etd"]).days <= 9:
                final_sels.append(c)
        if len(final_sels) == 3:
            break

    if not final_sels:
        print("  ⚠️ Không có lịch tàu nào thoả mãn ETD!")
        return {"POL": row_data[2], "POD": row_data[3], "Status": "No Valid ETD"}

    # Format Excel
    etd_strs = [f"{c['etd'].day}-{c['etd'].strftime('%b')}" for c in final_sels]
    if len(etd_strs) == 1:
        etd_excel = etd_strs[0]
    elif len(etd_strs) == 2:
        etd_excel = f"{etd_strs[0]} & {etd_strs[1]}"
    else:
        month = final_sels[0]["etd"].strftime("%b")
        days  = [str(c["etd"].day) for c in final_sels]
        etd_excel = f"{', '.join(days[:-1])}, {days[-1]}-{month}"

    transits      = [c["transit"] for c in final_sels]
    transit_excel = str(transits[0]) if len(set(transits)) == 1 else f"{transits[0]}-{transits[-1]}"
    valid_excel   = get_valid_date([c["etd"] for c in final_sels])

    # ── TÌM LẠI ELEMENT CỦA target_card BẰNG INDEX (tránh stale) ──
    target_data = final_sels[0]
    target_idx  = target_data["card_index"]

    def get_fresh_cards():
        """Lấy lại toàn bộ card element tươi từ DOM"""
        return driver.find_elements(
            By.XPATH, "//div[contains(@class, 'NewQuoteSummary_summary-card__')]"
        )

    def get_details_btn(card_index):
        """Lấy lại nút Details của card theo index"""
        fresh = get_fresh_cards()
        if card_index < len(fresh):
            return fresh[card_index].find_element(
                By.XPATH, ".//button[contains(@class, 'NewQuoteSummary_breakdown-button__')]"
            )
        return None
    
    # Format Excel
    etd_strs = [f"{c['etd'].day}-{c['etd'].strftime('%b')}" for c in final_sels]
    if len(etd_strs) == 1:
        etd_excel = etd_strs[0]
    elif len(etd_strs) == 2:
        etd_excel = f"{etd_strs[0]} & {etd_strs[1]}"
    else:
        month = final_sels[0]["etd"].strftime("%b")
        days  = [str(c["etd"].day) for c in final_sels]
        etd_excel = f"{', '.join(days[:-1])}, {days[-1]}-{month}"

    transits      = [c["transit"] for c in final_sels]
    transit_excel = str(transits[0]) if len(set(transits)) == 1 else f"{transits[0]}-{transits[-1]}"
    valid_excel   = get_valid_date([c["etd"] for c in final_sels])

    # Bóc phí Details
    # Tìm lại details_btn tươi theo index (tránh stale)
    details_btn = get_details_btn(target_idx)
    if not details_btn:
        print("  ⚠️ Không tìm lại được nút Details!")
        return {"POL": row_data[2], "POD": row_data[3], "Status": "Details btn not found"}

    driver.execute_script(
        "arguments[0].scrollIntoView({behavior:'smooth',block:'center'});", details_btn
    )
    time.sleep(SPEED)
    js_click(details_btn)
    time.sleep(1.5)

    final_prices = {"DRY 20": 0.0, "DRY 40": 0.0, "DRY 40H": 0.0}
    formula_parts = {"DRY 20": [], "DRY 40": [], "DRY 40H": []}
    pod_upper     = str(row_data[3]).strip().upper()
    country_upper = str(row_data[0]).strip().upper()
    china_route = is_china_destination(country_upper, pod_upper)
    has_thc, has_ens_ams, has_ows = False, False, False

    charge_items = driver.find_elements(By.XPATH, "//ul[contains(@class, 'ChargeBreakdownItem_p-sub-detail__') or contains(@class, 'ChargeBreakdownItem_p-sub-detail__LYkLW')]")
    for ul in charge_items:
        try:
            fee_name = ul.find_element(By.XPATH, "./preceding-sibling::div[1]//span[contains(@class, 'ChargeBreakdownItem_p-sub-title__')]").text.strip().lower()
        except: continue
        group = one_selenium_charge_group(ul)
        if one_is_origin_thc_charge(fee_name): has_thc = True
        if "entry summary declaration surcharge" in fee_name: has_ens_ams = True
        if one_is_ows_charge(fee_name): has_ows = True
        if not one_should_include_charge(fee_name, group, include_origin_thc=china_route): continue
        is_discount = "special promotion service" in fee_name
        equipment_amount_found = False
        common_amounts = []
        for line in ul.find_elements(By.XPATH, "./li"):
            text_line = line.text.strip()
            if " x 1" not in text_line: continue
            m = re.search(r'(DRY 40H|DRY 40|DRY 20).*?\((USD|EUR|CHF|VND)\s*([0-9,.]+)\)', text_line, re.I)
            if m:
                equipment_amount_found = True
                currency = m.group(2).upper()
                amount = float(m.group(3).replace(',','')) * get_live_exchange_rate(currency, "USD")
                signed_amount = -amount if is_discount else amount
                final_prices[m.group(1)] += signed_amount
                formula_parts[m.group(1)].append(signed_amount)
                continue

            # Per shipment/B/L charges do not show a DRY equipment label.
            common_match = re.search(r'\((USD|EUR|CHF|VND)\s*([0-9,.]+)\)', text_line, re.I)
            if common_match:
                currency = common_match.group(1).upper()
                amount = float(common_match.group(2).replace(',', '')) * get_live_exchange_rate(currency, "USD")
                common_amounts.append(-amount if is_discount else amount)

        if not equipment_amount_found:
            for signed_amount in common_amounts:
                for equipment_key in final_prices:
                    final_prices[equipment_key] += signed_amount
                    formula_parts[equipment_key].append(signed_amount)

    # Xác định remark phụ dựa theo POD + country
    manifest_code = get_manifest_code(country_upper, pod_upper)
    check_str     = pod_upper + " " + country_upper  # gộp để tìm kiếm

    def is_region(check, keyword_list):
        # Dùng \b để đảm bảo chỉ khớp nguyên 1 từ độc lập, không khớp chuỗi con
        return any(re.search(r'\b' + re.escape(kw) + r'\b', check) for kw in keyword_list)

    remark_str = "SUBJECT TO THC, BILL, SEAL" if (has_thc and not china_route) else "INCLUDED O.THC, SUBJECT TO BILL, SEAL"

    remark_str = apply_manifest_rule(remark_str, country_upper, pod_upper)
    if has_ens_ams and not manifest_code:
        # Trường hợp ngoài các nước trên nhưng web có fee ENS/AMS → thêm AMS
        remark_str += ", AMS"

    if has_ows:
        remark_str += ", OWS"

    # ── FREE TIME: tìm lại element tươi theo card_index ──
    freetime_result = "N/A"
    try:
        # Lấy lại card tươi theo index
        fresh_cards = get_fresh_cards()
        if target_idx < len(fresh_cards):
            fresh_card_el = fresh_cards[target_idx]
        else:
            fresh_card_el = None
            print("  ⚠️ Không tìm lại được card cho free time")

        if fresh_card_el:
            # Cuộn đến card
            driver.execute_script(
                "arguments[0].scrollIntoView({behavior:'smooth', block:'center'});",
                fresh_card_el
            )
            time.sleep(0.2)

            # Tìm nút free time BÊN TRONG card tươi
            free_time_btn = driver.execute_script("""
                var card = arguments[0];
                return card.querySelector('[data-aoq-v2-free-time-tag]')
                    || card.querySelector('[class*="free-time"]')
                    || card.querySelector('[class*="FreeTime"]')
                    || card.querySelector('[class*="ChipsPopover"]');
            """, fresh_card_el)

            if free_time_btn:
                time.sleep(0.2)
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", free_time_btn
                )
                time.sleep(0.2)
                # Click MỘT LẦN duy nhất bằng JS
                driver.execute_script("arguments[0].click();", free_time_btn)

                # Đợi panel FreeTimeInfor xuất hiện
                try:
                    WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH,
                            "//div[contains(@class,'FreeTimeInfor_body')]"
                        ))
                    )
                    time.sleep(0.4)

                    # Đọc đúng section DESTINATION
                    col_elements = driver.find_elements(By.XPATH,
                        "//div[contains(@class,'FreeTimeInfor_col-element')]"
                    )
                    dest_section = None
                    for col in col_elements:
                        try:
                            title = col.find_element(By.XPATH,
                                ".//p[contains(@class,'FreeTimeInfor_title')]"
                            ).text.strip().upper()
                            if title == "DESTINATION":
                                dest_section = col
                                break
                        except:
                            continue

                    if dest_section:
                        dest_text = dest_section.text.upper()
                        print(f"  🔍 Destination free time text: {dest_text}")

                        freetime_result = parse_one_free_time_text(dest_text) or "Xem thủ công"
                    else:
                        print("  ⚠️ Không tìm thấy section Destination")
                        panel_text = driver.find_element(By.XPATH,
                            "//div[contains(@class,'FreeTimeInfor_body')]").text
                        freetime_result = parse_one_free_time_text(panel_text) or freetime_result

                    print(f"  ⏱️ Free time (Destination): {freetime_result}")

                except TimeoutException:
                    print("  ⚠️ Panel free time không xuất hiện sau khi click")

                # Đóng panel
                ActionChains(driver).send_keys(Keys.ESCAPE).perform()
                time.sleep(0.4)

            else:
                print("  ⚠️ Không tìm thấy nút free time trong card")

    except Exception as e:
        print(f"  ⚠️ Lỗi khi lấy free time: {e}")
        try:
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        except:
            pass
        time.sleep(0.3)

        # Tìm nút free time nằm trong CÙNG card với target_card["btn"]
        # Đi lên ancestor summary-card rồi tìm xuống
        free_time_btn = driver.execute_script("""
            var card = arguments[0];
            return card.querySelector('[data-aoq-v2-free-time-tag]')
                || card.querySelector('[class*="free-time"]')
                || card.querySelector('[class*="FreeTime"]')
                || card.querySelector('[class*="ChipsPopover"]');
        """, fresh_card_el)

        if free_time_btn:
            time.sleep(0.2)
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", free_time_btn
            )
            time.sleep(0.2)
            # Click MỘT LẦN duy nhất
            driver.execute_script("arguments[0].click();", free_time_btn)

            # Đợi panel FreeTimeInfor xuất hiện
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH,
                        "//div[contains(@class,'FreeTimeInfor_body')]"
                    ))
                )
                time.sleep(0.4)

                # ── ĐỌC ĐÚNG SECTION DESTINATION ──
                # Dựa theo HTML: FreeTimeInfor_col-element chứa title "Destination"
                dest_section = None
                col_elements = driver.find_elements(By.XPATH,
                    "//div[contains(@class,'FreeTimeInfor_col-element')]"
                )
                for col in col_elements:
                    try:
                        title = col.find_element(By.XPATH,
                            ".//p[contains(@class,'FreeTimeInfor_title')]"
                        ).text.strip().upper()
                        if title == "DESTINATION":
                            dest_section = col
                            break
                    except:
                        continue

                if dest_section:
                    dest_text = dest_section.text.upper()
                    print(f"  🔍 Destination free time text: {dest_text}")

                    freetime_result = parse_one_free_time_text(dest_text) or "Xem thủ công"
                else:
                    print("  ⚠️ Không tìm thấy section Destination trong free time panel")
                    panel_text = driver.find_element(By.XPATH,
                        "//div[contains(@class,'FreeTimeInfor_body')]").text
                    freetime_result = parse_one_free_time_text(panel_text) or freetime_result

                print(f"  ⏱️ Free time (Destination): {freetime_result}")

            except TimeoutException:
                print("  ⚠️ Panel free time không xuất hiện sau khi click")

            # Đóng panel bằng ESC
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            time.sleep(0.4)

        else:
            print("  ⚠️ Không tìm thấy nút free time trong card mục tiêu")

    except Exception as e:
        print(f"  ⚠️ Lỗi khi lấy free time: {e}")
        try:
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        except:
            pass
        time.sleep(0.3)        

    # Tàu & Transshipment
    schedule_items = driver.find_elements(By.XPATH, "//li[contains(@class, 'ScheduleDepartureItem_li-boat__') or contains(@class, 'ScheduleArrival_li-location__')]")
    ports_in_route, vessels_dict = [], {}
    for item in schedule_items:
        try:
            pname = item.find_element(By.XPATH, ".//div[contains(@class, 'ScheduleItemDetails_title__')]").text.split('(')[0].strip()
            ports_in_route.append(pname)
            try: vessels_dict[pname] = item.find_element(By.XPATH, ".//span[contains(@class, 'ScheduleDepartureItem_transport-name-text__')]").text.strip()
            except: pass
        except: continue

    vessel_final, ts_ports = "", []
    if "HO CHI MINH" in pol_name or "CAI MEP" in pol_name:
        vessel_final = vessels_dict.get("CAI MEP", vessels_dict.get("HO CHI MINH", "TBA"))
        ts_ports = [p for p in ports_in_route[1:-1] if p != "CAI MEP"]
    else:
        vessel_final = vessels_dict.get(pol_name, "TBA")
        ts_ports = list(ports_in_route[1:-1])

    ts_excel     = " + ".join(ts_ports) if ts_ports else "DIRECT"
    vessel_excel = f"{vessel_final} / ETD: {target_data['etd'].day}-{target_data['etd'].strftime('%b')} / Transit time: {target_data['transit']} Days / Transshipment Port: {ts_excel}"

    print(f"  ✅ Xong: {row_data[2]} → {row_data[3]}")
    print(f"  💰 DRY20:{final_prices['DRY 20']:,.0f} | DRY40:{final_prices['DRY 40']:,.0f} | DRY40H:{final_prices['DRY 40H']:,.0f} USD")
    print(f"  📝 {remark_str} | Free: {freetime_result}")
    print(f"  🚢 {vessel_excel} | ⚓ {ts_excel}")

    return {
        "POL": row_data[2], "POD": row_data[3],
        "20 DRY": final_prices["DRY 20"], "40 DRY": final_prices["DRY 40"], "40 HC": final_prices["DRY 40H"],
        "20 FORMULA": _excel_formula_from_parts(formula_parts["DRY 20"]),
        "40 FORMULA": _excel_formula_from_parts(formula_parts["DRY 40"]),
        "40HC FORMULA": _excel_formula_from_parts(formula_parts["DRY 40H"]),
        "ETD": etd_excel, "Transit Time": transit_excel, "Valid": valid_excel,
        "Remark": remark_str, "Free Time": freetime_result,
        "Vessel": vessel_excel, "Transshipment": ts_excel
    }


# ==========================================
# HÀM GHI KẾT QUẢ VÀO EXCEL (dùng để gọi sau MỖI batch → không mất dữ liệu)
# ==========================================
def save_results_to_excel(results, file_name=None):
    if file_name is None:
        file_name = os.environ.get("EXCEL_PATH", "input_gia.xlsx")
    """
    Ghi toàn bộ results vào file Excel. Gọi sau mỗi batch để bảo vệ
    dữ liệu khỏi mất nếu Edge/Selenium crash.
    """
    if not results:
        print("  ℹ️ Chưa có kết quả nào để ghi.")
        return
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.worksheets[0]

        for res in results:
            idx = res.get("orig_index")
            if idx is None:
                continue
            excel_row = idx + 2

            status   = res.get("Status", "")
            c_20     = res.get("20 DRY", "-")
            c_40     = res.get("40 DRY", "-")
            c_40hc   = res.get("40 HC", "-")
            f_20     = res.get("20 FORMULA")
            f_40     = res.get("40 FORMULA")
            f_40hc   = res.get("40HC FORMULA")
            etd      = res.get("ETD", "-")
            tt       = res.get("Transit Time", "-")
            valid    = res.get("Valid", "-")
            remark   = res.get("Remark", "-")
            free_tm  = res.get("Free Time", "-")
            vessel   = res.get("Vessel", "-")
            trans    = res.get("Transshipment", "-")

            if status and status not in ["OK", "-", ""]:
                remark = status if remark == "-" else f"{status} | {remark}"

            ws.cell(row=excel_row, column=6).value  = f_20 or c_20
            ws.cell(row=excel_row, column=7).value  = f_40 or c_40
            ws.cell(row=excel_row, column=8).value  = f_40hc or c_40hc
            ws.cell(row=excel_row, column=9).value  = etd
            ws.cell(row=excel_row, column=10).value = tt
            ws.cell(row=excel_row, column=11).value = valid
            ws.cell(row=excel_row, column=13).value = remark
            ws.cell(row=excel_row, column=14).value = free_tm
            ws.cell(row=excel_row, column=15).value = vessel
            ws.cell(row=excel_row, column=16).value = trans

        try:
            wb.save(file_name)
            print(f"  💾 Đã lưu {len(results)} dòng vào {file_name}")
        except PermissionError:
            backup_name = "input_gia_KETQUA.xlsx"
            wb.save(backup_name)
            print(f"  ⚠️ File '{file_name}' đang mở trong Excel → lưu vào '{backup_name}'")
    except Exception as e:
        print(f"  ❌ Lỗi ghi Excel: {e}")
        # Dự phòng cuối cùng: cố gắng dump CSV
        try:
            import csv
            csv_name = "input_gia_BACKUP.csv"
            with open(csv_name, "w", newline="", encoding="utf-8-sig") as f:
                if results:
                    w = csv.DictWriter(f, fieldnames=list(results[0].keys()))
                    w.writeheader()
                    w.writerows(results)
            print(f"  💾 Backup CSV: {csv_name}")
        except Exception as e2:
            print(f"  ❌ Lỗi dự phòng CSV: {e2}")


# ==========================================
# MAIN: ĐỌC EXCEL → CHẠY TRICK 10 TAB + RELOAD ĐỒNG LOẠT
# ==========================================
results_list    = []
original_window = driver.current_window_handle

print("📂 Đang đọc file Excel...")
_ONE_EXCEL = os.environ.get("EXCEL_PATH", "input_gia.xlsx")
FILTER_POL = os.environ.get("FILTER_POL", "").strip().upper()
FILTER_POD = os.environ.get("FILTER_POD", "").strip().upper()
SINGLE_ROW = os.environ.get("SINGLE_ROW", "").strip()
import openpyxl
_wb_temp = openpyxl.load_workbook(_ONE_EXCEL, read_only=True)
_active_sheet = _wb_temp.active.title
_wb_temp.close()
df = pd.read_excel(_ONE_EXCEL, sheet_name=_active_sheet)

# 🟢 Tạo bản sao và lưu lại index gốc để lát map vào đúng dòng trong Excel
df_temp = df.copy()
df_temp['orig_index'] = df_temp.index

# Lọc chỉ lấy dòng có cột E (index 4) chứa chữ "ONE"
df_filtered = df_temp[
    df_temp.iloc[:, 4].astype(str).str.upper().str.contains("ONE", na=False)
]
# Lọc theo FILTER_POL/FILTER_POD nếu có (mode_route)
if FILTER_POL:
    df_filtered = df_filtered[df_filtered.iloc[:, 2].astype(str).str.upper().str.strip() == FILTER_POL]
if FILTER_POD:
    df_filtered = df_filtered[df_filtered.iloc[:, 3].astype(str).str.upper().str.strip() == FILTER_POD]
if SINGLE_ROW:
    try:
        target_excel_row = int(SINGLE_ROW)
        target_orig_index = target_excel_row - 2
        df_filtered = df_filtered[df_filtered["orig_index"] == target_orig_index]
        print(f"[SINGLE_ROW] Chi chay dong {target_excel_row} theo lenh tu main.py")
    except Exception:
        print(f"SINGLE_ROW khong hop le: {SINGLE_ROW}")
raw_data = df_filtered.dropna(
    subset=[df_filtered.columns[2], df_filtered.columns[3]]
).values.tolist()

# Port mapping: Excel name → ONE search name
ONE_PORT_MAPPING = {
    "TIANJIN": "XINGANG",
    "FOS SUR MER": "FOS-SUR-MER",
    "VENICE": "VENEZIA",
}
# Áp dụng mapping cho POD (row[3]) trong raw_data — giữ tên gốc trong Excel output
for _rd in raw_data:
    _pod_up = str(_rd[3]).strip().upper()
    if _pod_up in ONE_PORT_MAPPING:
        _rd[3] = ONE_PORT_MAPPING[_pod_up]

print(f"📋 Sau khi lọc hãng ONE: {len(raw_data)} tuyến (bỏ qua các hãng khác)")
BATCH_SIZE    = 10 # Số tab tối đa mỗi batch (có thể điều chỉnh tuỳ theo hiệu năng máy và web)
total_batches = math.ceil(len(raw_data) / BATCH_SIZE)
print(f"Tổng cộng {len(raw_data)} tuyến → {total_batches} lượt (tối đa 10 tab/lượt)")

tabs = []  # Giữ xuyên suốt tất cả batch, chỉ đóng sau batch cuối
ONE_API_FALLBACK_SELENIUM_MODE = False

if ONE_USE_API and raw_data:
    print("\n[ONE-API] Bat che do API-first. Row nao API loi moi fallback Selenium.")
    api_fallback_rows = []
    try:
        driver.get(ONE_URL)
        ensure_current_tab_ready(timeout=20)
    except Exception as e:
        print(f"[ONE-API] Khong chuan bi duoc tab/token ONE: {e}")
    for api_idx, row in enumerate(raw_data, start=1):
        orig_index = row[-1]
        print(f"\n[ONE-API {api_idx}/{len(raw_data)}] {row[2]} -> {row[3]}")
        try:
            result = scrape_one_api(row)
            result["orig_index"] = orig_index
            results_list.append(result)
            save_results_to_excel(results_list)
        except Exception as e:
            print(f"[ONE-API] Row {orig_index + 2} loi API: {e}")
            if ONE_API_FALLBACK_SELENIUM:
                api_fallback_rows.append(row)
            else:
                results_list.append({
                    "POL": row[2], "POD": row[3], "orig_index": orig_index,
                    "Status": f"Loi ONE API: {e}",
                    "20 DRY": "-", "40 DRY": "-", "40 HC": "-",
                    "ETD": "-", "Transit Time": "-", "Valid": "-",
                    "Remark": "-", "Free Time": "-", "Vessel": "-", "Transshipment": "-"
                })
                save_results_to_excel(results_list)
    if not api_fallback_rows:
        print("\n[ONE-API] Hoan tat toan bo bang API.")
        save_results_to_excel(results_list)
        sys.exit(0)
    print(f"\n[ONE-API] Co {len(api_fallback_rows)} row can fallback Selenium.")
    raw_data = api_fallback_rows
    total_batches = math.ceil(len(raw_data) / BATCH_SIZE)
    ONE_API_FALLBACK_SELENIUM_MODE = True
    tabs = []

for batch_idx in range(total_batches):
    start_idx  = batch_idx * BATCH_SIZE
    end_idx    = start_idx + BATCH_SIZE
    batch_data = raw_data[start_idx:end_idx]
    is_last    = (batch_idx == total_batches - 1)

    print(f"\n{'='*54}")
    print(f"🚀 LƯỢT CHẠY {batch_idx+1}/{total_batches}  ({len(batch_data)} tuyến)")
    print(f"{'='*54}")
    route_setup_errors = {}

    # -------------------------------------------------------
    # BƯỚC 1: KIỂM TRA TAB HIỆN CÓ, MATCHING ROUTE VÀ NHẬP POL/POD
    # -------------------------------------------------------
    if batch_idx == 0 and ONE_API_FALLBACK_SELENIUM_MODE:
        print("📂 ONE-API fallback: bỏ qua quét tab cũ, tự mở tab sạch cho Selenium fallback...")
        tabs = []
        for row in batch_data:
            target_pol = str(row[2]).strip().upper()
            target_pod = str(row[3]).strip().upper()
            target_country = str(row[0]).strip()
            setup_idx = len(tabs)
            print(f"\n🔁 Fallback Selenium tab mới: {target_pol} → {target_pod}")
            try:
                handle = open_tab_and_ensure_ready()
                tabs.append(handle)
                fill_tab_ports(row)
                time.sleep(0.6)
                if has_no_port_pair_error():
                    print(f"  🔁 Web báo No port pair cho {target_pol} → {target_pod}; retry POD bằng query rút ngắn...")
                    select_port(2, target_pod, target_country, allow_trim=True)
                    time.sleep(0.8)
                if has_no_port_pair_error():
                    print(f"  ⚠️ Web báo No port pair cho {target_pol} → {target_pod}. Bỏ qua row này.")
                    route_setup_errors[len(tabs) - 1] = "No port pair"
            except Exception as e:
                print(f"  ❌ Lỗi chuẩn bị tab fallback {target_pol} → {target_pod}: {e}")
                if len(tabs) == setup_idx:
                    try:
                        tabs.append(driver.current_window_handle)
                    except Exception:
                        tabs.append(None)
                route_setup_errors[setup_idx] = str(e)

    elif batch_idx == 0:
        print("📂 Batch đầu: Đọc các tab hiện có để matching tuyến...")
        existing_handles = driver.window_handles
        
        # Bước 1.1: Đọc POL/POD đang có sẵn trên từng tab
        tab_routes = {} # Lưu {handle: {"POL": pol_name, "POD": pod_name}}
        print(f"  🔍 Đang quét {len(existing_handles)} tab đang mở...")
        
        for handle in existing_handles:
            try:
                switch_to_one_handle(handle)
            except Exception:
                continue
            # Chỉ xử lý các tab đang ở trang ONE
            if ONE_URL in driver.current_url:
                try:
                    # Lấy text trong ô POL
                    pol_val = driver.find_element(By.XPATH, "(//input[@placeholder='Please search location'])[1]").get_attribute("value") or ""
                    # Lấy text trong ô POD
                    pod_val = driver.find_element(By.XPATH, "(//input[@placeholder='Please search location'])[2]").get_attribute("value") or ""
                    
                    # Cắt chuỗi lấy phần tên cảng trước dấu phẩy hoặc khoảng trắng (nếu có)
                    pol_clean = pol_val.split(',')[0].strip().upper() if pol_val else ""
                    pod_clean = pod_val.split(',')[0].strip().upper() if pod_val else ""
                    
                    if pol_clean and pod_clean:
                        tab_routes[handle] = {"POL": pol_clean, "POD": pod_clean}
                        print(f"    - Tab {handle[:8]}... có sẵn: {pol_clean} → {pod_clean}")
                except Exception:
                    pass # Tab này có thể chưa load xong form, bỏ qua

        # Bước 1.2: Gán tab cho các tuyến trong Excel
        unused_handles = list(existing_handles) # Các tab chưa được dùng
        for row in batch_data:
            target_pol = str(row[2]).strip().upper()
            target_pod = str(row[3]).strip().upper()
            target_country = str(row[0]).strip()
            
            matched_handle = None
            
            # Tìm xem có tab nào khớp POL và POD không
            for handle, route in tab_routes.items():
                if handle in unused_handles and target_pol in route["POL"] and target_pod in route["POD"]:
                    matched_handle = handle
                    break
            
            if matched_handle:
                print(f"\n✅ Đã match tab cho: {target_pol} → {target_pod}")
                print("  🔄 Reuse tab cũ nhưng reset form để tránh state dở sau lần chạy bị ngắt...")
                try:
                    switch_to_one_handle(matched_handle)
                    driver.get(ONE_URL)
                    if ensure_current_tab_ready(timeout=20):
                        fill_tab_ports(row)
                    else:
                        route_setup_errors[len(tabs)] = "Không reset được tab reuse về form ONE"
                except Exception as e:
                    print(f"  ⚠️ Reset tab reuse lỗi: {e}")
                    route_setup_errors[len(tabs)] = f"Reset tab reuse lỗi: {e}"
                tabs.append(matched_handle)
                unused_handles.remove(matched_handle)
            else:
                # Nếu không match được, lấy một tab trống/thừa hoặc mở tab mới để nhập
                print(f"\n⚠️ Chưa có tab cho: {target_pol} → {target_pod}. Chuẩn bị nhập mới...")
                if unused_handles:
                    handle = unused_handles.pop(0)
                    try:
                        switch_to_one_handle(handle)
                        driver.get(ONE_URL)
                        ensure_current_tab_ready(timeout=20)
                    except Exception as dead_e:
                        print(f"  ⚠️ Tab dư đã chết ({type(dead_e).__name__}) -> mở tab mới")
                        handle = open_tab_and_ensure_ready()
                    tabs.append(handle)
                else:
                    handle = open_tab_and_ensure_ready()
                    tabs.append(handle)
                
                # Gọi hàm select_port cho tab này
                # (đã có try/except bên trong select_port, nhưng ta bọc thêm để tránh văng lỗi cả hệ thống)
                try:
                    select_port(1, target_pol, target_country)
                    select_port(2, target_pod, target_country)
                    time.sleep(0.6)
                    if has_no_port_pair_error():
                        print(f"  🔁 Web báo No port pair cho {target_pol} → {target_pod}; retry POD bằng query rút ngắn...")
                        select_port(2, target_pod, target_country, allow_trim=True)
                        time.sleep(0.8)
                    if has_no_port_pair_error():
                        print(f"  ⚠️ Web báo No port pair cho {target_pol} → {target_pod}. Bỏ qua row này.")
                        route_setup_errors[len(tabs) - 1] = "No port pair"
                except Exception as e:
                    print(f"  ❌ Lỗi khi nhập cảng {target_pol} → {target_pod}: {e}")
                    route_setup_errors[len(tabs) - 1] = str(e)

    # -------------------------------------------------------
    # BƯỚC 2: SWITCH TAB LIÊN TỤC → ĐỢI HẾT POPUP LOADING
    # -------------------------------------------------------
    switch_tab_trick_until_clear(tabs[:len(batch_data)], rounds=5, pause=0.2)

    # -------------------------------------------------------
    # BƯỚC 3: TỪNG TAB → NHẬP CONT + GET QUOTE + BÓC GIÁ (VỚI HÀNG ĐỢI RETRY)
    # -------------------------------------------------------
    print("\n▶️  Bắt đầu nhập cont và bóc giá từng tab...")
    
    pending_tabs = list(range(len(batch_data))) # Hàng đợi các tab cần xử lý
    retries_count = {i: 0 for i in pending_tabs}
    commodity_retry_count = {i: 0 for i in pending_tabs}
    icon_add_retry_count = {i: 0 for i in pending_tabs}
    MAX_RETRIES = ONE_PIPELINE_MAX_RETRIES # Tổng lượt quay lại; mỗi lượt wait ngắn để còn xử lý tab khác

    while pending_tabs:
        i = pending_tabs.pop(0) # Lấy tab đầu tiên trong hàng đợi ra làm
        row_data = batch_data[i]
        orig_index = row_data[-1] 
        try:
            switch_to_one_handle(tabs[i])
        except Exception as dead_e:
            try:
                replace_dead_one_tab(i, row_data, type(dead_e).__name__)
            except Exception as repl_e:
                print(f"  ❌ Không thay được ONE tab {i+1}: {repl_e}")
                results_list.append({
                    "POL": row_data[2], "POD": row_data[3],
                    "orig_index": orig_index, "Status": f"Lỗi: tab ONE bị đóng ({type(dead_e).__name__})"
                })
                continue
        if i in route_setup_errors:
            setup_error = str(route_setup_errors[i])
            if setup_error == "No port pair":
                result = one_no_port_pair_result(row_data)
                result["orig_index"] = orig_index
                results_list.append(result)
            else:
                results_list.append({
                    "POL": row_data[2], "POD": row_data[3],
                    "orig_index": orig_index, "Status": f"Lỗi nhập cảng: {setup_error}"
                })
            try:
                ONE_TAB_PROGRESS.pop(tabs[i], None)
                driver.get(ONE_URL)
            except Exception:
                pass
            continue
        
        print(f"\n[Tab {i+1}/{len(batch_data)}] {row_data[2]} → {row_data[3]} (Lần thử: {retries_count[i] + 1})")
        try:
            result = scrape_tab(row_data)
            result["orig_index"] = orig_index 
            results_list.append(result)
            try:
                ONE_TAB_PROGRESS.pop(tabs[i], None)
                driver.get(ONE_URL)
                print(f"  ↩️  Reload tab {i+1} về form ONE để chuẩn bị row/batch khác.")
            except Exception as reload_e:
                print(f"  ⚠️ Reload tab {i+1} sau khi đọc giá lỗi: {reload_e}")
        except Exception as e:
            error_msg = str(e)
            if "ONE_COMMODITY_MISMATCH_RETRY" in error_msg:
                if commodity_retry_count[i] < 1:
                    print(f"  🔁 Commodity không match. Reset tab và chạy lại row {orig_index + 2} lần 2...")
                    commodity_retry_count[i] += 1
                    try:
                        reset_current_tab_for_row(row_data)
                        switch_tab_trick_until_clear([tabs[i]], rounds=3, pause=0.2)
                    except Exception as reset_e:
                        print(f"  ⚠️ Reset row {orig_index + 2} lỗi: {reset_e}")
                    pending_tabs.append(i)
                else:
                    print(f"  ❌ Row {orig_index + 2} vẫn lỗi commodity sau lần chạy lại.")
                    results_list.append({
                        "POL": row_data[2], "POD": row_data[3],
                        "orig_index": orig_index, "Status": "Lỗi: Commodity không match sau khi chạy lại"
                    })
            elif "icon-add" in error_msg:
                if icon_add_retry_count[i] < 1:
                    print(f"  🔁 Lỗi icon-add. Reset tab và retry row {orig_index + 2} thêm 1 lần...")
                    icon_add_retry_count[i] += 1
                    try:
                        reset_current_tab_for_row(row_data)
                        switch_tab_trick_until_clear([tabs[i]], rounds=3, pause=0.2)
                    except Exception as reset_e:
                        print(f"  ⚠️ Reset row {orig_index + 2} lỗi: {reset_e}")
                    pending_tabs.append(i)
                else:
                    print(f"  ❌ Row {orig_index + 2} vẫn lỗi icon-add sau retry.")
                    results_list.append({
                        "POL": row_data[2], "POD": row_data[3],
                        "orig_index": orig_index, "Status": f"Lỗi: {error_msg}"
                    })
            elif "WEB_LAG_RETRY" in error_msg:
                if retries_count[i] < MAX_RETRIES:
                    print(f"  ⏳ Tab {i+1} đang chờ web render ({error_msg}). Chuyển tab khác rồi quay lại.")
                    retries_count[i] += 1
                    if retries_count[i] == max(6, MAX_RETRIES // 2):
                        print(f"  🔄 Tab {i+1} chờ quá lâu. Reset form và nhập lại row {orig_index + 2} trước khi retry tiếp...")
                        try:
                            ONE_TAB_PROGRESS.pop(tabs[i], None)
                            reset_current_tab_for_row(row_data)
                            switch_tab_trick_until_clear([tabs[i]], rounds=3, pause=0.2)
                        except Exception as reset_e:
                            print(f"  ⚠️ Reset row {orig_index + 2} lỗi: {reset_e}")
                    pending_tabs.append(i) # Đẩy tab này xuống cuối hàng đợi
                else:
                    print(f"  ❌ Đã quay lại {MAX_RETRIES} lượt nhưng tab {i+1} vẫn lag. Bỏ qua hoàn toàn tab này.")
                    results_list.append({
                        "POL": row_data[2], "POD": row_data[3], 
                        "orig_index": orig_index, "Status": "Lỗi: Lịch/Commodity load quá chậm"
                    })
            else:
                import traceback
                print(f"  ❌ Lỗi cứng tab {i+1}: {e}")
                print(traceback.format_exc())
                capture_error_artifacts(f"row_{orig_index + 2}_{row_data[2]}_{row_data[3]}_hard_error")
                results_list.append({
                    "POL": row_data[2], "POD": row_data[3], 
                    "orig_index": orig_index, "Status": f"Lỗi: {error_msg}"
                })

    # -------------------------------------------------------
    # BƯỚC 3.5: GHI KẾT QUẢ VÀO EXCEL NGAY SAU MỖI BATCH (FIX: tránh mất dữ liệu)
    # -------------------------------------------------------
    print(f"\n💾 Lưu Excel tạm sau batch {batch_idx+1}/{total_batches}...")
    save_results_to_excel(results_list)

    # -------------------------------------------------------
    # BƯỚC 4: SAU KHI BÓC XONG BATCH
    # -------------------------------------------------------
    if not is_last:
        next_batch = raw_data[end_idx: end_idx + BATCH_SIZE]

        reload_all_tabs_simultaneously(tabs[:len(batch_data)])

        print(f"\n📝 Nhập liệu cho batch {batch_idx+2} ({len(next_batch)} tuyến)...")
        for i, row in enumerate(next_batch):
            try:
                switch_to_one_handle(tabs[i])
            except Exception as dead_e:
                replace_dead_one_tab(i, None, type(dead_e).__name__)
            ONE_TAB_PROGRESS.pop(tabs[i], None)
            print(f"  Tab {i+1}: {str(row[2]).strip()} → {str(row[3]).strip()}")
            try:
                fill_tab_ports(row)
            except Exception as e:
                print(f"  ⚠️ Tab {i+1} nhập lỗi: {e}")

        print("\n🔄 Switch tab đợi hết popup loading cho batch tiếp...")
        switch_tab_trick_until_clear(tabs[:len(next_batch)], rounds=5, pause=0.2)
        print(f"✅ Batch {batch_idx+2} sẵn sàng!")

    else:
        print("\n🧹 Batch cuối xong. Giữ nguyên trình duyệt và toàn bộ tabs theo yêu cầu...")


# ==========================================
# XUẤT EXCEL CUỐI CÙNG (lưu tổng hợp sau khi xong tất cả batch)
# ==========================================
print(f"\n🎉 ĐÃ CHẠY XONG! Tổng {len(results_list)} kết quả. Ghi file Excel lần cuối...")
save_results_to_excel(results_list)

# Lệnh này giúp cửa sổ console không bị tự tắt ngay lập tức
if not os.environ.get("EXCEL_PATH"):
    input("\nBấm nút Enter để thoát chương trình...")
