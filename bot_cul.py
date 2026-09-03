"""
CUL schedule checker.

Reads input_gia.xlsx, filters CUL rows, calls the CULines point-to-point
schedule endpoints, then writes schedule fields in the same columns used by
the other schedule-only bots:
  I: ETD, J: transit time, O: vessel detail, P: transshipment port.
"""

import json
import math
import os
import re
import sys
import time
from datetime import datetime, timedelta

import openpyxl
import requests
from openpyxl.styles import Alignment
from bot_cli import parse_date_offset_days, etd_within_max, max_etd_date, max_etd_date_only


try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


EXCEL_PATH = os.environ.get("EXCEL_PATH") or os.path.join(os.getcwd(), "input_gia.xlsx")
FILTER_POL = (os.environ.get("FILTER_POL") or "").strip().upper()
FILTER_POD = (os.environ.get("FILTER_POD") or "").strip().upper()
FILTER_COUNTRY = (os.environ.get("FILTER_COUNTRY") or "").strip().upper()
SINGLE_ROW = (os.environ.get("SINGLE_ROW") or "").strip()
DATE_OFFSET_DAYS = parse_date_offset_days()

CARRIER_TARGETS = {"CUL", "CU LINE", "CU LINES", "CULINES", "CHINA UNITED LINES"}
BASE = "https://eservice.culines.com/gnoss"
HOME_URL = "https://eservice.culines.com/gnoss/index.jsp"
PORT_SEARCH_URL = f"{BASE}/CUP_HOM_3000GS.do?is_ajax=Y"
SCHEDULE_URL = f"{BASE}/CUP_HOM_3001GS.do"
TODAY = datetime.now()

PORT_ALIASES = {
    "HO CHI MINH": ["HO CHI MINH", "HOCHIMINH", "SAIGON"],
    "HOCHIMINH": ["HO CHI MINH", "HOCHIMINH", "SAIGON"],
    "HCM": ["HO CHI MINH", "HOCHIMINH", "SAIGON"],
    "SAIGON": ["HO CHI MINH", "HOCHIMINH", "SAIGON"],
    "HAI PHONG": ["HAIPHONG", "HAI PHONG"],
    "HAIPHONG": ["HAIPHONG", "HAI PHONG"],
    "PORT KLANG": ["PORT KLANG", "PORT KELANG"],
    "PORTKLANG": ["PORT KLANG", "PORT KELANG"],
    "NHAVA SHEVA": ["NHAVA SHEVA", "JAWAHARLAL NEHRU"],
    "NHAVASHEVA": ["NHAVA SHEVA", "JAWAHARLAL NEHRU"],
    "JEBEL ALI": ["JEBEL ALI", "JEBELALI"],
    "JEBELALI": ["JEBEL ALI", "JEBELALI"],
}

MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


def log(msg):
    print(f"[CUL] {msg}")


def norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip()).upper()


def parse_valid_date(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if hasattr(raw, "year") and hasattr(raw, "month") and hasattr(raw, "day"):
        return datetime(raw.year, raw.month, raw.day)

    s = str(raw).strip()
    if not s:
        return None

    m = re.fullmatch(r"(\d{1,2})-([A-Za-z]{3})", s)
    if m:
        mon = MONTHS.get(m.group(2).upper())
        if mon:
            return datetime(TODAY.year, mon, int(m.group(1)))

    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})", s)
    if m:
        return datetime(TODAY.year, int(m.group(2)), int(m.group(1)))

    for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d", "%d %b %Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            if dt.year == 1900:
                dt = dt.replace(year=TODAY.year)
            return dt
        except Exception:
            pass

    m = re.match(r"(\d{1,2})[-/\s]+([A-Za-z]{3})", s)
    if m:
        mon = MONTHS.get(m.group(2).upper())
        if mon:
            return datetime(TODAY.year, mon, int(m.group(1)))
    return None


def compute_valid_window(valid_dt):
    if not valid_dt:
        return None, None
    start = (TODAY + timedelta(days=DATE_OFFSET_DAYS)).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    end = valid_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    return start, end


def make_session():
    s = requests.Session()
    s.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0"
            ),
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Accept-Language": "en-US,en;q=0.9",
            "Origin": "https://eservice.culines.com",
            "Referer": HOME_URL,
            "X-Requested-With": "XMLHttpRequest",
        }
    )
    try:
        s.get(f"{BASE}/CUP_HOM_3000.do?sessLocale=en", timeout=25)
    except Exception:
        pass
    return s


def load_json_response(resp):
    text = resp.text.strip()
    try:
        return resp.json()
    except Exception:
        pass
    try:
        return json.loads(text)
    except Exception:
        m = re.search(r"(\{.*\})", text, re.S)
        if m:
            return json.loads(m.group(1))
    raise ValueError(f"Cannot parse CUL JSON: HTTP {resp.status_code} {text[:200]}")


def alias_candidates(name):
    key = norm(name)
    candidates = PORT_ALIASES.get(key, [name])
    out = []
    for item in candidates + [name]:
        item = str(item or "").strip()
        if item and item.upper() not in {x.upper() for x in out}:
            out.append(item)
    return out


def resolve_port(session, name, country_hint=""):
    last_items = []
    for query in alias_candidates(name):
        resp = session.post(PORT_SEARCH_URL, data={"f_cmd": "123", "loc_nm": query}, timeout=25)
        data = load_json_response(resp)
        items = data.get("list") or []
        if items:
            last_items = items
            chosen = choose_port(items, name, country_hint)
            log(f"   Port '{name}' -> {chosen.get('locNm')} ({chosen.get('locCd')})")
            return chosen
    sample = ", ".join(str(x.get("locNm") or x.get("locCd") or "") for x in last_items[:5])
    raise ValueError(f"Port not found on CUL: {name}" + (f" | sample: {sample}" if sample else ""))


def choose_port(items, original_name, country_hint=""):
    country = norm(country_hint or FILTER_COUNTRY)
    original = norm(original_name)

    def score(item):
        loc_nm = norm(item.get("locNm"))
        loc_cd = norm(item.get("locCd"))
        s = 0
        if country and country in loc_nm:
            s += 100
        if original and loc_nm.startswith(original):
            s += 30
        if original and original in loc_nm:
            s += 15
        if loc_cd.endswith("SGN") or loc_cd.endswith("HPH"):
            s += 3
        return s

    return sorted(items, key=score, reverse=True)[0]


def date_chunks(start_dt, end_dt, max_days=30):
    cur = start_dt
    while cur <= end_dt:
        chunk_end = min(end_dt, cur + timedelta(days=max_days))
        yield cur, chunk_end
        cur = chunk_end + timedelta(days=1)


def fetch_schedules(session, pol_cd, pod_cd, start_dt, end_dt):
    rows = []
    for chunk_start, chunk_end in date_chunks(start_dt, end_dt):
        payload = {
            "f_cmd": "3",
            "por_cd": pol_cd,
            "por_nde_cd": "",
            "del_cd": pod_cd,
            "del_nde_cd": "",
            "frm_dt": chunk_start.strftime("%Y-%m-%d"),
            "to_dt": chunk_end.strftime("%Y-%m-%d"),
            "ts_ind": "",
            "svc_flg": "C",
            "time_flg": "D",
            "tran_tm": "60",
        }
        resp = session.post(SCHEDULE_URL, data=payload, timeout=35)
        data = load_json_response(resp)
        rows.extend(data.get("list") or [])
    return rows


def parse_cul_datetime(raw):
    s = str(raw or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:16] if "%H" in fmt else s[:10], fmt)
        except Exception:
            pass
    return None


def clean_port_name(raw):
    s = re.sub(r"\([^)]*\)", "", str(raw or "")).strip()
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def parse_transshipment(row, pol_name, pod_name):
    pol = clean_port_name(pol_name)
    pod = clean_port_name(pod_name)
    candidates = []
    for key in ("n2ndLocNm", "n2ndPolLocNm", "n3rdPolLocNm", "n4thPolLocNm", "n5thPolLocNm"):
        val = clean_port_name(row.get(key))
        if not val or val in {"DIRECT", "T/S", "TS", "-"}:
            continue
        if val in {pol, pod}:
            continue
        if val not in candidates:
            candidates.append(val)
    return " + ".join(candidates) if candidates else "DIRECT"


def build_vessel_name(row):
    vessel = str(row.get("n1stVslNm") or row.get("vslNm") or "TBA").strip()
    voyage = str(row.get("consVoyNr") or row.get("n1stVvd") or row.get("skdVoyNo") or "").strip()
    if voyage and voyage.upper() not in vessel.upper():
        return f"{vessel} {voyage}".strip()
    return vessel or "TBA"


def parse_entries(rows, pol_name, pod_name):
    entries = []
    for row in rows:
        try:
            etd = parse_cul_datetime(row.get("polEtdDt") or row.get("porEtdDt") or row.get("etdDt"))
            eta = parse_cul_datetime(row.get("lstPodEtaDt") or row.get("podEtaDt") or row.get("etaDt"))
            if not etd:
                continue
            tt_raw = row.get("ttlTzDys") or row.get("ocnTzDys") or row.get("tzDys")
            try:
                tt_days = int(float(str(tt_raw).strip()))
            except Exception:
                if eta:
                    tt_days = max(0, math.ceil((eta - etd).total_seconds() / 86400))
                else:
                    tt_days = 0
            entries.append(
                {
                    "etd_dt": etd,
                    "eta_dt": eta,
                    "tt_days": tt_days,
                    "vessel": build_vessel_name(row),
                    "ts_port": parse_transshipment(row, pol_name, pod_name),
                }
            )
        except Exception as exc:
            log(f"   Warn parse CUL row: {exc}")
    return entries


def apply_etd_rules(entries, valid_dt):
    min_etd, _ = compute_valid_window(valid_dt)
    future = [
        e for e in entries
        if e["etd_dt"] >= min_etd and etd_within_max(e["etd_dt"]) and (valid_dt is None or e["etd_dt"] <= valid_dt)
    ]
    future.sort(key=lambda e: (e["etd_dt"], int(e.get("tt_days") or 9999)))
    selected = []
    for e in future:
        if len(selected) >= 3:
            break
        if selected and (e["etd_dt"] - selected[-1]["etd_dt"]).days < 1:
            continue
        selected.append(e)
    return selected


def fmt_date_short(dt):
    return f"{dt.day}-{dt.strftime('%b')}"


def format_etd_text(entries):
    if not entries:
        return ""
    fmt = [(e["etd_dt"].day, e["etd_dt"].strftime("%b")) for e in entries]
    if len(fmt) == 1:
        return f"{fmt[0][0]}-{fmt[0][1]}"
    if len(fmt) == 2:
        return f"{fmt[0][0]}-{fmt[0][1]} & {fmt[1][0]}-{fmt[1][1]}"
    months = [m for _, m in fmt]
    if months[0] == months[1] == months[2]:
        return f"{fmt[0][0]}, {fmt[1][0]}, {fmt[2][0]}-{fmt[2][1]}"
    if months[0] == months[1]:
        return f"{fmt[0][0]}, {fmt[1][0]}-{fmt[1][1]} & {fmt[2][0]}-{fmt[2][1]}"
    if months[1] == months[2]:
        return f"{fmt[0][0]}-{fmt[0][1]}, {fmt[1][0]} & {fmt[2][0]}-{fmt[2][1]}"
    return f"{fmt[0][0]}-{fmt[0][1]}, {fmt[1][0]}-{fmt[1][1]} & {fmt[2][0]}-{fmt[2][1]}"


def format_tt_text(entries):
    tts = [int(e.get("tt_days") or 0) for e in entries if e.get("tt_days") is not None]
    if not tts:
        return ""
    if len(set(tts)) == 1:
        return str(tts[0])
    return f"{min(tts)}-{max(tts)}"


def format_vessel_block(entries):
    lines = []
    ts_seen = []
    for e in entries:
        ts = e.get("ts_port") or "DIRECT"
        lines.append(
            f"{e.get('vessel') or 'TBA'} / ETD: {fmt_date_short(e['etd_dt'])}"
            f" / Transit time: {int(e.get('tt_days') or 0)} Days / Transshipment Port: {ts}"
        )
        if ts not in ts_seen:
            ts_seen.append(ts)
    return "\n".join(lines), " or\n".join(ts_seen) if ts_seen else "DIRECT"


def write_excel_row(row_i, etd_text, tt_text, vessel_text, ts_text, error=None, valid_dt=None):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        if ws.cell(row=row_i, column=6).value is None:
            ws.cell(row=row_i, column=6).value = "-"
        if valid_dt and not ws.cell(row=row_i, column=11).value:
            ws.cell(row=row_i, column=11).value = f"{valid_dt.day}-{valid_dt.strftime('%b')}"
        if error:
            ws.cell(row=row_i, column=9).value = error
        else:
            ws.cell(row=row_i, column=9).value = etd_text
            ws.cell(row=row_i, column=10).value = tt_text
            ws.cell(row=row_i, column=15).value = vessel_text
            ws.cell(row=row_i, column=16).value = ts_text
            wrap = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_i, column=15).alignment = wrap
            ws.cell(row=row_i, column=16).alignment = wrap
        wb.save(EXCEL_PATH)
        log(f"   Wrote Excel row {row_i}")
    except PermissionError:
        log("   Cannot write Excel: close input_gia.xlsx then run again.")
    except Exception as exc:
        log(f"   write_excel_row error: {exc}")


def search_one(session, pol_excel, pod_excel, country_hint, valid_dt):
    start_dt, end_dt = compute_valid_window(valid_dt)
    if not start_dt or not end_dt:
        return {"error": "VALID rong/khong parse duoc"}
    if end_dt < start_dt:
        return {"error": f"VALID truoc ETD toi thieu date +{DATE_OFFSET_DAYS}"}

    pol = resolve_port(session, pol_excel, "")
    pod = resolve_port(session, pod_excel, country_hint)
    log(f"   Search API: {pol.get('locCd')} -> {pod.get('locCd')} | {start_dt:%Y-%m-%d}..{end_dt:%Y-%m-%d}")
    rows = fetch_schedules(session, pol.get("locCd"), pod.get("locCd"), start_dt, end_dt)
    entries = parse_entries(rows, pol.get("locNm"), pod.get("locNm"))
    log(f"   API rows={len(rows)} parsed={len(entries)}")
    selected = apply_etd_rules(entries, valid_dt)
    if not selected:
        return {"error": "Khong co lich tau hop le"}
    etd_text = format_etd_text(selected)
    tt_text = format_tt_text(selected)
    vessel_text, ts_text = format_vessel_block(selected)
    return {
        "etd_text": etd_text,
        "tt_text": tt_text,
        "vessel_text": vessel_text,
        "ts_text": ts_text,
    }


def main():
    if not os.path.exists(EXCEL_PATH):
        log(f"Excel not found: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    target_single_row = None
    if SINGLE_ROW:
        try:
            target_single_row = int(SINGLE_ROW)
            log(f"[SINGLE_ROW] Chi chay dong {target_single_row} theo main.py")
        except Exception:
            log(f"[SINGLE_ROW] Khong hop le: {SINGLE_ROW}")

    targets = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if target_single_row is not None and i != target_single_row:
            continue
        country = str(row[1] or "").strip() if len(row) > 1 else ""
        pol = str(row[2] or "").strip() if len(row) > 2 else ""
        pod = str(row[3] or "").strip() if len(row) > 3 else ""
        carrier = str(row[4] or "").strip().upper() if len(row) > 4 else ""
        valid_raw = row[10] if len(row) > 10 else None
        if not pol or not pod or carrier not in CARRIER_TARGETS:
            continue
        if FILTER_POL and pol.upper() != FILTER_POL:
            continue
        if FILTER_POD and pod.upper() != FILTER_POD:
            continue
        targets.append((i, country, pol, pod, valid_raw))
    wb.close()

    log(f"Co {len(targets)} dong can check (carrier = CUL)")
    if not targets:
        return

    session = make_session()
    for idx, (row_i, country, pol, pod, valid_raw) in enumerate(targets, start=1):
        log(f"\n========== [{idx}/{len(targets)}] DONG {row_i}: {pol} -> {pod} | VALID={valid_raw}")
        valid_dt = parse_valid_date(valid_raw)
        if not valid_dt:
            log("   Valid rong -> bo qua row, khong goi API.")
            continue
        try:
            result = search_one(session, pol, pod, country, valid_dt)
        except Exception as exc:
            result = {"error": f"Exception: {exc}"}
        if result.get("error"):
            log(f"   ERROR: {result['error']}")
            write_excel_row(row_i, "", "", "", "", error=result["error"])
        else:
            log(f"   ETD={result['etd_text']} | TT={result['tt_text']} | TS={result['ts_text']}")
            write_excel_row(
                row_i,
                result["etd_text"],
                result["tt_text"],
                result["vessel_text"],
                result["ts_text"],
            )
        time.sleep(0.8)

    log("\nDONE bot CUL")


if __name__ == "__main__":
    main()
